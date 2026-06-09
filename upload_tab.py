# ============================================================
# UPLOAD SYSTEM — Phase 11 addition to dashboard_phase10.py
# ============================================================
# WHAT TO DO:
#   1. Add "📤 Upload Planning" to the tabs list

import streamlit as st
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
#   2. Add a `depot_status` table in Supabase:
#
#   create table depot_status (
#     commercial   text primary key,
#     statut       text default 'en_attente',
#     nb_agri      int  default 0,
#     tonnage      numeric default 0,
#     depose_le    timestamp,
#     fichier_nom  text
#   );
#   alter table depot_status disable row level security;
#
#   3. Paste the tab content below at the end of dashboard_phase10.py
# ============================================================

# ── TEMPLATE MODEL Excel — columns each commercial must fill ──

# ── Besoin usines par région — source: TONNAGE_PAR_REGION juin26 ─────────
BESOIN_USINE_REGION = {
    "CAP BON 1":        {"SICAM":22000, "TUCAL":11000, "COMOCAP":12500, "ABIDA":0,    "ELFALLEH":4000},
    "CAP BON 2":        {"SICAM":22000, "TUCAL":11000, "COMOCAP":12500, "ABIDA":0,    "ELFALLEH":4000},
    "CAP BON":          {"SICAM":22000, "TUCAL":11000, "COMOCAP":12500, "ABIDA":0,    "ELFALLEH":4000},
    "NORD":             {"SICAM":3500,  "TUCAL":4500,  "COMOCAP":3000,  "ABIDA":1000, "ELFALLEH":500},
    "GAFSA / KASSRINE": {"SICAM":11000, "TUCAL":2500,  "COMOCAP":500,   "ABIDA":3500, "ELFALLEH":0},
    "KAIROUAN":         {"SICAM":4500,  "TUCAL":1200,  "COMOCAP":1500,  "ABIDA":1500, "ELFALLEH":500},
    "SIDI BOUZID":      {"SICAM":2500,  "TUCAL":500,   "COMOCAP":1500,  "ABIDA":2000, "ELFALLEH":0},
    "BOUFICHA":         {"SICAM":1000,  "TUCAL":1000,  "COMOCAP":1000,  "ABIDA":0,    "ELFALLEH":0},
}

REGION_NORM_BESOIN = {
    "CAPB1":"CAP BON 1","CAP B1":"CAP BON 1","NABEUL":"CAP BON 2",
    "CAPB2":"CAP BON 2","CAP B2":"CAP BON 2",
    "BEJA":"NORD","MANOUBA":"NORD",
    "GAFSA":"GAFSA / KASSRINE","KASSRINE":"GAFSA / KASSRINE",
}

def divide_by_besoin(tonnage, usines, region_raw):
    """Divise le tonnage selon le besoin de chaque usine dans cette région."""
    region = str(region_raw).strip().upper()
    region = REGION_NORM_BESOIN.get(region, region)
    besoins = BESOIN_USINE_REGION.get(region, {})
    relevant = {u: max(besoins.get(u, 0), 1) for u in usines}  # min 1 pour éviter 0
    total_b = sum(relevant.values())
    result = {}
    remaining = tonnage
    for i, u in enumerate(usines):
        if i == len(usines) - 1:
            result[u] = round(remaining, 1)
        else:
            t = round(tonnage * relevant[u] / total_b, 1)
            result[u] = t
            remaining -= t
    return result

TEMPLATE_COLUMNS = [
    "NOM_AGRICULTEUR",
    "TONNAGE",
    "USINE",           # SICAM / COMOCAP / TUCAL / ABIDA / ELFALLEH
    "ACCESSIBILITE",   # PL/PPL or PL/SEMI
    "DATE_DEBUT",      # YYYY-MM-DD
    "DATE_FIN",        # YYYY-MM-DD
    "REGION",          # optional
    "ZONE",            # optional
]

USINES_VALIDES = {"SICAM", "COMOCAP", "TUCAL", "ABIDA", "ELFALLEH"}
# Séparateur officiel = "/" uniquement. PL-PPL / PL-SEMI n'existent pas.
ACCESS_VALIDES = {"PL/PPL", "PL/SEMI", "RM", "TRC/PPL", "TRC/PPL/PL", "PL/PPL/TRC",
                  "PL", "PPL", "SEMI"}  # ✅ PL seul et PPL seul sont valides

# ── Alias usines ─────────────────────────────────────────────
# Toutes les abréviations/fautes rencontrées dans les fichiers commerciaux
USINE_ALIASES = {
    "SI":       "SICAM",
    "SIC":      "SICAM",
    "COM":      "COMOCAP",
    "COCO":     "COMOCAP",
    "TUC":      "TUCAL",
    "ABI":      "ABIDA",
    "ABD":      "ABIDA",
    "FAL":      "ELFALLEH",
    "FALL":     "ELFALLEH",
    "FELLEH":   "ELFALLEH",
    "ELFAL":    "ELFALLEH",
    "ELF":      "ELFALLEH",
}

def normalize_usine(token: str) -> str:
    """Normalise un token usine vers le nom officiel."""
    t = token.strip().upper()
    if t in USINES_VALIDES:
        return t
    return USINE_ALIASES.get(t, "")

def parse_usines(raw_value) -> list:
    """
    Parse la colonne USINE — accepte tous les séparateurs et abréviations.

    Exemples gérés :
      'SICAM'                       → ['SICAM']
      'SICAM/COMOCAP'               → ['SICAM','COMOCAP']
      'COMOCAP SICAM'               → ['COMOCAP','SICAM']
      'SI/COM/FAL'                  → ['SICAM','COMOCAP','ELFALLEH']
      'SICAM FALL'                  → ['SICAM','ELFALLEH']
      'SICAM TUCAL'                 → ['SICAM','TUCAL']
      'SICAM /COMOCAP/TUCAL/FALL/ABIDA' → ['SICAM','COMOCAP','TUCAL','ELFALLEH','ABIDA']
    """
    import re
    raw = str(raw_value).strip().upper()
    # Séparateurs : / , ; et espaces multiples
    parts = re.split(r'[/,;\s]+', raw)
    valid = []
    seen  = set()
    for p in parts:
        p = p.strip()
        if not p:
            continue
        norm = normalize_usine(p)
        if norm and norm not in seen:
            valid.append(norm)
            seen.add(norm)
    return valid

def normalize_access(raw_value: str) -> str:
    """
    Normalise la colonne ACCESSIBILITE.
    ✅ RÈGLE PRINCIPALE: respecter EXACTEMENT ce que le commercial a écrit.
       PL seul → PL  |  PPL seul → PPL  |  PL/PPL → PL/PPL
    """
    raw = str(raw_value).strip().upper().replace("-", "/")
    
    # Déjà valide → retourner sans modifier
    if raw in ACCESS_VALIDES:
        return raw
    if raw == "RM":
        return "RM"
    
    import re
    parts = set(re.split(r"[/,;\s]+", raw))
    parts = {p.strip() for p in parts if p.strip()}
    
    has_semi = "SEMI" in parts
    has_ppl  = "PPL" in parts
    has_pl   = "PL" in parts
    has_trc  = "TRC" in parts or "TRACTEUR" in parts or "TRAC" in parts
    
    # TRC présent → garder TRC/PPL
    if has_trc and has_pl and has_ppl:
        return "TRC/PPL/PL"
    if has_trc and (has_ppl or has_pl):
        return "TRC/PPL"
    if has_trc:
        return "TRC/PPL"
    
    # SEMI combiné
    if has_semi and has_pl:
        return "PL/SEMI"
    if has_semi and has_ppl:
        return "PL/PPL/SEMI" if "PL/PPL/SEMI" in ACCESS_VALIDES else "PL/SEMI"
    if has_semi:
        return "PL/SEMI"
    
    # Combinaisons PL + PPL
    if has_pl and has_ppl:
        return "PL/PPL"
    
    # ✅ Types simples — RESPECTER EXACTEMENT
    if has_pl:
        return "PL"
    if has_ppl:
        return "PPL"
    
    return raw

def fix_date(raw_date, col_name: str = "", nom: str = "") -> tuple:
    """
    Corrige une date avec année invalide.
    Détecte auto le format YYYY-MM-DD vs DD/MM/YYYY.
    """
    import pandas as pd, re as _re
    if raw_date is None: return None, None
    raw_str = str(raw_date).strip()
    if raw_str in ("", "nan", "NaT", "None"): return None, None
    # Format YYYY-MM-DD → dayfirst=False obligatoire
    if _re.match(r"^\d{4}-\d{2}-\d{2}", raw_str):
        d = pd.to_datetime(raw_str, errors="coerce", dayfirst=False)
    else:
        d = pd.to_datetime(raw_date, errors="coerce", dayfirst=True)
    if pd.isna(d):
        return None, f"{nom}: {col_name} illisible (valeur: '{raw_str}')"
    if d.year < 2026:
        d_fixed = d.replace(year=2026)
        return d_fixed, f"{nom}: {col_name} corrigée {d.date()} → {d_fixed.date()}"
    return d, None

def generate_template_excel() -> bytes:
    """Generate downloadable empty Excel template for commercial."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Mon Planning"

    # Header
    headers = TEMPLATE_COLUMNS
    fill    = PatternFill("solid", start_color="1F4E79")
    font    = Font(bold=True, color="FFFFFF", name="Calibri")
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = font; c.fill = fill
        c.alignment = Alignment(horizontal="center")
        ws.column_dimensions[chr(64+ci)].width = 18

    # Example rows
    examples = [
        ["EXEMPLE BEN ALI", 150, "SICAM", "PL/PPL", "2026-07-01", "2026-07-20", "NABEUL", "KORBA"],
        ["EXEMPLE MULTI USINE", 200, "SICAM/COMOCAP", "PL/SEMI", "2026-07-05", "2026-07-25", "NORD", ""],
        ["EXEMPLE SEMI ONLY", 300, "COMOCAP", "RM", "2026-07-10", "2026-08-01", "CAP BON 1", "DAR ALLOUCH"],
    ]
    for row_idx, example in enumerate(examples, 2):
        for ci, val in enumerate(example, 1):
            c = ws.cell(row=row_idx, column=ci, value=val)
            c.font = Font(italic=True, color="888888", name="Calibri", size=9)

    # Instructions sheet
    ws2 = wb.create_sheet("Instructions")
    instructions = [
        ("Colonne", "Description", "Valeurs acceptées"),
        ("NOM_AGRICULTEUR", "Nom complet de l'agriculteur", "Texte libre"),
        ("TONNAGE", "Tonnage total prévu (en tonnes)", "Nombre > 0"),
        ("USINE", "Usine(s) de livraison", "SICAM / COMOCAP / TUCAL / ABIDA / ELFALLEH — plusieurs: SICAM/COMOCAP"),
        ("ACCESSIBILITE", "Type d'accès à la ferme", "PL/PPL · PL/SEMI · RM (=100% Semi, 3-5/jour)"),
        ("DATE_DEBUT", "Date début récolte", "Format: 2026-07-01"),
        ("DATE_FIN", "Date fin récolte", "Format: 2026-07-31"),
        ("REGION", "Région (optionnel)", "NABEUL / NORD / KAIROUAN / etc."),
        ("ZONE", "Zone précise (optionnel)", "Ex: KORBA / DAR ALLOUCH / etc."),
        ("", "", ""),
        ("CAPS JOURNALIERS", "Limites par usine", "SICAM:1300 · TUCAL:750 · COMOCAP:700 · ABIDA:150 · ELFALLEH:100"),
        ("CAPS COMMERCIAUX", "Limites par commercial", "FEDI:850 · MAKKI:800 · KHALIL:800 · ACHREF:500 · JILANI:50"),
        ("COMOCAP TRANSPORT", "Extra tracteur", "+100 tracteurs COMOCAP pour caisses (automatique)"),
    ]
    for ri, row in enumerate(instructions, 1):
        for ci, val in enumerate(row, 1):
            c = ws2.cell(row=ri, column=ci, value=val)
            if ri == 1:
                c.font = Font(bold=True, color="FFFFFF", name="Calibri")
                c.fill = PatternFill("solid", start_color="1F4E79")
            c.alignment = Alignment(horizontal="left")
            ws2.column_dimensions[chr(64+ci)].width = 30

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def validate_upload(df_upload: pd.DataFrame, commercial_name: str) -> tuple:
    """
    Valide et normalise le fichier uploadé.
    Gère automatiquement :
    - Abréviations usines (SI→SICAM, FAL→ELFALLEH, COM→COMOCAP…)
    - Séparateurs variés (/ espace , ;)
    - Plusieurs usines par agriculteur (→ expansion en plusieurs lignes)
    - Casse mixte accessibilité (trc/ppl, pl/ppl/TRC → PL/PPL)
    - TRC/TRACTEUR dans accessibilité → converti en PL/PPL
    - SEMI/pl/ppl → PL/PPL
    - Années invalides dans les dates (1902, 1905, 2025 → 2026)
    Returns (is_valid, errors, warnings, cleaned_df)
    """
    import pandas as pd, re

    errors   = []
    warnings = []   # corrections non-bloquantes

    # ── Nettoyer les noms de colonnes (espaces, casse) ───────
    # Gère les cas comme 'TONNAGE ' (espace final), 'nom_agriculteur' (minuscule)
    df_upload = df_upload.copy()
    df_upload.columns = [str(c).strip() for c in df_upload.columns]
    
    # ✅ Mapping automatique des noms de colonnes (gère espaces, accents, variantes)
    RENAME_MAP = {
        # NOM
        "NOM AGRICULTEUR": "NOM_AGRICULTEUR", "NOM_AGRICULTEUR": "NOM_AGRICULTEUR",
        "AGRICULTEUR": "NOM_AGRICULTEUR", "NOM": "NOM_AGRICULTEUR",
        # TONNAGE
        "TONNAGE (T)": "TONNAGE", "TONNAGE(T)": "TONNAGE", "TONNAGE": "TONNAGE",
        # HECTARES
        "HECTARES": "NBR_HECTARES", "NBR_HECTARES": "NBR_HECTARES",
        "NBR,HECTAR": "NBR_HECTARES", "NBR HECTAR": "NBR_HECTARES",
        "NBR HECTARES": "NBR_HECTARES",
        # T/HA (à ignorer, on recalcule)
        "T / HA": "_TPHA_IGNORE", "T/HA": "_TPHA_IGNORE",
        # USINE
        "USINE": "USINE", "DESTINATION": "USINE",
        # ACCESSIBILITE
        "ACCESSIBILITÉ": "ACCESSIBILITE", "ACCESSIBILITE": "ACCESSIBILITE",
        "ACCESSBILITE": "ACCESSIBILITE", "ACCES": "ACCESSIBILITE",
        # REGION
        "RÉGION": "REGION", "REGION": "REGION",
        # ZONE
        "ZONE": "ZONE", "ZONNE": "ZONE", "LOCALISATION": "ZONE",
        # DATES
        "DATE DÉBUT": "DATE_DEBUT", "DATE_DEBUT": "DATE_DEBUT",
        "DATE DEBUT": "DATE_DEBUT", "DEBUT RECOLTE": "DATE_DEBUT",
        "DATE FIN": "DATE_FIN", "DATE_FIN": "DATE_FIN",
        "FIN RECOLTE": "DATE_FIN",
        # CENTRE
        "CENTRE": "CENTRE", "CENTER": "CENTRE",
        # COMMERCIAL
        "COMMERCIAL": "COMMERCIAL",
    }
    new_cols = {}
    for col in df_upload.columns:
        col_up = col.strip().upper()
        if col_up in RENAME_MAP:
            new_cols[col] = RENAME_MAP[col_up]
    df_upload = df_upload.rename(columns=new_cols)

    # ── Colonnes requises ────────────────────────────────────
    required = ["NOM_AGRICULTEUR","TONNAGE","USINE","ACCESSIBILITE","DATE_DEBUT","DATE_FIN"]
    missing  = [c for c in required if c not in df_upload.columns]
    if missing:
        return False, [f"Colonnes manquantes : {', '.join(missing)}"], [], None

    # ── Lignes valides ───────────────────────────────────────
    df = df_upload.copy()
    _nom_upper = df["NOM_AGRICULTEUR"].astype(str).str.strip().str.upper()
    df = df[~_nom_upper.isin(["EXEMPLE BEN ALI", "NAN", "NONE"])]
    # Filtrer les lignes de TOTAL (résumé généré automatiquement)
    df = df[~_nom_upper.str.startswith("TOTAL")]
    df = df[~_nom_upper.str.startswith("SOUS-TOTAL")]
    # Filtrer lignes sans NOM valide
    df = df[_nom_upper.str.len() > 2]
    df = df.dropna(subset=["NOM_AGRICULTEUR","TONNAGE"])
    df = df[pd.to_numeric(df["TONNAGE"], errors="coerce") > 0].copy()
    df["TONNAGE"] = pd.to_numeric(df["TONNAGE"], errors="coerce")

    if len(df) == 0:
        return False, ["Aucune ligne valide dans le fichier."], [], None

    # ── Expansion multi-usines ───────────────────────────────
    expanded = []
    skipped  = []
    for _, row in df.iterrows():
        nom    = str(row["NOM_AGRICULTEUR"])
        usines = parse_usines(row["USINE"])

        if len(usines) == 0:
            skipped.append(f"{nom}: usine '{row['USINE']}' non reconnue — ligne ignorée")
            continue

        if len(usines) == 1:
            r = row.copy()
            r["USINE"] = usines[0]
            expanded.append(r)
        else:
            # Diviser le tonnage selon le BESOIN de chaque usine dans cette région
            ton_base = float(row["TONNAGE"])
            region   = str(row.get("REGION", "")).strip()
            division = divide_by_besoin(ton_base, usines, region)
            for u in usines:
                r = row.copy()
                r["USINE"]   = u
                r["TONNAGE"] = division[u]
                expanded.append(r)
            div_str = ", ".join(f"{u}={division[u]:.0f}t" for u in usines)
            warnings.append(
                f"{nom}: usines {usines} → {len(usines)} lignes "
                f"({ton_base:.0f}t divisé par besoin: {div_str})"
            )

    if skipped:
        warnings.extend(skipped)

    if len(expanded) == 0:
        return False, ["Aucune ligne avec usine reconnue."], warnings, None

    df = pd.DataFrame(expanded).reset_index(drop=True)

    # ── Normalisation ACCESSIBILITE ──────────────────────────
    original_access = df["ACCESSIBILITE"].copy()
    df["ACCESSIBILITE"] = df["ACCESSIBILITE"].apply(normalize_access)
    for i, (orig, norm) in enumerate(zip(original_access, df["ACCESSIBILITE"])):
        if str(orig).upper().strip() != norm and norm in ACCESS_VALIDES:
            nom = df.iloc[i]["NOM_AGRICULTEUR"]
            warnings.append(f"{nom}: accessibilité '{orig}' → '{norm}'")
    # Lignes toujours invalides après normalisation
    bad_access = df[~df["ACCESSIBILITE"].isin(ACCESS_VALIDES)]
    if len(bad_access) > 0:
        errors.append(
            f"Accessibilités non reconnues (après correction auto) : "
            f"{bad_access['ACCESSIBILITE'].unique().tolist()}"
        )

    # ── Correction dates ─────────────────────────────────────
    dates_debut, dates_fin = [], []
    for _, row in df.iterrows():
        nom = str(row["NOM_AGRICULTEUR"])
        d, w = fix_date(row["DATE_DEBUT"], "DATE_DEBUT", nom)
        dates_debut.append(d)
        if w: warnings.append(w)

        f, w = fix_date(row["DATE_FIN"], "DATE_FIN", nom)
        dates_fin.append(f)
        if w: warnings.append(w)

    df["DATE_DEBUT"] = pd.to_datetime(dates_debut)
    df["DATE_FIN"]   = pd.to_datetime(dates_fin)

    # Dates illisibles
    bad_dates = df[df["DATE_DEBUT"].isna() | df["DATE_FIN"].isna()]
    if len(bad_dates) > 0:
        errors.append(f"{len(bad_dates)} lignes avec dates illisibles même après correction")

    # DATE_FIN <= DATE_DEBUT
    bad_order = df[df["DATE_FIN"] <= df["DATE_DEBUT"]]
    if len(bad_order) > 0:
        noms = bad_order["NOM_AGRICULTEUR"].tolist()
        errors.append(f"DATE_FIN ≤ DATE_DEBUT pour : {noms} — vérifiez les dates")

    if errors:
        return False, errors, warnings, None

    # ── Finalisation ─────────────────────────────────────────
    df["NOM_AGRICULTEUR"] = df["NOM_AGRICULTEUR"].astype(str).str.strip()
    df["REGION"]   = df.get("REGION", pd.Series(dtype=str)).fillna("").astype(str).str.strip()
    # ✅ ZONE : chercher dans "ZONE" d'abord, puis "ZONNE" (ancien nom)
    _zone_series = df.get("ZONE", df.get("ZONNE", pd.Series([""] * len(df), index=df.index, dtype=str)))
    df["ZONE"]    = _zone_series.fillna("").astype(str).str.strip()
    
    # ✅ HECTARES : chercher dans plusieurs noms possibles
    _ha_series = None
    for col_name in ["NBR_HECTARES","HECTARES","NBR HECTARES","NBR,HECTAR","NBR HECTAR","HECTARE"]:
        if col_name in df.columns:
            _ha_series = df[col_name]
            break
    if _ha_series is not None:
        df["NBR_HECTARES"] = pd.to_numeric(_ha_series, errors="coerce")
    else:
        df["NBR_HECTARES"] = pd.NA
    
    # ✅ CENTRE : chercher dans plusieurs noms
    _ctr_series = None
    for col_name in ["CENTRE","CENTER","CENTRE DE COLLECTE","CENTRE_COLLECTE"]:
        if col_name in df.columns:
            _ctr_series = df[col_name]
            break
    if _ctr_series is not None:
        df["CENTRE"] = _ctr_series.fillna("").astype(str).str.strip()
    else:
        df["CENTRE"] = ""
    
    df["COMMERCIAL"] = commercial_name

    return True, [], warnings, df


# ── TAB: UPLOAD PLANNING ─────────────────────────────────────

# ── Division intelligente multi-usines ──────────────────────────────
# Capacités journalières officielles (t/jour pendant PIC)
USINE_CAPS = {
    "SICAM": 1300, "COMOCAP": 700, "TUCAL": 750,
    "ABIDA": 150,  "ELFALLEH": 100,
}
USINE_ALIASES = {
    "FALL": "ELFALLEH", "FELLA": "ELFALLEH", "FELLEH": "ELFALLEH",
    "FAL":  "ELFALLEH", "SI": "SICAM", "COM": "COMOCAP",
    "TUC":  "TUCAL",    "ABI": "ABIDA",
}

def split_usines_intelligent(usine_str, tonnage, hectares=None):
    """
    Divise tonnage entre plusieurs usines proportionnellement à leurs capacités.
    usine_str peut être: "SICAM", "SICAM/COMOCAP", "SI/COM/FAL", "SICAM TUCAL", etc.
    Retourne: [(usine, tonnage, hectares), ...]
    """
    import re as _re
    
    # Parser les usines
    parts = [p.strip().upper() for p in _re.split(r"[/,;\s]+", usine_str) if p.strip()]
    usines = []
    for p in parts:
        u = USINE_ALIASES.get(p, p)
        if u in USINE_CAPS and u not in usines:
            usines.append(u)
    
    if len(usines) <= 1:
        u = usines[0] if usines else usine_str.strip().upper()
        return [(u, round(float(tonnage), 1), round(float(hectares), 2) if hectares else None)]
    
    # Division proportionnelle selon capacités
    total_cap   = sum(USINE_CAPS[u] for u in usines)
    result      = []
    rem_ton     = float(tonnage)
    rem_ha      = float(hectares) if hectares else None
    
    for i, u in enumerate(usines):
        is_last = (i == len(usines) - 1)
        ratio   = USINE_CAPS[u] / total_cap
        
        if is_last:
            ton = round(rem_ton, 1)
            ha  = round(rem_ha, 2) if rem_ha is not None else None
        else:
            ton = round(float(tonnage) * ratio, 1)
            ha  = round(float(hectares) * ratio, 2) if hectares else None
            rem_ton = round(rem_ton - ton, 1)
            if rem_ha is not None:
                rem_ha = round(rem_ha - ha, 2)
        
        result.append((u, ton, ha))
    
    return result


def render_upload_tab(sb, CURRENT_ROLE, CURRENT_NAME, CURRENT_FILTER,
                      GLOBAL_COMMERCIAL_FARMERS, GLOBAL_COMMERCIAL_TONS,
                      df_to_csv):
    """
    Render the upload tab.
    Call this with: render_upload_tab(sb, CURRENT_ROLE, CURRENT_NAME, ...)
    """

    # ── ADMIN VIEW ──────────────────────────────────────────
    if CURRENT_ROLE == "directeur":
        st.subheader("📊 Tableau de bord — Dépôts des commerciaux")
        
        # ✅ Bouton de réinitialisation complète
        with st.expander("⚙️ Actions administrateur", expanded=False):
            col_r1, col_r2 = st.columns(2)
            with col_r1:
                if st.button("🗑️ Vider TOUS les commerciaux", type="secondary",
                             help="Supprime toutes les données de tous les commerciaux"):
                    try:
                        sb.table("agriculteurs").delete().neq("commercial","__NONE__").execute()
                        sb.table("depot_status").delete().neq("commercial","__NONE__").execute()
                        st.cache_data.clear()
                        st.success("✅ Toutes les données supprimées. Les commerciaux peuvent re-uploader.")
                        st.rerun()
                    except Exception as e:
                        st.error(f"Erreur: {e}")
            with col_r2:
                comm_to_del = st.selectbox("Vider un seul commercial",
                    ["--","FEDI","MAKKI BEN SALAH","KHALIL","ACHREF AJLANI","JILANI OBAY"])
                if comm_to_del != "--":
                    if st.button(f"🗑️ Vider {comm_to_del}"):
                        try:
                            sb.table("agriculteurs").delete().eq("commercial", comm_to_del).execute()
                            sb.table("depot_status").delete().eq("commercial", comm_to_del).execute()
                            st.cache_data.clear()
                            st.success(f"✅ {comm_to_del} supprimé.")
                            st.rerun()
                        except Exception as e:
                            st.error(f"Erreur: {e}")

        commercials = ["FEDI","MAKKI BEN SALAH","KHALIL","ACHREF AJLANI","JILANI OBAY"]

        # Load depot status from Supabase (pour DATE upload et FICHIER)
        try:
            status_data = sb.table("depot_status").select("*").execute().data
            status_map  = {r["commercial"]: r for r in status_data}
        except Exception:
            status_map = {}
        
        # ✅ Charger TONNAGE RÉEL depuis agriculteurs (source de vérité)
        # Seuls les commerciaux ayant des données ACTUELLES sont "déposé"
        try:
            agri_data = sb.table("agriculteurs").select("commercial,nom,tonnage_total").execute().data or []
            # Commerciaux avec données réelles = ceux dans agriculteurs
            active_commercials = {r.get("commercial","") for r in agri_data if r.get("commercial")}
            real_tonnage = {}
            real_nb_agri = {}
            for r in agri_data:
                comm = r.get("commercial","")
                ton  = float(r.get("tonnage_total",0) or 0)
                if comm and ton > 0:
                    real_tonnage[comm] = real_tonnage.get(comm, 0) + ton
                    real_nb_agri[comm] = real_nb_agri.get(comm, set())
                    real_nb_agri[comm].add(r.get("nom",""))
            # convertir sets en counts
            real_nb_agri = {k: len(v) for k, v in real_nb_agri.items()}
        except Exception:
            real_tonnage = {}
            real_nb_agri = {}

        # Summary cards — utiliser TONNAGE RÉEL (agriculteurs) pas depot_status
        n_done = sum(1 for c in commercials if real_tonnage.get(c, 0) > 0)
        total_tons_received = sum(real_tonnage.get(c, 0) for c in commercials)

        c1, c2, c3 = st.columns(3)
        c1.metric("Commerciaux ayant déposé", f"{n_done}/{len(commercials)}")
        c2.metric("Tonnes reçues", f"{total_tons_received:,.0f} t")
        c3.metric("Statut global",
                  "✅ Prêt à générer" if n_done == len(commercials) else f"⏳ {len(commercials)-n_done} en attente")

        st.markdown("---")

        # Status table
        status_rows = []
        for comm in commercials:
            s = status_map.get(comm, {})
            statut = s.get("statut", "en_attente")
            depose_le = s.get("depose_le", "—")
            if depose_le and depose_le != "—":
                try:
                    depose_le = pd.to_datetime(depose_le).strftime("%d/%m %H:%M")
                except Exception:
                    pass
            # ✅ Utiliser TONNAGE RÉEL depuis agriculteurs (pas depot_status)
            real_ton = real_tonnage.get(comm, 0)
            real_n   = real_nb_agri.get(comm, 0)
            status_rows.append({
                "Commercial":      comm,
                "Statut":          "✅ Déposé"    if real_ton > 0           else
                                   "⏳ En attente" if statut == "en_attente" else "❌ Erreur",
                "Agriculteurs":    real_n if real_ton > 0 else "—",
                "Tonnage (réel)":  f"{real_ton:,.0f}t" if real_ton > 0 else "—",
                "Fichier":         s.get("fichier_nom", "—"),
                "Déposé le":       depose_le,
            })

        st.dataframe(
            pd.DataFrame(status_rows),
            use_container_width=True,
            hide_index=True,
            column_config={
                "Statut": st.column_config.TextColumn(width="medium"),
            }
        )

        st.markdown("---")

        # Generate full planning button
        if n_done < len(commercials):
            missing = [c for c in commercials
                       if status_map.get(c, {}).get("statut") != "depose"]
            st.warning(f"⚠️ En attente de : **{', '.join(missing)}**")
            force = st.checkbox("Générer quand même avec les données disponibles")
            can_generate = force
        else:
            st.success("✅ Tous les commerciaux ont déposé leur planning !")
            can_generate = True

        if can_generate:
            if st.button("🚀 Générer le planning complet (OR-Tools)",
                         type="primary", use_container_width=True):
                # First: export all uploaded data to original Excel format
                try:
                    all_agri = sb.table("agriculteurs").select("*").execute().data
                    st.info(f"📥 {len(all_agri)} agriculteurs chargés depuis Supabase")
                    st.info("👉 Lance maintenant : `python optimizer_v2.py` puis `python migrate.py`")
                    st.success("✅ Les données sont prêtes. Lance optimizer_v2.py pour générer le planning.")
                except Exception as e:
                    st.error(f"Erreur: {e}")

        # ── Live totals from Supabase ──
        st.markdown("---")
        st.subheader("📊 État actuel de Supabase")
        try:
            all_agri = sb.table("agriculteurs").select("commercial,nom,tonnage_total").execute().data
            _df_live = pd.DataFrame(all_agri) if all_agri else pd.DataFrame()
            if not _df_live.empty:
                _df_live["tonnage_total"] = pd.to_numeric(_df_live["tonnage_total"], errors="coerce")
                _live_total = _df_live["tonnage_total"].sum()
                _live_rows  = len(_df_live)
                _live_uniq  = _df_live["nom"].nunique()
                cc1, cc2, cc3 = st.columns(3)
                cc1.metric("Lignes dans agriculteurs", f"{_live_rows}")
                cc2.metric("Agriculteurs uniques", f"{_live_uniq}")
                cc3.metric("Tonnage total Supabase", f"{_live_total:,.0f} t")

                # Tonnage théorique objectif (plan officiel commerciaux)
                st.caption(f"📊 Total RÉEL en BD: {_live_total:,.0f}t — Tonnage théorique objectif: 90,310t (FEDI 31,690 + MAKKI 21,365 + KHALIL 15,705 + ACHREF 14,140 + JILANI 7,410). L'écart vient des doublons supprimés par --fix qui n'ont pas pu tous être restaurés.")
            else:
                st.warning("Table agriculteurs vide dans Supabase")
        except Exception as e:
            st.error(f"Erreur lecture Supabase: {e}")

        # Reset button (admin only)
        st.markdown("---")
        with st.expander("⚠️ Réinitialiser les dépôts"):
            comm_to_reset = st.selectbox("Commercial à réinitialiser",
                                          ["Tous"] + commercials)
            if st.button("🗑️ Réinitialiser", type="secondary"):
                try:
                    if comm_to_reset == "Tous":
                        sb.table("depot_status").delete().neq("commercial","").execute()
                    else:
                        sb.table("depot_status").delete().eq("commercial", comm_to_reset).execute()
                    st.success("Réinitialisé.")
                    st.rerun()
                except Exception as e:
                    st.error(f"Erreur: {e}")

    # ── COMMERCIAL VIEW ──────────────────────────────────────
    elif CURRENT_ROLE == "commercial":
        st.subheader(f"📤 Déposer mon planning — {CURRENT_NAME}")

        # Current status
        try:
            my_status = sb.table("depot_status").select("*").eq(
                "commercial", CURRENT_NAME).execute().data
            my_status = my_status[0] if my_status else {}
        except Exception:
            my_status = {}

        if my_status.get("statut") == "depose":
            st.success(f"✅ Planning déposé le {my_status.get('depose_le','—')[:16]} — "
                       f"{my_status.get('nb_agri',0)} agriculteurs, "
                       f"{my_status.get('tonnage',0):,.0f}t")
            st.info("Tu peux déposer un nouveau fichier pour remplacer l'ancien.")
        else:
            st.warning("⏳ Tu n'as pas encore déposé ton planning pour cette saison.")

        st.markdown("---")

        # Download template
        col1, col2 = st.columns(2)
        with col1:
            st.markdown("**Étape 1 — Télécharge le modèle Excel**")
            template_bytes = generate_template_excel()
            st.download_button(
                "⬇️ Télécharger le modèle vide",
                data=template_bytes,
                file_name=f"Modele_Planning_{CURRENT_NAME.replace(' ','_')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
            st.caption("Remplis le fichier avec tes agriculteurs, puis uploade-le ci-contre.")

        with col2:
            st.markdown("**Étape 2 — Uploade ton fichier rempli**")
            uploaded = st.file_uploader(
                "Choisir le fichier Excel",
                type=["xlsx","xls"],
                help="Le fichier doit avoir les colonnes du modèle"
            )

        if uploaded is not None:
            try:
                df_upload = pd.read_excel(uploaded)
                st.markdown(f"**Aperçu — {len(df_upload)} lignes détectées :**")
                st.dataframe(df_upload.head(5), use_container_width=True)

                is_valid, errors, warnings, df_clean = validate_upload(df_upload, CURRENT_NAME)

                if not is_valid:
                    for err in errors:
                        st.error(f"❌ {err}")
                    if warnings:
                        with st.expander("⚠️ Avertissements de correction automatique"):
                            for w in warnings:
                                st.warning(w)
                else:
                    st.success(f"✅ Fichier valide — {len(df_clean)} lignes, "
                               f"{df_clean['TONNAGE'].sum():,.0f}t détectées")
                    if warnings:
                        with st.expander(f"ℹ️ {len(warnings)} corrections automatiques appliquées (cliquez pour voir)"):
                            for w in warnings:
                                st.info(f"🔧 {w}")

                    col_a, col_b = st.columns(2)
                    with col_a:
                        st.metric("Lignes traitées", len(df_clean))
                    with col_b:
                        st.metric("Tonnage total", f"{df_clean['TONNAGE'].sum():,.0f} t")

                    if st.button("✅ Confirmer et déposer",
                                 type="primary", use_container_width=True):

                        # ✅ VALIDATION EN PREMIER — avant toute suppression
                        # validate_upload retourne colonnes NOM_AGRICULTEUR/TONNAGE/USINE (majuscules)
                        df_val = df_clean.copy()
                        
                        # Détecter les noms de colonnes (majuscules ou minuscules)
                        _nom_col = next((c for c in df_val.columns 
                                        if c.upper() in ("NOM_AGRICULTEUR","NOM")), None)
                        _ton_col = next((c for c in df_val.columns 
                                        if c.upper() in ("TONNAGE","TONNAGE_TOTAL")), None)
                        _usi_col = next((c for c in df_val.columns 
                                        if c.upper() == "USINE"), None)
                        
                        # Filtrer les lignes vides et TOTAL
                        if _nom_col:
                            _n = df_val[_nom_col].astype(str).str.strip().str.upper()
                            df_val = df_val[_n.str.len() > 1]
                            df_val = df_val[~_n.str.startswith("TOTAL")]
                            df_val = df_val[~_n.str.startswith("SOUS-TOTAL")]
                            df_val = df_val[~_n.isin(["NAN","NONE",""])]
                        if _ton_col:
                            df_val = df_val[pd.to_numeric(df_val[_ton_col], errors="coerce").fillna(0) > 0]
                        if _usi_col:
                            df_val = df_val[df_val[_usi_col].astype(str).str.strip().str.len() > 0]
                        df_clean = df_val

                        validation_errors = []
                        for _, row in df_clean.iterrows():
                            nom = str(row.get(_nom_col or "NOM_AGRICULTEUR","") or "").strip()
                            ton = float(row.get(_ton_col or "TONNAGE", 0) or 0)
                            usn = str(row.get(_usi_col or "USINE","") or "").strip().upper()
                            if not nom or len(nom) < 2:
                                validation_errors.append(f"Nom invalide: '{nom}'")
                            if ton <= 0 or ton > 50000:
                                validation_errors.append(f"Tonnage invalide: {ton}t pour {nom}")
                            if usn not in ["SICAM","COMOCAP","TUCAL","ABIDA","ELFALLEH"]:
                                validation_errors.append(f"Usine invalide: '{usn}' pour {nom}")
                        if validation_errors:
                            st.error(f"❌ {len(validation_errors)} erreur(s) de validation (données NON modifiées) :")
                            for e in validation_errors[:10]:
                                st.write(f"  • {e}")
                            st.stop()

                        # ✅ Validation OK → maintenant supprimer et insérer
                        # 1. Supprimer les anciens agriculteurs de ce commercial
                        sb.table("agriculteurs").delete().eq(
                            "commercial", CURRENT_NAME).execute()
                        
                        # 2. Insérer les nouveaux agriculteurs
                        # Normaliser region complètement
                        REGION_NORM_UPLOAD = {
                            # CAP BON
                            "NABEUL":"CAP BON 2","CAPB1":"CAP BON 1","CAPB2":"CAP BON 2",
                            "CAP B1":"CAP BON 1","CAP B2":"CAP BON 2",
                            "CAP BON":"CAP BON 1",
                            "CAP BON 1":"CAP BON 1","CAP BON 2":"CAP BON 2",
                            # GAFSA / KASSERINE
                            "GAFSA":"GAFSA / KASSRINE","KASSRINE":"GAFSA / KASSRINE",
                            "KASSERINE":"GAFSA / KASSRINE","KASRINE":"GAFSA / KASSRINE",
                            "KASSARINE":"GAFSA / KASSRINE","SBEITLA":"GAFSA / KASSRINE",
                            # NORD
                            "BEJA":"NORD","MANOUBA":"NORD","BIZERTE":"NORD",
                            "JENDOUBA":"NORD","BIR LAHFAY":"NORD",
                            "BOR AMRI":"NORD","BORJ AMRI":"NORD",
                            "MEDJEZ EL BAB":"NORD","MEJEZ EL BAB":"NORD","MEDJEZ BEB":"NORD",
                            "TESTOUR":"NORD","BOUSSALEM":"NORD",
                            # KAIROUAN
                            "KAIROUAN":"KAIROUAN","KAIRAOUAN":"KAIROUAN",
                            # SIDI BOUZID
                            "SIDI BOUZID":"SIDI BOUZID","SIDIBOUZID":"SIDI BOUZID",
                            # BOUFICHA
                            "BOUFICHA":"BOUFICHA","SOUSSE":"BOUFICHA","ENFIDHA":"BOUFICHA",
                            "HAMMAMET":"CAP BON 1",
                            # AUTRE
                            "AUTRE":"AUTRE",
                        }
                        KNOWN_REGIONS = {"CAP BON 1","CAP BON 2","NORD","KAIROUAN",
                                         "SIDI BOUZID","GAFSA / KASSRINE","BOUFICHA","AUTRE"}

                        rows_agri = []
                        for _, row in df_clean.iterrows():
                            region_raw = str(row.get("REGION","") or "").strip()
                            # Si vide/NaN → AUTRE
                            if not region_raw or region_raw.lower() in ("nan","none",""):
                                region_norm = "AUTRE"
                            else:
                                # Essayer mapping (raw puis uppercase)
                                region_norm = REGION_NORM_UPLOAD.get(
                                    region_raw, REGION_NORM_UPLOAD.get(
                                    region_raw.upper(), region_raw.upper()))
                                # Si pas reconnue → AUTRE
                                if region_norm.upper() not in {r.upper() for r in KNOWN_REGIONS}:
                                    region_norm = "AUTRE"
                            rows_agri.append({
                                "commercial":    CURRENT_NAME,
                                "nom":           str(row["NOM_AGRICULTEUR"]),
                                "region":        region_norm,
                                "zone":          str(row.get("ZONE","") or "").strip(),
                                "centre":        str(row.get("CENTRE","") or "").strip().upper() or None,
                                "usine":         str(row["USINE"]).strip().upper(),
                                "accessibilite": str(row["ACCESSIBILITE"]).strip(),
                                "tonnage_total": float(row["TONNAGE"]),
                                "date_debut":    str(row["DATE_DEBUT"])[:10] if pd.notna(row["DATE_DEBUT"]) else "2026-07-01",
                                "date_fin":      str(row["DATE_FIN"])[:10] if pd.notna(row["DATE_FIN"]) else "2026-08-17",
                                "nbr_hectares":  float(row["NBR_HECTARES"]) if "NBR_HECTARES" in row.index and pd.notna(row.get("NBR_HECTARES")) else None,
                            })

                        # Batch insert par 100
                        for i in range(0, len(rows_agri), 100):
                            sb.table("agriculteurs").insert(
                                rows_agri[i:i+100]).execute()

                        # 3. Mettre à jour depot_status
                        sb.table("depot_status").upsert({
                            "commercial":  CURRENT_NAME,
                            "statut":      "depose",
                            "nb_agri":     len(df_clean),
                            "tonnage":     float(df_clean["TONNAGE"].sum()),
                            "depose_le":   pd.Timestamp.now().isoformat(),
                            "fichier_nom": uploaded.name,
                        }).execute()

                        st.success(f"🎉 Planning déposé ! "
                                   f"{len(df_clean)} lignes, "
                                   f"{df_clean['TONNAGE'].sum():,.0f}t sauvegardés.")
                        st.balloons()
                        st.cache_data.clear()
                        if "sb_refresh" in st.session_state:
                            st.session_state["sb_refresh"] += 1
                        st.rerun()

            except Exception as e:
                st.error(f"Erreur lecture fichier: {e}")

    else:
        st.warning("🔒 Accès réservé aux commerciaux et au directeur.")