# ============================================================
# ONGLET COMPARAISON v5 — Plans rectifiés vs OR-Tools
# SOURCE UNIQUE DE VERITE :
#   priorite = upload session > Supabase > reference interne (reelle)
#   build_effective_planning() applique cette cascade sur le planning p
#   -> tous les tabs du dashboard qui lisent p en heritent automatiquement
# ============================================================
import streamlit as st
import pandas as pd
import plotly.graph_objects as go
import plotly.express as px
import io
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
try:
    import transport_calc as _tc
    TRANSPORT_CALC_AVAILABLE = True
except ImportError:
    TRANSPORT_CALC_AVAILABLE = False

SEASON  = pd.date_range("2026-06-20", "2026-08-25", freq="D")
PIC_S   = pd.Timestamp("2026-07-01").date()
PIC_E   = pd.Timestamp("2026-07-15").date()
DAYS_FR = ["Lun","Mar","Mer","Jeu","Ven","Sam","Dim"]

COMM_COLORS = {
    "FEDI":"#3b82f6","MAKKI BEN SALAH":"#00e5a0","KHALIL":"#f5a623",
    "ACHREF AJLANI":"#8b5cf6","JILANI OBAY":"#e8543a",
}
COMM_HEX = {
    "FEDI":"1F5FA6","MAKKI BEN SALAH":"1A6B3C","KHALIL":"B45309",
    "ACHREF AJLANI":"5B21B6","JILANI OBAY":"C0392B",
}
USINE_HEX = {
    "SICAM":"1F3864","TUCAL":"4A235A","COMOCAP":"0B4F6C","ELFALLEH":"196F3D","ABIDA":"922B21",
}
MANUAL_STATS = {
    "FEDI":           {"total":32736,"max_j":1216,"n_jours":55,"cap":1300},
    "MAKKI BEN SALAH":{"total":24310,"max_j":1315,"n_jours":44,"cap":1200},
    "KHALIL":         {"total":17455,"max_j":1060,"n_jours":46,"cap":1100},
    "ACHREF AJLANI":  {"total":17486,"max_j":570, "n_jours":61,"cap":700},
    "JILANI OBAY":    {"total":7000, "max_j":450, "n_jours":25,"cap":150},
}
USINE_STATS = {
    "SICAM":   {"total":44871,"max_j":1650,"n_jours":67,"cap_officiel":1300},
    "TUCAL":   {"total":20525,"max_j":700, "n_jours":52,"cap_officiel":700},
    "COMOCAP": {"total":20391,"max_j":875, "n_jours":50,"cap_officiel":700},
    "ELFALLEH":{"total":5190, "max_j":240, "n_jours":38,"cap_officiel":150},
    "ABIDA":   {"total":8010, "max_j":320, "n_jours":55,"cap_officiel":200},
}
USINE_CAPS      = {"SICAM":1300,"TUCAL":700,"COMOCAP":700,"ELFALLEH":150,"ABIDA":200}
USINE_CAPS_PLAN = {"SICAM":1300,"TUCAL":700,"COMOCAP":700,"ELFALLEH":240,"ABIDA":320}

# Profils usines reels (paries depuis PDFs/Excel reception_prevue 13/06/2026)
USINE_DAILY_PDF = {
    "SICAM":[0,0,140,170,220,220,260,320,320,415,515,625,560,910,1090,1170,1285,
             1425,1410,1540,1490,1440,1440,1455,1650,1600,1595,1550,1385,1345,1125,
             1000,935,835,840,575,456,660,750,730,660,760,535,615,580,610,520,470,
             530,500,380,320,290,240,200,140,170,140,210,180,150,150,150,120,120,120,120,120],
    "TUCAL":[0,0,0,0,15,15,15,85,85,105,230,235,285,465,560,590,590,625,610,700,
             670,690,645,660,620,655,585,665,645,595,490,500,405,405,395,345,365,475,
             490,410,430,380,380,350,350,330,315,310,340,350,300,270,200,120,90,90,
             0,0,0,0,0,0,0,0,0,0,0],
    "COMOCAP":[0,0,0,0,0,15,15,15,65,145,195,305,305,345,375,425,435,480,515,615,
               760,760,755,785,775,785,826,875,805,760,755,625,610,650,595,540,500,
               490,490,455,455,390,340,280,200,150,100,100,100,80,80,80,80,70,40,
               0,0,0,0,0,0,0,0,0,0,0],
    "ELFALLEH":[0,0,0,0,0,0,0,0,40,70,70,105,105,115,125,165,215,215,230,230,240,
                225,235,215,175,155,155,140,130,120,130,155,140,140,130,110,130,130,
                130,110,110,80,70,70,60,20,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
    "ABIDA":[0,0,0,0,0,0,0,0,30,80,80,120,140,140,140,170,170,200,200,200,240,270,
             250,230,290,320,190,180,190,220,220,210,110,110,110,110,110,70,70,20,20,
             80,140,170,170,110,90,60,60,60,60,60,60,60,120,210,270,300,270,210,60,
             90,90,0,0,0,0],
}

# Profils OR-Tools (comparaison uniquement — jamais affiches comme source)
ORT_COMM_BASE = {
    "FEDI":          {str((pd.Timestamp("2026-06-28")+pd.Timedelta(days=i)).date()):v
                     for i,v in enumerate([80,120,180,250,350,420,540,650,780,850,950,1050,
                     1100,1150,1175,1100,1050,900,750,600,500,450,400,350,300,250,200,180,
                     160,150,140,130,120,100,80,60,40,20,10]) if v>0},
    "MAKKI BEN SALAH":{str((pd.Timestamp("2026-06-22")+pd.Timedelta(days=i)).date()):v
                     for i,v in enumerate([50,100,150,200,280,380,480,580,650,750,850,950,
                     1050,1095,1000,900,780,680,580,480,380,300,250,200,160,130,100,80,60,
                     40,30,20,10]) if v>0},
    "KHALIL":        {str((pd.Timestamp("2026-06-20")+pd.Timedelta(days=i)).date()):v
                     for i,v in enumerate([200,280,350,430,520,630,720,820,920,1000,1020,980,
                     900,800,700,600,500,420,350,280,220,180,150,120,90,70,50,30,20,10]) if v>0},
    "ACHREF AJLANI": {str((pd.Timestamp("2026-06-22")+pd.Timedelta(days=i)).date()):v
                     for i,v in enumerate([30,60,120,180,240,300,360,420,480,540,600,650,700,
                     750,790,810,810,790,750,700,640,570,490,400,320,250,190,140,100,70,50,
                     30,20,10,5]) if v>0},
    "JILANI OBAY":   {str((pd.Timestamp("2026-07-23")+pd.Timedelta(days=i)).date()):v
                     for i,v in enumerate([80,130,180,240,310,380,440,440,410,370,320,270,220,
                     180,140,100,70,50,30,15]) if v>0},
}
ORT_USINE_BASE = {
    "SICAM":   {str((pd.Timestamp("2026-06-20")+pd.Timedelta(days=i)).date()):v
               for i,v in enumerate([100,120,180,250,250,420,440,500,560,653,743,898,895,1000,
               1105,1128,1213,1285,1340,1490,1530,1540,1510,1495,1440,1350,1280,1210,1140,
               1015,970,960,965,930,930,1155,1180,1220,1270,1280,1180,1090,1045,1040,940,860,
               850,835,795,578,600,540,520,380,260,210,150,90,30]) if v>0},
    "TUCAL":   {str((pd.Timestamp("2026-06-20")+pd.Timedelta(days=i)).date()):v
               for i,v in enumerate([20,60,80,80,110,195,195,245,315,320,325,330,365,390,420,
               415,395,380,350,370,340,300,290,320,300,230,210,205,185,255,240,240,140,160,235,
               315,345,380,410,410,385,385,375,370,330,370,350,340,270,260,270,255,250,230,190,
               190,90,100,90]) if v>0},
    "COMOCAP": {str((pd.Timestamp("2026-06-20")+pd.Timedelta(days=i)).date()):v
               for i,v in enumerate([0,0,0,0,0,0,0,0,20,55,55,60,135,145,160,165,165,165,170,
               400,520,500,505,585,600,515,558,545,580,542,520,421,380,463,343,293,275,245,290,
               306,293,250,247,197,160,118,45,40,15,15,8]) if v>0},
    "ELFALLEH":{}, "ABIDA":{},
}

# Courbes reelles parsees depuis reference_interne 13/06/2026 (fallback le plus bas
# de la cascade — utilisees uniquement si rien n'a ete uploade ni sauvegarde dans Supabase)
_REFERENCE_REAL_DAILY = {
    "FEDI": [0,0,0,0,0,40,40,80,180,230,220,340,420,470,540,580,590,700,720,920,1045,1010,1050,1130,1200,1170,1216,1150,1070,1120,1135,980,1010,970,970,950,890,800,900,820,770,700,650,610,500,410,370,340,300,270,200,160,160,150,150,110,80,60,60,30,0,0,0,0,0,0,0],
    "MAKKI BEN SALAH": [0,0,35,35,55,160,160,225,235,260,285,555,575,585,670,840,895,1030,1115,1095,1015,1085,1215,1315,1180,1230,1150,1185,970,865,695,655,565,425,285,280,280,230,200,165,160,115,100,80,35,20,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
    "KHALIL": [140,170,200,200,220,220,190,240,435,545,510,650,745,865,935,1020,995,1060,940,870,770,765,655,570,560,505,390,330,300,175,160,180,180,125,125,125,110,70,70,20,20,20,20,20,20,20,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
    "ACHREF AJLANI": [0,0,0,0,0,0,30,30,60,120,180,360,450,450,390,390,360,360,330,270,480,510,570,450,480,570,430,480,390,390,420,340,210,150,126,300,390,390,300,300,270,270,300,330,300,210,240,240,210,150,150,120,90,90,150,270,330,450,390,330,210,240,210,120,120,120,120],
    "JILANI OBAY": [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,60,60,60,130,150,150,260,310,340,310,420,280,400,410,410,370,380,450,430,330,310,310,270,220,180,0,0,0,0,0,0,0,0,0,0,0,0,0],
}

# ── openpyxl helpers ─────────────────────────────────────────
def _hf(h): return PatternFill("solid", start_color=h, end_color=h)
def _ft(bold=False, color="1F1F1F", size=10, white=False):
    return Font(bold=bold, name="Calibri", size=size, color="FFFFFF" if white else color)
def _tint_hex(hex_color, amount=0.8):
    """Eclaircit une couleur hex en la melangeant avec du blanc (amount=1 -> blanc pur)."""
    h = hex_color.lstrip("#")
    r, g, b = int(h[0:2],16), int(h[2:4],16), int(h[4:6],16)
    r = int(r + (255-r)*amount); g = int(g + (255-g)*amount); b = int(b + (255-b)*amount)
    return f"{r:02X}{g:02X}{b:02X}"
_CTR=Alignment(horizontal="center",vertical="center")
_LFT=Alignment(horizontal="left",vertical="center")
_THIN=Side(style="thin",color="CCCCCC")
_BORD=Border(left=_THIN,right=_THIN,top=_THIN,bottom=_THIN)
F_MAN=_hf("E2EFDA");F_ORT=_hf("DEEBF7");F_POS=_hf("C6EFCE");F_NEG=_hf("FFC7CE")
F_NEU=_hf("F2F2F2");F_PIC=_hf("FFF2CC");F_ALT1=_hf("FFFFFF");F_ALT2=_hf("F8F9FA")

# ═══ SUPABASE PERSISTENCE ════════════════════════════════════
def _save_rectifie(sb, entity_name, daily_dict, entity_type="commercial"):
    if sb is None: return False
    try:
        sb.table("plan_rectifie").delete().eq("entity_type",entity_type).eq("entity_name",entity_name).execute()
        rows=[{"entity_type":entity_type,"entity_name":entity_name,"date":str(d),"tonnes":float(t)}
              for d,t in daily_dict.items() if float(t)>0]
        if rows:
            for i in range(0,len(rows),500):
                sb.table("plan_rectifie").insert(rows[i:i+500]).execute()
        return True
    except Exception as e:
        st.warning(f"Supabase save: {e}"); return False

def _delete_rectifie(sb, entity_name, entity_type="commercial"):
    """Supprime les lignes Supabase d'une entite -> la cascade retombe sur la reference interne reelle.
    Utile quand Supabase contient une vieille valeur perimee (ex: ancien upload errone)."""
    if sb is None: return False
    try:
        sb.table("plan_rectifie").delete().eq("entity_type",entity_type).eq("entity_name",entity_name).execute()
        sb.table("plan_rectifie_detail").delete().eq("entity_name",entity_name).execute()
        return True
    except Exception as e:
        st.warning(f"Supabase delete: {e}"); return False

def _save_flotte_reelle(sb, real_fleet):
    """Sauvegarde la flotte reelle parsee (transport_etat_final.xlsx) dans Supabase,
    une ligne par camion (usine, type, capacite)."""
    if sb is None or not real_fleet: return False
    try:
        sb.table("flotte_reelle").delete().neq("usine", "__never__").execute()
        rows = []
        for usine, vtypes in real_fleet.items():
            for vtype, caps in vtypes.items():
                for cap in caps:
                    rows.append({"usine": usine, "type": vtype, "capacite": float(cap)})
        if rows:
            for i in range(0, len(rows), 500):
                sb.table("flotte_reelle").insert(rows[i:i+500]).execute()
        return True
    except Exception as e:
        st.warning(f"Supabase save flotte: {e}"); return False

def _load_flotte_reelle(sb):
    """Charge la flotte reelle depuis Supabase -> {usine: {type: [capacites triees desc]}}."""
    if sb is None: return {}
    try:
        data = sb.table("flotte_reelle").select("usine,type,capacite").execute().data
        if not data: return {}
        out = {}
        for row in data:
            out.setdefault(row["usine"], {}).setdefault(row["type"], []).append(float(row["capacite"]))
        return {u: {t: sorted(caps, reverse=True) for t, caps in vt.items()} for u, vt in out.items()}
    except Exception:
        return {}


def _save_rectifie_detail(sb, comm, df_raw):
    """Sauvegarde le detail GRANULAIRE (par agriculteur+date, valeurs EXACTES de l'upload)
    dans la table plan_rectifie_detail. C'est cette table qui permet a
    build_effective_planning de restituer les chiffres EXACTEMENT identiques
    a ceux uploades (au lieu d'un re-scaling proportionnel qui produisait des
    valeurs parasites comme 11/13/21 qui n'existent pas dans le fichier source)."""
    if sb is None or df_raw is None or df_raw.empty: return False
    try:
        sb.table("plan_rectifie_detail").delete().eq("entity_name", comm).execute()
        rows = [{"entity_name": comm, "agriculteur": str(r["agriculteur"]),
                 "date": str(pd.Timestamp(r["date"]).date()), "tonnes": float(r["tonnes"]),
                 "region": str(r.get("region","") or ""), "accessibilite": str(r.get("accessibilite","") or "")}
                for _, r in df_raw.iterrows() if float(r["tonnes"]) > 0]
        if rows:
            for i in range(0, len(rows), 500):
                sb.table("plan_rectifie_detail").insert(rows[i:i+500]).execute()
        return True
    except Exception as e:
        st.warning(f"Supabase save detail: {e}"); return False

def _load_all_rectifie_detail(sb):
    """Charge le detail granulaire de tous les commerciaux depuis Supabase.
    Avec PAGINATION (Supabase limite a 1000 lignes par requete, et le detail
    granulaire contient ~2800 lignes pour 5 commerciaux).
    Retourne {commercial: DataFrame(agriculteur,date,tonnes,region,accessibilite)}."""
    if sb is None: return {}
    try:
        data = []
        offset = 0
        page_size = 1000
        while True:
            batch = sb.table("plan_rectifie_detail").select(
                "entity_name,agriculteur,date,tonnes,region,accessibilite"
            ).range(offset, offset + page_size - 1).execute().data
            if not batch:
                break
            data.extend(batch)
            if len(batch) < page_size:
                break
            offset += page_size
        if not data: return {}
        out = {}
        for row in data:
            comm = row["entity_name"]
            out.setdefault(comm, []).append({
                "agriculteur": row["agriculteur"], "date": pd.Timestamp(row["date"]),
                "tonnes": float(row["tonnes"]), "region": row.get("region",""),
                "accessibilite": row.get("accessibilite",""),
            })
        return {c: pd.DataFrame(rows) for c, rows in out.items()}
    except Exception:
        return {}

def _load_all_rectifie(sb):
    if sb is None: return {},{}
    try:
        data=sb.table("plan_rectifie").select("entity_type,entity_name,date,tonnes").execute().data
        if not data: return {},{}
        cp={};up={}
        for row in data:
            et=row.get("entity_type","commercial"); en=row["entity_name"]
            d=str(row["date"]); t=float(row["tonnes"])
            if et=="commercial": cp.setdefault(en,{})[d]=t
            else: up.setdefault(en,{})[d]=t
        return cp,up
    except Exception: return {},{}

# ═══ HELPERS DONNEES — CASCADE: upload session > Supabase > reference reelle ═══
def _get_man_dict(comm, uploaded, sb_comm):
    if comm in uploaded:
        return {str(k):v for k,v in uploaded[comm].items()},"fichier uploade"
    if comm in sb_comm:
        return {str(k):v for k,v in sb_comm[comm].items()},"Supabase"
    return _build_reference_curve(comm),"reference interne (13/06/2026)"

def _get_usine_dict(usine, sb_usine):
    if usine in sb_usine:
        return {str(k):v for k,v in sb_usine[usine].items()},"Supabase"
    vals=USINE_DAILY_PDF.get(usine,[])
    d={str((pd.Timestamp("2026-06-20")+pd.Timedelta(days=i)).date()):v for i,v in enumerate(vals) if v>0}
    return d,"reference interne"

def _ortools_profile(comm_or_usine, planning_df, entity_type="commercial"):
    if planning_df is not None and not planning_df.empty:
        col="Commercial" if entity_type=="commercial" else "Usine"
        if col in planning_df.columns:
            sub=planning_df[planning_df[col]==comm_or_usine].copy()
            if not sub.empty and "Date" in sub.columns:
                sub["Date"]=pd.to_datetime(sub["Date"],errors="coerce")
                daily=sub.groupby("Date")["Tonnes/Jour"].sum().reset_index()
                r={str(r["Date"].date()):float(r["Tonnes/Jour"]) for _,r in daily.iterrows()}
                if r: return r
    if entity_type=="commercial":
        return {str(k):v for k,v in ORT_COMM_BASE.get(comm_or_usine,{}).items()}
    if comm_or_usine in ORT_USINE_BASE and ORT_USINE_BASE[comm_or_usine]:
        return {str(k):v for k,v in ORT_USINE_BASE[comm_or_usine].items()}
    vals=USINE_DAILY_PDF.get(comm_or_usine,[])
    return {str((pd.Timestamp("2026-06-20")+pd.Timedelta(days=i)).date()):v for i,v in enumerate(vals) if v>0}

def _build_reference_curve(comm):
    """Retourne la VRAIE courbe parsee depuis les fichiers reference_interne (pas une approximation)."""
    vals = _REFERENCE_REAL_DAILY.get(comm)
    if vals:
        return {str((pd.Timestamp("2026-06-20")+pd.Timedelta(days=i)).date()):float(v)
                for i,v in enumerate(vals) if v>0}
    stat=MANUAL_STATS.get(comm,{"n_jours":40,"max_j":500})
    n=stat["n_jours"];mx=stat["max_j"]
    starts={"FEDI":"2026-06-28","MAKKI BEN SALAH":"2026-06-22",
            "KHALIL":"2026-06-20","ACHREF AJLANI":"2026-06-20","JILANI OBAY":"2026-07-23"}
    start=pd.Timestamp(starts.get(comm,"2026-06-20"))
    p1=int(n*0.30);p2=int(n*0.50);p3=n-p1-p2;result={}
    for i in range(n):
        if i<p1:       v=mx*0.2+(mx*0.6)*(i/max(p1,1))
        elif i<p1+p2:  v=mx*0.8+(mx*0.2)*((i-p1)/max(p2,1))
        else:          v=mx*(1-((i-p1-p2)/max(p3,1))*0.8)
        result[str((start+pd.Timedelta(days=i)).date())]=round(max(0,v))
    return result

# ═══ HELPERS STREAMLIT ═══════════════════════════════════════
def _detect_comm_from_filename(filename):
    fn=filename.upper()
    for key,comm in {"FEDI":"FEDI","MEKKI":"MAKKI BEN SALAH","MAKKI":"MAKKI BEN SALAH",
                     "KHALIL":"KHALIL","ACHRAF":"ACHREF AJLANI","ACHREF":"ACHREF AJLANI",
                     "JILENI":"JILANI OBAY","JILANI":"JILANI OBAY"}.items():
        if key in fn: return comm
    return None

def _parse_rectification(uploaded_file):
    try: df=pd.read_excel(uploaded_file,header=0)
    except Exception as e: return None,None,str(e)
    header_row=0
    for i in range(min(5,len(df))):
        if any("agriculteur" in str(v).lower() for v in df.iloc[i].values):
            header_row=i;break
    if header_row>0:
        uploaded_file.seek(0);df=pd.read_excel(uploaded_file,header=header_row)
    cols=df.columns.tolist()
    col_comm=next((c for c in cols if "responsable" in str(c).lower() or "commercial" in str(c).lower()),cols[0] if cols else None)
    col_agri=next((c for c in cols if "agriculteur" in str(c).lower()),None)
    col_tonnage=next((c for c in cols if "tonnage" in str(c).lower()),None)
    def _norm_col(c): return str(c).strip().lower().replace("é","e")
    col_region=next((c for c in cols if _norm_col(c)=="region"),None)
    if col_region is None:
        col_region=next((c for c in cols if "region" in _norm_col(c) and "regional" not in _norm_col(c)),None)
    col_access=next((c for c in cols if "accessib" in str(c).lower()),None)
    parsed_dates=[]
    for c in cols[5:]:
        if isinstance(c,pd.Timestamp): parsed_dates.append((c,c));continue
        cs=str(c).strip()
        if "/" in cs and len(cs)<=5:
            try:
                p=cs.split("/");d=pd.Timestamp(f"2026-{int(p[1]):02d}-{int(p[0]):02d}")
                parsed_dates.append((c,d));continue
            except: pass
        try:
            d=pd.to_datetime(str(c).split(" ")[0],errors="coerce")
            if pd.notna(d) and d.year==2026: parsed_dates.append((c,d))
        except: pass
    if not parsed_dates or col_agri is None: return None,None,"Format non reconnu"
    rows=[];comm_detected=None
    for _,row in df.iterrows():
        comm=str(row.get(col_comm,"") or "").strip()
        agri=str(row.get(col_agri,"") or "").strip()
        if not agri or agri.upper() in ("NAN","","AGRICULTEUR","TOTAL"): continue
        if "sous-total" in agri.lower() or "total" in agri.lower(): continue
        if comm and comm.upper() not in ("NAN",""):
            for known in MANUAL_STATS:
                if known.upper() in comm.upper() or comm.upper() in known.upper():
                    comm_detected=known;break
        ton=pd.to_numeric(row.get(col_tonnage,0),errors="coerce") if col_tonnage else 0
        region=str(row.get(col_region,"") or "").strip() if col_region else ""
        access=str(row.get(col_access,"") or "").strip() if col_access else ""
        for oc,date in parsed_dates:
            val=pd.to_numeric(row.get(oc,0),errors="coerce")
            if pd.notna(val) and val>0:
                rows.append({"agriculteur":agri,"date":date,"tonnes":float(val),
                             "tonnage_total":float(ton) if pd.notna(ton) else 0,
                             "region":region,"accessibilite":access})
    if not rows: return comm_detected,None,"Aucune donnee trouvee"
    df_rows=pd.DataFrame(rows)
    daily=df_rows.groupby("date")["tonnes"].sum().reset_index().sort_values("date")
    daily_dict={str(r["date"].date()):float(r["tonnes"]) for _,r in daily.iterrows()}
    return comm_detected,daily_dict,df_rows

def _kpi_row(name,man_dict,ort_dict,cap):
    ms={str(k):v for k,v in man_dict.items()} if man_dict else {}
    os_={str(k):v for k,v in ort_dict.items()} if ort_dict else {}
    mt=sum(ms.values()) if ms else 0;mm=max(ms.values()) if ms else 0
    ot=sum(os_.values()) if os_ else 0;om=max(os_.values()) if os_ else 0
    k1,k2,k3,k4=st.columns(4)
    k1.metric("Total Rectifie",f"{int(mt)}t",delta=f"{int(mt-ot):+d}t vs OR-T",delta_color="off")
    k2.metric("Max/j Rectifie",f"{int(mm)}t",delta=f"OR-T max: {int(om)}t",delta_color="inverse" if mm>cap else "off")
    k3.metric("Jours actifs",f"{sum(1 for v in ms.values() if v>0)}j" if ms else "0j")
    k4.metric("Cap PIC",f"{cap}t/j",delta="OK" if mm<=cap else "DEPASSE",delta_color="normal" if mm<=cap else "inverse")

def _build_chart(name,man_dict,ort_dict,color,cap):
    ms={str(k):v for k,v in man_dict.items()} if man_dict else {}
    os_={str(k):v for k,v in ort_dict.items()} if ort_dict else {}
    all_d=sorted(set(list(ms.keys())+list(os_.keys())))
    dates=[pd.Timestamp(d) for d in all_d]
    mv=[ms.get(d,0) for d in all_d];ov=[os_.get(d,0) for d in all_d]
    fig=go.Figure()
    fig.add_trace(go.Scatter(x=dates,y=mv,name="Plan Rectifie (principal)",
        line=dict(color=color,width=2.5),fill="tozeroy",
        fillcolor="rgba(59,130,246,0.08)",mode="lines"))
    if any(v>0 for v in ov):
        fig.add_trace(go.Scatter(x=dates,y=ov,name="OR-Tools (comparaison)",
            line=dict(color="#888888",width=1.5,dash="dot"),mode="lines"))
    fig.add_hline(y=cap,line_dash="dash",line_color="#e8543a",line_width=1,
                  annotation_text=f"Cap: {cap}t/j",annotation_position="top right",
                  annotation_font_color="#e8543a")
    fig.add_vrect(x0=pd.Timestamp("2026-07-01"),x1=pd.Timestamp("2026-07-15"),
                  fillcolor="rgba(245,166,35,0.07)",line_width=0,
                  annotation_text="PIC",annotation_position="top left",annotation_font_color="#f5a623")
    fig.update_layout(title=f"{name} — Plan Rectifie (principal) vs OR-Tools (comparaison)",
        template="plotly_dark",plot_bgcolor="#0d1117",paper_bgcolor="#161b22",
        height=360,hovermode="closest",
        legend=dict(orientation="h",yanchor="bottom",y=1.02),
        xaxis=dict(title="Date",gridcolor="#21262d"),yaxis=dict(title="Tonnes/jour",gridcolor="#21262d"))
    return fig

# ═══ EXCEL ═══════════════════════════════════════════════════
def _build_entity_sheet(ws,name,man_dict,ort_dict,cap,hex_color):
    def _to_date_dict(d):
        out={}
        for k,v in d.items():
            try: out[pd.Timestamp(k).date()]=v
            except: pass
        return out
    man_d=_to_date_dict(man_dict);ort_d=_to_date_dict(ort_dict)
    HDR=_hf(hex_color);N=8
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=N)
    ws.cell(1,1).value=f"Comparaison Planning  —  {name}  —  Saison 2026"
    ws.cell(1,1).font=_ft(bold=True,size=12,white=True);ws.cell(1,1).fill=HDR;ws.cell(1,1).alignment=_CTR
    ws.row_dimensions[1].height=30
    ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=N)
    ws.cell(2,1).value=f"Vert=Rectifie>OR-T | Rouge=Rectifie<OR-T | Jaune=PIC | Cap={cap}t/j"
    ws.cell(2,1).font=_ft(size=9,color="595959");ws.cell(2,1).fill=_hf("F0F4F8");ws.cell(2,1).alignment=_LFT
    for ci,h in enumerate(["Date","Jour","Rectifie (t)","OR-Tools (t)","Ecart (t)","Ecart %","Statut",f"Cap {cap}t/j"],1):
        ws.cell(3,ci).value=h;ws.cell(3,ci).font=_ft(bold=True,size=9,white=True)
        ws.cell(3,ci).fill=HDR;ws.cell(3,ci).alignment=_CTR;ws.cell(3,ci).border=_BORD
    data_rows=[]
    for d in SEASON:
        dk=d.date();mv=man_d.get(dk,0);ov=ort_d.get(dk,0)
        if mv==0 and ov==0: continue
        data_rows.append((d,dk,mv,ov))
    for ri,(d,dk,mv,ov) in enumerate(data_rows):
        r=ri+4;is_pic=(PIC_S<=dk<=PIC_E);is_alt=ri%2==0
        base=F_PIC if is_pic else (F_ALT1 if is_alt else F_ALT2)
        ecart=mv-ov;pct=round(ecart/ov*100,1) if ov>0 else None
        ws.cell(r,1).value=d.strftime("%d/%m/%Y");ws.cell(r,1).fill=base
        ws.cell(r,1).border=_BORD;ws.cell(r,1).alignment=_CTR
        ws.cell(r,1).font=_ft(bold=is_pic,color="7D4F00" if is_pic else "1F1F1F")
        ws.cell(r,2).value=DAYS_FR[d.weekday()];ws.cell(r,2).fill=base
        ws.cell(r,2).border=_BORD;ws.cell(r,2).alignment=_CTR;ws.cell(r,2).font=_ft(size=9,color="595959")
        ws.cell(r,3).value=int(mv) if mv>0 else "";ws.cell(r,3).fill=F_MAN if mv>0 else base
        ws.cell(r,3).border=_BORD;ws.cell(r,3).alignment=_CTR;ws.cell(r,3).number_format="#,##0"
        ws.cell(r,4).value=int(ov) if ov>0 else "";ws.cell(r,4).fill=F_ORT if ov>0 else base
        ws.cell(r,4).border=_BORD;ws.cell(r,4).alignment=_CTR;ws.cell(r,4).number_format="#,##0"
        if mv>0 or ov>0:
            ws.cell(r,5).value=int(ecart)
            ws.cell(r,5).fill=F_POS if ecart>50 else (F_NEG if ecart<-50 else F_NEU)
            ws.cell(r,5).number_format="+#,##0;-#,##0;0"
        ws.cell(r,5).border=_BORD;ws.cell(r,5).alignment=_CTR
        if pct is not None:
            ws.cell(r,6).value=pct/100
            ws.cell(r,6).fill=F_POS if pct>5 else (F_NEG if pct<-5 else F_NEU)
            ws.cell(r,6).number_format="+0.0%;-0.0%;0%"
        else:
            ws.cell(r,6).value="—" if mv>0 else ""
        ws.cell(r,6).border=_BORD;ws.cell(r,6).alignment=_CTR
        if mv>0 and ov==0:   txt,fc,tc="Rect only","E8F5E9","2E7D32"
        elif mv==0 and ov>0: txt,fc,tc="OR-T only","FFF3E0","E65100"
        elif abs(ecart)<=50: txt,fc,tc="OK","E8F5E9","2E7D32"
        elif ecart>0:        txt,fc,tc="Rect +","E3F2FD","0D47A1"
        else:                txt,fc,tc="OR-T +","FCE4D6","993200"
        ws.cell(r,7).value=txt;ws.cell(r,7).fill=_hf(fc);ws.cell(r,7).font=_ft(size=9,color=tc)
        ws.cell(r,7).border=_BORD;ws.cell(r,7).alignment=_CTR
        if is_pic:
            peak=max(mv,ov)
            ws.cell(r,8).value=f"{'DEPASSE' if peak>cap else 'OK'} {int(peak)}t"
            ws.cell(r,8).fill=F_NEG if peak>cap else F_POS
        else:
            ws.cell(r,8).value="—";ws.cell(r,8).fill=base
        ws.cell(r,8).border=_BORD;ws.cell(r,8).alignment=_CTR
        ws.row_dimensions[r].height=17
    tr=len(data_rows)+4
    man_t=sum(r[2] for r in data_rows);ort_t=sum(r[3] for r in data_rows)
    ec_t=man_t-ort_t;pct_t=round(ec_t/ort_t*100,1) if ort_t>0 else 0
    man_m=max((r[2] for r in data_rows),default=0);ort_m=max((r[3] for r in data_rows),default=0)
    mj=sum(1 for r in data_rows if r[2]>0);oj=sum(1 for r in data_rows if r[3]>0)
    ws.merge_cells(start_row=tr,start_column=1,end_row=tr,end_column=2)
    for ci in range(1,N+1):
        ws.cell(tr,ci).fill=HDR;ws.cell(tr,ci).border=_BORD;ws.cell(tr,ci).alignment=_CTR
        ws.cell(tr,ci).font=_ft(bold=True,size=10,white=True)
    ws.cell(tr,1).value="TOTAL SAISON"
    ws.cell(tr,3).value=int(man_t);ws.cell(tr,3).number_format="#,##0"
    ws.cell(tr,4).value=int(ort_t);ws.cell(tr,4).number_format="#,##0"
    ws.cell(tr,5).value=int(ec_t);ws.cell(tr,5).number_format="+#,##0;-#,##0;0"
    ws.cell(tr,6).value=pct_t/100;ws.cell(tr,6).number_format="+0.0%;-0.0%;0%"
    ws.cell(tr,7).value=f"R:{mj}j|OT:{oj}j";ws.cell(tr,8).value=f"MaxR:{int(man_m)}t OT:{int(ort_m)}t"
    ws.row_dimensions[tr].height=24
    for ci,w in enumerate([12,6,13,13,12,10,12,16],1):
        ws.column_dimensions[get_column_letter(ci)].width=w
    ws.freeze_panes="A4"
    return len(data_rows)

def _build_synthese_sheet(ws,all_man_comm,all_ort_comm,all_man_usine,all_ort_usine):
    HDR=_hf("1F3864");N=11
    def ns(d): return {str(k):v for k,v in d.items()}
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=N)
    ws.cell(1,1).value="Synthese — Plan Rectifie vs OR-Tools — Saison 2026"
    ws.cell(1,1).font=_ft(bold=True,size=14,white=True);ws.cell(1,1).fill=HDR;ws.cell(1,1).alignment=_CTR
    ws.row_dimensions[1].height=30
    for ci,h in enumerate(["Entite","Type","Total Rect(t)","Total OR-T(t)","Ecart(t)",
                            "Ecart%","Max Rect(t)","Max OR-T(t)","Jours Rect","Jours OR-T","Cap PIC"],1):
        ws.cell(3,ci).value=h;ws.cell(3,ci).font=_ft(bold=True,size=9,white=True)
        ws.cell(3,ci).fill=HDR;ws.cell(3,ci).alignment=_CTR;ws.cell(3,ci).border=_BORD
    ws.row_dimensions[3].height=20
    row=4
    comm_order=["FEDI","MAKKI BEN SALAH","KHALIL","ACHREF AJLANI","JILANI OBAY"]
    usine_order=["SICAM","TUCAL","COMOCAP","ELFALLEH","ABIDA"]
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=N)
    ws.cell(row,1).value="COMMERCIAUX";ws.cell(row,1).font=_ft(bold=True,size=10,white=True)
    ws.cell(row,1).fill=_hf("2F4F7F");ws.cell(row,1).alignment=_LFT;ws.row_dimensions[row].height=16;row+=1
    for ci,comm in enumerate(comm_order):
        man=ns(all_man_comm.get(comm,{}));ort=ns(all_ort_comm.get(comm,{}))
        alld=sorted(set(list(man.keys())+list(ort.keys())))
        mt=sum(man.get(d,0) for d in alld);ot=sum(ort.get(d,0) for d in alld)
        ec=mt-ot;pct=round(ec/ot*100,1) if ot>0 else 0
        mm=max((man.get(d,0) for d in alld),default=0);om=max((ort.get(d,0) for d in alld),default=0)
        mj=sum(1 for d in alld if man.get(d,0)>0);oj=sum(1 for d in alld if ort.get(d,0)>0)
        hx=COMM_HEX.get(comm,"1F3864");cap=MANUAL_STATS.get(comm,{}).get("cap",0)
        for j,v in enumerate([comm,"Commercial",int(mt),int(ot),int(ec),f"{pct:+.1f}%",int(mm),int(om),mj,oj,cap],1):
            ws.cell(row,j).value=v;ws.cell(row,j).border=_BORD;ws.cell(row,j).alignment=_LFT if j==1 else _CTR
            ws.cell(row,j).font=_ft(bold=(j==1),color="FFFFFF" if j==1 else "1F1F1F")
            ws.cell(row,j).fill=_hf(hx) if j==1 else (F_ALT1 if ci%2==0 else F_ALT2)
            if j==5: ws.cell(row,j).fill=F_POS if ec>=0 else F_NEG
        ws.row_dimensions[row].height=17;row+=1
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=N)
    ws.cell(row,1).value="USINES";ws.cell(row,1).font=_ft(bold=True,size=10,white=True)
    ws.cell(row,1).fill=_hf("0B4F6C");ws.cell(row,1).alignment=_LFT;ws.row_dimensions[row].height=16;row+=1
    for ui,usine in enumerate(usine_order):
        man=ns(all_man_usine.get(usine,{}));ort=ns(all_ort_usine.get(usine,{}))
        alld=sorted(set(list(man.keys())+list(ort.keys())))
        mt=sum(man.get(d,0) for d in alld);ot=sum(ort.get(d,0) for d in alld)
        ec=mt-ot;pct=round(ec/ot*100,1) if ot>0 else 0
        mm=max((man.get(d,0) for d in alld),default=0);om=max((ort.get(d,0) for d in alld),default=0)
        mj=sum(1 for d in alld if man.get(d,0)>0);oj=sum(1 for d in alld if ort.get(d,0)>0)
        hx=USINE_HEX.get(usine,"1F3864");cap=USINE_CAPS.get(usine,0)
        for j,v in enumerate([usine,"Usine",int(mt),int(ot),int(ec),f"{pct:+.1f}%",int(mm),int(om),mj,oj,cap],1):
            ws.cell(row,j).value=v;ws.cell(row,j).border=_BORD;ws.cell(row,j).alignment=_LFT if j==1 else _CTR
            ws.cell(row,j).font=_ft(bold=(j==1),color="FFFFFF" if j==1 else "1F1F1F")
            ws.cell(row,j).fill=_hf(hx) if j==1 else (F_ALT1 if ui%2==0 else F_ALT2)
            if j==5: ws.cell(row,j).fill=F_POS if ec>=0 else F_NEG
        ws.row_dimensions[row].height=17;row+=1

def _build_legende_sheet(ws):
    HDR=_hf("1F3864")
    ws.merge_cells("A1:E1");ws.cell(1,1).value="Legende — Codes couleur"
    ws.cell(1,1).font=_ft(bold=True,size=13,white=True);ws.cell(1,1).fill=HDR;ws.cell(1,1).alignment=_CTR
    for ci,h in enumerate(["Couleur","Code Hex","Signification","Usage"],1):
        ws.cell(3,ci).value=h;ws.cell(3,ci).font=_ft(bold=True,size=10,white=True)
        ws.cell(3,ci).fill=HDR;ws.cell(3,ci).alignment=_CTR;ws.cell(3,ci).border=_BORD
    items=[("E2EFDA","Vert clair","Plan Rectifie (source principale)","Colonne Rectifie"),
           ("DEEBF7","Bleu clair","OR-Tools (comparaison)","Colonne OR-Tools"),
           ("C6EFCE","Vert moyen","Rectifie>OR-T de +50t","Ecart positif"),
           ("FFC7CE","Rose","Rectifie<OR-T de +50t","Ecart negatif"),
           ("FFF2CC","Jaune","Periode PIC 1-15 juillet","Caps actifs")]
    for ri,(hex_c,nom,desc,usage) in enumerate(items,4):
        ws.cell(ri,1).fill=_hf(hex_c);ws.cell(ri,1).border=_BORD;ws.cell(ri,1).value=nom
        ws.cell(ri,2).value=f"#{hex_c}";ws.cell(ri,2).fill=_hf(hex_c);ws.cell(ri,2).border=_BORD
        ws.cell(ri,3).value=desc;ws.cell(ri,3).border=_BORD
        ws.cell(ri,4).value=usage;ws.cell(ri,4).border=_BORD

def generate_comparison_excel(man_comm,ort_comm,man_usine,ort_usine):
    wb=Workbook();wb.remove(wb.active)
    comm_order=["FEDI","MAKKI BEN SALAH","KHALIL","ACHREF AJLANI","JILANI OBAY"]
    usine_order=["SICAM","TUCAL","COMOCAP","ELFALLEH","ABIDA"]
    ws=wb.create_sheet("Synthese Globale");_build_synthese_sheet(ws,man_comm,ort_comm,man_usine,ort_usine)
    for comm in comm_order:
        ws=wb.create_sheet(f"C - {comm[:14]}")
        _build_entity_sheet(ws,comm,man_comm.get(comm,{}),ort_comm.get(comm,{}),
                            MANUAL_STATS.get(comm,{}).get("cap",800),COMM_HEX.get(comm,"1F3864"))
    for usine in usine_order:
        ws=wb.create_sheet(f"U - {usine}")
        _build_entity_sheet(ws,usine,man_usine.get(usine,{}),ort_usine.get(usine,{}),
                            USINE_CAPS.get(usine,500),USINE_HEX.get(usine,"1F3864"))
    ws=wb.create_sheet("Legende");_build_legende_sheet(ws)
    buf=io.BytesIO();wb.save(buf);buf.seek(0);return buf.read()

# ═══════════════════════════════════════════════════════════════
# build_effective_planning — SOURCE UNIQUE DE VERITE
# Corrige le planning OR-Tools `p` avec les Plans Rectifies (cascade:
# upload session > Supabase > reference interne reelle). Tous les tabs
# du dashboard qui lisent `p` heritent automatiquement de la correction.
# ═══════════════════════════════════════════════════════════════
def build_effective_planning(p, sb=None):
    """
    p: DataFrame planning (colonnes Date, Commercial, Usine, Tonnes/Jour, ...)
    sb: client Supabase (ou None)
    Retourne une COPIE de p avec Tonnes/Jour corrige selon la cascade des
    Plans Rectifies.

      1) Par commercial — SUBSTITUTION EXACTE quand un detail granulaire
         (par agriculteur+date) est disponible (upload session > Supabase).
         Les lignes OR-Tools de ce commercial sont remplacees par les
         valeurs EXACTES de l'upload (aucun scaling, aucun arrondi —
         garantit que chaque chiffre journalier affiche est identique au
         fichier source, sans valeur parasite type 11/13/21). L'Usine de
         chaque agriculteur est deduite de l'assignation OR-Tools la plus
         frequente pour cet agriculteur (a cette date si connue, sinon sa
         destination dominante toutes dates confondues).
         Si aucun detail granulaire n'existe (cas rare : rien uploade
         encore pour ce commercial), on retombe sur l'ancien scaling
         proportionnel a partir du total journalier (upload > Supabase >
         reference interne).
      2) Par usine — UNIQUEMENT si une correction explicite existe dans
         Supabase pour cette usine (override intentionnel de l'admin).
         Jamais via le fallback de reference statique (sinon il ecraserait
         silencieusement la Passe 1).
    """
    if p is None or p.empty or "Commercial" not in p.columns or "Date" not in p.columns:
        return p

    p = p.copy()
    p["Date"] = pd.to_datetime(p["Date"], errors="coerce")
    p = p.dropna(subset=["Date"])
    if "Tonnes/Jour" not in p.columns:
        return p
    p["Tonnes/Jour"] = pd.to_numeric(p["Tonnes/Jour"], errors="coerce").fillna(0)

    uploaded_sess = st.session_state.get("comp_uploaded", {})
    uploaded_raw  = st.session_state.get("comp_raw", {})
    sb_comm = st.session_state.get("_sb_comm")
    sb_usine = st.session_state.get("_sb_usine")
    if sb_comm is None or sb_usine is None:
        sb_comm, sb_usine = _load_all_rectifie(sb)
        st.session_state["_sb_comm"] = sb_comm
        st.session_state["_sb_usine"] = sb_usine
    sb_detail = st.session_state.get("_sb_detail")
    if sb_detail is None or not st.session_state.get("comp_raw"):
        sb_detail = _load_all_rectifie_detail(sb)
        st.session_state["_sb_detail"] = sb_detail

    def _get_detail_df(comm):
        """Detail granulaire (agriculteur,date,tonnes,region,accessibilite) —
        priorite upload session (le plus recent) > Supabase."""
        d = uploaded_raw.get(comm)
        if isinstance(d, pd.DataFrame) and not d.empty:
            return d
        d = sb_detail.get(comm)
        if isinstance(d, pd.DataFrame) and not d.empty:
            return d
        return None

    def _build_agri_usine_map(comm):
        """Deduit l'usine de chaque agriculteur depuis l'assignation OR-Tools
        d'origine: mapping precis (agriculteur,date) si connu, sinon usine
        dominante de cet agriculteur toutes dates confondues."""
        sub = p[p["Commercial"] == comm]
        if sub.empty or "Usine" not in sub.columns:
            return {}, {}
        s = sub.copy()
        s["_k"] = s["Agriculteur"].astype(str).str.upper().str.strip()
        s["_d"] = s["Date"].dt.date
        g = s.groupby(["_k", "_d", "Usine"])["Tonnes/Jour"].sum().reset_index()
        g = g.sort_values("Tonnes/Jour", ascending=False).drop_duplicates(["_k", "_d"])
        exact_map = {(row["_k"], row["_d"]): row["Usine"] for _, row in g.iterrows()}
        g2 = s.groupby(["_k", "Usine"])["Tonnes/Jour"].sum().reset_index()
        g2 = g2.sort_values("Tonnes/Jour", ascending=False).drop_duplicates("_k")
        global_map = dict(zip(g2["_k"], g2["Usine"]))
        return exact_map, global_map

    EXTRA_COLS = ["Région","Accessibilité","Type Véhicule","Véhicules Requis",
                  "Disponibles","Manquants (à louer)","Nb Voyages","Pic de Récolte"]
    NUMERIC_EXTRA = {"Nb Voyages", "Disponibles", "Manquants (à louer)"}
    for c in EXTRA_COLS:
        if c not in p.columns:
            p[c] = 0 if c in NUMERIC_EXTRA else ""

    comms_with_detail = []
    new_rows_by_comm = {}
    for comm in p["Commercial"].dropna().unique():
        detail_df = _get_detail_df(comm)
        if detail_df is None:
            continue
        comms_with_detail.append(comm)
        exact_map, global_map = _build_agri_usine_map(comm)
        comm_sub = p[p["Commercial"] == comm]
        comm_default_usine = ""
        if "Usine" in comm_sub.columns and not comm_sub.empty:
            _m = comm_sub.groupby("Usine")["Tonnes/Jour"].sum().sort_values(ascending=False)
            comm_default_usine = _m.index[0] if not _m.empty else ""
        recs = []
        for _, r in detail_df.iterrows():
            agri = str(r["agriculteur"]).strip()
            d = pd.Timestamp(r["date"])
            key = agri.upper().strip()
            usine = (exact_map.get((key, d.date())) or global_map.get(key)
                     or comm_default_usine or "(usine non assignee)")
            recs.append({
                "Date": d, "Commercial": comm, "Agriculteur": agri, "Usine": usine,
                "Tonnes/Jour": float(r["tonnes"]),
                "Région": r.get("region","") or "", "Accessibilité": r.get("accessibilite","") or "",
                "Type Véhicule": "", "Véhicules Requis": "", "Disponibles": 0,
                "Manquants (à louer)": 0, "Nb Voyages": 0,
                "Pic de Récolte": "🟡 PIC" if PIC_S <= d.date() <= PIC_E else "",
            })
        new_rows_by_comm[comm] = recs

    if comms_with_detail:
        p = p[~p["Commercial"].isin(comms_with_detail)].copy()
        all_new = [rec for recs in new_rows_by_comm.values() for rec in recs]
        if all_new:
            df_new = pd.DataFrame(all_new)
            if TRANSPORT_CALC_AVAILABLE:
                real_fleet = st.session_state.get("_real_fleet", {})
                for idx, row in df_new.iterrows():
                    tons = float(row.get("Tonnes/Jour", 0) or 0)
                    if tons <= 0:
                        continue
                    acc_raw = row.get("Accessibilité", "")
                    acc = _tc.normalize_acc(acc_raw)
                    allowed = _tc.ACCESS_VEHICLES.get(acc, _tc.ACCESS_VEHICLES["NAN"])
                    usine = str(row.get("Usine", "") or "")
                    region = str(row.get("Région", "") or "")
                    vehicles = _tc.choose_vehicles(tons, allowed, usine=usine,
                                                    region=region, real_fleet=real_fleet)
                    if vehicles:
                        veh_type = vehicles[0].get("vehicle", "PL")
                        parts = []
                        for v in vehicles:
                            load = v.get("tons_each", 0)
                            n = v.get("trips", 1)
                            disp = int(load) if float(load) == int(float(load)) else round(float(load), 1)
                            parts.append(f"{v['vehicle']}({disp}t)" if n == 1 else f"{v['vehicle']} x{n}({disp}t)")
                        trips = sum(v.get("trips", 1) for v in vehicles)
                        df_new.at[idx, "Type Véhicule"] = veh_type
                        df_new.at[idx, "Véhicules Requis"] = " | ".join(parts)
                        df_new.at[idx, "Nb Voyages"] = trips
            p = pd.concat([p, df_new], ignore_index=True)

    # ── Commerciaux SANS detail granulaire (rien uploade encore) : ancien
    # comportement de scaling proportionnel a partir du total journalier,
    # pour ne jamais laisser un commercial non corrige. ─────────────────
    extra_rows = []

    def _apply_correction(mask, group_df, rect_dict, default_col, default_val):
        if not rect_dict:
            return
        rect_dict = {str(k): float(v) for k, v in rect_dict.items()}
        cur_daily = group_df.groupby(group_df["Date"].dt.date)["Tonnes/Jour"].sum()
        all_dates = set(rect_dict.keys()) | {str(d) for d in cur_daily.index}
        for dstr in all_dates:
            try:
                d = pd.Timestamp(dstr).date()
            except Exception:
                continue
            target = rect_dict.get(dstr, 0.0)
            base = float(cur_daily.get(d, 0.0))
            row_mask = mask & (p["Date"].dt.date == d)
            n_rows = int(row_mask.sum())
            if n_rows > 0 and base > 0:
                ratio = target / base
                p.loc[row_mask, "Tonnes/Jour"] = p.loc[row_mask, "Tonnes/Jour"] * ratio
            elif n_rows > 0 and base == 0 and target > 0:
                p.loc[row_mask, "Tonnes/Jour"] = target / n_rows
            elif n_rows == 0 and target > 0:
                row = {
                    "Date": pd.Timestamp(d), "Commercial": "", "Agriculteur": "(ajustement plan rectifie)",
                    "Usine": "", "Tonnes/Jour": target, "Région": "", "Accessibilité": "",
                    "Type Véhicule": "", "Véhicules Requis": 0, "Nb Voyages": 0,
                    "Pic de Récolte": "🟡 PIC" if PIC_S <= d <= PIC_E else "", "Note": "ajustement rectifie",
                }
                row[default_col] = default_val
                extra_rows.append(row)

    correction_mode = {c: "exact (detail granulaire)" for c in comms_with_detail}
    for comm in p["Commercial"].dropna().unique():
        if comm in comms_with_detail:
            continue
        rect_dict, _src = _get_man_dict(comm, uploaded_sess, sb_comm)
        correction_mode[comm] = (f"scaling proportionnel (pas de detail exact — source: {_src})"
                                  if rect_dict else "aucune correction (OR-Tools brut)")
        mask = p["Commercial"] == comm
        sub = p[mask]
        usine_default = ""
        if "Usine" in sub.columns and not sub.empty:
            m = sub["Usine"].mode()
            usine_default = m.iloc[0] if not m.empty else ""
        _apply_correction(mask, sub, rect_dict, "Usine", usine_default)
        for r in extra_rows:
            if r["Commercial"] == "":
                r["Commercial"] = comm

    # ── Passe 2 : correction par USINE — UNIQUEMENT si override
    # explicite Supabase (jamais via le fallback reference statique
    # USINE_DAILY_PDF, qui sinon ecrase silencieusement la Passe 1
    # par commercial et fait deriver les totaux uploades) ──────────
    if "Usine" in p.columns:
        for usine in p["Usine"].dropna().unique():
            if not usine or usine not in sb_usine:
                continue
            rect_u = {str(k): float(v) for k, v in sb_usine[usine].items()}
            mask_u = p["Usine"] == usine
            sub_u = p[mask_u]
            comm_default = ""
            if "Commercial" in sub_u.columns and not sub_u.empty:
                m = sub_u["Commercial"].mode()
                comm_default = m.iloc[0] if not m.empty else ""
            _apply_correction(mask_u, sub_u, rect_u, "Commercial", comm_default)

    if extra_rows:
        for cols_needed in ["Région","Accessibilité","Type Véhicule","Véhicules Requis",
                             "Nb Voyages","Pic de Récolte","Note"]:
            if cols_needed not in p.columns:
                p[cols_needed] = ""
        p = pd.concat([p, pd.DataFrame(extra_rows)], ignore_index=True)

    st.session_state["_correction_mode"] = correction_mode
    for _nc in ["Tonnes/Jour", "Nb Voyages", "Disponibles", "Manquants (à louer)"]:
        if _nc in p.columns:
            p[_nc] = pd.to_numeric(p[_nc], errors="coerce").fillna(0).astype(float)
    st.session_state["_p_full_with_acc"] = p.copy()
    return p


def _round_preserving_total(df, group_cols=("Commercial", "Date")):
    """
    Arrondit la colonne 'Tonnes/Jour' a l'entier (methode des plus grands
    restes) en garantissant que la somme par group_cols (typiquement par
    Commercial+Date) reste EXACTEMENT egale a l'entier le plus proche du
    total non arrondi. Necessaire car arrondir chaque ligne separement
    (comme avant) fait deriver le total importe de +/-1 a 2 tonnes selon
    le nombre de lignes, ce qui n'est pas acceptable: les totaux exportes
    doivent etre identiques au tonnage uploade, au tonnage pres.
    """
    import math
    vals = df["Tonnes/Jour"].astype(float)
    out = pd.Series(0, index=df.index, dtype=int)
    for _, idx in df.groupby(list(group_cols)).groups.items():
        v = vals.loc[idx]
        target = int(round(v.sum()))
        floors = v.apply(math.floor).astype(int)
        remainder = target - int(floors.sum())
        fracs = (v - floors).sort_values(ascending=False)
        result = floors.copy()
        if remainder > 0:
            for i in fracs.index[:remainder]:
                result[i] += 1
        elif remainder < 0:
            for i in fracs.index[::-1][:abs(remainder)]:
                result[i] = max(0, result[i] - 1)
        out.loc[idx] = result
    return out


def _write_transport_period_sheet(wb, sheet_name, title, df_usine, df_detail, period_label):
    """Construit une feuille Transport (Jour ou Semaine) avec 2 tables:
    A) Disponibilite flotte par usine (Date/Semaine | Type Vehicule | Tonnes
       requises | Voyages requis | Camions disponibles | Manquants a louer)
    B) Detail par commercial et region (qui a besoin de quoi)."""
    ws = wb.create_sheet(sheet_name)
    HDR = _hf("1F3864")
    REDF = _hf("FFC7CE"); GRNF = _hf("E2EFDA")

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    ws.cell(1, 1).value = title
    ws.cell(1, 1).font = _ft(bold=True, size=12, white=True)
    ws.cell(1, 1).fill = HDR; ws.cell(1, 1).alignment = _CTR
    ws.row_dimensions[1].height = 26

    row = 3
    ws.cell(row, 1).value = "A) Disponibilité de la flotte par usine"
    ws.cell(row, 1).font = _ft(bold=True, size=11)
    row += 1
    HDRS_A = [period_label, "Type Véhicule", "Tonnes requises", "Voyages requis",
              "Camions disponibles", "Camions manquants (à louer)", "Tonnage manquant (à louer)"]
    for ci, h in enumerate(HDRS_A, 1):
        c = ws.cell(row, ci); c.value = h; c.fill = HDR; c.font = _ft(bold=True, size=9, white=True)
        c.alignment = _CTR; c.border = _BORD
    row += 1

    if df_usine is not None and not df_usine.empty:
        usine_order = list(USINE_HEX.keys())
        usines = sorted(df_usine["Usine"].dropna().unique(),
                         key=lambda u: (usine_order.index(u) if u in usine_order else len(usine_order), str(u)))
        for usine in usines:
            usine_hex = USINE_HEX.get(usine, "404040")
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(HDRS_A))
            ws.cell(row, 1).value = f"📍 {usine}"
            for ci in range(1, len(HDRS_A)+1):
                ws.cell(row, ci).fill = _hf(usine_hex); ws.cell(row, ci).font = _ft(bold=True, size=10, white=True)
                ws.cell(row, ci).border = _BORD; ws.cell(row, ci).alignment = _CTR if ci > 1 else _LFT
            row += 1
            start_data = row
            sub = df_usine[df_usine["Usine"] == usine].sort_values(["Periode", "Type Véhicule"])
            for _, r in sub.iterrows():
                manquant = int(r["Camions manquants (a louer)"])
                vals = [r["Periode"].strftime("%d/%m/%Y"), r["Type Véhicule"], round(r["Tonnes requises"], 1),
                        int(r["Voyages requis"]), int(r["Camions disponibles"]), manquant,
                        round(r["Tonnage manquant (a louer)"], 1)]
                for ci, v in enumerate(vals, 1):
                    c = ws.cell(row, ci); c.value = v; c.border = _BORD; c.alignment = _CTR
                    if ci == 6:
                        c.fill = REDF if manquant > 0 else GRNF
                row += 1
            end_data = row - 1
            if end_data >= start_data:
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
                ws.cell(row, 1).value = f"Sous-total {usine}"
                ws.cell(row, 1).font = Font(bold=True, italic=True, name="Calibri", size=9)
                for ci in [3, 4, 6, 7]:
                    col_l = get_column_letter(ci)
                    ws.cell(row, ci).value = f"=SUM({col_l}{start_data}:{col_l}{end_data})"
                    ws.cell(row, ci).font = _ft(bold=True, size=9)
                for ci in range(1, len(HDRS_A)+1):
                    ws.cell(row, ci).border = _BORD
                    ws.cell(row, ci).fill = _hf(_tint_hex(usine_hex, 0.85))
                row += 1
            row += 1

    row += 1
    ws.cell(row, 1).value = "B) Détail par commercial et région"
    ws.cell(row, 1).font = _ft(bold=True, size=11)
    row += 1
    HDRS_B = [period_label, "Usine", "Commercial", "Région", "Type Véhicule", "Tonnes requises", "Voyages requis"]
    for ci, h in enumerate(HDRS_B, 1):
        c = ws.cell(row, ci); c.value = h; c.fill = HDR; c.font = _ft(bold=True, size=9, white=True)
        c.alignment = _CTR; c.border = _BORD
    row += 1

    if df_detail is not None and not df_detail.empty:
        comm_order = list(COMM_HEX.keys())
        sub = df_detail.sort_values(["Usine", "Commercial", "Periode", "Région", "Type Véhicule"])
        for _, r in sub.iterrows():
            comm_hex = COMM_HEX.get(r["Commercial"], "595959")
            vals = [r["Periode"].strftime("%d/%m/%Y"), r["Usine"], r["Commercial"], r["Région"],
                    r["Type Véhicule"], round(r["Tonnes requises"], 1), int(r["Voyages requis"])]
            for ci, v in enumerate(vals, 1):
                c = ws.cell(row, ci); c.value = v; c.border = _BORD; c.alignment = _CTR
                c.fill = _hf(_tint_hex(comm_hex, 0.88))
            row += 1

    for ci, w in enumerate([13, 14, 16, 14, 17, 22, 17], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A5"
    return ws


def _write_transport_semaine_detaillee(wb, detail, real_fleet):
    """
    Feuille 'Transport Semaine' detaillee, organisee par Commercial -> Semaine
    -> Jour -> Usine -> Agriculteur, avec colonnes PPL/PL/SEMI requis,
    disponibles (flotte de l'usine ce jour-la) et manquants (a louer).
    Une ligne 'TOTAL SEMAINE <Usine>' par usine livree, a la fin de chaque
    semaine, pour ce commercial.
    """
    real_fleet = real_fleet or {}
    ws = wb.create_sheet("Transport Semaine")
    VEH_COLS = ["TRACTEUR", "PPL", "PL", "SEMI"]
    HDRS = ["Date", "Agriculteur", "Usine",
            "TRACTEUR requis", "PPL requis", "PL requis", "SEMI requis",
            "TRACTEUR disponible", "PPL disponible", "PL disponible", "SEMI disponible",
            "TRACTEUR manquant (à louer)", "PPL manquant (à louer)", "PL manquant (à louer)", "SEMI manquant (à louer)"]
    N = len(HDRS)
    HDR = _hf("1F3864"); REDF = _hf("FFC7CE"); GRNF = _hf("E2EFDA")

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=N)
    ws.cell(1, 1).value = "Besoins Transport par Semaine — detail par Commercial (calcule sur le Plan Rectifie)"
    ws.cell(1, 1).font = _ft(bold=True, size=12, white=True)
    ws.cell(1, 1).fill = HDR; ws.cell(1, 1).alignment = _CTR
    ws.row_dimensions[1].height = 26
    row = 3

    if detail is None or detail.empty:
        ws.cell(row, 1).value = "Aucune donnee."
        return wb
    detail = detail.copy()
    detail["Date"] = pd.to_datetime(detail["Date"], errors="coerce")
    detail = detail.dropna(subset=["Date"])
    detail = detail[detail["Type Véhicule"].isin(VEH_COLS)]

    # Disponibilite par (jour, usine, type) — globale tous commerciaux confondus
    # (la flotte d'une usine est partagee, jamais sous-allouee par commercial)
    usine_day = _tc.summarize_transport_usine(detail, real_fleet=real_fleet, period="day")
    avail_lookup = {}
    for _, r in usine_day.iterrows():
        avail_lookup[(r["Periode"].date(), r["Usine"], r["Type Véhicule"])] = (
            int(r["Camions disponibles"]), int(r["Camions manquants (a louer)"]))

    def _fleet_count(usine, veh):
        return len(real_fleet.get(usine, {}).get(veh, []))

    comm_order = list(COMM_HEX.keys())
    commercials = sorted(detail["Commercial"].dropna().unique(),
                          key=lambda c: (comm_order.index(c) if c in comm_order else len(comm_order), str(c)))

    for comm in commercials:
        comm_hex = COMM_HEX.get(comm, "404040")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=N)
        ws.cell(row, 1).value = f"📍 {comm}"
        for ci in range(1, N + 1):
            ws.cell(row, ci).fill = _hf(comm_hex); ws.cell(row, ci).font = _ft(bold=True, size=12, white=True)
            ws.cell(row, ci).border = _BORD
        ws.cell(row, 1).alignment = _LFT
        ws.row_dimensions[row].height = 22
        row += 1

        sub_comm = detail[detail["Commercial"] == comm]
        dates_comm = sorted(sub_comm["Date"].dt.date.unique())
        if not dates_comm:
            continue

        weeks = []
        cur = dates_comm[0]
        while cur <= dates_comm[-1]:
            wk_end = cur + pd.Timedelta(days=6)
            weeks.append((cur, wk_end))
            cur = wk_end + pd.Timedelta(days=1)

        for wk_start, wk_end in weeks:
            week_dates = [d for d in dates_comm if wk_start <= d <= wk_end]
            if not week_dates:
                continue
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=N)
            ws.cell(row, 1).value = f"Semaine du {wk_start.strftime('%d/%m')} au {wk_end.strftime('%d/%m/%Y')}"
            for ci in range(1, N + 1):
                ws.cell(row, ci).fill = _hf(_tint_hex(comm_hex, 0.55)); ws.cell(row, ci).font = _ft(bold=True, size=10, white=True)
                ws.cell(row, ci).border = _BORD
            ws.cell(row, 1).alignment = _LFT
            row += 1
            for ci, h in enumerate(HDRS, 1):
                c = ws.cell(row, ci); c.value = h; c.fill = _hf(_tint_hex(comm_hex, 0.8))
                c.font = _ft(bold=True, size=9); c.alignment = _CTR; c.border = _BORD
            row += 1

            week_totals = defaultdict(lambda: defaultdict(int))   # [usine][veh] -> voyages requis semaine
            week_manq   = defaultdict(lambda: defaultdict(int))   # [usine][veh] -> somme manquants quotidiens semaine

            for d in week_dates:
                day_sub = sub_comm[sub_comm["Date"].dt.date == d]
                usines_today = sorted(day_sub["Usine"].dropna().unique())
                for usine in usines_today:
                    us_sub = day_sub[day_sub["Usine"] == usine]
                    agris = sorted(us_sub["Agriculteur"].dropna().unique())
                    day_usine_tot = {v: 0 for v in VEH_COLS}
                    for agri in agris:
                        ag_sub = us_sub[us_sub["Agriculteur"] == agri]
                        counts = {v: int(ag_sub[ag_sub["Type Véhicule"] == v]["Voyages"].sum()) for v in VEH_COLS}
                        if sum(counts.values()) == 0:
                            continue
                        vals = [d.strftime("%d/%m/%Y"), agri, usine] + \
                                [counts[v] or "" for v in VEH_COLS] + \
                                [""] * (len(VEH_COLS) * 2)
                        for ci, v in enumerate(vals, 1):
                            c = ws.cell(row, ci); c.value = v; c.border = _BORD; c.alignment = _CTR
                            if ci == 2: c.alignment = _LFT
                        row += 1
                        for v in VEH_COLS:
                            day_usine_tot[v] += counts[v]
                            week_totals[usine][v] += counts[v]
                    disp = {v: avail_lookup.get((d, usine, v), (_fleet_count(usine, v), 0))[0] for v in VEH_COLS}
                    manq = {v: avail_lookup.get((d, usine, v), (_fleet_count(usine, v), 0))[1] for v in VEH_COLS}
                    for v in VEH_COLS:
                        week_manq[usine][v] += manq[v]
                    manq_col_start = 3 + len(VEH_COLS) * 2 + 1
                    manq_col_end = 3 + len(VEH_COLS) * 3
                    vals = ["", "", f"   ↳ Sous-total {usine}"] + \
                            [day_usine_tot[v] or "" for v in VEH_COLS] + \
                            [disp[v] for v in VEH_COLS] + \
                            [manq[v] for v in VEH_COLS]
                    for ci, v in enumerate(vals, 1):
                        c = ws.cell(row, ci); c.value = v; c.border = _BORD; c.alignment = _CTR
                        if ci == 3:
                            c.font = Font(bold=True, italic=True, name="Calibri", size=9)
                        if manq_col_start <= ci <= manq_col_end:
                            try: c.fill = REDF if float(v) > 0 else GRNF
                            except (ValueError, TypeError): pass
                    row += 1

            for usine, totals in week_totals.items():
                disp_fixed = {v: _fleet_count(usine, v) for v in VEH_COLS}
                manq_sum = {v: week_manq[usine][v] for v in VEH_COLS}
                manq_col_start = 3 + len(VEH_COLS) * 2 + 1
                manq_col_end = 3 + len(VEH_COLS) * 3
                vals = ["", "", f"TOTAL SEMAINE {usine}"] + \
                        [totals[v] or "" for v in VEH_COLS] + \
                        [disp_fixed[v] for v in VEH_COLS] + \
                        [manq_sum[v] for v in VEH_COLS]
                for ci, v in enumerate(vals, 1):
                    c = ws.cell(row, ci); c.value = v; c.border = _BORD; c.alignment = _CTR
                    c.fill = _hf(_tint_hex(comm_hex, 0.7))
                    c.font = _ft(bold=True, size=9, white=True)
                    if manq_col_start <= ci <= manq_col_end:
                        try: c.fill = REDF if float(v) > 0 else GRNF; c.font = _ft(bold=True, size=9)
                        except (ValueError, TypeError): pass
                row += 1
            row += 1
        row += 1

    for ci, w in enumerate([12, 24, 16, 11, 11, 10, 11, 13, 13, 11, 13, 16, 16, 14, 16], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A4"
    return wb


def generate_transport_sheets(wb, p_corrige, real_fleet=None):
    """Ajoute les feuilles 'Transport Jour' et 'Transport Semaine' au workbook
    `wb` deja ouvert, calculees a partir du planning RECTIFIE (p_corrige),
    jamais a partir des valeurs OR-Tools d'origine."""
    if not TRANSPORT_CALC_AVAILABLE or p_corrige is None or p_corrige.empty:
        return wb
    real_fleet = real_fleet or {}
    detail = _tc.build_transport_detail(p_corrige, real_fleet=real_fleet)
    if detail.empty:
        return wb
    df_usine = _tc.summarize_transport_usine(detail, real_fleet=real_fleet, period="day")
    df_detail = _tc.summarize_transport_detail(detail, period="day")
    _write_transport_period_sheet(wb, "Transport Jour",
                                   "Besoins Transport — par JOUR (calcule sur le Plan Rectifie)",
                                   df_usine, df_detail, "Date")
    _write_transport_semaine_detaillee(wb, detail, real_fleet)
    return wb


def generate_planning_wide_excel(p_display, real_fleet=None):
    """
    Genere l'Excel "colonnes dates", organise en sections par Commercial
    puis par Usine, avec sous-totaux (formules Excel) et regroupement
    pliable (boutons +/- Excel). Une ligne de donnees par
    (Commercial, Usine, Agriculteur).

    Structure de chaque section Commercial :
      - bandeau d'entete colore (couleur de marque du commercial)
      - pour chaque Usine livree : lignes agriculteurs (teinte claire de
        la couleur du commercial) puis ligne "Sous-total <Usine>"
      - ligne "TOTAL <Commercial>" (couleur pleine, formules SOMME)
    Suivi d'une ligne "TOTAL GENERAL" en bas de feuille.
    """
    import io as _io
    if real_fleet is None:
        real_fleet = st.session_state.get("_real_fleet", {})
    if p_display is None or p_display.empty:
        wb = Workbook(); ws = wb.active; ws.title = "Vide"
        ws["A1"] = "Aucune donnee a exporter"
        buf = _io.BytesIO(); wb.save(buf); buf.seek(0); return buf.read()

    df = p_display.copy()
    df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
    df = df.dropna(subset=["Date"])
    if df.empty:
        wb = Workbook(); ws = wb.active; ws.title = "Vide"
        ws["A1"] = "Aucune donnee a exporter"
        buf = _io.BytesIO(); wb.save(buf); buf.seek(0); return buf.read()

    # ── Enrichir avec Accessibilité et Région depuis _p_full_with_acc ──
    # p_display (issu du dashboard Tab1) ne contient pas Accessibilité/Région
    # car le dashboard les retire lors de la construction de p_display.
    # Sans ces colonnes, choose_vehicles() dans build_transport_detail() ne peut
    # pas déterminer le bon type de véhicule (SEMI pour ACHREF, TRACTEUR pour FEDI).
    _p_full = st.session_state.get("_p_full_with_acc")
    if _p_full is not None and not _p_full.empty:
        for _col in ["Accessibilité", "Région"]:
            if _col not in df.columns and _col in _p_full.columns:
                # Merge sur (Date, Commercial, Agriculteur, Usine)
                _key_cols = [c for c in ["Date","Commercial","Agriculteur","Usine"] if c in df.columns and c in _p_full.columns]
                _src = _p_full[_key_cols + [_col]].copy()
                _src["Date"] = pd.to_datetime(_src["Date"], errors="coerce")
                df = df.merge(_src.drop_duplicates(_key_cols), on=_key_cols, how="left")
            elif _col not in df.columns:
                df[_col] = ""

    for c in ["Commercial","Agriculteur","Usine","Tonnes/Jour","Type Véhicule",
              "Véhicules Requis","Disponibles","Manquants (à louer)","Nb Voyages","Pic de Récolte",
              "Accessibilité","Région"]:
        if c not in df.columns:
            df[c] = ""
    df["Tonnes/Jour"] = pd.to_numeric(df["Tonnes/Jour"], errors="coerce").fillna(0)
    df["_T_int"] = _round_preserving_total(df, group_cols=("Commercial", "Date"))
    df["Usine"] = df["Usine"].fillna("").astype(str).replace({"nan": "", "": "(usine non assignee)"})

    season_dates = list(pd.date_range("2026-06-20", "2026-08-25", freq="D"))
    FIXED_HDRS = ["Date","Commercial","Agriculteur","Usine","Tonnes/Jour","Type Véhicule",
                  "Véhicules Requis","Disponibles","Manquants (à louer)","Nb Voyages","Pic de Récolte"]
    DATE_HDRS = [d.strftime("%d/%m/%Y") for d in season_dates]
    ALL_HDRS = FIXED_HDRS + DATE_HDRS
    N = len(ALL_HDRS)
    NUM_COLS = [5, 8, 9, 10]                       # colonnes numeriques sommables
    DATE_COL_START = len(FIXED_HDRS) + 1
    DATE_COL_END = len(FIXED_HDRS) + len(season_dates)

    wb = Workbook(); wb.remove(wb.active)
    ws = wb.create_sheet("Planning Rectifie")
    ws.sheet_properties.outlinePr.summaryBelow = True
    ws.sheet_view.showOutlineSymbols = True

    HDR  = _hf("1F3864")
    PICF = _hf("FFF2CC")
    GRNF = _hf("E2EFDA")
    REDF = _hf("FFC7CE")

    COMM_ORDER  = list(COMM_HEX.keys())
    USINE_ORDER = list(USINE_HEX.keys())
    _ckey = lambda c: (COMM_ORDER.index(c), c) if c in COMM_ORDER else (len(COMM_ORDER), str(c))
    _ukey = lambda u: (USINE_ORDER.index(u), u) if u in USINE_ORDER else (len(USINE_ORDER), str(u))

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=min(N, 30))
    ws.cell(1,1).value = "Planning Rectifie — base sur le plan corrige (Plans Rectifies prioritaires sur OR-Tools)"
    ws.cell(1,1).font = _ft(bold=True, size=12, white=True)
    ws.cell(1,1).fill = HDR; ws.cell(1,1).alignment = _CTR
    ws.row_dimensions[1].height = 28

    for ci, h in enumerate(ALL_HDRS, 1):
        c = ws.cell(2, ci); c.value = h; c.border = _BORD; c.alignment = _CTR
        if ci <= len(FIXED_HDRS):
            c.fill = HDR; c.font = _ft(bold=True, size=9, white=True)
        else:
            d_obj = season_dates[ci - len(FIXED_HDRS) - 1].date()
            is_pic = PIC_S <= d_obj <= PIC_E
            c.fill = _hf("7D6608") if is_pic else _hf("1F4E79")
            c.font = _ft(bold=True, size=8, white=True)
    ws.row_dimensions[2].height = 36
    ws.row_dimensions[1].outline_level = 0
    ws.row_dimensions[2].outline_level = 0

    def _sum_formula(col_letter, rows):
        """SOMME Excel referencant une liste de lignes (contigues ou non)."""
        if not rows:
            return None
        return f"=SUM({','.join(f'{col_letter}{r}' for r in rows)})"

    def _write_label_row(r, label, fill_hex, text_color, bold, size, italic=False, outline=0, height=18):
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        for ci in range(1, N+1):
            cell = ws.cell(r, ci)
            cell.fill = _hf(fill_hex); cell.border = _BORD; cell.alignment = _CTR
            cell.font = Font(bold=bold, italic=italic, name="Calibri", size=size, color=text_color)
        ws.cell(r, 2).value = label
        ws.cell(r, 2).alignment = _LFT
        ws.row_dimensions[r].height = height
        ws.row_dimensions[r].outline_level = outline

    def _write_sum_row(r, label, data_rows, fill_hex, text_color, bold, size, outline, italic=False, height=18):
        """Ecrit une ligne de sous-total/total avec formules SOMME sur data_rows
        (liste de numeros de ligne, contigus ou non) pour chaque colonne numerique
        et chaque colonne date."""
        _write_label_row(r, label, fill_hex, text_color, bold, size, italic, outline, height)
        for col in NUM_COLS:
            f = _sum_formula(get_column_letter(col), data_rows)
            if f:
                cell = ws.cell(r, col); cell.value = f
                cell.number_format = "#,##0"
        for ci2 in range(DATE_COL_START, DATE_COL_END + 1):
            f = _sum_formula(get_column_letter(ci2), data_rows)
            if f:
                cell = ws.cell(r, ci2); cell.value = f
                cell.number_format = "#,##0;;\"\""

    comm_list = sorted(df["Commercial"].dropna().unique(), key=_ckey)
    row = 3
    comm_total_rows = []   # lignes "TOTAL <commercial>" -> pour le total general

    for comm in comm_list:
        comm_hex = COMM_HEX.get(comm, "404040")
        comm_tint = _tint_hex(comm_hex, 0.78)
        comm_tint_alt = _tint_hex(comm_hex, 0.88)
        sub_df = df[df["Commercial"] == comm]

        header_row = row
        _write_label_row(header_row, f"📍 {comm}", comm_hex, "FFFFFF", True, 11, outline=0, height=22)
        row += 1

        usine_list = sorted(sub_df["Usine"].dropna().unique(), key=_ukey)
        usine_subtotal_rows = []

        for usine in usine_list:
            usine_df = sub_df[sub_df["Usine"] == usine]
            groups = usine_df.groupby("Agriculteur", dropna=False)
            usine_start = row
            for agri, grp in sorted(groups, key=lambda kv: str(kv[0])):
                grp_daily = grp.groupby(grp["Date"].dt.date)["_T_int"].sum()
                total = int(grp_daily.sum())
                r0 = grp.iloc[0]
                is_alt = (row - usine_start) % 2 == 1
                base_fill = _hf(comm_tint_alt if is_alt else comm_tint)
                is_pic_any = any(PIC_S <= d <= PIC_E and v > 0 for d, v in grp_daily.items())
                manq_val = r0.get("Manquants (à louer)", "")
                ws.cell(row, 1).fill = _hf(comm_hex)
                ws.cell(row, 1).border = _BORD
                fixed_vals = ["", comm, agri, usine, total if total else "",
                              r0.get("Type Véhicule",""), r0.get("Véhicules Requis",""),
                              r0.get("Disponibles",""), manq_val,
                              r0.get("Nb Voyages",""), "⚡ PIC" if is_pic_any else ""]
                for ci, val in enumerate(fixed_vals, 1):
                    if ci == 1:
                        continue
                    cell = ws.cell(row, ci); cell.value = val; cell.border = _BORD; cell.alignment = _CTR
                    if ci == 9:
                        try:
                            cell.fill = REDF if float(val) > 0 else base_fill
                        except (ValueError, TypeError):
                            cell.fill = base_fill
                    else:
                        cell.fill = base_fill
                    cell.font = _ft(size=9, bold=(ci == 3))
                for di, d in enumerate(season_dates):
                    val = int(grp_daily.get(d.date(), 0) or 0)
                    ci2 = len(FIXED_HDRS) + di + 1
                    cell = ws.cell(row, ci2)
                    cell.value = val if val > 0 else ""
                    cell.border = _BORD; cell.alignment = _CTR
                    is_pic_d = PIC_S <= d.date() <= PIC_E
                    cell.fill = PICF if (is_pic_d and val > 0) else (GRNF if val > 0 else base_fill)
                    cell.font = _ft(size=8)
                ws.row_dimensions[row].height = 15
                ws.row_dimensions[row].outline_level = 2
                row += 1
            usine_end = row - 1
            usine_hex = USINE_HEX.get(usine, "595959")
            _write_sum_row(row, f"   ↳ Sous-total {usine}", list(range(usine_start, usine_end+1)),
                            _tint_hex(usine_hex, 0.35), "FFFFFF", True, 9, outline=1, italic=True, height=16)
            usine_subtotal_rows.append(row)
            row += 1

        comm_total_row = row
        _write_sum_row(comm_total_row, f"TOTAL {comm}", usine_subtotal_rows,
                        comm_hex, "FFFFFF", True, 10, outline=0, height=20)
        comm_total_rows.append(comm_total_row)
        row += 1

    grand_total_row = row
    _write_sum_row(grand_total_row, "TOTAL GENERAL", comm_total_rows,
                    "0B132B", "FFFFFF", True, 12, outline=0, height=24)
    row += 1

    for ci, w in enumerate([4,13,34,16,11,12,11,11,13,10,11], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    for di in range(len(season_dates)):
        ws.column_dimensions[get_column_letter(len(FIXED_HDRS) + di + 1)].width = 8
    ws.freeze_panes = "E3"
    ws.auto_filter.ref = f"B2:{get_column_letter(N)}2"

    if TRANSPORT_CALC_AVAILABLE:
        try:
            p_for_transport = st.session_state.get("_p_full_with_acc", p_display)
            generate_transport_sheets(wb, p_for_transport, real_fleet or {})
        except Exception as e:
            st.warning(f"Feuilles Transport non generees: {e}")

    buf = _io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.read()

# ═══ POINT D'ENTREE PRINCIPAL ════════════════════════════════
def render_comparaison_tab(planning_df=None, df_to_xlsx_styled=None, sb=None):

    st.markdown("""
    <div style='background:#1a2332;border:1px solid #21262d;border-radius:12px;
    padding:16px 20px;margin-bottom:20px'>
      <div style='font-size:1.1rem;font-weight:700;color:#f0f6fc;margin-bottom:4px'>
        Plans Rectifies — Source principale de donnees
      </div>
      <div style='font-size:.82rem;color:#8b949e'>
        Plan rectifie = version corrigee du planning OR-Tools. Cascade : fichier uploade ici
        > Supabase (persistant) > reference interne (donnees reelles du 13/06/2026).
        Cette page corrige automatiquement TOUS les onglets du dashboard (Planning, Par Commercial,
        Par Usine...) car ils lisent le meme planning corrige.
      </div>
    </div>""", unsafe_allow_html=True)

    if "comp_uploaded" not in st.session_state: st.session_state["comp_uploaded"]={}
    if "comp_raw"      not in st.session_state: st.session_state["comp_raw"]={}

    if "comp_sb_loaded" not in st.session_state:
        _sb_comm,_sb_usine=_load_all_rectifie(sb)
        st.session_state["_sb_comm"]=_sb_comm
        st.session_state["_sb_usine"]=_sb_usine
        if _sb_comm:
            for _k,_v in _sb_comm.items():
                st.session_state["comp_uploaded"].setdefault(_k,_v)
            try: st.toast(f"Donnees chargees depuis Supabase ({len(_sb_comm)} commerciaux)")
            except: pass
        st.session_state["comp_sb_loaded"]=True

    if "_real_fleet" not in st.session_state:
        st.session_state["_real_fleet"] = _load_flotte_reelle(sb) if TRANSPORT_CALC_AVAILABLE else {}

    sb_comm=st.session_state.get("_sb_comm",{})
    sb_usine=st.session_state.get("_sb_usine",{})

    if st.session_state["comp_uploaded"] and "comp_excel_cache" not in st.session_state:
        try:
            _mc={};_oc={}
            for _c in MANUAL_STATS:
                _mc[_c],_=_get_man_dict(_c,st.session_state["comp_uploaded"],sb_comm)
                _oc[_c]=_ortools_profile(_c,planning_df,"commercial")
            _mu={};_ou={}
            for _u in USINE_CAPS:
                _mu[_u],_=_get_usine_dict(_u,sb_usine)
                _ou[_u]=_ortools_profile(_u,planning_df,"usine")
            st.session_state["comp_excel_cache"]=generate_comparison_excel(_mc,_oc,_mu,_ou)
        except: pass

    st.subheader("Deposer / Mettre a jour les plans rectifies")
    st.caption("Sauvegardes dans Supabase — persistant entre sessions et devices. Priorite sur la reference interne.")
    col_up,col_st=st.columns([3,2])
    with col_up:
        uploaded=st.file_uploader("Fichiers",type=["xlsx","xls"],accept_multiple_files=True,
                                  label_visibility="collapsed")
    if uploaded:
        for f in uploaded:
            cfn=_detect_comm_from_filename(f.name)
            cfd,daily,raw_or_err=_parse_rectification(f)
            comm=cfn or cfd
            if comm and daily:
                st.session_state["comp_uploaded"][comm]=daily
                if isinstance(raw_or_err,pd.DataFrame):
                    st.session_state["comp_raw"][comm]=raw_or_err
                    _save_rectifie_detail(sb, comm, raw_or_err)
                    st.session_state.setdefault("_sb_detail",{})[comm]=raw_or_err
                ok=_save_rectifie(sb,comm,daily,"commercial")
                if ok: st.session_state.get("_sb_comm",{})[comm]=daily
                if "comp_excel_cache" in st.session_state: del st.session_state["comp_excel_cache"]
                tot=sum(daily.values())
                st.success(f"{'Supabase sauvegarde' if ok else 'Charge'}: {comm} — {len(daily)}j, {int(tot)}t")
            elif comm: st.warning(f"{f.name}: {raw_or_err}")
            else: st.warning(f"{f.name}: commercial non detecte. Renommer: Rectification_FEDI_...")

    with col_st:
        st.markdown("**Statut :**")
        sb_detail_cache = st.session_state.get("_sb_detail", {})
        any_scaling = False
        for c in MANUAL_STATS:
            d,src=_get_man_dict(c,st.session_state["comp_uploaded"],sb_comm)
            t=sum(d.values());n=sum(1 for v in d.values() if v>0)
            icon = "🟢" if src=="fichier uploade" else ("🔵" if src=="Supabase" else "⚪")
            has_detail = (isinstance(st.session_state["comp_raw"].get(c), pd.DataFrame) and not st.session_state["comp_raw"].get(c).empty) or \
                         (isinstance(sb_detail_cache.get(c), pd.DataFrame) and not sb_detail_cache.get(c).empty)
            detail_tag = " · chiffres EXACTS" if has_detail else " · ⚠️ SCALING (detail manquant)"
            if not has_detail and src != "reference interne": any_scaling = True
            col_a,col_b = st.columns([4,1])
            with col_a:
                st.markdown(f"{icon} **{c.split()[0]}** — {n}j | {int(t)}t ({src}){detail_tag}")
            with col_b:
                if src=="Supabase":
                    if st.button("↺", key=f"reset_{c}", help=f"Reinitialiser {c} -> reference interne (si Supabase contient une valeur perimee)"):
                        _delete_rectifie(sb, c, "commercial")
                        st.session_state["_sb_comm"].pop(c, None)
                        st.session_state["comp_uploaded"].pop(c, None)
                        st.session_state.get("_sb_detail",{}).pop(c, None)
                        st.session_state["comp_raw"].pop(c, None)
                        if "comp_excel_cache" in st.session_state: del st.session_state["comp_excel_cache"]
                        st.rerun()
        if any_scaling:
            st.warning("⚠️ Au moins un commercial utilise le scaling proportionnel au lieu des "
                       "chiffres exacts (detail granulaire absent de Supabase/session). "
                       "Reimporte son fichier ci-dessus pour corriger.")
        st.markdown("---")
        st.markdown("**Usines :**")
        for u in USINE_CAPS:
            d,src=_get_usine_dict(u,sb_usine)
            t=sum(d.values())
            icon = "🔵" if src=="Supabase" else "⚪"
            col_a,col_b = st.columns([4,1])
            with col_a:
                st.markdown(f"{icon} **{u}** — {int(t)}t ({src})")
            with col_b:
                if src=="Supabase":
                    if st.button("↺", key=f"reset_u_{u}", help=f"Reinitialiser {u} -> reference interne"):
                        _delete_rectifie(sb, u, "usine")
                        st.session_state["_sb_usine"].pop(u, None)
                        if "comp_excel_cache" in st.session_state: del st.session_state["comp_excel_cache"]
                        st.rerun()
        if st.session_state["comp_uploaded"]:
            if st.button("Effacer uploads session",use_container_width=True):
                st.session_state["comp_uploaded"]={}
                st.session_state["comp_raw"]={}
                if "comp_excel_cache" in st.session_state: del st.session_state["comp_excel_cache"]
                st.rerun()

    st.markdown("---")
    st.subheader("Flotte de transport reelle (pour le tableau Transport)")
    st.caption("Fichier transport_etat_final.xlsx (ou transport_disponible.xlsx) — feuille "
               "'liste confirmé' avec colonnes Usine/Tonnage/Type Vehicule/Confirmation/Contrat. "
               "Sauvegarde dans Supabase, persistant entre sessions.")
    if TRANSPORT_CALC_AVAILABLE:
        col_fu, col_fs = st.columns([3, 2])
        with col_fu:
            fleet_file = st.file_uploader("Fichier flotte", type=["xlsx", "xls"],
                                           label_visibility="collapsed", key="fleet_uploader")
        if fleet_file:
            parsed, err = _tc.parse_real_fleet_file(fleet_file)
            if err:
                st.warning(f"Lecture flotte: {err}")
            elif parsed:
                st.session_state["_real_fleet"] = parsed
                ok = _save_flotte_reelle(sb, parsed)
                nb = sum(len(v) for vt in parsed.values() for v in vt.values())
                st.success(f"{'Supabase sauvegarde' if ok else 'Charge (session)'}: {nb} camions sur {len(parsed)} usines")
                if "comp_excel_cache" in st.session_state: del st.session_state["comp_excel_cache"]
            else:
                st.warning("Aucune flotte detectee dans ce fichier.")
        with col_fs:
            rf = st.session_state.get("_real_fleet", {})
            if rf:
                for u, vt in sorted(rf.items()):
                    n = sum(len(v) for v in vt.values())
                    st.markdown(f"🔵 **{u}** — {n} camions")
            else:
                st.markdown("⚪ Aucune flotte chargee — le tableau Transport utilisera des "
                             "capacites theoriques (0 camion disponible affiche) jusqu'a l'import.")
    else:
        st.info("Module transport_calc.py non deploye — tableau Transport indisponible pour l'instant.")

    st.divider()
    col_exp1,col_exp2=st.columns([2,3])
    with col_exp1:
        if "comp_excel_cache" in st.session_state and st.session_state["comp_excel_cache"]:
            st.download_button("Telecharger Comparaison_Rectifie_vs_ORT_2026.xlsx",
                data=st.session_state["comp_excel_cache"],
                file_name="Comparaison_ORT_vs_Manuel_2026.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,type="primary")
            st.caption("Fichier pret (genere automatiquement)")
            if st.button("Regenerer Excel",use_container_width=True):
                del st.session_state["comp_excel_cache"];st.rerun()
        else:
            if st.button("Generer et telecharger Excel",type="primary",use_container_width=True):
                with st.spinner("Generation..."):
                    mc={};oc={}
                    for comm in MANUAL_STATS:
                        mc[comm],_=_get_man_dict(comm,st.session_state["comp_uploaded"],sb_comm)
                        oc[comm]=_ortools_profile(comm,planning_df,"commercial")
                    mu={};ou={}
                    for usine in USINE_CAPS:
                        mu[usine],_=_get_usine_dict(usine,sb_usine)
                        ou[usine]=_ortools_profile(usine,planning_df,"usine")
                    xl=generate_comparison_excel(mc,oc,mu,ou)
                st.session_state["comp_excel_cache"]=xl;st.rerun()
    with col_exp2:
        st.markdown("""<div style='background:#161b22;border:1px solid #21262d;border-radius:8px;
        padding:12px 16px;font-size:12px;color:#8b949e'>
          <b style='color:#f0f6fc'>Contenu :</b> Synthese + 5 commerciaux + 5 usines + Legende<br>
          <b style='color:#3b82f6'>Source :</b> Plan Rectifie (principal)<br>
          <b style='color:#888'>Comparaison :</b> OR-Tools en pointille<br>
          <b style='color:#f5a623'>Astuce :</b> bouton ↺ a cote d'une entite = efface la valeur
          Supabase perimee et retombe sur la reference interne reelle
        </div>""",unsafe_allow_html=True)

    st.divider()

    c1,c2,c3,c4,c5=st.tabs(["Courbes par commercial","Par usine","Statistiques globales","Etat optimizer","🚛 Transport Semaine"])

    with c1:
        sel_comm=st.selectbox("Choisir un commercial",list(MANUAL_STATS.keys()),key="comp_sel_comm")
        color=COMM_COLORS[sel_comm];cap=MANUAL_STATS[sel_comm]["cap"]
        man_dict,src_lbl=_get_man_dict(sel_comm,st.session_state["comp_uploaded"],sb_comm)
        ort_dict=_ortools_profile(sel_comm,planning_df,"commercial")
        _kpi_row(sel_comm,man_dict,ort_dict,cap)
        st.caption(f"Source plan rectifie : **{src_lbl}**")
        st.plotly_chart(_build_chart(sel_comm,man_dict,ort_dict,color,cap),use_container_width=True)
        st.markdown(f"**Tableau jour par jour — {sel_comm}**")
        ms={str(k):v for k,v in man_dict.items()};os_={str(k):v for k,v in ort_dict.items()}
        pivot_rows=[]
        for d in SEASON:
            dk=d.date();mv=ms.get(str(dk),0);ov=os_.get(str(dk),0)
            if mv==0 and ov==0: continue
            ec=mv-ov
            pivot_rows.append({"Date":d.strftime("%d/%m/%Y"),"Jour":DAYS_FR[d.weekday()],
                               "PIC":"⚡" if PIC_S<=dk<=PIC_E else "",
                               "Rectifie (t)":int(mv) if mv>0 else "",
                               "OR-Tools (t)":int(ov) if ov>0 else "",
                               "Ecart (t)":f"{int(ec):+d}" if (mv>0 or ov>0) else "",
                               "Statut":"OK" if abs(ec)<=50 else ("Rect+" if ec>0 else "OR-T+")})
        if pivot_rows:
            st.dataframe(pd.DataFrame(pivot_rows),use_container_width=True,height=320,hide_index=True,
                         column_config={"Rectifie (t)":st.column_config.NumberColumn("Rectifie (t)",format="%d t"),
                                        "OR-Tools (t)":st.column_config.NumberColumn("OR-Tools (t)",format="%d t")})
        if sel_comm in st.session_state["comp_raw"]:
            st.markdown("---")
            st.markdown(f"**Profil par agriculteur — {sel_comm}**")
            df_raw=st.session_state["comp_raw"][sel_comm]
            agri_stats=df_raw.groupby("agriculteur").agg(
                total=("tonnes","sum"),max_jour=("tonnes","max"),jours=("date","nunique")
            ).reset_index().sort_values("total",ascending=False)
            agri_stats.columns=["Agriculteur","Total (t)","Max/jour (t)","Jours actifs"]
            agri_stats[["Total (t)","Max/jour (t)"]]=agri_stats[["Total (t)","Max/jour (t)"]].round(0).astype(int)
            fig_a=px.bar(agri_stats,x="Total (t)",y="Agriculteur",orientation="h",
                         height=max(350,len(agri_stats)*28+80),color="Total (t)",
                         color_continuous_scale="Viridis",title=f"Tonnage par agriculteur — {sel_comm}",
                         template="plotly_dark",text="Total (t)")
            fig_a.update_traces(textposition="outside",texttemplate="%{x:.0f}t")
            fig_a.update_layout(paper_bgcolor="#161b22",showlegend=False,margin=dict(l=230,r=60,t=50,b=40))
            st.plotly_chart(fig_a,use_container_width=True)

    with c2:
        sel_usine=st.selectbox("Choisir une usine",["SICAM","TUCAL","COMOCAP","ELFALLEH","ABIDA"],key="comp_usine_sel")
        cap=USINE_CAPS[sel_usine];us=USINE_STATS.get(sel_usine,{})
        man_u,src_u=_get_usine_dict(sel_usine,sb_usine)
        ort_u=_ortools_profile(sel_usine,planning_df,"usine")
        total_u=us.get("total",sum(man_u.values()) if man_u else 0)
        max_u=us.get("max_j",max(man_u.values()) if man_u else 0)
        jours_u=us.get("n_jours",sum(1 for v in man_u.values() if v>0))
        ot_u=sum(ort_u.values()) if ort_u else 0
        k1,k2,k3,k4,k5=st.columns(5)
        k1.metric("Total saison",f"{int(total_u)}t")
        k2.metric("Max journalier",f"{max_u}t",
                  delta=f"cap {cap}t/j" if max_u<=cap else f"+{max_u-cap}t vs cap {cap}t/j",
                  delta_color="normal" if max_u<=cap else "inverse")
        k3.metric("Cap officiel",f"{cap}t/j")
        k4.metric("Total OR-Tools",f"{int(ot_u)}t",delta=f"{int(total_u-ot_u):+d}t ecart")
        k5.metric("Jours actifs",f"{jours_u}j")
        st.caption(f"Source : {src_u}")
        dates_u=[pd.Timestamp(k) for k in sorted(man_u.keys())]
        vals_u=[man_u[k] for k in sorted(man_u.keys())]
        fig_u=go.Figure()
        fig_u.add_trace(go.Bar(x=dates_u,y=vals_u,name="Plan rectifie",marker_color="#3b82f6",marker_line_width=0))
        fig_u.add_hline(y=cap,line_dash="dash",line_color="#e8543a",line_width=1.5,
                        annotation_text=f"Cap:{cap}t/j",annotation_position="top right",annotation_font_color="#e8543a")
        fig_u.add_vrect(x0=pd.Timestamp("2026-07-01"),x1=pd.Timestamp("2026-07-15"),
                        fillcolor="rgba(245,166,35,0.08)",line_width=0)
        if ort_u:
            fig_u.add_trace(go.Scatter(x=[pd.Timestamp(d) for d in sorted(ort_u.keys())],
                y=[ort_u[d] for d in sorted(ort_u.keys())],
                name="OR-Tools",line=dict(color="#f5a623",width=2,dash="dash"),mode="lines"))
        fig_u.update_layout(title=f"{sel_usine} — Plan Rectifie vs OR-Tools",
            template="plotly_dark",plot_bgcolor="#0d1117",paper_bgcolor="#161b22",
            height=360,hovermode="closest",legend=dict(orientation="h",yanchor="bottom",y=1.02))
        st.plotly_chart(fig_u,use_container_width=True)
        st.markdown(f"**Tableau jour par jour — {sel_usine}**")
        man_u_s={str(k):v for k,v in man_u.items()};ort_u_s={str(k):v for k,v in ort_u.items()}
        pivot_u=[]
        for d in SEASON:
            dk=d.date();mv=man_u_s.get(str(dk),0);ov=ort_u_s.get(str(dk),0)
            if mv==0 and ov==0: continue
            ec=mv-ov
            pivot_u.append({"Date":d.strftime("%d/%m/%Y"),"Jour":DAYS_FR[d.weekday()],
                            "PIC":"⚡" if PIC_S<=dk<=PIC_E else "",
                            "Rectifie (t)":int(mv) if mv>0 else "",
                            "OR-Tools (t)":int(ov) if ov>0 else "",
                            "Ecart (t)":f"{int(ec):+d}" if (mv>0 or ov>0) else "",
                            "Cap":"DEPASSE" if (PIC_S<=dk<=PIC_E and max(mv,ov)>cap) else ("OK" if PIC_S<=dk<=PIC_E else "")})
        if pivot_u:
            st.dataframe(pd.DataFrame(pivot_u),use_container_width=True,height=320,hide_index=True)
        rules={"SICAM":["Montee progressive juin (140-625t)","Pic 1-15 juillet (910-1650t)","Declin rapide aout"],
               "TUCAL":["Demarrage 24 juin","Plateau 590-700t/j juillet"],
               "COMOCAP":["Pic tardif 27/07 (875t)","Declin net aout"],
               "ELFALLEH":["Total reel 5190t / 38 jours / max 240t/j","Cap officiel 150t/j (jours doubles autorises)"],
               "ABIDA":["Total reel 8010t / max 320t/j","Demarrage tardif juillet"]}
        st.markdown("**Regles terrain :**")
        for rule in rules.get(sel_usine,[]): st.markdown(f"• {rule}")

    with c3:
        st.markdown("### Comparaison globale — tous commerciaux et usines")
        rows_cmp=[];mt_list=[];ot_list=[]
        for comm in MANUAL_STATS:
            d,src=_get_man_dict(comm,st.session_state["comp_uploaded"],sb_comm)
            ds={str(k):v for k,v in d.items()}
            od=_ortools_profile(comm,planning_df,"commercial")
            ods={str(k):v for k,v in od.items()}
            mt=sum(ds.values());ot=sum(ods.values()) if ods else 0
            mm=max(ds.values()) if ds else 0;om=max(ods.values()) if ods else 0
            ec=mt-ot;pct=round(ec/ot*100,1) if ot>0 else 0
            mt_list.append(mt);ot_list.append(ot)
            rows_cmp.append({"Commercial":comm,"Source":src,"Total Rectifie(t)":int(mt),
                             "Total OR-T(t)":int(ot),"Ecart(t)":f"{int(ec):+d}",
                             "Max Rect":f"{int(mm)}t","Max OR-T":f"{int(om)}t",
                             "Jours Rect":sum(1 for v in ds.values() if v>0),
                             "Cap PIC":f"{MANUAL_STATS[comm]['cap']}t/j"})
        st.dataframe(pd.DataFrame(rows_cmp),use_container_width=True,hide_index=True)
        fig_cmp=go.Figure()
        comm_order=list(MANUAL_STATS.keys())
        fig_cmp.add_trace(go.Bar(name="Plan Rectifie",x=comm_order,y=mt_list,marker_color="#3b82f6",
                                 text=[f"{int(v)}t" for v in mt_list],textposition="outside"))
        fig_cmp.add_trace(go.Bar(name="OR-Tools",x=comm_order,y=ot_list,marker_color="#888888",
                                 text=[f"{int(v)}t" for v in ot_list],textposition="outside"))
        fig_cmp.update_layout(barmode="group",template="plotly_dark",paper_bgcolor="#161b22",
                              plot_bgcolor="#0d1117",height=380,yaxis_title="Tonnes",
                              legend=dict(orientation="h",yanchor="bottom",y=1.02))
        st.plotly_chart(fig_cmp,use_container_width=True)

    with c4:
        st.markdown("### Etat actuel optimizer_v2.py — Configuration OPTIMALE")
        st.success("Configuration stable — ne pas modifier les parametres ci-dessous.")
        st.markdown("""
| Parametre | Valeur actuelle | Statut |
|---|---|---|
| Borne journaliere (`_ub_day`) | `day_planned x SCALE x 2.0` | OPTIMAL |
| `FACTORY_OVERFLOW_WEIGHT` | `2000` | Applique |
| Arrondi tonnage | `10t` | Stable |
| Correction post-traitement | Desactivee (`continue`) | OK |
        """)
        st.warning("**x1.1 et x1.5 causent INFEASIBLE** — teste et prouve. Ne pas appliquer.")
        st.code("""Status: OPTIMAL (~2s) | 2630 rows | 96 676t (-0.32%)
FEDI   : -234t (-0.7%)
MAKKI  : +115t (+0.5%)
ACHREF : 0t (parfait)
KHALIL : -139t (-0.8%)
JILANI : -125t (-1.8%)""")

    # ══════════════════════════════════════════════════════════════
    # ONGLET 5 — TRANSPORT SEMAINE
    # ══════════════════════════════════════════════════════════════
    with c5:
        if not TRANSPORT_CALC_AVAILABLE:
            st.info("Module transport_calc.py non deploye — onglet Transport indisponible.")
        else:
            TRACTEUR_MAX = 6
            TRACTEUR_CAP = 10.0

            p_tr = build_effective_planning(planning_df, sb=sb)
            p_full_tr = st.session_state.get("_p_full_with_acc", p_tr)
            real_fleet_tr = st.session_state.get("_real_fleet", {})

            @st.cache_data(show_spinner=False, ttl=300)
            def _compute_transport_weekly(p_hash):
                detail = _tc.build_transport_detail(p_full_tr, real_fleet=real_fleet_tr)
                detail["Date"] = pd.to_datetime(detail["Date"])
                # Max journalier par (date, commercial, type)
                daily = detail.groupby(["Date","Commercial","Type Véhicule"])["Voyages"].sum().reset_index()
                daily["Week_start"] = daily["Date"].dt.to_period("W-SUN").apply(lambda w: w.start_time.date())
                daily["Week_end"]   = daily["Date"].dt.to_period("W-SUN").apply(lambda w: w.end_time.date())
                # Cap TRACTEUR = 6, excédent -> PPL
                rows_c = []
                for _, row in daily.iterrows():
                    vt = row["Type Véhicule"]; trips = row["Voyages"]
                    if vt == "TRACTEUR" and trips > TRACTEUR_MAX:
                        excess = trips - TRACTEUR_MAX
                        ppl_extra = -(-int(excess * TRACTEUR_CAP) // 14)
                        rows_c.append({**row.to_dict(), "Voyages": TRACTEUR_MAX})
                        rows_c.append({**row.to_dict(), "Type Véhicule": "PPL", "Voyages": ppl_extra})
                    else:
                        rows_c.append(row.to_dict())
                dc = pd.DataFrame(rows_c)
                wk_max = dc.groupby(["Week_start","Week_end","Commercial","Type Véhicule"])["Voyages"].max().reset_index()
                wk_max["Voyages"] = wk_max["Voyages"].astype(int)

                # Tableau par usine : MAX journalier par (date, comm, usine, type)
                daily_u = detail.groupby(["Date","Commercial","Usine","Type Véhicule"])["Voyages"].sum().reset_index()
                daily_u["Week_start"] = daily_u["Date"].dt.to_period("W-SUN").apply(lambda w: w.start_time.date())
                daily_u["Week_end"]   = daily_u["Date"].dt.to_period("W-SUN").apply(lambda w: w.end_time.date())
                rows_u = []
                for _, row in daily_u.iterrows():
                    vt = row["Type Véhicule"]; trips = row["Voyages"]
                    if vt == "TRACTEUR" and trips > TRACTEUR_MAX:
                        excess = trips - TRACTEUR_MAX
                        ppl_extra = -(-int(excess * TRACTEUR_CAP) // 14)
                        rows_u.append({**row.to_dict(), "Voyages": TRACTEUR_MAX})
                        rows_u.append({**row.to_dict(), "Type Véhicule": "PPL", "Voyages": ppl_extra})
                    else:
                        rows_u.append(row.to_dict())
                dcu = pd.DataFrame(rows_u)
                wk_usine = dcu.groupby(["Week_start","Week_end","Commercial","Usine","Type Véhicule"])["Voyages"].max().reset_index()
                wk_usine["Voyages"] = wk_usine["Voyages"].astype(int)
                return wk_max, wk_usine

            try:
                _ph = str(len(p_full_tr)) if p_full_tr is not None and not p_full_tr.empty else "0"
                wk_max, wk_usine = _compute_transport_weekly(_ph)
            except Exception as _e:
                st.error(f"Erreur calcul transport: {_e}")
                wk_max, wk_usine = pd.DataFrame(), pd.DataFrame()

            # ── Constantes style ──
            _VEH_ORDER  = ["TRACTEUR","PPL","PL","SEMI"]
            _VEH_EMOJI  = {"TRACTEUR":"🚜","PPL":"🟣","PL":"🔵","SEMI":"🟢"}
            _VEH_COLOR  = {"TRACTEUR":"#833C00","PPL":"#7030A0","PL":"#2E75B6","SEMI":"#1E8449"}
            _COMM_COLOR = {
                "FEDI":"#1A5276","MAKKI BEN SALAH":"#1F7A1F",
                "KHALIL":"#7D3C98","ACHREF AJLANI":"#C0392B","JILANI OBAY":"#D4AC0D",
            }
            _COMM_ORDER = ["FEDI","MAKKI BEN SALAH","KHALIL","ACHREF AJLANI","JILANI OBAY"]
            _USINE_ORDER = ["SICAM","TUCAL","COMOCAP","ELFALLEH","ABIDA"]

            # ─────────────────────────────────────────────
            # TABLEAU 1 — BESOIN PAR SEMAINE PAR COMMERCIAL
            # ─────────────────────────────────────────────
            st.markdown("""
<div style='background:#1a2332;border:1px solid #2E75B6;border-radius:10px;
padding:14px 18px;margin-bottom:16px'>
  <div style='font-size:1.05rem;font-weight:700;color:#f0f6fc'>
    📊 Tableau 1 — Besoin en bennes par semaine par commercial
  </div>
  <div style='font-size:.8rem;color:#8b949e;margin-top:4px'>
    MAX journalier de la semaine = nb bennes simultanées requises •
    TRACTEUR limité à <b style='color:#FFC107'>6 max</b> • excédent redistributé en PPL
  </div>
</div>""", unsafe_allow_html=True)

            if not wk_max.empty:
                all_weeks = sorted(wk_max["Week_start"].unique())
                week_labels = {wd: f"{str(wd)[5:]} → {str((pd.Timestamp(wd)+pd.Timedelta(days=6)).date())[5:]}"
                               for wd in all_weeks}

                # Sélecteur commercial
                sel_comm_tr = st.selectbox("Filtrer par commercial (ou Tous)",
                                           ["Tous"] + _COMM_ORDER,
                                           key="tr_sel_comm")
                comms_to_show = _COMM_ORDER if sel_comm_tr == "Tous" else [sel_comm_tr]

                for comm in comms_to_show:
                    sub = wk_max[wk_max["Commercial"] == comm]
                    if sub.empty:
                        continue
                    vehs = [v for v in _VEH_ORDER if v in sub["Type Véhicule"].values]
                    comm_hex = _COMM_COLOR.get(comm, "#404040")

                    # En-tête commercial
                    st.markdown(f"""
<div style='background:{comm_hex};border-radius:8px 8px 0 0;
padding:8px 14px;margin-top:12px;margin-bottom:0'>
  <span style='color:white;font-weight:700;font-size:.95rem'>{comm}</span>
</div>""", unsafe_allow_html=True)

                    # Construction DataFrame pivot
                    pivot_rows = []
                    for vt in vehs:
                        vt_sub = sub[sub["Type Véhicule"] == vt]
                        row_d = {"Type": f"{_VEH_EMOJI.get(vt,'')} {vt}"}
                        for wd in all_weeks:
                            wk_row = vt_sub[vt_sub["Week_start"] == wd]
                            val = int(wk_row["Voyages"].sum()) if not wk_row.empty else 0
                            row_d[week_labels[wd]] = val if val > 0 else ""
                        pivot_rows.append(row_d)

                    df_piv = pd.DataFrame(pivot_rows)
                    # Style: cellules non vides colorées selon type véhicule
                    week_cols = [c for c in df_piv.columns if c != "Type"]

                    def _style_tr_row(row):
                        vt_name = str(row["Type"]).split()[-1] if row["Type"] else ""
                        col = _VEH_COLOR.get(vt_name, "#444")
                        styles = [f"background-color:{col}33;color:white;font-weight:600"] + [
                            (f"background-color:{col}22;font-weight:700;text-align:center"
                             if str(v) not in ("","0") else "color:#555;text-align:center")
                            for v in row[week_cols]
                        ]
                        return styles

                    styled = df_piv.style.apply(_style_tr_row, axis=1)
                    st.dataframe(styled, use_container_width=True, hide_index=True,
                                 height=min(38 * len(pivot_rows) + 40, 200))

                # Note TRACTEUR
                st.markdown("""
<div style='background:#2D1B00;border:1px solid #FFC107;border-radius:6px;
padding:8px 14px;margin-top:12px;font-size:.82rem;color:#FFC107'>
  ⚠️ <b>TRACTEUR</b> : limité à 6 disponibles max — l'excédent est comptabilisé en PPL (≈14t/voyage)
</div>""", unsafe_allow_html=True)

                # Bouton export
                if st.button("📥 Exporter Tableau 1 (Excel)", key="exp_tr1"):
                    import io as _io, openpyxl as _oxl
                    from openpyxl.styles import PatternFill as _PF, Font as _FN, Alignment as _AL, Border as _BD, Side as _SD
                    from openpyxl.utils import get_column_letter as _gcl
                    def _hf(c): return _PF("solid",start_color=c,end_color=c)
                    def _bd(): s=_SD(style="thin",color="CCCCCC"); return _BD(left=s,right=s,top=s,bottom=s)
                    _CTR=_AL(horizontal="center",vertical="center")
                    wb_e=_oxl.Workbook(); ws_e=wb_e.active; ws_e.title="Besoin par Semaine"
                    _all_w=sorted(wk_max["Week_start"].unique())
                    _wlbls=[f"{str(w)[5:]}→{str((pd.Timestamp(w)+pd.Timedelta(days=6)).date())[5:]}" for w in _all_w]
                    _nc=2+len(_wlbls)
                    ws_e.merge_cells(f"A1:{_gcl(_nc)}1")
                    ws_e["A1"]="BESOIN TRANSPORT PAR SEMAINE — Cap TRACTEUR: 6 max"
                    ws_e["A1"].fill=_hf("0B132B"); ws_e["A1"].font=_FN(bold=True,color="FFFFFF",name="Calibri",size=13); ws_e["A1"].alignment=_CTR
                    ws_e.row_dimensions[1].height=28
                    for _ci,_t in enumerate(["Commercial","Type Véhicule"],1):
                        _c=ws_e.cell(2,_ci); _c.value=_t; _c.fill=_hf("2C3E50"); _c.font=_FN(bold=True,color="FFFFFF",name="Calibri",size=10); _c.alignment=_CTR; _c.border=_bd()
                    for _ci,_l in enumerate(_wlbls,3):
                        _c=ws_e.cell(2,_ci); _c.value=_l; _c.fill=_hf("2C3E50"); _c.font=_FN(bold=True,color="FFFFFF",name="Calibri",size=10); _c.alignment=_CTR; _c.border=_bd()
                    ws_e.row_dimensions[2].height=22
                    _VLC={"TRACTEUR":"833C00","PPL":"7030A0","PL":"2E75B6","SEMI":"1E8449"}
                    _VLL={"TRACTEUR":"F5E6DF","PPL":"EDE7F6","PL":"DEEBF7","SEMI":"E2EFDA"}
                    _CCX={"FEDI":"1A5276","MAKKI BEN SALAH":"1F7A1F","KHALIL":"7D3C98","ACHREF AJLANI":"C0392B","JILANI OBAY":"D4AC0D"}
                    _er=3
                    for _cm in _COMM_ORDER:
                        _sb=wk_max[wk_max["Commercial"]==_cm]
                        if _sb.empty: continue
                        _chx=_CCX.get(_cm,"404040"); _vs=[v for v in _VEH_ORDER if v in _sb["Type Véhicule"].values]
                        _ms=_er
                        for _vi,_vt in enumerate(_vs):
                            _vs2=_sb[_sb["Type Véhicule"]==_vt]; _vhx=_VLC.get(_vt,"404040"); _vlt=_VLL.get(_vt,"F5F5F5")
                            ws_e.cell(_er,1).value=_cm if _vi==0 else ""; ws_e.cell(_er,1).fill=_hf(_chx); ws_e.cell(_er,1).font=_FN(bold=True,color="FFFFFF",name="Calibri",size=10); ws_e.cell(_er,1).alignment=_CTR; ws_e.cell(_er,1).border=_bd()
                            ws_e.cell(_er,2).value=_vt; ws_e.cell(_er,2).fill=_hf(_vhx); ws_e.cell(_er,2).font=_FN(bold=True,color="FFFFFF",name="Calibri",size=10); ws_e.cell(_er,2).alignment=_CTR; ws_e.cell(_er,2).border=_bd()
                            for _ci2,_wd in enumerate(_all_w,3):
                                _wr=_vs2[_vs2["Week_start"]==_wd]; _vl=int(_wr["Voyages"].sum()) if not _wr.empty else 0
                                _cc=ws_e.cell(_er,_ci2); _cc.border=_bd()
                                if _vl>0: _cc.value=_vl; _cc.fill=_hf(_vlt); _cc.font=_FN(bold=True,name="Calibri",size=11); _cc.alignment=_CTR
                                else: _cc.value="—"; _cc.fill=_hf("F8F8F8"); _cc.font=_FN(color="CCCCCC",name="Calibri",size=9); _cc.alignment=_CTR
                            ws_e.row_dimensions[_er].height=20; _er+=1
                        if len(_vs)>1:
                            ws_e.merge_cells(start_row=_ms,start_column=1,end_row=_er-1,end_column=1)
                            ws_e.cell(_ms,1).alignment=_AL(horizontal="center",vertical="center",wrap_text=True)
                        for _ci2 in range(1,_nc+1): ws_e.cell(_er,_ci2).fill=_hf("ECF0F1"); _er+=1
                    ws_e.column_dimensions["A"].width=20; ws_e.column_dimensions["B"].width=12
                    for _ci in range(3,_nc+1): ws_e.column_dimensions[_gcl(_ci)].width=14
                    ws_e.freeze_panes="C3"
                    _buf=_io.BytesIO(); wb_e.save(_buf); _buf.seek(0)
                    st.download_button("💾 Telecharger Tableau 1", data=_buf.getvalue(),
                                       file_name="transport_besoin_semaines.xlsx",
                                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

            st.divider()

            # ─────────────────────────────────────────────────────
            # TABLEAU 2 — BESOIN PAR USINE PAR SEMAINE PAR COMMERCIAL
            # ─────────────────────────────────────────────────────
            st.markdown("""
<div style='background:#1a2332;border:1px solid #1E8449;border-radius:10px;
padding:14px 18px;margin-bottom:16px'>
  <div style='font-size:1.05rem;font-weight:700;color:#f0f6fc'>
    🏭 Tableau 2 — Besoin par Usine × Semaine × Commercial
  </div>
  <div style='font-size:.8rem;color:#8b949e;margin-top:4px'>
    Pour chaque usine : nb bennes max par semaine décomposé par commercial et type de véhicule
  </div>
</div>""", unsafe_allow_html=True)

            if not wk_usine.empty:
                all_weeks_u = sorted(wk_usine["Week_start"].unique())
                wlbls_u = {wd: f"{str(wd)[5:]} → {str((pd.Timestamp(wd)+pd.Timedelta(days=6)).date())[5:]}"
                           for wd in all_weeks_u}

                # Filtres
                col_fu1, col_fu2, col_fu3 = st.columns(3)
                with col_fu1:
                    sel_usine_tr = st.selectbox("Usine", ["Toutes"] + _USINE_ORDER, key="tr_sel_usine")
                with col_fu2:
                    sel_veh_tr = st.selectbox("Type Véhicule", ["Tous"] + _VEH_ORDER, key="tr_sel_veh")
                with col_fu3:
                    sel_comm_tr2 = st.selectbox("Commercial", ["Tous"] + _COMM_ORDER, key="tr_sel_comm2")

                usines_show = _USINE_ORDER if sel_usine_tr == "Toutes" else [sel_usine_tr]
                vehs_show   = _VEH_ORDER   if sel_veh_tr   == "Tous"   else [sel_veh_tr]
                comms_show  = _COMM_ORDER  if sel_comm_tr2 == "Tous"   else [sel_comm_tr2]

                for usine in usines_show:
                    sub_u = wk_usine[
                        (wk_usine["Usine"] == usine) &
                        (wk_usine["Type Véhicule"].isin(vehs_show)) &
                        (wk_usine["Commercial"].isin(comms_show))
                    ]
                    if sub_u.empty:
                        continue

                    # Couleur usine
                    _USINE_HEX = {"SICAM":"0B3954","TUCAL":"2C0A3B","COMOCAP":"0B4F6C",
                                  "ELFALLEH":"145A32","ABIDA":"641E16"}
                    uhex = _USINE_HEX.get(usine, "2C3E50")

                    st.markdown(f"""
<div style='background:{uhex};border-radius:8px 8px 0 0;
padding:8px 14px;margin-top:14px'>
  <span style='color:white;font-weight:700;font-size:.95rem'>🏭 {usine}</span>
</div>""", unsafe_allow_html=True)

                    # Construire pivot: lignes = (Commercial, Type), colonnes = semaines
                    pivot_u_rows = []
                    for comm in comms_show:
                        for vt in vehs_show:
                            sub_cv = sub_u[(sub_u["Commercial"]==comm) & (sub_u["Type Véhicule"]==vt)]
                            if sub_cv.empty:
                                continue
                            row_u = {"Commercial": comm, "Type": f"{_VEH_EMOJI.get(vt,'')} {vt}"}
                            any_val = False
                            for wd in all_weeks_u:
                                wk_r = sub_cv[sub_cv["Week_start"]==wd]
                                val = int(wk_r["Voyages"].sum()) if not wk_r.empty else 0
                                row_u[wlbls_u[wd]] = val if val > 0 else ""
                                if val > 0: any_val = True
                            if any_val:
                                pivot_u_rows.append(row_u)

                    if not pivot_u_rows:
                        st.caption("— Aucune donnée pour cette combinaison —")
                        continue

                    df_upiv = pd.DataFrame(pivot_u_rows)
                    week_cols_u = [c for c in df_upiv.columns if c not in ("Commercial","Type")]

                    def _style_u_row(row):
                        comm_name = str(row.get("Commercial",""))
                        vt_name   = str(row.get("Type","")).split()[-1]
                        cc  = _COMM_COLOR.get(comm_name,"#444")
                        vc  = _VEH_COLOR.get(vt_name,"#444")
                        styles = [
                            f"background-color:{cc}44;color:white;font-weight:600",
                            f"background-color:{vc}33;color:white;font-weight:600",
                        ] + [
                            (f"background-color:{vc}22;font-weight:700;text-align:center"
                             if str(v) not in ("","0") else "color:#555;text-align:center")
                            for v in row[week_cols_u]
                        ]
                        return styles

                    styled_u = df_upiv.style.apply(_style_u_row, axis=1)
                    st.dataframe(styled_u, use_container_width=True, hide_index=True,
                                 height=min(38 * len(pivot_u_rows) + 42, 320))

                # Bouton export Tableau 2
                if st.button("📥 Exporter Tableau 2 (Excel)", key="exp_tr2"):
                    import io as _io2, openpyxl as _oxl2
                    from openpyxl.styles import PatternFill as _PF2, Font as _FN2, Alignment as _AL2, Border as _BD2, Side as _SD2
                    from openpyxl.utils import get_column_letter as _gcl2
                    def _hf2(c): return _PF2("solid",start_color=c,end_color=c)
                    def _bd2(): s=_SD2(style="thin",color="CCCCCC"); return _BD2(left=s,right=s,top=s,bottom=s)
                    _CTR2=_AL2(horizontal="center",vertical="center")
                    _VLC2={"TRACTEUR":"833C00","PPL":"7030A0","PL":"2E75B6","SEMI":"1E8449"}
                    _VLL2={"TRACTEUR":"F5E6DF","PPL":"EDE7F6","PL":"DEEBF7","SEMI":"E2EFDA"}
                    _CCX2={"FEDI":"1A5276","MAKKI BEN SALAH":"1F7A1F","KHALIL":"7D3C98","ACHREF AJLANI":"C0392B","JILANI OBAY":"D4AC0D"}
                    _UHX2={"SICAM":"0B3954","TUCAL":"2C0A3B","COMOCAP":"0B4F6C","ELFALLEH":"145A32","ABIDA":"641E16"}
                    wb2=_oxl2.Workbook(); ws2e=wb2.active; ws2e.title="Usine x Semaine"
                    _aw2=sorted(wk_usine["Week_start"].unique())
                    _wl2=[f"{str(w)[5:]}→{str((pd.Timestamp(w)+pd.Timedelta(days=6)).date())[5:]}" for w in _aw2]
                    _nc2=3+len(_wl2)
                    ws2e.merge_cells(f"A1:{_gcl2(_nc2)}1")
                    ws2e["A1"]="BESOIN TRANSPORT PAR USINE x SEMAINE x COMMERCIAL — Saison 2026"
                    ws2e["A1"].fill=_hf2("0B132B"); ws2e["A1"].font=_FN2(bold=True,color="FFFFFF",name="Calibri",size=12); ws2e["A1"].alignment=_CTR2
                    ws2e.row_dimensions[1].height=26
                    for _ci3,_t3 in enumerate(["Usine","Commercial","Type Véhicule"],1):
                        _c3=ws2e.cell(2,_ci3); _c3.value=_t3; _c3.fill=_hf2("2C3E50"); _c3.font=_FN2(bold=True,color="FFFFFF",name="Calibri",size=10); _c3.alignment=_CTR2; _c3.border=_bd2()
                    for _ci3,_l3 in enumerate(_wl2,4):
                        _c3=ws2e.cell(2,_ci3); _c3.value=_l3; _c3.fill=_hf2("2C3E50"); _c3.font=_FN2(bold=True,color="FFFFFF",name="Calibri",size=10); _c3.alignment=_CTR2; _c3.border=_bd2()
                    ws2e.row_dimensions[2].height=22
                    _er2=3
                    for _us2 in _USINE_ORDER:
                        _uhx2=_UHX2.get(_us2,"2C3E50")
                        for _cm2 in _COMM_ORDER:
                            for _vt2 in _VEH_ORDER:
                                _sv2=wk_usine[(wk_usine["Usine"]==_us2)&(wk_usine["Commercial"]==_cm2)&(wk_usine["Type Véhicule"]==_vt2)]
                                if _sv2.empty or _sv2["Voyages"].sum()==0: continue
                                _vhx2=_VLC2.get(_vt2,"404040"); _vlt2=_VLL2.get(_vt2,"F5F5F5"); _chx2=_CCX2.get(_cm2,"404040")
                                ws2e.cell(_er2,1).value=_us2; ws2e.cell(_er2,1).fill=_hf2(_uhx2); ws2e.cell(_er2,1).font=_FN2(bold=True,color="FFFFFF",name="Calibri",size=10); ws2e.cell(_er2,1).alignment=_CTR2; ws2e.cell(_er2,1).border=_bd2()
                                ws2e.cell(_er2,2).value=_cm2; ws2e.cell(_er2,2).fill=_hf2(_chx2); ws2e.cell(_er2,2).font=_FN2(bold=True,color="FFFFFF",name="Calibri",size=10); ws2e.cell(_er2,2).alignment=_CTR2; ws2e.cell(_er2,2).border=_bd2()
                                ws2e.cell(_er2,3).value=_vt2; ws2e.cell(_er2,3).fill=_hf2(_vhx2); ws2e.cell(_er2,3).font=_FN2(bold=True,color="FFFFFF",name="Calibri",size=10); ws2e.cell(_er2,3).alignment=_CTR2; ws2e.cell(_er2,3).border=_bd2()
                                for _ci4,_wd2 in enumerate(_aw2,4):
                                    _wr2=_sv2[_sv2["Week_start"]==_wd2]; _vl2=int(_wr2["Voyages"].sum()) if not _wr2.empty else 0
                                    _cc2=ws2e.cell(_er2,_ci4); _cc2.border=_bd2()
                                    if _vl2>0: _cc2.value=_vl2; _cc2.fill=_hf2(_vlt2); _cc2.font=_FN2(bold=True,name="Calibri",size=11); _cc2.alignment=_CTR2
                                    else: _cc2.value="—"; _cc2.fill=_hf2("F8F8F8"); _cc2.font=_FN2(color="CCCCCC",name="Calibri",size=9); _cc2.alignment=_CTR2
                                ws2e.row_dimensions[_er2].height=18; _er2+=1
                    ws2e.column_dimensions["A"].width=12; ws2e.column_dimensions["B"].width=20; ws2e.column_dimensions["C"].width=12
                    for _ci in range(4,_nc2+1): ws2e.column_dimensions[_gcl2(_ci)].width=14
                    ws2e.freeze_panes="D3"
                    _buf2=_io2.BytesIO(); wb2.save(_buf2); _buf2.seek(0)
                    st.download_button("💾 Telecharger Tableau 2", data=_buf2.getvalue(),
                                       file_name="transport_usine_semaine.xlsx",
                                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")