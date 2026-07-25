# ============================================================
# DASHBOARD PHASE 10 — Tomate Planning 2026
#
# FONCTIONNALITES :
#   - Login par rôle (directeur / commercial / usine)
#   - Données depuis Supabase + fallback Excel local
#   - OR-Tools optimizer_v2.py (distances + caps)
#   - Upload Excel par commercial (upload_tab.py)
#   - Gestion agriculteurs directement dans le dashboard
#   - Historique 2025 vs plan 2026
#
# FICHIERS REQUIS dans le même dossier :
#   optimizer_v2.py  migrate.py  upload_tab.py
#   Planning_Tomate_2026.xlsx
#   Recap_tonnage_pre_vu_ajuste__mai26.xlsx
#
# LANCEMENT :
#   streamlit run dashboard_phase10.py
# ============================================================

import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import datetime, os, subprocess, sys, io, zipfile, hashlib

# Import upload system (upload_tab.py must be in the same folder)
try:
    from upload_tab import render_upload_tab, generate_template_excel
    UPLOAD_AVAILABLE = True
except ImportError:
    UPLOAD_AVAILABLE = False

try:
    from comparaison_tab import render_comparaison_tab
    COMPARAISON_AVAILABLE = True
except ImportError:
    COMPARAISON_AVAILABLE = False

try:
    from agroeco_dashboard import render_agroeco_tab
    AGROECO_AVAILABLE = True
except ImportError:
    AGROECO_AVAILABLE = False


# set_page_config MUST be the very first Streamlit call
st.set_page_config(
    page_title="🍅 Tomate Planning 2026",
    layout="wide",
    initial_sidebar_state="expanded",
    page_icon="🍅"
)

# ── CSV / Excel export helpers ───────────────────────────────
def df_to_csv(df):
    """Convert dataframe to CSV bytes for download button."""
    return df.to_csv(index=False, encoding="utf-8-sig").encode("utf-8-sig")

def df_to_xlsx_styled(df, sheet_name="Données", title=None):
    """
    Convert ONE dataframe to a styled Excel file (single sheet).
    Used for download buttons in the dashboard.
    """
    return dfs_to_excel({sheet_name: df})

# ── Plotly: réduire la sensibilité du hover ──────────────────
def smooth_hover(fig, mobile_ready=True):
    """
    Rend les courbes Plotly MOINS sensibles au survol :
    - hoverdistance = 50px (défaut Plotly = 20) → il faut s'approcher davantage
    - spikedistance = -1 → désactive les lignes verticales ultra-sensibles
    - hovermode = "closest" → tooltip seulement sur le point le plus proche
                              (au lieu de "x unified" qui affiche tout d'un coup)
    Sur mobile, désactive complètement le hover (clic uniquement).
    """
    fig.update_layout(
        hoverdistance=50,
        spikedistance=-1,
        hovermode="closest",
        # Sur mobile, désactive complètement le hover (problème tactile)
        # L'utilisateur clique pour voir les détails
    )
    # Désactiver les spike lines pour ne pas suivre la souris partout
    fig.update_xaxes(showspikes=False)
    fig.update_yaxes(showspikes=False)
    return fig

def dfs_to_zip(sheets: dict) -> bytes:
    """
    Convert multiple dataframes to a ZIP of CSVs.
    sheets = {"filename.csv": dataframe, ...}
    Returns ZIP bytes.
    """
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for fname, df in sheets.items():
            csv_bytes = df.to_csv(index=False, encoding="utf-8-sig").encode("utf-8-sig")
            zf.writestr(fname, csv_bytes)
    buf.seek(0)
    return buf.read()

def _sanitize_sheet_name(name: str) -> str:
    """
    Excel limite les noms d'onglets à 31 chars et interdit certains caractères.
    Caractères interdits: [ ] : * ? / \\
    """
    if not name:
        return "Feuille1"
    name = str(name)
    for bad in ['[', ']', ':', '*', '?', '/', '\\']:
        name = name.replace(bad, '-')
    name = name.strip("'").strip()
    name = name[:31].strip()
    if not name:
        name = "Feuille1"
    return name


def _sanitize_dataframe_for_excel(df):
    """
    Nettoie un DataFrame pour qu'il soit sérialisable dans Excel.
    Convertit dates en string, gère NaN, supprime types complexes.
    """
    if df is None or df.empty:
        return pd.DataFrame()
    out = df.copy()
    
    for col in out.columns:
        s = out[col]
        if pd.api.types.is_datetime64_any_dtype(s):
            try:
                out[col] = s.dt.strftime("%d/%m/%Y")
            except Exception:
                out[col] = s.astype(str)
        elif s.dtype == "object":
            try:
                out[col] = s.apply(
                    lambda x: x.strftime("%d/%m/%Y") if isinstance(x, pd.Timestamp)
                    else (str(x) if isinstance(x, (list, dict, tuple)) else x)
                )
            except Exception:
                pass
    
    out = out.fillna("")
    for col in out.columns:
        if out[col].dtype == "object":
            out[col] = out[col].astype(str).replace("nan", "").replace("NaT", "")
    
    return out


def dfs_to_excel(sheets: dict) -> bytes:
    """
    Convert multiple dataframes to a single styled Excel file with multiple sheets.
    sheets = {"Sheet Name": dataframe, ...}
    Returns Excel bytes.
    """
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule

    HEADER_FILL = PatternFill("solid", start_color="1F3864", end_color="1F3864")
    HEADER_FONT = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
    ROW_ALT_FILL = PatternFill("solid", start_color="EAF1FA", end_color="EAF1FA")
    BORDER_THIN = Border(left=Side(style="thin", color="D0D7DE"),
                          right=Side(style="thin", color="D0D7DE"),
                          top=Side(style="thin", color="D0D7DE"),
                          bottom=Side(style="thin", color="D0D7DE"))
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    RIGHT  = Alignment(horizontal="right",  vertical="center")
    PIC_FILL = PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC")

    COMM_FILLS = {
        "FEDI":            PatternFill("solid", start_color="DEEBF7", end_color="DEEBF7"),
        "MAKKI BEN SALAH": PatternFill("solid", start_color="E2EFDA", end_color="E2EFDA"),
        "KHALIL":          PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC"),
        "ACHREF AJLANI":   PatternFill("solid", start_color="EDEDED", end_color="EDEDED"),
        "JILANI OBAY":     PatternFill("solid", start_color="FCE4D6", end_color="FCE4D6"),
    }

    def _is_numeric_value(v):
        if v is None or v == "":
            return False
        try:
            float(v)
            return True
        except (ValueError, TypeError):
            return False

    if not sheets:
        sheets = {"Vide": pd.DataFrame({"Info": ["Aucune donnée à exporter"]})}
    
    used_names = set()
    def _unique_sheet_name(base):
        clean = _sanitize_sheet_name(base)
        if clean not in used_names:
            used_names.add(clean)
            return clean
        for i in range(2, 100):
            candidate = f"{clean[:28]}_{i}"
            if candidate not in used_names:
                used_names.add(candidate)
                return candidate
        return f"Sheet{len(used_names)+1}"

    buf = io.BytesIO()
    
    try:
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            for sheet_name, df in sheets.items():
                df_clean = _sanitize_dataframe_for_excel(df)
                if df_clean.empty:
                    df_clean = pd.DataFrame({"Info": ["Aucune donnée"]})
                
                clean_sheet = _unique_sheet_name(sheet_name)
                df_clean.to_excel(writer, sheet_name=clean_sheet, index=False)
                ws = writer.sheets[clean_sheet]

                n_rows, n_cols = df_clean.shape
                if n_cols == 0:
                    continue

                # Headers
                for col_idx in range(1, n_cols + 1):
                    cell = ws.cell(row=1, column=col_idx)
                    cell.fill = HEADER_FILL
                    cell.font = HEADER_FONT
                    cell.alignment = CENTER
                    cell.border = BORDER_THIN
                ws.row_dimensions[1].height = 30

                # Identifier colonnes spéciales
                comm_col_idx = None
                pic_col_idx  = None
                manquants_col_idx = None
                dispo_col_idx = None
                for i, col in enumerate(df_clean.columns, start=1):
                    col_str = str(col).strip().lower()
                    if col_str == "commercial":
                        comm_col_idx = i
                    if "pic" in col_str:
                        pic_col_idx = i
                    if "manquant" in col_str or "à louer" in col_str:
                        manquants_col_idx = i
                    if "disponible" in col_str:
                        dispo_col_idx = i

                # Style cellules
                from openpyxl.styles import PatternFill as _PF, Font as _F
                RED_FILL = _PF("solid", start_color="FFC7CE", end_color="FFC7CE")
                RED_FONT = _F(bold=True, color="9C0006", name="Calibri", size=10)
                GREEN_FILL = _PF("solid", start_color="C6EFCE", end_color="C6EFCE")
                GREEN_FONT = _F(bold=True, color="006100", name="Calibri", size=10)
                
                for r in range(2, n_rows + 2):
                    for c in range(1, n_cols + 1):
                        cell = ws.cell(row=r, column=c)
                        cell.border = BORDER_THIN
                        cell.alignment = RIGHT if _is_numeric_value(cell.value) else LEFT
                        if r % 2 == 0:
                            cell.fill = ROW_ALT_FILL

                    if comm_col_idx:
                        comm_val = str(ws.cell(row=r, column=comm_col_idx).value or "").strip().upper()
                        if comm_val in COMM_FILLS:
                            ws.cell(row=r, column=comm_col_idx).fill = COMM_FILLS[comm_val]
                            ws.cell(row=r, column=comm_col_idx).font = Font(bold=True, name="Calibri", size=10)

                    if pic_col_idx:
                        pic_val = str(ws.cell(row=r, column=pic_col_idx).value or "")
                        if "PIC" in pic_val.upper():
                            ws.cell(row=r, column=pic_col_idx).fill = PIC_FILL
                            ws.cell(row=r, column=pic_col_idx).font = Font(bold=True, color="996600")
                    
                    # ✅ Colorer ROUGE la colonne "Manquants" si valeur > 0
                    if manquants_col_idx:
                        m_cell = ws.cell(row=r, column=manquants_col_idx)
                        try:
                            v = float(m_cell.value) if m_cell.value not in (None, "") else 0
                            if v > 0:
                                m_cell.fill = RED_FILL
                                m_cell.font = RED_FONT
                            elif v == 0 and dispo_col_idx:
                                # 0 manquant = OK → vert subtil
                                m_cell.fill = GREEN_FILL
                                m_cell.font = GREEN_FONT
                        except (ValueError, TypeError):
                            pass

                # Largeur colonnes
                for col_idx, col in enumerate(df_clean.columns, start=1):
                    col_letter = get_column_letter(col_idx)
                    try:
                        max_len = max(
                            [len(str(col))] +
                            [len(str(v)) for v in df_clean.iloc[:, col_idx - 1].head(200).tolist()]
                        )
                    except Exception:
                        max_len = 15
                    ws.column_dimensions[col_letter].width = min(max(12, max_len + 3), 35)

                ws.freeze_panes = "A2"

                if n_rows > 0:
                    last_col_letter = get_column_letter(n_cols)
                    ws.auto_filter.ref = f"A1:{last_col_letter}{n_rows + 1}"

                # Color scale pour tonnage
                for col_idx, col in enumerate(df_clean.columns, start=1):
                    col_lower = str(col).lower()
                    if ("tonne" in col_lower or "tonnage" in col_lower) and n_rows > 1:
                        col_letter = get_column_letter(col_idx)
                        rng = f"{col_letter}2:{col_letter}{n_rows + 1}"
                        try:
                            rule = ColorScaleRule(
                                start_type="min", start_color="FFFFFF",
                                mid_type="percentile", mid_value=50, mid_color="9EC3E6",
                                end_type="max", end_color="1F3864",
                            )
                            ws.conditional_formatting.add(rng, rule)
                        except Exception:
                            pass
    except Exception as e:
        # ✅ Fallback : Excel basique sans style
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            for sheet_name, df in sheets.items():
                clean = _sanitize_sheet_name(sheet_name)
                df_safe = _sanitize_dataframe_for_excel(df) if df is not None else pd.DataFrame()
                if df_safe.empty:
                    df_safe = pd.DataFrame({"Erreur": [f"Erreur génération: {str(e)[:200]}"]})
                df_safe.to_excel(writer, sheet_name=clean, index=False)
    
    buf.seek(0)
    return buf.read()


def _build_usine_excel(planning_df, usine_name, agri_df=None):
    """
    Génère un fichier Excel riche pour une usine donnée.
    Structure :
      - Colonnes fixes : Commercial | Région | Agriculteur | Tonnage Total | Accessibilité
      - Puis UNE COLONNE PAR JOUR de la saison (20/06/2026, 21/06/2026, ..., 25/08/2026)
      - Lignes : une par agriculteur, avec le tonnage livré ce jour-là
      - Sous-total par commercial (somme de tous ses agriculteurs par jour)
      - TOTAL JOUR en bas (somme de tous les commerciaux par jour)
    """
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule
    import io as _io

    # ── Couleurs par usine ─────────────────────────────────────────
    USINE_COLORS = {
        "SICAM":    "1F3864",
        "TUCAL":    "4A235A",
        "COMOCAP":  "0B4F6C",
        "ABIDA":    "922B21",
        "ELFALLEH": "196F3D",
    }
    HDR_COLOR  = USINE_COLORS.get(usine_name.upper(), "1F3864")
    COMM_FILLS = {
        "FEDI":            "DEEBF7",
        "MAKKI BEN SALAH": "E2EFDA",
        "KHALIL":          "FFF2CC",
        "ACHREF AJLANI":   "EDEDED",
        "JILANI OBAY":     "FCE4D6",
    }
    ALT_ROW    = "F5F8FF"
    SUBTOT_CLR = "D6EAF8"   # bleu clair pour sous-total
    TOTAL_CLR  = "D5F5E3"   # vert clair pour total jour
    PIC_CLR    = "FFF2CC"   # jaune pour PIC (1-15 juillet)
    GRAND_CLR  = HDR_COLOR  # même couleur que l'en-tête pour grand total

    THIN   = Side(style="thin", color="CCCCCC")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    def hf(color): return PatternFill("solid", start_color=color, end_color=color)
    def bf(bold=True, color="000000", size=10, white=False):
        return Font(bold=bold, name="Calibri", size=size,
                    color="FFFFFF" if white else color)
    CTR  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT = Alignment(horizontal="left",   vertical="center")
    RGT  = Alignment(horizontal="right",  vertical="center")

    # ── Préparer données ────────────────────────────────────────────
    df = planning_df.copy()
    if df.empty:
        return df_to_xlsx_styled(df, sheet_name=usine_name)

    col_comm = next((c for c in df.columns if c.lower()=="commercial"), "Commercial")
    col_agri = next((c for c in df.columns if c.lower()=="agriculteur"), "Agriculteur")
    col_tons = next((c for c in df.columns if "tonne" in c.lower() and "jour" in c.lower()), "Tonnes/Jour")
    col_date = next((c for c in df.columns if c.lower()=="date"), "Date")
    col_reg  = next((c for c in df.columns if "region" in c.lower() or "région" in c.lower()), None)
    col_acc  = next((c for c in df.columns if "access" in c.lower()), None)

    df[col_date] = pd.to_datetime(df[col_date], errors="coerce")
    df[col_tons] = pd.to_numeric(df[col_tons], errors="coerce").fillna(0)

    # Récupérer accessibilité et région depuis agri_df si dispo
    agri_lookup = {}
    if agri_df is not None and not agri_df.empty:
        for _, row in agri_df.iterrows():
            k = (str(row.get("commercial","")).strip().upper(),
                 str(row.get("nom","")).strip().upper())
            agri_lookup[k] = {
                "accessibilite": str(row.get("accessibilite","") or ""),
                "region":        str(row.get("region","") or ""),
            }

    def _info(comm, nom, field):
        k  = (str(comm).strip().upper(), str(nom).strip().upper())
        k2 = (str(comm).strip().upper(), str(nom).split(" (")[0].strip().upper())
        return (agri_lookup.get(k) or agri_lookup.get(k2) or {}).get(field, "")

    # ── TABLEAU PIVOT ───────────────────────────────────────────────
    # ✅ Forcer TOUTE la saison 20/06 → 25/08 (même les jours sans livraison = 0)
    SEASON_START_XL = pd.Timestamp("2026-06-20")
    SEASON_END_XL   = pd.Timestamp("2026-08-25")
    all_season_dates = pd.date_range(SEASON_START_XL, SEASON_END_XL, freq="D")

    pivot = df.pivot_table(
        index=[col_comm, col_agri],
        columns=col_date,
        values=col_tons,
        aggfunc="sum",
        fill_value=0,
    )
    pivot = pivot.reset_index()
    pivot.columns.name = None

    # Colonnes de dates = TOUTE la saison (pas seulement les jours avec données)
    existing_date_cols = set(c for c in pivot.columns if isinstance(c, pd.Timestamp))
    date_cols = sorted(all_season_dates)
    # Ajouter les colonnes manquantes à 0 dans le pivot
    for d in date_cols:
        if d not in existing_date_cols:
            pivot[d] = 0

    # Construire la table finale avec les colonnes fixes en premier
    rows_data = []
    for _, row in pivot.iterrows():
        comm = str(row[col_comm])
        agri = str(row[col_agri])
        total_agri = sum(row[d] for d in date_cols if d in row.index)
        region = _info(comm, agri, "region")
        if not region and col_reg and col_reg in df.columns:
            sub = df[(df[col_comm]==comm) & (df[col_agri]==agri)]
            region = sub[col_reg].iloc[0] if not sub.empty else ""
        access = _info(comm, agri, "accessibilite")
        if not access and col_acc and col_acc in df.columns:
            sub = df[(df[col_comm]==comm) & (df[col_agri]==agri)]
            access = sub[col_acc].iloc[0] if not sub.empty else ""
        entry = {
            "Commercial":   comm,
            "Région":       region,
            "Agriculteur":  agri,
            "Tonnage Total (t)": round(total_agri, 0),
            "Accessibilité": access,
        }
        for d in date_cols:
            entry[d] = row[d] if d in row.index else 0
        rows_data.append(entry)

    # Trier par commercial puis agriculteur
    rows_data.sort(key=lambda r: (r["Commercial"], r["Agriculteur"]))

    # ── CONSTRUIRE L'EXCEL ─────────────────────────────────────────
    buf = _io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:

        # Feuille unique
        sheet_name = usine_name[:31]
        # Créer une feuille vide
        dummy = pd.DataFrame()
        dummy.to_excel(writer, sheet_name=sheet_name, index=False)
        ws = writer.sheets[sheet_name]

        # ── Nombre de colonnes ─────────────────────────────────────
        N_FIXED = 5  # Commercial | Région | Agriculteur | Tonnage Total | Accessibilité
        N_DATES = len(date_cols)
        N_TOTAL = N_FIXED + N_DATES

        # ── LIGNE 1 : Titre de l'usine ─────────────────────────────
        ws.merge_cells(start_row=1, start_column=1,
                       end_row=1, end_column=N_TOTAL)
        t = ws.cell(1, 1)
        t.value     = f"🏭  Planning {usine_name}  —  Saison 2026"
        t.font      = bf(white=True, size=14)
        t.fill      = hf(HDR_COLOR)
        t.alignment = CTR
        ws.row_dimensions[1].height = 34

        # ── LIGNE 2 : En-têtes ─────────────────────────────────────
        fixed_headers = ["Commercial", "Région", "Agriculteur",
                         "Tonnage Total (t)", "Accessibilité"]
        for ci, h in enumerate(fixed_headers, 1):
            c = ws.cell(2, ci)
            c.value     = h
            c.font      = bf(white=True, size=11)
            c.fill      = hf(HDR_COLOR)
            c.alignment = CTR
            c.border    = BORDER
        # En-têtes dates
        for di, d in enumerate(date_cols):
            ci = N_FIXED + di + 1
            c = ws.cell(2, ci)
            c.value     = d.strftime("%d/%m")
            # Surligner les jours PIC (1-15 juillet)
            is_pic_col = (d.month == 7 and 1 <= d.day <= 15)
            c.fill      = hf("7D6608" if is_pic_col else HDR_COLOR)
            c.font      = bf(white=True, size=9)
            c.alignment = CTR
            c.border    = BORDER
        ws.row_dimensions[2].height = 36

        # ── LIGNES DONNÉES avec sous-totaux par commercial ─────────
        current_row = 3
        current_comm = None
        comm_start_row = 3

        def _write_subtotal(ws, row_idx, comm_label, rows_data_comm, date_cols, N_FIXED, N_TOTAL, hf, bf, BORDER, CTR, RGT, LEFT, SUBTOT_CLR, HDR_COLOR):
            """Écrire une ligne de sous-total pour un commercial."""
            ws.merge_cells(start_row=row_idx, start_column=1,
                           end_row=row_idx, end_column=3)
            ws.cell(row_idx, 1).value     = f"Sous-total {comm_label}"
            ws.cell(row_idx, 1).font      = bf(size=10, color="0B4F6C")
            ws.cell(row_idx, 1).fill      = hf(SUBTOT_CLR)
            ws.cell(row_idx, 1).alignment = LEFT
            # Tonnage total
            tot = sum(r["Tonnage Total (t)"] for r in rows_data_comm)
            ws.cell(row_idx, 4).value     = round(tot, 0)
            ws.cell(row_idx, 4).font      = bf(size=10, color="0B4F6C")
            ws.cell(row_idx, 4).fill      = hf(SUBTOT_CLR)
            ws.cell(row_idx, 4).alignment = RGT
            ws.cell(row_idx, 5).fill = hf(SUBTOT_CLR)
            ws.cell(row_idx, 5).border = BORDER
            # Totaux par jour
            for di, d in enumerate(date_cols):
                ci = N_FIXED + di + 1
                day_tot = sum(r.get(d, 0) for r in rows_data_comm)
                c = ws.cell(row_idx, ci)
                c.value     = round(day_tot, 0) if day_tot > 0 else ""
                c.font      = bf(size=9, color="0B4F6C")
                c.fill      = hf(SUBTOT_CLR)
                c.alignment = CTR
                c.border    = BORDER
            for ci in range(1, N_TOTAL + 1):
                ws.cell(row_idx, ci).border = BORDER

        # Grouper par commercial pour insérer les sous-totaux
        from itertools import groupby as _groupby
        rows_data_sorted = sorted(rows_data, key=lambda r: (r["Commercial"], r["Agriculteur"]))
        comm_groups = {}
        for r in rows_data_sorted:
            comm_groups.setdefault(r["Commercial"], []).append(r)

        all_comms_sorted = sorted(comm_groups.keys())
        for comm in all_comms_sorted:
            comm_rows = comm_groups[comm]
            comm_upper = comm.strip().upper()
            fill_comm = COMM_FILLS.get(comm_upper, "F0F0F0")

            for ri, row_data in enumerate(comm_rows):
                is_alt = ri % 2 == 0
                # Colonnes fixes
                vals_fixed = [
                    row_data["Commercial"],
                    row_data["Région"],
                    row_data["Agriculteur"],
                    row_data["Tonnage Total (t)"],
                    row_data["Accessibilité"],
                ]
                for ci, val in enumerate(vals_fixed, 1):
                    c = ws.cell(current_row, ci)
                    c.value     = val
                    c.font      = bf(bold=(ci==1), size=10)
                    c.alignment = RGT if ci == 4 else LEFT
                    c.border    = BORDER
                    c.fill      = hf(fill_comm) if ci == 1 else hf(ALT_ROW if is_alt else "FFFFFF")

                # Colonnes jours
                for di, d in enumerate(date_cols):
                    ci = N_FIXED + di + 1
                    val = row_data.get(d, 0)
                    c  = ws.cell(current_row, ci)
                    c.value     = round(val, 0) if val > 0 else ""
                    c.font      = Font(name="Calibri", size=9)
                    c.alignment = CTR
                    c.border    = BORDER
                    c.fill      = hf(ALT_ROW if is_alt else "FFFFFF")

                current_row += 1

            # Sous-total du commercial
            _write_subtotal(ws, current_row, comm, comm_rows, date_cols,
                            N_FIXED, N_TOTAL, hf, bf, BORDER, CTR, RGT, LEFT,
                            SUBTOT_CLR, HDR_COLOR)
            current_row += 1

        # ── LIGNE TOTAL GÉNÉRAL PAR JOUR ───────────────────────────
        ws.merge_cells(start_row=current_row, start_column=1,
                       end_row=current_row, end_column=3)
        ws.cell(current_row, 1).value     = "TOTAL GÉNÉRAL / JOUR"
        ws.cell(current_row, 1).font      = bf(white=True, size=11)
        ws.cell(current_row, 1).fill      = hf(HDR_COLOR)
        ws.cell(current_row, 1).alignment = LEFT
        ws.cell(current_row, 1).border    = BORDER

        grand_total = sum(r["Tonnage Total (t)"] for r in rows_data)
        ws.cell(current_row, 4).value     = round(grand_total, 0)
        ws.cell(current_row, 4).font      = bf(white=True, size=11)
        ws.cell(current_row, 4).fill      = hf(HDR_COLOR)
        ws.cell(current_row, 4).alignment = RGT
        ws.cell(current_row, 4).border    = BORDER
        ws.cell(current_row, 5).fill   = hf(HDR_COLOR)
        ws.cell(current_row, 5).border = BORDER

        for di, d in enumerate(date_cols):
            ci = N_FIXED + di + 1
            day_grand = sum(r.get(d, 0) for r in rows_data)
            c  = ws.cell(current_row, ci)
            c.value     = round(day_grand, 0) if day_grand > 0 else ""
            c.font      = bf(white=True, size=9)
            c.fill      = hf(HDR_COLOR)
            c.alignment = CTR
            c.border    = BORDER

        ws.row_dimensions[current_row].height = 26

        # ── AJUSTEMENT COLONNES ────────────────────────────────────
        ws.column_dimensions["A"].width = 18  # Commercial
        ws.column_dimensions["B"].width = 14  # Région
        ws.column_dimensions["C"].width = 26  # Agriculteur
        ws.column_dimensions["D"].width = 14  # Tonnage Total
        ws.column_dimensions["E"].width = 13  # Accessibilité
        # Colonnes dates : étroites (5 chars : 01/07)
        for di in range(N_DATES):
            ci = N_FIXED + di + 1
            ws.column_dimensions[get_column_letter(ci)].width = 6

        # Figer les colonnes fixes (on reste en ligne 3 pour voir les en-têtes)
        ws.freeze_panes = "F3"

        # Color scale sur la colonne "Tonnage Total (t)"
        last_data_row = current_row - 1
        if last_data_row >= 3:
            ws.conditional_formatting.add(
                f"D3:D{last_data_row}",
                ColorScaleRule(
                    start_type="min", start_color="FFFFFF",
                    mid_type="percentile", mid_value=50, mid_color="9EC3E6",
                    end_type="max", end_color=HDR_COLOR,
                )
            )

        # ══════════════════════════════════════════════════════════
        # ONGLETS JOURNALIERS : un onglet par jour AVEC des données
        # Colonnes : Commercial | Région | Agriculteur | Tonnage (t) | Accessibilité | Pic
        # + Sous-total par commercial + Total général du jour
        # ══════════════════════════════════════════════════════════
        days_fr  = ["Lun","Mar","Mer","Jeu","Ven","Sam","Dim"]
        PEAK_S   = pd.Timestamp("2026-07-01")
        PEAK_E   = pd.Timestamp("2026-07-15")

        # Garder seulement les dates qui ont des livraisons réelles
        dates_with_data = sorted([d for d in date_cols
                                   if any(r.get(d, 0) > 0 for r in rows_data)])

        col_acc_src  = col_acc  if (col_acc  and col_acc  in df.columns) else None
        col_reg_src  = col_reg  if (col_reg  and col_reg  in df.columns) else None
        col_pic_src  = next((c for c in df.columns if "pic" in c.lower()), None)

        for date_val in dates_with_data:
            # Nom de l'onglet : 01-07 (Mer)
            try:
                sheet_d = f"{date_val.day:02d}-{date_val.month:02d} ({days_fr[date_val.weekday()]})"
            except Exception:
                sheet_d = str(date_val)[:31]

            # ── Construire les données du jour ─────────────────────
            day_rows = []
            for row_d in rows_data:
                tons_day = row_d.get(date_val, 0)
                if tons_day <= 0:
                    continue
                # Récupérer pic depuis df original
                sub = df[(df[col_comm] == row_d["Commercial"]) &
                         (df[col_agri] == row_d["Agriculteur"]) &
                         (df[col_date] == date_val)]
                pic_val = ""
                if col_pic_src and not sub.empty:
                    pic_val = str(sub[col_pic_src].iloc[0] or "")
                day_rows.append({
                    "Commercial":    row_d["Commercial"],
                    "Région":        row_d["Région"],
                    "Agriculteur":   row_d["Agriculteur"],
                    "Tonnage (t)":   round(tons_day, 0),
                    "Accessibilité": row_d["Accessibilité"],
                    "Pic":           pic_val,
                })

            if not day_rows:
                continue

            day_rows.sort(key=lambda r: (r["Commercial"], r["Agriculteur"]))
            N_COLS_D = 6  # Commercial | Région | Agriculteur | Tonnage | Accès | Pic

            # Créer feuille
            dummy_d = pd.DataFrame()
            dummy_d.to_excel(writer, sheet_name=sheet_d, index=False)
            wsd = writer.sheets[sheet_d]

            # ── Titre du jour ──────────────────────────────────────
            wsd.merge_cells(start_row=1, start_column=1,
                            end_row=1,   end_column=N_COLS_D)
            td = wsd.cell(1, 1)
            is_pic_day = (PEAK_S <= date_val <= PEAK_E)
            try:
                day_label = date_val.strftime("%A %d/%m/%Y")
            except Exception:
                day_label = sheet_d
            td.value     = f"🏭  {usine_name}  —  {day_label}" + (" ⚡ PIC" if is_pic_day else "")
            td.font      = bf(white=True, size=12)
            td.fill      = hf("7D6608" if is_pic_day else HDR_COLOR)
            td.alignment = CTR
            wsd.row_dimensions[1].height = 30

            # ── En-têtes ligne 2 ──────────────────────────────────
            hdrs_d = ["Commercial","Région","Agriculteur","Tonnage (t)","Accessibilité","Pic"]
            for ci, h in enumerate(hdrs_d, 1):
                c = wsd.cell(2, ci)
                c.value     = h
                c.font      = bf(white=True, size=11)
                c.fill      = hf("7D6608" if is_pic_day else HDR_COLOR)
                c.alignment = CTR
                c.border    = BORDER
            wsd.row_dimensions[2].height = 26

            # ── Données ────────────────────────────────────────────
            for ri, row_d in enumerate(day_rows):
                r_idx = ri + 3
                comm_u = row_d["Commercial"].strip().upper()
                fill_c = COMM_FILLS.get(comm_u, ALT_ROW if ri%2==0 else "FFFFFF")
                vals   = [row_d["Commercial"], row_d["Région"], row_d["Agriculteur"],
                          row_d["Tonnage (t)"], row_d["Accessibilité"], row_d["Pic"]]
                for ci, val in enumerate(vals, 1):
                    c = wsd.cell(r_idx, ci)
                    c.value     = val
                    c.border    = BORDER
                    c.font      = bf(bold=(ci==1), size=10)
                    c.alignment = RGT if ci == 4 else LEFT
                    c.fill      = hf(fill_c) if ci == 1 else hf(ALT_ROW if ri%2==0 else "FFFFFF")

            # ── Sous-totaux par commercial ─────────────────────────
            comms_day = dict()
            for row_d in day_rows:
                comms_day.setdefault(row_d["Commercial"], 0)
                comms_day[row_d["Commercial"]] += row_d["Tonnage (t)"]

            for comm_d, sub_t in sorted(comms_day.items()):
                rr = wsd.max_row + 1
                for ci in range(1, N_COLS_D + 1):
                    c = wsd.cell(rr, ci)
                    c.value     = None
                    c.fill      = hf(SUBTOT_CLR)
                    c.font      = bf(size=10, color="0B4F6C")
                    c.border    = BORDER
                    c.alignment = RGT if ci == 4 else LEFT
                wsd.cell(rr, 1).value = f"Sous-total {comm_d}"
                wsd.cell(rr, 4).value = round(sub_t, 0)
                wsd.row_dimensions[rr].height = 20

            # ── Total général du jour ──────────────────────────────
            last_r = wsd.max_row + 1
            for ci in range(1, N_COLS_D + 1):
                c = wsd.cell(last_r, ci)
                c.value     = None
                c.fill      = hf(HDR_COLOR)
                c.font      = bf(white=True, size=11)
                c.border    = BORDER
                c.alignment = RGT if ci == 4 else LEFT
            wsd.merge_cells(start_row=last_r, start_column=1,
                            end_row=last_r,   end_column=3)
            wsd.cell(last_r, 1).value = f"TOTAL  {sheet_d}"
            wsd.cell(last_r, 4).value = round(sum(r["Tonnage (t)"] for r in day_rows), 0)
            wsd.row_dimensions[last_r].height = 26

            # ── Largeurs ───────────────────────────────────────────
            for ci, w in enumerate([18, 14, 26, 12, 14, 8], 1):
                wsd.column_dimensions[get_column_letter(ci)].width = w
            wsd.freeze_panes = "A3"

    buf.seek(0)
    return buf.read()


def df_to_xlsx_by_day(df, date_col="Date", base_name="Jour"):
    """
    ✅ Convertit un DataFrame en Excel avec UN ONGLET PAR JOUR + onglet récap.
    Chaque jour = un onglet séparé, triés chronologiquement.
    """
    if df is None or df.empty:
        return df_to_xlsx_styled(pd.DataFrame({"Info": ["Aucune donnée"]}))
    
    # Trouver la colonne de date
    actual_date_col = None
    for c in df.columns:
        if str(c).strip().lower() == date_col.lower():
            actual_date_col = c
            break
    
    if actual_date_col is None:
        return df_to_xlsx_styled(df, sheet_name="Données")
    
    df_work = df.copy()
    try:
        df_work[actual_date_col] = pd.to_datetime(df_work[actual_date_col], errors="coerce")
    except Exception:
        pass
    
    sheets = {}
    # Onglet récap (toutes les données)
    sheets["📋 Tout le planning"] = df.copy()
    
    days_fr = ["Lun","Mar","Mer","Jeu","Ven","Sam","Dim"]
    
    # Trier par date pour un ordre cohérent des onglets
    df_sorted = df_work.sort_values(actual_date_col)
    
    for date_val, day_df in df_sorted.groupby(actual_date_col, dropna=False, sort=False):
        if pd.isna(date_val):
            sheet_name = "Sans date"
        else:
            try:
                d = pd.Timestamp(date_val)
                day_str = days_fr[d.weekday()]
                sheet_name = f"{d.day:02d}-{d.month:02d} ({day_str})"
            except Exception:
                sheet_name = str(date_val)[:31]
        
        day_clean = day_df.copy()
        if actual_date_col in day_clean.columns:
            day_clean = day_clean.drop(columns=[actual_date_col])
        # Trier par tonnage décroissant si applicable
        for c in day_clean.columns:
            if "tonne" in str(c).lower() or "tonnage" in str(c).lower():
                try:
                    day_clean = day_clean.sort_values(by=c, ascending=False)
                except Exception:
                    pass
                break
        sheets[sheet_name] = day_clean
    
    return dfs_to_excel(sheets)

# ── Page config ──────────────────────────────────────────────
# ============================================================
# LOGIN SYSTEM
# ============================================================

# Rôles disponibles :
#   directeur  → voit tout, peut régénérer le planning
#   commercial → voit seulement ses agriculteurs et ses données
#   usine      → voit seulement les livraisons à son usine

def hash_password(password: str) -> str:
    """Hash simple SHA-256 pour ne pas stocker les mots de passe en clair."""
    return hashlib.sha256(password.encode()).hexdigest()

def load_users() -> dict:
    """
    Charge les utilisateurs depuis st.secrets (Streamlit Cloud)
    ou depuis un dict local si secrets non disponibles (développement local).

    Format dans .streamlit/secrets.toml :
    [users.directeur]
    password = "hash_sha256_ici"
    role     = "directeur"
    name     = "Directeur Général"
    filter   = ""

    [users.fedi]
    password = "hash_sha256_ici"
    role     = "commercial"
    name     = "FEDI"
    filter   = "FEDI"

    [users.comocap]
    password = "hash_sha256_ici"
    role     = "usine"
    name     = "COMOCAP"
    filter   = "COMOCAP"
    """
    try:
        # Production : lire depuis secrets.toml
        users = {}
        for username, info in st.secrets.get("users", {}).items():
            users[username] = {
                "password": info["password"],
                "role":     info["role"],
                "name":     info["name"],
                "filter":   info.get("filter", ""),
            }
        if users:
            return users
    except Exception:
        pass

    # Fallback local pour développement (mots de passe en clair hashés)
    # CHANGE CES MOTS DE PASSE avant de mettre en production !
    return {
        "directeur": {
            "password": "6051fc84a7a0d74c225fb18a496b09952da5642e60723ecae543298edd7d82d6",
            "role":     "directeur",
            "name":     "Directeur General",
            "filter":   "",
        },
        "fedi": {
            "password": "4c9806477c34233bde7f927e4c32e71f61f18a03d8f1dbbea7eed62b52f68e80",
            "role":     "commercial",
            "name":     "FEDI",
            "filter":   "FEDI",
        },
        "makki": {
            "password": "dea689708af4c45ef9c7bd4a84a7be0c0cfaf86ea61a25946c91e6372d0b9bd5",
            "role":     "commercial",
            "name":     "MAKKI BEN SALAH",
            "filter":   "MAKKI BEN SALAH",
        },
        "khalil": {
            "password": "8673b5cfef6d807071848ed5087de701a011ee30cfd8c319a5441ab6c529767b",
            "role":     "commercial",
            "name":     "KHALIL",
            "filter":   "KHALIL",
        },
        "achref": {
            "password": "9963e33a2017503ba137009a825b2967177532f5ad1454f199cbcae49a6df156",
            "role":     "commercial",
            "name":     "ACHREF AJLANI",
            "filter":   "ACHREF AJLANI",
        },
        "jilani": {
            "password": "2bcc02eb315a8e31e9c00380096925bd4dd0fbb23c463380bfcdf748040e41b8",
            "role":     "commercial",
            "name":     "JILANI OBAY",
            "filter":   "JILANI OBAY",
        },
        "comocap": {
            "password": "3f276a837113a5f2d1d39351ec763e37f4b254d9148b5c4cec7f1121183d276f",
            "role":     "usine",
            "name":     "COMOCAP",
            "filter":   "COMOCAP",
        },
        "sicam": {
            "password": "f6e1df72ff9f1c25236226624b47c6417ee18031eb35f7179aaf12a73bdf24ab",
            "role":     "usine",
            "name":     "SICAM",
            "filter":   "SICAM",
        },
        "tucal": {
            "password": "9fd85359508dd65535e69b45d7ca357b1b3ae0d36ed3d8438dbf677f5ca83abd",
            "role":     "usine",
            "name":     "TUCAL",
            "filter":   "TUCAL",
        },
        "baccara": {
            "password": "a3997ef8b11c2d90fb5a05c4d807f6d530d1a3c0940fd7f61ed2a2b4cb0a5800",
            "role":     "centre",
            "name":     "STE BACCARA",
            "filter":   "STE BACCARA",
        },
        "kerkouane": {
            "password": "94776c6081efbbd6b24e28789add1912fc3e930f06bb58e4aeb6c41b24c12c46",
            "role":     "centre",
            "name":     "STE KERKOUANE S.A",
            "filter":   "STE KERKOUANE S.A",
        },
        "centre428": {
            "password": "e7e32a9ea2a259655a98ee0992de2c1f5585fe6bb81afe3c2117a4ba6273ac0b",
            "role":     "centre",
            "name":     "STE 428 SERVICES AGRICOLES",
            "filter":   "STE 428 SERVICES AGRICOLES",
        },
    }

def show_login_page():
    """Affiche la page de login — appelée si non connecté."""
    st.markdown("""
    <style>
    [data-testid="stAppViewContainer"]{background:#080b12}
    [data-testid="stSidebar"]{display:none}
    .login-wrap{max-width:420px;margin:80px auto 0;padding:0 16px}
    .login-logo{text-align:center;margin-bottom:32px}
    .login-logo h1{font-size:2rem;color:#f0f6fc;margin-top:12px}
    .login-logo p{color:#8b949e;font-size:.85rem}
    .login-card{background:#161b22;border:1px solid #21262d;border-radius:16px;padding:32px}
    .login-card h2{color:#f0f6fc;font-size:1.1rem;margin-bottom:24px;font-weight:600}
    </style>
    <div class="login-wrap">
      <div class="login-logo">
        <div style="font-size:3rem">🍅</div>
        <h1>Tomate Planning</h1>
        <p>Système de planification transport & récolte 2026</p>
      </div>
    </div>
    """, unsafe_allow_html=True)

    col_l, col_c, col_r = st.columns([1, 2, 1])
    with col_c:
        with st.container(border=True):
            st.markdown("#### 🔐 Connexion")
            username = st.text_input("Identifiant",
                                     placeholder="Entrez votre identifiant",
                                     label_visibility="visible")
            password = st.text_input("Mot de passe", type="password",
                                     placeholder="Entrez votre mot de passe")
            login_btn = st.button("Se connecter", use_container_width=True,
                                  type="primary")

            if login_btn:
                users = load_users()
                u = username.strip().lower()
                if u in users and users[u]["password"] == hash_password(password):
                    st.session_state["logged_in"]  = True
                    st.session_state["username"]   = u
                    st.session_state["role"]       = users[u]["role"]
                    st.session_state["name"]       = users[u]["name"]
                    st.session_state["filter"]     = users[u]["filter"]
                    # Sauvegarder dans l'URL pour survivre aux refreshs
                    try:
                        st.query_params["u"] = u
                        st.query_params["t"] = _make_token(u)
                    except Exception:
                        pass
                    st.rerun()
                else:
                    st.error("❌ Identifiant ou mot de passe incorrect.")

# ── Timeout de session : déconnexion après 30 min d'inactivité ─
SESSION_TIMEOUT_MIN = 30

def _check_session_timeout():
    """Déconnecte automatiquement après SESSION_TIMEOUT_MIN minutes."""
    import time
    now = time.time()
    last = st.session_state.get("last_activity", now)
    if now - last > SESSION_TIMEOUT_MIN * 60:
        for key in list(st.session_state.keys()):
            del st.session_state[key]
        try:
            st.query_params.clear()
        except Exception:
            pass
        st.warning("⏱️ Session expirée après inactivité. Reconnectez-vous.")
        st.stop()
    st.session_state["last_activity"] = now

# ── Helpers tokens URL pour persister la session entre refreshs ─
def _make_token(username):
    """Génère un token simple à partir du username + secret."""
    import hashlib
    secret = "tomate2026_secret_seed"
    return hashlib.sha256(f"{username}|{secret}".encode()).hexdigest()[:32]

def _verify_token(username, token):
    """Vérifie qu'un token correspond au username."""
    return token == _make_token(username)

# ── Check login state — vérifier d'abord les query params (refresh) ─
if "logged_in" not in st.session_state:
    st.session_state["logged_in"] = False

# Si pas connecté mais URL contient un token valide → reconnecter auto
if not st.session_state["logged_in"]:
    try:
        qp = st.query_params
        url_user  = qp.get("u", "")
        url_token = qp.get("t", "")
        if url_user and url_token:
            users = load_users()
            u = url_user.strip().lower()
            if u in users and _verify_token(u, url_token):
                # Restaurer la session depuis l'URL
                st.session_state["logged_in"] = True
                st.session_state["username"]  = u
                st.session_state["role"]      = users[u]["role"]
                st.session_state["name"]      = users[u]["name"]
                st.session_state["filter"]    = users[u]["filter"]
    except Exception:
        pass

if not st.session_state["logged_in"]:
    show_login_page()
    st.stop()   # ← Stop here if not logged in. Nothing below executes.

# ── At this point: user is logged in ─────────────────────────
_check_session_timeout()  # ← Vérifier timeout d'inactivité
CURRENT_USER   = st.session_state["username"]
CURRENT_ROLE   = st.session_state["role"]
CURRENT_NAME   = st.session_state["name"]
CURRENT_FILTER = st.session_state["filter"]

# ============================================================
# PAGE CONFIG — must be first Streamlit call
# ============================================================

st.markdown("""
<style>
  /* ── Fond global dark partout ───────────────────────── */
  [data-testid="stAppViewContainer"] { background: #0d1117; }
  [data-testid="stSidebar"]          { background: #161b22; border-right:1px solid #21262d; }
  [data-testid="stHeader"]           { background: #0d1117; }
  .block-container { padding-top: 1.5rem; }
  h1,h2,h3,h4 { color: #f0f6fc; }

  /* ── FIX TABLES BLANCHES (st.dataframe + st.table) ──── */
  [data-testid="stDataFrame"],
  [data-testid="stTable"],
  [data-testid="stDataFrameResizable"] {
    background: #161b22 !important;
    border-radius: 8px;
  }
  [data-testid="stDataFrame"] *,
  [data-testid="stTable"] * {
    background-color: transparent !important;
  }
  /* Cellules dataframe en dark */
  [data-testid="stDataFrame"] [role="row"] {
    background: #161b22 !important;
  }
  [data-testid="stDataFrame"] [role="row"]:nth-child(even) {
    background: #1c2128 !important;
  }
  [data-testid="stDataFrame"] [role="columnheader"] {
    background: #1F3864 !important;
    color: #ffffff !important;
    font-weight: 700 !important;
  }
  [data-testid="stDataFrame"] [role="gridcell"] {
    color: #e6edf3 !important;
    border-color: #30363d !important;
  }
  /* Boutons d'expansion/scroll dans les dataframes */
  [data-testid="stDataFrame"] button {
    background: #21262d !important;
    color: #e6edf3 !important;
  }

  /* ── KPI cards ──────────────────────────────────────── */
  .metric-row { display:flex; gap:16px; margin-bottom:20px; flex-wrap:wrap; }
  .kpi-box {
    background:#161b22; border:1px solid #21262d; border-radius:12px;
    padding:16px 20px; min-width:140px; flex:1;
    border-top: 2px solid var(--c,#e8543a);
  }
  .kpi-val { font-size:1.7rem; font-weight:700; color:#f0f6fc; }
  .kpi-lbl { font-size:.72rem; color:#8b949e; text-transform:uppercase; letter-spacing:.06em; margin-top:3px; }
  .kpi-sub { font-size:.7rem; color:#3dd68c; margin-top:4px; }
  .peak-box {
    background:linear-gradient(90deg,rgba(255,179,71,.1),rgba(232,84,58,.08));
    border:1px solid rgba(255,179,71,.25); border-radius:10px;
    padding:12px 18px; margin-bottom:18px; font-size:.85rem; color:#f0f6fc;
  }
  .stDownloadButton > button {
    background:#161b22 !important; border:1px solid #21262d !important;
    color:#f0f6fc !important; font-size:.8rem !important;
  }
  .stDownloadButton > button:hover {
    border-color:#3b82f6 !important; color:#3b82f6 !important;
  }

  /* ── Tabs en dark ───────────────────────────────────── */
  .stTabs [data-baseweb="tab-list"] {
    background: #161b22 !important;
    border-radius: 8px;
    padding: 4px;
  }
  .stTabs [data-baseweb="tab"] {
    color: #8b949e !important;
  }
  .stTabs [aria-selected="true"] {
    background: #21262d !important;
    color: #f0f6fc !important;
  }

  /* ── AMÉLIORATION TABLEAUX (lisibilité tous écrans) ── */
  [data-testid="stDataFrame"] {
    border: 1px solid #30363d !important;
    box-shadow: 0 4px 16px rgba(0,0,0,0.3);
  }
  /* Headers plus visibles */
  [data-testid="stDataFrame"] [role="columnheader"] {
    background: linear-gradient(180deg, #2d4a87, #1F3864) !important;
    color: #ffffff !important;
    font-weight: 700 !important;
    font-size: 13px !important;
    text-transform: uppercase;
    letter-spacing: 0.5px;
    border-right: 1px solid #1a2942 !important;
    padding: 10px 8px !important;
  }
  /* Cellules avec meilleur contraste */
  [data-testid="stDataFrame"] [role="gridcell"] {
    color: #e6edf3 !important;
    border-color: #30363d !important;
    font-size: 13px !important;
    padding: 8px !important;
  }
  /* Lignes paires/impaires plus distinctes */
  [data-testid="stDataFrame"] [role="row"]:nth-child(odd) [role="gridcell"] {
    background: #161b22 !important;
  }
  [data-testid="stDataFrame"] [role="row"]:nth-child(even) [role="gridcell"] {
    background: #1c2128 !important;
  }
  /* Hover row highlight */
  [data-testid="stDataFrame"] [role="row"]:hover [role="gridcell"] {
    background: #2d4a87 !important;
    color: #ffffff !important;
  }
  /* Colonne nom agriculteur en gras blanc */
  [data-testid="stDataFrame"] [role="gridcell"]:nth-child(2) {
    font-weight: 600 !important;
    color: #ffffff !important;
  }

  /* ── Inputs / selects en dark ───────────────────────── */
  [data-baseweb="input"], [data-baseweb="select"] {
    background: #0d1117 !important;
  }
  input, textarea, select {
    background: #0d1117 !important;
    color: #e6edf3 !important;
  }

  /* ── Expanders ──────────────────────────────────────── */
  [data-testid="stExpander"] {
    background: #161b22 !important;
    border: 1px solid #21262d !important;
    border-radius: 8px;
  }

  /* ───────────────────────────────────────────────────── */
  /* ── RESPONSIVE MOBILE (largeur < 768px) ────────────── */
  /* ───────────────────────────────────────────────────── */
  @media (max-width: 768px) {
    .block-container { padding: 1rem 0.5rem !important; }
    h1 { font-size: 1.3rem !important; }
    h2 { font-size: 1.1rem !important; }
    h3 { font-size: 1rem !important; }

    /* KPI cards en colonne sur mobile */
    .metric-row {
      flex-direction: column !important;
      gap: 8px !important;
    }
    .kpi-box {
      min-width: 100% !important;
      padding: 10px 14px !important;
    }
    .kpi-val { font-size: 1.3rem !important; }
    .kpi-lbl { font-size: .65rem !important; }

    /* Tableaux scrollables horizontalement */
    [data-testid="stDataFrame"], [data-testid="stTable"] {
      overflow-x: auto !important;
      font-size: 11px !important;
    }
    [data-testid="stDataFrame"] [role="columnheader"] {
      font-size: 10px !important;
      padding: 4px !important;
    }
    [data-testid="stDataFrame"] [role="gridcell"] {
      font-size: 11px !important;
      padding: 4px !important;
    }

    /* Tabs compacts */
    .stTabs [data-baseweb="tab"] {
      font-size: 11px !important;
      padding: 6px 8px !important;
    }

    /* Boutons full-width */
    .stButton > button, .stDownloadButton > button {
      width: 100% !important;
      font-size: .8rem !important;
    }

    /* Sidebar plus compacte */
    [data-testid="stSidebar"] {
      width: 85% !important;
      min-width: 280px !important;
    }

    /* Plotly responsive */
    .js-plotly-plot, .plotly {
      width: 100% !important;
    }
  }

  /* ── Graphiques Plotly : réduire la sensibilité hover ──
     hovermode="closest" est moins sensible que x/x unified */
  .js-plotly-plot .hovertext {
    pointer-events: none;
  }
</style>

<script>
/* ── Réduire la sensibilité des graphiques Plotly ────────
   Patche les figures pour utiliser hoverdistance plus grand
   = il faut être PLUS proche d'un point pour déclencher tooltip */
function fixPlotlyHover() {
  document.querySelectorAll('.js-plotly-plot').forEach(function(el) {
    if (el._fullLayout && !el._hoverFixed) {
      el._hoverFixed = true;
      try {
        window.Plotly.relayout(el, {
          'hoverdistance': 50,       // pixels minimum pour tooltip (défaut 20)
          'spikedistance': -1,       // désactive le spike line ultra-sensible
        });
      } catch(e) {}
    }
  });
}
setInterval(fixPlotlyHover, 1500);
</script>
""", unsafe_allow_html=True)

# ── Supabase connection ──────────────────────────────────────
from supabase import create_client, Client

SCRIPT_DIR    = os.path.dirname(os.path.abspath(__file__))
PHASE4_SCRIPT = os.path.join(SCRIPT_DIR, "optimizer_v2.py")

PEAK_START = datetime.date(2026, 7, 1)
PEAK_END   = datetime.date(2026, 7, 15)

COMM_COLORS = {
    "ACHREF AJLANI":  "#8b5cf6",
    "FEDI":           "#3b82f6",
    "JILANI OBAY":    "#e8543a",
    "KHALIL":         "#f5a623",
    "MAKKI BEN SALAH":"#00e5a0",
}
FACTORY_COLORS = {
    "ABIDA":    "#ff6b9d",
    "COMOCAP":  "#3b82f6",
    "ELFALLEH": "#00e5a0",
    "SICAM":    "#f5a623",
    "TUCAL":    "#8b5cf6",
}

@st.cache_resource
def get_supabase() -> Client:
    """Create Supabase client once and reuse it."""
    try:
        url = st.secrets["supabase"]["url"]
        key = st.secrets["supabase"]["key"]
    except Exception:
        # Fallback: read from environment or hardcoded (local dev only)
        url = os.environ.get("SUPABASE_URL", "")
        key = os.environ.get("SUPABASE_KEY", "")
    if not url or not key:
        st.error("Supabase credentials not found. Add them to .streamlit/secrets.toml")
        st.stop()
    return create_client(url, key)

EXCEL_PHASE4 = os.path.join(SCRIPT_DIR, "Planning_Tomate_2026.xlsx")

def load_from_excel():
    """Load planning data directly from local Excel file."""
    if not os.path.exists(EXCEL_PHASE4):
        return None, None, None, None, None

    planning = pd.read_excel(EXCEL_PHASE4, sheet_name="Planning Journalier", header=0)
    planning["Date"]             = pd.to_datetime(planning["Date"],       errors="coerce")
    planning["Commercial"]       = planning.get("Commercial", pd.Series())
    planning["Agriculteur"]      = planning.get("Agriculteur", pd.Series())
    planning["Usine"]            = planning.get("Usine", pd.Series())
    planning["Région"]           = planning.get("Region", planning.get("Région", ""))
    planning["Accessibilité"]    = planning.get("Accessibilite", planning.get("Accessibilité", ""))
    planning["Tonnes/Jour"]      = pd.to_numeric(planning.get("Tonnes/Jour", 0), errors="coerce")
    planning["Type Véhicule"]    = planning.get("Type Vehicule", planning.get("Type Véhicule", ""))
    planning["Véhicules Requis"] = planning.get("Vehicules", planning.get("Véhicules Requis", ""))
    planning["Nb Voyages"]       = pd.to_numeric(planning.get("Nb Voyages", 0), errors="coerce")
    planning["Date Début"]       = pd.to_datetime(planning.get("Date Debut", planning.get("Date Début", None)), errors="coerce")
    planning["Date Fin"]         = pd.to_datetime(planning.get("Date Fin", None), errors="coerce")
    planning["Total Tonnes"]     = pd.to_numeric(planning.get("Total Tonnes", 0), errors="coerce")
    pic_col = "Pic de Recolte" if "Pic de Recolte" in planning.columns else "Pic de Récolte"
    planning["Pic de Récolte"]   = planning.get(pic_col, "").apply(lambda x: "🟡 PIC" if "PIC" in str(x).upper() else "")
    planning["Note"]             = planning.get("Note", "")
    planning = planning.dropna(subset=["Date"])

    transport = pd.read_excel(EXCEL_PHASE4, sheet_name="Besoins Transport-Jour", header=0)
    transport["Date"]      = pd.to_datetime(transport["Date"], errors="coerce")
    transport["Commercial"]= transport.get("Commercial", "")
    transport["Total Tonnes"] = pd.to_numeric(transport.get("Total Tonnes", 0), errors="coerce")
    transport["Voyages TRACTEUR"]      = pd.to_numeric(transport.get("Voyages TRACTEUR", 0),      errors="coerce").fillna(0)
    transport["Voyages PETIT POILOUR"] = pd.to_numeric(transport.get("Voyages PETIT POILOUR", 0), errors="coerce").fillna(0)
    transport["Voyages POILOUR"]       = pd.to_numeric(transport.get("Voyages POILOUR", 0),       errors="coerce").fillna(0)
    transport["Voyages SEMI"]          = pd.to_numeric(transport.get("Voyages SEMI", 0),          errors="coerce").fillna(0)
    transport["Jours Double"]          = pd.to_numeric(transport.get("Jours Double", 0),          errors="coerce").fillna(0)
    transport = transport.dropna(subset=["Date"])

    double_j = pd.DataFrame(columns=[
        "Commercial","Agriculteur A (finit tôt)","Agriculteur B (reçoit véhicule)",
        "Véhicule Partagé","Jours Économisés","Fin Orig. A","Nouvelle Fin A",
        "Début Orig. B","Nouveau Début B","Risque Maladie","Action Requise"
    ])

    return planning, transport, pd.DataFrame(), double_j, None

@st.cache_data(ttl=60)
def load_data(_sb_version: int = 0):
    """
    Load data — tries Supabase first, falls back to local Excel
    if Supabase has incomplete data (stops before August).
    """
    sb = get_supabase()
    DATA_SOURCE = "supabase"

    def fetch(table, order_col=None, limit=10000):
        """Récupère TOUTES les lignes avec pagination (Supabase limite à 1000/req)."""
        all_rows = []
        page_size = 1000
        offset = 0
        while True:
            q = sb.table(table).select("*")
            if order_col:
                q = q.order(order_col)
            batch = q.range(offset, offset + page_size - 1).execute().data
            if not batch:
                break
            all_rows.extend(batch)
            if len(batch) < page_size or len(all_rows) >= limit:
                break
            offset += page_size
        return pd.DataFrame(all_rows)

    # ── Planning ──
    planning = fetch("planning", order_col="date")
    if not planning.empty:
        planning["Date"]       = pd.to_datetime(planning["date"],       errors="coerce")
        planning["Commercial"] = planning["commercial"]
        # ✅ Ne garder que les commerciaux qui ont des données dans agriculteurs
        # Évite d'afficher de vieilles données d'anciens commerciaux
        try:
            _active_comms = sb.table("agriculteurs").select("commercial").execute().data
            _active_comms = list({r["commercial"] for r in _active_comms if r.get("commercial")})
            if _active_comms:
                planning = planning[planning["commercial"].isin(_active_comms)]
        except Exception:
            pass  # si erreur → garder tout
        planning["Agriculteur"]= planning["agriculteur"]
        planning["Usine"]      = planning["usine"]
        planning["Région"]     = planning["region"]
        planning["Accessibilité"] = planning["accessibilite"]
        planning["Tonnes/Jour"]= pd.to_numeric(planning["tonnes_jour"], errors="coerce")
        planning["Type Véhicule"] = planning["type_vehicule"]
        planning["Véhicules Requis"] = planning["vehicules"]
        planning["Nb Voyages"] = pd.to_numeric(planning["nb_voyages"], errors="coerce")
        planning["Date Début"] = pd.to_datetime(planning["date_debut"], errors="coerce")
        planning["Date Fin"]   = pd.to_datetime(planning["date_fin"],   errors="coerce")
        planning["Total Tonnes"] = pd.to_numeric(planning["total_tonnes"], errors="coerce")
        planning["Pic de Récolte"] = planning["pic"].apply(lambda x: "🟡 PIC" if x else "")
        planning["Note"]       = planning["note"]
        planning = planning.dropna(subset=["Date"])

        # ── Check if Supabase data is complete (must reach August) ──
        if planning["Date"].max().date() < datetime.date(2026, 7, 31):
            DATA_SOURCE = "excel_fallback"
            planning = pd.DataFrame()  # force fallback

    # ── Fallback to local Excel if Supabase is empty or incomplete ──
    if planning.empty and os.path.exists(EXCEL_PHASE4):
        DATA_SOURCE = "excel"
        result = load_from_excel()
        if result[0] is not None:
            planning, transport, dispo, double_j, _ = result
            # build resume from Supabase agriculteurs (not hardcoded)
            try:
                _agri = sb.table("agriculteurs").select("commercial,nom,tonnage_total").execute().data
                _adf  = pd.DataFrame(_agri) if _agri else pd.DataFrame()
                if not _adf.empty:
                    _adf["tonnage_total"] = pd.to_numeric(_adf["tonnage_total"], errors="coerce")
                    _adf = _adf[_adf["tonnage_total"] > 0]
                    # Filtrer lignes TOTAL et noms invalides
                    _nom_e = _adf["nom"].astype(str).str.strip().str.upper()
                    _adf = _adf[~_nom_e.str.startswith("TOTAL")]
                    _adf = _adf[~_nom_e.str.startswith("SOUS-TOTAL")]
                    _adf = _adf[_nom_e.str.len() > 2]
                    if not _adf.empty:
                        _grp  = _adf.groupby("commercial")
                        resume = pd.DataFrame({
                            "Commercial":            _grp["tonnage_total"].sum().round(0).index,
                            "Tonnes Totales Saison": _grp["tonnage_total"].sum().round(0).values,
                            "Nb Agriculteurs":       _grp["nom"].nunique().values,
                            "Conflits Résolus":      0,
                            "Total Jours Double":    0,
                        })
                    else:
                        resume = pd.DataFrame()
                else:
                    resume = pd.DataFrame()
            except Exception:
                resume = pd.DataFrame()
            # Store source for display
            st.session_state["data_source"] = "📁 Local Excel"
            return planning, transport, pd.DataFrame(), double_j, resume

    st.session_state["data_source"] = "🗄️ Supabase" if DATA_SOURCE == "supabase" else "📁 Excel (Supabase incomplet)"

    # ── Transport ──
    transport = fetch("transport", order_col="date")
    if not transport.empty:
        transport["Date"]      = pd.to_datetime(transport["date"], errors="coerce")
        transport["Commercial"]= transport["commercial"]
        transport["Total Tonnes"] = pd.to_numeric(transport["total_tonnes"], errors="coerce")
        transport["Voyages TRACTEUR"]      = pd.to_numeric(transport["tracteur"],      errors="coerce").fillna(0)
        transport["Voyages PETIT POILOUR"] = pd.to_numeric(transport["petit_poilour"], errors="coerce").fillna(0)
        transport["Voyages POILOUR"]       = pd.to_numeric(transport["poilour"],       errors="coerce").fillna(0)
        transport["Voyages SEMI"]          = pd.to_numeric(transport["semi"],          errors="coerce").fillna(0)
        transport["Jours Double"]          = pd.to_numeric(transport["jours_double"],  errors="coerce").fillna(0)
        transport = transport.dropna(subset=["Date"])

    # ── Decalage (journal double transport) ──
    double_j = fetch("decalage")
    if not double_j.empty and "commercial" in double_j.columns:
        double_j["Commercial"] = double_j["commercial"]
        double_j["Agriculteur A (finit tôt)"]       = double_j.get("agriculteur_a", "")
        double_j["Agriculteur B (reçoit véhicule)"] = double_j.get("agriculteur_b", "")
        double_j["Véhicule Partagé"]  = double_j.get("vehicule", "")
        double_j["Jours Économisés"]  = double_j.get("shift_jours", 0)
        double_j["Fin Orig. A"]       = pd.to_datetime(double_j.get("fin_orig_a"),       errors="coerce")
        double_j["Nouvelle Fin A"]    = pd.to_datetime(double_j.get("nouvelle_fin_a"),   errors="coerce")
        double_j["Début Orig. B"]     = pd.to_datetime(double_j.get("debut_orig_b"),     errors="coerce")
        double_j["Nouveau Début B"]   = pd.to_datetime(double_j.get("nouveau_debut_b"),  errors="coerce")
        double_j["Risque Maladie"]    = double_j.get("risque", "")
        double_j["Action Requise"]    = double_j.get("action", "")
        double_j = double_j.dropna(subset=["Commercial"])
    else:
        # Empty table (optimizer handles conflicts internally — no decalage rows)
        double_j = pd.DataFrame(columns=[
            "Commercial","Agriculteur A (finit tôt)","Agriculteur B (reçoit véhicule)",
            "Véhicule Partagé","Jours Économisés","Fin Orig. A","Nouvelle Fin A",
            "Début Orig. B","Nouveau Début B","Risque Maladie","Action Requise"
        ])

    # ── Resume par commercial — from agriculteurs table (real uploaded data) ──
    resume = pd.DataFrame()
    try:
        agri_raw = fetch("agriculteurs")
        if not agri_raw.empty and "commercial" in agri_raw.columns:
            agri_raw["tonnage_total"] = pd.to_numeric(agri_raw["tonnage_total"], errors="coerce")
            agri_raw = agri_raw[agri_raw["tonnage_total"] > 0]
            # Filtrer les lignes TOTAL et noms invalides
            _nom_r = agri_raw["nom"].astype(str).str.strip().str.upper()
            agri_raw = agri_raw[~_nom_r.str.startswith("TOTAL")]
            agri_raw = agri_raw[~_nom_r.str.startswith("SOUS-TOTAL")]
            agri_raw = agri_raw[_nom_r.str.len() > 2]
            if not agri_raw.empty:
                agri_grp = agri_raw.groupby("commercial")
                resume   = agri_grp["tonnage_total"].sum().reset_index()
                resume.columns = ["Commercial", "Tonnes Totales Saison"]
                resume["Tonnes Totales Saison"] = resume["Tonnes Totales Saison"].round(0)
                resume["Nb Agriculteurs"] = agri_grp["nom"].nunique().values
                resume["Conflits Résolus"]   = 0
                resume["Total Jours Double"] = 0
    except Exception:
        pass

    # dispo not in Supabase yet — return empty
    dispo = pd.DataFrame()

    return planning, transport, dispo, double_j, resume

# ── Load data from Supabase ───────────────────────────────────
# sb_refresh_counter increments when user clicks "Régénérer"
if "sb_refresh" not in st.session_state:
    st.session_state["sb_refresh"] = 0

result = load_data(_sb_version=st.session_state["sb_refresh"])
planning, transport, dispo, double_j, resume = result
orig = None  # not needed for dashboard display

# ✅ CORRECTION POST-CHARGEMENT : ACHREF = 100% SEMI
# Ses agriculteurs sont à Gafsa/Kasserine (accessibilité SEMI uniquement)
# Si Supabase contient encore des lignes PL (vieux runs), on les corrige ici
if not planning.empty and "Type Véhicule" in planning.columns and "Commercial" in planning.columns:
    _ach_mask = planning["Commercial"].astype(str).str.upper().str.contains("ACHREF", na=False)
    if _ach_mask.any():
        planning.loc[_ach_mask, "Type Véhicule"] = planning.loc[_ach_mask, "Type Véhicule"].apply(
            lambda v: "SEMI" if str(v).strip().upper() in ("PL","PPL","TRACTEUR","PL/PPL") else v
        )

# ── GLOBAL CONSTANTS — read from Supabase agriculteurs ──────
@st.cache_data(ttl=30)
def load_global_stats(_version: int = 0):
    """
    Charge les stats globales depuis Supabase.
    ✅ Écrase les anciennes données avec les nouvelles uploads
    ✅ Normalise les régions incohérentes (nabeul→CAP BON 2, beja→NORD…)
    ✅ Filtre les données corrompues (tonnage=0, dates invalides)
    """
    try:
        sb = get_supabase()

        # Commerciaux qui ont uploadé via le dashboard
        try:
            depot_data = sb.table("depot_status").select(
                "commercial,statut,depose_le").execute().data
            deposited  = {r["commercial"]: r.get("depose_le","")
                          for r in depot_data if r.get("statut") == "depose"}
        except Exception:
            deposited = {}

        # Toutes les données agriculteurs
        data = sb.table("agriculteurs").select(
            "commercial,nom,tonnage_total,usine,region,zone,date_debut,nbr_hectares,centre").execute().data
        df = pd.DataFrame(data) if data else pd.DataFrame()
        if df.empty:
            raise ValueError("empty")

        df["tonnage_total"] = pd.to_numeric(df["tonnage_total"], errors="coerce")
        df = df[df["tonnage_total"] > 0]   # filtre tonnage nul
        # Filtrer les lignes de TOTAL qui ont pu être insérées par erreur
        _nom = df["nom"].astype(str).str.strip().str.upper()
        df = df[~_nom.str.startswith("TOTAL")]
        df = df[~_nom.str.startswith("SOUS-TOTAL")]
        df = df[_nom.str.len() > 2]

        # Normalisation régions (évite les doublons nabeul/NABEUL/beja/MANOUBA)
        REGION_NORM = {
            "nabeul": "CAP BON 2", "NABEUL": "CAP BON 2",
            "beja": "NORD",        "BEJA": "NORD",
            "manouba": "NORD",     "MANOUBA": "NORD",
            "gafsa": "GAFSA / KASSRINE", "GAFSA": "GAFSA / KASSRINE",
            "kassrine": "GAFSA / KASSRINE", "KASSRINE": "GAFSA / KASSRINE",
            "capb1": "CAP BON 1",  "CAPB1": "CAP BON 1",
            "capb2": "CAP BON 2",  "CAPB2": "CAP BON 2",
        }
        df["region"] = df["region"].fillna("").astype(str).str.strip()
        df["region"] = df["region"].replace(REGION_NORM)

        # Filtre dates invalides — permissif pour ne pas perdre des données valides
        # (On garde tout, les dates invalides ne suppriement pas les tonnages)
        df["date_debut"] = pd.to_datetime(df["date_debut"], errors="coerce")
        # Supprimer SEULEMENT les lignes avec des dates vraiment absurdes (avant 2000)
        _bad_dates = df["date_debut"].notna() & (df["date_debut"].dt.year < 2000)
        if _bad_dates.sum() < len(df):   # ne pas tout supprimer
            df = df[~_bad_dates]

        # Priorité upload récent : si un commercial a uploadé,
        # ses données Supabase sont les plus récentes — garder toutes
        # (le delete() lors de l'upload a déjà écrasé l'ancien)
        df_use = df

        return {
            "total_tons":          round(df_use["tonnage_total"].sum(), 0),
            "n_farmers":           int(df_use["nom"].nunique()),
            "commercial_tons":     df_use.groupby("commercial")["tonnage_total"].sum().round(0).to_dict(),
            "commercial_farmers":  df_use.groupby("commercial")["nom"].nunique().to_dict(),
            "usine_tons":          df_use.groupby("usine")["tonnage_total"].sum().round(0).to_dict(),
            "region_tons":         df_use.groupby("region")["tonnage_total"].sum().round(0).to_dict(),
            "deposited":           list(deposited.keys()),
            "total_rows":          len(df_use),
            "df_agri":             df_use,   # ← DataFrame complet partagé globalement
            "data_quality": {
                "regions_ok":   df_use["region"].isin([
                    "CAP BON 1","CAP BON 2","NORD","GAFSA / KASSRINE",
                    "KAIROUAN","SIDI BOUZID","BOUFICHA"
                ]).sum(),
                "total_rows":   len(df_use),
            }
        }
    except Exception as e:
        return {
            "total_tons": 0.0, "n_farmers": 0,
            "commercial_tons": {}, "commercial_farmers": {},
            "usine_tons": {}, "region_tons": {},
            "deposited": [], "total_rows": 0,
            "df_agri": pd.DataFrame(),   # DataFrame vide en cas d'erreur
            "data_quality": {"regions_ok": 0, "total_rows": 0},
        }

_stats = load_global_stats(_version=st.session_state["sb_refresh"])
GLOBAL_TOTAL_TONS         = _stats["total_tons"]
GLOBAL_N_FARMERS          = _stats["n_farmers"]
AGRI_DF                   = _stats.get("df_agri", pd.DataFrame())  # DataFrame global agriculteurs
GLOBAL_COMMERCIAL_TONS    = _stats["commercial_tons"]
GLOBAL_COMMERCIAL_FARMERS = _stats["commercial_farmers"]
GLOBAL_USINE_TONS         = _stats["usine_tons"]

GLOBAL_PEAK_TONS   = round(planning[
    (planning["Date"].dt.date >= PEAK_START) &
    (planning["Date"].dt.date <= PEAK_END)
]["Tonnes/Jour"].sum(), 0) if not planning.empty else 0
GLOBAL_N_CONFLICTS = len(double_j) if not double_j.empty else 0

# ── Apply role-based data filtering ──────────────────────────
if planning is not None and CURRENT_FILTER:
    if CURRENT_ROLE == "commercial":
        if not planning.empty and "Commercial" in planning.columns:
            planning = planning[planning["Commercial"] == CURRENT_FILTER]
        if not transport.empty and "Commercial" in transport.columns:
            transport = transport[transport["Commercial"] == CURRENT_FILTER]
        if dispo is not None and not dispo.empty and "Commercial" in dispo.columns:
            dispo = dispo[dispo["Commercial"] == CURRENT_FILTER]
        if double_j is not None and not double_j.empty and "Commercial" in double_j.columns:
            double_j = double_j[double_j["Commercial"] == CURRENT_FILTER]
        if resume is not None and not resume.empty and "Commercial" in resume.columns:
            resume = resume[resume["Commercial"] == CURRENT_FILTER]
    elif CURRENT_ROLE == "usine":
        if not planning.empty and "Usine" in planning.columns:
            planning = planning[planning["Usine"] == CURRENT_FILTER]
        # FIX: ne pas vider transport pour usine — garder tout le transport
        # (le transport est par commercial, pas filtrable par usine sans perte info)
        # avant: transport.iloc[0:0] rendait les courbes transport VIDES
        double_j = double_j.iloc[0:0]

# ── Sidebar ───────────────────────────────────────────────────
with st.sidebar:
    # ── User info + logout ────────────────────────────────────
    role_icons = {"directeur": "👑", "commercial": "👤", "usine": "🏭"}
    role_labels = {"directeur": "Directeur", "commercial": "Commercial", "usine": "Usine"}
    st.markdown(f"""
    <div style='background:#1c2333;border:1px solid #21262d;border-radius:10px;
    padding:14px 16px;margin-bottom:16px'>
      <div style='font-size:1.1rem;font-weight:700;color:#f0f6fc'>
        {role_icons.get(CURRENT_ROLE,"👤")} {CURRENT_NAME}
      </div>
      <div style='font-size:.72rem;color:#8b949e;margin-top:3px;text-transform:uppercase;letter-spacing:.07em'>
        {role_labels.get(CURRENT_ROLE, CURRENT_ROLE)}
      </div>
    </div>
    """, unsafe_allow_html=True)
    if st.button("🚪 Déconnexion", use_container_width=True):
        for key in list(st.session_state.keys()):
            del st.session_state[key]
        # Effacer le token de l'URL pour ne pas rester connecté au refresh
        try:
            st.query_params.clear()
        except Exception:
            pass
        st.rerun()
    st.divider()


    # Régénérer planning — directeur seulement
    if CURRENT_ROLE == "directeur":
        st.subheader("⚙️ Mise à jour données")
        st.caption("Recalcule le planning et met à jour Supabase.")
        if st.button("🔄 Régénérer le planning", use_container_width=True, type="primary"):
            # ✅ SÉCURITÉ: vérifier qu'aucune régénération n'est en cours
            _regen_locked = False
            try:
                _sb_lock = get_supabase()
                _lock = _sb_lock.table("app_locks").select("*").eq(
                    "lock_name", "planning_regen").execute().data
                import datetime as _dt
                if _lock and _lock[0].get("expires_at"):
                    _exp = _dt.datetime.fromisoformat(
                        _lock[0]["expires_at"].replace("Z","+00:00"))
                    if _exp > _dt.datetime.now(_dt.timezone.utc):
                        _regen_locked = True
                        _by = _lock[0].get("locked_by","?")
                        st.warning(f"⚠️ Régénération déjà en cours par {_by}. Réessaye dans 2 min.")
            except Exception:
                pass  # Si table n'existe pas, continuer

            if not _regen_locked and os.path.exists(PHASE4_SCRIPT):
                # Poser le verrou
                try:
                    import datetime as _dt2
                    _exp_time = (_dt2.datetime.now(_dt2.timezone.utc) + 
                                 _dt2.timedelta(minutes=10)).isoformat()
                    get_supabase().table("app_locks").upsert({
                        "lock_name":  "planning_regen",
                        "locked_by":  CURRENT_USER,
                        "locked_at":  _dt2.datetime.now(_dt2.timezone.utc).isoformat(),
                        "expires_at": _exp_time,
                    }).execute()
                except Exception:
                    pass
                
                migrate_script = os.path.join(SCRIPT_DIR, "migrate.py")
                with st.spinner("Etape 1/2 : Calcul du planning..."):
                    r1 = subprocess.run(
                        [sys.executable, PHASE4_SCRIPT],
                        capture_output=True, text=True, timeout=600,
                        cwd=SCRIPT_DIR, encoding="utf-8", errors="replace",
                    )
                if r1.returncode != 0:
                    st.error("Erreur optimizer_v2.py :")
                    st.code(r1.stderr[-400:], language="text")
                elif os.path.exists(migrate_script):
                    with st.spinner("Etape 2/2 : Mise à jour Supabase..."):
                        r2 = subprocess.run(
                            [sys.executable, migrate_script],
                            capture_output=True, text=True, timeout=300,
                            cwd=SCRIPT_DIR, encoding="utf-8", errors="replace",
                        )
                    if r2.returncode == 0:
                        # Libérer le verrou
                        try:
                            get_supabase().table("app_locks").delete().eq(
                                "lock_name", "planning_regen").execute()
                            # Audit log
                            get_supabase().table("audit_log").insert({
                                "user_name": CURRENT_USER,
                                "user_role": CURRENT_ROLE,
                                "action":    "planning_regenerated",
                                "details":   f"Planning régénéré avec succès",
                            }).execute()
                        except Exception:
                            pass
                        st.success("✅ Planning recalculé et Supabase mis à jour !")
                        st.session_state["sb_refresh"] += 1
                        st.cache_data.clear()
                        st.rerun()
                    else:
                        st.error("Erreur migrate.py :")
                        st.code(r2.stderr[-400:], language="text")
                else:
                    st.success("✅ optimizer_v2.py terminé (migrate.py non trouvé)")
                    st.cache_data.clear()
                    st.rerun()
            else:
                st.warning(f"❌ optimizer_v2.py introuvable dans : {SCRIPT_DIR}")
        st.divider()

    # Filters — directeur sees all, commercial/usine see only their data
    st.subheader("🔍 Filtres")
    if planning is not None and not planning.empty:
        if CURRENT_ROLE == "directeur":
            comms_all = sorted(planning["Commercial"].dropna().unique())
            sel_comms = st.multiselect("Commercial(s)", comms_all, default=comms_all)
            if "Usine" in planning.columns:
                facts_all = sorted(planning["Usine"].dropna().unique())
                sel_facts = st.multiselect("Usine(s)", facts_all, default=facts_all)
            else:
                sel_facts = []
        else:
            # Commercial or usine: filter already applied above, no choice
            sel_comms = list(planning["Commercial"].dropna().unique()) if "Commercial" in planning.columns else []
            sel_facts = list(planning["Usine"].dropna().unique()) if "Usine" in planning.columns else []
            if CURRENT_ROLE == "commercial":
                st.info(f"👤 Vue filtrée : **{CURRENT_NAME}**")
            elif CURRENT_ROLE == "usine":
                st.info(f"🏭 Vue filtrée : **{CURRENT_NAME}**")

        # Use full season range — not limited by what's in Supabase
        d_min = datetime.date(2026, 6, 15)
        d_max = datetime.date(2026, 8, 31)
        # Default value: actual data range from planning
        d_data_min = planning["Date"].min().date() if not planning.empty else d_min
        d_data_max = planning["Date"].max().date() if not planning.empty else d_max
        date_range = st.date_input("Période", value=(d_data_min, d_data_max),
                                   min_value=d_min, max_value=d_max)
        peak_only = st.checkbox("⚡ Pic seulement (1–15 Jul)")
    else:
        sel_comms, sel_facts = [], []
        date_range = None
        peak_only = False

    st.divider()

    # Fleet inventory — directeur only
    if CURRENT_ROLE == "directeur":
        st.subheader("🚛 Votre flotte")
        st.caption("ℹ️ Ces valeurs servent uniquement aux **alertes de capacité** dans l'onglet Transport — elles ne recalculent pas le planning. Pour modifier le planning, il faut relancer `optimizer_v2.py`.")
        fl_trac = st.number_input("TRACTEUR",       0, 20, 0)
        fl_ppl  = st.number_input("PETIT POILOUR", 0, 30, 3)
        fl_pl   = st.number_input("POILOUR",       0, 30, 6)
        fl_semi = st.number_input("SEMI",          0, 20, 4)
        st.divider()
    else:
        fl_trac, fl_ppl, fl_pl, fl_semi = 0, 3, 6, 4  # defaults, not shown

    # ── EXPORT ALL — ZIP or Excel ─────────────────────────────
    st.subheader("📥 Exporter tout")
    st.caption("Télécharge toutes les 5 feuilles du planning Phase 4.")
    if planning is not None:
        export_sheets = {
            "Planning Journalier":      planning,
            "Transport-Jour":           transport,
            "Disponibilité Véhicules":  dispo,
            "Journal Décalage":         double_j,
            "Résumé Commercial":        resume,
        }
        # Excel export (all sheets in one file)
        xlsx_bytes = dfs_to_excel(export_sheets)
        st.download_button(
            "⬇️ Excel complet (5 feuilles)",
            data=xlsx_bytes,
            file_name="Tomate_Planning_Export.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
        # ZIP of CSVs
        zip_sheets = {f"{k.replace(' ','_')}.csv": v for k,v in export_sheets.items()}
        zip_bytes = dfs_to_zip(zip_sheets)
        st.download_button(
            "⬇️ ZIP de CSV (5 fichiers)",
            data=zip_bytes,
            file_name="Tomate_Planning_CSV.zip",
            mime="application/zip",
            use_container_width=True,
        )

    st.divider()
    if planning is not None and not planning.empty:
        last_date = planning["Date"].max().strftime("%d/%m/%Y")
        st.caption(f"📅 Données jusqu'au {last_date}")
        source = st.session_state.get("data_source", "🗄️ Supabase")
        st.caption(f"Source : {source}")

# ── Guard: Supabase empty ─────────────────────────────────────
if planning is None or planning.empty:
    st.markdown("---")
    if CURRENT_ROLE == "directeur":
        st.warning("⚠️ Aucune donnée de planning dans Supabase pour le moment.")
        st.markdown("### 🚀 Pour démarrer — suivez ces 3 étapes :")

        col1, col2, col3 = st.columns(3)
        with col1:
            st.markdown("""
            **Étape 1 — Commerciaux uploadent**
            - Chaque commercial se connecte
            - Onglet **📤 Upload Planning**
            - Télécharge le modèle Excel
            - Le remplit et l'uploade
            """)
        with col2:
            st.markdown("""
            **Étape 2 — Générer le planning**
            - Dans le terminal, lance :
            ```
            python optimizer_v2.py
            python migrate.py
            ```
            - Ou clique le bouton ci-dessous ↓
            """)
        with col3:
            st.markdown("""
            **Étape 3 — Rafraîchir**
            - Le dashboard se met à jour
            - Toutes les données s'affichent
            - Graphiques et statistiques OK
            """)

        st.markdown("---")

        # Allow directeur to trigger generation even when empty
        col_a, col_b = st.columns(2)
        with col_a:
            if st.button("🔄 Générer le planning maintenant",
                         type="primary", use_container_width=True):
                optimizer_script = os.path.join(SCRIPT_DIR, "optimizer_v2.py")
                migrate_script   = os.path.join(SCRIPT_DIR, "migrate.py")
                if os.path.exists(optimizer_script):
                    with st.spinner("Étape 1/2 : Calcul OR-Tools en cours (peut prendre 2 min)..."):
                        r1 = subprocess.run(
                            [sys.executable, optimizer_script],
                            capture_output=True, text=True, timeout=300,
                            cwd=SCRIPT_DIR, encoding="utf-8", errors="replace",
                        )
                    if r1.returncode != 0:
                        st.error("Erreur optimizer_v2.py :")
                        st.code(r1.stderr[-600:], language="text")
                    elif os.path.exists(migrate_script):
                        with st.spinner("Étape 2/2 : Mise à jour Supabase..."):
                            r2 = subprocess.run(
                                [sys.executable, migrate_script],
                                capture_output=True, text=True, timeout=300,
                                cwd=SCRIPT_DIR, encoding="utf-8", errors="replace",
                            )
                        if r2.returncode == 0:
                            st.success("✅ Planning généré et Supabase mis à jour !")
                            st.session_state["sb_refresh"] += 1
                            st.cache_data.clear()
                            st.rerun()
                        else:
                            st.error("Erreur migrate.py :")
                            st.code(r2.stderr[-400:], language="text")
                else:
                    st.error(f"optimizer_v2.py introuvable dans : {SCRIPT_DIR}")
        with col_b:
            if st.button("🔄 Rafraîchir les données",
                         use_container_width=True):
                st.session_state["sb_refresh"] += 1
                st.cache_data.clear()
                st.rerun()

        # Still show upload tab
        st.markdown("---")
        st.info("💡 Les commerciaux peuvent déposer leurs fichiers via l'onglet **📤 Upload Planning** même si le planning n'est pas encore généré.")

    elif CURRENT_ROLE == "commercial":
        st.info(f"👋 Bonjour **{CURRENT_NAME}** — Le planning n'est pas encore disponible.")
        st.markdown("Déposez votre fichier Excel via l'onglet **📤 Upload Planning** pour commencer.")

    elif CURRENT_ROLE == "usine":
        st.info(f"👋 Bonjour **{CURRENT_NAME}** — Le planning n'est pas encore disponible.")
        st.markdown("Contactez le directeur pour qu'il génère le planning.")

    # Show upload tab even when no planning data
    if CURRENT_ROLE in ("directeur", "commercial"):
        st.markdown("---")
        try:
            with st.expander("📤 Accéder à l'upload de planning", expanded=True):
                from upload_tab import render_upload_tab
                render_upload_tab(
                    sb=get_supabase(),
                    CURRENT_ROLE=CURRENT_ROLE,
                    CURRENT_NAME=CURRENT_NAME,
                    CURRENT_FILTER=CURRENT_FILTER,
                    GLOBAL_COMMERCIAL_FARMERS=GLOBAL_COMMERCIAL_FARMERS,
                    GLOBAL_COMMERCIAL_TONS=GLOBAL_COMMERCIAL_TONS,
                    df_to_csv=df_to_csv,
                )
        except Exception:
            pass
    st.stop()

# ── Apply filters ─────────────────────────────────────────────
p = planning[planning["Commercial"].isin(sel_comms)].copy()
t = transport[transport["Commercial"].isin(sel_comms)].copy()

# ── SOURCE UNIQUE: priorite Plans Rectifies > OR-Tools ──────
# Cascade automatiquement vers tous les tabs qui lisent 'p' ensuite.
if COMPARAISON_AVAILABLE:
    try:
        from comparaison_tab import build_effective_planning
        p = build_effective_planning(p, get_supabase())
    except Exception:
        pass

if "Usine" in p.columns and sel_facts:
    p = p[p["Usine"].isin(sel_facts)]

if isinstance(date_range, (list, tuple)) and len(date_range) == 2:
    d0, d1 = date_range
    p = p[(p["Date"].dt.date >= d0) & (p["Date"].dt.date <= d1)]
    t = t[(t["Date"].dt.date >= d0) & (t["Date"].dt.date <= d1)]

if peak_only:
    p = p[(p["Date"].dt.date >= PEAK_START) & (p["Date"].dt.date <= PEAK_END)]
    t = t[(t["Date"].dt.date >= PEAK_START) & (t["Date"].dt.date <= PEAK_END)]

# ── Header + KPIs ─────────────────────────────────────────────
st.title("🍅 Tomate Planning 2026")
st.caption("Phase 10 — Connecté à Supabase · Login par rôle · OR-Tools Optimizer")

# KPIs — GLOBAL_N_FARMERS uses unique names from agriculteurs table
total_tons   = GLOBAL_TOTAL_TONS

# Statut OR-Tools: lire depuis planning (si FEASIBLE = solution trouvée, sinon vide)
_n_planning_days = len(planning["Date"].dt.date.unique()) if not planning.empty else 0
if _n_planning_days >= 60:
    ortools_status = "FEASIBLE"
    ortools_sub    = f"{_n_planning_days} jours planifiés ✅"
elif _n_planning_days > 0:
    ortools_status = "PARTIEL"
    ortools_sub    = f"⚠️ {_n_planning_days}/67 jours — relancer optimizer"
else:
    ortools_status = "—"
    ortools_sub    = "Planning non généré"

# Tonnage planifié (somme planning) vs déclaré
planned_tons = round(planning["Tonnes/Jour"].sum(), 0) if not planning.empty else 0
ecart_pct    = round((planned_tons - total_tons) / total_tons * 100, 1) if total_tons > 0 else 0
total_trips  = int(p["Nb Voyages"].sum()) if not p.empty and "Nb Voyages" in p.columns else 0
# Unique farmer names — not rows (one farmer can deliver to 2 usines = 2 rows)
# Toujours utiliser le nombre réel d'agriculteurs déclarés (table agriculteurs)
# pas le nombre dans le planning (qui peut être incomplet ou filtré)
n_farmers = GLOBAL_N_FARMERS
peak_tons    = GLOBAL_PEAK_TONS
n_conflicts  = GLOBAL_N_CONFLICTS
dbl_days     = int(t["Jours Double"].sum()) if not t.empty and "Jours Double" in t.columns else 0

st.markdown(f"""
<div class="metric-row">
  <div class="kpi-box" style="--c:#e8543a"><div class="kpi-val">{int(total_tons):,} t</div><div class="kpi-lbl">Déclaré (saison)</div><div class="kpi-sub">{n_farmers} agriculteurs</div></div>
  <div class="kpi-box" style="--c:#22c55e"><div class="kpi-val">{int(planned_tons):,} t</div><div class="kpi-lbl">Planifié OR-Tools</div><div class="kpi-sub">Écart {ecart_pct:+.1f}% vs déclaré</div></div>
  <div class="kpi-box" style="--c:#f5a623"><div class="kpi-val">{int(peak_tons):,} t</div><div class="kpi-lbl">Tonnes période pic</div><div class="kpi-sub">1–15 Juillet</div></div>
  <div class="kpi-box" style="--c:#8b5cf6"><div class="kpi-val">{total_trips:,}</div><div class="kpi-lbl">Voyages (vue actuelle)</div><div class="kpi-sub">PPL + PL + SEMI</div></div>
  <div class="kpi-box" style="--c:#00e5a0"><div class="kpi-val">{ortools_status}</div><div class="kpi-lbl">Statut OR-Tools</div><div class="kpi-sub">{ortools_sub}</div></div>
  <div class="kpi-box" style="--c:#3b82f6"><div class="kpi-val">{len(sel_comms)}</div><div class="kpi-lbl">Commerciaux</div><div class="kpi-sub">{_n_planning_days} jours planifiés</div></div>
</div>
<div class="peak-box">⚡ <b>Pic 1–15 Juillet :</b> Les caps usines et commerciaux s'appliquent UNIQUEMENT pendant cette période. Hors pic: distribution libre selon les fenêtres de maturité.</div>
""", unsafe_allow_html=True)

# ── Warning planning incomplet (UNIQUEMENT pour directeur) ──
# Pour usine/commercial, le planning est NATURELLEMENT filtré :
#   - COMOCAP voit ~52j (ses jours de réception)
#   - JILANI voit ~27j (ses fermiers commencent Jul 23)
#   → Ce n'est PAS un planning incomplet, c'est la vue filtrée !
if CURRENT_ROLE == "directeur":
    if _n_planning_days > 0 and _n_planning_days < 60:
        st.warning(
            f"⚠️ **Planning incomplet** : seulement **{_n_planning_days} jours** dans Supabase "
            f"au lieu de 67. Les graphiques et calculs sont partiels. "
            f"**Solution :** Sidebar → 🔄 Régénérer le planning"
        )
    elif _n_planning_days == 0:
        st.error("❌ Aucun planning dans Supabase. Lance : `python optimizer_v2.py` puis `python migrate.py`")
else:
    # Pour usine/commercial : afficher seulement si VRAIMENT vide (= 0 jour)
    if _n_planning_days == 0:
        st.warning(f"⚠️ Pas de planning trouvé pour {CURRENT_NAME}. "
                   f"Demande au directeur de relancer optimizer + migrate.")

# ── Tabs — visibility depends on role ────────────────────────
# Usine    : voit seulement Par Usine + Transport (pas de commercial, décalage, historique, admin)
# Commercial: voit Planning + Par Commercial + Transport + Gestion de ses agriculteurs
# Directeur : voit tout

if CURRENT_ROLE == "centre":
    # ── Session CENTRE (BACCARA, KERKOUANE, 428) ──────────────
    # Rend un dashboard dédié et stoppe l'exécution du dashboard global
    try:
        from centre_tab import render_centre_dashboard
        render_centre_dashboard(get_supabase(), CURRENT_FILTER, CURRENT_NAME)
    except ImportError:
        st.error("Module centre_tab.py introuvable. Vérifier le déploiement.")
    except Exception as e:
        st.error(f"Erreur centre dashboard: {e}")
    st.stop()

if CURRENT_ROLE == "usine":
    tab3, tab4 = st.tabs([
        "🏭 Par Usine",
        "🚛 Transport & Alertes",
    ])
    tab1 = tab2 = tab3
    tab5 = tab6 = tab9 = tab10 = tab4
    tab7 = tab8 = tab4
    tab_agroeco = tab4  # usine : redirige vers Transport (tab non visible)

elif CURRENT_ROLE == "commercial":
    tab1, tab2, tab6, tab4, tab9, tab10, tab_comp, tab_agroeco = st.tabs([
        "📅 Planning Journalier",
        "👤 Par Commercial",
        "📈 Mes Prévisions 25/26",
        "🚛 Transport & Alertes",
        "🌾 Mes Agriculteurs",
        "📤 Upload Planning",
        "📊 Comparaison Plans",
        "📊 Dashboard Agroéco",
    ])
    tab3 = tab1
    tab5 = tab7 = tab8 = tab4

else:
    # Directeur : all tabs
    tab1, tab2, tab3, tab4, tab5, tab6, tab7, tab8, tab9, tab10, tab_comp, tab_agroeco = st.tabs([
        "📅 Planning Journalier",
        "👤 Par Commercial",
        "🏭 Par Usine",
        "🚛 Transport & Alertes",
        "⚙️ Décalage & Conflits",
        "📈 Comparaison par années",
        "🗺️ Tonnage par Région",
        "📊 Comparaison par prévision",
        "🌾 Gestion Agriculteurs",
        "📤 Upload Planning",
        "📊 Comparaison Plans",
        "📊 Dashboard Agroéco",
    ])

# ── TAB 1: DAILY PLANNING ────────────────────────────────────
with tab1:
    daily = p.groupby("Date")["Tonnes/Jour"].sum().reset_index()
    daily["Période"] = daily["Date"].apply(
        lambda d: "⚡ Pic (1-15 Jul)" if PEAK_START <= d.date() <= PEAK_END else "Normal"
    )

    fig = px.bar(
        daily, x="Date", y="Tonnes/Jour", color="Période",
        color_discrete_map={"⚡ Pic (1-15 Jul)":"#f5a623", "Normal":"#3b82f6"},
        title="Tonnes récoltées par jour — toute la saison",
        labels={"Tonnes/Jour":"Tonnes/jour"},
        template="plotly_dark",
    )
    fig.update_layout(
        plot_bgcolor="#0d1117", paper_bgcolor="#161b22",
        legend_title="Période", hovermode="closest", height=420,
        font_color="#8b949e",
    )
    fig.add_vrect(x0=str(PEAK_START), x1=str(PEAK_END),
                  fillcolor="gold", opacity=0.06, line_width=0,
                  annotation_text="⚡ PIC", annotation_position="top left",
                  annotation_font_color="#f5a623")
    st.plotly_chart(fig, use_container_width=True)

    c1, c2 = st.columns(2)
    with c1:
        # Use DECLARED tonnage from agriculteurs (correct) not planning sum (filtered/incomplete)
        if GLOBAL_COMMERCIAL_TONS:
            comm_df = pd.DataFrame(list(GLOBAL_COMMERCIAL_TONS.items()),
                                    columns=["Commercial","Tonnes/Jour"])
        else:
            comm_df = p.groupby("Commercial")["Tonnes/Jour"].sum().reset_index()
        fig2 = px.pie(
            comm_df,
            names="Commercial", values="Tonnes/Jour",
            title="Répartition par commercial (tonnage déclaré)", hole=0.45,
            color="Commercial",
            color_discrete_map=COMM_COLORS,
            template="plotly_dark",
        )
        fig2.update_layout(paper_bgcolor="#161b22", height=320)
        st.plotly_chart(fig2, use_container_width=True)
    with c2:
        if "Usine" in p.columns:
            # Use declared tonnage from agriculteurs (correct) not planning sum (wrong)
            if GLOBAL_USINE_TONS:
                usine_df = pd.DataFrame(list(GLOBAL_USINE_TONS.items()),
                                         columns=["Usine","Tonnes/Jour"])
            else:
                usine_df = p.groupby("Usine")["Tonnes/Jour"].sum().reset_index()
            fig3 = px.pie(
                usine_df,
                names="Usine", values="Tonnes/Jour",
                title="Répartition par usine", hole=0.45,
                color="Usine", color_discrete_map=FACTORY_COLORS,
                template="plotly_dark",
            )
            fig3.update_layout(paper_bgcolor="#161b22", height=320)
            st.plotly_chart(fig3, use_container_width=True)

    st.subheader("📋 Données détaillées")
    
    # Barre de recherche par nom d'agriculteur
    col_search1, col_search2 = st.columns([3, 1])
    with col_search1:
        search_farmer = st.text_input(
            "🔍 Rechercher un agriculteur",
            placeholder="Tapez le nom (ex: ZOUHAIR ou SOUHAIL)",
            key="search_planning_farmer"
        )
    with col_search2:
        st.write("")  # spacer
        st.caption(f"Total: {len(p)} lignes")
    
    display_cols = [c for c in ["Date","Commercial","Agriculteur","Usine",
                                "Tonnes/Jour","Type Véhicule","Véhicules Requis","Nb Voyages",
                                "Pic de Récolte","Note"] if c in p.columns]
    p_display = p[display_cols].sort_values("Date").reset_index(drop=True)
    
    # ✅ NOUVEAU : Ajouter colonnes "Disponibles" et "Manquants" pour chaque ligne
    # Permet de voir d'un coup d'œil combien de véhicules manquent à louer
    _FLEET_AVAILABILITY = {
        # Source: transport_etat_final.xlsx (nb de véhicules confirmés par usine)
        "SICAM":    {"PL":48, "PPL":6,  "SEMI":13, "TRACTEUR":0},
        "TUCAL":    {"PL":17, "PPL":0,  "SEMI":2,  "TRACTEUR":0, "PL_bourak":6},   # +6 PL via BOURAK
        "COMOCAP":  {"PL":6,  "PPL":14, "SEMI":3,  "TRACTEUR":10, "PL_luimeme":3, "PPL_luimeme":4},  # +3PL +4PPL via LUIMEME
        "ABIDA":    {"PL":1,  "PPL":0,  "SEMI":2,  "TRACTEUR":0},
        "ELFALLEH": {"PL":0,  "PPL":2,  "SEMI":0,  "TRACTEUR":0},
    }
    
    def _compute_availability(row):
        """Calcule disponibles + manquants pour une ligne du planning."""
        usine = str(row.get("Usine","")).upper().strip()
        veh_type = str(row.get("Type Véhicule","")).upper().strip()
        requis = pd.to_numeric(row.get("Véhicules Requis", 0), errors="coerce")
        if pd.isna(requis):
            requis = 0
        requis = int(requis)
        
        # Récupérer la dispo pour cette usine + ce véhicule
        fleet = _FLEET_AVAILABILITY.get(usine, {})
        dispo = fleet.get(veh_type, 0)
        # Ajouter les jokers selon le véhicule
        if veh_type == "PL":
            dispo += fleet.get("PL_bourak", 0) + fleet.get("PL_luimeme", 0)
        elif veh_type == "PPL":
            dispo += fleet.get("PPL_luimeme", 0)
        
        manque = max(0, requis - dispo)
        return pd.Series({
            "Disponibles": dispo if requis > 0 else 0,
            "Manquants (à louer)": manque,
        })
    
    # Appliquer le calcul si les colonnes existent
    if "Véhicules Requis" in p_display.columns and "Type Véhicule" in p_display.columns and "Usine" in p_display.columns:
        _availability = p_display.apply(_compute_availability, axis=1)
        p_display["Disponibles"] = _availability["Disponibles"].astype(int)
        p_display["Manquants (à louer)"] = _availability["Manquants (à louer)"].astype(int)
        # Réordonner les colonnes : placer Disponibles + Manquants juste après "Véhicules Requis"
        new_order = []
        for c in p_display.columns:
            new_order.append(c)
            if c == "Véhicules Requis":
                if "Disponibles" not in new_order:
                    new_order.extend(["Disponibles", "Manquants (à louer)"])
        # Dédoublonner en gardant l'ordre
        seen = set()
        new_order = [c for c in new_order if not (c in seen or seen.add(c))]
        p_display = p_display[new_order]
    
    # Appliquer le filtre de recherche
    if search_farmer.strip():
        mask = p_display["Agriculteur"].astype(str).str.upper().str.contains(
            search_farmer.upper().strip(), na=False
        )
        p_display = p_display[mask]
        if len(p_display) == 0:
            st.warning(f"Aucun agriculteur trouvé pour '{search_farmer}'")
        else:
            total_tons_found = p_display["Tonnes/Jour"].sum()
            n_unique = p_display["Agriculteur"].nunique()
            st.success(f"✅ {n_unique} agriculteur(s) trouvé(s) — {len(p_display)} lignes — {total_tons_found:,.0f}t")
    
    st.dataframe(
        p_display,
        use_container_width=True, height=280,
    )
    # ✅ Exporter p_display (qui contient Disponibles + Manquants), pas l'original
    _planning_export = p_display.copy()
    col_dl0, col_dl1, col_dl2, col_dl3 = st.columns(4)
    with col_dl0:
        # Export RECTIFIE — base sur le planning EFFECTIF (p, deja corrige par build_effective_planning)
        if COMPARAISON_AVAILABLE:
            try:
                from comparaison_tab import generate_planning_wide_excel
                st.download_button(
                    "📆 Planning RECTIFIÉ (colonnes dates)",
                    data=generate_planning_wide_excel(_planning_export),
                    file_name="planning_rectifie_colonnes_dates.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary",
                    use_container_width=True,
                    help="Colonnes: Date|Commercial|Agriculteur|Usine|...|20/06|21/06|...|25/08 — Source: Plans Rectifies (priorite) + OR-Tools",
                )
            except Exception as _e:
                st.caption(f"Export rectifie indisponible: {_e}")
        else:
            st.info("comparaison_tab.py requis pour export rectifié")
    with col_dl1:
        st.download_button(
            "📅 Excel SÉPARÉ PAR JOUR",
            data=df_to_xlsx_by_day(_planning_export, date_col="Date"),
            file_name="planning_par_jour.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            help="Un onglet par jour + un onglet récapitulatif",
        )
    with col_dl2:
        st.download_button(
            "📊 Excel TOUT EN 1 ONGLET",
            data=df_to_xlsx_styled(_planning_export, sheet_name="Planning"),
            file_name="planning_journalier.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    with col_dl3:
        st.download_button(
            "⬇️ CSV brut",
            data=df_to_csv(_planning_export),
            file_name="planning_journalier.csv",
            mime="text/csv",
            use_container_width=True,
        )

# ── TAB 2: PAR COMMERCIAL ────────────────────────────────────
with tab2:
    # Line chart per commercial
    comm_daily = (p.groupby(["Date","Commercial"])["Tonnes/Jour"]
                  .sum().reset_index())
    fig4 = px.line(
        comm_daily, x="Date", y="Tonnes/Jour", color="Commercial",
        color_discrete_map=COMM_COLORS,
        title="Tonnes/jour par commercial",
        labels={"Tonnes/Jour":"Tonnes/j"},
        template="plotly_dark",
    )
    fig4.update_traces(line_width=2)
    fig4.update_layout(
        plot_bgcolor="#0d1117", paper_bgcolor="#161b22",
        height=400, hovermode="closest",
    )
    fig4.add_vrect(x0=str(PEAK_START), x1=str(PEAK_END),
                   fillcolor="gold", opacity=0.06, line_width=0)
    st.plotly_chart(fig4, use_container_width=True)

    c1, c2, c3 = st.columns(3)
    with c1:
        # Use DECLARED tonnage from agriculteurs table (source of truth)
        # NOT planning sum which is filtered/partial
        if GLOBAL_COMMERCIAL_TONS:
            comm_tot = pd.DataFrame(list(GLOBAL_COMMERCIAL_TONS.items()),
                                     columns=["Commercial","Tonnes/Jour"])
        else:
            comm_tot = p.groupby("Commercial")["Tonnes/Jour"].sum().reset_index()
        fig5 = px.bar(
            comm_tot, x="Commercial", y="Tonnes/Jour",
            color="Commercial", color_discrete_map=COMM_COLORS,
            title="Tonnes totales (déclarées par chaque commercial)",
            template="plotly_dark",
            text_auto=".3s",
        )
        fig5.update_layout(paper_bgcolor="#161b22", showlegend=False, height=300)
        st.plotly_chart(fig5, use_container_width=True)
    with c2:
        # Use DECLARED unique farmer count from agriculteurs table
        if GLOBAL_COMMERCIAL_FARMERS:
            farmers_ct = pd.DataFrame(list(GLOBAL_COMMERCIAL_FARMERS.items()),
                                       columns=["Commercial","Agriculteur"])
        else:
            farmers_ct = p.groupby("Commercial")["Agriculteur"].nunique().reset_index()
        fig6 = px.bar(
            farmers_ct, x="Commercial", y="Agriculteur",
            color="Commercial", color_discrete_map=COMM_COLORS,
            title="Nb agriculteurs (déclarés)", template="plotly_dark",
            text_auto=True,
        )
        fig6.update_layout(paper_bgcolor="#161b22", showlegend=False, height=300)
        st.plotly_chart(fig6, use_container_width=True)
    with c3:
        if not resume.empty:
            fig7 = px.bar(
                resume, x="Commercial", y="Conflits Résolus",
                color="Commercial", color_discrete_map=COMM_COLORS,
                title="Conflits résolus", template="plotly_dark",
            )
            fig7.update_layout(paper_bgcolor="#161b22", showlegend=False, height=300)
            st.plotly_chart(fig7, use_container_width=True)

    # Per-commercial drill-down
    st.subheader("🔎 Détail par commercial")
    selected = st.selectbox("Choisir un commercial", sel_comms)
    one = p[p["Commercial"] == selected]
    one_daily = one.groupby("Date")["Tonnes/Jour"].sum().reset_index()
    fig8 = px.area(
        one_daily, x="Date", y="Tonnes/Jour",
        title=f"Tonnes/jour — {selected}",
        color_discrete_sequence=[COMM_COLORS.get(selected,"#3b82f6")],
        template="plotly_dark",
    )
    fig8.update_layout(paper_bgcolor="#161b22", height=280)
    fig8.add_vrect(x0=str(PEAK_START), x1=str(PEAK_END),
                   fillcolor="gold", opacity=0.08, line_width=0)
    st.plotly_chart(fig8, use_container_width=True)

    show_cols = [c for c in ["Date","Agriculteur","Usine","Tonnes/Jour",
                              "Type Véhicule","Véhicules Requis","Nb Voyages","Note"] if c in one.columns]
    st.dataframe(one[show_cols].sort_values("Date").reset_index(drop=True),
                 use_container_width=True, height=240)
    
    # ── 📊 NOUVEAU: Courbes Agriculteurs × Tonnage pour ce commercial ─
    st.markdown("---")
    st.subheader(f"📈 Agriculteurs de {selected} — Répartition des tonnages")
    
    if "Agriculteur" in one.columns and "Tonnes/Jour" in one.columns:
        # 1. Tonnage DÉCLARÉ depuis AGRI_DF (chargé une seule fois au démarrage)
        # AGRI_DF = table agriculteurs complète en mémoire — fiable, rapide, pas de pagination
        if not AGRI_DF.empty and "commercial" in AGRI_DF.columns:
            _comm_mask = AGRI_DF["commercial"] == selected
            _sel_agri  = AGRI_DF[_comm_mask].copy()
            _sel_agri["tonnage_total"] = pd.to_numeric(
                _sel_agri["tonnage_total"], errors="coerce").fillna(0)
            # ✅ FIX: Grouper par (nom, usine, accessibilite) pour distinguer les lots
            # d'un même agriculteur (ex: AMOR KHECHIN RM-SICAM vs PL-TUCAL)
            _has_usine = "usine" in _sel_agri.columns
            _has_acc   = "accessibilite" in _sel_agri.columns
            _grp_cols = ["nom"]
            if _has_usine:
                _grp_cols.append("usine")
            if _has_acc:
                _grp_cols.append("accessibilite")
            agri_totals = _sel_agri.groupby(_grp_cols, dropna=False)["tonnage_total"].sum().reset_index()
            
            # Construire un nom d'affichage avec suffixe si nécessaire
            # Pré-calculer le nombre de lots par nom (1 seul groupby au lieu d'un par ligne)
            if _has_usine and _has_acc:
                _lots_per_name = (_sel_agri.groupby("nom", dropna=False)
                                  .apply(lambda g: g[["usine","accessibilite"]].drop_duplicates().shape[0])
                                  .to_dict())
            else:
                _lots_per_name = {}
            
            def _make_display_name(row):
                base = str(row.get("nom", ""))
                n_lots = _lots_per_name.get(base, 1)
                if n_lots > 1 and _has_usine and _has_acc:
                    acc = str(row.get("accessibilite", "")).upper() or "?"
                    us  = str(row.get("usine", "")).upper() or "?"
                    return f"{base} ({acc}-{us})"
                return base
            
            agri_totals["Agriculteur"] = agri_totals.apply(_make_display_name, axis=1)
            agri_totals = agri_totals[["Agriculteur", "tonnage_total"]].rename(
                columns={"tonnage_total": "Tonnage Total (t)"})
            # ✅ Trier par tonnage DÉCROISSANT (plus gros en haut visuellement)
            agri_totals = agri_totals.sort_values("Tonnage Total (t)", ascending=True)
            agri_totals["Tonnage Total (t)"] = agri_totals["Tonnage Total (t)"].round(0).astype(int)
            
            # ✅ Ajouter NBR_HECTARES et t/ha
            if "nbr_hectares" in _sel_agri.columns:
                ha_map = _sel_agri.groupby("nom")["nbr_hectares"].first()
                _ha = pd.to_numeric(
                    agri_totals["Agriculteur"].apply(
                        lambda x: ha_map.get(x.split(" (")[0], 0)),
                    errors="coerce")
                agri_totals["Hectares"] = _ha.round(2)
                agri_totals["t/ha"] = (agri_totals["Tonnage Total (t)"] / _ha.replace(0, pd.NA)).round(1)
        else:
            # Fallback uniquement si AGRI_DF vide (connexion échouée au démarrage)
            agri_totals = one.groupby("Agriculteur")["Tonnes/Jour"].sum().reset_index()
            agri_totals.columns = ["Agriculteur", "Tonnage Total (t)"]
            agri_totals = agri_totals.sort_values("Tonnage Total (t)", ascending=True)
            agri_totals["Tonnage Total (t)"] = agri_totals["Tonnage Total (t)"].round(0).astype(int)
        
        n_agri = len(agri_totals)
        # ✅ FIX: Hauteur plus grande pour TOUS afficher (35px par agri minimum)
        bar_height = max(400, n_agri * 35 + 100)
        
        col_a, col_b = st.columns([3, 1])
        with col_a:
            fig_bar = px.bar(
                agri_totals, x="Tonnage Total (t)", y="Agriculteur",
                orientation="h",
                title=f"Tonnage total par agriculteur — {selected} ({n_agri} agriculteurs)",
                color="Tonnage Total (t)", color_continuous_scale="Viridis",
                template="plotly_dark", height=bar_height,
                text="Tonnage Total (t)",
            )
            fig_bar.update_traces(
                textposition="outside",
                texttemplate="%{x:,.0f}t",   # ✅ Format: 1,234t
                textfont=dict(size=11, color="#f0f6fc"),
            )
            fig_bar.update_layout(
                paper_bgcolor="#161b22",
                xaxis_title="Tonnes (t)",
                yaxis_title="",
                showlegend=False,
                margin=dict(l=250, r=80, t=60, b=40),    # ✅ Plus de marge gauche pour noms longs
                yaxis=dict(automargin=True, tickfont=dict(size=11)),
                # ✅ Forcer affichage de TOUS les ticks (pas de skip)
                xaxis=dict(automargin=True),
            )
            # ✅ Forcer Plotly à afficher TOUS les noms (pas de troncature)
            fig_bar.update_yaxes(showticklabels=True, tickmode="array",
                                  tickvals=list(range(len(agri_totals))),
                                  ticktext=agri_totals["Agriculteur"].tolist())
            st.plotly_chart(fig_bar, use_container_width=True)
        
        with col_b:
            if agri_totals.empty or agri_totals['Tonnage Total (t)'].sum() == 0:
                st.info("Aucune donnée")
            else:
                st.metric("Nb agriculteurs", n_agri)
                st.metric("Tonnage total",   f"{int(agri_totals['Tonnage Total (t)'].sum()):,}t")
                st.metric("Moyenne / agri",  f"{int(agri_totals['Tonnage Total (t)'].mean()):,}t")
                st.metric("Plus gros",       f"{int(agri_totals['Tonnage Total (t)'].max()):,}t")
                st.caption(f"Top: **{agri_totals.iloc[-1]['Agriculteur']}**")
        
        # 2. Évolution temporelle TOP 8 agriculteurs (courbes superposées)
        st.markdown(f"#### 📉 Courbes journalières — TOP 8 agriculteurs de {selected}")
        top_agri = agri_totals.tail(8)["Agriculteur"].tolist()
        one_top = one[one["Agriculteur"].isin(top_agri)].copy()
        
        if not one_top.empty:
            fig_lines = px.line(
                one_top.sort_values("Date"), x="Date", y="Tonnes/Jour",
                color="Agriculteur", template="plotly_dark", height=380,
                title=f"Évolution journalière des tonnes — TOP 8 agriculteurs de {selected}",
                markers=True,
            )
            fig_lines.update_layout(
                paper_bgcolor="#161b22",
                hovermode="closest",
                legend=dict(orientation="h", yanchor="top", y=-0.15),
            )
            fig_lines.add_vrect(x0=str(PEAK_START), x1=str(PEAK_END),
                              fillcolor="gold", opacity=0.08, line_width=0,
                              annotation_text="PIC", annotation_position="top left")
            st.plotly_chart(fig_lines, use_container_width=True)
        
        # 3. Répartition par usine (donut)
        if "Usine" in one.columns:
            st.markdown(f"#### 🏭 Répartition par usine — {selected}")
            usine_dist = one.groupby("Usine")["Tonnes/Jour"].sum().reset_index()
            usine_dist["Tonnes/Jour"] = usine_dist["Tonnes/Jour"].round(0).astype(int)
            usine_dist = usine_dist.sort_values("Tonnes/Jour", ascending=False)
            
            col_pie, col_table = st.columns([1, 1])
            with col_pie:
                fig_pie = px.pie(
                    usine_dist, names="Usine", values="Tonnes/Jour",
                    hole=0.45, template="plotly_dark",
                    title=f"Tonnage par usine — {selected}",
                )
                fig_pie.update_traces(textinfo="label+percent",
                                       textposition="inside")
                fig_pie.update_layout(paper_bgcolor="#161b22", height=380)
                st.plotly_chart(fig_pie, use_container_width=True)
            
            with col_table:
                usine_dist["Part"] = (usine_dist["Tonnes/Jour"] / 
                                      usine_dist["Tonnes/Jour"].sum() * 100).round(1)
                usine_dist["Part"] = usine_dist["Part"].astype(str) + "%"
                usine_dist["Tonnes/Jour"] = usine_dist["Tonnes/Jour"].apply(lambda x: f"{x:,}t")
                st.markdown("**Détail par usine**")
                st.dataframe(usine_dist.rename(columns={"Tonnes/Jour": "Tonnage"}),
                            use_container_width=True, hide_index=True)

    # ══════════════════════════════════════════════════════════════
    # 📅 TABLEAU JOURNALIER PAR AGRICULTEUR — vue calendrier pivot
    # ══════════════════════════════════════════════════════════════
    st.markdown("---")
    st.subheader(f"📅 Planning journalier détaillé — {selected}")
    
    # ✅ Afficher CLAIREMENT les 2 totaux : déclaré (agriculteurs) vs planifié (OR-Tools)
    _declared_tot = 0
    if not AGRI_DF.empty and "commercial" in AGRI_DF.columns:
        _sel_d = AGRI_DF[AGRI_DF["commercial"] == selected]
        _declared_tot = pd.to_numeric(_sel_d["tonnage_total"], errors="coerce").fillna(0).sum()
    _planned_tot = one["Tonnes/Jour"].sum() if not one.empty else 0
    _ecart = _planned_tot - _declared_tot
    _pct   = (_ecart/_declared_tot*100) if _declared_tot > 0 else 0
    
    cd1, cd2, cd3 = st.columns(3)
    cd1.metric("📋 Tonnage DÉCLARÉ (saison)", f"{_declared_tot:,.0f}t",
               help="Tonnage total que le commercial a déclaré dans son fichier d'upload")
    cd2.metric("📊 Tonnage PLANIFIÉ (OR-Tools)", f"{_planned_tot:,.0f}t",
               delta=f"{_pct:+.1f}% vs déclaré",
               help="Tonnage qu'OR-Tools a placé dans le calendrier (~95% du déclaré, tolérance ±5%)")
    cd3.metric("Écart placement", f"{_ecart:+,.0f}t",
               help="Écart entre déclaré et planifié — normal jusqu'à -5% (tolérance OR-Tools)")
    
    st.caption("📋 Déclaré = ce que le commercial a annoncé | 📊 Planifié = ce qu'OR-Tools place dans le calendrier "
               "(la différence est normale jusqu'à 5% — tolérance solveur)")
    st.markdown(" ")
    st.caption("⬇️ Le tableau ci-dessous montre le tonnage **planifié** journalier par agriculteur "
               "(0 = pas de livraison ce jour)")

    if not one.empty and "Agriculteur" in one.columns and "Tonnes/Jour" in one.columns:

        # Construire le pivot: lignes=Agriculteur, colonnes=Date
        one_pivot_src = one[["Date","Agriculteur","Tonnes/Jour","Usine"]].copy()
        one_pivot_src["Date_str"] = one_pivot_src["Date"].dt.strftime("%d/%m")

        # Aggregate (un agriculteur peut livrer à plusieurs usines le même jour)
        piv = one_pivot_src.groupby(["Agriculteur","Date_str"])["Tonnes/Jour"].sum().reset_index()
        piv["Tonnes/Jour"] = piv["Tonnes/Jour"].round(0).astype(int)

        # Pivot table
        piv_table = piv.pivot(index="Agriculteur", columns="Date_str", values="Tonnes/Jour").fillna(0).astype(int)

        # Trier les colonnes par ordre chronologique
        try:
            all_dates_sorted = sorted(
                one_pivot_src["Date"].dt.normalize().unique()
            )
            date_str_ordered = [d.strftime("%d/%m") for d in all_dates_sorted
                                 if d.strftime("%d/%m") in piv_table.columns]
            # Garder uniquement les colonnes existantes dans l'ordre
            piv_table = piv_table[[c for c in date_str_ordered if c in piv_table.columns]]
        except Exception:
            pass

        # Ajouter colonne TOTAL
        piv_table.insert(0, "TOTAL (t)", piv_table.sum(axis=1))
        piv_table = piv_table.sort_values("TOTAL (t)", ascending=False)

        # Ajouter ligne TOTAL
        totals_row = piv_table.sum(axis=0)
        totals_row.name = "── TOTAL JOUR ──"
        piv_table = pd.concat([piv_table, totals_row.to_frame().T])

        # Options d'affichage
        col_opt1, col_opt2 = st.columns([2, 1])
        with col_opt1:
            n_days_total = len([c for c in piv_table.columns if c != "TOTAL (t)"])
            n_farmers_piv = len(piv_table) - 1  # -1 pour la ligne TOTAL
            st.caption(f"📊 {n_farmers_piv} agriculteurs × {n_days_total} jours | "
                       f"Période: {one['Date'].min().strftime('%d/%m/%Y')} → {one['Date'].max().strftime('%d/%m/%Y')}")
        with col_opt2:
            # Filtre période rapide
            show_pic_only = st.checkbox("⚡ PIC uniquement (1-15 Jul)", key=f"piv_pic_{selected}")

        if show_pic_only:
            # Garder seulement les colonnes PIC
            import re as _re
            def _is_pic_col(col):
                if col == "TOTAL (t)": return True
                try:
                    d, m = col.split("/")
                    return int(m) == 7 and 1 <= int(d) <= 15
                except Exception:
                    return False
            piv_table = piv_table[[c for c in piv_table.columns if _is_pic_col(c)]]

        # Affichage avec couleur selon valeur (0=gris, >0=dégradé vert)
        def _color_cell(val):
            if isinstance(val, str):
                return "background-color: #1a1a2e; color: #888; font-weight: bold"
            if val == 0:
                return "background-color: #0d1117; color: #444"
            intensity = min(1.0, val / 300)  # normalise à 300t max
            r = int(14  + (0   - 14)  * intensity)
            g = int(149 + (229 - 149) * intensity)
            b = int(160 + (160 - 160) * intensity)
            return f"background-color: rgb({r},{g},{b}); color: {'white' if intensity > 0.4 else '#0d1117'}; font-weight: {'bold' if val > 50 else 'normal'}"

        try:
            styled = piv_table.style.applymap(_color_cell)
            st.dataframe(styled, use_container_width=True,
                         height=min(900, max(300, (n_farmers_piv + 2) * 35 + 40)))
        except Exception:
            st.dataframe(piv_table, use_container_width=True,
                         height=min(900, max(300, (n_farmers_piv + 2) * 35 + 40)))

        # Export planning journalier du commercial
        # ✅ Reset_index pour avoir Date en colonne (piv_table a Date en index)
        _piv_for_export = piv_table.reset_index()
        col_pe1, col_pe2 = st.columns(2)
        with col_pe1:
            st.download_button(
                f"📅 {selected} — Excel SÉPARÉ PAR JOUR",
                data=df_to_xlsx_by_day(_piv_for_export, date_col="Date"),
                file_name=f"planning_{selected.replace(' ','_')}_par_jour.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True,
                help="Un onglet par jour",
            )
        with col_pe2:
            st.download_button(
                f"📊 {selected} — Excel récap complet",
                data=df_to_xlsx_styled(_piv_for_export, sheet_name="Planning"),
                file_name=f"planning_{selected.replace(' ','_')}_complet.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        
        # ── 🚛 NOUVEAU: Export TRANSPORT du commercial ─────────────────
        st.markdown("---")
        st.subheader(f"🚛 Transport & Voyages — {selected}")
        
        if "Véhicules Requis" in one.columns or "Type Véhicule" in one.columns:
            # Construire le tableau transport: agriculteur × date × véhicules × voyages
            transp_cols = [c for c in ["Date","Agriculteur","Usine","Tonnes/Jour",
                                       "Type Véhicule","Véhicules Requis","Nb Voyages"] 
                          if c in one.columns]
            df_transport = one[transp_cols].sort_values(["Date","Agriculteur"]).reset_index(drop=True)
            df_transport["Date"] = pd.to_datetime(df_transport["Date"]).dt.strftime("%d/%m/%Y")
            
            # Affichage tableau
            st.dataframe(df_transport, use_container_width=True, height=400, hide_index=True)
            
            # Stats résumées
            col_t1, col_t2, col_t3, col_t4 = st.columns(4)
            total_voyages = int(df_transport["Nb Voyages"].sum()) if "Nb Voyages" in df_transport.columns else 0
            total_jours   = df_transport["Date"].nunique()
            avg_per_day   = total_voyages / total_jours if total_jours > 0 else 0
            
            col_t1.metric("Total voyages saison", f"{total_voyages:,}")
            col_t2.metric("Jours actifs", total_jours)
            col_t3.metric("Voyages/jour (moy)", f"{avg_per_day:.1f}")
            
            # Comptage par type véhicule
            if "Type Véhicule" in df_transport.columns:
                veh_count = df_transport["Type Véhicule"].value_counts().to_dict()
                veh_str = " | ".join(f"{v}: {c}" for v, c in veh_count.items())
                col_t4.metric("Répartition véhicules", veh_str[:50])
            
            # Bouton export Excel (formaté)
            import io
            buffer = io.BytesIO()
            with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
                df_transport.to_excel(writer, sheet_name="Transport_jour", index=False)
                # Résumé par jour
                daily_summary = one.groupby(one["Date"].dt.date if "Date" in one.columns else "Date").agg({
                    "Tonnes/Jour": "sum",
                    "Nb Voyages": "sum" if "Nb Voyages" in one.columns else "count",
                    "Agriculteur": "nunique",
                }).reset_index()
                daily_summary.columns = ["Date", "Total tonnes", "Total voyages", "Nb agriculteurs"]
                daily_summary.to_excel(writer, sheet_name="Resume_par_jour", index=False)
            buffer.seek(0)
            
            st.download_button(
                f"⬇️ Exporter Transport & Voyages {selected} (Excel)",
                data=buffer,
                file_name=f"transport_{selected.replace(' ','_')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
            )
            
            # Aussi CSV simple
            st.download_button(
                f"📊 Exporter Transport {selected} (Excel)",
                data=df_to_xlsx_styled(df_transport),
                file_name=f"transport_{selected.replace(' ','_')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        else:
            st.info("Colonnes véhicules non disponibles dans le planning.")
    else:
        st.info("Aucune donnée disponible pour ce commercial.")
    
    # ═══════════════════════════════════════════════════════════════
    # 🔁 SECTION JOURS DOUBLES
    # ═══════════════════════════════════════════════════════════════
    st.markdown("---")
    st.subheader("🔁 Jours doubles — détails par commercial")
    st.caption("Un **jour double** = 2 livraisons dans la même journée (matin + après-midi). "
               "Le système les utilise quand le tonnage à livrer dépasse le cap normal.")
    
    if not p.empty and "Note" in p.columns:
        # Filtrer livraisons marquées JOUR DOUBLE
        dbl_mask = p["Note"].astype(str).str.contains("JOUR DOUBLE", case=False, na=False)
        df_dbl = p[dbl_mask].copy()
        
        if df_dbl.empty:
            st.success("✅ Aucun jour double dans ce planning ! Tous les commerciaux respectent leur cap normal.")
        else:
            # KPIs en haut
            nb_lignes = len(df_dbl)
            nb_jours_unique = df_dbl.groupby(["Commercial","Date"]).ngroups
            tonnage_total = df_dbl["Tonnes/Jour"].sum()
            nb_commerciaux = df_dbl["Commercial"].nunique()
            
            k1, k2, k3, k4 = st.columns(4)
            k1.metric("🔁 Jours doubles", f"{nb_jours_unique}")
            k2.metric("📦 Livraisons concernées", f"{nb_lignes:,}")
            k3.metric("⚖️ Tonnage en double", f"{tonnage_total:,.0f}t")
            k4.metric("👥 Commerciaux concernés", f"{nb_commerciaux}")
            
            # ── Récapitulatif par commercial ──
            st.markdown("##### Récapitulatif par commercial")
            recap = df_dbl.groupby("Commercial").agg(
                jours_doubles=("Date", "nunique"),
                livraisons=("Date", "count"),
                tonnage=("Tonnes/Jour", "sum")
            ).reset_index()
            recap.columns = ["Commercial", "Nb jours doubles", "Livraisons", "Tonnage (t)"]
            recap["Tonnage (t)"] = recap["Tonnage (t)"].round(0).astype(int)
            recap = recap.sort_values("Nb jours doubles", ascending=False)
            st.dataframe(recap, hide_index=True, use_container_width=True)
            
            # ── Liste détaillée des jours doubles par commercial ──
            st.markdown("##### 📅 Liste détaillée des jours doubles")
            
            # Sélecteur de commercial
            comm_list = sorted(df_dbl["Commercial"].unique().tolist())
            if CURRENT_ROLE == "commercial":
                # Si commercial connecté, n'afficher que le sien
                if CURRENT_NAME in comm_list:
                    sel_comm = CURRENT_NAME
                else:
                    st.info(f"✅ {CURRENT_NAME} n'a pas de jour double dans ce plan.")
                    sel_comm = None
            else:
                sel_comm = st.selectbox("Choisir un commercial", ["Tous"] + comm_list,
                                        key="sel_comm_dbl")
            
            if sel_comm:
                if sel_comm == "Tous":
                    df_show = df_dbl.copy()
                else:
                    df_show = df_dbl[df_dbl["Commercial"]==sel_comm].copy()
                
                # Agrégation par (Commercial, Date) : total tonnage + nb agriculteurs + usines
                jour_summary = df_show.groupby(["Commercial","Date"]).agg(
                    tonnage_jour=("Tonnes/Jour", "sum"),
                    nb_agris=("Agriculteur", "nunique"),
                    nb_livraisons=("Agriculteur", "count"),
                    usines=("Usine", lambda x: ", ".join(sorted(set(x))))
                ).reset_index()
                jour_summary["Date"] = pd.to_datetime(jour_summary["Date"]).dt.strftime("%Y-%m-%d")
                jour_summary["tonnage_jour"] = jour_summary["tonnage_jour"].round(0).astype(int)
                jour_summary.columns = ["Commercial", "Date", "Tonnage (t)",
                                       "Nb agriculteurs", "Livraisons", "Usines"]
                jour_summary = jour_summary.sort_values(["Commercial","Date"])
                
                st.dataframe(jour_summary, hide_index=True, use_container_width=True,
                    column_config={
                        "Tonnage (t)": st.column_config.ProgressColumn(
                            "Tonnage (t)", min_value=0, max_value=1400, format="%d t"),
                    })
                
                # Détail livraisons d'un jour précis
                st.markdown("##### 🔍 Détail des livraisons d'un jour double")
                jours_uniques = jour_summary["Date"].unique().tolist()
                if jours_uniques:
                    sel_date = st.selectbox("Choisir un jour", jours_uniques, 
                                            key="sel_date_dbl")
                    
                    detail = df_show[df_show["Date"].astype(str).str[:10] == sel_date][
                        ["Agriculteur","Usine","Tonnes/Jour","Type Véhicule","Véhicules Requis"]
                    ].copy()
                    detail["Tonnes/Jour"] = detail["Tonnes/Jour"].astype(int)
                    detail = detail.sort_values("Tonnes/Jour", ascending=False)
                    
                    if sel_comm != "Tous":
                        st.caption(f"**{sel_comm}** — {sel_date} — Total : "
                                  f"**{detail['Tonnes/Jour'].sum()}t** sur **{len(detail)} livraisons**")
                    st.dataframe(detail, hide_index=True, use_container_width=True)
    else:
        st.info("Pas d'information sur les jours doubles dans ce planning.")

# ── TAB 3: PAR USINE ─────────────────────────────────────────
with tab3:
    if "Usine" not in p.columns:
        st.info("Colonne 'Usine' absente du planning.")
    else:
        # For usine role: show only data for their factory
        # Header message for usine users
        if CURRENT_ROLE == "usine":
            st.info(f"🏭 Vue **{CURRENT_NAME}** — Tonnage et voyages vous concernant.")

        factories = sorted(p["Usine"].dropna().unique())

        # Cards — total per factory (usine sees only theirs)
        if CURRENT_ROLE != "usine":
            cols_f = st.columns(len(factories))
            for i, f in enumerate(factories):
                # Use declared tonnage from agriculteurs table (more accurate)
                ft = GLOBAL_USINE_TONS.get(f, p[p["Usine"]==f]["Tonnes/Jour"].sum())
                with cols_f[i]:
                    st.metric(f, f"{ft:,.0f} t")

        # Line chart per factory
        fact_daily = p.groupby(["Date","Usine"])["Tonnes/Jour"].sum().reset_index()
        fig9 = px.line(
            fact_daily, x="Date", y="Tonnes/Jour", color="Usine",
            color_discrete_map=FACTORY_COLORS,
            title="Tonnes/jour reçues" + (f" — {CURRENT_NAME}" if CURRENT_ROLE == "usine" else " par usine"),
            labels={"Tonnes/Jour":"Tonnes/j"},
            template="plotly_dark",
        )
        fig9.update_traces(line_width=2)
        fig9.update_layout(
            plot_bgcolor="#0d1117", paper_bgcolor="#161b22",
            height=420, hovermode="closest",
        )
        fig9.add_vrect(x0=str(PEAK_START), x1=str(PEAK_END),
                       fillcolor="gold", opacity=0.06, line_width=0)
        st.plotly_chart(fig9, use_container_width=True)

        c1, c2 = st.columns(2)
        with c1:
            peak_fact = (fact_daily.groupby("Usine")["Tonnes/Jour"]
                         .max().reset_index().rename(columns={"Tonnes/Jour":"Pic/Jour"}))
            fig10 = px.bar(
                peak_fact, x="Usine", y="Pic/Jour",
                color="Usine", color_discrete_map=FACTORY_COLORS,
                title="Pic journalier max", template="plotly_dark",
                text_auto=".3s",
            )
            fig10.update_layout(paper_bgcolor="#161b22", showlegend=False, height=300)
            st.plotly_chart(fig10, use_container_width=True)
        with c2:
            # Use declared tonnage from agriculteurs (correct totals per usine)
            if GLOBAL_USINE_TONS:
                fact_tot = pd.DataFrame(list(GLOBAL_USINE_TONS.items()),
                                         columns=["Usine","Tonnes/Jour"])
            else:
                fact_tot = p.groupby("Usine")["Tonnes/Jour"].sum().reset_index()
            fig11 = px.bar(
                fact_tot, x="Tonnes/Jour", y="Usine", orientation="h",
                color="Usine", color_discrete_map=FACTORY_COLORS,
                title="Total tonnes déclarées (saison)", template="plotly_dark",
                text_auto=".3s",
            )
            fig11.update_layout(paper_bgcolor="#161b22", showlegend=False, height=300)
            st.plotly_chart(fig11, use_container_width=True)

        # Drill-down
        st.subheader("📋 Détail journalier")
        if CURRENT_ROLE == "usine":
            # Usine sees: date, total tonnes/jour, nb voyages — NO commercial column
            fact_one = p[p["Usine"] == CURRENT_NAME].sort_values("Date")
            daily_usine = (fact_one.groupby("Date")
                           .agg(Tonnes_Jour=("Tonnes/Jour","sum"),
                                Nb_Voyages=("Nb Voyages","sum"))
                           .reset_index())
            daily_usine["Pic"] = daily_usine["Date"].apply(
                lambda d: "⚡ PIC" if PEAK_START <= d.date() <= PEAK_END else ""
            )
            st.dataframe(daily_usine, use_container_width=True, height=300)
            st.download_button(
                f"📊 Télécharger planning {CURRENT_NAME} (Excel)",
                data=_build_usine_excel(fact_one, CURRENT_NAME, AGRI_DF),
                file_name=f"planning_{CURRENT_NAME}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
            )
        else:
            sel_fact = st.selectbox("Choisir une usine", factories)
            fact_one = p[p["Usine"]==sel_fact].sort_values("Date")
            # Directeur/commercial sees commercial column too
            show_cols_usine = [c for c in ["Date","Commercial","Agriculteur","Tonnes/Jour",
                                            "Nb Voyages","Pic de Récolte"] if c in fact_one.columns]
            st.dataframe(
                fact_one[show_cols_usine].reset_index(drop=True),
                use_container_width=True, height=260,
            )
            # ✅ Bouton export riche avec colonnes organisées + totaux
            st.download_button(
                f"📊 Télécharger planning {sel_fact} (Excel complet)",
                data=_build_usine_excel(fact_one, sel_fact, AGRI_DF),
                file_name=f"planning_{sel_fact}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True,
            )

# ── TAB 4: TRANSPORT & ALERTES ───────────────────────────────
with tab4:
    # ALL 4 vehicle types — TRACTEUR shown even if 0
    ALL_VEH_COLS = ["Voyages TRACTEUR","Voyages PETIT POILOUR","Voyages POILOUR","Voyages SEMI"]
    VEH_COLORS   = {
        "TRACTEUR":      "#a16207",   # Tracteur caisses (COMOCAP)
        "PETIT POILOUR": "#f5a623",   # PPL (7-12t) = Petit Poilour
        "POILOUR":       "#3b82f6",   # PL  (13-25t) = Poilour
        "SEMI":          "#00e5a0",   # Semi (25-40t)
    }
    VEH_DISPLAY = {
        "TRACTEUR":      "TRACTEUR (caisses)",
        "PETIT POILOUR": "PPL / Petit Poilour (7-12t)",
        "POILOUR":       "PL / Poilour (13-25t)",
        "SEMI":          "SEMI (25-40t)",
    }
    # Ensure all 4 columns exist, fill missing with 0
    for vc in ALL_VEH_COLS:
        if vc not in t.columns:
            t[vc] = 0

    veh_daily = t.groupby("Date")[ALL_VEH_COLS + ["Total Tonnes"]].sum().reset_index()

    # Stacked bar — all 4 vehicles
    veh_long = veh_daily.melt(id_vars="Date", value_vars=ALL_VEH_COLS,
                               var_name="Véhicule", value_name="Voyages")
    veh_long["Véhicule"] = veh_long["Véhicule"].str.replace("Voyages ", "")
    fig12 = px.bar(
        veh_long, x="Date", y="Voyages", color="Véhicule",
        barmode="stack",
        color_discrete_map={k.replace("Voyages ",""): v for k,v in
                            zip(ALL_VEH_COLS, VEH_COLORS.values())},
        title="Voyages par type de véhicule — chaque jour (4 types)",
        template="plotly_dark",
    )
    fig12.update_layout(
        plot_bgcolor="#0d1117", paper_bgcolor="#161b22",
        height=400, hovermode="closest",
    )
    fig12.add_vrect(x0=str(PEAK_START), x1=str(PEAK_END),
                    fillcolor="gold", opacity=0.06, line_width=0)
    st.plotly_chart(fig12, use_container_width=True)

    # Export transport table
    _ct1, _ct2 = st.columns(2)
    with _ct1:
        st.download_button(
            "📊 Transport — Excel stylé",
            data=df_to_xlsx_styled(veh_daily, sheet_name="Transport"),
            file_name="transport_journalier.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True,
        )
    with _ct2:
        st.download_button(
            "📊 Transport — Excel",
            data=df_to_xlsx_styled(veh_daily),
            file_name="transport_journalier.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

    # ── Vehicle count summary ──
    st.subheader("📦 Total voyages par type — saison complète")
    veh_totals_all = {
        vc.replace("Voyages ",""): int(t[vc].sum())
        for vc in ALL_VEH_COLS
    }
    v1, v2, v3, v4 = st.columns(4)
    v1.metric("🚜 TRACTEUR",       f"{veh_totals_all.get('TRACTEUR',0):,} voyages")
    v2.metric("🚛 PETIT POILOUR",  f"{veh_totals_all.get('PETIT POILOUR',0):,} voyages")
    v3.metric("🚚 POILOUR",        f"{veh_totals_all.get('POILOUR',0):,} voyages")
    v4.metric("🚜 SEMI",           f"{veh_totals_all.get('SEMI',0):,} voyages")
    st.markdown("---")

    c1, c2 = st.columns(2)
    with c1:
        # Total unique vehicle counts needed for the season
        st.subheader("🚛 Véhicules nécessaires — saison")
        # Rotations par jour par type de véhicule (basé sur distances moyennes Tunisia)
        # SEMI: 3-5 rot/j (longs trajets KAIROUAN/GAFSA), PPL/PL: 5-7 rot/j (courts)
        ROT_PAR_VEH = {
            "TRACTEUR":      6,   # 6 rotations/j (courte distance, ferme locale)
            "PETIT POILOUR": 5,   # 5 rotations/j (moyen rayon)
            "POILOUR":       5,   # 5 rotations/j (moyen rayon)
            "SEMI":          4,   # 4 rotations/j (longs trajets RM + GAFSA)
        }
        fleet_inventory = {
            "TRACTEUR":      fl_trac,
            "PETIT POILOUR": fl_ppl,
            "POILOUR":       fl_pl,
            "SEMI":          fl_semi,
        }
        # Véhicules nécessaires = ceil(voyages_pic_journée / rotations_par_véhicule)
        import math
        for vc in ALL_VEH_COLS:
            vname = vc.replace("Voyages ", "")
            rot   = ROT_PAR_VEH.get(vname, 5)
            total_voyages = int(t[vc].sum())
            # Peak = max sur la période 1-15 juillet uniquement (cap s'applique là)
            t_peak = t[(t["Date"].dt.date >= PEAK_START) & (t["Date"].dt.date <= PEAK_END)]
            peak_day = int(t_peak[vc].max()) if not t_peak.empty and vc in t_peak.columns else 0
            if peak_day == 0 and not t.empty:
                peak_day = int(t[vc].max())  # fallback sur toute la période
            needed = math.ceil(peak_day / rot) if peak_day > 0 else 0
            owned  = fleet_inventory[vname]
            if total_voyages == 0:
                st.info(f"⚪ **{vname}** : non utilisé cette saison")
            elif needed > owned:
                st.error(f"🔴 **{vname}** : besoin **{needed} véhicules** × {rot} rot/j (pic={peak_day} voyages) | vous avez {owned} → manque {needed-owned}")
            else:
                st.success(f"✅ **{vname}** : besoin **{needed} véhicules** × {rot} rot/j (pic={peak_day} voyages) | vous avez {owned} → marge {owned-needed}")

    with c2:
        # Bar chart total voyages by type
        veh_totals_all = {vc.replace("Voyages ",""): int(t[vc].sum()) for vc in ALL_VEH_COLS}
        fig13 = px.bar(
            x=list(veh_totals_all.values()),
            y=list(veh_totals_all.keys()),
            orientation="h",
            color=list(veh_totals_all.keys()),
            color_discrete_map={k: VEH_COLORS[k] for k in VEH_COLORS},
            title="Total voyages saison",
            template="plotly_dark",
            text_auto=True,
        )
        fig13.update_layout(paper_bgcolor="#161b22", showlegend=False, height=300)
        st.plotly_chart(fig13, use_container_width=True)

    # ──── NOUVEAU : Tableau récapitulatif Transport par USINE ────
    st.markdown("---")
    st.subheader("🚛 Disponibilité Transport par Usine (caps PIC 1-15 Jul)")
    
    # Caps & transport confirmé (référence: transport_12_mai.xlsx)
    # Caps officiels + transport confirmé + joker + tracteur COMOCAP
    # COMOCAP : 100t/j en TRACTEUR (~10 voyages × 10t) en plus du transport confirmé
    # ✅ Source: transport_etat_final.xlsx (10/06/2026)
    TRANSPORT_DATA_USINE = {
        "SICAM":    {"cap": 1500, "conf": 1381, "bennes_conf": 67, "joker": 0,   "joker_bennes": 0, "tracteur": 0},
        "TUCAL":    {"cap": 800,  "conf": 363,  "bennes_conf": 19, "joker": 114, "joker_bennes": 6, "tracteur": 0},
        "COMOCAP":  {"cap": 800,  "conf": 328,  "bennes_conf": 23, "joker": 101, "joker_bennes": 7, "tracteur": 0},
        "ABIDA":    {"cap": 200,  "conf": 80,   "bennes_conf": 3,  "joker": 0,   "joker_bennes": 0, "tracteur": 0},
        "ELFALLEH": {"cap": 150,  "conf": 24,   "bennes_conf": 2,  "joker": 0,   "joker_bennes": 0, "tracteur": 0},
    }
    
    transport_summary = []
    for usine, data in TRANSPORT_DATA_USINE.items():
        cap_official = data["cap"]
        confirm_tons = data["conf"]
        bennes_conf  = data["bennes_conf"]
        joker_tons   = data["joker"]
        joker_bennes = data["joker_bennes"]
        tracteur_tons = data.get("tracteur", 0)
        total_dispo  = confirm_tons + joker_tons + tracteur_tons
        manque       = max(0, cap_official - total_dispo)
        # Estimation bennes nécessaires pour combler le manque (~25t/benne moyen)
        bennes_manquantes = math.ceil(manque / 25) if manque > 0 else 0
        couverture = round(total_dispo / cap_official * 100, 1)
        
        statut = "✅ Complet" if couverture >= 100 else (
                 "⚠️ Limité" if couverture >= 50 else "🔴 Critique")
        
        transport_summary.append({
            "Usine":            usine,
            "Cap officiel":     f"{cap_official}t/j",
            "Confirmé":         f"{confirm_tons}t/j",
            "Bennes confirmées":bennes_conf,
            "Joker":            f"{joker_tons}t/j" if joker_tons > 0 else "—",
            "Tracteur":         f"{tracteur_tons}t/j (~{tracteur_tons//10} voyages)" if tracteur_tons > 0 else "—",
            "Total disponible": f"{total_dispo}t/j",
            "Manque":           f"{manque}t/j" if manque > 0 else "—",
            "Bennes nécessaires en plus": bennes_manquantes if bennes_manquantes > 0 else "—",
            "Couverture":       f"{couverture}%",
            "Statut":           statut,
        })
    
    df_ts = pd.DataFrame(transport_summary)
    st.dataframe(df_ts, use_container_width=True, hide_index=True)
    
    # Stats globales
    total_cap    = sum(d["cap"] for d in TRANSPORT_DATA_USINE.values())
    total_dispo  = sum(d["conf"] + d["joker"] + d.get("tracteur", 0) for d in TRANSPORT_DATA_USINE.values())
    total_bennes = sum(d["bennes_conf"] + d["joker_bennes"] for d in TRANSPORT_DATA_USINE.values())
    total_manque = total_cap - total_dispo
    
    cs1, cs2, cs3, cs4 = st.columns(4)
    cs1.metric("Cap officiel total",      f"{total_cap:,} t/j")
    cs2.metric("Transport disponible",    f"{total_dispo:,} t/j",
               delta=f"{round(total_dispo/total_cap*100,1)}% couvert")
    cs3.metric("Bennes confirmées",       f"{total_bennes}")
    cs4.metric("Manque global",           f"{total_manque:,} t/j",
               delta=f"~{math.ceil(total_manque/25)} bennes manquantes",
               delta_color="inverse")
    
    st.caption("📌 Joker = camions polyvalents (BOURAK pour TUCAL, LUI-MEME pour COMOCAP). "
               "TRACTEUR = vrai transport 10t/voyage (uniquement COMOCAP, ~100t/j = 10 voyages). "
               "Source: transport_12_mai.xlsx (confirmé). Manque = caps officiel - transport disponible.")
    
    st.markdown("---")
    
    # Double transport days chart
    if "Jours Double" in t.columns:
        dbl_d = t.groupby("Date")["Jours Double"].sum().reset_index()
        fig14 = px.bar(
            dbl_d, x="Date", y="Jours Double",
            color_discrete_sequence=["#e8543a"],
            title="Jours de double transport par date",
            template="plotly_dark",
        )
        fig14.update_layout(paper_bgcolor="#161b22", height=280)
        st.plotly_chart(fig14, use_container_width=True)

# ── SECTION NÉCESSITÉ TRANSPORT (fin de tab4) ───────────────────
    st.markdown("---")
    st.subheader("📦 Disponibilité Transport & Besoin Restant par Usine")
    st.caption("Comparaison flotte propre disponible vs tonnage max planifié pendant le PIC")
    # Le dashboard utilise toujours FLEET_EXACT (valeurs hardcodées depuis Etat_Transport_Final_2026.xlsx)
    # transport_disponible.xlsx est utilisé comme source de vérification uniquement

    # ── Flotte propre : lecture depuis transport_disponible.xlsx ──
    @st.cache_data(ttl=300)
    def _load_fleet():
        import os
        # ✅ Source: transport_etat_final.xlsx (10/06/2026)
        FALLBACK = {
            "SICAM":    {"SEMI":390,"PL":927,"PPL": 64,"TRACTEUR":  0,"BOURAK":  0,"total":1381,"nb_bennes": 67},
            "COMOCAP":  {"SEMI": 90,"PL": 91,"PPL":147,"TRACTEUR":  0,"BOURAK":  0,"total": 328,"nb_bennes": 23},
            "TUCAL":    {"SEMI": 60,"PL":303,"PPL":  0,"TRACTEUR":  0,"BOURAK":114,"total": 477,"nb_bennes": 25},
            "ABIDA":    {"SEMI": 60,"PL": 20,"PPL":  0,"TRACTEUR":  0,"BOURAK":  0,"total":  80,"nb_bennes":  3},
            "ELFALLEH": {"SEMI":  0,"PL":  0,"PPL": 24,"TRACTEUR":  0,"BOURAK":  0,"total":  24,"nb_bennes":  2},
        }
        paths = ["transport_etat_final.xlsx",
                 os.path.join(os.path.dirname(__file__), "transport_etat_final.xlsx"),
                 "transport_disponible.xlsx",
                 os.path.join(os.path.dirname(__file__), "transport_disponible.xlsx")]
        df_t = None
        for p in paths:
            if os.path.exists(p):
                try:
                    df_t = pd.read_excel(p, sheet_name=0)
                    break
                except Exception:
                    pass
        if df_t is None:
            return FALLBACK, False
        try:
            # Auto-détecter colonnes (case-insensitive)
            col_lower = {str(c).strip().lower(): str(c).strip() for c in df_t.columns}
            usine_col = col_lower.get("usine")
            ton_col   = col_lower.get("tonnage")
            type_col  = next((col_lower[k] for k in col_lower
                              if "type" in k and "vehicule" in k), None)
            conf_col  = col_lower.get("confirmation", col_lower.get("actif"))
            if not (usine_col and ton_col and type_col):
                return FALLBACK, False
            ALIASES = {"EL FALLEH":"ELFALLEH","FELLEH":"ELFALLEH","FELLA":"ELFALLEH",
                       "LUI-MEME":"LUIMEME","LUI-MÊME":"LUIMEME","TOTAL":"SKIP"}
            df_t["_u"] = df_t[usine_col].astype(str).str.strip().str.upper().map(
                lambda x: ALIASES.get(x, x))
            df_t["_t"] = pd.to_numeric(df_t[ton_col], errors="coerce")
            def _vt(t):
                t = str(t).strip().upper()
                if "SEMI" in t or "DOUBLE" in t: return "SEMI"
                if t.startswith("PPL"):           return "PPL"
                if t.startswith("PL"):            return "PL"
                return t
            df_t["_v"] = df_t[type_col].apply(_vt)
            df_t["_a"] = df_t[conf_col].astype(str).str.strip().str.lower() \
                         if conf_col else pd.Series("ok", index=df_t.index)
            df_ok = df_t[(df_t["_a"]=="ok") & df_t["_t"].notna() &
                         (df_t["_t"]>0) & (df_t["_u"]!="SKIP")]
            bour = int(df_ok[df_ok["_u"]=="BOURAK"]["_t"].sum())
            fleet = {}
            for usine in ["SICAM","COMOCAP","TUCAL","ABIDA","ELFALLEH"]:
                sub  = df_ok[df_ok["_u"]==usine]
                semi = int(sub[sub["_v"]=="SEMI"]["_t"].sum())
                pl   = int(sub[sub["_v"]=="PL"]["_t"].sum())
                ppl  = int(sub[sub["_v"]=="PPL"]["_t"].sum())
                trac = 100 if usine=="COMOCAP" else 0
                b    = bour if usine=="TUCAL" else 0
                fleet[usine] = {"SEMI":semi,"PL":pl,"PPL":ppl,
                                "TRACTEUR":trac,"BOURAK":b,
                                "total":semi+pl+ppl+trac+b,"nb_bennes":len(sub)}
            return fleet, True
        except Exception as e:
            return FALLBACK, False
    
    FLEET_DISPO, _fleet_from_file = _load_fleet()
    # ✅ Valeurs réelles depuis Etat_Transport_Final_2026.xlsx
    # Logique: Confirmation=ok/OK ET Contrat ≠ "En attente"
    FLEET_EXACT = {
        # Source: transport_etat_final.xlsx (10/06/2026)
        "SICAM":    {"PL":927,"PPL":64,"SEMI":390,"BOURAK":0,"LUIMEME":0,
                     "nb_PL":48,"nb_PPL":6,"nb_SEMI":13,"nb_total":67,"total":1381},
        "TUCAL":    {"PL":303,"PPL":0,"SEMI":60,"BOURAK":114,"LUIMEME":0,
                     "nb_PL":17,"nb_PPL":0,"nb_SEMI":2,"nb_total":19,"total":363,
                     "bourak_nb":6,"bourak_t":114},
        "COMOCAP":  {"PL":91,"PPL":147,"SEMI":90,"BOURAK":0,"LUIMEME":101,
                     "nb_PL":6,"nb_PPL":14,"nb_SEMI":3,"nb_total":23,"total":328,
                     "luimeme_nb":7,"luimeme_t":101},
        "ABIDA":    {"PL":20,"PPL":0,"SEMI":60,"BOURAK":0,"LUIMEME":0,
                     "nb_PL":1,"nb_PPL":0,"nb_SEMI":2,"nb_total":3,"total":80},
        "ELFALLEH": {"PL":0,"PPL":24,"SEMI":0,"BOURAK":0,"LUIMEME":0,
                     "nb_PL":0,"nb_PPL":2,"nb_SEMI":0,"nb_total":2,"total":24},
    }
    CAP_OFFICIEL = {"SICAM":1500,"COMOCAP":800,"TUCAL":800,"ABIDA":200,"ELFALLEH":150}
    CAP_VEH      = {"SEMI":(27,33),"PL":(15,25),"PPL":(6,14),"TRACTEUR":(9,11)}

    # Calculer le max planifié par usine depuis le planning chargé
    if not p.empty and "Usine" in p.columns:
        max_par_usine = (p.groupby(["Date","Usine"])["Tonnes/Jour"]
                         .sum().reset_index()
                         .groupby("Usine")["Tonnes/Jour"].max()
                         .to_dict())
    else:
        max_par_usine = {u: CAP_OFFICIEL[u] for u in CAP_OFFICIEL}

    # ── Tableau résumé ───────────────────────────────────────────────
    import plotly.graph_objects as go

    rows_nec = []
    for usine in ["SICAM","TUCAL","COMOCAP","ABIDA","ELFALLEH"]:
        cap_off  = CAP_OFFICIEL[usine]
        fl       = FLEET_EXACT[usine]
        joker_t  = fl.get("bourak_t",0) + fl.get("luimeme_t",0)
        joker_nb = fl.get("bourak_nb",0) + fl.get("luimeme_nb",0)
        f_total  = fl["total"]
        f_total_avec_jokers = f_total + joker_t
        max_plan = max_par_usine.get(usine, cap_off)
        besoin   = max(0, cap_off - f_total_avec_jokers)
        pct      = round(f_total_avec_jokers / cap_off * 100, 1)

        rows_nec.append({
            "Usine":             usine,
            "Cap. Officielle":   cap_off,
            "Nb bennes propres": fl["nb_total"],
            "Nb PL":             fl["nb_PL"],
            "Nb PPL":            fl["nb_PPL"],
            "Nb Semi":           fl["nb_SEMI"],
            "t/j PL":            fl["PL"],
            "t/j PPL":           fl["PPL"],
            "t/j Semi":          fl["SEMI"],
            "Propres (t/j)":     f_total,
            "Jokers (t/j)":      joker_t,
            "Jokers (bennes)":   joker_nb,
            "TOTAL (t/j)":       f_total_avec_jokers,
            "Manque (t/j)":      besoin,
            "Couverture %":      pct,
        })

    df_nec = pd.DataFrame(rows_nec)

    # ── 5 cartes KPI par usine ────────────────────────────────────
    cols_usine = st.columns(5)
    for i, row in df_nec.iterrows():
        usine  = row["Usine"]
        besoin = row["Manque (t/j)"]
        pct    = row["Couverture %"]
        total  = row["TOTAL (t/j)"]
        cap    = row["Cap. Officielle"]
        fl     = FLEET_EXACT[usine]
        color  = "#1E8449" if pct >= 90 else ("#F39C12" if pct >= 60 else "#C0392B")
        joker_str = ""
        if fl.get("bourak_t",0) > 0:
            joker_str += f"🚛 BOURAK +{fl['bourak_t']}t<br>"
        if fl.get("luimeme_t",0) > 0:
            joker_str += f"🚛 LUIMÈME +{fl['luimeme_t']}t<br>"
        with cols_usine[i]:
            st.markdown(f"""
            <div style="background:#1a2332;border-radius:8px;padding:10px;text-align:center;
                        border-left:4px solid {color};">
              <div style="font-size:14px;font-weight:bold;color:#ccc">{usine}</div>
              <div style="font-size:22px;font-weight:bold;color:{color}">{pct:.0f}%</div>
              <div style="font-size:11px;color:#aaa">couvert ({total:.0f}t / {cap}t)</div>
              <div style="font-size:11px;color:#8cf">
                🔵 PL: {fl['nb_PL']}×≈{fl['PL']//fl['nb_PL'] if fl['nb_PL']>0 else 0}t 
                &nbsp; {'🟢 PPL: '+str(fl['nb_PPL'])+'×≈'+str(fl['PPL']//fl['nb_PPL'] if fl['nb_PPL']>0 else 0)+'t' if fl['nb_PPL']>0 else ''}
                &nbsp; {'🟡 Semi: '+str(fl['nb_SEMI'])+'×≈'+str(fl['SEMI']//fl['nb_SEMI'] if fl['nb_SEMI']>0 else 0)+'t' if fl['nb_SEMI']>0 else ''}
              </div>
              <div style="font-size:11px;color:#fa8">{joker_str}</div>
              <div style="font-size:12px;font-weight:bold;color:{'#C0392B' if besoin>0 else '#1E8449'}">
                {'⚠️ Manque: +'+str(int(besoin))+'t' if besoin > 0 else '✅ Suffisant'}</div>
            </div>""", unsafe_allow_html=True)

    st.markdown(" ")

    # ── Tableau détaillé bennes ──────────────────────────────────
    st.markdown("**📋 Détail flotte par usine (Etat_Transport_Final_2026)**")
    
    col_disp = ["Usine","Nb bennes propres","Nb PL","Nb PPL","Nb Semi",
                "t/j PL","t/j PPL","t/j Semi","Propres (t/j)",
                "Jokers (bennes)","Jokers (t/j)","TOTAL (t/j)",
                "Cap. Officielle","Manque (t/j)","Couverture %"]
    df_show = df_nec[col_disp].copy()
    df_show["Couverture %"] = df_show["Couverture %"].apply(lambda x: f"{x:.0f}%")
    df_show["Manque (t/j)"] = df_show["Manque (t/j)"].apply(
        lambda x: f"✅ 0t" if x == 0 else f"❌ -{int(x)}t")
    
    st.dataframe(df_show, use_container_width=True, hide_index=True, height=220,
        column_config={
            "Nb bennes propres": st.column_config.NumberColumn("🚛 Bennes", format="%d"),
            "TOTAL (t/j)": st.column_config.ProgressColumn(
                "TOTAL (t/j)", min_value=0, max_value=1600, format="%d t"),
        })

    # ── Détail BOURAK et LUI-MÊME ─────────────────────────────────
    st.markdown("**🔧 Jokers disponibles (renforts toutes usines)**")
    c1, c2 = st.columns(2)
    with c1:
        st.info("🚛 **BOURAK** — 6 bennes PL | **114 t/j** → renforce TUCAL principalement")
    with c2:
        st.info("🔧 **LUI-MÊME** — 7 bennes (3 PL + 4 PPL) | **101 t/j** → renforce COMOCAP principalement")

    # ── Graphique barres groupées ──────────────────────────────────
    fig_nec = go.Figure()
    fig_nec.add_trace(go.Bar(
        name="PL propres",
        x=df_nec["Usine"], y=df_nec["t/j PL"],
        marker_color="#2196F3",
        text=df_nec.apply(lambda r: f"{r['Nb PL']}×PL<br>{r['t/j PL']:.0f}t" if r['Nb PL']>0 else "", axis=1),
        textposition="inside", textfont_size=9,
    ))
    fig_nec.add_trace(go.Bar(
        name="PPL propres",
        x=df_nec["Usine"], y=df_nec["t/j PPL"],
        marker_color="#9C27B0",
        text=df_nec.apply(lambda r: f"{r['Nb PPL']}×PPL<br>{r['t/j PPL']:.0f}t" if r['Nb PPL']>0 else "", axis=1),
        textposition="inside", textfont_size=9,
    ))
    fig_nec.add_trace(go.Bar(
        name="Semi propres",
        x=df_nec["Usine"], y=df_nec["t/j Semi"],
        marker_color="#4CAF50",
        text=df_nec.apply(lambda r: f"{r['Nb Semi']}×Semi<br>{r['t/j Semi']:.0f}t" if r['Nb Semi']>0 else "", axis=1),
        textposition="inside", textfont_size=9,
    ))
    fig_nec.add_trace(go.Bar(
        name="Jokers (BOURAK/LUI-MÊME)",
        x=df_nec["Usine"], y=df_nec["Jokers (t/j)"],
        marker_color="#FF9800",
        text=df_nec.apply(lambda r: f"+{r['Jokers (bennes)']:.0f}bn<br>+{r['Jokers (t/j)']:.0f}t" if r['Jokers (t/j)']>0 else "", axis=1),
        textposition="inside", textfont_size=9,
    ))
    fig_nec.add_trace(go.Bar(
        name="Manque (à recruter)",
        x=df_nec["Usine"], y=df_nec["Manque (t/j)"],
        marker_color="#F44336",
        text=df_nec["Manque (t/j)"].apply(lambda x: f"❌ -{int(x)}t" if x>0 else "✅"),
        textposition="inside", textfont_size=9,
    ))
    for i, row in df_nec.iterrows():
        fig_nec.add_annotation(
            x=row["Usine"], y=row["Cap. Officielle"],
            text=f"Cap: {row['Cap. Officielle']}t",
            showarrow=False, font=dict(color="#FFD700", size=10),
            yshift=10,
        )
    fig_nec.add_trace(go.Scatter(
        name="Cap. officielle (PIC)",
        x=df_nec["Usine"], y=df_nec["Cap. Officielle"],
        mode="markers", marker_symbol="diamond",
        marker_size=12, marker_color="#FFD700",
    ))
    fig_nec.update_layout(
        barmode="stack",
        title="Flotte disponible par usine — Détail PL / PPL / Semi / Jokers",
        template="plotly_dark",
        paper_bgcolor="#161b22",
        plot_bgcolor="#0d1117",
        height=420,
        legend=dict(orientation="h", y=-0.25),
        yaxis_title="Tonnes/jour",
    )
    st.plotly_chart(fig_nec, use_container_width=True)

    # ── Export ────────────────────────────────────────────────────
    st.download_button(
        "📊 Exporter détail transport (Excel)",
        data=df_to_xlsx_styled(df_nec),
        file_name="transport_detail_2026.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    # ════════════════════════════════════════════════════════════════
    # SECTION : TRANSPORT À LOUER
    # ════════════════════════════════════════════════════════════════
    st.divider()
    st.markdown("## 🔴 Transport à recruter / louer")
    st.caption("Calcul automatique du manque par usine selon les règles de répartition par type de véhicule")

    # ── Données de base ─────────────────────────────────────────────
    _CONF = {
        # Source: transport_etat_final.xlsx (10/06/2026)
        "SICAM":    {"total":1381,"PL":927, "PPL":64, "SEMI":390,"nb_PL":48,"nb_PPL":6, "nb_SEMI":13,"nb_total":67},
        "TUCAL":    {"total":363, "PL":303, "PPL":0,  "SEMI":60, "nb_PL":17,"nb_PPL":0, "nb_SEMI":2, "nb_total":19},
        "COMOCAP":  {"total":328, "PL":91,  "PPL":147,"SEMI":90, "nb_PL":6, "nb_PPL":14,"nb_SEMI":3, "nb_total":23},
        "ABIDA":    {"total":80,  "PL":20,  "PPL":0,  "SEMI":60, "nb_PL":1, "nb_PPL":0, "nb_SEMI":2, "nb_total":3},
        "ELFALLEH": {"total":24,  "PL":0,   "PPL":24, "SEMI":0,  "nb_PL":0, "nb_PPL":2, "nb_SEMI":0, "nb_total":2},
    }
    _JOKERS = {"TUCAL":114, "COMOCAP":101}
    _JOKER_LBL = {"TUCAL":"BOURAK (6 PL)", "COMOCAP":"LUIMÈME (7 bn)"}
    _CAPS  = {"SICAM":1500,"TUCAL":800,"COMOCAP":800,"ABIDA":200,"ELFALLEH":150}
    # Règles de répartition du manque par type
    _REGLES = {
        "SICAM":    {"PPL":0.00,"PL":0.00,"SEMI":1.00},   # 100% Semi
        "TUCAL":    {"PPL":0.00,"PL":0.30,"SEMI":0.70},   # 30% PL + 70% Semi
        "COMOCAP":  {"PPL":0.50,"PL":0.30,"SEMI":0.20},   # 50% PPL + 30% PL + 20% Semi
        "ABIDA":    {"PPL":0.00,"PL":0.50,"SEMI":0.50},   # 50% PL + 50% Semi
        "ELFALLEH": {"PPL":0.70,"PL":0.30,"SEMI":0.00},   # 70% PPL + 30% PL
    }
    _CAP_BENNE = {"PPL":10,"PL":20,"SEMI":30}
    _VEH_LBL   = {"PPL":"Petit PL","PL":"Grand PL","SEMI":"Semi"}
    _VEH_COLOR = {"PPL":"#9C27B0","PL":"#2196F3","SEMI":"#4CAF50"}
    _USINE_ORDER = ["SICAM","TUCAL","COMOCAP","ABIDA","ELFALLEH"]

    # ── Calcul ──────────────────────────────────────────────────────
    rows_louer = []
    for u in _USINE_ORDER:
        c     = _CONF[u]
        cap   = _CAPS[u]
        joker = _JOKERS.get(u,0)
        dispo = c["total"] + joker
        manque= max(0, cap - dispo)
        for vt in ["PPL","PL","SEMI"]:
            pct   = _REGLES[u][vt]
            tonnes= int(round(manque * pct))
            nb    = int(round(tonnes / _CAP_BENNE[vt])) if tonnes > 0 else 0
            rows_louer.append({
                "Usine":u,"vtype":vt,"cap":cap,"conf":c["total"],
                "nb_conf":c["nb_total"],"nb_ppl":c["nb_PPL"],"nb_pl":c["nb_PL"],"nb_semi":c["nb_SEMI"],
                "joker":joker,"joker_lbl":_JOKER_LBL.get(u,"—"),
                "dispo":dispo,"manque":manque,"pct":pct,"tonnes":tonnes,"nb":nb,
            })
    df_louer = pd.DataFrame(rows_louer)

    # ── KPI cards — résumé par usine ────────────────────────────────
    st.markdown("**Résumé par usine**")
    cols_k = st.columns(5)
    _USINE_BG = {"SICAM":"#0d3b6e","TUCAL":"#1a4731","COMOCAP":"#5c3a00","ABIDA":"#4a1500","ELFALLEH":"#3b0a45"}
    for i, u in enumerate(_USINE_ORDER):
        sub    = df_louer[df_louer["Usine"]==u]
        cap    = _CAPS[u]
        dispo  = sub.iloc[0]["dispo"]
        manque = sub.iloc[0]["manque"]
        pct_ok = round(dispo/cap*100)
        total_bn = int(sub["nb"].sum())
        ppl_bn = int(sub[sub["vtype"]=="PPL"]["nb"].sum())
        pl_bn  = int(sub[sub["vtype"]=="PL"]["nb"].sum())
        semi_bn= int(sub[sub["vtype"]=="SEMI"]["nb"].sum())
        joker  = sub.iloc[0]["joker"]
        color  = "#E74C3C" if manque>300 else ("#F39C12" if manque>100 else "#27AE60")
        bg     = _USINE_BG.get(u,"#1a2332")
        parts  = []
        if ppl_bn>0:  parts.append(f"<span style='color:#CE93D8'>{ppl_bn} PPL</span>")
        if pl_bn>0:   parts.append(f"<span style='color:#90CAF9'>{pl_bn} PL</span>")
        if semi_bn>0: parts.append(f"<span style='color:#A5D6A7'>{semi_bn} Semi</span>")
        with cols_k[i]:
            st.markdown(f"""
            <div style="background:{bg};border-radius:10px;padding:12px;
                        border:2px solid {color};text-align:center;">
              <div style="font-size:15px;font-weight:900;color:#fff;letter-spacing:1px">{u}</div>
              <div style="margin:4px 0;font-size:11px;color:#aaa">{dispo}t dispo / {cap}t cap</div>
              <div style="font-size:26px;font-weight:900;color:{color}">{manque}t</div>
              <div style="font-size:10px;color:#aaa">à louer/jour</div>
              <div style="margin-top:6px;font-size:11px;color:#ffd">
                {' + '.join(parts) if parts else '<span style="color:#4CAF50">✅ OK</span>'}
              </div>
              {'<div style="font-size:10px;color:#80CBC4;margin-top:4px">🔧 +'+str(joker)+'t joker</div>' if joker>0 else ''}
            </div>""", unsafe_allow_html=True)

    st.markdown(" ")

    # ── Tableau détaillé ────────────────────────────────────────────
    st.markdown("**📋 Détail par usine et type de véhicule**")

    tbl_rows = []
    for u in _USINE_ORDER:
        sub   = df_louer[df_louer["Usine"]==u]
        manque= sub.iloc[0]["manque"]
        joker = sub.iloc[0]["joker"]
        dispo = sub.iloc[0]["dispo"]
        for _, r in sub.iterrows():
            if r["pct"]==0: continue
            tbl_rows.append({
                "Usine":         r["Usine"],
                "Type":          _VEH_LBL[r["vtype"]],
                "Cap (t/j)":     r["cap"],
                "Confirmé (t/j)":r["conf"],
                "Nb bennes conf":r["nb_conf"],
                "Joker (t/j)":   joker if joker>0 else "—",
                "Disponible (t/j)":dispo,
                "Manque (t/j)":  manque,
                "% règle":       f"{int(r['pct']*100)}%",
                "À louer (t/j)": r["tonnes"],
                "Nb bennes":     r["nb"],
                "t/benne":       _CAP_BENNE[r["vtype"]],
            })
    df_tbl = pd.DataFrame(tbl_rows)

    # Coloration conditionnelle
    def color_nb(v):
        if not isinstance(v, (int,float)) or v==0: return ""
        if v >= 15: return "background-color:#C0392B;color:white;font-weight:bold"
        if v >= 8:  return "background-color:#E67E22;color:white;font-weight:bold"
        if v >= 3:  return "background-color:#F9E79F;color:black"
        return "background-color:#D5F5E3;color:black"

    styled = df_tbl.style.map(color_nb, subset=["Nb bennes"])
    st.dataframe(styled, use_container_width=True, hide_index=True, height=280)

    # ── Graphique 2 : Transport à louer par usine et type ───────────
    st.markdown("**📊 Visualisation — bennes à louer par usine**")
    fig_louer = go.Figure()

    for vt in ["PPL","PL","SEMI"]:
        sub_vt = df_louer[df_louer["vtype"]==vt]
        nb_vals= [int(sub_vt[sub_vt["Usine"]==u]["nb"].sum()) for u in _USINE_ORDER]
        t_vals = [int(sub_vt[sub_vt["Usine"]==u]["tonnes"].sum()) for u in _USINE_ORDER]
        texts  = [f"{nb}bn<br>{t}t" if nb>0 else "" for nb,t in zip(nb_vals,t_vals)]
        fig_louer.add_trace(go.Bar(
            name=_VEH_LBL[vt],
            x=_USINE_ORDER, y=nb_vals,
            marker_color=_VEH_COLOR[vt],
            text=texts, textposition="inside", textfont_size=10,
        ))

    # Ligne total bennes à louer
    total_by_usine = [int(df_louer[df_louer["Usine"]==u]["nb"].sum()) for u in _USINE_ORDER]
    fig_louer.add_trace(go.Scatter(
        name="Total bennes",
        x=_USINE_ORDER, y=total_by_usine,
        mode="markers+text",
        text=[f"<b>{v}</b>" for v in total_by_usine],
        textposition="top center",
        marker=dict(symbol="diamond", size=12, color="#FFD700"),
    ))
    fig_louer.update_layout(
        barmode="stack",
        title="Bennes à recruter par usine et type de véhicule",
        template="plotly_dark",
        paper_bgcolor="#161b22",
        plot_bgcolor="#0d1117",
        height=380,
        legend=dict(orientation="h", y=-0.2),
        yaxis_title="Nombre de bennes",
        yaxis=dict(dtick=5),
    )
    st.plotly_chart(fig_louer, use_container_width=True)

    # ── Récapitulatif global ─────────────────────────────────────────
    st.markdown("**📊 Récapitulatif global à louer**")
    c1, c2, c3, c4 = st.columns(4)
    tot_manque  = int(df_louer.groupby("Usine")["manque"].first().sum())
    tot_bennes  = int(df_louer["nb"].sum())
    tot_ppl_bn  = int(df_louer[df_louer["vtype"]=="PPL"]["nb"].sum())
    tot_pl_bn   = int(df_louer[df_louer["vtype"]=="PL"]["nb"].sum())
    tot_semi_bn = int(df_louer[df_louer["vtype"]=="SEMI"]["nb"].sum())
    tot_tonnes  = int(df_louer["tonnes"].sum())
    with c1:
        st.metric("Manque total (t/j)",    f"{tot_manque}t")
    with c2:
        st.metric("Total bennes à louer",  f"{tot_bennes} bennes")
    with c3:
        st.metric("dont Semi",             f"{tot_semi_bn} Semi")
    with c4:
        st.metric("dont PL + PPL",         f"{tot_pl_bn+tot_ppl_bn} bennes")

    # ── Export CSV ──────────────────────────────────────────────────
    csv_louer = df_tbl.to_csv(index=False, sep=";").encode("utf-8-sig")
    st.download_button(
        "📊 Exporter transport à louer (Excel)",
        data=csv_louer,
        file_name="transport_a_louer_2026.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

# ── TAB 5: DÉCALAGE & OPTIMISATION ──────────────────────────
with tab5:

    # Explain clearly why these are empty
    st.success("✅ **OR-Tools a optimisé le planning automatiquement** — aucun décalage manuel ni double transport nécessaire.")

    st.markdown("""
    > Avec l'optimiseur OR-Tools, tous les conflits de transport sont résolus **pendant le calcul**.
    > L'algorithme distribue les tonnages sur les bons jours dès le départ,
    > respectant simultanément tous les caps. Il n'y a donc pas de "décalage après coup".
    """)

    st.divider()

    # ── Instead of empty charts: show constraint verification ──
    st.subheader("📊 Vérification des contraintes — résultat optimizer")

    # Caps NORMAUX (1 livraison/jour)
    COMMERCIAL_CAPS = {
        "FEDI": 850, "MAKKI BEN SALAH": 850, "KHALIL": 900,
        "ACHREF AJLANI": 450, "JILANI OBAY": 100,
    }
    # ✅ Caps DOUBLES (matin + après-midi) — autorisés pendant PIC
    COMMERCIAL_CAPS_DOUBLE = {
        "FEDI": 1300, "MAKKI BEN SALAH": 1200, "KHALIL": 1100,
        "ACHREF AJLANI": 700, "JILANI OBAY": 150,
    }
    FACTORY_CAPS = {
        "SICAM": 1500, "TUCAL": 800, "COMOCAP": 800,
        "ABIDA": 200, "ELFALLEH": 150,
    }
    # ✅ Transport confirmé réel — Source: transport_etat_final.xlsx (10/06/2026)
    # SICAM:    48 PL(927t) + 6 PPL(64t) + 13 SEMI(390t) = 1381t/j / 67 bennes
    # TUCAL:    17 PL(303t) + 2 SEMI(60t) = 363t/j / 19 bennes
    # COMOCAP:  6 PL(91t) + 14 PPL(147t) + 3 SEMI(90t) = 328t/j / 23 bennes
    # ABIDA:    1 PL(20t) + 2 SEMI(60t) = 80t/j / 3 bennes
    # ELFALLEH: 2 PPL(24t) = 24t/j / 2 bennes
    TRANSPORT_CONF = {
        "SICAM": 1381, "TUCAL": 363, "COMOCAP": 328,
        "ABIDA": 80,   "ELFALLEH": 24,
    }
    # Jokers: BOURAK=114t (6 PL) + LUIMEME=101t (3 PL + 4 PPL)
    # BOURAK  → renforce TUCAL principalement
    # LUIMEME → renforce COMOCAP principalement
    JOKER_ALLOC_DASH = {
        "TUCAL":    114,   # BOURAK 114t → renforce TUCAL
        "COMOCAP":  101,   # LUIMEME 101t → renforce COMOCAP
        "ELFALLEH": 0,
        "SICAM":    0,
        "ABIDA":    0,
    }
    # Cap réel = min(cap_théorique, transport_confirmé + jokers)
    FACTORY_REAL_CAP = {
        u: min(FACTORY_CAPS[u], TRANSPORT_CONF.get(u, FACTORY_CAPS[u]) + JOKER_ALLOC_DASH.get(u, 0))
        for u in FACTORY_CAPS
    }
    # Manque de bennes (info)
    MANQUE_BENNES = {
        u: max(0, FACTORY_CAPS[u] - (TRANSPORT_CONF.get(u,0) + JOKER_ALLOC_DASH.get(u,0)))
        for u in FACTORY_CAPS
    }

    if not p.empty:
        c1, c2 = st.columns(2)

        with c1:
            st.markdown("**Caps commerciaux — max pendant PIC (1-15 Juillet)**")
            st.caption("Les caps ne s'appliquent que du 1 au 15 juillet (JILANI/KHALIL: 1-12 juil)")
            # Filtrer uniquement la période de pic pour la vérification des caps
            p_pic = p[(p["Date"].dt.date >= PEAK_START) & (p["Date"].dt.date <= PEAK_END)]
            if p_pic.empty:
                st.warning("⚠️ Aucune donnée pour le pic 1-15 juillet — planning peut-être incomplet")
                comm_max = p.groupby(["Date","Commercial"])["Tonnes/Jour"].sum().reset_index()
            else:
                comm_max = p_pic.groupby(["Date","Commercial"])["Tonnes/Jour"].sum().reset_index()
            comm_peak = comm_max.groupby("Commercial")["Tonnes/Jour"].max().reset_index()
            # Ajouter commerciaux absents pendant pic (ex: JILANI commence Jul 23)
            _all_c  = sorted(p["Commercial"].dropna().unique())
            _exist_c = set(comm_peak["Commercial"])
            _miss_c  = [cc for cc in _all_c if cc not in _exist_c]
            if _miss_c:
                comm_peak = pd.concat([comm_peak,
                    pd.DataFrame({"Commercial":_miss_c,"Tonnes/Jour":[0]*len(_miss_c)})],
                    ignore_index=True)
            comm_peak.columns = ["Commercial", "Max réel (t/j)"]
            # ✅ Pendant PIC, utiliser CAP DOUBLE (jours doubles autorisés)
            comm_peak["Cap normal"]    = comm_peak["Commercial"].map(COMMERCIAL_CAPS).fillna(800)
            comm_peak["Limite (t/j)"]  = comm_peak["Commercial"].map(COMMERCIAL_CAPS_DOUBLE).fillna(1200)
            comm_peak["Marge (t/j)"]   = comm_peak["Limite (t/j)"] - comm_peak["Max réel (t/j)"]
            # Détecter si jours doubles (max > cap normal)
            comm_peak["Jours doubles"] = comm_peak.apply(
                lambda r: "🔁 OUI" if r["Max réel (t/j)"] > r["Cap normal"] else "—", axis=1)
            comm_peak["Statut"] = comm_peak["Marge (t/j)"].apply(
                lambda m: "✅ OK" if m >= 0 else "❌ DÉPASSÉ"
            )
            comm_peak["Max réel (t/j)"]  = comm_peak["Max réel (t/j)"].round(0).astype(int)
            comm_peak["Cap normal"]      = comm_peak["Cap normal"].astype(int)
            comm_peak["Limite (t/j)"]    = comm_peak["Limite (t/j)"].astype(int)
            comm_peak["Marge (t/j)"]     = comm_peak["Marge (t/j)"].round(0).astype(int)
            comm_peak = comm_peak[["Commercial","Max réel (t/j)","Cap normal",
                                   "Limite (t/j)","Marge (t/j)","Jours doubles","Statut"]]
            st.caption("💡 **Cap normal** = 1 livraison/jour | **Limite** = cap avec jours doubles "
                       "(2 livraisons matin + après-midi)")
            st.dataframe(comm_peak, use_container_width=True, hide_index=True,
                column_config={
                    "Statut": st.column_config.TextColumn(width="small"),
                    "Max réel (t/j)": st.column_config.ProgressColumn(
                        "Max réel (t/j)", min_value=0, max_value=1600, format="%d t"),
                })

            # Bar chart: actual vs cap
            import plotly.graph_objects as go
            fig_cap = go.Figure()
            fig_cap.add_trace(go.Bar(
                name="Max réel", x=comm_peak["Commercial"],
                y=comm_peak["Max réel (t/j)"],
                marker_color="#3b82f6", text=comm_peak["Max réel (t/j)"],
                textposition="outside",
            ))
            fig_cap.add_trace(go.Bar(
                name="Limite", x=comm_peak["Commercial"],
                y=comm_peak["Limite (t/j)"],
                marker_color="rgba(255,77,28,0.3)",
                marker_line_color="#e8543a", marker_line_width=2,
            ))
            fig_cap.update_layout(
                barmode="overlay", template="plotly_dark",
                paper_bgcolor="#161b22", height=280,
                title="Max journalier réel vs limite par commercial",
                legend=dict(orientation="h", yanchor="bottom", y=1.02),
            )
            st.plotly_chart(fig_cap, use_container_width=True)

        with c2:
            st.markdown("**Caps usines — PIC 1-15 Juillet uniquement ✅**")
            if "Usine" in p.columns:
                # ⚠️ OBLIGATOIRE: filtrer sur PIC seulement !
                # Les caps s'appliquent UNIQUEMENT 1-15 juillet.
                # Hors pic, OR-Tools libère les contraintes pour placer
                # le tonnage → les dépassements hors-pic sont NORMAUX et attendus.
                p_pic_u = p[(p["Date"].dt.date >= PEAK_START) &
                            (p["Date"].dt.date <= PEAK_END)]
                src_u = p_pic_u if not p_pic_u.empty else p
                usine_max  = src_u.groupby(["Date","Usine"])["Tonnes/Jour"].sum().reset_index()
                usine_peak = usine_max.groupby("Usine")["Tonnes/Jour"].max().reset_index()
                # Ajouter les usines à 0t pendant pic (ex: celles sans livraisons pic)
                _all_u = sorted(p["Usine"].dropna().unique())
                _exist_u = set(usine_peak["Usine"])
                _miss_u  = [u for u in _all_u if u not in _exist_u]
                if _miss_u:
                    usine_peak = pd.concat([usine_peak,
                        pd.DataFrame({"Usine":_miss_u,"Tonnes/Jour":[0]*len(_miss_u)})],
                        ignore_index=True)
                usine_peak.columns = ["Usine", "Max réel PIC (t/j)"]
                usine_peak["Limite (t/j)"]       = usine_peak["Usine"].map(FACTORY_CAPS).fillna(500)
                usine_peak["Cap transport (t/j)"] = usine_peak["Usine"].map(FACTORY_REAL_CAP).fillna(500)
                usine_peak["Marge (t/j)"]  = usine_peak["Limite (t/j)"] - usine_peak["Max réel PIC (t/j)"]
                usine_peak["Statut"]       = usine_peak["Marge (t/j)"].apply(
                    lambda m: "✅ OK" if m >= 0 else "❌ DÉPASSÉ"
                )
                usine_peak["Max réel PIC (t/j)"] = usine_peak["Max réel PIC (t/j)"].round(0).astype(int)
                usine_peak["Limite (t/j)"]   = usine_peak["Limite (t/j)"].astype(int)
                usine_peak["Marge (t/j)"]    = usine_peak["Marge (t/j)"].round(0).astype(int)
                st.dataframe(usine_peak, use_container_width=True, hide_index=True)

                fig_usine = go.Figure()
                fig_usine.add_trace(go.Bar(
                    name="Max réel PIC", x=usine_peak["Usine"],
                    y=usine_peak["Max réel PIC (t/j)"],
                    marker_color="#00e5a0", text=usine_peak["Max réel PIC (t/j)"],
                    textposition="outside",
                ))
                fig_usine.add_trace(go.Bar(
                    name="Limite", x=usine_peak["Usine"],
                    y=usine_peak["Limite (t/j)"],
                    marker_color="rgba(255,77,28,0.3)",
                    marker_line_color="#e8543a", marker_line_width=2,
                ))
                fig_usine.update_layout(
                    barmode="overlay", template="plotly_dark",
                    paper_bgcolor="#161b22", height=280,
                    title="Max journalier réel vs limite par usine",
                    legend=dict(orientation="h", yanchor="bottom", y=1.02),
                )
                st.plotly_chart(fig_usine, use_container_width=True)

    st.divider()

    # ── Peak period analysis ──
    st.subheader("📅 Analyse période de pic — 1–15 Juillet")
    if not p.empty:
        peak_p = p[(p["Date"].dt.date >= PEAK_START) & (p["Date"].dt.date <= PEAK_END)]
        if not peak_p.empty:
            peak_daily = peak_p.groupby("Date")["Tonnes/Jour"].sum().reset_index()
            fig_pk = px.area(
                peak_daily, x="Date", y="Tonnes/Jour",
                title="Tonnage journalier pendant le pic (1-15 Juillet)",
                color_discrete_sequence=["#f5a623"],
                template="plotly_dark",
            )
            fig_pk.update_layout(paper_bgcolor="#161b22", height=260)
            st.plotly_chart(fig_pk, use_container_width=True)

            c1, c2, c3 = st.columns(3)
            c1.metric("Total pic", f"{peak_p['Tonnes/Jour'].sum():,.0f} t")
            c2.metric("Max en 1 jour", f"{peak_daily['Tonnes/Jour'].max():,.0f} t")
            c3.metric("Moyenne/jour", f"{peak_daily['Tonnes/Jour'].mean():,.0f} t")

    # Summary table still shown
    if not resume.empty:
        st.divider()
        st.subheader("📊 Résumé par commercial")
        st.dataframe(resume.reset_index(drop=True), use_container_width=True)
        st.download_button(
            "📊 Exporter résumé commercial (Excel)",
            data=df_to_xlsx_styled(resume.reset_index(drop=True)),
            file_name="resume_commercial.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

# ── TAB 6: COMPARAISON PAR ANNÉES ────────────────────────────
with tab6:
    import plotly.graph_objects as go
    import pandas as _pd
    import pickle as _pkl
    import os as _os

    st.subheader("📈 Comparaison par années — Tomate industrielle")

    # ══════════════════════════════════════════════════════════════════
    #  PERSISTANCE DISQUE — survit aux fermetures de session
    # ══════════════════════════════════════════════════════════════════
    try:
        _PERSIST_DIR = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), ".dashboard_data")
        _os.makedirs(_PERSIST_DIR, exist_ok=True)
    except Exception:
        _PERSIST_DIR = _os.path.join(_os.path.expanduser("~"), ".dashboard_data")
        _os.makedirs(_PERSIST_DIR, exist_ok=True)
    _ANNEES_FILE = _os.path.join(_PERSIST_DIR, "annees_data.pkl")

    def _save_annees_to_disk():
        try:
            with open(_ANNEES_FILE, "wb") as _f:
                _pkl.dump({
                    "data": st.session_state.annees_data,
                    "apply": st.session_state.annees_apply,
                }, _f)
        except Exception as _e:
            st.warning(f"⚠️ Sauvegarde disque échouée : {_e}")

    def _load_annees_from_disk():
        try:
            if _os.path.exists(_ANNEES_FILE):
                with open(_ANNEES_FILE, "rb") as _f:
                    _d = _pkl.load(_f)
                    return _d.get("data", {}), _d.get("apply", False)
        except Exception:
            pass
        return {}, False

    # ══════════════════════════════════════════════════════════════════
    #  IMPORT MULTI-ANNÉES (session_state + persistance)
    # ══════════════════════════════════════════════════════════════════
    if "annees_data" not in st.session_state:
        _data_disk, _apply_disk = _load_annees_from_disk()
        st.session_state.annees_data = _data_disk
        st.session_state.annees_apply = _apply_disk
        if _data_disk:
            st.info(f"🔄 {len(_data_disk)} année(s) rechargée(s) automatiquement : {sorted(_data_disk.keys())}")
    if "annees_apply" not in st.session_state:
        st.session_state.annees_apply = False

    with st.expander("📤 Importer / Gérer les années", expanded=False):
        st.caption(
            "Format attendu : colonnes **Date** et **Tonnes** au minimum. "
            "Colonne optionnelle **Usine** : si présente, le total journalier sera lu "
            "depuis les lignes `Usine=TOTAL` (les autres lignes servent au détail par usine)."
        )
        c1, c2, c3 = st.columns(3)
        with c1:
            _annee_new = st.number_input("Année", min_value=2020, max_value=2035, value=2024, step=1,
                                          key="annee_new_input")
        with c2:
            _annee_couleur = st.color_picker("Couleur", value="#3b82f6", key="annee_new_color")
        with c3:
            _annee_style = st.selectbox("Style", ["Solide","Pointillé"], index=0, key="annee_new_style")

        _file_annee = st.file_uploader(
            f"Fichier pour {_annee_new}", type=["xlsx","xls","csv"], key=f"annee_up_{_annee_new}"
        )
        if _file_annee is not None:
            try:
                _df_up = _pd.read_excel(_file_annee) if _file_annee.name.lower().endswith(("xlsx","xls")) \
                         else _pd.read_csv(_file_annee)
                _df_up.columns = [str(c).strip() for c in _df_up.columns]
                _mp = {}
                for c in _df_up.columns:
                    lc = str(c).lower().strip()
                    if lc in ("date","jour","day"):        _mp[c] = "Date"
                    elif lc in ("tonnes","tonnage","t/jour","tonnes/jour"): _mp[c] = "Tonnes"
                    elif lc in ("usine","factory","site"): _mp[c] = "Usine"
                    elif lc in ("commercial","responsable","commerciale","commerciaux"): _mp[c] = "Commercial"
                _df_up = _df_up.rename(columns=_mp)
                if "Date" in _df_up.columns and "Tonnes" in _df_up.columns:
                    _df_up["Tonnes"] = _pd.to_numeric(_df_up["Tonnes"], errors="coerce").fillna(0)
                    _df_up["_couleur"] = _annee_couleur
                    _df_up["_style"] = _annee_style
                    _df_up["_annee"] = int(_annee_new)
                    st.session_state.annees_data[int(_annee_new)] = _df_up
                    _save_annees_to_disk()
                    # ── FIX: afficher le vrai total (lignes TOTAL uniquement si Usine présente) ──
                    if "Usine" in _df_up.columns:
                        _mask_total = _df_up["Usine"].astype(str).str.upper().str.strip() == "TOTAL"
                        _tot_verif = _df_up.loc[_mask_total, "Tonnes"].sum() if _mask_total.any() \
                                     else _df_up["Tonnes"].sum()
                    else:
                        _tot_verif = _df_up["Tonnes"].sum()
                    st.success(f"✅ Année {_annee_new} chargée — {len(_df_up)} lignes • Total : **{_tot_verif:,.0f} t**")
                else:
                    st.error("❌ Le fichier doit contenir au minimum les colonnes **Date** et **Tonnes**")
            except Exception as _e:
                st.error(f"❌ Erreur lecture : {_e}")

        if st.session_state.annees_data:
            st.markdown("**Années chargées :**")
            _row_cols = st.columns(min(len(st.session_state.annees_data)+1, 6))
            for _i, _an in enumerate(sorted(st.session_state.annees_data.keys())):
                with _row_cols[_i % len(_row_cols)]:
                    _dfa = st.session_state.annees_data[_an]
                    if "Usine" in _dfa.columns:
                        _mt = _dfa["Usine"].astype(str).str.upper().str.strip() == "TOTAL"
                        _t = _dfa.loc[_mt, "Tonnes"].sum() if _mt.any() else _dfa["Tonnes"].sum()
                    else:
                        _t = _dfa["Tonnes"].sum()
                    st.markdown(f"**{_an}** — {_t:,.0f} t")
                    if st.button(f"🗑️ Retirer {_an}", key=f"del_annee_{_an}"):
                        del st.session_state.annees_data[_an]
                        _save_annees_to_disk()
                        st.rerun()

        st.markdown("---")
        cA, cB = st.columns([1,3])
        with cA:
            if st.button("✅ Appliquer", type="primary", use_container_width=True,
                         disabled=not st.session_state.annees_data):
                st.session_state.annees_apply = True
                _save_annees_to_disk()
                st.success("Graphique mis à jour")
        with cB:
            if st.button("🔄 Réinitialiser aux données par défaut", use_container_width=True):
                st.session_state.annees_data = {}
                st.session_state.annees_apply = False
                _save_annees_to_disk()
                st.rerun()

    st.markdown("---")

    # ══════════════════════════════════════════════════════════════════
    #  GRAPHIQUE — style photo (Plan Rectifié)
    # ══════════════════════════════════════════════════════════════════
    def _to_daily(df, usine_filter=None, commercial_filter=None):
        """Retourne [(mmdd, tonnes)] triée pour SUPERPOSITION multi-années.
        Filtres cumulatifs :
        - usine_filter :  TOTAL | SICAM | TUCAL | ... 
        - commercial_filter : TOUS | FEDI | ACHREF AJLANI | ...
        Règle : ligne TOTAL (Usine=TOTAL, Commercial=TOTAL) uniquement quand aucun filtre.
        Sinon on somme les lignes détail correspondantes.
        """
        if df is None or df.empty:
            return []
        d = df.copy()
        u_col = d["Usine"].astype(str).str.upper().str.strip() if "Usine" in d.columns else None
        c_col = d["Commercial"].astype(str).str.upper().str.strip() if "Commercial" in d.columns else None
        
        u_choix = str(usine_filter or "TOTAL").upper()
        c_choix = str(commercial_filter or "TOUS").upper()
        
        # Cas 1: aucun filtre → prendre les lignes TOTAL / TOTAL
        if u_choix == "TOTAL" and c_choix == "TOUS":
            if u_col is not None and c_col is not None:
                d = d[(u_col == "TOTAL") & (c_col == "TOTAL")]
            elif u_col is not None:
                d = d[u_col == "TOTAL"]
            # Fallback si pas de ligne TOTAL
            if len(d) == 0:
                d = df.copy()
                if u_col is not None:
                    d = d[d["Usine"].astype(str).str.upper() != "TOTAL"]
        # Cas 2: filtres actifs → agréger détail
        else:
            d = df.copy()
            if u_col is not None and u_choix != "TOTAL":
                d = d[d["Usine"].astype(str).str.upper().str.strip() == u_choix]
            elif u_col is not None:
                # Toutes usines sauf ligne TOTAL
                d = d[d["Usine"].astype(str).str.upper().str.strip() != "TOTAL"]
            if c_col is not None and c_choix != "TOUS":
                d = d[d["Commercial"].astype(str).str.upper().str.strip() == c_choix]
            elif c_col is not None:
                # Tous sauf ligne TOTAL
                d = d[d["Commercial"].astype(str).str.upper().str.strip() != "TOTAL"]
        
        d["Tonnes"] = _pd.to_numeric(d["Tonnes"], errors="coerce").fillna(0)
        def _mmdd(x):
            s = str(x).strip()
            if len(s) >= 10 and s[4] in "-/":
                return s[5:10].replace("/", "-")
            return s
        d["_mmdd"] = d["Date"].apply(_mmdd)
        agg = d.groupby("_mmdd", as_index=False)["Tonnes"].sum()
        return sorted([(str(r["_mmdd"]), float(r["Tonnes"])) for _, r in agg.iterrows()])

    # Filtres usine + commercial + PIC
    _all_usines = set()
    _all_comms = set()
    for _dfa in st.session_state.annees_data.values():
        if "Usine" in _dfa.columns:
            for u in _dfa["Usine"].dropna().astype(str).str.strip().unique():
                if u.upper() != "TOTAL":
                    _all_usines.add(u.upper())
        if "Commercial" in _dfa.columns:
            for c in _dfa["Commercial"].dropna().astype(str).str.strip().unique():
                if c.upper() not in ("TOTAL","HISTORIQUE","INCONNU","NAN"):
                    _all_comms.add(c.upper())
    _usines_opts = ["TOTAL"] + sorted(_all_usines)
    _comms_opts  = ["TOUS"] + sorted(_all_comms) + (["HISTORIQUE"] if any("HISTORIQUE" in _dfa.get("Commercial", _pd.Series()).astype(str).str.upper().values for _dfa in st.session_state.annees_data.values()) else [])

    cF1, cF2, cF3, cF4 = st.columns([1, 1, 1, 2])
    with cF1:
        _usine_choix = st.selectbox("🏭 Filtre usine", _usines_opts, index=0, key="usine_filter_tab6")
    with cF2:
        _comm_choix = st.selectbox("👤 Filtre commercial", _comms_opts, index=0, key="comm_filter_tab6")
    with cF3:
        _min_tonnes = st.number_input("Cacher jours < (t)", min_value=0, max_value=500, value=50,
                                       step=10, key="min_tonnes_tab6",
                                       help="Cache les jours avec un tonnage négligeable (queues de saison)")
    with cF4:
        _pic_dates = st.text_input("Zone PIC (MM-JJ → MM-JJ)",
                                    value="07-01 → 07-15", key="pic_dates_tab6")

    st.subheader("📊 Courbes journalières comparées — superposition année par année")

    if st.session_state.annees_apply and st.session_state.annees_data:
        fig = go.Figure()
        annees_tri = sorted(st.session_state.annees_data.keys(), reverse=True)
        for _idx, _an in enumerate(annees_tri):
            _df = st.session_state.annees_data[_an]
            _serie = _to_daily(_df, usine_filter=_usine_choix, commercial_filter=_comm_choix)
            # ✅ FIX: filtrer les jours avec tonnage négligeable (queues de saison)
            if _min_tonnes and _min_tonnes > 0:
                _serie = [(d, t) for d, t in _serie if t >= _min_tonnes]
            if not _serie: continue
            _xs = [d for d,_ in _serie]
            _ys = [t for _,t in _serie]
            _col = _df["_couleur"].iloc[0] if "_couleur" in _df.columns else "#3b82f6"
            _sty = _df["_style"].iloc[0]   if "_style"   in _df.columns else "Solide"
            _dash = "dot" if _sty == "Pointillé" else None
            _is_ref = (_idx == 0)
            fig.add_trace(go.Scatter(
                x=_xs, y=_ys,
                name=f"{_an} — {sum(_ys):,.0f} t",
                customdata=[_an]*len(_xs),
                line=dict(color=_col, width=(3.0 if _is_ref else 2.0), dash=_dash,
                          shape="spline", smoothing=0.6),
                mode="lines",
                hovertemplate=f"<b>%{{customdata}}-%{{x}}</b><br>%{{y:,.0f}} t/jour<extra></extra>",
            ))

        # Ligne capacité fixée à 1300 t/j
        fig.add_hline(y=1300, line_dash="dash", line_color="#e8543a", line_width=1.7,
                      annotation_text="Cap: 1,300t/j",
                      annotation_position="top right",
                      annotation_font_color="#e8543a", annotation_font_size=11)

        # Zone PIC
        try:
            _p = [s.strip() for s in _pic_dates.replace("→","->").split("->")]
            if len(_p) == 2 and _p[0] and _p[1]:
                fig.add_vrect(x0=_p[0], x1=_p[1],
                              fillcolor="#f4c430", opacity=0.09, line_width=0,
                              annotation_text="PIC", annotation_position="top left",
                              annotation_font_color="#f4c430", annotation_font_size=13)
        except Exception: pass

        _titre_filtre = ""
        if _usine_choix != "TOTAL": _titre_filtre += f" · usine {_usine_choix}"
        if _comm_choix != "TOUS":   _titre_filtre += f" · commercial {_comm_choix}"

        fig.update_layout(
            template="plotly_dark",
            plot_bgcolor="#0d1117", paper_bgcolor="#0d1117",
            height=460, hovermode="x unified",
            title=dict(text=f"Comparaison{_titre_filtre}", font=dict(size=13, color="#94a3b8"), x=0.01),
            legend=dict(orientation="h", yanchor="bottom", y=1.05, xanchor="left", x=0,
                        bgcolor="rgba(0,0,0,0)", font=dict(size=11)),
            yaxis=dict(title="Tonnes/jour", gridcolor="rgba(255,255,255,0.08)",
                       zerolinecolor="rgba(255,255,255,0.15)", tickfont=dict(size=11)),
            xaxis=dict(title="Jour de la saison (MM-JJ) — superposition multi-années",
                       gridcolor="rgba(255,255,255,0)", tickfont=dict(size=11),
                       type="category", categoryorder="category ascending"),
            margin=dict(l=70, r=40, t=60, b=60),
        )
        st.plotly_chart(fig, use_container_width=True)

        # ── Statistiques dynamiques ──
        st.markdown("### 📊 Statistiques par année")
        _stats_cols = st.columns(len(annees_tri))
        for _i, _an in enumerate(annees_tri):
            _df = st.session_state.annees_data[_an]
            _serie = _to_daily(_df, usine_filter=_usine_choix, commercial_filter=_comm_choix)
            _tot = sum(t for _,t in _serie) if _serie else 0
            _mx  = max(t for _,t in _serie) if _serie else 0
            _mx_date = max(_serie, key=lambda x: x[1])[0] if _serie else "-"
            _n_jours = len(_serie)
            _col = _df["_couleur"].iloc[0] if "_couleur" in _df.columns else "#3b82f6"
            with _stats_cols[_i]:
                _label = f"{_usine_choix}" + (f" · {_comm_choix}" if _comm_choix != "TOUS" else "")
                st.markdown(f"<div style='background:{_col}22; border-left:4px solid {_col}; "
                            f"padding:8px 12px; border-radius:4px; margin-bottom:8px;'>"
                            f"<b style='color:{_col}; font-size:15px;'>{_an}</b> "
                            f"<span style='color:#888; font-size:11px;'>· {_label}</span></div>",
                            unsafe_allow_html=True)
                st.metric("Total saison", f"{_tot:,.0f} t")
                st.metric("Pic max/jour", f"{_mx:,.0f} t", delta=str(_mx_date))
                st.metric("Jours d'activité", f"{_n_jours}")

        # ── Comparaison par commercial (barres empilées) ──
        if _comm_choix == "TOUS" and any("Commercial" in _dfa.columns for _dfa in st.session_state.annees_data.values()):
            st.markdown("---")
            st.markdown("### 👤 Comparaison par commercial × année")
            _fig_comm = go.Figure()
            _all_comm_list = sorted(_all_comms)
            for _an in annees_tri:
                _df = st.session_state.annees_data[_an]
                if "Commercial" not in _df.columns: continue
                _col = _df["_couleur"].iloc[0] if "_couleur" in _df.columns else "#3b82f6"
                _tots_comm = []
                for _cn in _all_comm_list:
                    _s = _to_daily(_df, usine_filter=_usine_choix, commercial_filter=_cn)
                    _tots_comm.append(sum(t for _,t in _s))
                _fig_comm.add_trace(go.Bar(
                    name=str(_an), x=_all_comm_list, y=_tots_comm,
                    marker_color=_col,
                    text=[f"{v:,.0f}" for v in _tots_comm], textposition="outside",
                ))
            _fig_comm.update_layout(
                barmode="group", template="plotly_dark",
                paper_bgcolor="#0d1117", plot_bgcolor="#0d1117",
                height=400, yaxis_title="Tonnes totales saison",
                legend=dict(orientation="h", yanchor="bottom", y=1.02),
            )
            st.plotly_chart(_fig_comm, use_container_width=True)

        # ── NOUVELLE SECTION : Comparaison par usine (5 mini-graphiques) ──
        if _all_usines:
            st.markdown("---")
            st.markdown("### 🏭 Comparaison par usine × année")
            st.caption("Un mini-graphique par usine, superposition des 3 années — même style que le graphique principal.")
            _usine_order = [u for u in ["SICAM","TUCAL","COMOCAP","ABIDA","ELFALLEH"] if u in _all_usines]
            if _usine_order:
                # Afficher 2 mini-graphiques par ligne
                for _row_idx in range(0, len(_usine_order), 2):
                    _cols = st.columns(2)
                    for _ci, _u in enumerate(_usine_order[_row_idx:_row_idx+2]):
                        with _cols[_ci]:
                            _fig_u = go.Figure()
                            for _idx_yr, _an in enumerate(annees_tri):
                                _df = st.session_state.annees_data[_an]
                                _serie = _to_daily(_df, usine_filter=_u,
                                                    commercial_filter=_comm_choix)
                                # Appliquer le filtre min_tonnes
                                if _min_tonnes and _min_tonnes > 0:
                                    _serie = [(d, t) for d, t in _serie if t >= _min_tonnes]
                                if not _serie: continue
                                _xs = [d for d,_ in _serie]
                                _ys = [t for _,t in _serie]
                                _col = _df["_couleur"].iloc[0] if "_couleur" in _df.columns else "#3b82f6"
                                _sty = _df["_style"].iloc[0]   if "_style"   in _df.columns else "Solide"
                                _dash = "dot" if _sty == "Pointillé" else None
                                _is_ref_u = (_idx_yr == 0)
                                _fig_u.add_trace(go.Scatter(
                                    x=_xs, y=_ys,
                                    name=f"{_an} ({sum(_ys):,.0f} t)",
                                    customdata=[_an]*len(_xs),
                                    line=dict(color=_col,
                                              width=(2.5 if _is_ref_u else 1.8),
                                              dash=_dash, shape="spline", smoothing=0.6),
                                    mode="lines",
                                    hovertemplate=f"<b>{_u} · %{{customdata}}-%{{x}}</b><br>%{{y:,.0f}} t/jour<extra></extra>",
                                ))
                            _fig_u.update_layout(
                                template="plotly_dark",
                                plot_bgcolor="#0d1117", paper_bgcolor="#0d1117",
                                height=300, hovermode="x unified",
                                title=dict(text=f"🏭 {_u}", font=dict(size=13, color="#e5e7eb"), x=0.02),
                                legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="left", x=0,
                                            bgcolor="rgba(0,0,0,0)", font=dict(size=9)),
                                yaxis=dict(title="t/jour", gridcolor="rgba(255,255,255,0.06)",
                                           tickfont=dict(size=9)),
                                xaxis=dict(gridcolor="rgba(255,255,255,0)", tickfont=dict(size=9),
                                           type="category", categoryorder="category ascending"),
                                margin=dict(l=50, r=20, t=40, b=40),
                            )
                            st.plotly_chart(_fig_u, use_container_width=True)

    else:
        st.info("💡 Importez des fichiers dans la section « Importer / Gérer les années » "
                "puis cliquez **Appliquer** pour afficher le graphique.")


# ── TAB 7: TONNAGE PAR RÉGION ────────────────────────────────
with tab7:
    st.subheader("🗺️ Tonnage par Région — Saison 2026")

    import math
    def ceil100(x):
        if not x or x <= 0: return 0
        return math.ceil(x / 100) * 100

    USINES_ORD = ["SICAM","COMOCAP","TUCAL","ELFALLEH","ABIDA"]
    COMMS_ORD  = ["FEDI","MAKKI BEN SALAH","KHALIL","ACHREF AJLANI","JILANI OBAY"]
    # Exact order requested — GAFSA+KASSRINE merged, no AUTRE
    REG_ORD    = ["CAP BON 2","CAP BON 1","NORD","GAFSA / KASSRINE",
                  "KAIROUAN","SIDI BOUZID","BOUFICHA"]

    REG_NORM = {
        "NABEUL":   "CAP BON 2",
        "BEJA":     "NORD",
        "MANOUBA":  "NORD",
        "GAFSA":    "GAFSA / KASSRINE",
        "KASSRINE": "GAFSA / KASSRINE",
    }

    # ── Detect if a date filter is active ──────────────────
    filter_active = False
    filter_label  = "Saison complète"
    d_filter_0    = None
    d_filter_1    = None

    if peak_only:
        filter_active = True
        filter_label  = "⚡ Pic 1–15 Juillet"
        d_filter_0    = PEAK_START
        d_filter_1    = PEAK_END
    elif isinstance(date_range, (list, tuple)) and len(date_range) == 2:
        d0, d1 = date_range
        season_start = datetime.date(2026, 6, 15)
        season_end   = datetime.date(2026, 8, 31)
        if d0 != season_start or d1 != season_end:
            filter_active = True
            filter_label  = f"{d0.strftime('%d/%m')} → {d1.strftime('%d/%m/%Y')}"
            d_filter_0    = d0
            d_filter_1    = d1

    st.caption(f"📅 Période affichée : **{filter_label}**")

    # ── Build df_reg — TOUJOURS depuis agriculteurs (region 100% normalisée) ──
    # NOTE: On N'utilise PAS planning.region car get_region(zone) de l'optimizer
    # ne reconnaît qu'un nombre limité de zones → la plupart deviennent
    # "AUTRE" → toutes les régions sauf KAIROUAN seraient filtrées.
    # Le tonnage par région est une vue SAISON COMPLÈTE — le filtre date 
    # ne s'applique pas ici.
    if False:  # if False → toujours aller dans else (load_region_data)
        pass
    else:
        # Use agriculteurs from Supabase — full season declared tonnage
        @st.cache_data(ttl=30)
        def load_region_data(_v=0):
            try:
                sb = get_supabase()

                # ── Source 1: Table agriculteurs ────────────────────────
                data_agri = sb.table("agriculteurs").select(
                    "commercial,nom,tonnage_total,usine,zone,region"
                ).execute().data
                df_agri = pd.DataFrame(data_agri) if data_agri else pd.DataFrame()

                # ── Source 2: Table planning (region aussi disponible) ──
                # Pagination obligatoire (Supabase limite à 1000/requête)
                data_plan = []
                _offset = 0
                while True:
                    _b = sb.table("planning").select(
                        "commercial,agriculteur,tonnes_jour,usine,region"
                    ).range(_offset, _offset+999).execute().data
                    if not _b:
                        break
                    data_plan.extend(_b)
                    if len(_b) < 1000:
                        break
                    _offset += 1000
                df_plan = pd.DataFrame(data_plan) if data_plan else pd.DataFrame()

                # ── Normalisation helper (étendue) ────────────────────────
                # Mapping complet incluant CAP BON sans numéro, casse mixte, etc.
                NORM = {
                    # Variantes Cap Bon (sans numéro = par défaut CAP BON 1)
                    "CAP BON":"CAP BON 1","cap bon":"CAP BON 1","Cap Bon":"CAP BON 1",
                    "CAPBON":"CAP BON 1","capbon":"CAP BON 1",
                    "CAP BON 1":"CAP BON 1","cap bon 1":"CAP BON 1",
                    "CAP BON 2":"CAP BON 2","cap bon 2":"CAP BON 2",
                    "CAPB1":"CAP BON 1","capb1":"CAP BON 1","CAP B1":"CAP BON 1",
                    "CAPB2":"CAP BON 2","capb2":"CAP BON 2","CAP B2":"CAP BON 2",
                    "CAPBON1":"CAP BON 1","CAPBON2":"CAP BON 2",
                    # Nabeul → CAP BON 2 (zone sud)
                    "NABEUL":"CAP BON 2","nabeul":"CAP BON 2","Nabeul":"CAP BON 2",
                    # NORD (Beja, Manouba, Tunis, Jandouba)
                    "BEJA":"NORD","beja":"NORD","Beja":"NORD","BÉJA":"NORD",
                    "MANOUBA":"NORD","manouba":"NORD","Manouba":"NORD",
                    "JANDOUBA":"NORD","jandouba":"NORD",
                    "TUNIS":"NORD","tunis":"NORD",
                    "Nord":"NORD","nord":"NORD",
                    # GAFSA / KASSRINE (toutes variantes)
                    "GAFSA":"GAFSA / KASSRINE","gafsa":"GAFSA / KASSRINE",
                    "KASSRINE":"GAFSA / KASSRINE","kassrine":"GAFSA / KASSRINE",
                    "KASSERINE":"GAFSA / KASSRINE","kasserine":"GAFSA / KASSRINE",
                    "GAFSA/KASSRINE":"GAFSA / KASSRINE",
                    "GAFSA / KASSERINE":"GAFSA / KASSRINE",
                    "Gafsa":"GAFSA / KASSRINE","Kassrine":"GAFSA / KASSRINE",
                    # KAIROUAN
                    "Kairouan":"KAIROUAN","kairouan":"KAIROUAN",
                    # SIDI BOUZID
                    "Sidi Bouzid":"SIDI BOUZID","sidi bouzid":"SIDI BOUZID",
                    "SIDI-BOUZID":"SIDI BOUZID","SIDIBOUZID":"SIDI BOUZID",
                    # BOUFICHA / Sousse
                    "Boufiche":"BOUFICHA","boufiche":"BOUFICHA",
                    "SOUSSE":"BOUFICHA","sousse":"BOUFICHA",
                    "Boufiche":"BOUFICHA","Bouficha":"BOUFICHA",
                }

                def normalize_reg(series):
                    s = series.fillna("").astype(str).str.strip()
                    # Premier passage: matching exact
                    s = s.replace(NORM)
                    # Deuxième passage: UPPER pour ce qui reste
                    mask = ~s.isin(REG_ORD) & (s != "")
                    if mask.any():
                        s_upper = s[mask].str.upper()
                        s[mask] = s_upper.replace(NORM)
                    # Troisième passage: contient "CAP BON" → CAP BON 1
                    mask2 = ~s.isin(REG_ORD) & (s != "")
                    if mask2.any():
                        cap_mask = s[mask2].str.upper().str.contains("CAP", na=False) & s[mask2].str.upper().str.contains("BON", na=False)
                        if cap_mask.any():
                            idx = s[mask2][cap_mask].index
                            s.loc[idx] = "CAP BON 1"
                    return s

                # ── Traiter agriculteurs ────────────────────────────────
                result_frames = []
                if not df_agri.empty:
                    df_agri["tonnage_total"] = pd.to_numeric(
                        df_agri["tonnage_total"], errors="coerce")
                    df_agri = df_agri[df_agri["tonnage_total"] > 0]
                    df_agri["REGION"] = normalize_reg(df_agri["region"])
                    df_agri["nom_key"] = df_agri["nom"].fillna("").astype(str).str.upper()
                    df_agri_ok = df_agri[df_agri["REGION"].isin(REG_ORD)].copy()
                    if not df_agri_ok.empty:
                        df_agri_ok = df_agri_ok.rename(columns={"nom":"nom"})
                        result_frames.append(df_agri_ok[
                            ["commercial","nom","tonnage_total","usine","REGION"]])

                # ── Traiter planning si agriculteurs vide ou sans régions ─
                if (not result_frames) and not df_plan.empty:
                    df_plan["tonnage_total"] = pd.to_numeric(
                        df_plan["tonnes_jour"], errors="coerce")
                    df_plan = df_plan[df_plan["tonnage_total"] > 0]
                    df_plan["REGION"] = normalize_reg(df_plan["region"])
                    df_plan_ok = df_plan[df_plan["REGION"].isin(REG_ORD)].copy()
                    if not df_plan_ok.empty:
                        df_plan_ok = df_plan_ok.rename(columns={"agriculteur":"nom"})
                        # Agréger par agriculteur (sum des tonnes/jour)
                        df_plan_agg = df_plan_ok.groupby(
                            ["commercial","nom","usine","REGION"],
                            as_index=False
                        )["tonnage_total"].sum()
                        result_frames.append(df_plan_agg)

                if not result_frames:
                    return None

                df_final = pd.concat(result_frames, ignore_index=True)
                df_final = df_final[df_final["REGION"].isin(REG_ORD)]
                return df_final if not df_final.empty else None

            except Exception as e:
                return None

        df_reg = load_region_data(_v=st.session_state["sb_refresh"])

        # Fallback to planning if agriculteurs empty
        if (df_reg is None or df_reg.empty) and not p.empty and "Région" in p.columns:
            p_all = p.copy()
            p_all["REGION"] = p_all["Région"].fillna("").astype(str).str.strip().str.upper()
            p_all["REGION"] = p_all["REGION"].replace(REG_NORM)
            p_all = p_all[p_all["REGION"].isin(REG_ORD)]
            p_all["usine"]       = p_all["Usine"].fillna("").astype(str).str.strip().str.upper()
            p_all["commercial"]  = p_all["Commercial"].fillna("").astype(str).str.strip()
            p_all["nom"]         = p_all["Agriculteur"].fillna("").astype(str).str.strip()
            p_all["tonnage_total"] = pd.to_numeric(p_all["Tonnes/Jour"], errors="coerce").fillna(0)
            df_reg = p_all[["commercial","nom","tonnage_total","usine","REGION"]].copy()

        data_source_label = "saison complète (tonnage déclaré)"

    # ── Display ─────────────────────────────────────────────
    if df_reg is None or df_reg.empty:
        st.warning("⚠️ Aucune donnée de région disponible dans Supabase.")
        st.info("""
**Causes possibles :**
- La colonne `region` dans Supabase est vide ou non normalisée
- Les commerciaux n'ont pas encore uploadé leurs fichiers

**Solution :**
1. Vérifiez l'onglet 🌾 Gestion Agriculteurs → Diagnostic Supabase
2. Re-uploadez les fichiers des commerciaux avec les régions correctes
3. Ou lancez `python migrate.py` après `python optimizer_v2.py`
        """)
    else:
        # KPI cards
        total_reg = df_reg["tonnage_total"].sum()
        k1, k2, k3, k4 = st.columns(4)
        k1.metric("Régions actives", df_reg["REGION"].nunique())
        k2.metric("Tonnage", f"{total_reg:,.0f} t")
        k3.metric("Région principale",
                  df_reg.groupby("REGION")["tonnage_total"].sum().idxmax()
                  if not df_reg.empty else "—")
        k4.metric("Usine principale",
                  df_reg.groupby("usine")["tonnage_total"].sum().idxmax()
                  if not df_reg.empty else "—")
        st.caption(f"Source: {data_source_label}")
        st.markdown("---")

        # ── Pivot Région × Usine ──────────────────────────
        st.subheader("📊 Tonnage par Région × Usine")
        pv_usine = df_reg.groupby(["REGION","usine"])["tonnage_total"].sum().round(0).unstack(fill_value=0)
        for u in USINES_ORD:
            if u not in pv_usine.columns: pv_usine[u] = 0
        pv_usine = pv_usine[[u for u in USINES_ORD if u in pv_usine.columns]]
        # Apply ceil100 to each cell
        for col in pv_usine.columns:
            pv_usine[col] = pv_usine[col].apply(ceil100)
        pv_usine["TOTAL"] = pv_usine.sum(axis=1)
        pv_usine = pv_usine.reindex([r for r in REG_ORD if r in pv_usine.index])
        total_u = pv_usine.sum(); total_u.name = "TOTAL"
        pv_display = pd.concat([pv_usine, total_u.to_frame().T]).astype(int)
        pv_display.index.name = "Région"
        st.dataframe(pv_display, use_container_width=True,
            column_config={"TOTAL": st.column_config.NumberColumn("TOTAL", format="%d t"),
                **{u: st.column_config.NumberColumn(u, format="%d t")
                   for u in USINES_ORD if u in pv_display.columns}})

        pv_long = pv_usine.drop(columns="TOTAL").reset_index().melt(
            id_vars="REGION", var_name="Usine", value_name="Tonnes")
        pv_long = pv_long[pv_long["Tonnes"] > 0]
        fig_ru = px.bar(pv_long, x="REGION", y="Tonnes", color="Usine",
            barmode="stack", color_discrete_map=FACTORY_COLORS,
            title=f"Tonnage par région × usine — {filter_label}",
            template="plotly_dark", text_auto=".2s")
        fig_ru.update_layout(paper_bgcolor="#161b22", plot_bgcolor="#0d1117",
                              height=400, hovermode="closest")
        st.plotly_chart(fig_ru, use_container_width=True)
        st.markdown("---")

        # ── Pivot Région × Commercial ─────────────────────
        st.subheader("📊 Tonnage par Région × Commercial")
        pv_comm = df_reg.groupby(["REGION","commercial"])["tonnage_total"].sum().round(0).unstack(fill_value=0)
        for c in COMMS_ORD:
            if c not in pv_comm.columns: pv_comm[c] = 0
        pv_comm = pv_comm[[c for c in COMMS_ORD if c in pv_comm.columns]]
        for col in pv_comm.columns:
            pv_comm[col] = pv_comm[col].apply(ceil100)
        pv_comm["TOTAL"] = pv_comm.sum(axis=1)
        pv_comm = pv_comm.reindex([r for r in REG_ORD if r in pv_comm.index])
        total_c = pv_comm.sum(); total_c.name = "TOTAL"
        pv_comm_display = pd.concat([pv_comm, total_c.to_frame().T]).astype(int)
        pv_comm_display.index.name = "Région"
        st.dataframe(pv_comm_display, use_container_width=True)

        pv_comm_long = pv_comm.drop(columns="TOTAL").reset_index().melt(
            id_vars="REGION", var_name="Commercial", value_name="Tonnes")
        pv_comm_long = pv_comm_long[pv_comm_long["Tonnes"] > 0]
        fig_rc = px.bar(pv_comm_long, x="REGION", y="Tonnes", color="Commercial",
            barmode="stack", color_discrete_map=COMM_COLORS,
            title=f"Tonnage par région × commercial — {filter_label}",
            template="plotly_dark", text_auto=".2s")
        fig_rc.update_layout(paper_bgcolor="#161b22", plot_bgcolor="#0d1117",
                              height=400, hovermode="closest")
        st.plotly_chart(fig_rc, use_container_width=True)
        st.markdown("---")

        # ── Pie + Summary table ───────────────────────────
        c1, c2 = st.columns(2)
        with c1:
            reg_tot = df_reg.groupby("REGION")["tonnage_total"].sum().reset_index()
            fig_pie = px.pie(reg_tot, names="REGION", values="tonnage_total",
                title=f"Répartition par région — {filter_label}", hole=0.4,
                template="plotly_dark", color_discrete_sequence=px.colors.qualitative.Set3)
            fig_pie.update_layout(paper_bgcolor="#161b22", height=350)
            st.plotly_chart(fig_pie, use_container_width=True)
        with c2:
            reg_ton  = df_reg.groupby("REGION")["tonnage_total"].sum().round(0).reset_index()
            reg_ton.columns = ["Région","Tonnage (t)"]
            nom_col = "nom" if "nom" in df_reg.columns else "Agriculteur"
            reg_uniq = df_reg.groupby("REGION")[nom_col].nunique().reset_index()
            reg_uniq.columns = ["Région","Agriculteurs"]
            reg_usine = df_reg.groupby("REGION").apply(
                lambda x: x.groupby("usine")["tonnage_total"].sum().idxmax()
                if not x.empty else "—").reset_index()
            reg_usine.columns = ["Région","Usine principale"]
            summary = reg_ton.merge(reg_uniq, on="Région").merge(reg_usine, on="Région")
            summary = summary.sort_values("Tonnage (t)", ascending=False)
            summary["% Total"] = (summary["Tonnage (t)"] / summary["Tonnage (t)"].sum() * 100).round(1).astype(str) + "%"
            st.dataframe(summary, use_container_width=True, hide_index=True, height=350)

        # ── Download ──────────────────────────────────────
        st.markdown("---")
        col_dl1, col_dl2 = st.columns(2)
        with col_dl1:
            st.download_button("⬇️ Région × Usine (CSV)",
                data=df_to_xlsx_styled(pv_display.reset_index()),
                file_name=f"tonnage_region_usine_{filter_label.replace(' ','_')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
        with col_dl2:
            st.download_button("⬇️ Région × Commercial (CSV)",
                data=df_to_xlsx_styled(pv_comm_display.reset_index()),
                file_name=f"tonnage_region_commercial_{filter_label.replace(' ','_')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)

# ── TAB 8: COMPARAISON PAR PRÉVISION ─────────────────────────
with tab8:
    import pandas as _pd
    import plotly.graph_objects as go
    import pickle as _pkl
    import os as _os
    st.subheader("📊 Comparaison par prévision — Mai / Juin / Réalisé (multi-années)")
    st.caption("Importer pour chaque année : Prévision Mai + Prévision Juin + Réalisé — comparer entre saisons")

    # ══════════════════════════════════════════════════════════════════
    #  PERSISTANCE DISQUE — pour survivre aux fermetures de session
    # ══════════════════════════════════════════════════════════════════
    try:
        _PDIR = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), ".dashboard_data")
        _os.makedirs(_PDIR, exist_ok=True)
    except Exception:
        _PDIR = _os.path.join(_os.path.expanduser("~"), ".dashboard_data")
        _os.makedirs(_PDIR, exist_ok=True)
    _PREV_FILE = _os.path.join(_PDIR, "prev_data.pkl")

    def _save_prev_to_disk():
        try:
            with open(_PREV_FILE, "wb") as _f:
                _pkl.dump({
                    "data": st.session_state.prev_data,
                    "apply": st.session_state.prev_apply,
                }, _f)
        except Exception as _e:
            st.warning(f"⚠️ Sauvegarde disque échouée : {_e}")

    def _load_prev_from_disk():
        try:
            if _os.path.exists(_PREV_FILE):
                with open(_PREV_FILE, "rb") as _f:
                    _d = _pkl.load(_f)
                    return _d.get("data", {}), _d.get("apply", False)
        except Exception:
            pass
        return {}, False

    if "prev_data" not in st.session_state:
        _dprev, _aprev = _load_prev_from_disk()
        st.session_state.prev_data = _dprev
        st.session_state.prev_apply = _aprev
        if _dprev:
            st.info(f"🔄 Prévisions de {len(_dprev)} année(s) rechargées : {sorted(_dprev.keys())}")
    if "prev_apply" not in st.session_state:
        st.session_state.prev_apply = False

    with st.expander("📤 Importer / Gérer les prévisions par année", expanded=False):
        st.caption("Colonnes attendues : **Region** et **Tonnage**")
        cP1, cP2 = st.columns(2)
        with cP1:
            _prev_annee = st.number_input("Année", min_value=2020, max_value=2035, value=2025,
                                           step=1, key="prev_annee_new")
        with cP2:
            _prev_type = st.selectbox("Type", ["Prévision Mai","Prévision Juin","Réalisé"],
                                       key="prev_type_new")
        _file_prev = st.file_uploader(
            f"Charger {_prev_type} — {_prev_annee}",
            type=["xlsx","xls","csv"], key=f"prev_up_{_prev_annee}_{_prev_type}"
        )
        if _file_prev is not None:
            try:
                _df_prev = _pd.read_excel(_file_prev) if _file_prev.name.lower().endswith(("xlsx","xls")) \
                           else _pd.read_csv(_file_prev)
                _df_prev.columns = [str(c).strip() for c in _df_prev.columns]
                _mp = {}
                for c in _df_prev.columns:
                    lc = str(c).lower().strip()
                    if lc in ("region","région","zone"):        _mp[c] = "Region"
                    elif lc in ("tonnage","tonnes","total"):    _mp[c] = "Tonnage"
                _df_prev = _df_prev.rename(columns=_mp)
                if "Region" in _df_prev.columns and "Tonnage" in _df_prev.columns:
                    _df_prev["Tonnage"] = _pd.to_numeric(_df_prev["Tonnage"], errors="coerce").fillna(0)
                    _key = {"Prévision Mai":"mai","Prévision Juin":"juin","Réalisé":"reel"}[_prev_type]
                    st.session_state.prev_data.setdefault(int(_prev_annee), {})[_key] = _df_prev
                    _save_prev_to_disk()
                    st.success(f"✅ {_prev_type} {_prev_annee} : {len(_df_prev)} régions — Total {_df_prev['Tonnage'].sum():,.0f} t")
                else:
                    st.error("❌ Colonnes **Region** et **Tonnage** requises")
            except Exception as _e:
                st.error(f"❌ Erreur : {_e}")

        if st.session_state.prev_data:
            st.markdown("**Prévisions chargées :**")
            for _an in sorted(st.session_state.prev_data.keys()):
                _d = st.session_state.prev_data[_an]
                _tags = []
                if "mai"  in _d: _tags.append(f"Mai ({_d['mai']['Tonnage'].sum():,.0f}t)")
                if "juin" in _d: _tags.append(f"Juin ({_d['juin']['Tonnage'].sum():,.0f}t)")
                if "reel" in _d: _tags.append(f"Réel ({_d['reel']['Tonnage'].sum():,.0f}t)")
                cL, cR = st.columns([4,1])
                with cL: st.markdown(f"**{_an}** : {' • '.join(_tags)}")
                with cR:
                    if st.button("🗑️", key=f"del_prev_{_an}"):
                        del st.session_state.prev_data[_an]
                        _save_prev_to_disk()
                        st.rerun()

        st.markdown("---")
        cP3, cP4 = st.columns([1,3])
        with cP3:
            if st.button("✅ Appliquer prévisions", type="primary", use_container_width=True,
                         disabled=not st.session_state.prev_data, key="btn_apply_prev"):
                st.session_state.prev_apply = True
                _save_prev_to_disk()
                st.success("Mis à jour")
        with cP4:
            if st.button("🔄 Réinitialiser", use_container_width=True, key="btn_reset_prev"):
                st.session_state.prev_data = {}
                st.session_state.prev_apply = False
                _save_prev_to_disk()
                st.rerun()

    st.markdown("---")

    if st.session_state.prev_apply and st.session_state.prev_data:
        _regions = set()
        for _an, _d in st.session_state.prev_data.items():
            for _k, _dfr in _d.items():
                _regions.update(_dfr["Region"].astype(str).unique())
        _regions = sorted(_regions)
        _rows = []
        for _r in _regions:
            _row = {"Région": _r}
            for _an in sorted(st.session_state.prev_data.keys()):
                _d = st.session_state.prev_data[_an]
                for _k, _label in [("mai",f"Mai {_an}"),("juin",f"Juin {_an}"),("reel",f"Réel {_an}")]:
                    if _k in _d:
                        _v = _d[_k][_d[_k]["Region"].astype(str) == _r]["Tonnage"].sum()
                        _row[_label] = _v
            _rows.append(_row)
        _df_all_p = _pd.DataFrame(_rows)
        _tot = {"Région":"TOTAL"}
        for c in _df_all_p.columns[1:]:
            _tot[c] = _df_all_p[c].sum()
        _df_all_p = _pd.concat([_df_all_p, _pd.DataFrame([_tot])], ignore_index=True)
        st.dataframe(_df_all_p, use_container_width=True, hide_index=True)

        fig_p = go.Figure()
        COULEURS_ANNEE = {2023:"#9C27B0", 2024:"#3B82F6", 2025:"#00C896", 2026:"#F5A623", 2027:"#F97316"}
        LAB = {"mai":"Prév. Mai","juin":"Prév. Juin","reel":"Réel"}
        for _an in sorted(st.session_state.prev_data.keys()):
            _color = COULEURS_ANNEE.get(_an, "#94a3b8")
            _d = st.session_state.prev_data[_an]
            for _k in ["mai","juin","reel"]:
                if _k in _d:
                    _dfr = _d[_k]
                    fig_p.add_trace(go.Bar(
                        name=f"{LAB[_k]} {_an}", x=_dfr["Region"].astype(str),
                        y=_dfr["Tonnage"], marker_color=_color,
                        opacity=(0.9 if _k=="mai" else (0.65 if _k=="juin" else 0.4)),
                        text=[f"{v:,.0f}" for v in _dfr["Tonnage"]], textposition="outside",
                    ))
        fig_p.update_layout(barmode="group", template="plotly_dark",
                            paper_bgcolor="#161b22", plot_bgcolor="#0d1117",
                            height=500, yaxis_title="Tonnes", hovermode="x unified",
                            legend=dict(orientation="h", yanchor="bottom", y=1.02))
        st.plotly_chart(fig_p, use_container_width=True)
    else:
        st.info("💡 Importez au moins une prévision puis cliquez **Appliquer**.")

# ── TAB 9: GESTION AGRICULTEURS ──────────────────────────────
with tab9:

    # Directeur voit tout, commercial voit seulement ses agriculteurs
    if CURRENT_ROLE not in ("directeur", "commercial"):
        st.warning("🔒 Accès réservé au directeur et aux commerciaux.")
        st.stop()

    sb = get_supabase()
    IS_COMMERCIAL = CURRENT_ROLE == "commercial"

    # ── Constantes formulaire ──
    COMMERCIALS  = ["FEDI","MAKKI BEN SALAH","KHALIL","ACHREF AJLANI","JILANI OBAY"]
    USINES       = ["SICAM","COMOCAP","TUCAL","ABIDA","ELFALLEH"]
    ACCESS_CODES = ["PL/PPL","PL/SEMI","RM"]     # '/' uniquement, pas de '-'
    REGIONS      = ["CAP BON 1","CAP BON 2","NORD","GAFSA / KASSRINE",
                    "KAIROUAN","SIDI BOUZID","BOUFICHA"]

    # ── DIAGNOSTIC SUPABASE (directeur uniquement) ─────────────
    if CURRENT_ROLE == "directeur":
        with st.expander("🔍 Diagnostic Supabase — Qualité des données", expanded=False):
            try:
                sb_diag = get_supabase()
                diag_data = sb_diag.table("agriculteurs").select(
                    "commercial,region,usine,tonnage_total,date_debut,accessibilite"
                ).execute().data
                df_diag = pd.DataFrame(diag_data) if diag_data else pd.DataFrame()

                if df_diag.empty:
                    st.warning("Table agriculteurs vide.")
                else:
                    df_diag["tonnage_total"] = pd.to_numeric(df_diag["tonnage_total"], errors="coerce")
                    df_diag["date_debut"]    = pd.to_datetime(df_diag["date_debut"], errors="coerce")

                    d1, d2, d3, d4 = st.columns(4)
                    d1.metric("Total lignes", len(df_diag))
                    d2.metric("Total tonnage", f"{df_diag['tonnage_total'].sum():,.0f} t")
                    d3.metric("Agriculteurs uniques", df_diag["commercial"].nunique())

                    # Problèmes détectés
                    problems = []

                    # Régions invalides
                    GOOD_REGIONS = {"CAP BON 1","CAP BON 2","NORD","GAFSA / KASSRINE",
                                    "KAIROUAN","SIDI BOUZID","BOUFICHA"}
                    bad_reg = df_diag[~df_diag["region"].fillna("").str.strip().isin(GOOD_REGIONS)]
                    if len(bad_reg) > 0:
                        problems.append(f"⚠️ {len(bad_reg)} lignes avec région non normalisée: {bad_reg['region'].unique().tolist()}")

                    # Dates invalides (avant 2026)
                    bad_dates = df_diag[df_diag["date_debut"].dt.year < 2026]
                    if len(bad_dates) > 0:
                        problems.append(f"⚠️ {len(bad_dates)} lignes avec date_debut < 2026")

                    # Accessibilités invalides
                    GOOD_ACCESS = {"PL/PPL","PL/SEMI","RM","TRC/PPL","TRC/PPL/PL","PPL","PL","SEMI","TRAC+P,PLD","PL/PPL/SEMI"}
                    bad_acc = df_diag[~df_diag["accessibilite"].fillna("").isin(GOOD_ACCESS)]
                    if len(bad_acc) > 0:
                        vals = bad_acc['accessibilite'].unique().tolist()
                        problems.append(f"ℹ️ {len(bad_acc)} lignes avec accessibilité non répertoriée: {vals} — vérifier si valide")

                    # Tonnage nul
                    bad_ton = df_diag[df_diag["tonnage_total"].isna() | (df_diag["tonnage_total"] <= 0)]
                    if len(bad_ton) > 0:
                        problems.append(f"⚠️ {len(bad_ton)} lignes avec tonnage nul ou invalide")

                    d4.metric("Problèmes détectés", len(problems),
                              delta="OK" if len(problems)==0 else f"{len(problems)} anomalies",
                              delta_color="normal" if len(problems)==0 else "inverse")

                    if problems:
                        st.markdown("**Anomalies détectées :**")
                        for p in problems:
                            st.warning(p)
                        st.markdown("**ℹ️ Note :** Les codes TRC/PPL, TRC/PPL/PL sont valides. Si une autre valeur est signalée, vérifier avec le commercial concerné.")
                    else:
                        st.success("✅ Toutes les données Supabase sont propres et cohérentes.")

                    # Par commercial
                    st.markdown("**Par commercial :**")
                    comm_stats = df_diag.groupby("commercial").agg(
                        Lignes=("tonnage_total","count"),
                        Tonnage=("tonnage_total","sum"),
                    ).round(0).reset_index()
                    st.dataframe(comm_stats, use_container_width=True, hide_index=True)

                    # Nettoyage en un clic (directeur)
                    st.markdown("---")
                    if st.button("🧹 Nettoyer données corrompues (régions/dates invalides)",
                                 type="secondary"):
                        try:
                            # Supprimer beja, manouba et toutes régions invalides
                            for bad_region in ["beja","BEJA","manouba","MANOUBA","nabeul"]:
                                sb_diag.table("agriculteurs").delete().eq(
                                    "region", bad_region).execute()
                            # Supprimer dates invalides
                            sb_diag.table("agriculteurs").delete().lt(
                                "date_debut", "2026-01-01").execute()
                            st.success("✅ Données corrompues supprimées. Rechargez la page.")
                            st.cache_data.clear()
                            st.session_state["sb_refresh"] += 1
                            st.rerun()
                        except Exception as e:
                            st.error(f"Erreur nettoyage: {e}")

            except Exception as e:
                st.error(f"Erreur diagnostic Supabase: {e}")

    # ── Helper : charger tous les agriculteurs depuis Supabase ──
    @st.cache_data(ttl=10)
    def load_agriculteurs():
        try:
            q = sb.table("agriculteurs").select("*").order("commercial")
            data = q.execute().data
            df_a = pd.DataFrame(data) if data else pd.DataFrame()
            # Commercial voit seulement ses agriculteurs
            if IS_COMMERCIAL and not df_a.empty:
                df_a = df_a[df_a["commercial"] == CURRENT_NAME]
            return df_a
        except Exception as e:
            st.error(f"Erreur chargement agriculteurs : {e}")
            return pd.DataFrame()

    # ── Titre + bouton refresh ──
    col_title, col_refresh = st.columns([5, 1])
    with col_title:
        if IS_COMMERCIAL:
            st.subheader(f"🌾 Mes Agriculteurs — {CURRENT_NAME}")
            st.caption("Ajoutez, modifiez ou supprimez vos propres agriculteurs. "
                       "Contactez le directeur pour régénérer le planning.")
        else:
            st.subheader("🌾 Gestion des Agriculteurs")
            st.caption("Ajouter, modifier ou supprimer un agriculteur. "
                       "Les changements sont sauvegardés dans Supabase. "
                       "Cliquez 'Régénérer le planning' ensuite pour recalculer.")
    with col_refresh:
        if st.button("🔄", help="Rafraîchir la liste"):
            st.cache_data.clear()
            st.rerun()

    st.divider()

    # ── Sous-onglets : Ajouter / Modifier / Supprimer / Liste ──
    a1, a2, a3, a4 = st.tabs([
        "➕ Ajouter",
        "✏️ Modifier",
        "🗑️ Supprimer",
        "📋 Liste complète",
    ])

    # ════════════════════════════════════════════════════════
    # SOUS-ONGLET 1 : AJOUTER UN AGRICULTEUR
    # ════════════════════════════════════════════════════════
    with a1:
        st.subheader("Ajouter un nouvel agriculteur")

        with st.form("form_add", clear_on_submit=True):
            c1, c2 = st.columns(2)
            with c1:
                nom         = st.text_input("Nom de l'agriculteur *",
                                            placeholder="ex: MOHAMED BEN ALI")
                # Commercial voit seulement son propre nom dans la liste
                if IS_COMMERCIAL:
                    commercial = CURRENT_NAME
                    st.info(f"Commercial : **{CURRENT_NAME}**")
                else:
                    commercial = st.selectbox("Commercial *", COMMERCIALS)
                usine       = st.selectbox("Usine *", USINES)
                tonnage     = st.number_input("Tonnage total (tonnes) *",
                                              min_value=10.0, max_value=10000.0,
                                              value=200.0, step=25.0)
            with c2:
                region      = st.selectbox("Région", REGIONS)
                zone        = st.text_input("Zone / Localisation",
                                            placeholder="ex: dar allouch")
                access      = st.selectbox("Accessibilité véhicule *", ACCESS_CODES,
                                           help="PL/PPL = route normale | PL/SEMI = grande route")
                date_debut  = st.date_input("Date début récolte *",
                                            value=pd.Timestamp("2026-07-01"),
                                            min_value=pd.Timestamp("2026-06-15"),
                                            max_value=pd.Timestamp("2026-08-31"))
                date_fin    = st.date_input("Date fin récolte *",
                                            value=pd.Timestamp("2026-07-20"),
                                            min_value=pd.Timestamp("2026-06-15"),
                                            max_value=pd.Timestamp("2026-09-16"))

            st.markdown("")
            submitted = st.form_submit_button("✅ Enregistrer l'agriculteur",
                                              use_container_width=True,
                                              type="primary")

        if submitted:
            # Validation
            errors = []
            if not nom.strip():
                errors.append("Le nom est obligatoire.")
            if date_fin <= date_debut:
                errors.append("La date de fin doit être après la date de début.")
            if tonnage <= 0:
                errors.append("Le tonnage doit être positif.")

            if errors:
                for e in errors:
                    st.error(e)
            else:
                try:
                    row = {
                        "commercial":    commercial,
                        "nom":           nom.strip().upper(),
                        "region":        region,
                        "zone":          zone.strip() or None,
                        "usine":         usine,
                        "accessibilite": access,
                        "tonnage_total": float(tonnage),
                        "date_debut":    str(date_debut),
                        "date_fin":      str(date_fin),
                    }
                    sb.table("agriculteurs").insert(row).execute()
                    st.success(f"✅ Agriculteur **{nom.upper()}** ajouté avec succès !")
                    st.info("👉 Cliquez 'Régénérer le planning' dans la sidebar pour recalculer.")
                    st.cache_data.clear()
                except Exception as e:
                    st.error(f"Erreur lors de l'insertion : {e}")

    # ════════════════════════════════════════════════════════
    # SOUS-ONGLET 2 : MODIFIER UN AGRICULTEUR
    # ════════════════════════════════════════════════════════
    with a2:
        st.subheader("Modifier un agriculteur existant")

        df_agri = load_agriculteurs()
        if df_agri.empty:
            st.info("Aucun agriculteur dans la base.")
        else:
            # Search / filter
            search = st.text_input("🔍 Rechercher par nom ou commercial",
                                   placeholder="ex: MOHAMED ou FEDI")
            filtered = df_agri.copy()
            if search.strip():
                mask = (filtered["nom"].str.upper().str.contains(search.upper(), na=False) |
                        filtered["commercial"].str.upper().str.contains(search.upper(), na=False))
                filtered = filtered[mask]

            if filtered.empty:
                st.warning("Aucun résultat pour cette recherche.")
            else:
                # Select from filtered list
                options = [
                    f"[{r['id']}] {r['commercial']} — {r['nom']} ({r['tonnage_total']}t → {r['usine']})"
                    for _, r in filtered.iterrows()
                ]
                selected_label = st.selectbox("Sélectionner l'agriculteur à modifier",
                                               options)
                selected_id = int(selected_label.split("]")[0].replace("[",""))
                row_data = df_agri[df_agri["id"] == selected_id].iloc[0]

                st.markdown("---")
                st.caption(f"Modification de : **{row_data['nom']}** (ID {selected_id})")

                with st.form("form_edit"):
                    c1, c2 = st.columns(2)
                    with c1:
                        e_nom        = st.text_input("Nom", value=str(row_data["nom"]))
                        if IS_COMMERCIAL:
                            e_commercial = CURRENT_NAME
                            st.info(f"Commercial : **{CURRENT_NAME}**")
                        else:
                            e_commercial = st.selectbox("Commercial",
                                                         COMMERCIALS,
                                                         index=COMMERCIALS.index(row_data["commercial"])
                                                         if row_data["commercial"] in COMMERCIALS else 0)
                        e_usine      = st.selectbox("Usine",
                                                     USINES,
                                                     index=USINES.index(row_data["usine"])
                                                     if row_data["usine"] in USINES else 0)
                        e_tonnage    = st.number_input("Tonnage total",
                                                        min_value=10.0, max_value=10000.0,
                                                        value=float(row_data["tonnage_total"]),
                                                        step=25.0)
                    with c2:
                        e_region     = st.selectbox("Région",
                                                     REGIONS,
                                                     index=REGIONS.index(str(row_data["region"]).upper())
                                                     if str(row_data["region"]).upper() in REGIONS else 0)
                        e_zone       = st.text_input("Zone", value=str(row_data["zone"] or ""))
                        e_access     = st.selectbox("Accessibilité",
                                                     ACCESS_CODES,
                                                     index=ACCESS_CODES.index(str(row_data["accessibilite"]))
                                                     if str(row_data["accessibilite"]) in ACCESS_CODES else 0)
                        try:
                            dd_parsed = pd.to_datetime(row_data["date_debut"], errors="coerce")
                            dd = dd_parsed.date() if not pd.isna(dd_parsed) else pd.Timestamp("2026-07-01").date()
                            df_parsed = pd.to_datetime(row_data["date_fin"], errors="coerce")
                            df_ = df_parsed.date() if not pd.isna(df_parsed) else pd.Timestamp("2026-07-20").date()
                        except Exception:
                            dd  = pd.Timestamp("2026-07-01").date()
                            df_ = pd.Timestamp("2026-07-20").date()

                        e_debut = st.date_input("Date début", value=dd)
                        e_fin   = st.date_input("Date fin",   value=df_)

                    save_edit = st.form_submit_button("💾 Sauvegarder les modifications",
                                                      use_container_width=True,
                                                      type="primary")

                if save_edit:
                    if e_fin <= e_debut:
                        st.error("La date de fin doit être après la date de début.")
                    else:
                        try:
                            sb.table("agriculteurs").update({
                                "commercial":    e_commercial,
                                "nom":           e_nom.strip().upper(),
                                "region":        e_region,
                                "zone":          e_zone.strip() or None,
                                "usine":         e_usine,
                                "accessibilite": e_access,
                                "tonnage_total": float(e_tonnage),
                                "date_debut":    str(e_debut),
                                "date_fin":      str(e_fin),
                            }).eq("id", selected_id).execute()
                            st.success(f"✅ Agriculteur **{e_nom.upper()}** modifié !")
                            st.info("👉 Cliquez 'Régénérer le planning' dans la sidebar.")
                            st.cache_data.clear()
                        except Exception as e:
                            st.error(f"Erreur lors de la modification : {e}")

    # ════════════════════════════════════════════════════════
    # SOUS-ONGLET 3 : SUPPRIMER UN AGRICULTEUR
    # ════════════════════════════════════════════════════════
    with a3:
        st.subheader("Supprimer un agriculteur")
        st.warning("⚠️ La suppression est définitive. Le planning associé sera recalculé "
                   "à la prochaine régénération.")

        df_agri = load_agriculteurs()
        if df_agri.empty:
            st.info("Aucun agriculteur dans la base.")
        else:
            search_del = st.text_input("🔍 Rechercher l'agriculteur à supprimer",
                                        placeholder="ex: HSSINE BRINI")
            filtered_del = df_agri.copy()
            if search_del.strip():
                mask = (filtered_del["nom"].str.upper().str.contains(search_del.upper(), na=False) |
                        filtered_del["commercial"].str.upper().str.contains(search_del.upper(), na=False))
                filtered_del = filtered_del[mask]

            if filtered_del.empty:
                st.warning("Aucun résultat.")
            else:
                options_del = [
                    f"[{r['id']}] {r['commercial']} — {r['nom']} ({r['tonnage_total']}t)"
                    for _, r in filtered_del.iterrows()
                ]
                selected_del = st.selectbox("Sélectionner l'agriculteur à supprimer",
                                             options_del)
                sel_id_del = int(selected_del.split("]")[0].replace("[",""))
                sel_nom_del = df_agri[df_agri["id"] == sel_id_del].iloc[0]["nom"]

                st.markdown(f"**Tu vas supprimer :** {sel_nom_del} (ID {sel_id_del})")

                # Double confirmation
                confirm = st.checkbox(f"Je confirme la suppression de **{sel_nom_del}**")
                if confirm:
                    if st.button("🗑️ Supprimer définitivement",
                                 type="primary",
                                 use_container_width=True):
                        try:
                            sb.table("agriculteurs").delete().eq("id", sel_id_del).execute()
                            st.success(f"✅ Agriculteur **{sel_nom_del}** supprimé.")
                            st.info("👉 Cliquez 'Régénérer le planning' dans la sidebar.")
                            st.cache_data.clear()
                            st.rerun()
                        except Exception as e:
                            st.error(f"Erreur lors de la suppression : {e}")

    # ════════════════════════════════════════════════════════
    # SOUS-ONGLET 4 : LISTE COMPLÈTE
    # ════════════════════════════════════════════════════════
    with a4:
        st.subheader("Liste complète des agriculteurs")

        df_agri = load_agriculteurs()
        if df_agri.empty:
            st.info("Aucun agriculteur dans la base. Utilisez l'onglet 'Ajouter'.")
        else:
            # Stats rapides
            c1, c2, c3, c4 = st.columns(4)
            c1.metric("Total agriculteurs", len(df_agri))
            c2.metric("Tonnage total", f"{df_agri['tonnage_total'].sum():,.0f} t")
            c3.metric("Commercials", df_agri["commercial"].nunique())
            c4.metric("Usines", df_agri["usine"].nunique())

            st.markdown("")

            # Filtres
            col_f1, col_f2, col_f3 = st.columns(3)
            with col_f1:
                filt_comm = st.multiselect("Filtrer par commercial",
                                            sorted(df_agri["commercial"].unique()),
                                            default=sorted(df_agri["commercial"].unique()))
            with col_f2:
                filt_usine = st.multiselect("Filtrer par usine",
                                             sorted(df_agri["usine"].unique()),
                                             default=sorted(df_agri["usine"].unique()))
            with col_f3:
                sort_col = st.selectbox("Trier par",
                                         ["commercial","nom","tonnage_total","usine"])

            display_df = df_agri[
                df_agri["commercial"].isin(filt_comm) &
                df_agri["usine"].isin(filt_usine)
            ].sort_values(sort_col).reset_index(drop=True)

            # Show table
            show_cols = [c for c in
                         ["id","commercial","nom","usine","tonnage_total",
                          "accessibilite","region","zone","date_debut","date_fin"]
                         if c in display_df.columns]
            st.dataframe(
                display_df[show_cols],
                use_container_width=True,
                height=400,
                column_config={
                    "id":            st.column_config.NumberColumn("ID", width="small"),
                    "commercial":    st.column_config.TextColumn("Commercial"),
                    "nom":           st.column_config.TextColumn("Agriculteur"),
                    "usine":         st.column_config.TextColumn("Usine"),
                    "tonnage_total": st.column_config.NumberColumn("Tonnage (t)", format="%.0f t"),
                    "accessibilite": st.column_config.TextColumn("Accès"),
                    "region":        st.column_config.TextColumn("Région"),
                    "zone":          st.column_config.TextColumn("Zone"),
                    "date_debut":    st.column_config.TextColumn("Début récolte"),
                    "date_fin":      st.column_config.TextColumn("Fin récolte"),
                }
            )

            # Export
            st.download_button(
                "📊 Exporter la liste (Excel)",
                data=df_to_xlsx_styled(display_df[show_cols]),
                file_name="agriculteurs_supabase.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

    # ════════════════════════════════════════════════════════════════
    # VUE PAR CENTRE & RENDEMENT (t/ha)
    # ════════════════════════════════════════════════════════════════
    st.divider()
    st.markdown("## 🏘️ Vue par Centre & Rendement (t/ha)")
    st.caption("Détail des agriculteurs regroupés par centre de collecte, avec rendement par hectare")

    import json as _json, os as _os
    _centres_path = _os.path.join(_os.path.dirname(__file__), "centres_data.json")
    _CENTRES = None
    try:
        with open(_centres_path, encoding="utf-8") as _f:
            _CENTRES = _json.load(_f)
    except Exception:
        _CENTRES = None

    if not _CENTRES:
        st.info("ℹ️ Fichier centres_data.json non trouvé. Ajoutez-le dans le repo GitHub "
                "(à côté de dashboard_phase10.py) pour activer la vue par centre.")
    else:
        # Filtrer selon le rôle
        if IS_COMMERCIAL:
            my = st.session_state.get("name","").upper()
            comm_list = [c for c in _CENTRES if c.upper() == my] or list(_CENTRES.keys())
        else:
            comm_list = list(_CENTRES.keys())

        # Sélection commercial (admin) ou auto (commercial)
        if IS_COMMERCIAL:
            sel_comm = comm_list[0]
            st.markdown(f"**Commercial : {sel_comm}**")
        else:
            sel_comm = st.selectbox("👤 Choisir un commercial", comm_list, key="centre_comm")

        centres = _CENTRES.get(sel_comm, {})
        if not centres:
            st.warning("Aucun centre pour ce commercial.")
        else:
            # ── Récap rendement par centre ──
            rows_c = []
            for centre, info in centres.items():
                rows_c.append({
                    "Centre":          centre,
                    "Nb agriculteurs": len(info["agriculteurs"]),
                    "Hectares":        info["total_ha"],
                    "Tonnage (t)":     info["total_ton"],
                    "Rendement (t/ha)":info["rend"],
                })
            df_c = pd.DataFrame(rows_c).sort_values("Tonnage (t)", ascending=False)

            # KPI globaux du commercial
            tot_ha  = sum(c["total_ha"] for c in centres.values())
            tot_ton = sum(c["total_ton"] for c in centres.values())
            tot_agri= sum(len(c["agriculteurs"]) for c in centres.values())
            k1,k2,k3,k4 = st.columns(4)
            k1.metric("Centres", len(centres))
            k2.metric("Agriculteurs", tot_agri)
            k3.metric("Hectares", f"{tot_ha:.0f} ha")
            k4.metric("Rendement moyen", f"{tot_ton/tot_ha:.1f} t/ha" if tot_ha>0 else "—")

            # Graphique rendement par centre
            import plotly.graph_objects as go
            fig_c = go.Figure()
            fig_c.add_trace(go.Bar(
                x=df_c["Centre"], y=df_c["Rendement (t/ha)"],
                marker_color="#4CAF50",
                text=[f"{r:.0f} t/ha" for r in df_c["Rendement (t/ha)"]],
                textposition="outside",
            ))
            fig_c.update_layout(
                title=f"Rendement par centre — {sel_comm}",
                template="plotly_dark", paper_bgcolor="#161b22", plot_bgcolor="#0d1117",
                height=320, yaxis_title="t/ha",
            )
            st.plotly_chart(fig_c, use_container_width=True)

            st.dataframe(df_c, use_container_width=True, hide_index=True)

            # ── Détail agriculteurs d'un centre ──
            st.markdown("### 🌾 Agriculteurs d'un centre")
            sel_centre = st.selectbox("🏘️ Choisir un centre", list(centres.keys()), key="centre_sel")
            info = centres[sel_centre]
            st.caption(f"**{sel_centre}** — {len(info['agriculteurs'])} agriculteurs | "
                       f"{info['total_ha']} ha | {info['total_ton']:,}t | {info['rend']} t/ha".replace(",", " "))

            df_agri = pd.DataFrame(info["agriculteurs"])
            df_agri.columns = ["Agriculteur","Hectares","Tonnage (t)","Rendement (t/ha)","Région"]
            st.dataframe(df_agri, use_container_width=True, hide_index=True, height=320)

            st.download_button(
                f"📊 Exporter {sel_centre} (Excel)",
                data=df_to_xlsx_styled(df_agri),
                file_name=f"centre_{sel_centre.replace(' ','_')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"dl_centre_{sel_centre}",
            )

    # ════════════════════════════════════════════════════════════════
    # SECTION : BESOIN DE TRANSPORT PAR USINE
    # ════════════════════════════════════════════════════════════════
    st.divider()
    st.markdown("## 🚛 Mon besoin de transport par usine")
    st.caption("Tonnage à livrer par usine, type de véhicule recommandé et estimation bennes au pic")

    # ── Données embarquées ──────────────────────────────────────────
    # ── Besoins PIC validés par les commerciaux (guide terrain) ───────
    # Chiffres = nb de bennes SIMULTANÉES au jour de POINTE
    # ACHREF: 22 Semi | KHALIL: 40 PL + 10 Semi | JILANI: 83 PL
    # MAKKI: 40 PL + 20 PPL | FEDI: 35 PL + 15 PPL
    _BESOINS_PIC = {
        "FEDI":            [{"type":"PL",   "nb":35,"t_par_benne":20,"cap_j":700},
                            {"type":"PPL",  "nb":15,"t_par_benne":10,"cap_j":150}],
        "MAKKI BEN SALAH": [{"type":"PL",   "nb":40,"t_par_benne":20,"cap_j":800},
                            {"type":"PPL",  "nb":20,"t_par_benne":10,"cap_j":200}],
        "KHALIL":          [{"type":"PL",   "nb":40,"t_par_benne":20,"cap_j":800},
                            {"type":"SEMI", "nb":10,"t_par_benne":30,"cap_j":300}],
        "ACHREF AJLANI":   [{"type":"SEMI", "nb":22,"t_par_benne":30,"cap_j":660}],
        "JILANI OBAY":     [{"type":"PL",   "nb":83,"t_par_benne":20,"cap_j":1660}],
    }
    _TONNAGES = {
        "FEDI":32312,"MAKKI BEN SALAH":22525,"KHALIL":16510,
        "ACHREF AJLANI":16373,"JILANI OBAY":6965,
    }

    # Détail par usine (inchangé — pour référence)
    _TRANSPORT_BESOINS = {
        "FEDI": [
            {"usine":"SICAM",    "tonnage":11860,"nb_agri":12,"vehicule":"PL/PPL","long":False},
            {"usine":"TUCAL",    "tonnage":5900, "nb_agri":6, "vehicule":"PL/PPL","long":False},
            {"usine":"COMOCAP",  "tonnage":12002,"nb_agri":50,"vehicule":"PL",    "long":False},
            {"usine":"ABIDA",    "tonnage":1000, "nb_agri":1, "vehicule":"PL",    "long":False},
            {"usine":"ELFALLEH", "tonnage":1100, "nb_agri":2, "vehicule":"PPL",   "long":False},
        ],
        "MAKKI BEN SALAH": [
            {"usine":"SICAM",    "tonnage":10368,"nb_agri":40,"vehicule":"PL",    "long":False},
            {"usine":"TUCAL",    "tonnage":6715, "nb_agri":19,"vehicule":"PL",    "long":False},
            {"usine":"COMOCAP",  "tonnage":3162, "nb_agri":19,"vehicule":"PL",    "long":False},
            {"usine":"ELFALLEH", "tonnage":2280, "nb_agri":3, "vehicule":"PPL",   "long":False},
        ],
        "KHALIL": [
            {"usine":"SICAM",    "tonnage":8955, "nb_agri":11,"vehicule":"SEMI/PL","long":True},
            {"usine":"TUCAL",    "tonnage":3320, "nb_agri":4, "vehicule":"SEMI/PL","long":True},
            {"usine":"COMOCAP",  "tonnage":1320, "nb_agri":3, "vehicule":"PL",    "long":True},
            {"usine":"ABIDA",    "tonnage":2915, "nb_agri":4, "vehicule":"PL",    "long":True},
        ],
        "ACHREF AJLANI": [
            {"usine":"SICAM",    "tonnage":10768,"nb_agri":43,"vehicule":"SEMI",  "long":True},
            {"usine":"TUCAL",    "tonnage":2100, "nb_agri":13,"vehicule":"SEMI",  "long":True},
            {"usine":"COMOCAP",  "tonnage":1575, "nb_agri":1, "vehicule":"SEMI",  "long":True},
            {"usine":"ABIDA",    "tonnage":1930, "nb_agri":11,"vehicule":"SEMI",  "long":True},
        ],
        "JILANI OBAY": [
            {"usine":"SICAM",    "tonnage":3795, "nb_agri":7, "vehicule":"PL",    "long":False},
            {"usine":"TUCAL",    "tonnage":1325, "nb_agri":2, "vehicule":"PL",    "long":False},
            {"usine":"COMOCAP",  "tonnage":1845, "nb_agri":2, "vehicule":"PL",    "long":False},
        ],
    }
    _VEH_COLORS = {
        "SEMI":"#1E8449","PL":"#2F75B6","PPL":"#7030A0","PL/PPL":"#2F75B6","SEMI/PL":"#1E8449"
    }
    _USINE_COLORS = {
        "SICAM":"#0d3b6e","TUCAL":"#1a4731","COMOCAP":"#5c3a00",
        "ABIDA":"#4a1500","ELFALLEH":"#3b0a45"
    }

    # Filtrer selon le rôle
    if IS_COMMERCIAL:
        _my_name = st.session_state.get("name","").upper()
        _comm_list_t = [c for c in _TRANSPORT_BESOINS if c.upper()==_my_name]
        if not _comm_list_t: _comm_list_t = list(_TRANSPORT_BESOINS.keys())
    else:
        _comm_list_t = list(_TRANSPORT_BESOINS.keys())
        _sel_comm_t = st.selectbox("👤 Commercial", _comm_list_t, key="transp_comm_sel")
        _comm_list_t = [_sel_comm_t]

    for _comm_t in _comm_list_t:
        _pic = _BESOINS_PIC.get(_comm_t, [])
        _usine_detail = _TRANSPORT_BESOINS.get(_comm_t, [])
        if not _pic: continue

        _tot_bn  = sum(p["nb"] for p in _pic)
        _tot_cap = sum(p["cap_j"] for p in _pic)
        _tonnage = _TONNAGES.get(_comm_t, 0)

        if not IS_COMMERCIAL:
            st.markdown(f"### {_comm_t}")

        # ── Bandeau récap PIC ──────────────────────────────────────────
        st.markdown(f"""
        <div style="background:#1a2332;border-radius:10px;padding:12px 18px;
                    border-left:5px solid #FFD700;margin-bottom:12px;">
          <span style="font-size:13px;color:#FFD700;font-weight:900">
            🚛 BESOIN AU PIC — {_comm_t}
          </span>
          <span style="font-size:12px;color:#aaa;margin-left:16px">
            {_tot_bn} bennes simultanées · {_tot_cap}t/j capacité · {_tonnage:,}t saison
          </span>
        </div>""".replace(",", " "), unsafe_allow_html=True)

        # ── Cartes par type de véhicule ────────────────────────────────
        _cols_pic = st.columns(len(_pic))
        for i, p in enumerate(_cols_pic):
            b = _pic[i]
            vc = _VEH_COLORS.get(b["type"], "#888")
            nb_total  = b["nb"]
            cap_total = b["cap_j"]
            with p:
                st.markdown(f"""
                <div style="background:#0d1117;border-radius:12px;padding:14px;
                            text-align:center;border:3px solid {vc};">
                  <div style="font-size:42px;font-weight:900;color:{vc}">{nb_total}</div>
                  <div style="font-size:16px;font-weight:900;color:#fff;margin:4px 0">
                    {b['type']}
                  </div>
                  <div style="font-size:12px;color:#aaa">{b['t_par_benne']}t × {nb_total} = 
                    <b style="color:#fff">{cap_total}t/j</b>
                  </div>
                  <div style="margin-top:8px;font-size:10px;color:#888">
                    bennes simultanées au pic
                  </div>
                </div>""", unsafe_allow_html=True)

        st.markdown(" ")

        # ── Tableau récap par type ─────────────────────────────────────
        _rows_pic = []
        for p in _pic:
            _rows_pic.append({
                "Type véhicule": p["type"],
                "Nb bennes au pic": p["nb"],
                "Capacité/benne (t)": p["t_par_benne"],
                "Capacité totale (t/j)": p["cap_j"],
            })
        _tot_row_p = {
            "Type véhicule":"TOTAL",
            "Nb bennes au pic":_tot_bn,
            "Capacité/benne (t)":"—",
            "Capacité totale (t/j)":_tot_cap,
        }
        _df_pic = pd.DataFrame(_rows_pic + [_tot_row_p])
        st.dataframe(_df_pic, use_container_width=True, hide_index=True)

        # ── Détail par usine (tableau secondaire) ──────────────────────
        if _usine_detail:
            with st.expander("📋 Détail du tonnage par usine"):
                _df_u = pd.DataFrame([{
                    "Usine":b["usine"],"Tonnage (t)":b["tonnage"],
                    "Nb agriculteurs":b["nb_agri"],"Véhicule":b["vehicule"],
                    "Distance":"🔴 Longue" if b["long"] else "🟢 Courte",
                } for b in _usine_detail])
                _tot_u = {"Usine":"TOTAL","Tonnage (t)":sum(b["tonnage"] for b in _usine_detail),
                          "Nb agriculteurs":sum(b["nb_agri"] for b in _usine_detail),
                          "Véhicule":"—","Distance":"—"}
                _df_u = pd.concat([_df_u, pd.DataFrame([_tot_u])], ignore_index=True)
                st.dataframe(_df_u, use_container_width=True, hide_index=True)

        # ── Graphique barres horizontales bennes par type ───────────────
        _fig_pic = go.Figure()
        for p in _pic:
            _fig_pic.add_trace(go.Bar(
                name=p["type"], y=[_comm_t.split()[0]], x=[p["nb"]],
                orientation="h",
                marker_color=_VEH_COLORS.get(p["type"],"#888"),
                text=[f"{p['nb']} {p['type']} ({p['cap_j']}t)"],
                textposition="inside", textfont_size=12,
            ))
        _fig_pic.update_layout(
            barmode="stack",
            title=f"Bennes au pic — {_comm_t}  ({_tot_bn} bennes | {_tot_cap}t/j)",
            template="plotly_dark",paper_bgcolor="#161b22",plot_bgcolor="#0d1117",
            height=130, showlegend=True,
            xaxis=dict(title="Nb bennes"),
            margin=dict(l=10,r=10,t=40,b=10),
            legend=dict(orientation="h",y=-0.5),
        )
        st.plotly_chart(_fig_pic, use_container_width=True)

        # Export CSV
        st.download_button(
            f"📊 Exporter besoin transport {_comm_t} (Excel)",
            data=df_to_xlsx_styled(_df_pic),
            file_name=f"transport_{_comm_t.split()[0]}_2026.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"dl_transport_{_comm_t}",
        )
        st.markdown("---")

# ── TAB 11: UPLOAD PLANNING ──────────────────────────────────
with tab10:
    if not UPLOAD_AVAILABLE:
        st.error("❌ Fichier `upload_tab.py` introuvable dans le dossier.")
        st.info("Mets `upload_tab.py` dans le même dossier que `dashboard_phase10.py`.")
    else:
        render_upload_tab(
            sb=get_supabase(),
            CURRENT_ROLE=CURRENT_ROLE,
            CURRENT_NAME=CURRENT_NAME,
            CURRENT_FILTER=CURRENT_FILTER,
            GLOBAL_COMMERCIAL_FARMERS=GLOBAL_COMMERCIAL_FARMERS,
            GLOBAL_COMMERCIAL_TONS=GLOBAL_COMMERCIAL_TONS,
            df_to_csv=df_to_csv,
        )

# ONGLET COMPARAISON PLANS
with tab_comp:
    if COMPARAISON_AVAILABLE:
        render_comparaison_tab(planning_df=planning, df_to_xlsx_styled=df_to_xlsx_styled, sb=get_supabase())
    else:
        st.error("comparaison_tab.py introuvable")
        st.info("Mets comparaison_tab.py dans le meme dossier que dashboard_phase10.py")
# ── TAB DASHBOARD AGROÉCONOMIQUE (4 fichiers) ───────────────
with tab_agroeco:
    if CURRENT_ROLE == "usine":
        st.info("🔒 Onglet réservé aux commerciaux et au directeur.")
    elif AGROECO_AVAILABLE:
        render_agroeco_tab(
            sb=get_supabase(),
            CURRENT_ROLE=CURRENT_ROLE,
            CURRENT_NAME=CURRENT_NAME,
        )
    else:
        st.error("❌ agroeco_dashboard.py introuvable.")
        st.info(
            "Télécharge `agroeco_dashboard.py` depuis Claude "
            "et place-le dans le même dossier que dashboard_phase10.py, "
            "puis relance l'application."
        )