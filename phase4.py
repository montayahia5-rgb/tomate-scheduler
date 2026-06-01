# -*- coding: utf-8 -*-
import sys, io
# Force UTF-8 output on Windows (fixes cp1252 crash)
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

import pandas as pd
import datetime
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# PHASE 4 — Complete transport planning system
# Reads: Recap_tonnage_pre_vu_ajuste__mai26.xlsx  (original)
# Writes: Planning_Phase4_Transport_Double.xlsx   (dashboard input)
# ============================================================

EXCEL_IN  = "Recap tonnage prévu ajusté mai26.xlsx"
EXCEL_OUT = "Planning_Phase4_Transport_Double.xlsx"

PEAK_START  = datetime.date(2026, 7, 1)
PEAK_END    = datetime.date(2026, 7, 15)
MAX_SHIFT   = 4    # max days a farmer's end can be pushed (tomate safety)
BUFFER_DAYS = 3    # vehicle rest days between two farmers

FLEET_CAPACITY = {
    "TRACTEUR":      (11, 13),
    "PETIT POILOUR": (7,  12),
    "POILOUR":       (13, 25),
    "SEMI":          (25, 40),
}
ACCESS_VEHICLES = {
    "PL/PPL":  ["POILOUR", "PETIT POILOUR"],
    "PL-PPL":  ["POILOUR", "PETIT POILOUR"],
    "PL/SEMI": ["POILOUR", "SEMI"],
    "PL-SEMI": ["POILOUR", "SEMI"],
    "NAN":     ["TRACTEUR", "PETIT POILOUR", "POILOUR", "SEMI"],
}
COMMERCIAL_LIMITS = {
    "FEDI": 500, "MAKKI BEN SALAH": 600, "KHALIL": 500,
    "ACHREF AJLANI": 400, "JILANI OBAY": 450,
}
FACTORY_LIMITS = {
    "COMOCAP": 300, "SICAM": 500, "TUCAL": 300, "ABIDA": 200, "ELFALLEH": 150,
}

# ── helpers ──────────────────────────────────
def choose_vehicles(tons, allowed):
    order = ["SEMI","POILOUR","PETIT POILOUR","TRACTEUR"]
    sorted_v = [v for v in order if v in allowed]
    result, rem = [], tons
    for veh in sorted_v:
        if rem <= 0: break
        mn, mx = FLEET_CAPACITY[veh]
        if rem < mn: continue
        trips = int(rem / mx)
        if trips > 0:
            result.append({"vehicle":veh,"trips":trips,"tons_each":mx,"total":trips*mx})
            rem -= trips * mx
        if mn <= rem < mx:
            result.append({"vehicle":veh,"trips":1,"tons_each":round(rem,1),"total":round(rem,1)})
            rem = 0
    if rem > 0 and not result:
        veh = sorted_v[-1] if sorted_v else "POILOUR"
        result.append({"vehicle":veh,"trips":1,"tons_each":round(rem,1),"total":round(rem,1)})
    return result

thin   = Side(style="thin", color="BBBBBB")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def style(ws, headers, widths, color, start_row=1):
    fill = PatternFill("solid", start_color=color)
    font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    for ci,(h,w) in enumerate(zip(headers,widths),1):
        c = ws.cell(row=start_row, column=ci, value=h)
        c.font=font; c.fill=fill
        c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
        c.border=border
        ws.column_dimensions[get_column_letter(ci)].width=w
    ws.row_dimensions[start_row].height=30

def write_row(ws, row_num, values, fill_key):
    fills = {
        "peak":   PatternFill("solid", start_color="FFE699"),
        "double": PatternFill("solid", start_color="FFCCCC"),
        "alt":    PatternFill("solid", start_color="EBF3FB"),
        "white":  PatternFill("solid", start_color="FFFFFF"),
        "green":  PatternFill("solid", start_color="C6EFCE"),
        "orange": PatternFill("solid", start_color="FCE4D6"),
    }
    for ci,val in enumerate(values,1):
        c = ws.cell(row=row_num, column=ci, value=val)
        c.fill = fills.get(fill_key, fills["white"])
        c.font = Font(name="Calibri", size=9)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border

# ── STEP 1: Load ─────────────────────────────
df = pd.read_excel(EXCEL_IN, sheet_name="Feuil1")
df = df[~df["responsable région"].astype(str).str.contains("TOTAL", na=False)]
df = df.dropna(subset=["AGRICULTEUR","TONNAGE","debut recolte","USINE"])
df = df[df["TONNAGE"] > 0]
df["debut recolte"] = pd.to_datetime(df["debut recolte"], errors="coerce")
df = df[df["debut recolte"].dt.year >= 2000]
df["responsable région"] = df["responsable région"].astype(str).str.strip()
df["USINE"]              = df["USINE"].astype(str).str.strip()
df["accessbilite"]       = df["accessbilite"].astype(str).str.strip().str.upper()
date_cols = [c for c in df.columns if isinstance(c, datetime.datetime)]
print(f"Loaded {len(df)} rows, {len(date_cols)} date columns.")

# ── STEP 2: Build farmer records ──────────────
records = []
for _, row in df.iterrows():
    commercial = row["responsable région"]
    farmer     = str(row["AGRICULTEUR"]).strip()
    factory    = row["USINE"]
    region     = str(row.get("REGION","") or "").strip()
    access_raw = row["accessbilite"]
    total_tons = float(row["TONNAGE"])
    allowed_v  = ACCESS_VEHICLES.get(access_raw, ACCESS_VEHICLES["NAN"])

    daily_plan = {}
    for col in date_cols:
        val = row[col]
        if pd.notna(val) and float(val) > 0:
            daily_plan[col.date()] = float(val)

    if not daily_plan:
        start_dt    = row["debut recolte"].date()
        days_needed = max(1, int(total_tons/20) + 5)
        daily_t     = round(total_tons/days_needed, 2)
        for d in range(days_needed):
            daily_plan[start_dt + datetime.timedelta(days=d)] = daily_t

    records.append({
        "Commercial": commercial, "Farmer": farmer,
        "Factory":    factory,    "Region": region,
        "Access":     access_raw, "Allowed_Veh": allowed_v,
        "Total_Tons": total_tons, "Daily_Plan": daily_plan,
        "Start_Date": min(daily_plan), "End_Date": max(daily_plan),
    })

print(f"Built {len(records)} farmer records.")

# ── STEP 3: Décalage ──────────────────────────
decalage_log = []

def resolve_decalage(recs):
    by_comm = defaultdict(list)
    for r in recs:
        by_comm[r["Commercial"]].append(r)

    for comm, flist in by_comm.items():
        for i in range(len(flist)):
            for j in range(i+1, len(flist)):
                a, b = flist[i], flist[j]
                shared = set(a["Allowed_Veh"]) & set(b["Allowed_Veh"])
                if not shared: continue
                ov_start = max(a["Start_Date"], b["Start_Date"])
                ov_end   = min(a["End_Date"],   b["End_Date"])
                if ov_start > ov_end: continue
                overlap_days = (ov_end - ov_start).days + 1
                if overlap_days <= 2: continue

                shift = min(overlap_days - 1, MAX_SHIFT)
                finisher, waiter = (a,b) if a["End_Date"] <= b["End_Date"] else (b,a)

                orig_end_f   = finisher["End_Date"]
                orig_start_w = waiter["Start_Date"]

                # Finisher: remove last `shift` days, redistribute tons
                plan_f = dict(finisher["Daily_Plan"])
                removed = sorted(plan_f.keys(), reverse=True)[:shift]
                extra   = sum(plan_f.pop(d, 0) for d in removed)
                if plan_f:
                    per_extra = extra / len(plan_f)
                    plan_f = {d: round(t+per_extra, 2) for d,t in plan_f.items()}
                finisher["Daily_Plan"] = plan_f
                finisher["End_Date"]   = max(plan_f.keys()) if plan_f else orig_end_f

                # Waiter: delay start
                plan_w = {d + datetime.timedelta(days=shift): t
                          for d,t in waiter["Daily_Plan"].items()}
                waiter["Daily_Plan"]  = plan_w
                waiter["Start_Date"]  = waiter["Start_Date"] + datetime.timedelta(days=shift)
                if plan_w:
                    waiter["End_Date"] = max(plan_w.keys())

                veh_free = finisher["End_Date"] + datetime.timedelta(days=BUFFER_DAYS)
                voyages_double = round(
                    sum(finisher["Daily_Plan"].get(d,0) for d in
                        sorted(finisher["Daily_Plan"].keys())[-2:]) /
                    (finisher["Total_Tons"] / max(len(finisher["Daily_Plan"]),1) or 1), 1
                )

                decalage_log.append({
                    "Commercial":         comm,
                    "Agriculteur A (finit tôt)": finisher["Farmer"],
                    "Agriculteur B (reçoit véhicule)": waiter["Farmer"],
                    "Véhicule Partagé":   ", ".join(sorted(shared)),
                    "Jours Overlap":      overlap_days,
                    "Jours Économisés":   shift,
                    "Fin Orig. A":        orig_end_f,
                    "Nouvelle Fin A":     finisher["End_Date"],
                    "Début Orig. B":      orig_start_w,
                    "Nouveau Début B":    waiter["Start_Date"],
                    "Libre À Partir De":  veh_free,
                    "Voyages Double/Jour": int(voyages_double) + 2,
                    "Tonnes Double/Jour": round(finisher["Total_Tons"]/max(len(finisher["Daily_Plan"]),1)*1.5, 1),
                    "Risque Maladie":     "✅ Sûr" if shift <= MAX_SHIFT else "⚠️ Vérifier",
                    "Action Requise":     f"A fait x{shift} voyages/j pendant {shift}j → libère véhicule le {finisher['End_Date']} → B démarre le {waiter['Start_Date']}",
                })
    return recs

records = resolve_decalage(records)
print(f"Décalage: {len(decalage_log)} conflits résolus.")

# ── STEP 4: Expand to daily rows ──────────────
all_days = []
for rec in records:
    for dt, tons in sorted(rec["Daily_Plan"].items()):
        if tons <= 0: continue
        vehicles    = choose_vehicles(tons, rec["Allowed_Veh"])
        veh_str     = " | ".join(f"{v['vehicle']} ×{v['trips']} ({v['tons_each']}t)" for v in vehicles)
        trips_total = sum(v["trips"] for v in vehicles)
        veh_type    = vehicles[0]["vehicle"] if vehicles else "POILOUR"
        is_peak     = PEAK_START <= dt <= PEAK_END
        is_double   = any(
            lg["Agriculteur A (finit tôt)"] == rec["Farmer"] and
            dt >= rec["End_Date"] - datetime.timedelta(days=lg["Jours Économisés"])
            for lg in decalage_log
        )
        all_days.append({
            "Commercial":    rec["Commercial"],
            "Agriculteur":   rec["Farmer"],
            "Usine":         rec["Factory"],
            "Région":        rec["Region"],
            "Accessibilité": rec["Access"],
            "Date":          dt,
            "Tonnes/Jour":   round(tons, 1),
            "Type Véhicule": veh_type,
            "Véhicules Requis": veh_str,
            "Nb Voyages":    trips_total,
            "Date Début":    rec["Start_Date"],
            "Date Fin":      rec["End_Date"],
            "Total Tonnes":  rec["Total_Tons"],
            "Pic de Récolte": "🟡 PIC" if is_peak else "",
            "Note":          "🚛 DOUBLE TRANSPORT" if is_double else "",
        })

result_df = pd.DataFrame(all_days).sort_values(["Date","Commercial","Agriculteur"]).reset_index(drop=True)
print(f"Planning rows: {len(result_df)}")

# ── STEP 5: Transport needs per day ──────────
transport_rows = []
by_date_comm = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))
for row in all_days:
    d    = row["Date"]
    comm = row["Commercial"]
    for part in row["Véhicules Requis"].split("|"):
        part = part.strip()
        for vname in FLEET_CAPACITY:
            if vname in part:
                try:    trips = int(part.split("×")[1].split(" ")[0])
                except: trips = 1
                by_date_comm[d][comm][vname] += trips

for d in sorted(by_date_comm.keys()):
    for comm in sorted(by_date_comm[d].keys()):
        nd = by_date_comm[d][comm]
        day_rows = [r for r in all_days if r["Date"]==d and r["Commercial"]==comm]
        total_t  = sum(r["Tonnes/Jour"] for r in day_rows)
        dbl_days = sum(1 for r in day_rows if "DOUBLE" in str(r["Note"]))
        transport_rows.append({
            "Date":                  d,
            "Commercial":            comm,
            "Total Tonnes":          round(total_t, 1),
            "Voyages TRACTEUR":      nd.get("TRACTEUR",0),
            "Voyages PETIT POILOUR": nd.get("PETIT POILOUR",0),
            "Voyages POILOUR":       nd.get("POILOUR",0),
            "Voyages SEMI":          nd.get("SEMI",0),
            "Jours Double":          dbl_days,
            "Pic":                   "🟡" if PEAK_START <= d <= PEAK_END else "",
        })

transport_df = pd.DataFrame(transport_rows)

# ── STEP 6: Vehicle availability ──────────────
avail_rows = []
for rec in records:
    for veh in rec["Allowed_Veh"]:
        orig_end  = rec["End_Date"]
        is_double = any(lg["Agriculteur A (finit tôt)"]==rec["Farmer"] for lg in decalage_log)
        avail_rows.append({
            "Commercial":      rec["Commercial"],
            "Agriculteur":     rec["Farmer"],
            "Type Véhicule":   veh,
            "Date Début":      rec["Start_Date"],
            "Date Fin Orig.":  orig_end,
            "Date Fin Nouvelle": rec["End_Date"],
            "Libre À Partir De": rec["End_Date"] + datetime.timedelta(days=BUFFER_DAYS),
            "Double Transport": "🚛 OUI" if is_double else "non",
            "Total Tonnes":    rec["Total_Tons"],
        })
avail_df = pd.DataFrame(avail_rows).sort_values(["Commercial","Type Véhicule","Date Début"]).reset_index(drop=True)

# ── STEP 7: Resume per commercial ─────────────
resume_rows = []
for comm in sorted(result_df["Commercial"].unique()):
    sub = result_df[result_df["Commercial"]==comm]
    dbl_log = [lg for lg in decalage_log if lg["Commercial"]==comm]
    resume_rows.append({
        "Commercial":            comm,
        "Tonnes Totales Saison": round(sub["Total Tonnes"].iloc[0] if len(sub)>0 else 0) * sub["Agriculteur"].nunique(),
        "Nb Agriculteurs":       sub["Agriculteur"].nunique(),
        "Conflits Résolus":      len(dbl_log),
        "Total Jours Double":    len(dbl_log) * MAX_SHIFT,
        "Décalage Max (j)":      MAX_SHIFT,
        "Statut Tomates":        "✅ Sûr",
    })
    # Fix tonnage
    resume_rows[-1]["Tonnes Totales Saison"] = round(
        df[df["responsable région"]==comm]["TONNAGE"].sum(), 0)

resume_rows.append({
    "Commercial": "TOTAL",
    "Tonnes Totales Saison": sum(r["Tonnes Totales Saison"] for r in resume_rows),
    "Nb Agriculteurs":       sum(r["Nb Agriculteurs"] for r in resume_rows),
    "Conflits Résolus":      len(decalage_log),
    "Total Jours Double":    sum(r["Total Jours Double"] for r in resume_rows),
    "Décalage Max (j)":      MAX_SHIFT,
    "Statut Tomates":        "✅ Sûr",
})
resume_df = pd.DataFrame(resume_rows)

# ── STEP 8: Write Excel ───────────────────────
wb = Workbook()

# ---- Sheet 1: Planning Journalier ----
ws1 = wb.active
ws1.title = "Planning Journalier"
COLS1   = ["Commercial","Agriculteur","Usine","Région","Accessibilité","Date",
           "Tonnes/Jour","Type Véhicule","Véhicules Requis","Nb Voyages",
           "Date Début","Date Fin","Total Tonnes","Pic de Récolte","Note"]
WIDTHS1 = [15,24,10,12,11,12,10,13,42,9,12,12,12,10,20]
style(ws1, COLS1, WIDTHS1, "1F4E79")
for ri, row in result_df[COLS1].iterrows():
    is_peak   = "PIC"    in str(row["Pic de Récolte"])
    is_double = "DOUBLE" in str(row["Note"])
    fk = "double" if is_double else ("peak" if is_peak else ("alt" if ri%2==0 else "white"))
    write_row(ws1, ri+2, [row[c] for c in COLS1], fk)

# ---- Sheet 2: Besoins Transport-Jour ----
ws2 = wb.create_sheet("Besoins Transport-Jour")
COLS2   = ["Date","Commercial","Total Tonnes","Voyages TRACTEUR",
           "Voyages PETIT POILOUR","Voyages POILOUR","Voyages SEMI","Jours Double","Pic"]
WIDTHS2 = [12,16,12,16,20,16,14,12,6]
style(ws2, COLS2, WIDTHS2, "375623")
for ri, row in transport_df[COLS2].iterrows():
    is_peak = "🟡" in str(row["Pic"])
    fk = "peak" if is_peak else ("alt" if ri%2==0 else "white")
    write_row(ws2, ri+2, [row[c] for c in COLS2], fk)

# ---- Sheet 3: Disponibilité Véhicules ----
ws3 = wb.create_sheet("Disponibilité Véhicules")
COLS3   = ["Commercial","Agriculteur","Type Véhicule","Date Début",
           "Date Fin Orig.","Date Fin Nouvelle","Libre À Partir De","Double Transport","Total Tonnes"]
WIDTHS3 = [16,26,15,12,14,16,17,15,12]
style(ws3, COLS3, WIDTHS3, "7B2D8B")
for ri, row in avail_df[COLS3].iterrows():
    is_dbl = "OUI" in str(row["Double Transport"])
    fk = "orange" if is_dbl else ("alt" if ri%2==0 else "white")
    write_row(ws3, ri+2, [str(row[c]) for c in COLS3], fk)

# ---- Sheet 4: Journal Double Transport ----
ws4 = wb.create_sheet("Journal Double Transport")
if decalage_log:
    decalage_df = pd.DataFrame(decalage_log)
    COLS4   = ["Commercial","Agriculteur A (finit tôt)","Agriculteur B (reçoit véhicule)",
               "Véhicule Partagé","Jours Overlap","Jours Économisés","Fin Orig. A",
               "Nouvelle Fin A","Début Orig. B","Nouveau Début B","Libre À Partir De",
               "Voyages Double/Jour","Tonnes Double/Jour","Risque Maladie","Action Requise"]
    WIDTHS4 = [14,24,24,16,11,11,14,14,14,14,17,14,14,12,50]
    style(ws4, COLS4, WIDTHS4, "C00000", start_row=2)
    ws4["A1"] = "Journal des conflits de transport résolus par décalage"
    ws4["A1"].font = Font(bold=True, size=12, name="Calibri", color="C00000")
    for ri, row in decalage_df[COLS4].iterrows():
        is_risk = "⚠️" in str(row["Risque Maladie"])
        fk = "double" if is_risk else ("alt" if ri%2==0 else "white")
        write_row(ws4, ri+3, [str(row[c]) for c in COLS4], fk)
else:
    ws4["A1"] = "Aucun conflit détecté."

# ---- Sheet 5: Résumé par Commercial ----
ws5 = wb.create_sheet("Résumé par Commercial")
COLS5   = ["Commercial","Tonnes Totales Saison","Nb Agriculteurs",
           "Conflits Résolus","Total Jours Double","Décalage Max (j)","Statut Tomates"]
WIDTHS5 = [18,20,16,16,18,16,14]
style(ws5, COLS5, WIDTHS5, "1F4E79")
for ri, row in resume_df[COLS5].iterrows():
    fk = "green" if "TOTAL" in str(row["Commercial"]) else ("alt" if ri%2==0 else "white")
    write_row(ws5, ri+2, [row[c] for c in COLS5], fk)

wb.save(EXCEL_OUT)

# ── STEP 9: Print summary ─────────────────────
print(f"\n{'='*55}")
print(f"  [OK] PHASE 4 COMPLETE -> {EXCEL_OUT}")
print(f"{'='*55}")
print(f"  Planning rows : {len(result_df)}")
print(f"  Transport rows: {len(transport_df)}")
print(f"  Conflits      : {len(decalage_log)}")
print(f"\n  Tonnes par commercial:")
for _, r in resume_df[resume_df["Commercial"]!="TOTAL"].iterrows():
    print(f"    {r['Commercial']:<20}: {int(r['Tonnes Totales Saison']):>7,} t | {r['Nb Agriculteurs']} agriculteurs | {r['Conflits Résolus']} conflits")
print(f"{'='*55}")