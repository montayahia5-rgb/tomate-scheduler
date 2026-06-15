"""
Patch dashboard_phase10.py pour utiliser données rectifiées dans tous les tabs.
Usage: python apply_patch.py dashboard_phase10.py
Sortie: dashboard_phase10.py (backup sauvegardé en dashboard_phase10_BACKUP.py)
"""
import sys, shutil, re, os

def apply_patch(filepath):
    with open(filepath, 'r', encoding='utf-8') as f:
        src = f.read()

    # Backup
    shutil.copy2(filepath, filepath.replace('.py', '_BACKUP.py'))
    print(f"Backup: {filepath.replace('.py','_BACKUP.py')}")

    original_len = len(src)
    patches_applied = []

    # ─────────────────────────────────────────────────────────
    # PATCH 0 — Ajouter import RECTIF depuis comparaison_tab
    # ─────────────────────────────────────────────────────────
    OLD_IMPORT = '''try:
    from comparaison_tab import render_comparaison_tab
    COMPARAISON_AVAILABLE = True
except ImportError:
    COMPARAISON_AVAILABLE = False'''

    NEW_IMPORT = '''try:
    from comparaison_tab import render_comparaison_tab
    COMPARAISON_AVAILABLE = True
except ImportError:
    COMPARAISON_AVAILABLE = False

# Import données rectifiées (SOURCE PRINCIPALE — reference_interne 13/06/2026)
try:
    from comparaison_tab import (
        RECTIF_COMM_DAILY, RECTIF_USINE_DAILY,
        RECTIF_COMM_DICT, RECTIF_USINE_DICT,
        RECTIF_STATS_COMM, RECTIF_STATS_USINE,
    )
    RECTIF_AVAILABLE = True
except ImportError:
    RECTIF_AVAILABLE = False
    # Fallback vide si comparaison_tab absent
    RECTIF_COMM_DAILY = {}; RECTIF_USINE_DAILY = {}
    RECTIF_COMM_DICT  = {}; RECTIF_USINE_DICT  = {}
    RECTIF_STATS_COMM = {}; RECTIF_STATS_USINE = {}'''

    if OLD_IMPORT in src:
        src = src.replace(OLD_IMPORT, NEW_IMPORT, 1)
        patches_applied.append("PATCH 0 — Import RECTIF")
    else:
        print("AVERTISSEMENT: PATCH 0 — string import non trouvé")

    # ─────────────────────────────────────────────────────────
    # Ajouter helper _build_rectif_daily_df juste après les imports
    # ─────────────────────────────────────────────────────────
    HELPER_ANCHOR = "# set_page_config MUST be the very first Streamlit call"

    HELPER_CODE = '''# ── Helpers données rectifiées ─────────────────────────────
def _rectif_total_daily():
    """DataFrame total journalier depuis données rectifiées (tous commerciaux)"""
    import pandas as _pd
    rows = []
    for comm, vals in RECTIF_COMM_DAILY.items():
        for i, v in enumerate(vals):
            if v > 0:
                d = _pd.Timestamp("2026-06-20") + _pd.Timedelta(days=i)
                rows.append({"Date": d, "Tonnes/Jour": float(v), "Commercial": comm})
    if not rows:
        return _pd.DataFrame(columns=["Date","Tonnes/Jour","Commercial"])
    df = _pd.DataFrame(rows)
    return df.groupby("Date")["Tonnes/Jour"].sum().reset_index()

def _rectif_comm_daily_df():
    """DataFrame par commercial depuis données rectifiées"""
    import pandas as _pd
    rows = []
    for comm, vals in RECTIF_COMM_DAILY.items():
        for i, v in enumerate(vals):
            if v > 0:
                d = _pd.Timestamp("2026-06-20") + _pd.Timedelta(days=i)
                rows.append({"Date": d, "Commercial": comm, "Tonnes/Jour": float(v)})
    if not rows:
        return _pd.DataFrame(columns=["Date","Commercial","Tonnes/Jour"])
    return _pd.DataFrame(rows)

def _rectif_one_comm_daily(comm):
    """DataFrame journalier pour un commercial depuis données rectifiées"""
    import pandas as _pd
    vals = RECTIF_COMM_DAILY.get(comm, [])
    rows = [{"Date": _pd.Timestamp("2026-06-20")+_pd.Timedelta(days=i), "Tonnes/Jour": float(v)}
            for i, v in enumerate(vals) if v > 0]
    return _pd.DataFrame(rows) if rows else _pd.DataFrame(columns=["Date","Tonnes/Jour"])

def _rectif_usine_daily_df():
    """DataFrame par usine depuis données rectifiées"""
    import pandas as _pd
    rows = []
    for usine, vals in RECTIF_USINE_DAILY.items():
        for i, v in enumerate(vals):
            if v > 0:
                d = _pd.Timestamp("2026-06-20") + _pd.Timedelta(days=i)
                rows.append({"Date": d, "Usine": usine, "Tonnes/Jour": float(v)})
    if not rows:
        return _pd.DataFrame(columns=["Date","Usine","Tonnes/Jour"])
    return _pd.DataFrame(rows)

def _build_rectif_wide_excel(planning_df_ort=None):
    """
    Génère Excel avec colonnes dates (structure wide) basé sur données rectifiées.
    Colonnes: Date | Commercial | Agriculteur | Usine | Tonnes/Jour | Type Véhicule |
              Véhicules Requis | Disponibles | Manquants (à louer) | Nb Voyages |
              Pic de Récolte | 20/06/2026 | 21/06/2026 | ... | 25/08/2026
    """
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io as _io, pandas as _pd

    SEASON_DATES = list(_pd.date_range("2026-06-20","2026-08-25",freq="D"))
    PIC_S_ = _pd.Timestamp("2026-07-01").date()
    PIC_E_ = _pd.Timestamp("2026-07-15").date()

    wb = Workbook(); wb.remove(wb.active)
    ws = wb.create_sheet("Planning Rectifié (Large)")

    HDR_FILL = PatternFill("solid", start_color="1F3864", end_color="1F3864")
    HDR_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=9)
    PIC_FILL = PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC")
    GRN_FILL = PatternFill("solid", start_color="E2EFDA", end_color="E2EFDA")
    ALT_FILL = PatternFill("solid", start_color="F8F9FA", end_color="F8F9FA")
    WHT_FILL = PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")
    RED_FILL = PatternFill("solid", start_color="FFC7CE", end_color="FFC7CE")
    THIN     = Side(style="thin", color="CCCCCC")
    BORD     = Border(left=THIN,right=THIN,top=THIN,bottom=THIN)
    CTR      = Alignment(horizontal="center",vertical="center")
    LFT      = Alignment(horizontal="left",  vertical="center")

    COMM_FILLS = {"FEDI":"DEEBF7","MAKKI BEN SALAH":"E2EFDA","KHALIL":"FFF2CC",
                  "ACHREF AJLANI":"EDEDED","JILANI OBAY":"FCE4D6"}
    FLEET      = {"SICAM":{"PL":48,"PPL":6,"SEMI":13},
                  "TUCAL":{"PL":17,"PPL":0,"SEMI":2},
                  "COMOCAP":{"PL":6,"PPL":14,"SEMI":3},
                  "ABIDA":{"PL":1,"PPL":0,"SEMI":2},
                  "ELFALLEH":{"PL":0,"PPL":2,"SEMI":0}}

    FIXED_HDRS = ["Date","Commercial","Agriculteur","Usine","Tonnes/Jour",
                  "Type Véhicule","Véhicules Requis","Disponibles","Manquants (à louer)",
                  "Nb Voyages","Pic de Récolte"]
    DATE_HDRS  = [d.strftime("%d/%m/%Y") for d in SEASON_DATES]
    ALL_HDRS   = FIXED_HDRS + DATE_HDRS
    N = len(ALL_HDRS)

    # Titre
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=min(N,35))
    ws.cell(1,1).value = "Planning Rectifié — SOURCE: reference_interne 13/06/2026"
    ws.cell(1,1).font  = Font(bold=True,color="FFFFFF",name="Calibri",size=12)
    ws.cell(1,1).fill  = HDR_FILL; ws.cell(1,1).alignment = CTR
    ws.row_dimensions[1].height = 28

    # En-têtes
    for ci, h in enumerate(ALL_HDRS, 1):
        c = ws.cell(2, ci); c.value = h; c.border = BORD; c.alignment = CTR
        if ci <= len(FIXED_HDRS):
            c.fill = HDR_FILL; c.font = HDR_FONT
        else:
            d_obj = SEASON_DATES[ci-len(FIXED_HDRS)-1].date()
            c.fill = PIC_FILL if PIC_S_<=d_obj<=PIC_E_ else PatternFill("solid",start_color="1F4E79",end_color="1F4E79")
            c.font = Font(bold=True,color="FFFFFF",name="Calibri",size=8)
    ws.row_dimensions[2].height = 36

    comm_order = ["FEDI","MAKKI BEN SALAH","KHALIL","ACHREF AJLANI","JILANI OBAY"]
    data_row   = 3

    for comm in comm_order:
        rectif = RECTIF_COMM_DICT.get(comm, {})
        cfill  = PatternFill("solid", start_color=COMM_FILLS.get(comm,"F0F0F0"), end_color=COMM_FILLS.get(comm,"F0F0F0"))

        # Utiliser planning OR-Tools pour structure agriculteurs si disponible
        if planning_df_ort is not None and not planning_df_ort.empty and "Commercial" in planning_df_ort.columns:
            sub = planning_df_ort[planning_df_ort["Commercial"]==comm].copy()
            if not sub.empty and "Agriculteur" in sub.columns:
                sub["Date"] = _pd.to_datetime(sub["Date"],errors="coerce")
                grps = sub.groupby(["Agriculteur","Usine"] if "Usine" in sub.columns else ["Agriculteur"])
                for key, grp in grps:
                    agri  = key[0] if isinstance(key,tuple) else key
                    usine = key[1] if isinstance(key,tuple) and len(key)>1 else ""
                    ort_d = {}
                    if "Date" in grp.columns and "Tonnes/Jour" in grp.columns:
                        for _, r in grp.iterrows():
                            if _pd.notna(r["Date"]):
                                ort_d[str(r["Date"].date())] = float(r.get("Tonnes/Jour",0) or 0)
                    r0 = grp.iloc[0]
                    tv  = str(r0.get("Type Véhicule","") or "").upper()
                    vr  = str(r0.get("Véhicules Requis","") or "")
                    nv  = str(r0.get("Nb Voyages","") or "")
                    vr_n= int(_pd.to_numeric(vr,errors="coerce") or 0)
                    dispo= FLEET.get(usine.upper(),{}).get(tv,0)
                    manq = max(0,vr_n-dispo)
                    total_ort = sum(ort_d.values())
                    is_pic = any(PIC_S_<=_pd.Timestamp(k).date()<=PIC_E_ for k,v in ort_d.items() if v>0)

                    alt = data_row%2==0; bf = ALT_FILL if alt else WHT_FILL
                    fv = ["",comm,agri,usine,round(total_ort,0) if total_ort>0 else "",
                          tv,vr,dispo if vr_n>0 else "",manq if vr_n>0 else "",nv,
                          "⚡ PIC" if is_pic else ""]
                    for ci, val in enumerate(fv,1):
                        cell=ws.cell(data_row,ci); cell.value=val; cell.border=BORD; cell.alignment=CTR
                        if ci==2: cell.fill=cfill; cell.font=Font(bold=True,name="Calibri",size=9)
                        elif ci==9 and isinstance(val,int) and val>0:
                            cell.fill=RED_FILL; cell.font=Font(bold=True,name="Calibri",size=9,color="9C0006")
                        else: cell.fill=bf; cell.font=Font(name="Calibri",size=9)
                    for di,d in enumerate(SEASON_DATES):
                        dk=str(d.date()); val=ort_d.get(dk,"")
                        if val==0: val=""
                        ci2=len(FIXED_HDRS)+di+1
                        cell=ws.cell(data_row,ci2)
                        cell.value=int(val) if val!="" else ""; cell.border=BORD; cell.alignment=CTR
                        is_pic_d=PIC_S_<=d.date()<=PIC_E_
                        cell.fill=PIC_FILL if (is_pic_d and val!="") else (GRN_FILL if val!="" else bf)
                        cell.font=Font(name="Calibri",size=8)
                    ws.row_dimensions[data_row].height=15; data_row+=1
                continue  # prochain commercial

        # Fallback: ligne agrégée rectifiée
        total_rect=sum(float(v) for v in rectif.values())
        alt=data_row%2==0; bf=ALT_FILL if alt else WHT_FILL
        fv=["",comm,"(total commercial)","Toutes usines",round(total_rect,0),"","","","","",
            "⚡ PIC" if any(PIC_S_<=_pd.Timestamp(k).date()<=PIC_E_ for k,v in rectif.items() if float(v)>0) else ""]
        for ci,val in enumerate(fv,1):
            cell=ws.cell(data_row,ci); cell.value=val; cell.border=BORD; cell.alignment=CTR
            cell.fill=cfill if ci==2 else bf; cell.font=Font(name="Calibri",size=9,bold=(ci==2))
        for di,d in enumerate(SEASON_DATES):
            dk=str(d.date()); val=float(rectif.get(dk,0) or 0)
            ci2=len(FIXED_HDRS)+di+1
            cell=ws.cell(data_row,ci2)
            cell.value=int(val) if val>0 else ""; cell.border=BORD; cell.alignment=CTR
            is_pic_d=PIC_S_<=d.date()<=PIC_E_
            cell.fill=PIC_FILL if (is_pic_d and val>0) else (GRN_FILL if val>0 else bf)
            cell.font=Font(name="Calibri",size=8)
        ws.row_dimensions[data_row].height=15; data_row+=1

    # Largeurs colonnes
    for ci,w in enumerate([12,16,28,12,11,12,11,11,13,10,11],1):
        ws.column_dimensions[get_column_letter(ci)].width=w
    for di in range(len(SEASON_DATES)):
        ws.column_dimensions[get_column_letter(len(FIXED_HDRS)+di+1)].width=8
    ws.freeze_panes="L3"

    # Onglet récap journalier
    ws2=wb.create_sheet("Recap Journalier Rectifié")
    ws2.merge_cells(start_row=1,start_column=1,end_row=1,end_column=8)
    ws2.cell(1,1).value="Récapitulatif journalier — Données Rectifiées (reference_interne)"
    ws2.cell(1,1).fill=HDR_FILL; ws2.cell(1,1).font=Font(bold=True,color="FFFFFF",name="Calibri",size=11)
    ws2.cell(1,1).alignment=CTR; ws2.row_dimensions[1].height=24

    comm_order2=["FEDI","MAKKI BEN SALAH","KHALIL","ACHREF AJLANI","JILANI OBAY"]
    hdrs2=["Date","Jour"]+[c.split()[0] for c in comm_order2]+["TOTAL"]
    for ci,h in enumerate(hdrs2,1):
        cell=ws2.cell(2,ci); cell.value=h; cell.fill=HDR_FILL; cell.font=HDR_FONT
        cell.alignment=CTR; cell.border=BORD
    ws2.row_dimensions[2].height=20
    DAYS_FR2=["Lun","Mar","Mer","Jeu","Ven","Sam","Dim"]
    for ri,d in enumerate(SEASON_DATES):
        dk=d.date(); is_pic=PIC_S_<=dk<=PIC_E_
        vals=[float(RECTIF_COMM_DICT.get(c,{}).get(str(dk),0) or 0) for c in comm_order2]
        total_j=sum(vals)
        if total_j==0: continue
        base=PIC_FILL if is_pic else (ALT_FILL if ri%2==0 else WHT_FILL)
        row_vals=[d.strftime("%d/%m/%Y"),DAYS_FR2[d.weekday()]]+[int(v) if v>0 else "" for v in vals]+[int(total_j)]
        for ci,val in enumerate(row_vals,1):
            cell=ws2.cell(ri+3,ci); cell.value=val; cell.border=BORD; cell.alignment=CTR
            cell.fill=base; cell.font=Font(name="Calibri",size=9,bold=(ci==8))
        ws2.row_dimensions[ri+3].height=15
    for ci,w in enumerate([12,6,12,12,12,12,12,14],1):
        ws2.column_dimensions[get_column_letter(ci)].width=w
    ws2.freeze_panes="A3"

    buf=_io.BytesIO(); wb.save(buf); buf.seek(0); return buf.read()

# ── Fin helpers ─────────────────────────────────────────────
'''

    if HELPER_ANCHOR in src:
        src = src.replace(HELPER_ANCHOR, HELPER_CODE + "\n" + HELPER_ANCHOR, 1)
        patches_applied.append("PATCH 0b — Helpers _rectif_*")
    else:
        print("AVERTISSEMENT: PATCH 0b — anchor non trouvé, helpers insérés différemment")
        # Insérer après l'import comparaison_tab
        src = src.replace(NEW_IMPORT, NEW_IMPORT + "\n\n" + HELPER_CODE, 1)
        patches_applied.append("PATCH 0b — Helpers (fallback insertion)")

    # ─────────────────────────────────────────────────────────
    # PATCH 1 — Tab 1: Graphique quotidien total (données rectifiées)
    # ─────────────────────────────────────────────────────────
    OLD_TAB1_DAILY = '''    daily = p.groupby("Date")["Tonnes/Jour"].sum().reset_index()
    daily["Période"] = daily["Date"].apply(
        lambda d: "⚡ Pic (1-15 Jul)" if PEAK_START <= d.date() <= PEAK_END else "Normal"
    )'''

    NEW_TAB1_DAILY = '''    # SOURCE = Données Rectifiées (reference_interne 13/06/2026)
    if RECTIF_AVAILABLE and RECTIF_COMM_DAILY:
        daily = _rectif_total_daily()
        daily["Période"] = daily["Date"].apply(
            lambda d: "⚡ Pic (1-15 Jul)" if PEAK_START <= d.date() <= PEAK_END else "Normal"
        )
    else:
        daily = p.groupby("Date")["Tonnes/Jour"].sum().reset_index()
        daily["Période"] = daily["Date"].apply(
            lambda d: "⚡ Pic (1-15 Jul)" if PEAK_START <= d.date() <= PEAK_END else "Normal"
        )'''

    if OLD_TAB1_DAILY in src:
        src = src.replace(OLD_TAB1_DAILY, NEW_TAB1_DAILY, 1)
        patches_applied.append("PATCH 1 — Tab1 daily chart → rectifié")
    else:
        print("AVERTISSEMENT: PATCH 1 non trouvé")

    # ─────────────────────────────────────────────────────────
    # PATCH 1b — Tab 1: Boutons export — ajouter bouton WIDE FORMAT
    # ─────────────────────────────────────────────────────────
    OLD_EXPORT_BTNS = '''    col_dl1, col_dl2, col_dl3 = st.columns(3)
    with col_dl1:
        st.download_button(
            "📅 Excel SÉPARÉ PAR JOUR",
            data=df_to_xlsx_by_day(_planning_export, date_col="Date"),
            file_name="planning_par_jour.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True,
            help="Un onglet par jour + un onglet récapitulatif",
        )
    with col_dl2:
        st.download_button(
            "📊 Excel TOUT EN 1 ONGLET",
            data=df_to_xlsx_styled(_planning_export, sheet_name="Planning"),
            file_name="planning_journalier.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    with col_dl3:
        st.download_button(
            "⬇️ CSV brut",
            data=df_to_csv(_planning_export),
            file_name="planning_journalier.csv",
            mime="text/csv",
            use_container_width=True,
        )'''

    NEW_EXPORT_BTNS = '''    col_dl0, col_dl1, col_dl2, col_dl3 = st.columns(4)
    with col_dl0:
        # ✅ NOUVEAU — Export RECTIFIÉ avec colonnes dates
        if RECTIF_AVAILABLE:
            st.download_button(
                "📆 Planning RECTIFIÉ (colonnes dates)",
                data=_build_rectif_wide_excel(p if not p.empty else None),
                file_name="planning_rectifie_colonnes_dates.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True,
                help="Colonnes: Date|Commercial|Agriculteur|Usine|...|20/06|21/06|...|25/08 — Source: Plans Rectifiés",
            )
        else:
            st.info("comparaison_tab.py requis pour export rectifié")
    with col_dl1:
        st.download_button(
            "📅 Excel SÉPARÉ PAR JOUR",
            data=df_to_xlsx_by_day(_planning_export, date_col="Date"),
            file_name="planning_par_jour.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            help="Un onglet par jour + un onglet récapitulatif",
        )
    with col_dl2:
        st.download_button(
            "📊 Excel TOUT EN 1 ONGLET",
            data=df_to_xlsx_styled(_planning_export, sheet_name="Planning"),
            file_name="planning_journalier.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    with col_dl3:
        st.download_button(
            "⬇️ CSV brut",
            data=df_to_csv(_planning_export),
            file_name="planning_journalier.csv",
            mime="text/csv",
            use_container_width=True,
        )'''

    if OLD_EXPORT_BTNS in src:
        src = src.replace(OLD_EXPORT_BTNS, NEW_EXPORT_BTNS, 1)
        patches_applied.append("PATCH 1b — Tab1 export → bouton RECTIFIÉ colonnes dates")
    else:
        print("AVERTISSEMENT: PATCH 1b (export buttons) non trouvé")

    # ─────────────────────────────────────────────────────────
    # PATCH 2 — Tab 2: Courbe ligne par commercial
    # ─────────────────────────────────────────────────────────
    OLD_TAB2_COMM = '''    comm_daily = (p.groupby(["Date","Commercial"])["Tonnes/Jour"]
                  .sum().reset_index())'''

    NEW_TAB2_COMM = '''    # SOURCE = Données Rectifiées
    if RECTIF_AVAILABLE and RECTIF_COMM_DAILY:
        comm_daily = _rectif_comm_daily_df()
    else:
        comm_daily = (p.groupby(["Date","Commercial"])["Tonnes/Jour"]
                      .sum().reset_index())'''

    if OLD_TAB2_COMM in src:
        src = src.replace(OLD_TAB2_COMM, NEW_TAB2_COMM, 1)
        patches_applied.append("PATCH 2 — Tab2 comm_daily → rectifié")
    else:
        print("AVERTISSEMENT: PATCH 2 non trouvé")

    # ─────────────────────────────────────────────────────────
    # PATCH 3 — Tab 2: Drill-down un commercial
    # ─────────────────────────────────────────────────────────
    OLD_TAB2_ONE = '''    one_daily = one.groupby("Date")["Tonnes/Jour"].sum().reset_index()
    fig8 = px.area('''

    NEW_TAB2_ONE = '''    # SOURCE = Données Rectifiées pour le drill-down
    if RECTIF_AVAILABLE and selected in RECTIF_COMM_DAILY:
        one_daily = _rectif_one_comm_daily(selected)
    else:
        one_daily = one.groupby("Date")["Tonnes/Jour"].sum().reset_index()
    fig8 = px.area('''

    if OLD_TAB2_ONE in src:
        src = src.replace(OLD_TAB2_ONE, NEW_TAB2_ONE, 1)
        patches_applied.append("PATCH 3 — Tab2 drill-down one_daily → rectifié")
    else:
        print("AVERTISSEMENT: PATCH 3 non trouvé")

    # ─────────────────────────────────────────────────────────
    # PATCH 4 — Tab 3: Courbe par usine
    # ─────────────────────────────────────────────────────────
    OLD_TAB3_FACT = '''        fact_daily = p.groupby(["Date","Usine"])["Tonnes/Jour"].sum().reset_index()'''

    NEW_TAB3_FACT = '''        # SOURCE = Données Rectifiées
        if RECTIF_AVAILABLE and RECTIF_USINE_DAILY:
            fact_daily = _rectif_usine_daily_df()
        else:
            fact_daily = p.groupby(["Date","Usine"])["Tonnes/Jour"].sum().reset_index()'''

    if OLD_TAB3_FACT in src:
        src = src.replace(OLD_TAB3_FACT, NEW_TAB3_FACT, 1)
        patches_applied.append("PATCH 4 — Tab3 fact_daily (usine) → rectifié")
    else:
        print("AVERTISSEMENT: PATCH 4 non trouvé")

    # ─────────────────────────────────────────────────────────
    # PATCH 5 — KPIs globaux : utiliser totaux rectifiés
    # ─────────────────────────────────────────────────────────
    OLD_TOTAL_TONS = '''total_tons   = GLOBAL_TOTAL_TONS'''

    NEW_TOTAL_TONS = '''# ── KPIs: priorité aux totaux rectifiés ─────────────────
if RECTIF_AVAILABLE and RECTIF_STATS_COMM:
    # Totaux depuis données rectifiées (source principale)
    _rectif_comm_total = sum(s.get("total",0) for s in RECTIF_STATS_COMM.values())
    _rectif_usine_total= sum(s.get("total",0) for s in RECTIF_STATS_USINE.values())
    total_tons = max(_rectif_comm_total, GLOBAL_TOTAL_TONS)  # garder le plus précis
else:
    total_tons   = GLOBAL_TOTAL_TONS'''

    if OLD_TOTAL_TONS in src:
        src = src.replace(OLD_TOTAL_TONS, NEW_TOTAL_TONS, 1)
        patches_applied.append("PATCH 5 — KPI total_tons → priorité rectifié")
    else:
        print("AVERTISSEMENT: PATCH 5 non trouvé")

    # ─────────────────────────────────────────────────────────
    # PATCH 6 — Tab 1: Pie chart par commercial — utiliser rectifié
    # ─────────────────────────────────────────────────────────
    OLD_PIE_COMM = '''        if GLOBAL_COMMERCIAL_TONS:
            comm_df = pd.DataFrame(list(GLOBAL_COMMERCIAL_TONS.items()),
                                    columns=["Commercial","Tonnes/Jour"])
        else:
            comm_df = p.groupby("Commercial")["Tonnes/Jour"].sum().reset_index()'''

    NEW_PIE_COMM = '''        # SOURCE PIE = Rectifié en priorité
        if RECTIF_AVAILABLE and RECTIF_STATS_COMM:
            comm_df = pd.DataFrame([{"Commercial":c,"Tonnes/Jour":s.get("total",0)}
                                     for c,s in RECTIF_STATS_COMM.items()])
        elif GLOBAL_COMMERCIAL_TONS:
            comm_df = pd.DataFrame(list(GLOBAL_COMMERCIAL_TONS.items()),
                                    columns=["Commercial","Tonnes/Jour"])
        else:
            comm_df = p.groupby("Commercial")["Tonnes/Jour"].sum().reset_index()'''

    if OLD_PIE_COMM in src:
        src = src.replace(OLD_PIE_COMM, NEW_PIE_COMM, 1)
        patches_applied.append("PATCH 6 — Tab1 pie comm → rectifié")
    else:
        print("AVERTISSEMENT: PATCH 6 non trouvé")

    # ─────────────────────────────────────────────────────────
    # PATCH 7 — Tab 1: Pie chart par usine — utiliser rectifié
    # ─────────────────────────────────────────────────────────
    OLD_PIE_USINE = '''            if GLOBAL_USINE_TONS:
                usine_df = pd.DataFrame(list(GLOBAL_USINE_TONS.items()),
                                         columns=["Usine","Tonnes/Jour"])
            else:
                usine_df = p.groupby("Usine")["Tonnes/Jour"].sum().reset_index()'''

    NEW_PIE_USINE = '''            # SOURCE PIE USINE = Rectifié en priorité
            if RECTIF_AVAILABLE and RECTIF_STATS_USINE:
                usine_df = pd.DataFrame([{"Usine":u,"Tonnes/Jour":s.get("total",0)}
                                          for u,s in RECTIF_STATS_USINE.items()])
            elif GLOBAL_USINE_TONS:
                usine_df = pd.DataFrame(list(GLOBAL_USINE_TONS.items()),
                                         columns=["Usine","Tonnes/Jour"])
            else:
                usine_df = p.groupby("Usine")["Tonnes/Jour"].sum().reset_index()'''

    if OLD_PIE_USINE in src:
        src = src.replace(OLD_PIE_USINE, NEW_PIE_USINE, 1)
        patches_applied.append("PATCH 7 — Tab1 pie usine → rectifié")
    else:
        print("AVERTISSEMENT: PATCH 7 non trouvé")

    # ─────────────────────────────────────────────────────────
    # PATCH 8 — Tab 2: Bar chart tonnes totales par commercial
    # ─────────────────────────────────────────────────────────
    OLD_BAR_COMM = '''        if GLOBAL_COMMERCIAL_TONS:
            comm_tot = pd.DataFrame(list(GLOBAL_COMMERCIAL_TONS.items()),
                                     columns=["Commercial","Tonnes/Jour"])
        else:
            comm_tot = p.groupby("Commercial")["Tonnes/Jour"].sum().reset_index()'''

    NEW_BAR_COMM = '''        # SOURCE BAR COMM = Rectifié
        if RECTIF_AVAILABLE and RECTIF_STATS_COMM:
            comm_tot = pd.DataFrame([{"Commercial":c,"Tonnes/Jour":s.get("total",0)}
                                      for c,s in RECTIF_STATS_COMM.items()])
        elif GLOBAL_COMMERCIAL_TONS:
            comm_tot = pd.DataFrame(list(GLOBAL_COMMERCIAL_TONS.items()),
                                     columns=["Commercial","Tonnes/Jour"])
        else:
            comm_tot = p.groupby("Commercial")["Tonnes/Jour"].sum().reset_index()'''

    if OLD_BAR_COMM in src:
        src = src.replace(OLD_BAR_COMM, NEW_BAR_COMM, 1)
        patches_applied.append("PATCH 8 — Tab2 bar totaux comm → rectifié")
    else:
        print("AVERTISSEMENT: PATCH 8 non trouvé")

    # ─────────────────────────────────────────────────────────
    # PATCH 9 — Tab 3: Bar chart totaux usines
    # ─────────────────────────────────────────────────────────
    OLD_BAR_USINE = '''            if GLOBAL_USINE_TONS:
                fact_tot = pd.DataFrame(list(GLOBAL_USINE_TONS.items()),
                                         columns=["Usine","Tonnes/Jour"])
            else:
                fact_tot = p.groupby("Usine")["Tonnes/Jour"].sum().reset_index()'''

    NEW_BAR_USINE = '''            # SOURCE BAR USINE = Rectifié
            if RECTIF_AVAILABLE and RECTIF_STATS_USINE:
                fact_tot = pd.DataFrame([{"Usine":u,"Tonnes/Jour":s.get("total",0)}
                                          for u,s in RECTIF_STATS_USINE.items()])
            elif GLOBAL_USINE_TONS:
                fact_tot = pd.DataFrame(list(GLOBAL_USINE_TONS.items()),
                                         columns=["Usine","Tonnes/Jour"])
            else:
                fact_tot = p.groupby("Usine")["Tonnes/Jour"].sum().reset_index()'''

    if OLD_BAR_USINE in src:
        src = src.replace(OLD_BAR_USINE, NEW_BAR_USINE, 1)
        patches_applied.append("PATCH 9 — Tab3 bar totaux usine → rectifié")
    else:
        print("AVERTISSEMENT: PATCH 9 non trouvé")

    # ─────────────────────────────────────────────────────────
    # Sauvegarder
    # ─────────────────────────────────────────────────────────
    with open(filepath, 'w', encoding='utf-8') as f:
        f.write(src)

    new_len = len(src)
    print(f"\n✅ Patch terminé: {len(patches_applied)} patches appliqués")
    print(f"   Taille: {original_len} → {new_len} chars (+{new_len-original_len})")
    for p_name in patches_applied:
        print(f"   ✓ {p_name}")
    return patches_applied

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python apply_patch.py dashboard_phase10.py")
        sys.exit(1)
    apply_patch(sys.argv[1])

def apply_extra_patches(filepath):
    """Patches supplémentaires avec indentation alternative"""
    with open(filepath, 'r', encoding='utf-8') as f:
        src = f.read()
    
    applied = []
    
    # Pie comm (variante indentation)
    for old_var, new_var, label in [
        # Pie comm Tab1
        (
        '''    with c1:
        # Use DECLARED tonnage from agriculteurs (correct) not planning sum (filtered/incomplete)
        if GLOBAL_COMMERCIAL_TONS:
            comm_df = pd.DataFrame(list(GLOBAL_COMMERCIAL_TONS.items()),
                                    columns=["Commercial","Tonnes/Jour"])
        else:
            comm_df = p.groupby("Commercial")["Tonnes/Jour"].sum().reset_index()''',
        '''    with c1:
        # SOURCE PIE = Rectifié en priorité
        if RECTIF_AVAILABLE and RECTIF_STATS_COMM:
            comm_df = pd.DataFrame([{"Commercial":c,"Tonnes/Jour":s.get("total",0)}
                                     for c,s in RECTIF_STATS_COMM.items()])
        elif GLOBAL_COMMERCIAL_TONS:
            comm_df = pd.DataFrame(list(GLOBAL_COMMERCIAL_TONS.items()),
                                    columns=["Commercial","Tonnes/Jour"])
        else:
            comm_df = p.groupby("Commercial")["Tonnes/Jour"].sum().reset_index()''',
        "PATCH 6b — Tab1 pie comm (avec contexte)"
        ),
        # Pie usine Tab1
        (
        '''            if GLOBAL_USINE_TONS:
                usine_df = pd.DataFrame(list(GLOBAL_USINE_TONS.items()),
                                         columns=["Usine","Tonnes/Jour"])
            else:
                usine_df = p.groupby("Usine")["Tonnes/Jour"].sum().reset_index()
            fig3 = px.pie(''',
        '''            # SOURCE PIE USINE = Rectifié
            if RECTIF_AVAILABLE and RECTIF_STATS_USINE:
                usine_df = pd.DataFrame([{"Usine":u,"Tonnes/Jour":s.get("total",0)}
                                          for u,s in RECTIF_STATS_USINE.items()])
            elif GLOBAL_USINE_TONS:
                usine_df = pd.DataFrame(list(GLOBAL_USINE_TONS.items()),
                                         columns=["Usine","Tonnes/Jour"])
            else:
                usine_df = p.groupby("Usine")["Tonnes/Jour"].sum().reset_index()
            fig3 = px.pie(''',
        "PATCH 7b — Tab1 pie usine (avec contexte)"
        ),
        # Bar comm Tab2
        (
        '''    with c1:
        # Use DECLARED tonnage from agriculteurs table (source of truth)
        # NOT planning sum which is filtered/partial
        if GLOBAL_COMMERCIAL_TONS:
            comm_tot = pd.DataFrame(list(GLOBAL_COMMERCIAL_TONS.items()),
                                     columns=["Commercial","Tonnes/Jour"])
        else:
            comm_tot = p.groupby("Commercial")["Tonnes/Jour"].sum().reset_index()''',
        '''    with c1:
        # SOURCE BAR COMM = Rectifié
        if RECTIF_AVAILABLE and RECTIF_STATS_COMM:
            comm_tot = pd.DataFrame([{"Commercial":c,"Tonnes/Jour":s.get("total",0)}
                                      for c,s in RECTIF_STATS_COMM.items()])
        elif GLOBAL_COMMERCIAL_TONS:
            comm_tot = pd.DataFrame(list(GLOBAL_COMMERCIAL_TONS.items()),
                                     columns=["Commercial","Tonnes/Jour"])
        else:
            comm_tot = p.groupby("Commercial")["Tonnes/Jour"].sum().reset_index()''',
        "PATCH 8b — Tab2 bar totaux comm (avec contexte)"
        ),
        # Bar usine Tab3
        (
        '''        with c2:
            # Use declared tonnage from agriculteurs (correct totals per usine)
            if GLOBAL_USINE_TONS:
                fact_tot = pd.DataFrame(list(GLOBAL_USINE_TONS.items()),
                                         columns=["Usine","Tonnes/Jour"])
            else:
                fact_tot = p.groupby("Usine")["Tonnes/Jour"].sum().reset_index()''',
        '''        with c2:
            # SOURCE BAR USINE = Rectifié
            if RECTIF_AVAILABLE and RECTIF_STATS_USINE:
                fact_tot = pd.DataFrame([{"Usine":u,"Tonnes/Jour":s.get("total",0)}
                                          for u,s in RECTIF_STATS_USINE.items()])
            elif GLOBAL_USINE_TONS:
                fact_tot = pd.DataFrame(list(GLOBAL_USINE_TONS.items()),
                                         columns=["Usine","Tonnes/Jour"])
            else:
                fact_tot = p.groupby("Usine")["Tonnes/Jour"].sum().reset_index()''',
        "PATCH 9b — Tab3 bar totaux usine (avec contexte)"
        ),
    ]:
        if old_var in src:
            src = src.replace(old_var, new_var, 1)
            applied.append(label)
        else:
            print(f"  INFO: {label} — string non trouvé (peut-être déjà patché)")
    
    with open(filepath, 'w', encoding='utf-8') as f:
        f.write(src)
    
    for a in applied:
        print(f"  ✓ {a}")
    return applied