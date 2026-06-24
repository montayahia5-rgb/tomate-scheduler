# -*- coding: utf-8 -*-
"""
agro_performance.py — Module Performance Agronomique v2
=======================================================
3 sources de données :
  • ROYAL  / Bourak  → Plan Livré / Extra / Plan Actif (plants)
  • SOTUSFA          → Intrants (engrais, pesticides, irrigation)
  • Saisie manuelle  → Hectares + Tonnage récolté

Calculs clés :
  Taux de prise (%)   = Plan Actif / Plan Livré × 100
  kg tomate/plant     = Tonnage(t)×1000 / Plan Actif
  DAP/plant actif     = DAP(kg) / Plan Actif
  Coût intrants/tonne = Total Intrants(TND) / Tonnage(t)
  Rendement t/ha      = Tonnage(t) / Hectares

Intégration dashboard_phase10.py :
  from agro_performance import render_agro_tab
  with tab_agro:
      render_agro_tab(sb=get_supabase(), CURRENT_ROLE=..., CURRENT_NAME=...)
"""

import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
import io, re
from collections import defaultdict

# ══════════════════════════════════════════════════════════════
# CONSTANTES AGRONOMIQUES
# ══════════════════════════════════════════════════════════════
RENDEMENT_REF = {
    "CAP BON 1": 44.5, "CAP BON 2": 41.0, "NORD": 34.0,
    "KAIROUAN":  31.5, "BOUFICHA":  37.0,
    "GAFSA / KASSRINE": 28.5, "SIDI BOUZID": 27.0,
}
DENSITE_PLANTS = {
    "CAP BON 1": 25000, "CAP BON 2": 25000, "NORD": 22000,
    "KAIROUAN":  20000, "BOUFICHA":  22000,
    "GAFSA / KASSRINE": 18000, "SIDI BOUZID": 18000,
}
SCORE_SEUILS = [
    (1.15, "⭐ Excellent",   "#1E8449"),
    (1.05, "✅ Bon",          "#2E86C1"),
    (0.90, "🟡 Moyen",        "#D4AC0D"),
    (0.75, "🟠 Faible",       "#CA6F1E"),
    (0.00, "🔴 Sous-perf.",   "#C0392B"),
]
COMM_COLORS = {
    "FEDI": "#1A5276", "MAKKI BEN SALAH": "#1F7A1F",
    "KHALIL": "#7D3C98", "ACHREF AJLANI": "#C0392B", "JILANI OBAY": "#D4AC0D",
}

# ══════════════════════════════════════════════════════════════
# PARSERS — lecture des fichiers Excel source
# ══════════════════════════════════════════════════════════════

def _find_header_row(df_raw, keywords):
    """Cherche la ligne d'en-tête contenant les mots-clés."""
    for i, row in df_raw.iterrows():
        vals = [str(v).strip().lower() for v in row if pd.notna(v)]
        if any(any(kw in v for v in vals) for kw in keywords):
            return i
    return 0

def _norm_col(c):
    return str(c).strip().lower().replace("é","e").replace("è","e")\
           .replace("ê","e").replace("â","a").replace("ô","o")\
           .replace("î","i").replace("û","u").replace(" ","_")\
           .replace("(","").replace(")","").replace("/","_")

def parse_bourak_file(file_obj):
    """
    Parse le fichier Bourak (livraisons de plants).
    Retourne DataFrame avec colonnes normalisées +
    récap par (Commercial, Agriculteur, Variété).
    """
    try:
        raw = pd.read_excel(file_obj, sheet_name=0, header=None)
        hr = _find_header_row(raw, ["commercial","client","article","qte"])
        df = pd.read_excel(file_obj, sheet_name=0, header=hr)
        df.columns = [_norm_col(c) for c in df.columns]

        # Mapping colonnes flexibles
        MAP = {
            "date":         ["date_livraison","date","date_livr"],
            "commercial":   ["commercial"],
            "agriculteur":  ["client","agriculteur","nom"],
            "pour_compte":  ["pour_compte","pour compte"],
            "variete":      ["article","variete","variété"],
            "qte":          ["qte","quantite","quantité","plants"],
            "pu":           ["p_u","pu","prix_unitaire","prix"],
            "remise":       ["remise"],
            "total":        ["total","total_tnd","montant"],
            "unite":        ["unite","unité"],
            "transporteur": ["transporteur"],
            "destination":  ["destination","destiantion","destintation"],
        }
        rename = {}
        for target, candidates in MAP.items():
            for c in candidates:
                if c in df.columns and target not in rename.values():
                    rename[c] = target
                    break
        df = df.rename(columns=rename)

        # Nettoyage
        for c in ["qte","pu","remise","total"]:
            if c in df.columns:
                df[c] = pd.to_numeric(df[c], errors="coerce")
        if "qte" in df.columns:
            df = df[df["qte"].notna() & (df["qte"] > 0)]
        if "commercial" in df.columns:
            df["commercial"] = df["commercial"].astype(str).str.strip()
        if "agriculteur" in df.columns:
            df["agriculteur"] = df["agriculteur"].astype(str).str.strip()

        # Récap plan livré par (commercial, agriculteur, variété)
        grp_cols = [c for c in ["commercial","agriculteur","variete"] if c in df.columns]
        if grp_cols and "qte" in df.columns:
            recap = df.groupby(grp_cols).agg(
                plan_livre=("qte", "sum"),
                cout_plants=("total", "sum") if "total" in df.columns else ("qte","count"),
                nb_livraisons=("qte", "count"),
            ).reset_index()
        else:
            recap = df.copy()

        return df, recap, None
    except Exception as e:
        return None, None, str(e)


def parse_sotusfa_file(file_obj):
    """
    Parse le fichier Sotusfa (intrants : engrais, pesticides, irrigation).
    Retourne DataFrame détail + récap par (Commercial, Agriculteur, Famille).
    """
    try:
        raw = pd.read_excel(file_obj, sheet_name=0, header=None)
        hr = _find_header_row(raw, ["agriculteur","famille","article","commercial"])
        df = pd.read_excel(file_obj, sheet_name=0, header=hr)
        df.columns = [_norm_col(c) for c in df.columns]

        MAP = {
            "date":         ["date"],
            "societe":      ["societe","société","soc"],
            "commercial":   ["commerciale","commercial"],
            "agriculteur":  ["agriculteur"],
            "famille":      ["famille"],
            "article":      ["article"],
            "qte":          ["qte","quantite"],
            "prix_ht":      ["prix_ht"],
            "prix_ttc":     ["prix_un_ttc","prix_ttc","prix"],
            "total_ttc":    ["total_ttc","total","montant"],
            "campagne":     ["compagne","campagne"],
        }
        rename = {}
        for target, candidates in MAP.items():
            for c in candidates:
                if c in df.columns and target not in rename.values():
                    rename[c] = target
                    break
        df = df.rename(columns=rename)

        for c in ["qte","prix_ht","prix_ttc","total_ttc"]:
            if c in df.columns:
                df[c] = pd.to_numeric(df[c], errors="coerce")
        if "total_ttc" in df.columns:
            df = df[df["total_ttc"].notna() & (df["total_ttc"] > 0)]
        if "agriculteur" in df.columns:
            df = df[df["agriculteur"].astype(str).str.strip().str.upper()
                    .isin(["", "NAN", "TOTAL", "SOUS-TOTAL"]) == False]

        # Normaliser famille
        FAM_NORM = {
            "engrais": "Engrais", "engrais ": "Engrais",
            "fertilissant": "Fertilisant", "fertilisant": "Fertilisant",
            "fongicide": "Fongicide",
            "insecticide": "Insecticide",
            "irrigations ": "Irrigation", "irrigations": "Irrigation",
            "irrigations turk": "Irrigation",
            "herbicide": "Herbicide",
            "divers": "Divers",
            "materiel": "Matériel",
            "traitement": "Traitement",
        }
        if "famille" in df.columns:
            df["famille_norm"] = df["famille"].astype(str).str.strip().str.lower()\
                                  .map(FAM_NORM).fillna("Autre")

        # Récap par agriculteur + famille
        grp = [c for c in ["commercial","agriculteur","famille_norm"] if c in df.columns]
        if grp and "total_ttc" in df.columns:
            recap = df.groupby(grp)["total_ttc"].sum().reset_index()
            pivot = recap.pivot_table(
                index=[c for c in ["commercial","agriculteur"] if c in recap.columns],
                columns="famille_norm", values="total_ttc", aggfunc="sum", fill_value=0
            ).reset_index()
            pivot.columns = [_norm_col(str(c)) for c in pivot.columns]
            pivot["total_intrants"] = pivot.select_dtypes("number").sum(axis=1)
        else:
            pivot = pd.DataFrame()

        return df, pivot, None
    except Exception as e:
        return None, None, str(e)


def parse_royal_file(file_obj):
    """
    Parse le fichier Royal/Plan (Plan Livré / Extra / Plan Actif).
    Accepte le format template généré ou format libre.
    """
    try:
        raw = pd.read_excel(file_obj, sheet_name=0, header=None)
        hr = _find_header_row(raw, ["plan","extra","actif","agriculteur"])
        df = pd.read_excel(file_obj, sheet_name=0, header=hr)
        df.columns = [_norm_col(c) for c in df.columns]

        MAP = {
            "commercial":   ["commercial"],
            "agriculteur":  ["agriculteur","client","nom"],
            "region":       ["region","région"],
            "variete":      ["variete","variété","article"],
            "hectares":     ["hectares_ha","hectares","ha"],
            "plan_livre":   ["plan_livre_plants","plan_livre","plan_livr","plants_livres","qte"],
            "extra":        ["extra_pertes","extra","pertes","plants_perdus"],
            "plan_actif":   ["plan_actif","plan_actifs","plants_actifs"],
            "tonnage":      ["tonnage_recolte_t","tonnage","tonnage_t","recolte"],
        }
        rename = {}
        for target, candidates in MAP.items():
            for c in candidates:
                if c in df.columns and target not in rename.values():
                    rename[c] = target
                    break
        df = df.rename(columns=rename)

        num_cols = ["hectares","plan_livre","extra","plan_actif","tonnage"]
        for c in num_cols:
            if c in df.columns:
                df[c] = pd.to_numeric(df[c], errors="coerce")

        # Calcul automatique si plan_actif absent
        if "plan_actif" not in df.columns and "plan_livre" in df.columns:
            extra_col = df["extra"] if "extra" in df.columns else 0
            df["plan_actif"] = df["plan_livre"] - extra_col

        if "agriculteur" in df.columns:
            df = df[df["agriculteur"].astype(str).str.strip().str.upper()
                    .isin(["", "NAN", "TOTAL", "SOUS-TOTAL"]) == False]

        df = df.dropna(subset=["plan_actif"] if "plan_actif" in df.columns
                       else ["plan_livre"])
        return df, None
    except Exception as e:
        return None, str(e)


# ══════════════════════════════════════════════════════════════
# FUSION DES 3 SOURCES
# ══════════════════════════════════════════════════════════════

def merge_sources(df_royal, df_bourak_recap, df_sotusfa_pivot):
    """
    Joint les 3 sources sur (commercial + agriculteur).
    Calcule toutes les métriques dérivées.
    """
    base = None

    # Source principale : Royal (Plan Livré/Extra/Actif)
    if df_royal is not None and not df_royal.empty:
        base = df_royal.copy()
        key_cols = [c for c in ["commercial","agriculteur"] if c in base.columns]
        for c in key_cols:
            base[c] = base[c].astype(str).str.strip().str.upper()

    # Ajouter infos Bourak si pas de Royal
    if base is None and df_bourak_recap is not None and not df_bourak_recap.empty:
        base = df_bourak_recap.copy()
        if "plan_livre" not in base.columns and "qte" in base.columns:
            base = base.rename(columns={"qte": "plan_livre"})
        key_cols = [c for c in ["commercial","agriculteur"] if c in base.columns]
        for c in key_cols:
            base[c] = base[c].astype(str).str.strip().str.upper()

    if base is None:
        return pd.DataFrame()

    # Merge Bourak → ajouter plan_livre si vient de Royal
    if df_bourak_recap is not None and not df_bourak_recap.empty and df_royal is not None:
        b = df_bourak_recap.copy()
        for c in ["commercial","agriculteur"]:
            if c in b.columns:
                b[c] = b[c].astype(str).str.strip().str.upper()
        b_grp = b.groupby([c for c in ["commercial","agriculteur"] if c in b.columns])\
                  .agg(plan_livre_bourak=("plan_livre","sum"),
                       cout_plants=("cout_plants","sum") if "cout_plants" in b.columns
                                   else ("plan_livre","count")).reset_index()
        key = [c for c in ["commercial","agriculteur"] if c in base.columns and c in b_grp.columns]
        if key:
            base = base.merge(b_grp, on=key, how="left")
            if "plan_livre" not in base.columns and "plan_livre_bourak" in base.columns:
                base["plan_livre"] = base["plan_livre_bourak"]

    # Merge Sotusfa → ajouter intrants
    if df_sotusfa_pivot is not None and not df_sotusfa_pivot.empty:
        s = df_sotusfa_pivot.copy()
        for c in ["commercial","agriculteur"]:
            if c in s.columns:
                s[c] = s[c].astype(str).str.strip().str.upper()
        key = [c for c in ["commercial","agriculteur"] if c in base.columns and c in s.columns]
        if key:
            base = base.merge(s, on=key, how="left")

    # ── Calculs dérivés ─────────────────────────────────────
    df = base.copy()

    # Taux de prise
    if "plan_livre" in df.columns and "plan_actif" in df.columns:
        df["taux_prise_pct"] = np.where(
            df["plan_livre"] > 0,
            (df["plan_actif"] / df["plan_livre"] * 100).round(2),
            np.nan
        )

    # Extra si absent
    if "extra" not in df.columns and "plan_livre" in df.columns and "plan_actif" in df.columns:
        df["extra"] = df["plan_livre"] - df["plan_actif"]

    # Rendement t/ha
    if "tonnage" in df.columns and "hectares" in df.columns:
        df["rendement_t_ha"] = np.where(
            df["hectares"] > 0,
            (df["tonnage"] / df["hectares"]).round(3),
            np.nan
        )

    # kg tomate par plant actif
    if "tonnage" in df.columns and "plan_actif" in df.columns:
        df["kg_par_plant_actif"] = np.where(
            df["plan_actif"] > 0,
            (df["tonnage"] * 1000 / df["plan_actif"]).round(4),
            np.nan
        )

    # Intrants par plant actif (si disponibles)
    for intrant_col in ["engrais","fertilisant","fongicide","insecticide","irrigation"]:
        col_src = next((c for c in df.columns if intrant_col in c.lower()), None)
        if col_src and "plan_actif" in df.columns:
            df[f"{intrant_col}_par_plant"] = np.where(
                df["plan_actif"] > 0,
                (df[col_src] / df["plan_actif"]).round(6),
                np.nan
            )

    # Coût intrants par tonne
    if "total_intrants" in df.columns and "tonnage" in df.columns:
        df["cout_intrants_par_tonne"] = np.where(
            df["tonnage"] > 0,
            (df["total_intrants"] / df["tonnage"]).round(2),
            np.nan
        )

    # Coût plants par tonne
    cost_col = next((c for c in df.columns if "cout_plant" in c.lower()), None)
    if cost_col and "tonnage" in df.columns:
        df["cout_plants_par_tonne"] = np.where(
            df["tonnage"] > 0,
            (df[cost_col] / df["tonnage"]).round(2),
            np.nan
        )

    # Score de performance
    if "rendement_t_ha" in df.columns:
        def _get_score(row):
            r = row.get("rendement_t_ha")
            reg = str(row.get("region","")).strip().upper()
            if not r or pd.isna(r):
                return ("—", "#888888", 0.0)
            ref = RENDEMENT_REF.get(reg, 32.0)
            ratio = r / ref if ref > 0 else 0
            for seuil, label, color in SCORE_SEUILS:
                if ratio >= seuil:
                    return (label, color, round(ratio, 3))
            return ("🔴 Sous-perf.", "#C0392B", round(ratio, 3))

        scores = df.apply(_get_score, axis=1)
        df["score_label"] = scores.apply(lambda x: x[0])
        df["score_color"] = scores.apply(lambda x: x[1])
        df["score_ratio"] = scores.apply(lambda x: x[2])

    return df


# ══════════════════════════════════════════════════════════════
# TEMPLATE EXCEL À TÉLÉCHARGER
# ══════════════════════════════════════════════════════════════

def generate_template_royal():
    """Template Plan Livré / Extra / Plan Actif pour Royal/ingénieurs."""
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Plans Actifs"
    HDR = PatternFill("solid", start_color="1F3864", end_color="1F3864")
    ORG = PatternFill("solid", start_color="7B3F00", end_color="7B3F00")
    GRN = PatternFill("solid", start_color="1A5C2A", end_color="1A5C2A")
    BF  = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)
    BD  = Border(*[Side(style="thin", color="CCCCCC")] * 0,
                 left=Side(style="thin",color="CCCCCC"),
                 right=Side(style="thin",color="CCCCCC"),
                 top=Side(style="thin",color="CCCCCC"),
                 bottom=Side(style="thin",color="CCCCCC"))

    COLS = [
        ("Commercial",       14, HDR), ("Agriculteur",   26, HDR),
        ("Région",           16, HDR), ("Zone",          16, HDR),
        ("Variété",          14, HDR), ("Hectares (ha)", 12, GRN),
        ("Plan Livré (plants)", 16, GRN), ("Extra (pertes)", 14, ORG),
        ("Plan Actif",       14, GRN), ("Tonnage Récolté (t)", 16, GRN),
        ("Notes",            22, HDR),
    ]
    for ci, (h, w, fill) in enumerate(COLS, 1):
        c = ws.cell(1, ci, value=h)
        c.font = BF; c.fill = fill
        c.alignment = CTR; c.border = BD
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 36

    # Ligne exemple
    EX = ["KHALIL","NEJI ZAAFOURI","KAIROUAN","Zaafria","Savera",
          62, 1236444, 111000, "=G2-H2", 1820, "RM mécanique"]
    for ci, val in enumerate(EX, 1):
        c = ws.cell(2, ci, value=val)
        c.border = BD
        c.alignment = Alignment(horizontal="center", vertical="center")

    ws.cell(3,1).value = "← Remplir à partir de la ligne 2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def generate_template_saisie():
    """Template saisie manuelle hectares + tonnage (si pas de Royal)."""
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Saisie Hectares Tonnage"
    HDR = PatternFill("solid", start_color="4A235A", end_color="4A235A")
    BF = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)
    BD = Border(left=Side(style="thin",color="CCCCCC"),right=Side(style="thin",color="CCCCCC"),
                top=Side(style="thin",color="CCCCCC"),bottom=Side(style="thin",color="CCCCCC"))

    COLS = [("Commercial",14),("Agriculteur",26),("Région",16),
            ("Variété",14),("Hectares (ha)",12),("Tonnage Récolté (t)",16),("Notes",22)]
    for ci,(h,w) in enumerate(COLS,1):
        c = ws.cell(1,ci,value=h)
        c.font=BF; c.fill=HDR; c.alignment=CTR; c.border=BD
        ws.column_dimensions[get_column_letter(ci)].width=w
    ws.row_dimensions[1].height=30

    for ci,val in enumerate(["MAKKI BEN SALAH","KHALED BELHAJ","CAP BON 1","Savera",63,2800,""],1):
        c = ws.cell(2,ci,value=val); c.border=BD
        c.alignment=Alignment(horizontal="center",vertical="center")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ══════════════════════════════════════════════════════════════
# EXPORT EXCEL RÉSULTATS
# ══════════════════════════════════════════════════════════════

def export_resultats_excel(df):
    """Génère un Excel de résultats avec toutes les métriques calculées."""
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule

    wb = Workbook()
    ws = wb.active
    ws.title = "Résultats Agro"
    ws.sheet_view.showGridLines = False

    HDR = PatternFill("solid", start_color="0D3349", end_color="0D3349")
    BF  = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)
    BD  = Border(left=Side(style="thin",color="CCCCCC"),
                 right=Side(style="thin",color="CCCCCC"),
                 top=Side(style="thin",color="CCCCCC"),
                 bottom=Side(style="thin",color="CCCCCC"))

    EXPORT_COLS = [
        ("commercial","Commercial",14), ("agriculteur","Agriculteur",26),
        ("region","Région",16), ("variete","Variété",14),
        ("hectares","Hectares",10),
        ("plan_livre","Plan Livré",14), ("extra","Extra",12),
        ("plan_actif","Plan Actif",14), ("taux_prise_pct","Taux Prise %",12),
        ("tonnage","Tonnage (t)",13), ("rendement_t_ha","Rend. t/ha",12),
        ("kg_par_plant_actif","kg/plant actif",13),
        ("total_intrants","Total Intrants TND",16),
        ("cout_intrants_par_tonne","Coût/tonne TND",14),
        ("score_label","Score",14),
    ]
    avail = [(col, lbl, w) for col, lbl, w in EXPORT_COLS if col in df.columns]

    for ci, (_, lbl, w) in enumerate(avail, 1):
        c = ws.cell(1, ci, value=lbl)
        c.font=BF; c.fill=HDR; c.alignment=CTR; c.border=BD
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 34

    for ri, (_, row) in enumerate(df.iterrows()):
        r = ri + 2
        fill = PatternFill("solid", start_color="F0F5FF" if ri%2==0 else "FFFFFF",
                           end_color="F0F5FF" if ri%2==0 else "FFFFFF")
        for ci, (col, _, _) in enumerate(avail, 1):
            val = row.get(col, "")
            if pd.isna(val): val = ""
            c = ws.cell(r, ci, value=val)
            c.border = BD; c.alignment = CTR; c.fill = fill
            if col in ("plan_livre","extra","plan_actif"):
                c.number_format = "#,##0"
            elif col in ("taux_prise_pct",):
                c.number_format = "0.0"
            elif col in ("tonnage","rendement_t_ha","kg_par_plant_actif"):
                c.number_format = "0.000"
            elif col in ("total_intrants","cout_intrants_par_tonne"):
                c.number_format = "#,##0.00"

    last = len(df) + 2
    rend_col = next((ci+1 for ci,(col,_,_) in enumerate(avail) if col=="rendement_t_ha"), None)
    taux_col = next((ci+1 for ci,(col,_,_) in enumerate(avail) if col=="taux_prise_pct"), None)
    if rend_col:
        ws.conditional_formatting.add(
            f"{get_column_letter(rend_col)}2:{get_column_letter(rend_col)}{last}",
            ColorScaleRule(start_type="min",start_color="FFC7CE",
                           mid_type="percentile",mid_value=50,mid_color="FFEB9C",
                           end_type="max",end_color="C6EFCE"))
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ══════════════════════════════════════════════════════════════
# SQL SUPABASE
# ══════════════════════════════════════════════════════════════
SQL_CREATE = """
-- ============================================================
-- Tables Supabase pour le module Performance Agronomique
-- Exécuter une seule fois dans l'éditeur SQL Supabase
-- ============================================================

-- Table principale
CREATE TABLE IF NOT EXISTS agri_performance (
    id              BIGSERIAL PRIMARY KEY,
    commercial      TEXT NOT NULL,
    agriculteur     TEXT NOT NULL,
    region          TEXT,
    zone            TEXT,
    variete         TEXT,
    hectares        NUMERIC,
    plan_livre      NUMERIC,   -- plants livrés par Royal/Bourak
    extra           NUMERIC,   -- plants perdus (saisie ingénieur)
    plan_actif      NUMERIC,   -- plan_livre - extra
    tonnage_t       NUMERIC,   -- tonnage récolté (t)
    saison          TEXT DEFAULT '2026',
    -- Intrants Sotusfa
    engrais_tnd     NUMERIC,
    fertilisant_tnd NUMERIC,
    fongicide_tnd   NUMERIC,
    insecticide_tnd NUMERIC,
    irrigation_tnd  NUMERIC,
    divers_tnd      NUMERIC,
    cout_plants_tnd NUMERIC,   -- coût plants Bourak
    notes           TEXT,
    created_at      TIMESTAMPTZ DEFAULT NOW()
);

CREATE INDEX IF NOT EXISTS idx_agri_perf_comm ON agri_performance(commercial);
CREATE INDEX IF NOT EXISTS idx_agri_perf_reg  ON agri_performance(region);
CREATE INDEX IF NOT EXISTS idx_agri_perf_sais ON agri_performance(saison);

-- Vue avec tous les calculs automatiques
CREATE OR REPLACE VIEW agri_performance_calcule AS
SELECT *,
  CASE WHEN plan_livre > 0
       THEN ROUND(plan_actif::NUMERIC / plan_livre * 100, 2) END  AS taux_prise_pct,
  CASE WHEN plan_actif > 0
       THEN ROUND(tonnage_t * 1000 / plan_actif, 4) END           AS kg_par_plant_actif,
  CASE WHEN hectares  > 0
       THEN ROUND(tonnage_t / hectares, 3) END                     AS rendement_t_ha,
  COALESCE(engrais_tnd,0)+COALESCE(fertilisant_tnd,0)
    +COALESCE(fongicide_tnd,0)+COALESCE(insecticide_tnd,0)
    +COALESCE(irrigation_tnd,0)+COALESCE(divers_tnd,0)            AS total_intrants_tnd,
  CASE WHEN tonnage_t > 0 THEN ROUND(
    (COALESCE(engrais_tnd,0)+COALESCE(fertilisant_tnd,0)
     +COALESCE(fongicide_tnd,0)+COALESCE(insecticide_tnd,0)
     +COALESCE(irrigation_tnd,0)+COALESCE(divers_tnd,0))
    / tonnage_t, 2) END                                            AS cout_intrants_par_tonne
FROM agri_performance;
"""


def save_supabase(sb, df):
    """Sauvegarde le DataFrame fusionné dans agri_performance."""
    if sb is None or df is None or df.empty:
        return False, "Supabase non disponible"
    try:
        COLS_MAP = {
            "commercial":"commercial","agriculteur":"agriculteur",
            "region":"region","zone":"zone","variete":"variete",
            "hectares":"hectares","plan_livre":"plan_livre","extra":"extra",
            "plan_actif":"plan_actif","tonnage":"tonnage_t",
            "engrais":"engrais_tnd","fertilisant":"fertilisant_tnd",
            "fongicide":"fongicide_tnd","insecticide":"insecticide_tnd",
            "irrigation":"irrigation_tnd","divers":"divers_tnd",
            "cout_plants":"cout_plants_tnd","notes":"notes",
        }
        comms = df["commercial"].dropna().unique().tolist() if "commercial" in df.columns else []
        for comm in comms:
            sb.table("agri_performance").delete()\
              .eq("commercial", comm).eq("saison","2026").execute()

        rows = []
        for _, row in df.iterrows():
            rec = {"saison": "2026"}
            for src, tgt in COLS_MAP.items():
                col = next((c for c in df.columns if src in c.lower()), None)
                if col:
                    val = row.get(col)
                    rec[tgt] = None if pd.isna(val) else val
            rows.append(rec)

        for i in range(0, len(rows), 500):
            sb.table("agri_performance").insert(rows[i:i+500]).execute()
        return True, f"{len(rows)} agriculteurs sauvegardés"
    except Exception as e:
        return False, str(e)


def load_supabase(sb):
    """Charge et recalcule les données depuis Supabase."""
    if sb is None:
        return pd.DataFrame()
    try:
        data = sb.table("agri_performance").select("*")\
                 .eq("saison","2026").execute().data
        if not data:
            return pd.DataFrame()
        df = pd.DataFrame(data)
        renames = {"tonnage_t":"tonnage","engrais_tnd":"engrais",
                   "fertilisant_tnd":"fertilisant","fongicide_tnd":"fongicide",
                   "insecticide_tnd":"insecticide","irrigation_tnd":"irrigation",
                   "divers_tnd":"divers","cout_plants_tnd":"cout_plants"}
        df = df.rename(columns=renames)
        intrant_cols = ["engrais","fertilisant","fongicide","insecticide","irrigation","divers"]
        existing = [c for c in intrant_cols if c in df.columns]
        if existing:
            df["total_intrants"] = df[existing].fillna(0).sum(axis=1)
        return merge_sources(df, None, None)
    except Exception:
        return pd.DataFrame()


# ══════════════════════════════════════════════════════════════
# RENDER PRINCIPAL
# ══════════════════════════════════════════════════════════════

def render_agro_tab(sb=None, planning_df=None,
                    CURRENT_ROLE="directeur", CURRENT_NAME=""):

    st.markdown("""
<div style='background:#0a1a0a;border:1px solid #1E8449;border-radius:12px;
padding:16px 20px;margin-bottom:18px'>
  <div style='font-size:1.05rem;font-weight:700;color:#f0f6fc;margin-bottom:5px'>
    🌱 Performance Agronomique — Tomate 2026
  </div>
  <div style='font-size:.82rem;color:#8b949e;line-height:1.6'>
    <b style='color:#4CAF50'>Royal / Bourak</b> → Plan Livré / Extra / Plan Actif &nbsp;|&nbsp;
    <b style='color:#2196F3'>Sotusfa</b> → Engrais / Pesticides / Irrigation &nbsp;|&nbsp;
    <b style='color:#9C27B0'>Calculs</b> : kg/plant actif · Rendement t/ha · Coût/tonne
  </div>
</div>""", unsafe_allow_html=True)

    # ── Session state ──────────────────────────────────────
    for key in ["agro_royal","agro_bourak_raw","agro_bourak_recap",
                "agro_sotusfa_raw","agro_sotusfa_pivot","agro_merged",
                "agro_saisie"]:
        if key not in st.session_state:
            st.session_state[key] = None

    # Charger Supabase au démarrage
    if (st.session_state["agro_merged"] is None and sb is not None):
        df_sb = load_supabase(sb)
        if not df_sb.empty:
            st.session_state["agro_merged"] = df_sb

    # ── Tabs ───────────────────────────────────────────────
    t1, t2, t3, t4, t5, t6 = st.tabs([
        "📥 Import fichiers",
        "🌱 Plan Actif & Taux prise",
        "💊 Efficacité Intrants",
        "🏆 Classements",
        "🔬 Analyse par variété",
        "⚙️  SQL / Export",
    ])

    # ══════════════════════════════════════════
    # TAB 1 — IMPORT
    # ══════════════════════════════════════════
    with t1:
        st.markdown("### Import des 3 sources de données")

        c1, c2, c3 = st.columns(3)

        # ── Bourak ──────────────────────────────────────────
        with c1:
            st.markdown("""<div style='background:#1a1000;border:1px solid #8B3A00;
border-radius:8px;padding:10px 14px;margin-bottom:10px'>
<b style='color:#FF9800'>🚛 BOURAK</b><br>
<span style='font-size:.8rem;color:#aaa'>
Livraisons plants (Plan Livré)<br>
Colonnes : Commercial · Client · Article (variété) · Qte · P.U · Total
</span></div>""", unsafe_allow_html=True)
            f_bourak = st.file_uploader("Fichier Bourak", type=["xlsx","xls"],
                                         key="up_bourak", label_visibility="collapsed")
            st.download_button("📥 Voir format attendu",
                data=b"Commercial,Agriculteur,Variete,Qte,PU,Total\nKHALIL,NEJI ZAAFOURI,Savera,1236444,0.065,80369",
                file_name="format_bourak.csv", mime="text/csv",
                use_container_width=True)
            if f_bourak:
                df_raw, df_recap, err = parse_bourak_file(f_bourak)
                if err:
                    st.error(f"Erreur Bourak: {err}")
                else:
                    st.session_state["agro_bourak_raw"]   = df_raw
                    st.session_state["agro_bourak_recap"] = df_recap
                    n = df_recap["plan_livre"].sum() if df_recap is not None and "plan_livre" in df_recap.columns else 0
                    st.success(f"✅ {len(df_recap)} agriculteurs · {int(n):,} plants")
            if st.session_state["agro_bourak_recap"] is not None:
                st.caption(f"Chargé : {len(st.session_state['agro_bourak_recap'])} lignes")

        # ── Sotusfa ─────────────────────────────────────────
        with c2:
            st.markdown("""<div style='background:#001a00;border:1px solid #1A5C2A;
border-radius:8px;padding:10px 14px;margin-bottom:10px'>
<b style='color:#4CAF50'>🌿 SOTUSFA</b><br>
<span style='font-size:.8rem;color:#aaa'>
Intrants (engrais, pesticides, irrigation)<br>
Colonnes : Agriculteur · Famille · Article · Qte · Total TTC
</span></div>""", unsafe_allow_html=True)
            f_sotusfa = st.file_uploader("Fichier Sotusfa", type=["xlsx","xls"],
                                          key="up_sotusfa", label_visibility="collapsed")
            if f_sotusfa:
                df_raw_s, df_piv, err = parse_sotusfa_file(f_sotusfa)
                if err:
                    st.error(f"Erreur Sotusfa: {err}")
                else:
                    st.session_state["agro_sotusfa_raw"]   = df_raw_s
                    st.session_state["agro_sotusfa_pivot"] = df_piv
                    tot = df_piv["total_intrants"].sum() if df_piv is not None and "total_intrants" in df_piv.columns else 0
                    st.success(f"✅ {len(df_piv) if df_piv is not None else 0} agriculteurs · {tot:,.0f} TND")
            if st.session_state["agro_sotusfa_pivot"] is not None:
                st.caption(f"Chargé : {len(st.session_state['agro_sotusfa_pivot'])} lignes")

        # ── Royal ────────────────────────────────────────────
        with c3:
            st.markdown("""<div style='background:#0a001a;border:1px solid #4A1A6B;
border-radius:8px;padding:10px 14px;margin-bottom:10px'>
<b style='color:#9C27B0'>🌾 ROYAL</b><br>
<span style='font-size:.8rem;color:#aaa'>
Plan Livré / Extra / Plan Actif + Hectares + Tonnage<br>
Remplir le template ci-dessous
</span></div>""", unsafe_allow_html=True)
            st.download_button("📥 Télécharger template Royal",
                data=generate_template_royal(),
                file_name="template_royal_plan_actif_2026.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True, type="primary")
            f_royal = st.file_uploader("Fichier Royal", type=["xlsx","xls"],
                                        key="up_royal", label_visibility="collapsed")
            if f_royal:
                df_r, err = parse_royal_file(f_royal)
                if err:
                    st.error(f"Erreur Royal: {err}")
                else:
                    st.session_state["agro_royal"] = df_r
                    st.success(f"✅ {len(df_r)} agriculteurs chargés")
            if st.session_state["agro_royal"] is not None:
                st.caption(f"Chargé : {len(st.session_state['agro_royal'])} lignes")

        st.divider()

        # ── Saisie manuelle (hectares + tonnage si pas de Royal) ─
        with st.expander("✏️ Saisie manuelle Hectares + Tonnage (si pas de fichier Royal)"):
            st.download_button("📥 Template saisie manuelle",
                data=generate_template_saisie(),
                file_name="saisie_hectares_tonnage.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            f_saisie = st.file_uploader("Fichier saisie", type=["xlsx","xls"],
                                         key="up_saisie", label_visibility="collapsed")
            if f_saisie:
                try:
                    df_s = pd.read_excel(f_saisie, header=0)
                    df_s.columns = [_norm_col(c) for c in df_s.columns]
                    st.session_state["agro_saisie"] = df_s
                    st.success(f"✅ {len(df_s)} lignes chargées")
                except Exception as e:
                    st.error(str(e))

        # ── Bouton FUSIONNER ─────────────────────────────────
        st.markdown("---")
        if st.button("🔗 Fusionner les 3 sources et calculer",
                     type="primary", use_container_width=True):
            royal_src = (st.session_state["agro_royal"] or
                         st.session_state.get("agro_saisie"))
            df_merged = merge_sources(
                royal_src,
                st.session_state["agro_bourak_recap"],
                st.session_state["agro_sotusfa_pivot"],
            )
            if df_merged.empty:
                st.warning("Aucune donnée à fusionner — importez au moins un fichier.")
            else:
                st.session_state["agro_merged"] = df_merged
                ok, msg = save_supabase(sb, df_merged)
                st.success(f"✅ {len(df_merged)} agriculteurs fusionnés · Supabase: {msg}")

        # Aperçu
        df_m = st.session_state.get("agro_merged")
        if df_m is not None and not df_m.empty:
            st.markdown(f"**Aperçu données fusionnées ({len(df_m)} agriculteurs) :**")
            preview_cols = [c for c in ["commercial","agriculteur","region","variete",
                                         "plan_livre","extra","plan_actif","taux_prise_pct",
                                         "tonnage","rendement_t_ha","kg_par_plant_actif"]
                            if c in df_m.columns]
            st.dataframe(df_m[preview_cols].head(10),
                         use_container_width=True, hide_index=True)

    # ══════════════════════════════════════════
    # TAB 2 — PLAN ACTIF & TAUX DE PRISE
    # ══════════════════════════════════════════
    with t2:
        df = st.session_state.get("agro_merged")
        if df is None or df.empty:
            st.info("📥 Importez vos fichiers dans l'onglet 'Import' puis cliquez 'Fusionner'.")
        else:
            # Filtres
            fc1, fc2 = st.columns(2)
            comms = ["Tous"] + sorted(df["commercial"].dropna().unique().tolist()) \
                    if "commercial" in df.columns else ["Tous"]
            sel_c = fc1.selectbox("Commercial", comms, key="t2_comm")
            regs  = ["Toutes"] + sorted(df["region"].dropna().unique().tolist()) \
                    if "region" in df.columns else ["Toutes"]
            sel_r = fc2.selectbox("Région", regs, key="t2_reg")

            df_f = df.copy()
            if sel_c != "Tous" and "commercial" in df_f.columns:
                df_f = df_f[df_f["commercial"] == sel_c]
            if sel_r != "Toutes" and "region" in df_f.columns:
                df_f = df_f[df_f["region"] == sel_r]

            # KPIs
            k1,k2,k3,k4,k5 = st.columns(5)
            if "plan_livre" in df_f.columns:
                k1.metric("Plan Livré", f"{int(df_f['plan_livre'].sum()):,} plants")
            if "extra" in df_f.columns:
                k2.metric("Extra (pertes)", f"{int(df_f['extra'].sum()):,} plants",
                          delta=f"{df_f['extra'].sum()/df_f['plan_livre'].sum()*100:.1f}% perte"
                          if "plan_livre" in df_f.columns else "")
            if "plan_actif" in df_f.columns:
                k3.metric("Plan Actif", f"{int(df_f['plan_actif'].sum()):,} plants")
            if "taux_prise_pct" in df_f.columns:
                k4.metric("Taux de prise moyen",
                          f"{df_f['taux_prise_pct'].mean():.1f}%")
            if "tonnage" in df_f.columns:
                k5.metric("Tonnage total", f"{df_f['tonnage'].sum():,.1f} t")

            # Graphique Plan Livré vs Plan Actif vs Extra
            if all(c in df_f.columns for c in ["plan_livre","plan_actif","extra"]):
                st.markdown("#### 🌱 Plan Livré / Extra / Plan Actif par agriculteur")
                df_bars = df_f.dropna(subset=["plan_livre"]).sort_values("plan_livre", ascending=True)
                agri_col = "agriculteur" if "agriculteur" in df_bars.columns else df_bars.columns[0]
                fig = go.Figure()
                fig.add_trace(go.Bar(name="Plan Actif",
                    y=df_bars[agri_col], x=df_bars["plan_actif"],
                    orientation="h", marker_color="#1E8449"))
                fig.add_trace(go.Bar(name="Extra (pertes)",
                    y=df_bars[agri_col], x=df_bars["extra"],
                    orientation="h", marker_color="#E53935"))
                fig.update_layout(barmode="stack",
                    template="plotly_dark", paper_bgcolor="#161b22",
                    plot_bgcolor="#0d1117", height=max(350, len(df_bars)*28+80),
                    xaxis_title="Nombre de plants", yaxis_title="",
                    legend=dict(orientation="h", yanchor="bottom", y=1.02),
                    margin=dict(l=220,r=60,t=40,b=40))
                st.plotly_chart(fig, use_container_width=True)

            # Taux de prise par région
            if "taux_prise_pct" in df_f.columns and "region" in df_f.columns:
                st.markdown("#### 📊 Taux de prise moyen par région")
                tp_reg = df_f.groupby("region")["taux_prise_pct"].mean().reset_index()
                tp_reg.columns = ["Région","Taux prise (%)"]
                tp_reg = tp_reg.sort_values("Taux prise (%)", ascending=False)
                fig2 = px.bar(tp_reg, x="Région", y="Taux prise (%)",
                              color="Taux prise (%)",
                              color_continuous_scale=["#E53935","#FF9800","#1E8449"],
                              template="plotly_dark", text_auto=".1f",
                              title="Taux de prise moyen par région")
                fig2.update_layout(paper_bgcolor="#161b22", height=320)
                fig2.add_hline(y=90, line_dash="dash", line_color="#f5a623",
                               annotation_text="Seuil OK 90%")
                st.plotly_chart(fig2, use_container_width=True)

            # Tableau détaillé
            st.markdown("#### 📋 Tableau détaillé")
            tab_cols = [c for c in ["commercial","agriculteur","region","variete",
                                     "plan_livre","extra","plan_actif",
                                     "taux_prise_pct","tonnage","rendement_t_ha",
                                     "kg_par_plant_actif","score_label"]
                        if c in df_f.columns]
            df_disp = df_f[tab_cols].copy()
            df_disp.columns = [c.replace("_"," ").title() for c in df_disp.columns]
            st.dataframe(df_disp.sort_values("Plan Actif" if "Plan Actif" in df_disp.columns
                                             else df_disp.columns[0], ascending=False),
                         use_container_width=True, hide_index=True, height=380)

    # ══════════════════════════════════════════
    # TAB 3 — EFFICACITÉ INTRANTS
    # ══════════════════════════════════════════
    with t3:
        df = st.session_state.get("agro_merged")
        if df is None or df.empty:
            st.info("📥 Importez vos fichiers dans l'onglet 'Import'.")
        else:
            intrant_cols = [c for c in df.columns if any(
                x in c.lower() for x in ["engrais","fertilisant","fongicide",
                                          "insecticide","irrigation","total_intrant"])]
            if not intrant_cols:
                st.warning("Pas de données Sotusfa — importez le fichier intrants.")
            else:
                st.markdown("### 💊 Efficacité des intrants par agriculteur")

                sel_comm3 = st.selectbox("Commercial",
                    ["Tous"] + sorted(df["commercial"].dropna().unique().tolist())
                    if "commercial" in df.columns else ["Tous"], key="t3_comm")
                df_f3 = df[df["commercial"]==sel_comm3].copy() \
                        if sel_comm3 != "Tous" else df.copy()

                # KPIs intrants
                ks = st.columns(4)
                if "total_intrants" in df_f3.columns:
                    ks[0].metric("Total intrants",
                                 f"{df_f3['total_intrants'].sum():,.0f} TND")
                if "cout_intrants_par_tonne" in df_f3.columns:
                    ks[1].metric("Coût moy/tonne",
                                 f"{df_f3['cout_intrants_par_tonne'].mean():,.0f} TND/t")
                if "engrais" in df_f3.columns and "plan_actif" in df_f3.columns:
                    eng_plant = (df_f3["engrais"].sum() /
                                 df_f3["plan_actif"].sum() * 1000) if df_f3["plan_actif"].sum() > 0 else 0
                    ks[2].metric("Engrais/1000 plants", f"{eng_plant:.1f} TND")
                if "tonnage" in df_f3.columns:
                    ks[3].metric("Tonnage total", f"{df_f3['tonnage'].sum():,.1f} t")

                # Graphique empilé intrants par agriculteur
                fam_cols = [c for c in df_f3.columns if c in
                            ["engrais","fertilisant","fongicide",
                             "insecticide","irrigation","divers"]]
                if fam_cols and "agriculteur" in df_f3.columns:
                    df_melt = df_f3[["agriculteur"] + fam_cols].dropna(
                        subset=fam_cols, how="all")
                    df_melt = df_melt.melt(id_vars="agriculteur",
                                           value_vars=fam_cols,
                                           var_name="Intrant", value_name="TND")
                    df_melt["TND"] = df_melt["TND"].fillna(0)
                    fig_int = px.bar(df_melt, x="agriculteur", y="TND",
                                     color="Intrant", barmode="stack",
                                     template="plotly_dark",
                                     title="Dépenses intrants par agriculteur (TND)",
                                     color_discrete_sequence=px.colors.qualitative.Set2)
                    fig_int.update_layout(paper_bgcolor="#161b22",
                                          plot_bgcolor="#0d1117", height=420,
                                          xaxis_tickangle=-35)
                    st.plotly_chart(fig_int, use_container_width=True)

                # Scatter Coût/tonne vs Rendement
                if ("cout_intrants_par_tonne" in df_f3.columns and
                    "rendement_t_ha" in df_f3.columns):
                    st.markdown("#### 💡 Coût intrants/tonne vs Rendement")
                    st.caption("Idéal = en bas à droite (faible coût, haut rendement)")
                    fig_sc = px.scatter(
                        df_f3.dropna(subset=["cout_intrants_par_tonne","rendement_t_ha"]),
                        x="cout_intrants_par_tonne", y="rendement_t_ha",
                        color="commercial" if "commercial" in df_f3.columns else None,
                        size="tonnage" if "tonnage" in df_f3.columns else None,
                        hover_data=[c for c in ["agriculteur","region","variete"]
                                    if c in df_f3.columns],
                        labels={"cout_intrants_par_tonne":"Coût intrants (TND/t)",
                                "rendement_t_ha":"Rendement (t/ha)"},
                        color_discrete_map=COMM_COLORS,
                        template="plotly_dark")
                    fig_sc.update_layout(paper_bgcolor="#161b22",
                                         plot_bgcolor="#0d1117", height=400)
                    st.plotly_chart(fig_sc, use_container_width=True)

    # ══════════════════════════════════════════
    # TAB 4 — CLASSEMENTS
    # ══════════════════════════════════════════
    with t4:
        df = st.session_state.get("agro_merged")
        if df is None or df.empty:
            st.info("📥 Importez vos fichiers dans l'onglet 'Import'.")
        else:
            st.markdown("### 🏆 Classement des commerciaux")
            if "commercial" in df.columns and "rendement_t_ha" in df.columns:
                rows_r = []
                for comm in sorted(df["commercial"].dropna().unique()):
                    sub = df[df["commercial"]==comm]
                    tot_t    = sub["tonnage"].sum() if "tonnage" in sub.columns else 0
                    avg_rend = sub["rendement_t_ha"].mean() \
                               if "rendement_t_ha" in sub.columns else 0
                    avg_tp   = sub["taux_prise_pct"].mean() \
                               if "taux_prise_pct" in sub.columns else 0
                    cout_t   = sub["cout_intrants_par_tonne"].mean() \
                               if "cout_intrants_par_tonne" in sub.columns else 0
                    kg_plant = sub["kg_par_plant_actif"].mean() \
                               if "kg_par_plant_actif" in sub.columns else 0
                    exc_pct  = (len(sub[sub.get("score_ratio",
                                pd.Series([0]*len(sub))) >= 1.05]) /
                               max(len(sub),1) * 100) if "score_ratio" in sub.columns else 0
                    rows_r.append({
                        "Commercial":     comm,
                        "Nb agriculteurs":len(sub),
                        "Tonnage (t)":    round(tot_t,0),
                        "Rend. moy. t/ha":round(avg_rend,2),
                        "Taux prise %":   round(avg_tp,1),
                        "Coût/tonne TND": round(cout_t,0),
                        "kg/plant actif": round(kg_plant,4),
                        "% Excellents":   round(exc_pct,1),
                    })
                df_rank = pd.DataFrame(rows_r)\
                          .sort_values("Rend. moy. t/ha", ascending=False)\
                          .reset_index(drop=True)
                df_rank.index += 1
                medals = ["🥇","🥈","🥉"] + [""]*10
                df_rank.insert(0,"",medals[:len(df_rank)])

                st.dataframe(df_rank, use_container_width=True,
                             column_config={
                                 "Rend. moy. t/ha": st.column_config.ProgressColumn(
                                     "Rend. moy. t/ha", min_value=0, max_value=50,format="%.2f"),
                                 "Taux prise %": st.column_config.ProgressColumn(
                                     "Taux prise %", min_value=80, max_value=100,format="%.1f%%"),
                             })

                # Radar chart
                fig_r = go.Figure()
                for comm in df_rank["Commercial"].tolist():
                    row_r = df_rank[df_rank["Commercial"]==comm].iloc[0]
                    fig_r.add_trace(go.Scatterpolar(
                        r=[min(row_r["Rend. moy. t/ha"]/50,1),
                           min(row_r["Taux prise %"]/100,1),
                           min(row_r["% Excellents"]/100,1),
                           1-min(row_r["Coût/tonne TND"]/2000,1)],
                        theta=["Rendement","Taux prise","% Excellents","Efficacité coût"],
                        fill="toself", name=comm,
                        line_color=COMM_COLORS.get(comm,"#888")))
                fig_r.update_layout(polar=dict(bgcolor="#161b22"),
                    template="plotly_dark", paper_bgcolor="#161b22", height=420)
                st.plotly_chart(fig_r, use_container_width=True)

    # ══════════════════════════════════════════
    # TAB 5 — ANALYSE PAR VARIÉTÉ
    # ══════════════════════════════════════════
    with t5:
        df = st.session_state.get("agro_merged")
        if df is None or df.empty:
            st.info("📥 Importez vos fichiers dans l'onglet 'Import'.")
        elif "variete" not in df.columns:
            st.warning("Colonne 'variété' absente — ajoutez-la dans le fichier Royal.")
        else:
            st.markdown("### 🔬 Analyse par variété et région")

            sel_reg5 = st.selectbox("Région",
                ["Toutes"] + sorted(df["region"].dropna().unique().tolist())
                if "region" in df.columns else ["Toutes"], key="t5_reg")
            df_f5 = df[df["region"]==sel_reg5].copy() if sel_reg5 != "Toutes" else df.copy()
            df_f5 = df_f5.dropna(subset=["variete"])

            if "rendement_t_ha" in df_f5.columns:
                var_stats = df_f5.groupby("variete").agg(
                    Nb_agri=("variete","count"),
                    Rend_moy=("rendement_t_ha","mean"),
                    Taux_prise=("taux_prise_pct","mean") if "taux_prise_pct" in df_f5.columns else ("variete","count"),
                    Tonnage=("tonnage","sum") if "tonnage" in df_f5.columns else ("variete","count"),
                    Plan_actif=("plan_actif","sum") if "plan_actif" in df_f5.columns else ("variete","count"),
                ).reset_index()
                var_stats.columns = ["Variété","Nb agri.","Rend. moy (t/ha)",
                                     "Taux prise %","Tonnage (t)","Plan Actif"]
                var_stats = var_stats.sort_values("Rend. moy (t/ha)", ascending=False)

                fig_v = px.bar(var_stats, x="Variété", y="Rend. moy (t/ha)",
                               color="Rend. moy (t/ha)",
                               color_continuous_scale=["#E53935","#FF9800","#1E8449"],
                               text_auto=".1f", template="plotly_dark",
                               title=f"Rendement moyen par variété — {sel_reg5}")
                fig_v.update_layout(paper_bgcolor="#161b22",
                                    plot_bgcolor="#0d1117", height=360)
                st.plotly_chart(fig_v, use_container_width=True)
                st.dataframe(var_stats, use_container_width=True, hide_index=True)

                st.markdown("---")
                st.markdown("#### 💡 Recommandation variété × région")
                if not var_stats.empty:
                    best = var_stats.iloc[0]
                    st.success(f"✅ **Meilleure variété** pour {sel_reg5} : "
                               f"**{best['Variété']}** avec {best['Rend. moy (t/ha)']:.1f} t/ha "
                               f"({best['Nb agri.']} agriculteurs)")

    # ══════════════════════════════════════════
    # TAB 6 — SQL / EXPORT
    # ══════════════════════════════════════════
    with t6:
        st.markdown("### ⚙️ SQL Supabase — à exécuter une seule fois")
        st.code(SQL_CREATE, language="sql")

        df = st.session_state.get("agro_merged")
        if df is not None and not df.empty:
            st.markdown("---")
            st.markdown("### 📤 Exporter les résultats calculés")
            st.download_button(
                "📥 Exporter tout (Excel formaté)",
                data=export_resultats_excel(df),
                file_name="agro_performance_resultats_2026.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary", use_container_width=True)

            col_e1, col_e2 = st.columns(2)
            with col_e1:
                csv = df.to_csv(index=False, encoding="utf-8-sig").encode("utf-8-sig")
                st.download_button("⬇️ CSV brut",data=csv,
                    file_name="agro_performance_2026.csv", mime="text/csv",
                    use_container_width=True)
            with col_e2:
                if st.button("🔄 Recharger depuis Supabase", use_container_width=True):
                    df_sb = load_supabase(sb)
                    if not df_sb.empty:
                        st.session_state["agro_merged"] = df_sb
                        st.success(f"✅ {len(df_sb)} agriculteurs rechargés")
                    else:
                        st.warning("Aucune donnée Supabase — importez d'abord.")