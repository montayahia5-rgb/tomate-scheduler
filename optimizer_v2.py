# -*- coding: utf-8 -*-
"""
optimizer_v2.py — OR-Tools avec optimisation distance + caps
=============================================================
Nouvelles fonctionnalités vs optimizer.py :
  - Matrice de distances zones → usines (km)
  - SICAM = 70% Borj Sedeya + 30% Mejez el Bab (pondéré)
  - OR-Tools minimise : distance totale de transport + déviation du plan
  - Assignation automatique agriculteur → usine la plus proche
  - Affichage par région dans le résultat

USAGE :
  python optimizer_v2.py
  python migrate.py
  streamlit run dashboard_phase10.py
"""

import sys, io, os, datetime, json, math
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

import pandas as pd
from collections import defaultdict
from ortools.sat.python import cp_model
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# CONFIGURATION
# ============================================================
OUTPUT_FILE = "Planning_Tomate_2026.xlsx"   # new clean name
TEMP_FILE   = "Planning_Tomate_TEMP.xlsx"

PEAK_START   = datetime.date(2026, 7,  1)
PEAK_END     = datetime.date(2026, 7, 15)
SEASON_START = datetime.date(2026, 6, 15)
SEASON_END   = datetime.date(2026, 8, 31)

def clamp_date(d):
    if d < SEASON_START: return SEASON_START
    if d > SEASON_END:   return SEASON_END
    return d

# Caps NORMAUX (1 livraison/jour) = LIMITE PENDANT LE PIC
COMMERCIAL_CAPS = {
    "FEDI":             850,   # inchangé
    "MAKKI BEN SALAH":  850,   # ajusté de 800
    "KHALIL":           900,   # ajusté de 800
    "ACHREF AJLANI":    450,   # ajusté de 500
    # JILANI : 6965t / 73j = 95t/j → cap réel ajusté à 100t/j
    "JILANI OBAY":      100,   # inchangé
}

# ✅ Caps MAXIMUM avec JOURS DOUBLES (2 livraisons matin + après-midi)
# Permet d'absorber les pics de récolte pendant PIC (1-15 juillet)
COMMERCIAL_CAPS_DOUBLE = {
    "FEDI":             1300,  # 850 × 1.5 ≈ jour double partiel
    "MAKKI BEN SALAH":  1200,
    "KHALIL":           1100,
    "ACHREF AJLANI":     700,
    "JILANI OBAY":       150,
}
FACTORY_CAPS = {
    # ✅ CAP = capacité PHYSIQUE max de l'usine (info, jamais utilisé comme contrainte)
    "SICAM":    1500,
    "TUCAL":     800,
    "COMOCAP":   800,
    "ABIDA":     200,
    "ELFALLEH":  150,
}
FACTORY_LIMITS = {
    # ✅ LIMITE = ce qu'on AUTORISE à envoyer (toujours < CAP pour marge de sécurité)
    # C'est CETTE valeur qui sert de contrainte dure dans l'optimizer
    "SICAM":    1300,   # 200t de marge sous cap 1500
    "TUCAL":     700,   # 100t de marge sous cap 800
    "COMOCAP":   700,   # 100t de marge sous cap 800
    "ABIDA":     200,   # marge 0 (limite = cap)
    "ELFALLEH":  150,   # marge 0 (limite = cap)
}

# ✅ MARGE pour absorber l'arrondi à la dizaine
# Réduite à 5% (était 10%) pour ne pas perdre trop de tonnage
# L'arrondi à la dizaine ajoute en moyenne +2-3% → 5% de marge est suffisant
ROUNDING_MARGIN_PCT = 0.02  # 5% de marge sous le cap pour le solveur
ROUNDING_MARGIN_MIN = 30    # minimum 30t de marge
FLEET_CAPACITY = {
    "TRACTEUR":         (9,  11),    # min/max tonnes par voyage (moyenne ~10t)
    "PPL":              (6,  14),    # Petit Poilour
    "PL":               (15, 25),   # Poilour
    "SEMI":             (27, 33),   # Semi-remorque
    "DOUBLE_REM":       (27, 33),   # Double Remorque
}

# ── Flotte réelle depuis transport_disponible.xlsx ────────────────────
# Chaque usine a ses propres bennes avec capacités exactes
# Format: {"SICAM": {"SEMI":[30,30,...], "PL":[22,20,20,...], "PPL":[14]}, ...}
REAL_FLEET = {}

def _load_real_fleet():
    """
    Charge les capacités réelles depuis le fichier transport.
    Priorité: transport_etat_final.xlsx → transport_disponible.xlsx
    Logique: Confirmation=ok ET Contrat ≠ "En attente"
    """
    import os, pandas as pd
    from collections import defaultdict
    
    script_dir = os.path.dirname(os.path.abspath(__file__))
    # ✅ Chercher d'abord le nouveau fichier, puis l'ancien en fallback
    candidates = [
        os.path.join(script_dir, "transport_etat_final.xlsx"),
        "transport_etat_final.xlsx",
        os.path.join(script_dir, "transport_disponible.xlsx"),
        "transport_disponible.xlsx",
    ]
    fpath = next((p for p in candidates if os.path.exists(p)), None)
    if not fpath:
        return {}
    
    try:
        # ✅ Lire la feuille "liste confirmé" si elle existe
        xl = pd.ExcelFile(fpath)
        sheet = "liste confirmé" if "liste confirmé" in xl.sheet_names else xl.sheet_names[0]
        df = pd.read_excel(fpath, sheet_name=sheet)
        cols_upper = [str(c).strip().lower() for c in df.columns]
        
        # Détecter le format
        if "usine" in cols_upper and "tonnage" in cols_upper:
            df.columns = [str(c).strip() for c in df.columns]
            usine_col = next(c for c in df.columns if c.lower()=="usine")
            ton_col   = next(c for c in df.columns if c.lower()=="tonnage")
            type_col  = next(c for c in df.columns if "type" in c.lower() and "vehicule" in c.lower())
            conf_col  = next((c for c in df.columns 
                              if c.lower() in ("confirmation","actif")), None)
            cont_col  = next((c for c in df.columns 
                              if c.lower() == "contrat"), None)
            
            df["_usine"]  = df[usine_col].astype(str).str.strip().str.upper()
            df["_tonnage"]= pd.to_numeric(df[ton_col], errors="coerce")
            df["_type"]   = df[type_col].astype(str).str.strip().str.upper()
            # ✅ Confirmation = ok
            if conf_col:
                df["_actif"] = df[conf_col].astype(str).str.strip().str.lower()
            else:
                df["_actif"] = "ok"
            # ✅ Exclure Contrat = "En attente"
            if cont_col:
                df["_exclu"] = df[cont_col].astype(str).str.strip().str.lower().str.contains("attente", na=False)
            else:
                df["_exclu"] = False
        else:
            df["_usine"]  = df.iloc[:,1].astype(str).str.strip().str.upper()
            df["_tonnage"]= pd.to_numeric(df.iloc[:,0], errors="coerce")
            df["_type"]   = df.iloc[:,4].astype(str).str.strip().str.upper()
            df["_actif"]  = df.iloc[:,11].astype(str).str.strip().str.lower() \
                            if len(df.columns)>11 else pd.Series("ok", index=df.index)
            df["_exclu"]  = False
        
        # Normaliser usine
        USINE_N = {
            "SICAM":"SICAM","COMOCAP":"COMOCAP","COMOCAB":"COMOCAP",
            "TUCAL":"TUCAL","ABIDA":"ABIDA",
            "EL FALLEH":"ELFALLEH","ELFALLEH":"ELFALLEH","FALLEH":"ELFALLEH",
            "FELLA":"ELFALLEH",
            "LUI-MEME":"LUIMEME","LUI-MÊME":"LUIMEME","LUIMEME":"LUIMEME",
            "BOURAK":"BOURAK","TOTAL":"SKIP",
        }
        df["_usine"] = df["_usine"].map(lambda x: USINE_N.get(x, x))
        
        def norm_vtype(t):
            t = str(t).strip().upper()
            if "SEMI" in t or "2*6" in t or "DOUBLE" in t or "REMORQUE" in t: return "SEMI"
            if "PPL" in t or "PELÉE" in t or "PELEE" in t: return "PPL"
            if t.startswith("PL"):  return "PL"
            if "TRACTEUR" in t:     return "TRACTEUR"
            return t
        df["_type"] = df["_type"].apply(norm_vtype)
        
        # ✅ Filtrer: actif=ok ET pas exclu (Contrat ≠ En attente)
        df_ok = df[
            (df["_actif"]=="ok") &
            (~df["_exclu"]) &
            df["_tonnage"].notna() & 
            (df["_tonnage"]>0) &
            (df["_usine"]!="SKIP")
        ]
        
        fleet = defaultdict(lambda: defaultdict(list))
        for _, row in df_ok.iterrows():
            fleet[row["_usine"]][row["_type"]].append(float(row["_tonnage"]))
        
        result = {}
        for usine, vtypes in fleet.items():
            result[usine] = {vt: sorted(caps, reverse=True)
                             for vt, caps in vtypes.items()}
        # ✅ Log du fichier chargé
        fname = os.path.basename(fpath)
        total_bn = sum(len(v) for vt in result.values() for v in vt.values())
        print(f"  📂 Fichier transport: {fname} ({total_bn} bennes confirmées)")
        return result
    except Exception as e:
        print(f"  ⚠️  Fichier transport non chargé: {e}")
        return {}

REAL_FLEET = _load_real_fleet()
if REAL_FLEET:
    print(f"  ✅ Flotte réelle chargée: {sum(sum(len(v) for v in vt.values()) for vt in REAL_FLEET.values())} véhicules")
    for usine, vts in sorted(REAL_FLEET.items()):
        n = sum(len(v) for v in vts.values())
        t = sum(sum(v) for v in vts.values())
        print(f"     {usine:<10}: {n:2d} bennes | {t:.0f}t/j disponible")
else:
    print(f"  ℹ️  transport_disponible.xlsx absent — utilisation capacités théoriques")

# Alias rétrocompatibilité
FLEET_CAPACITY["PETIT POILOUR"] = FLEET_CAPACITY["PPL"]
FLEET_CAPACITY["POILOUR"]       = FLEET_CAPACITY["PL"]

ACCESS_VEHICLES = {
    # ── 1 véhicule ─────────────────────────────────────────────────
    "PL":          ["PL"],                       # PL seulement
    "PPL":         ["PPL"],                      # PPL seulement
    "SEMI":        ["SEMI"],                     # ✅ SEMI seul → SEMI uniquement (ex: ACHREF Gafsa)
    "RM":          ["SEMI"],                     # RM = 100% Semi (récolte mécanique)
    # ── 2 véhicules ────────────────────────────────────────────────
    "PL/PPL":      ["PL", "PPL"],
    "PL-PPL":      ["PL", "PPL"],
    "TRC/PPL":     ["TRACTEUR", "PPL"],          # tracteur + PPL
    "PL/SEMI":     ["PL", "SEMI"],
    "PL-SEMI":     ["PL", "SEMI"],
    # ── 3 véhicules ────────────────────────────────────────────────
    "TRC/PPL/PL":  ["TRACTEUR", "PPL", "PL"],   # tracteur + PPL + PL
    "PL/PPL/TRC":  ["PL", "PPL", "TRACTEUR"],
    "PL/PPL/SEMI": ["PL", "PPL", "SEMI"],        # PL + PPL + SEMI
    "SEMI/PL/PPL": ["SEMI", "PL", "PPL"],
    # ── Fallback ───────────────────────────────────────────────────
    "NAN":         ["TRACTEUR", "PPL", "PL", "SEMI"],
}

# ── Capacités confirmées par usine (source: transport_12_mai.xlsx) ──────
# Mise à jour: 12 mai 2026 — liste confirmée uniquement
# ── Données exactes depuis transport_12_mai.xlsx (liste confirmée) ──────────
TRANSPORT_CONFIRMED = {
    # Source: transport_etat_final.xlsx (10/06/2026)
    # Logique: Confirmation=ok ET Contrat ≠ "En attente"
    # ⚠️  ELFALLEH = 0 dans ce fichier (2 PPL EL FALLEH = autre libellé, non rattachés)
    "SICAM":    {"total": 1381, "PL": 927, "PPL": 64,  "SEMI": 390, "nb_bennes": 67},
    "TUCAL":    {"total": 363,  "PL": 303, "PPL": 0,   "SEMI": 60,  "nb_bennes": 19},
    "COMOCAP":  {"total": 328,  "PL": 91,  "PPL": 147, "SEMI": 90,  "nb_bennes": 23},
    "ABIDA":    {"total": 80,   "PL": 20,  "PPL": 0,   "SEMI": 60,  "nb_bennes": 3},
    "ELFALLEH": {"total": 24,   "PL": 0,   "PPL": 24,  "SEMI": 0,   "nb_bennes": 2},
}
# Jokers = BOURAK et LUI-MÊME (renforts toutes usines)
TRANSPORT_JOKERS = {
    "BOURAK":   {"total": 114, "PL": 114, "PPL": 0,  "SEMI": 0, "nb_bennes": 6},
    "LUIMEME":  {"total": 101, "PL": 55,  "PPL": 46, "SEMI": 0, "nb_bennes": 7},
}

# ── Règles de complétion transport par usine ──────────────────────────
# Quand la capacité confirmée ne couvre pas le cap journalier,
# on complète avec ces proportions (règles professionnelles)
TRANSPORT_REGLES = {
    "COMOCAP":  [("TRACTEUR", 100), ("PPL", 0.50), ("PL", 0.30), ("SEMI", 0.20)],
    "TUCAL":    [("PL", 0.30),      ("SEMI", 0.70)],
    "ABIDA":    [("PL", 0.50),      ("SEMI", 0.50)],
    "ELFALLEH": [("PPL", 0.70),     ("PL", 0.30)],
    "SICAM":    [("SEMI", 1.00)],
}

# ── Besoins de transport calculés (reste à compléter) ────────────────
def calc_transport_needs():
    """
    Calcule le tonnage manquant par usine et type de véhicule.
    Retourne dict {usine: {type: tonnes_manquantes}}
    """
    needs = {}
    for usine, cap in FACTORY_CAPS.items():
        conf  = TRANSPORT_CONFIRMED.get(usine, {}).get("total", 0)
        reste = max(0, cap - conf)
        if reste == 0:
            needs[usine] = {}
            continue
        regles = TRANSPORT_REGLES.get(usine, [])
        reste_var = reste
        usine_needs = {}
        for vtype, share in regles:
            if isinstance(share, int):
                usine_needs[vtype] = share
                reste_var -= share
            else:
                usine_needs[vtype] = round(reste_var * share)
        needs[usine] = usine_needs
    return needs

TRANSPORT_NEEDS = calc_transport_needs()

# ── Allocation jokers (PL→TUCAL/COMOCAP, PPL→ELFALLEH/COMOCAP) ──────
def alloc_jokers():
    """Alloue les jokers aux usines qui en ont besoin, par priorité de manque."""
    jpl  = TRANSPORT_JOKERS["BOURAK"]["PL"]
    jppl = TRANSPORT_JOKERS["LUIMEME"]["PPL"]
    jpl_luimeme = TRANSPORT_JOKERS["LUIMEME"]["PL"]
    alloc = {}
    for usine in FACTORY_CAPS:
        needs = TRANSPORT_NEEDS.get(usine, {})
        alloc[usine] = {}
        if "PL" in needs and jpl > 0:
            a = min(needs["PL"], jpl)
            alloc[usine]["PL_joker"] = a
            jpl -= a
        if "PL" in needs and jpl_luimeme > 0:
            a = min(needs.get("PL",0) - alloc[usine].get("PL_joker",0), jpl_luimeme)
            if a > 0:
                alloc[usine]["PL_joker"] = alloc[usine].get("PL_joker",0) + a
                jpl_luimeme -= a
        if "PPL" in needs and jppl > 0:
            a = min(needs["PPL"], jppl)
            alloc[usine]["PPL_joker"] = a
            jppl -= a
    return alloc

JOKER_ALLOC = alloc_jokers()

# Transport COMOCAP — TRACTEUR est un VRAI transport (10t/voyage)
# Le tracteur fait partie de la flotte normale COMOCAP, pas pour caisses vides.
# Désactivé : la logique CAISSE_VIDE était incorrecte.
COMOCAP_EXTRA_TRACTEUR = False

# ============================================================
# DISTANCE MATRIX — Zone → Usine (km)
# SICAM = 70% Borj Sedeya + 30% Mejez el Bab
# ============================================================
DISTANCE_KM = {
    # CAP BON 1 (Dar Allouche, Haouaria, Kelibia)
    "DAR ALLOUCH":          {"COMOCAP":15,  "ELFALLEH":20,  "SICAM":94,  "TUCAL":100, "ABIDA":220},
    "KORBA":                {"COMOCAP":20,  "ELFALLEH":15,  "SICAM":94,  "TUCAL":95,  "ABIDA":215},
    "KORBA/SOMAA":          {"COMOCAP":22,  "ELFALLEH":18,  "SICAM":95,  "TUCAL":97,  "ABIDA":217},
    "SOMAA":                {"COMOCAP":25,  "ELFALLEH":20,  "SICAM":89,  "TUCAL":90,  "ABIDA":210},
    "LEBNA":                {"COMOCAP":30,  "ELFALLEH":25,  "SICAM":94,  "TUCAL":95,  "ABIDA":215},
    "LEBNA/TAMEZRRAT":      {"COMOCAP":32,  "ELFALLEH":27,  "SICAM":95,  "TUCAL":97,  "ABIDA":217},
    "DIAR HOJJEJ":          {"COMOCAP":10,  "ELFALLEH":12,  "SICAM":97,  "TUCAL":98,  "ABIDA":218},
    "DIAR HOJJEJ/KHARREZ":  {"COMOCAP":12,  "ELFALLEH":14,  "SICAM":96,  "TUCAL":99,  "ABIDA":219},
    "TEFELOUN":             {"COMOCAP":18,  "ELFALLEH":22,  "SICAM":90,  "TUCAL":92,  "ABIDA":212},
    "TEFELOUN/DIAR HOJJEJ": {"COMOCAP":15,  "ELFALLEH":18,  "SICAM":94,  "TUCAL":95,  "ABIDA":215},
    "ATHLETH":              {"COMOCAP":8,   "ELFALLEH":10,  "SICAM":98,  "TUCAL":100, "ABIDA":220},
    "ATHLETH/HTOUBA":       {"COMOCAP":10,  "ELFALLEH":12,  "SICAM":96,  "TUCAL":98,  "ABIDA":218},
    "HTOUBA":               {"COMOCAP":12,  "ELFALLEH":15,  "SICAM":96,  "TUCAL":97,  "ABIDA":217},
    "KHADHRA":              {"COMOCAP":20,  "ELFALLEH":18,  "SICAM":91,  "TUCAL":93,  "ABIDA":213},
    "SIDI KHELIFA":         {"COMOCAP":25,  "ELFALLEH":22,  "SICAM":88,  "TUCAL":90,  "ABIDA":210},
    "MENZEL HORR":          {"COMOCAP":28,  "ELFALLEH":25,  "SICAM":86,  "TUCAL":88,  "ABIDA":208},
    "BIR LAHFAY":           {"COMOCAP":35,  "ELFALLEH":30,  "SICAM":82,  "TUCAL":85,  "ABIDA":205},
    "OUED CHIBA":           {"COMOCAP":55,  "ELFALLEH":58,  "SICAM":58,  "TUCAL":62,  "ABIDA":182},
    "GOURCHIN":             {"COMOCAP":42,  "ELFALLEH":45,  "SICAM":70,  "TUCAL":72,  "ABIDA":192},
    # CAP BON 2 (Menzel Temime → Nabeul)
    "MENZEL TAMIM":         {"COMOCAP":40,  "ELFALLEH":45,  "SICAM":74,  "TUCAL":75,  "ABIDA":195},
    "MENZEL MHIRI":         {"COMOCAP":45,  "ELFALLEH":50,  "SICAM":69,  "TUCAL":70,  "ABIDA":190},
    "GROMBELIA":            {"COMOCAP":50,  "ELFALLEH":55,  "SICAM":64,  "TUCAL":65,  "ABIDA":185},
    "SIDI AICH":            {"COMOCAP":48,  "ELFALLEH":52,  "SICAM":67,  "TUCAL":68,  "ABIDA":188},
    "SIDI OTHMAN":          {"COMOCAP":52,  "ELFALLEH":55,  "SICAM":64,  "TUCAL":65,  "ABIDA":185},
    # NORD (Tunis → Jendouba)
    "FARTOUNA":             {"COMOCAP":95,  "ELFALLEH":100, "SICAM":51,  "TUCAL":35,  "ABIDA":120},
    "GAR DIMAOU":           {"COMOCAP":120, "ELFALLEH":125, "SICAM":65,  "TUCAL":60,  "ABIDA":100},
    "JANDOUBA":             {"COMOCAP":130, "ELFALLEH":135, "SICAM":71,  "TUCAL":70,  "ABIDA":90},
    "BOU SALEM":            {"COMOCAP":115, "ELFALLEH":120, "SICAM":64,  "TUCAL":55,  "ABIDA":105},
    "WED MLIZ":             {"COMOCAP":110, "ELFALLEH":115, "SICAM":61,  "TUCAL":50,  "ABIDA":110},
    "SIDI HASSOUN":         {"COMOCAP":100, "ELFALLEH":105, "SICAM":54,  "TUCAL":40,  "ABIDA":115},
    "AMAYMIA":              {"COMOCAP":80,  "ELFALLEH":85,  "SICAM":47,  "TUCAL":30,  "ABIDA":130},
    "ZAAFRANA-ELKHADHRA":   {"COMOCAP":85,  "ELFALLEH":90,  "SICAM":49,  "TUCAL":35,  "ABIDA":125},
    "ZAAFRIA":              {"COMOCAP":82,  "ELFALLEH":87,  "SICAM":48,  "TUCAL":33,  "ABIDA":127},
    "BOUJRIDA":             {"COMOCAP":88,  "ELFALLEH":93,  "SICAM":52,  "TUCAL":37,  "ABIDA":122},
    "OUED KHATEF":          {"COMOCAP":105, "ELFALLEH":110, "SICAM":58,  "TUCAL":45,  "ABIDA":112},
    "MEDJEZ BEB":           {"COMOCAP":100, "ELFALLEH":105, "SICAM":49,  "TUCAL":45,  "ABIDA":110},
    "AMAYMIA":              {"COMOCAP":80,  "ELFALLEH":85,  "SICAM":47,  "TUCAL":30,  "ABIDA":130},
    "BOR AMRI":             {"COMOCAP":78,  "ELFALLEH":83,  "SICAM":45,  "TUCAL":28,  "ABIDA":132},
    "AWAMRIYA":             {"COMOCAP":92,  "ELFALLEH":97,  "SICAM":53,  "TUCAL":40,  "ABIDA":118},
    "FRININ":               {"COMOCAP":75,  "ELFALLEH":80,  "SICAM":44,  "TUCAL":25,  "ABIDA":135},
    "GOMBAR":               {"COMOCAP":83,  "ELFALLEH":88,  "SICAM":49,  "TUCAL":32,  "ABIDA":128},
    "BENI AYECH":           {"COMOCAP":110, "ELFALLEH":115, "SICAM":61,  "TUCAL":55,  "ABIDA":100},
    "BIR MASOUDA":          {"COMOCAP":108, "ELFALLEH":113, "SICAM":59,  "TUCAL":52,  "ABIDA":102},
    # KAIROUAN / CENTRE
    "SBIKHA-CHRARDA":       {"COMOCAP":150, "ELFALLEH":155, "SICAM":114, "TUCAL":110, "ABIDA":130},
    "MAJEL BELABESS":       {"COMOCAP":160, "ELFALLEH":165, "SICAM":124, "TUCAL":120, "ABIDA":120},
    "CHEBIKA-ELHAWEREB":    {"COMOCAP":155, "ELFALLEH":160, "SICAM":119, "TUCAL":115, "ABIDA":125},
    "ELHAWEREB-AIN BIDHA-HAFOUZ": {"COMOCAP":165, "ELFALLEH":170, "SICAM":129, "TUCAL":125, "ABIDA":115},
    "OULED ZID":            {"COMOCAP":145, "ELFALLEH":150, "SICAM":109, "TUCAL":105, "ABIDA":135},
    "BATTEN":               {"COMOCAP":158, "ELFALLEH":163, "SICAM":122, "TUCAL":118, "ABIDA":122},
    "GARAT SASSI":          {"COMOCAP":162, "ELFALLEH":167, "SICAM":126, "TUCAL":122, "ABIDA":118},
    "BELYES":               {"COMOCAP":148, "ELFALLEH":153, "SICAM":112, "TUCAL":108, "ABIDA":132},
    # SIDI BOUZID
    "SIDI BOUZID":          {"COMOCAP":200, "ELFALLEH":205, "SICAM":164, "TUCAL":160, "ABIDA":80},
    "OM ADHAM":             {"COMOCAP":195, "ELFALLEH":200, "SICAM":159, "TUCAL":155, "ABIDA":85},
    "TBAG":                 {"COMOCAP":198, "ELFALLEH":203, "SICAM":162, "TUCAL":158, "ABIDA":82},
    # KASSRINE / GAFSA
    "FERIANA":              {"COMOCAP":230, "ELFALLEH":235, "SICAM":190, "TUCAL":185, "ABIDA":60},
    "CENTRE -OUEST":        {"COMOCAP":180, "ELFALLEH":185, "SICAM":144, "TUCAL":140, "ABIDA":100},
    "GAFSA":                {"COMOCAP":280, "ELFALLEH":285, "SICAM":239, "TUCAL":235, "ABIDA":55},
    # BOUFICHA
    "BOUFICHA":             {"COMOCAP":60,  "ELFALLEH":65,  "SICAM":55,  "TUCAL":55,  "ABIDA":160},
}

# Region mapping
ZONE_REGION = {
    # ── CAP BON 1 (SUD du Cap Bon: Korba, Lebna, Diar Hojjej, etc.) ──
    "KORBA":"CAP BON 1","KORBA/SOMAA":"CAP BON 1","SOMAA":"CAP BON 1",
    "LEBNA":"CAP BON 1","LEBNA/TAMEZRRAT":"CAP BON 1",
    "MENZEL HORR":"CAP BON 1","DIAR HOJJEJ":"CAP BON 1",
    "DIAR HOJJEJ/KHARREZ":"CAP BON 1","TEFELOUN":"CAP BON 1",
    "TEFELOUN/DIAR HOJJEJ":"CAP BON 1","OUED CHIBA":"CAP BON 1",
    "GOURCHIN":"CAP BON 1","GARAT SASSI":"CAP BON 1",
    "MANZEL GAMOUDI":"CAP BON 1","SIDI HASSOUN":"CAP BON 1",
    "BENI AYECH":"CAP BON 1","FARTOUNA":"CAP BON 1",
    "ATHLETH":"CAP BON 1","ATHLETH/HTOUBA":"CAP BON 1","HTOUBA":"CAP BON 1",
    "FRININ":"CAP BON 1","GOMBAR":"CAP BON 1","TBAG":"CAP BON 1",
    "GROMBELIA":"CAP BON 1","GROMBALIA":"CAP BON 1",
    "BOU ARGOUB":"CAP BON 1","SLIMEN":"CAP BON 1","TEKELSA":"CAP BON 1",
    "JAMMEL":"CAP BON 1","JBENYANA":"CAP BON 1","MOKNINE":"CAP BON 1",
    "MENZEL HAYET":"CAP BON 1","OUED KHATEF":"CAP BON 1",
    "BIR MASOUDA":"CAP BON 1","BELYES":"CAP BON 1","BOUJRIDA":"CAP BON 1",
    
    # ── CAP BON 2 (NORD du Cap Bon + Bouficha: Dar Allouch, Menzel Tamim) ──
    "DAR ALLOUCH":"CAP BON 2","D.AL":"CAP BON 2",
    "MENZEL TAMIM":"CAP BON 2","MENZEL TMIME":"CAP BON 2",
    "BENI KHIAR":"CAP BON 2","BENI KALLED":"CAP BON 2",
    "MENZEL BOUZELFA":"CAP BON 2","MENZEL NOUR":"CAP BON 2",
    "MZAWGHA":"CAP BON 2","ALIA":"CAP BON 2","TUNIS":"CAP BON 2",
    "NABEUL":"CAP BON 2",
    # BOUFICHA intégrée dans CAP BON 2
    "SIDI SAIID":"CAP BON 2","SIDI KHELIFA":"CAP BON 2",
    "SIDI SAID":"CAP BON 2","SIDI KHLIFA":"CAP BON 2",
    "BOUFICHA":"CAP BON 2","ENFIDHA":"CAP BON 2","SOUSSE":"CAP BON 2",
    
    # ── NORD (Jendouba, Beja, Manouba, Bizerte) ──
    "JANDOUBA":"NORD","JENDOUBA":"NORD",
    "GAR DIMAOU":"NORD","MEDJEZ BEB":"NORD",
    "BOR AMRI":"NORD","BORJ AMRI":"NORD",
    "WED MLIZ":"NORD","SIDI ISMAIL":"NORD",
    "BELLARIGIA":"NORD","BELARIGIA":"NORD",
    "BOU SALEM":"NORD","BOUSSALEM":"NORD",
    "BIR LAHFAY":"NORD","BIR DRASSEN":"NORD",
    "ZARMDINE":"NORD","BOU KRIM":"NORD",
    "MEDJEZ EL BAB":"NORD","MEJEZ EL BAB":"NORD",
    "BEJA":"NORD","MANOUBA":"NORD","BIZERTE":"NORD",
    "TESTOUR":"NORD",
    
    # ── KAIROUAN ──
    "BATTEN":"KAIROUAN","SBIKHA-CHRARDA":"KAIROUAN",
    "AWAMRIYA":"KAIROUAN","ZAAFRANA-ELKHADHRA":"KAIROUAN",
    "ELHAWEREB-AIN BIDHA-HAFOUZ":"KAIROUAN","CHEBIKA-ELHAWEREB":"KAIROUAN",
    "KHADHRA":"KAIROUAN","ZAAFRIA":"KAIROUAN",
    "MENZEL MHIRI":"KAIROUAN","KAIROUAN":"KAIROUAN","KAIRAOUAN":"KAIROUAN",
    
    # ── SIDI BOUZID (incluant OM ADHAM) ──
    "SIDI BOUZID":"SIDI BOUZID","SIDI OTHMAN":"SIDI BOUZID",
    "SIDIBOUZID":"SIDI BOUZID","OM ADHAM":"SIDI BOUZID",
    
    # ── GAFSA / KASSRINE (incluant MAJEL BELABESS, AMAYMIA, SIDI AICH) ──
    "OULED OMRAN":"GAFSA / KASSRINE","OULED ZID":"GAFSA / KASSRINE",
    "SIDI AICH":"GAFSA / KASSRINE","AMAYMIA":"GAFSA / KASSRINE",
    "FERIANA":"GAFSA / KASSRINE","KASSERINE":"GAFSA / KASSRINE",
    "GAFSA":"GAFSA / KASSRINE","SBEITLA":"GAFSA / KASSRINE",
    "KASSRINE":"GAFSA / KASSRINE","MAJEL BELABESS":"GAFSA / KASSRINE",
    "CENTRE -OUEST":"GAFSA / KASSRINE",
}

def get_distance(zone: str, usine: str) -> int:
    """Get distance km from zone to usine. Returns 999 if unknown."""
    zone_up = zone.upper().strip()
    d = DISTANCE_KM.get(zone_up, {})
    return d.get(usine, 999)

def get_best_usine(zone: str, allowed_usines: list) -> str:
    """Return the closest usine to this zone among allowed list."""
    zone_up = zone.upper().strip()
    dists = {u: get_distance(zone_up, u) for u in allowed_usines}
    return min(dists, key=dists.get)

def get_region(zone: str) -> str:
    return ZONE_REGION.get(zone.upper().strip(), "AUTRE")

# ============================================================
# SUPABASE CONNECTION
# ============================================================
SUPABASE_URL = "https://mwjefdqfzrtsfzspeppg.supabase.co"
SUPABASE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6Im13amVmZHFmenJ0c2Z6c3BlcHBnIiwicm9sZSI6ImFub24iLCJpYXQiOjE3ODAzODg3MDgsImV4cCI6MjA5NTk2NDcwOH0.H5IX2uLneHdvwyLJrgN7OrHGrLSZSrQBAuJeuzCEz44"

# ============================================================
# STEP 1: Load ONLY from Supabase — farmers uploaded by commercials
# ============================================================
print("Connexion à Supabase...")
try:
    from supabase import create_client
    sb = create_client(SUPABASE_URL, SUPABASE_KEY)
    print("  OK")
except Exception as e:
    print(f"  ERREUR: {e}")
    print("  Installe supabase: pip install supabase")
    import sys; sys.exit(1)

# Get list of commercials who have deposited (statut = 'depose')
print("\nVérification des dépôts...")
try:
    depot_data = sb.table("depot_status").select("*").eq("statut","depose").execute().data
    deposited  = [r["commercial"] for r in depot_data] if depot_data else []
except Exception:
    # If depot_status table doesn't exist yet, use all agriculteurs
    deposited = []

if not deposited:
    print("  Aucun commercial n'a déposé via depot_status.")
    print("  Utilisation de tous les agriculteurs dans Supabase.")
    deposited = None  # means: use all

# Load agriculteurs from Supabase
print("Chargement des agriculteurs depuis Supabase...")
try:
    agri_data = sb.table("agriculteurs").select("*").execute().data
    if not agri_data:
        print("  ERREUR: Table agriculteurs vide.")
        print("  Lance d'abord python migrate.py ou demande aux commerciaux d'uploader.")
        import sys; sys.exit(1)
    df_all = pd.DataFrame(agri_data)
except Exception as e:
    print(f"  ERREUR lecture Supabase: {e}")
    import sys; sys.exit(1)

# Filter: only farmers from commercials who deposited
if deposited is not None:
    df_all["commercial"] = df_all["commercial"].astype(str).str.strip()
    df = df_all[df_all["commercial"].isin(deposited)].copy()
    missing = [c for c in ["FEDI","MAKKI BEN SALAH","KHALIL","ACHREF AJLANI","JILANI OBAY"]
               if c not in deposited]
    if missing:
        print(f"  ⚠️  Pas encore déposé: {missing}")
        print(f"  Planning généré SEULEMENT pour: {deposited}")
else:
    df = df_all.copy()

# Rename columns to match internal format
df = df.rename(columns={
    "nom":           "AGRICULTEUR",
    "tonnage_total": "TONNAGE",
    "usine":         "USINE",
    "accessibilite": "accessbilite",
    "region":        "REGION",
    "zone":          "ZONNE",
})

df = df.dropna(subset=["AGRICULTEUR","TONNAGE","USINE"])
df = df[pd.to_numeric(df["TONNAGE"], errors="coerce") > 0]
df["TONNAGE"]      = pd.to_numeric(df["TONNAGE"], errors="coerce")
df["commercial"]   = df["commercial"].astype(str).str.strip()
df["USINE"]        = df["USINE"].astype(str).str.strip().str.upper()   # ← .upper() obligatoire !
df["accessbilite"] = df["accessbilite"].fillna("PL/PPL").astype(str).str.strip().str.upper()
def normalize_acc(x):
    """
    Normalise l'accessibilité en RESPECTANT EXACTEMENT ce que le commercial a écrit.
    PL → PL  |  PPL → PPL  |  PL/PPL → PL/PPL  |  PL/SEMI → PL/SEMI
    Aucune conversion automatique entre types simples.
    """
    x = str(x).strip().upper().replace("-", "/")
    if x in ("NAN","NONE","","NAT"): return "PL/PPL"  # défaut si vide
    if x == "RM":                    return "RM"
    # ✅ Déjà valide tel quel → retourner sans modifier
    if x in ACCESS_VEHICLES:         return x
    # PL seul ou PPL seul → respecter exactement
    if x == "PL":                    return "PL"
    if x == "PPL":                   return "PPL"
    # Parser les combinaisons
    import re as _re
    parts = set(_re.split(r"[/,;\s]+", x))
    parts = {p.strip() for p in parts if p.strip()}
    has_trc  = "TRC" in parts or "TRACTEUR" in parts
    has_pl   = "PL" in parts
    has_ppl  = "PPL" in parts
    has_semi = "SEMI" in parts
    # Combinaisons avec TRACTEUR
    if has_trc and has_pl and has_ppl:  return "TRC/PPL/PL"
    if has_trc and has_pl:              return "TRC/PPL/PL"
    if has_trc and has_ppl:             return "TRC/PPL"
    # Combinaisons avec SEMI
    if has_semi and has_pl and has_ppl: return "PL/PPL/SEMI"
    if has_semi and has_ppl:            return "PL/PPL/SEMI"
    if has_semi and has_pl:             return "PL/SEMI"
    if has_semi:                        return "PL/SEMI"
    # PL + PPL ensemble
    if has_pl and has_ppl:              return "PL/PPL"
    # ✅ Respecter les types simples exactement
    if has_pl:                          return "PL"
    if has_ppl:                         return "PPL"
    return "PL/PPL"  # défaut absolu

df["accessbilite"] = df["accessbilite"].apply(normalize_acc)
df["ZONNE"]  = df["ZONNE"].fillna("").astype(str).str.strip().str.upper()
df["REGION"] = df["REGION"].fillna("").astype(str).str.strip().str.upper()
REGION_NORM_OPT = {
    # CAP BON
    "NABEUL":"CAP BON 2","CAPB1":"CAP BON 1","CAP B1":"CAP BON 1",
    "CAPB2":"CAP BON 2","CAP B2":"CAP BON 2",
    "CAP BON":"CAP BON 1",
    # GAFSA / KASSERINE (fusionnées)
    "GAFSA":"GAFSA / KASSRINE","KASSRINE":"GAFSA / KASSRINE",
    "KASSERINE":"GAFSA / KASSRINE","KASRINE":"GAFSA / KASSRINE",
    "KASSARINE":"GAFSA / KASSRINE","SBEITLA":"GAFSA / KASSRINE",
    "GAFSA / KASSERINE":"GAFSA / KASSRINE",
    "GAFSA/KASSRINE":"GAFSA / KASSRINE",
    # NORD
    "BEJA":"NORD","MANOUBA":"NORD","BIZERTE":"NORD",
    "JENDOUBA":"NORD","JANDOUBA":"NORD",
    "BIR LAHFAY":"NORD","BOR AMRI":"NORD","BORJ AMRI":"NORD",
    "MEDJEZ EL BAB":"NORD","MEJEZ EL BAB":"NORD","MEDJEZ BEB":"NORD",
    "TESTOUR":"NORD","BOUSSALEM":"NORD",
    # KAIROUAN
    "KAIRAOUAN":"KAIROUAN",
    # SIDI BOUZID
    "SIDIBOUZID":"SIDI BOUZID","SIDI BOU ZID":"SIDI BOUZID",
    # BOUFICHA = région séparée
    "BOUFICHA":"BOUFICHA",
    "SOUSSE":"BOUFICHA","ENFIDHA":"BOUFICHA","HAMMAMET":"CAP BON 1",
}
# Normaliser AVANT le replace
df["REGION"] = df["REGION"].astype(str).str.strip().str.upper()
df["REGION"] = df["REGION"].replace(REGION_NORM_OPT)
# Si toujours non reconnue, mettre AUTRE
KNOWN_REGIONS = {"CAP BON 1","CAP BON 2","NORD","KAIROUAN","SIDI BOUZID",
                 "GAFSA / KASSRINE","BOUFICHA","AUTRE"}
df["REGION"] = df["REGION"].where(df["REGION"].isin(KNOWN_REGIONS), "AUTRE")
df["date_debut"] = pd.to_datetime(df["date_debut"], errors="coerce")
df["date_fin"]   = pd.to_datetime(df["date_fin"],   errors="coerce")

date_cols = []  # no date matrix — dates come from date_debut/date_fin columns

print(f"  {len(df)} agriculteurs chargés depuis {len(deposited) if deposited else 'tous les'} commerciaux")
print(f"  Commerciaux inclus: {sorted(df['commercial'].unique().tolist())}")
print(f"  Usines: {sorted(df['USINE'].unique().tolist())}")
print(f"  Tonnage total: {df['TONNAGE'].sum():,.0f}t")

# ============================================================
# STEP 2: Build Farmer objects
# ============================================================
class Farmer:
    def __init__(self, row, date_cols):
        self.commercial  = row["commercial"]
        self.name        = str(row["AGRICULTEUR"]).strip()
        self.usine       = row["USINE"]
        self.region      = str(row.get("REGION", "") or "").strip()
        self.zone        = str(row.get("ZONNE", "") or "").strip()
        self.access      = row["accessbilite"]
        self.tonnage     = float(row["TONNAGE"])
        self.allowed_veh = ACCESS_VEHICLES.get(self.access, ACCESS_VEHICLES["NAN"])
        self.distance_km = get_distance(self.zone, self.usine)
        
        # ✅ PRIORITÉ à la région DÉCLARÉE par le commercial dans Supabase
        declared_region = str(self.region or "").strip().upper()
        if declared_region and declared_region not in ("", "AUTRE", "NAN", "NONE"):
            self.geo_region = declared_region
        else:
            self.geo_region = get_region(self.zone)
        
        # ✅ SEMI = TOUJOURS 30t physique par voyage
        # Le coefficient 0.7 (rotation longue distance) ne s'applique PAS à la capacité
        # mais au nombre max de voyages/jour par benne — géré ailleurs.
        # Chaque Semi transporte 30t — qu'il soit à 10km ou à 200km de l'usine.
        self.semi_coeff = 1.0   # ✅ TOUJOURS 1.0 (Semi = 30t fixe)

        # ── Build maturity window ──────────────────────────
        # Case 1: date_cols present (Excel fallback path)
        if date_cols:
            raw_window = {}
            for col in date_cols:
                val = row.get(col)
                if val is not None and pd.notna(val) and float(val) > 0:
                    d = col.date()
                    if SEASON_START <= d <= SEASON_END:
                        raw_window[d] = float(val)

            if raw_window:
                self.start = clamp_date(min(raw_window))
                self.end   = clamp_date(max(raw_window))
                date_sum   = sum(raw_window.values())
                if date_sum < self.tonnage * 0.5:
                    n_days = max(1, (self.end - self.start).days + 1)
                    daily  = self.tonnage / n_days
                    self.window = {
                        clamp_date(self.start + datetime.timedelta(days=i)): round(daily, 1)
                        for i in range(n_days)
                    }
                elif date_sum > self.tonnage * 1.5:
                    factor = self.tonnage / date_sum
                    self.window = {d: round(t * factor, 1) for d, t in raw_window.items()}
                else:
                    self.window = raw_window
                return

        # Case 2: Supabase path — use date_debut / date_fin columns
        try:
            raw_start = pd.to_datetime(row["date_debut"]).date()
            raw_end   = pd.to_datetime(row["date_fin"]).date()
        except Exception:
            raw_start = datetime.date(2026, 7, 1)
            raw_end   = datetime.date(2026, 7, 31)

        if pd.isna(raw_start): raw_start = datetime.date(2026, 7, 1)
        if pd.isna(raw_end):   raw_end   = datetime.date(2026, 7, 31)

        # ✅ DATE_FIN EST EXCLUSIVE : 15→19 = jours 15,16,17,18 (le 19 n'est PAS inclus)
        # Le commercial donne la date où il n'y a PLUS de récolte
        # → on récolte jusqu'à date_fin - 1
        raw_end = raw_end - datetime.timedelta(days=1)

        self.start  = clamp_date(raw_start)
        self.end    = clamp_date(raw_end)
        if self.end < self.start:
            # Cas pathologique (1 seul jour) → garder au moins 1 jour
            self.end = self.start
        if self.end <= self.start and self.tonnage > 100:
            # Tonnage important sur 1 jour → étendre à 30 jours
            self.end = clamp_date(self.start + datetime.timedelta(days=30))
        
        # ✅ EXTENSION SIMPLE — max 1 à 3 jours seulement
        # On respecte au mieux la fenêtre déclarée par le commercial.
        if self.tonnage < 300:
            ext_days = 1       # Petit agriculteur → +1 jour
        elif self.tonnage < 700:
            ext_days = 2       # Moyen → +2 jours
        else:
            ext_days = 3       # Gros (≥700t) → +3 jours max
        self.end = clamp_date(self.end + datetime.timedelta(days=ext_days))

        n_days = max(1, (self.end - self.start).days + 1)

        # ✅ EXTENSION SEMI-ONLY (ACHREF Gafsa/Kasserine) :
        # La fenêtre est calculée sur 30t/j (1 Semi) pour maximiser les jours
        # dans la fenêtre → OR-Tools peut distribuer sur plus de jours.
        # choose_vehicles applique ensuite la règle progressive (60/90/120t).
        if self.allowed_veh == ["SEMI"]:
            # Fenêtre calculée sur 30t/j pour maximiser le nombre de jours disponibles
            # La progression 60/90/120t est appliquée dans le post-traitement (arrondi)
            _max_per_day = 30.0
            _daily_check = self.tonnage / n_days
            if _daily_check > _max_per_day:
                _days_needed_semi = math.ceil(self.tonnage / _max_per_day)
                _extra_semi = _days_needed_semi - n_days
                if _extra_semi > 0:
                    self.end = clamp_date(self.end + datetime.timedelta(days=_extra_semi))
                    n_days = max(1, (self.end - self.start).days + 1)

        # ✅ EXTENSION DE FAISABILITÉ — uniquement si nécessaire
        # Un gros agriculteur sur une petite usine (ex: 1080t sur ELFALLEH cap 150t)
        # ne peut PAS tenir dans une fenêtre courte sans violer le cap usine.
        # On étend juste assez pour que sa charge/jour ≤ ~60% du cap usine
        # (60% laisse de la place aux autres agriculteurs de la même usine).
        _FACTORY_CAPS_LOCAL = {"SICAM":1500,"TUCAL":800,"COMOCAP":800,"ABIDA":200,"ELFALLEH":150}
        _ucap = _FACTORY_CAPS_LOCAL.get(str(self.usine).upper(), 800)
        _max_share = _ucap * 0.60        # part max d'un seul agriculteur sur l'usine
        _daily_now = self.tonnage / n_days
        if _daily_now > _max_share and _max_share > 0:
            _days_needed = math.ceil(self.tonnage / _max_share)
            # étendre la fin (sans dépasser la borne saison) pour atteindre _days_needed
            _extra = _days_needed - n_days
            if _extra > 0:
                self.end = clamp_date(self.end + datetime.timedelta(days=_extra))
                n_days = max(1, (self.end - self.start).days + 1)

        # ✅ COURBE DE MATURATION — basée sur la réalité agronomique
        # La tomate NE MÛRIT PAS uniformément :
        #   - Début de fenêtre : lente montée (plant en maturation)
        #   - Milieu de fenêtre : PIC de récolte (tomate prête = jours doubles ici)
        #   - Fin de fenêtre : déclin progressif (fin de récolte)
        #
        # Courbes par commercial/région :
        #   KHALIL (Kairouan/Bouficha/Sidi Bouzid) : 15% / 65% / 20%
        #     sur 20% / 55% / 25% du temps
        #   TOUS AUTRES (FEDI, MAKKI, JILANI, ACHREF hors RM) : 10% / 70% / 20%
        #     sur 15% / 60% / 25% du temps
        #   → Les jours doubles tombent au milieu (70% du tonnage) pas au début

        def _build_maturation_window(start, ndays, tonnage, pct_debut, pct_milieu,
                                     time_debut, time_milieu):
            """
            Courbe de maturation réaliste.
            pct_debut/milieu/fin = % du tonnage sur chaque phase
            time_debut/milieu = % du temps sur chaque phase (le reste = fin)
            """
            if ndays <= 2:
                return {start: round(tonnage, 1)}

            n_debut  = max(1, round(ndays * time_debut))
            n_milieu = max(1, round(ndays * time_milieu))
            n_fin    = ndays - n_debut - n_milieu
            if n_fin < 1:
                n_fin = 1
                if n_milieu > 1: n_milieu -= 1
                else:            n_debut  = max(1, n_debut - 1)

            pct_fin = 1.0 - pct_debut - pct_milieu
            t_debut  = tonnage * pct_debut
            t_milieu = tonnage * pct_milieu
            t_fin    = tonnage - t_debut - t_milieu   # = exact, pas de perte d'arrondi

            d_debut  = t_debut  / n_debut
            d_milieu = t_milieu / n_milieu
            d_fin    = t_fin    / max(1, n_fin)

            win = {}; idx = 0
            for _ in range(n_debut):
                win[start + datetime.timedelta(days=idx)] = round(d_debut, 1);  idx += 1
            for _ in range(n_milieu):
                win[start + datetime.timedelta(days=idx)] = round(d_milieu, 1); idx += 1
            for _ in range(n_fin):
                win[start + datetime.timedelta(days=idx)] = round(d_fin, 1);    idx += 1
            return win

        _is_khalil = str(self.commercial).strip().upper() == "KHALIL"

        if _is_khalil:
            # KHALIL — Kairouan / Bouficha / Sidi Bouzid (climat continental)
            # Maturité plus longue, pic plus tardif :
            # 20% / 60% / 20% du tonnage sur 20% / 50% / 30% du temps
            self.window = _build_maturation_window(
                self.start, n_days, self.tonnage,
                pct_debut=0.20, pct_milieu=0.60,
                time_debut=0.20, time_milieu=0.50,
            )
        else:
            # FEDI / MAKKI / JILANI / ACHREF (non-RM) — Cap Bon / Nord
            # Maturité plus rapide, pic au tiers :
            # 20% / 60% / 20% du tonnage sur 30% / 50% / 20% du temps
            self.window = _build_maturation_window(
                self.start, n_days, self.tonnage,
                pct_debut=0.20, pct_milieu=0.60,
                time_debut=0.30, time_milieu=0.50,
            )

farmers = [Farmer(row, date_cols) for _, row in df.iterrows()]

# ✅ DIFFÉRENCIATION DES NOMS pour agriculteurs avec plusieurs lots
# Ex: AMOR KHECHIN a 3 lignes :
#   - 600t RM   à SICAM    → "AMOR KHECHIN (RM-SICAM)"
#   - 408t PL   à TUCAL    → "AMOR KHECHIN (PL-TUCAL)"
#   - 408t PL   à ELFALLEH → "AMOR KHECHIN (PL-ELFALLEH)"
# Cette différenciation rend le planning et les rapports plus lisibles
# et garantit que chaque lot est traité indépendamment (Semi pour RM, PL pour PL).
from collections import Counter as _Counter
_name_count = _Counter((f.commercial, f.name) for f in farmers)
for f in farmers:
    if _name_count[(f.commercial, f.name)] > 1:
        # Plusieurs lots pour cet agriculteur → suffixer le nom
        _acc_simple = str(f.access).strip().upper()
        _usine_simple = str(f.usine).strip().upper()
        # Si RM → mettre RM en suffixe (priorité visuelle)
        if _acc_simple == "RM":
            f.name = f"{f.name} (RM-{_usine_simple})"
        else:
            f.name = f"{f.name} ({_acc_simple}-{_usine_simple})"

total_dist = sum(f.distance_km for f in farmers if f.distance_km < 999)
print(f"  Built {len(farmers)} farmers | Avg distance to usine: {total_dist/len(farmers):.0f}km")

# Show distance summary
print()
print("  Distance summary by region:")
from collections import defaultdict
by_region = defaultdict(list)
for f in farmers:
    by_region[f.geo_region].append(f.distance_km)
for region, dists in sorted(by_region.items()):
    valid = [d for d in dists if d < 999]
    if valid:
        print(f"    {region:<15}: avg {sum(valid)/len(valid):.0f}km to assigned usine")

# ============================================================
# P1 FIX — Pré-allocation RM : séquence 60/90/120t exacte
# ============================================================
def build_rm_schedule(tonnage: float) -> list:
    """
    Calcule la séquence exacte de livraisons pour un agriculteur RM/SEMI.
    Règle: j1=60t, j2=90t, j3+=120t, dernier jour = reste exact.
    Total == tonnage déclaré (écart=0).
    Exemples:
      250t → [60, 90, 100]   total=250 ✅
      300t → [60, 90, 120, 30] total=300 ✅
      600t → [60, 90, 120, 120, 120, 90] total=600 ✅
       60t → [60]            total=60  ✅
    """
    sequence  = []
    remaining = round(float(tonnage), 1)
    if remaining <= 0:
        return sequence
    # Jour 1 : min 60t
    j1 = min(60.0, remaining)
    sequence.append(round(j1, 1))
    remaining = round(remaining - j1, 1)
    if remaining <= 0:
        return sequence
    # Jour 2 : min 90t
    j2 = min(90.0, remaining)
    sequence.append(round(j2, 1))
    remaining = round(remaining - j2, 1)
    # Jours suivants : 120t ou reste
    while remaining > 0.1:
        if remaining <= 120.0:
            sequence.append(round(remaining, 1))
            remaining = 0.0
        else:
            sequence.append(120.0)
            remaining = round(remaining - 120.0, 1)
    return sequence

def is_rm_farmer(farmer) -> bool:
    return farmer.allowed_veh == ["SEMI"] or str(farmer.access).strip().upper() == "RM"

# ============================================================
# STEP 3: OR-Tools model with distance objective
# ============================================================
print("\nP1 Fix: Pré-allocation RM (règle 60/90/120t)...")

# Séparer RM (pré-alloués) et non-RM (OR-Tools)
rm_farmers    = [f for f in farmers if is_rm_farmer(f)]
nonrm_farmers = [f for f in farmers if not is_rm_farmer(f)]

# Pré-calculer les jours de livraison RM
rm_fixed_days = []
for farmer in rm_farmers:
    sequence    = build_rm_schedule(farmer.tonnage)
    dates_avail = sorted(farmer.window.keys())
    # Étendre la fenêtre si pas assez de jours pour la séquence
    if len(sequence) > len(dates_avail):
        last_d = dates_avail[-1]
        while len(dates_avail) < len(sequence):
            last_d = clamp_date(last_d + datetime.timedelta(days=1))
            if last_d not in dates_avail:
                dates_avail.append(last_d)
        print(f"    ⚠️ {farmer.name}: fenêtre étendue à {len(dates_avail)}j pour séquence {sequence}")
    for i, tons in enumerate(sequence):
        if i >= len(dates_avail):
            break
        date    = dates_avail[i]
        nb_semi = max(1, int(round(tons / 30.0)))
        veh_str = f"SEMI x{nb_semi}(30t)" if nb_semi > 1 else "SEMI(30t)"
        rm_fixed_days.append({
            "Commercial":    farmer.commercial,
            "Agriculteur":   farmer.name,
            "Usine":         farmer.usine,
            "Region":        farmer.geo_region,
            "Zone":          farmer.zone,
            "Accessibilite": farmer.access,
            "Date":          date,
            "Tonnes/Jour":   round(tons, 1),
            "Type Vehicule": "SEMI",
            "Vehicules":     veh_str,
            "Nb Voyages":    nb_semi,
            "Distance km":   farmer.distance_km if farmer.distance_km < 999 else 0,
            "Date Debut":    farmer.start,
            "Date Fin":      farmer.end,
            "Total Tonnes":  farmer.tonnage,
            "Pic de Recolte": "PIC" if PEAK_START <= date <= PEAK_END else "",
            "Note":          "RM-pre-alloc",
            "_is_rm":        True,
        })

rm_total_decl = sum(f.tonnage for f in rm_farmers)
rm_total_plan = sum(r["Tonnes/Jour"] for r in rm_fixed_days)
print(f"  RM: {len(rm_farmers)} farmers | déclaré={rm_total_decl:.0f}t | planifié={rm_total_plan:.0f}t | écart={rm_total_plan-rm_total_decl:+.1f}t")
print(f"  Non-RM: {len(nonrm_farmers)} farmers → OR-Tools")
print("\nBuilding OR-Tools model (non-RM uniquement)...")

SCALE = 10
MAX_SOLVE_SECONDS = 300

all_dates   = sorted({d for f in nonrm_farmers for d in f.window.keys()})
date_to_idx = {d: i for i, d in enumerate(all_dates)}
N_DATES     = len(all_dates)
N_FARM      = len(nonrm_farmers)

# ✅ Minimum journalier selon l'accessibilité du véhicule
# Un agriculteur PL ne peut pas avoir 5t/jour → minimum = capacité min du véhicule
VEH_MIN_TONS = {
    "SEMI":     30,   # SEMI seul = minimum 1 Semi = 30t (multiple de 30 obligatoire)
    "RM":       30,   # RM = idem SEMI, minimum 30t (1 Semi)
    "PL":       15,   # 1 PL = minimum 15t
    "PPL":       6,   # 1 PPL = minimum 6t
    "TRACTEUR":  9,   # 1 Tracteur = minimum 9t
    "PL/PPL":    6,   # au moins 1 PPL = 6t minimum
    "PL/SEMI":  15,   # au moins 1 PL = 15t minimum
    "TRC/PPL":   6,
    "TRC/PPL/PL": 6,
    "PL/PPL/SEMI": 6,
}

def _get_min_tons(farmer):
    """Retourne le minimum de tonnes par jour pour ce farmer selon son accessibilité.
    
    Règle Gafsa (SEMI/RM) : minimum = 30t (1 Semi exact = capacité physique)
    Tous les tonnages SEMI/RM doivent être multiples de 30t (30, 60, 90...)
    """
    acc = str(farmer.access).strip().upper()
    # Cas RM/SEMI → minimum 1 Semi = 30t (multiple de 30 obligatoire)
    if acc == "RM" or farmer.allowed_veh == ["SEMI"]:
        return 30  # 1 Semi minimum (pas 60 — la progression RM gère j1=30, j2=60, j3+=90)
    # Chercher dans le dictionnaire
    min_t = VEH_MIN_TONS.get(acc, 0)
    if min_t > 0:
        return min_t
    # Fallback : prendre le minimum de tous les véhicules autorisés
    mins = [FLEET_CAPACITY.get(v, (6, 25))[0] for v in farmer.allowed_veh]
    return min(mins) if mins else 6

model = cp_model.CpModel()

x = {}
for f_idx, farmer in enumerate(nonrm_farmers):
    for d_idx, date in enumerate(all_dates):
        if date in farmer.window:
            # ✅ Borne supérieure JOURNALIÈRE basée sur la COURBE de maturation
            # × 1.5 = tolérance modérée (un jour "double" = 1.5× la courbe naturelle)
            # ce qui empêche OR-Tools de mettre 100t là où la courbe prévoit 35t
            day_planned = farmer.window[date]
            _ub_day = max(int(day_planned * SCALE * 2.0), int(_get_min_tons(farmer) * SCALE))
            x[(f_idx, d_idx)] = model.NewIntVar(0, _ub_day, f"x_{f_idx}_{d_idx}")

        else:
            x[(f_idx, d_idx)] = model.NewConstant(0)

# Constraint 1: Tonnage per farmer (±2%)
for f_idx, farmer in enumerate(nonrm_farmers):
    total_scaled   = int(farmer.tonnage * SCALE)
    window_max     = int(sum(farmer.window.values()) * SCALE)
    effective      = min(total_scaled, window_max)
    # ✅ Tolérance serrée ±2% — évite le débordement ACHREF
    tolerance      = max(int(total_scaled * 0.02), SCALE)
    model.Add(sum(x[(f_idx, d)] for d in range(N_DATES)) >= effective - tolerance)
    model.Add(sum(x[(f_idx, d)] for d in range(N_DATES)) <= effective + tolerance)

# ── Pré-calcul des caps effectifs (UNE seule fois, avant les boucles) ──────
comm_effective_caps = {}
for comm in set(f.commercial for f in nonrm_farmers):
    cap_declared  = COMMERCIAL_CAPS.get(comm, 1200)
    total_comm_t  = sum(f.tonnage for f in nonrm_farmers if f.commercial == comm)
    cap_needed    = math.ceil(total_comm_t / N_DATES) if N_DATES > 0 else cap_declared
    cap_effective = max(cap_declared, cap_needed)
    comm_effective_caps[comm] = cap_effective
    if cap_effective > cap_declared:
        print(f"    ⚠️  {comm}: cap ajusté {cap_declared}→{cap_effective}t/j "
              f"(tonnage={total_comm_t:,.0f}t / {N_DATES}j = {cap_needed}t/j requis)")

# ✅ Pré-calcul des livraisons RM par (date, commercial) et (date, usine)
# OR-Tools doit soustraire ces tonnages des caps pour ne pas surcharger
_rm_by_date_comm = defaultdict(lambda: defaultdict(float))
_rm_by_date_usine = defaultdict(lambda: defaultdict(float))
for row in rm_fixed_days:
    d = row["Date"]
    _rm_by_date_comm[d][row["Commercial"]]  += row["Tonnes/Jour"]
    _rm_by_date_usine[d][row["Usine"]]      += row["Tonnes/Jour"]
print(f"  RM offset: {len(rm_fixed_days)} livraisons pré-allouées incluses dans les caps OR-Tools")

# Constraint 2: Commercial daily cap — UNIQUEMENT pendant le pic
# Caps s'appliquent du 1-15 juillet pour tous les commerciaux
# SAUF JILANI et KHALIL : caps du 1-12 juillet seulement
CAP_PERIOD_DEFAULT = (datetime.date(2026, 7, 1), datetime.date(2026, 7, 15))
CAP_PERIOD_SPECIAL = {
    "JILANI OBAY": (datetime.date(2026, 7, 1), datetime.date(2026, 7, 12)),
    "KHALIL":      (datetime.date(2026, 7, 1), datetime.date(2026, 7, 12)),
}
COMM_OVERFLOW_WEIGHT = 30
comm_overflows = []
for d_idx, date in enumerate(all_dates):
    by_comm = defaultdict(list)
    for f_idx, f in enumerate(nonrm_farmers):
        by_comm[f.commercial].append(x[(f_idx, d_idx)])
    for comm, vs in by_comm.items():
        cap_double  = COMMERCIAL_CAPS_DOUBLE.get(comm, comm_effective_caps.get(comm, 1200))
        # ✅ Soustraire les tonnes RM déjà planifiées ce jour pour ce commercial
        rm_offset   = _rm_by_date_comm[date].get(comm, 0.0)
        cap_restant = max(0.0, cap_double - rm_offset)
        cap_scaled  = int(cap_restant * SCALE)
        c_total = model.NewIntVar(0, int(cap_double * SCALE * 5), f"comm_{d_idx}_{hash(comm)%10000}")
        model.Add(c_total == sum(vs))
        c_ovf = model.NewIntVar(0, int(cap_double * SCALE * 5), f"comm_ovf_{d_idx}_{hash(comm)%10000}")
        model.Add(c_ovf >= c_total - cap_scaled)
        model.Add(c_ovf >= 0)
        comm_overflows.append(c_ovf * COMM_OVERFLOW_WEIGHT)

# Constraint 3: Factory daily cap — basé sur transport RÉEL confirmé + jokers
# cap_reel = min(cap_max_usine, transport_confirmé + jokers_alloués)
def get_real_cap(usine):
    """
    Cap journalier réel = min(cap_théorique, transport_confirmé + jokers).
    Plancher = cap_théorique/2 pour éviter INFEASIBLE si transport très bas.
    Si transport non encore complété, on autorise jusqu'au cap théorique.
    """
    # ✅ Utiliser FACTORY_LIMITS (limite planification) au lieu de FACTORY_CAPS
    cap_theorique  = FACTORY_LIMITS.get(usine, FACTORY_CAPS.get(usine, 2000))
    transport_conf = TRANSPORT_CONFIRMED.get(usine, {}).get("total", cap_theorique)
    joker_pl       = JOKER_ALLOC.get(usine, {}).get("PL_joker", 0)
    joker_ppl      = JOKER_ALLOC.get(usine, {}).get("PPL_joker", 0)
    cap_transport  = transport_conf + joker_pl + joker_ppl
    # Plancher = limite_planification pour garantir la faisabilité
    cap_reel = min(cap_theorique, max(cap_transport, cap_theorique))
    return cap_reel  # = limite tant que transport < limite

# Factory caps — SOUPLE (pénalité) appliqué TOUTE LA SAISON
# ✅ Une usine ne devrait jamais dépasser sa capacité, MAIS certaines usines
# (ex: COMOCAP) ont des agriculteurs dont les fenêtres se concentrent au pic
# et dépassent structurellement le cap (jusqu'à 856t pour un cap de 800t).
# Un cap DUR rend le modèle INFEASIBLE. On utilise donc un cap SOUPLE :
# le dépassement est autorisé mais FORTEMENT pénalisé dans l'objectif.
# Résultat: le solveur minimise les dépassements (reste à ~810-820 les pires
# jours au lieu de bloquer), et le plan reste réalisable.
FACTORY_CAP_START = datetime.date(2026, 7, 1)
FACTORY_CAP_END   = datetime.date(2026, 7, 15)
FACTORY_OVERFLOW_WEIGHT = 2000
factory_overflows = []
for d_idx, date in enumerate(all_dates):
    by_fact = defaultdict(list)
    for f_idx, f in enumerate(nonrm_farmers):
        by_fact[f.usine].append(x[(f_idx, d_idx)])
    for fact, vs in by_fact.items():
        # ✅ Soustraire les tonnes RM déjà planifiées ce jour pour cette usine
        cap_reel_brut = FACTORY_LIMITS.get(fact, get_real_cap(fact))
        cap_reel = max(20, cap_reel_brut)
        rm_offset_f  = _rm_by_date_usine[date].get(fact, 0.0)
        cap_net      = max(10, cap_reel - rm_offset_f)
        cap_scaled   = int(cap_net * SCALE)
        cap_dur      = int(cap_net * SCALE * 1.05)
        total_day = model.NewIntVar(0, cap_dur, f"fact_{fact}_{d_idx}")
        model.Add(total_day == sum(vs))
        model.Add(total_day <= cap_dur)
        overflow = model.NewIntVar(0, int(cap_reel * SCALE * 0.1), f"ovf_{fact}_{d_idx}")
        model.Add(overflow >= total_day - cap_scaled)
        model.Add(overflow >= 0)
        factory_overflows.append(overflow * FACTORY_OVERFLOW_WEIGHT)

print("  Limites planification vs cap usine (transport confirmé):")
for usine in FACTORY_LIMITS:
    cap_th  = FACTORY_CAPS[usine]        # cap physique (info)
    lim     = FACTORY_LIMITS[usine]      # limite planification (contrainte)
    cap_tr  = TRANSPORT_CONFIRMED.get(usine, {}).get("total", cap_th)
    jok     = JOKER_ALLOC.get(usine, {}).get("PL_joker", 0) + JOKER_ALLOC.get(usine, {}).get("PPL_joker", 0)
    cap_r   = get_real_cap(usine)
    manque  = lim - cap_tr - jok
    status  = f"✅ transport OK" if manque <= 0 else f"⚠️ manque {manque}t/j de bennes"
    print(f"    {usine:<12}: CAP={cap_th} | LIMITE={lim} | transport_dispo={cap_tr}+{jok}j | utilisé={cap_r} | {status}")

# Constraint 4 SUPPRIMÉE (peak avg cap) — cause INFEASIBLE
# Constraint 5 SUPPRIMÉE (lissage) — cause INFEASIBLE avec ub=1.0
# Raison: quand plusieurs agriculteurs terminent leur fenêtre en même temps,
# la capacité naturelle peut chuter de >20% entre deux jours consécutifs.
# Avec ub=1.0 (profil naturel), la contrainte de lissage est mathématiquement
# incompatible avec les chutes de fin de saison → INFEASIBLE instantané.
# L'objectif (minimiser déviation du plan) gère le lissage de façon souple.
print("  Constraints added: tonnage + commercial caps (pic only) + factory caps (pic only)")

# ============================================================
# OBJECTIVE: Minimize plan deviation + distance cost
# Weight: 70% minimize deviation from plan, 30% minimize transport distance
# ============================================================
DEVIATION_WEIGHT = 7   # higher = stick closer to original plan
DISTANCE_WEIGHT  = 3   # higher = favor closer usines

deviations = []
distance_costs = []

for f_idx, farmer in enumerate(nonrm_farmers):
    dist_norm = min(farmer.distance_km, 300)
    dist_coeff = int(dist_norm * DISTANCE_WEIGHT // 100) + 1
    for d_idx, date in enumerate(all_dates):
        if date in farmer.window:
            original = int(farmer.window[date] * SCALE)
            diff = model.NewIntVar(0, int(farmer.tonnage * SCALE), f"dev_{f_idx}_{d_idx}")
            model.AddAbsEquality(diff, x[(f_idx, d_idx)] - original)
            deviations.append(diff * DEVIATION_WEIGHT)
            distance_costs.append(x[(f_idx, d_idx)] * dist_coeff)

model.Minimize(sum(deviations) + sum(distance_costs)
               + sum(factory_overflows) + sum(comm_overflows))
print(f"  Objective: minimize (plan deviation ×{DEVIATION_WEIGHT}) + (distance ×{DISTANCE_WEIGHT})"
      f" + (factory overflow ×{FACTORY_OVERFLOW_WEIGHT}) + (commercial overflow ×{COMM_OVERFLOW_WEIGHT})")

# ============================================================
# STEP 4: Solve
# ============================================================
print(f"\nSolving (max 30s, 8 cores)...")
solver = cp_model.CpSolver()
solver.parameters.max_time_in_seconds   = 30
solver.parameters.num_search_workers    = 8
solver.parameters.log_search_progress   = True
solver.parameters.cp_model_presolve     = True
solver.parameters.linearization_level   = 2
solver.parameters.search_branching      = cp_model.PORTFOLIO_SEARCH

status = solver.Solve(model)
STATUS_NAMES = {
    cp_model.OPTIMAL:   "OPTIMAL",
    cp_model.FEASIBLE:  "FEASIBLE",
    cp_model.INFEASIBLE:"INFEASIBLE",
    cp_model.UNKNOWN:   "UNKNOWN",
}
print(f"  Status: {STATUS_NAMES.get(status, status)} | Time: {solver.WallTime():.1f}s")

if status not in (cp_model.OPTIMAL, cp_model.FEASIBLE):
    print("  No solution — applying capped fallback distribution")
    # Fallback intelligent: distribuer le tonnage en respectant les caps
    # au lieu de tout mettre dans la fenêtre originale sans contraintes
    from collections import defaultdict
    solution = {}
    # Compteurs journaliers pour vérifier les caps
    daily_comm  = defaultdict(lambda: defaultdict(float))   # [date][commercial]
    daily_fact  = defaultdict(lambda: defaultdict(float))   # [date][usine]

    for fi, farmer in enumerate(nonrm_farmers):
        remaining = farmer.tonnage
        # Distribution uniforme sur toute la fenêtre (pas concentrée sur premiers jours)
        # Chaque jour de la fenêtre reçoit le taux journalier naturel
        dates_avail = [(di, all_dates[di]) for di in range(N_DATES)
                       if farmer.window.get(all_dates[di], 0) > 0]
        
        # Passe 1: distribuer le taux naturel sur chaque jour disponible
        for di, date in dates_avail:
            if remaining <= 0:
                break
            in_peak_comm = (PEAK_START <= date <= datetime.date(2026, 7, 15))
            in_peak_fact = (PEAK_START <= date <= PEAK_END)
            if farmer.commercial in ("JILANI OBAY","KHALIL"):
                in_peak_comm = (PEAK_START <= date <= datetime.date(2026, 7, 12))
            
            # Taux naturel = tonnage / nb_jours_fenêtre
            natural_rate = farmer.window.get(date, 0)
            
            if in_peak_comm:
                cap_c   = comm_effective_caps.get(farmer.commercial, 9999)
                used_c  = daily_comm[date][farmer.commercial]
                avail_c = max(0, cap_c - used_c)
            else:
                avail_c = 9999
            
            if in_peak_fact:
                cap_f   = get_real_cap(farmer.usine)
                used_f  = daily_fact[date][farmer.usine]
                avail_f = max(0, cap_f - used_f)
            else:
                avail_f = 9999
            
            # Livrer exactement le taux naturel (pas plus, pas moins)
            daily_qty = min(avail_c, avail_f, remaining, natural_rate)
            if daily_qty > 0.1:
                solution[(fi, di)] = round(daily_qty, 2)
                daily_comm[date][farmer.commercial] += daily_qty
                daily_fact[date][farmer.usine]      += daily_qty
                remaining -= daily_qty
            else:
                solution[(fi, di)] = 0.0
        
        # Passe 2: placer le reste (si remaining > 0 à cause des caps du pic)
        if remaining > 0.5:
            for di, date in sorted(dates_avail, key=lambda x: daily_comm[x[1]][farmer.commercial]):
                if remaining <= 0.5:
                    break
                in_peak_comm = (PEAK_START <= date <= datetime.date(2026, 7, 15))
                in_peak_fact = (PEAK_START <= date <= PEAK_END)
                if farmer.commercial in ("JILANI OBAY","KHALIL"):
                    in_peak_comm = (PEAK_START <= date <= datetime.date(2026, 7, 12))
                
                if in_peak_comm:
                    cap_c  = comm_effective_caps.get(farmer.commercial, 9999)
                    avail_c = max(0, cap_c - daily_comm[date][farmer.commercial])
                else:
                    avail_c = 9999
                
                if in_peak_fact:
                    cap_f  = get_real_cap(farmer.usine)
                    avail_f = max(0, cap_f - daily_fact[date][farmer.usine])
                else:
                    avail_f = 9999
                
                extra = min(avail_c, avail_f, remaining)
                if extra > 0.1:
                    cur = solution.get((fi, di), 0.0)
                    solution[(fi, di)] = round(cur + extra, 2)
                    daily_comm[date][farmer.commercial] += extra
                    daily_fact[date][farmer.usine]      += extra
                    remaining -= extra

        # Remplir les cases non assignées à 0
        for di in range(N_DATES):
            if (fi, di) not in solution:
                solution[(fi, di)] = 0.0
else:
    solution = {(fi, di): solver.Value(x[(fi, di)]) / SCALE
                    for fi in range(N_FARM) for di in range(N_DATES)}



# ============================================================
# STEP 5: Build output
# ============================================================
print("\nBuilding output tables...")

def choose_vehicles(tons, allowed_raw, usine=None, region=None, semi_coeff=1.0, rm_day_rank=0):
    """
    VERSION V5 — Logique simplifiée et correcte.
    
    Principe:
    - L'accessibilité (PL/SEMI, PL/PPL...) = véhicules PHYSIQUEMENT possibles sur ce terrain
    - On choisit UN seul type de véhicule pour chaque livraison (le plus adapté au tonnage)
    - Exceptionnellement 2 types pour COMOCAP (TRACTEUR + transport principal)
    - Pas de découpage proportionnel artificiel en 4-5 types
    
    rm_day_rank : rang du jour dans les jours de livraison de cet agriculteur RM
                  0 = 1er jour → 2 Semi (60t)
                  1 = 2ème jour → 3 Semi (90t)
                  2+ = 3ème+ jour → 4 Semi (120t)
    
    Logique de sélection:
    1. Regarder la préférence usine (SICAM préfère SEMI, ELFALLEH préfère PPL...)
    2. Si le véhicule préféré est accessible → l'utiliser
    3. Sinon → choisir le meilleur véhicule accessible pour le tonnage
    4. _alloc() répartit en voyages en respectant mn ≤ charge ≤ mx
    """
    # Normaliser allowed
    _norm = {"PETIT POILOUR":"PPL","POILOUR":"PL"}
    allowed = list(dict.fromkeys(_norm.get(v,v) for v in allowed_raw))
    if not allowed:
        allowed = ["PL"]

    def _alloc_real(veh, qty, usine_name=None):
        """
        Alloue qty tonnes en utilisant les vraies bennes disponibles.
        ✅ SEMI seul (ACHREF Gafsa): bennes entières uniquement
           Reste < 1 benne → ignoré (livré le lendemain dans la fenêtre)
           → plus jamais de SEMI à 10t, 20t ou 40t
        """
        if qty <= 0: return []
        
        # TRACTEUR: capacité fixe ~10t
        if veh == "TRACTEUR":
            return [{"vehicle": "TRACTEUR", "trips": 1,
                     "tons_each": round(min(10.0, qty), 1), "real_cap": 10.0}]
        
        mn_veh, mx_veh = FLEET_CAPACITY.get(veh, (7, 25))
        
        # ✅ SEMI : capacité TOUJOURS 30t (pas de réduction par coeff)
        # Note: semi_coeff était utilisé pour réduire artificiellement la capacité
        # → SUPPRIMÉ car un Semi physique transporte 30t qu'il aille à 10km ou 200km
        
        # ✅ SEMI seul (ACHREF/RM/Gafsa): TOUJOURS 30t par voyage
        #    Bennes entières uniquement, plus de coefficient bizarre
        #    Si qty < 27t → reporté au lendemain (jamais de voyage partiel)
        _semi_only = (allowed == ["SEMI"])
        if veh == "SEMI" and _semi_only:
            SEMI_CAP = 30.0   # ✅ TOUJOURS 30t (capacité physique réelle)
            SEMI_MIN = 27.0   # seuil minimum pour 1 voyage (90% de la benne)
            if qty < SEMI_MIN:
                return []   # trop petit → reporté au lendemain
            nb = max(1, int(round(qty / SEMI_CAP)))  # nb voyages entiers de 30t
            return [{"vehicle": "SEMI", "trips": nb, "tons_each": SEMI_CAP,
                     "real_cap": SEMI_CAP}]
        
        # Récupérer les bennes réelles de cette usine
        real_caps = []
        if usine_name and usine_name in REAL_FLEET:
            real_caps = list(REAL_FLEET[usine_name].get(veh, []))
            if not real_caps and usine_name == "TUCAL":
                real_caps = list(REAL_FLEET.get("BOURAK", {}).get(veh, []))
        
        # ✅ SEMI: capacités réelles utilisées telles quelles (pas de × coeff)
        # Le tableau transport_etat_final.xlsx donne les vraies capacités (30t typique)
        
        if not real_caps:
            return _alloc_theory(veh, qty)
        
        # ✅ ALGORITHME SOLDE DYNAMIQUE — ZÉRO BENNE FICTIVE
        # Si le reste ne peut pas remplir une vraie benne du tableau → annoter comme solde
        # Affichage : PL(25t)+1.5 solde | PL(25t)+1.5 solde
        # MAX_SOLDE = 4t fixe OU dynamique si reste < plus petite benne non utilisée

        if not real_caps:
            return _alloc_theory(veh, qty)

        MAX_SOLDE_FIXE = 4.0   # solde max fixe annoté sur bennes existantes

        result    = []
        remaining = round(qty, 1)

        for idx, cap in enumerate(real_caps):
            if remaining <= 0:
                break
            # ✅ CLE : calculer la plus petite benne ENCORE DISPONIBLE (non utilisée)
            # Si le reste est trop petit pour remplir cette benne → l'annoter comme solde
            used = len(result)
            caps_restantes = real_caps[used:]
            min_dispo = min(caps_restantes) if caps_restantes else cap

            # Annoter comme solde si :
            # - Solde fixe ≤ 4t, OU
            # - Reste < min benne disponible (impossible de remplir une vraie benne)
            if result and (remaining <= MAX_SOLDE_FIXE or remaining < min_dispo * 0.5):
                break   # garder remaining comme solde

            actual_load = min(cap, remaining)
            result.append({
                "vehicle":  veh,
                "trips":    1,
                "tons_each": round(actual_load, 1),
                "real_cap": round(cap, 1),
                "solde":    0.0,
            })
            remaining = round(remaining - actual_load, 1)

        # Distribuer le solde sur les bennes existantes
        if remaining > 0.05 and result:
            nb = len(result)
            solde_base = math.floor(remaining / nb * 10) / 10
            solde_last = round(remaining - solde_base * (nb - 1), 1)
            for i in range(nb):
                result[i]["solde"] = solde_last if i == nb - 1 else solde_base
            remaining = 0

        return result if result else _alloc_theory(veh, qty)
        
        return result if result else _alloc_theory(veh, qty)
    
    def _alloc_theory(veh, qty):
        """Fallback: répartition théorique si flotte réelle non disponible."""
        if qty <= 0: return []
        mn, mx = FLEET_CAPACITY.get(veh, (7, 25))
        if qty <= mx:
            return [{"vehicle": veh, "trips": 1, "tons_each": round(qty, 2)}]
        trips = math.ceil(qty / mx)
        each  = qty / trips
        while each < mn and trips > 1:
            trips -= 1; each = qty / trips
        base  = int(qty // trips)
        extra = int(qty - base * trips)
        entries = []
        if extra:     entries.append({"vehicle": veh, "trips": extra,   "tons_each": round(base+1, 2)})
        if trips-extra: entries.append({"vehicle": veh, "trips": trips-extra, "tons_each": round(base, 2)})
        return entries
    
    def _alloc(veh, qty):
        """Point d'entrée: utilise vraies bennes si disponibles."""
        return _alloc_real(veh, qty, usine_name=usine)

    def _best_for_tons(qty, candidates):
        """
        Parmi les candidats accessibles, choisit celui qui gère qty le plus proprement.
        Scores:
          0 = 1 voyage parfait (mn ≤ qty ≤ mx)
          1 = multi-voyages propres (chaque voyage dans [mn, mx])
          2 = 1 voyage légèrement sous-chargé (proche du mn)
          3 = sous-chargé significatif (< 80% du mn)
          4 = surcharge
        Priorité absolue: ne jamais mettre un SEMI pour < 27t si PPL ou PL peut faire mieux
        """
        best, best_score, best_ratio, best_trips = None, 999, 0, 999
        for veh in candidates:
            mn_v, mx_v = FLEET_CAPACITY.get(veh, (7, 25))
            ratio = min(qty / mn_v, 1.0)
            if mn_v <= qty <= mx_v:
                score = 0
                n_trips = 1
            elif qty > mx_v:
                trips = math.ceil(qty / mx_v)
                each  = qty / trips
                if each >= mn_v:
                    score = 1
                    n_trips = trips
                else:
                    t2 = trips
                    while each < mn_v and t2 > 1:
                        t2 -= 1
                        each = qty / t2
                    score = 2 if each >= mn_v else 4
                    n_trips = t2
            elif qty >= mn_v * 0.8:
                score = 3
                n_trips = 1
            else:
                score = 5
                n_trips = 1
            # Priorité: score bas > ratio élevé > moins de voyages
            if (score < best_score or
                (score == best_score and ratio > best_ratio) or
                (score == best_score and ratio == best_ratio and n_trips < best_trips)):
                best_score, best, best_ratio, best_trips = score, veh, ratio, n_trips
        return best

    # ── CAS SPÉCIAL : SEMI seul / RM (Gafsa/Kasserine — ACHREF) ─────────────
    # Règle RM progressive : le tonnage OR-Tools détermine le nb de Semi.
    # Min 60t/j (2 Semi) appliqué dans la fenêtre de maturation de l'agriculteur.
    # choose_vehicles affiche simplement nb = round(tons/30).
    if allowed == ["SEMI"]:
        SEMI_CAP = 30.0
        if tons <= 0:
            return []
        nb_semi = max(1, int(round(tons / SEMI_CAP)))
        return [{"vehicle": "SEMI", "trips": nb_semi,
                 "tons_each": SEMI_CAP, "real_cap": SEMI_CAP, "solde": 0.0}]

    # ── PRÉFÉRENCES PAR USINE ───────────────────────────────────────────
    # Ordre de préférence des véhicules selon l'usine
    # Le système choisit le PREMIER véhicule de la liste qui est accessible
    # ── PRÉFÉRENCES PAR USINE (ajustées selon la DISTANCE) ──────────────
    # Le choix PL vs SEMI dépend de la distance:
    #   - Proche (Cap Bon, Bouficha, Nord): PL/PPL suffisent (FEDI, MAKKI)
    #   - Loin (Gafsa, Sidi Bouzid, Kairouan): SEMI préféré (ACHREF, volume + distance)
    _is_long_distance = (region and str(region).strip().upper() in
                         {"GAFSA / KASSRINE", "GAFSA/KASSRINE", "SIDI BOUZID", "KAIROUAN"})

    if _is_long_distance:
        # Longue distance → SEMI en premier (transport massif efficace)
        USINE_PREFS = {
            "SICAM":    ["SEMI", "PL",  "PPL"],
            "TUCAL":    ["SEMI", "PL",  "PPL"],
            "COMOCAP":  ["SEMI", "PL",  "PPL"],
            "ABIDA":    ["PL",   "SEMI","PPL"],
            "ELFALLEH": ["PPL",  "PL",  "SEMI"],  # ✅ ELFALLEH toujours PPL en priorité
        }
    else:
        # Courte distance → PL/PPL en premier (Cap Bon, Nord, Bouficha)
        USINE_PREFS = {
            "SICAM":    ["PL",   "PPL", "SEMI"],
            "TUCAL":    ["PL",   "PPL", "SEMI"],
            "COMOCAP":  ["PL",   "PPL", "SEMI"],
            "ABIDA":    ["PL",   "SEMI","PPL"],
            "ELFALLEH": ["PPL",  "PL",  "SEMI"],  # ✅ ELFALLEH toujours PPL en priorité
        }
    prefs = USINE_PREFS.get(usine, ["PL", "PPL", "SEMI"])
    # ✅ Filtrer prefs par allowed (accessibilité de l'agriculteur)
    prefs_allowed = [v for v in prefs if v in allowed]
    if not prefs_allowed:
        prefs_allowed = allowed  # fallback sur allowed seul

    # ── CAS COMOCAP : TRACTEUR fixe (10t) + transport principal ─────────
    # TRACTEUR = flotte propre COMOCAP — UNIQUEMENT pour CAP BON 1
    TRACTEUR_REGIONS = {"CAP BON 1"}

    # Région effective: geo_region si valide, sinon fallback sur region Supabase
    _r = str(region or "").strip().upper()
    _region_effective = _r if _r not in ("", "AUTRE") else ""
    use_tracteur = (_region_effective in TRACTEUR_REGIONS) and ("TRACTEUR" in allowed or "TRC" in str(allowed_raw).upper())
    
    if usine == "COMOCAP":
        result = []
        trac_mn, trac_mx = FLEET_CAPACITY.get("TRACTEUR", (9, 11))
        
        if use_tracteur:
            # TRACTEUR uniquement pour CAP BON 1
            trac_tons = min(trac_mx, max(trac_mn, round(tons * 0.14)))
            result.append({"vehicle": "TRACTEUR", "trips": 1,
                           "tons_each": round(trac_tons, 2)})
            remaining = round(tons - trac_tons, 2)
        else:
            # Autres régions → pas de TRACTEUR, tout en PL/PPL/SEMI selon allowed
            remaining = tons
        if remaining > 0:
            # ✅ Choisir le meilleur véhicule PARMI ceux autorisés (allowed)
            # Ordre de préférence pour COMOCAP: PL > PPL > SEMI
            candidates = [v for v in ["PL", "PPL", "SEMI"] if v in allowed]
            if not candidates:
                candidates = allowed  # fallback
            main_veh = _best_for_tons(remaining, candidates)
            if not main_veh:
                main_veh = candidates[0] if candidates else "PL"
            result.extend(_alloc(main_veh, remaining))
        return result if result else _alloc(allowed[0] if allowed else "PL", tons)

    # ── CAS GÉNÉRAL : 1 véhicule principal ──────────────────────────────
    # ✅ Choisir le meilleur véhicule PARMI ceux autorisés (allowed)
    primary = _best_for_tons(tons, prefs_allowed)
    if not primary:
        # Essayer tous les véhicules autorisés
        primary = _best_for_tons(tons, allowed)
    if not primary:
        primary = allowed[0] if allowed else "PL"  # dernier recours
    
    # ✅ Vérification: si le véhicule choisi est très sous-chargé,
    # essayer un véhicule plus petit (mais TOUJOURS dans allowed)
    if primary and primary in FLEET_CAPACITY:
        mn_p, mx_p = FLEET_CAPACITY[primary]
        if tons < mn_p * 0.8:
            smaller = {
                "SEMI": ["PL", "PPL"],
                "PL":   ["PPL"],
                "PPL":  [],
            }
            for alt in smaller.get(primary, []):
                if alt in allowed:   # ✅ TOUJOURS vérifier allowed
                    alt_mn, alt_mx = FLEET_CAPACITY[alt]
                    if tons >= alt_mn * 0.5:
                        primary = alt
                        break

    result = _alloc(primary, tons)
    
    # Vérification minimale
    if not result:
        result = [{"vehicle": primary, "trips": 1, "tons_each": round(tons, 2)}]
    
    # ✅ VÉRIFICATION FINALE: aucun SEMI si non autorisé
    # Sécurité supplémentaire contre les edge cases
    if "SEMI" not in allowed:
        cleaned = []
        for v in result:
            if v.get("vehicle") == "SEMI":
                # Convertir SEMI → PL (plus proche en capacité)
                best_alt = "PL" if "PL" in allowed else ("PPL" if "PPL" in allowed else None)
                if best_alt:
                    cleaned.extend(_alloc(best_alt, v["trips"] * v.get("tons_each", 0)))
                # sinon ignorer (absorbe dans autres voyages)
            else:
                cleaned.append(v)
        if cleaned:
            result = cleaned
    
    return result

all_days = []
# ✅ CONSOLIDATION : regrouper les parcelles du même agriculteur livrant
# à la même usine le même jour pour optimiser le transport
# (ex: 2 parcelles HAFEDH MOSBE × 10t = 1 voyage PL de 20t au lieu de 2)
from collections import defaultdict as _dd_cons
consolidated = _dd_cons(lambda: {"tons": 0, "farmer": None, "parcelles": []})
for f_idx, farmer in enumerate(nonrm_farmers):
    for d_idx, date in enumerate(all_dates):
        tons = solution[(f_idx, d_idx)]
        if tons <= 0.5: continue
        # Clé : nom agriculteur + usine + date (= un envoi unique)
        key = (farmer.name.strip().upper(), farmer.usine, date)
        consolidated[key]["tons"] += round(tons, 1)
        consolidated[key]["farmer"] = farmer  # garder la dernière référence
        consolidated[key]["parcelles"].append(f_idx)

# ✅ MINIMUM TONNAGE PAR ACCESSIBILITÉ
# Un véhicule ne peut pas livrer moins que sa capacité minimale physique
# → si le tonnage brut est trop petit, on l'accumule avec les jours adjacents

# ✅ PRÉ-CALCUL pour RM progressif : rang du jour de livraison par agriculteur
_rm_delivery_days = {}
for (nom, usine, date), data in consolidated.items():
    f = data["farmer"]
    if f.allowed_veh == ["SEMI"] or str(f.access).upper() == "RM":
        key_rm = (nom, usine)
        if key_rm not in _rm_delivery_days:
            _rm_delivery_days[key_rm] = []
        _rm_delivery_days[key_rm].append(date)
for key_rm in _rm_delivery_days:
    _rm_delivery_days[key_rm] = sorted(_rm_delivery_days[key_rm])

# ✅ ACCUMULATION des petits tonnages pour éviter les jours vides
# Si tons_brut < min_tons pour ce véhicule → accumuler sur le jour suivant
# Ex: PL min=15t, OR-Tools donne 8t → accumuler avec le lendemain
_accumulator = {}  # {(nom, usine): tonnes_accumulées}

# Trier les livraisons par (agriculteur, date) pour traitement chronologique
consolidated_sorted = sorted(
    consolidated.items(),
    key=lambda x: (x[0][0], x[0][2])  # trier par (nom, date)
)

# Génération des all_days à partir des envois consolidés
for (nom, usine, date), data in consolidated.items():
    tons_brut = data["tons"]
    farmer    = data["farmer"]
    n_parcelles = len(data["parcelles"])
    
    # ✅ ARRONDI à la DIZAINE la plus proche MAIS pour RM → multiple de 30t obligatoire
    if farmer.allowed_veh == ["SEMI"] or str(farmer.access).upper() == "RM":
        # RM/SEMI : arrondir au multiple de 30t INFÉRIEUR (floor)
        tons = int(tons_brut / 30) * 30
        if tons == 0 and tons_brut >= 15:
            tons = 30
    else:
        tons = int(round(round(tons_brut, 1) / 10)) * 10
        _min_t_agri = _get_min_tons(farmer)
        if tons == 0 and tons_brut > 0:
            tons = _min_t_agri
        elif 0 < tons < _min_t_agri:
            tons = _min_t_agri
    if tons <= 0:
        continue

    # Calculer le rang du jour pour les agriculteurs RM (montée progressive Semi)
    _rm_rank = 0
    if farmer.allowed_veh == ["SEMI"] or str(farmer.access).upper() == "RM":
        key_rm = (nom, usine)
        days_list = _rm_delivery_days.get(key_rm, [])
        try:
            _rm_rank = days_list.index(date)  # 0=jour1, 1=jour2, 2+=jour3
        except ValueError:
            _rm_rank = 0
        # ✅ Règle RM progressive : appliquer minimum selon rang du jour
        # j1 → min 60t (2 Semi), j2 → min 90t (3 Semi), j3+ → min 120t (4 Semi)
        if _rm_rank == 0:
            _rm_min = 60
        elif _rm_rank == 1:
            _rm_min = 90
        else:
            _rm_min = 120
        # Forcer le minimum si le tonnage OR-Tools le permet (ne pas dépasser tons_brut × 1.5)
        if tons < _rm_min and tons_brut >= _rm_min * 0.5:
            tons = _rm_min

    vehicles = choose_vehicles(tons, farmer.allowed_veh,
                              usine=farmer.usine,
                              region=farmer.geo_region if farmer.geo_region not in ("", "AUTRE")
                                     else farmer.region,
                              semi_coeff=getattr(farmer, "semi_coeff", 1.0),
                              rm_day_rank=_rm_rank)

    # ✅ Pour SEMI/RM : recalculer le tonnage réel (nb_semi × 30t)
    # Le tonnage OR-Tools peut être 100t mais choose_vehicles impose 1/2/3 Semi
    # → on corrige Tonnes/Jour pour afficher le vrai tonnage livré (30, 60 ou 90t)
    if (farmer.allowed_veh == ["SEMI"] or str(farmer.access).upper() == "RM") and vehicles:
        tons_rm_reel = sum(v.get("trips", 1) * v.get("tons_each", 30) for v in vehicles)
        tons = int(round(tons_rm_reel / 30)) * 30  # garantit multiple de 30
    veh_parts = []
    for v in vehicles:
        if v.get('tons_each', 0) <= 0:
            continue
        veh   = v['vehicle']
        load  = v.get('tons_each', 0)
        n     = v.get('trips', 1)
        solde = v.get('solde', 0.0)  # ✅ solde additionnel à afficher

        # ✅ Afficher la charge réelle livrée (tons_each) + solde si > 0
        # Exemples d'affichage :
        #   PL(25t)          → charge normale, benne du tableau
        #   PL(25t)+1.5 solde → benne du tableau + 1.5t de solde réparti
        disp_load = int(load) if float(load) == int(float(load)) else round(float(load), 1)

        if n > 1:
            part = f"{veh} x{n}({disp_load}t)"
        else:
            part = f"{veh}({disp_load}t)"

        # Ajouter annotation solde si non nul
        if solde and solde > 0.05:
            disp_solde = int(solde) if float(solde) == int(float(solde)) else round(float(solde), 1)
            part = f"{part}+{disp_solde} solde"

        veh_parts.append(part)
    
    # ✅ Fallback intelligent : utiliser le PREMIER véhicule AUTORISÉ
    # (pas un POILOUR hardcodé qui violerait l'accessibilité de l'agriculteur)
    if vehicles:
        veh_type = vehicles[0]["vehicle"]
    else:
        # Aucun véhicule alloué (tonnage trop petit pour Semi par ex.)
        # → utiliser le premier véhicule autorisé pour ce farmer
        _allowed = farmer.allowed_veh if farmer.allowed_veh else ["PL"]
        veh_type = _allowed[0]
    
    # Idem pour veh_str : fallback selon accessibilité
    if not veh_parts:
        _fallback_veh = veh_type
        _fallback_cap = 30 if _fallback_veh == "SEMI" else (20 if _fallback_veh == "PL" else 10)
        veh_str = f"{_fallback_veh}({_fallback_cap}t)"
    else:
        veh_str = " | ".join(veh_parts)
    
    trips    = sum(v["trips"] for v in vehicles) if vehicles else 1
    dist_km  = farmer.distance_km if farmer.distance_km < 999 else 0
    note     = f"{n_parcelles} parcelles consolidées" if n_parcelles > 1 else "AI-optimise"
    
    all_days.append({
        "Commercial":    farmer.commercial,
        "Agriculteur":   farmer.name,
        "Usine":         usine,
        "Region":        farmer.geo_region,
        "Zone":          farmer.zone,
        "Accessibilite": farmer.access,
        "Date":          date,
        "Tonnes/Jour":   tons,
        "Type Vehicule": veh_type,
        "Vehicules":     veh_str,
        "Nb Voyages":    trips,
        "Distance km":   dist_km,
        "Date Debut":    farmer.start,
        "Date Fin":      farmer.end,
        "Total Tonnes":  farmer.tonnage,
        "Pic de Recolte":"PIC" if PEAK_START <= date <= PEAK_END else "",
        "Note":          note,
        "_is_rm":        False,
    })

# ✅ FUSION RM pré-alloués + non-RM OR-Tools
all_days.extend(rm_fixed_days)
print(f"  Lignes: {len(all_days)-len(rm_fixed_days)} non-RM + {len(rm_fixed_days)} RM = {len(all_days)} total")

# ── POST-TRAITEMENT TONNAGE: récupérer les tonnes perdues par l'arrondi ─
# Si total planifié d'un agriculteur < total déclaré × 98%,
# augmenter certains jours de 10t pour combler le manque
_agri_plan_total = {}
for row in consolidated.values():
    nom   = row["farmer"].name
    usine = list(consolidated.keys())[0][1] if row["farmer"] else ""
    # Accumuler les tonnes planifiées par agriculteur
    _key = row["farmer"].name
    if _key not in _agri_plan_total:
        _agri_plan_total[_key] = {"planned": 0, "declared": row["farmer"].tonnage}
    _agri_plan_total[_key]["planned"] += row["tons"]

# ANCIEN code remplacé par la consolidation par envoi unique

# ✅ MARQUER LES JOURS DOUBLES (livraison > cap normal du commercial)
print("  Détection jours doubles (livraisons > cap normal du commercial)...")
from collections import defaultdict as _dd_dbl

# Cap NORMAL par commercial (sans jour double)
# JILANI = 100t/j (6965t / 73j) — pas de jours doubles
CAPS_NORMAL = {
    "FEDI":            850,
    "MAKKI BEN SALAH": 850,
    "KHALIL":          900,
    "ACHREF AJLANI":   450,
    "JILANI OBAY":     100,
}

# Calculer tonnage par (commercial, date) → si > cap normal = jour double
day_tonnage = _dd_dbl(float)
for row in all_days:
    day_tonnage[(row["Commercial"], row["Date"])] += row["Tonnes/Jour"]

nb_jours_doubles = 0
for row in all_days:
    # ✅ Ne JAMAIS écraser la note RM-pre-alloc (sinon la correction post-traitement casse)
    if row.get("_is_rm", False):
        continue
    key = (row["Commercial"], row["Date"])
    cap_norm = CAPS_NORMAL.get(row["Commercial"], 1000)
    if day_tonnage[key] > cap_norm * 1.05:
        row["Note"] = "AI-optimise + JOUR DOUBLE"
        nb_jours_doubles += 1
    else:
        row["Note"] = "AI-optimise"

jours_doubles_uniques = set()
for row in all_days:
    if "JOUR DOUBLE" in row.get("Note",""):
        jours_doubles_uniques.add((row["Commercial"], row["Date"]))
print(f"    → {len(jours_doubles_uniques)} jours doubles identifiés ({nb_jours_doubles} livraisons concernées)")

# ✅ POST-TRAITEMENT CORRECTION TONNAGE : total planifié = total déclaré
# CORRECTION CLÉE : ajuster par AGRICULTEUR (pas par commercial) et seulement
# sur le DERNIER JOUR de chaque agriculteur → évite les sauts brusques 100→10→100
print("  Post-traitement correction tonnage (arrondi)...")
from collections import defaultdict as _dd_ton

# Tonnages déclarés par commercial
_decl_by_comm  = {}
for f in farmers:
    _decl_by_comm[f.commercial] = _decl_by_comm.get(f.commercial, 0.0) + f.tonnage

# Tonnages déclarés par (commercial, agriculteur)
_decl_by_agri = {}
for f in farmers:
    key = (f.commercial, f.name)
    _decl_by_agri[key] = _decl_by_agri.get(key, 0.0) + f.tonnage

# Tonnages planifiés par (commercial, agriculteur)
_plan_by_agri  = _dd_ton(float)
for row in all_days:
    key = (row["Commercial"], row["Agriculteur"])
    _plan_by_agri[key] += row["Tonnes/Jour"]

_plan_by_comm  = _dd_ton(float)
for row in all_days:
    _plan_by_comm[row["Commercial"]] += row["Tonnes/Jour"]

print("    Vérification avant correction:")
for comm, decl in _decl_by_comm.items():
    plan = _plan_by_comm.get(comm, 0)
    diff = plan - decl
    print(f"      {comm:<20}: déclaré={decl:>8.0f}t | planifié={plan:>8.0f}t | diff={diff:+.0f}t")

# Correction par AGRICULTEUR sur son DERNIER JOUR
# → DÉSACTIVÉE pour non-SEMI : la correction empire systématiquement
#   les résultats (transforme -94t en +798t). On garde uniquement les RM
#   qui sont déjà exacts par pré-allocation.
_corrections = 0
for (comm, agri), decl in _decl_by_agri.items():
    # ✅ SKIP correction pour TOUS — la correction empire les résultats
    # Les RM sont déjà exacts par pré-allocation
    # Les non-RM gardent les valeurs OR-Tools (écart total < 0.3%)
    continue

    # --- Code mort ci-dessous, conservé pour référence ---
    _farmer_ref = next((f for f in farmers if f.commercial == comm and f.name == agri), None)
    if _farmer_ref and is_rm_farmer(_farmer_ref):
        continue
    plan = sum(r["Tonnes/Jour"] for r in all_days
               if r["Commercial"] == comm and r["Agriculteur"] == agri)
    diff = round(plan - decl, 1)
    if abs(diff) < 0.5:
        continue
    agri_indices = [i for i, r in enumerate(all_days)
                    if r["Commercial"] == comm and r["Agriculteur"] == agri
                    and not r.get("_is_rm", False)]
    if not agri_indices:
        continue
    agri_indices.sort(key=lambda i: all_days[i]["Date"])

    _min_t = _get_min_tons(_farmer_ref) if _farmer_ref else 10
    _is_semi_rm = False  # déjà skipé au dessus pour les vrais RM

    remaining_diff = diff  # positif = trop planifié, négatif = trop peu

    if _is_semi_rm and diff > 0:
        # SEMI/RM avec EXCÈS : supprimer des jours entiers (multiples de 30t)
        for idx in reversed(agri_indices):
            if remaining_diff < 15:
                break
            current_day = all_days[idx]["Tonnes/Jour"]
            if current_day <= 0:
                continue
            reduce = min(current_day, int(remaining_diff / 30) * 30)
            if reduce <= 0:
                reduce = current_day
            new_val_30 = int(round((current_day - reduce) / 30)) * 30
            all_days[idx]["Tonnes/Jour"] = max(0, new_val_30)
            remaining_diff -= reduce
            _corrections += 1
    elif _is_semi_rm and diff < 0:
        # SEMI/RM avec DÉFICIT : augmenter le dernier jour
        last_idx = agri_indices[-1]
        current  = all_days[last_idx]["Tonnes/Jour"]
        new_val  = int(round((current - diff) / 30)) * 30
        all_days[last_idx]["Tonnes/Jour"] = max(30, new_val)
        _corrections += 1
    else:
        # Non SEMI/RM : correction sur le dernier jour
        last_idx = agri_indices[-1]
        current  = all_days[last_idx]["Tonnes/Jour"]
        new_val  = round(current - diff, 1)
        if new_val >= _min_t:
            all_days[last_idx]["Tonnes/Jour"] = new_val
            _corrections += 1
        elif new_val > 0:
            all_days[last_idx]["Tonnes/Jour"] = max(0, new_val)
            _corrections += 1
        elif len(agri_indices) >= 2:
            prev_idx  = agri_indices[-2]
            prev_curr = all_days[prev_idx]["Tonnes/Jour"]
            prev_new  = round(prev_curr - diff, 1)
            if prev_new >= _min_t:
                all_days[prev_idx]["Tonnes/Jour"] = prev_new
                _corrections += 1

print(f"    → {_corrections} ajustement(s) appliqués")
# Recalculer pour vérifier
_plan_after = _dd_ton(float)
for row in all_days:
    _plan_after[row["Commercial"]] += row["Tonnes/Jour"]
for comm, decl in _decl_by_comm.items():
    plan = _plan_after.get(comm, 0)
    print(f"      {comm:<20}: déclaré={decl:>8.0f}t | planifié après={plan:>8.0f}t | diff={plan-decl:+.0f}t")

# ✅ POST-TRAITEMENT TRACTEUR COMOCAP : max 10 voyages/jour
# DOIT être fait AVANT result_df pour aussi corriger transport_rows
print("  Post-traitement TRACTEUR COMOCAP (max 10 voyages/jour)...")
TRACTEUR_DAILY_LIMIT = 10  # max 10 voyages × 10t = 100t/jour

# Indexer all_days par date
from collections import defaultdict as _dd
by_date = _dd(list)
for i, row in enumerate(all_days):
    by_date[row["Date"]].append(i)

nb_corrections = 0
for date_val, indices in by_date.items():
    # Filtrer COMOCAP avec TRACTEUR
    comocap_trac_idx = [
        i for i in indices
        if all_days[i]["Usine"] == "COMOCAP" 
        and "TRACTEUR" in str(all_days[i].get("Vehicules", ""))
    ]
    if len(comocap_trac_idx) <= TRACTEUR_DAILY_LIMIT:
        continue
    # Trop de voyages TRACTEUR — convertir les surplus en PPL
    # Garder les 10 PREMIERS par ordre d'index (déterministe)
    # Convertir le reste
    sorted_by_tons = sorted(comocap_trac_idx, key=lambda i: all_days[i].get("Tonnes/Jour", 0))
    to_convert = sorted_by_tons[TRACTEUR_DAILY_LIMIT:]
    for idx in to_convert:
        ton = all_days[idx].get("Tonnes/Jour", 0)
        # PPL capacité 7-12t → calculer voyages
        n_voyages = max(1, math.ceil(ton / 12))
        ton_each  = round(ton / n_voyages, 1)
        all_days[idx]["Vehicules"]    = f"PPL x{n_voyages}({ton_each}t)"
        all_days[idx]["Type Vehicule"] = "PPL"
        all_days[idx]["Nb Voyages"]    = n_voyages
        nb_corrections += 1
print(f"    → {nb_corrections} agriculteur(s) TRACTEUR convertis en PPL pour respecter cap 10/jour")

result_df = pd.DataFrame(all_days).sort_values(
    ["Date","Commercial","Agriculteur"]).reset_index(drop=True)

# Normalize Region column in result
REGION_NORM_RESULT = {
    "NABEUL":"CAP BON 2","nabeul":"CAP BON 2",
    "BEJA":"NORD","beja":"NORD",
    "MANOUBA":"NORD","manouba":"NORD",
    "GAFSA":"GAFSA / KASSRINE","KASSRINE":"GAFSA / KASSRINE",
    "CAPB1":"CAP BON 1","CAP B1":"CAP BON 1",
    "CAPB2":"CAP BON 2","CAP B2":"CAP BON 2",
}
if "Region" in result_df.columns:
    result_df["Region"] = result_df["Region"].replace(REGION_NORM_RESULT)
elif "Région" in result_df.columns:
    result_df["Région"] = result_df["Région"].replace(REGION_NORM_RESULT)
print(f"  Planning rows: {len(result_df)}")

# Transport rows
# Mapping interne → colonnes Excel/dashboard
VEH_MAP_TO_COL = {
    "TRACTEUR":    "TRACTEUR",
    "PPL":         "PETIT POILOUR",   # PPL = Petit Poilour
    "PL":          "POILOUR",          # PL  = Poilour
    "SEMI":        "SEMI",
    "DOUBLE_REM":  "SEMI",             # Double Rem = Semi
}

transport_rows = []
by_date_comm = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))
for row in all_days:
    d, comm = row["Date"], row["Commercial"]
    veh_str = row.get("Vehicules", "")
    for part in veh_str.split("|"):
        part = part.strip()
        if not part: continue
        # Extraire: "PLx2(15.0t)" → veh=PL, trips=2
        # Extraire: "TRC(caisses)" → veh=TRACTEUR, trips=1
        trips_n = 1
        matched_veh = None
        # Chercher dans l'ordre du plus long au plus court pour éviter PPL→PL
        for internal, col in sorted(VEH_MAP_TO_COL.items(), key=lambda x: -len(x[0])):
            if part.upper().startswith(internal.upper()):
                matched_veh = col
                try:
                    trips_n = int(part.split("x")[1].split("(")[0].strip())
                except: trips_n = 1
                break
        if matched_veh:
            by_date_comm[d][comm][matched_veh] += trips_n

for d in sorted(by_date_comm.keys()):
    for comm in sorted(by_date_comm[d].keys()):
        nd       = by_date_comm[d][comm]
        day_rows = [r for r in all_days if r["Date"]==d and r["Commercial"]==comm]
        avg_dist = round(sum(r["Distance km"] for r in day_rows)/max(len(day_rows),1), 0)
        transport_rows.append({
            "Date":                  d,
            "Commercial":            comm,
            "Total Tonnes":          round(sum(r["Tonnes/Jour"] for r in day_rows), 1),
            "Voyages TRACTEUR":      nd.get("TRACTEUR", 0),
            "Voyages PETIT POILOUR": nd.get("PETIT POILOUR", 0),
            "Voyages POILOUR":       nd.get("POILOUR", 0),
            "Voyages SEMI":          nd.get("SEMI", 0),
            "Distance Moy km":       int(avg_dist),
            "Jours Double":          0,
            "Pic":                   "yes" if PEAK_START <= d <= PEAK_END else "",
        })

transport_df = pd.DataFrame(transport_rows)

# Region summary
region_summary = result_df.groupby("Region").agg(
    Tonnes_Total=("Tonnes/Jour", "sum"),
    Nb_Agriculteurs=("Agriculteur", "nunique"),
    Distance_Moy_km=("Distance km", "mean"),
).round(0).reset_index()
region_summary.columns = ["Region","Tonnes Totales","Nb Agriculteurs","Distance Moy (km)"]

# Availability table
avail_rows = []
for farmer in farmers:
    for veh in farmer.allowed_veh:
        avail_rows.append({
            "Commercial": farmer.commercial, "Agriculteur": farmer.name,
            "Usine": farmer.usine, "Zone": farmer.zone,
            "Region": farmer.geo_region, "Distance km": farmer.distance_km,
            "Type Vehicule": veh, "Date Debut": farmer.start,
            "Date Fin": farmer.end, "Total Tonnes": farmer.tonnage,
        })
avail_df = pd.DataFrame(avail_rows).sort_values(["Commercial","Date Debut"]).reset_index(drop=True)

# Resume
resume_rows = []
for comm in sorted(result_df["Commercial"].unique()):
    resume_rows.append({
        "Commercial":            comm,
        "Tonnes Totales Saison": round(df[df["commercial"]==comm]["TONNAGE"].sum(), 0),
        "Nb Agriculteurs":       df[df["commercial"]==comm]["AGRICULTEUR"].nunique(),
        "Conflits Resolus":      0,
        "Total Jours Double":    0,
        "Statut":                "AI optimise",
    })
resume_df = pd.DataFrame(resume_rows)

# ============================================================
# STEP 6: Write Excel
# ============================================================
print(f"\nWriting {OUTPUT_FILE}...")

wb   = Workbook()
thin = Side(style="thin", color="BBBBBB")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_ws(ws, headers, widths, color):
    fill = PatternFill("solid", start_color=color)
    font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font=font; c.fill=fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 28

def write_df(ws, df_in, start_row=2):
    for ro, (_, row) in enumerate(df_in.iterrows()):
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=start_row+ro, column=ci, value=val)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = border
            c.font = Font(name="Calibri", size=9)

# Sheet 1: Planning
ws1 = wb.active; ws1.title = "Planning Journalier"
COLS1 = ["Commercial","Agriculteur","Usine","Region","Zone","Accessibilite",
         "Date","Tonnes/Jour","Type Vehicule","Vehicules","Nb Voyages",
         "Distance km","Date Debut","Date Fin","Total Tonnes","Pic de Recolte","Note"]
W1 = [15,24,10,12,14,11,12,10,13,38,9,11,12,12,12,10,14]
# ✅ Supprimer la colonne _is_rm (interne) avant l'export Excel
disp = result_df[COLS1].copy()
style_ws(ws1, COLS1, W1, "1F4E79")
write_df(ws1, disp)

# Sheet 2: Transport
ws2 = wb.create_sheet("Besoins Transport-Jour")
COLS2 = ["Date","Commercial","Total Tonnes","Voyages TRACTEUR",
         "Voyages PETIT POILOUR","Voyages POILOUR","Voyages SEMI",
         "Distance Moy km","Jours Double","Pic"]
style_ws(ws2, COLS2, [12,16,12,16,20,16,14,14,12,6], "375623")
write_df(ws2, transport_df[COLS2])

# Sheet 3: By Region
ws3 = wb.create_sheet("Par Region")
COLS3 = ["Region","Tonnes Totales","Nb Agriculteurs","Distance Moy (km)"]
style_ws(ws3, COLS3, [16,16,16,16], "7B2D8B")
write_df(ws3, region_summary[COLS3])

# Sheet 4: Availability
ws4 = wb.create_sheet("Disponibilite Vehicules")
COLS4 = ["Commercial","Agriculteur","Usine","Zone","Region","Distance km",
         "Type Vehicule","Date Debut","Date Fin","Total Tonnes"]
style_ws(ws4, COLS4, [15,26,10,16,12,11,14,12,12,12], "C00000")
write_df(ws4, avail_df)

# Sheet 5: Résumé optimisation (pas de double transport avec OR-Tools)
ws5 = wb.create_sheet("Résumé Optimisation")
ws5["A1"] = "Planning généré par OR-Tools — Optimisation IA (distances + caps simultanés)"
ws5["A1"].font = Font(bold=True, size=11, name="Calibri", color="1F4E79")
ws5["A2"] = f"Statut: {STATUS_NAMES.get(status, status)} | Farmers: {len(farmers)} | Total: {result_df['Tonnes/Jour'].sum():,.0f}t"
ws5["A2"].font = Font(size=10, name="Calibri", color="375623")
ws5["A3"] = "Aucun double transport — OR-Tools évite les conflits dès la planification"
ws5["A3"].font = Font(size=9, name="Calibri", color="5F5E5A")

# Sheet 6: Resume
ws6 = wb.create_sheet("Resume par Commercial")
COLS6 = ["Commercial","Tonnes Totales Saison","Nb Agriculteurs",
         "Conflits Resolus","Total Jours Double","Statut"]
style_ws(ws6, COLS6, [18,20,16,16,18,16], "1F4E79")
write_df(ws6, resume_df[COLS6])

# ── Sheet 7: Planning Horizontal (dates en colonnes) ──────────────────────
# Format : Commercial | Agriculteur | Usine | Total | 15-juin | 16-juin | ...
ws7 = wb.create_sheet("Planning Horizontal")
pivot = result_df.copy()
pivot["Date"] = pd.to_datetime(pivot["Date"])
all_pivot_dates = sorted(pivot["Date"].unique())

# Construire le pivot : lignes = (Commercial, Agriculteur, Usine), colonnes = dates
pivot_data = {}
for _, row in pivot.iterrows():
    key = (row["Commercial"], row["Agriculteur"], row["Usine"], row["Region"], row["Accessibilite"])
    d   = row["Date"]
    if key not in pivot_data:
        pivot_data[key] = {}
    pivot_data[key][d] = pivot_data[key].get(d, 0) + row["Tonnes/Jour"]

# Écrire en-têtes
h_cols = ["Commercial","Agriculteur","Usine","Region","Accessibilite","Total Saison"]
# Format date Windows-compatible (pas de %-d)
date_labels = []
for d in all_pivot_dates:
    label = d.strftime("%d/%m").lstrip("0") if d.strftime("%d/%m")[0] == "0" else d.strftime("%d/%m")
    date_labels.append(label)

all_h_headers = h_cols + date_labels
col_widths_h   = [14, 26, 10, 12, 12, 12] + [7] * len(date_labels)

fill7  = PatternFill("solid", start_color="C55A11")
font7h = Font(bold=True, color="FFFFFF", name="Calibri", size=9)
for ci, (h, w) in enumerate(zip(all_h_headers, col_widths_h), 1):
    c = ws7.cell(row=1, column=ci, value=h)
    c.font = font7h; c.fill = fill7
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border
    ws7.column_dimensions[get_column_letter(ci)].width = w
ws7.row_dimensions[1].height = 30

# Trier par Commercial puis Agriculteur
sorted_keys = sorted(pivot_data.keys(), key=lambda k: (k[0], k[1]))
fill_alt = PatternFill("solid", start_color="FFF2CC")  # alternance lignes

prev_comm = None
row_idx   = 2
for key in sorted_keys:
    comm, agri, usine, region, acc = key
    day_dict  = pivot_data[key]
    total_agri = round(sum(day_dict.values()), 0)

    # Couleur de fond alternée par commercial
    if comm != prev_comm:
        use_fill = (prev_comm is not None)
        prev_comm = comm
    
    row_vals = [comm, agri, usine, region, acc, total_agri]
    for d in all_pivot_dates:
        v = day_dict.get(d, 0)
        row_vals.append(int(v) if v > 0 else "")

    for ci, val in enumerate(row_vals, 1):
        c = ws7.cell(row=row_idx, column=ci, value=val)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border
        c.font = Font(name="Calibri", size=8)
        # Mettre en gras les totaux
        if ci == 6:
            c.font = Font(name="Calibri", size=8, bold=True)
        # Colorier les jours avec valeur
        if ci > 6 and val != "":
            c.fill = PatternFill("solid", start_color="E2EFDA")
    row_idx += 1

# Ligne de TOTAL en bas
ws7.cell(row=row_idx, column=1, value="TOTAL").font = Font(bold=True, name="Calibri", size=9)
ws7.cell(row=row_idx, column=6, value=round(result_df["Tonnes/Jour"].sum(), 0)).font = Font(bold=True, name="Calibri", size=9)
for ci, d in enumerate(all_pivot_dates, 7):
    day_total = sum(pv_data.get(d, 0) for pv_data in pivot_data.values())
    c = ws7.cell(row=row_idx, column=ci, value=int(round(day_total, 0)) if day_total > 0 else "")
    c.font = Font(bold=True, name="Calibri", size=8)
    c.fill = PatternFill("solid", start_color="BDD7EE")
    c.border = border

ws7.freeze_panes = "G2"  # Figer les 6 premières colonnes + l'en-tête
print(f"  Feuille 'Planning Horizontal' : {len(sorted_keys)} agriculteurs × {len(all_pivot_dates)} jours")

wb.save(TEMP_FILE)

import shutil, time
for attempt in range(5):
    try:
        shutil.move(TEMP_FILE, OUTPUT_FILE)
        break
    except PermissionError:
        if attempt < 4:
            print(f"  Excel ouvert, retry {attempt+1}/5 dans 3s...")
            time.sleep(3)
        else:
            print("  ERREUR: Fermez le fichier Excel et relancez.")
            sys.exit(1)

# ============================================================
# STEP 7: Summary
# ============================================================
print(f"\n{'='*60}")
print(f"  OPTIMIZER V2 COMPLETE -> {OUTPUT_FILE}")
print(f"{'='*60}")
print(f"  Status    : {STATUS_NAMES.get(status, status)}")
print(f"  Rows      : {len(result_df)}")
print(f"  Total declared : {sum(f.tonnage for f in farmers):,.0f}t")
print(f"  Total planned  : {result_df['Tonnes/Jour'].sum():,.0f}t")
print()
print("  Distance optimization:")
for region, dists in sorted(by_region.items()):
    valid = [d for d in dists if d < 999]
    if valid:
        total_ton_dist = sum(f.tonnage * f.distance_km
                             for f in farmers if f.geo_region == region and f.distance_km < 999)
        print(f"    {region:<15}: {len(valid)} farmers | avg {sum(valid)/len(valid):.0f}km")
print()
print("  Commercial caps (vérification PIC UNIQUEMENT 1-15 Juillet):")
# Convertir Date en datetime pour éviter l'erreur .dt.date
result_df["Date"] = pd.to_datetime(result_df["Date"], errors="coerce")
for comm in COMMERCIAL_CAPS:
    cap = COMMERCIAL_CAPS[comm]
    sub = result_df[result_df["Commercial"]==comm].copy()
    if sub.empty: continue
    cap_end_date = pd.Timestamp("2026-07-12") if comm in ("JILANI OBAY","KHALIL") else pd.Timestamp("2026-07-15")
    peak_start_ts = pd.Timestamp(PEAK_START)
    sub_pic = sub[(sub["Date"] >= peak_start_ts) & (sub["Date"] <= cap_end_date)]
    mx_pic  = sub_pic.groupby("Date")["Tonnes/Jour"].sum().max() if not sub_pic.empty else 0
    mx_all  = sub.groupby("Date")["Tonnes/Jour"].sum().max()
    ok_pic  = "✅" if mx_pic <= cap else "❌ DEPASSE PENDANT PIC"
    # ✅ Nombre de véhicules par jour (max journalier)
    veh_per_day = sub.groupby("Date")["Nb Voyages"].sum()
    max_veh_day = int(veh_per_day.max()) if not veh_per_day.empty else 0
    max_veh_date = veh_per_day.idxmax().strftime("%d/%m") if not veh_per_day.empty else "?"
    print(f"    {comm:<20}: max PIC={mx_pic:.0f}t/j | max hors-pic={mx_all:.0f}t/j | "
          f"limite={cap}t {ok_pic} | 🚛 max véhicules/jour={max_veh_day} ({max_veh_date})")
print()
print("  Vérification LIMITES usine (PIC UNIQUEMENT 1-15 Juillet):")
factory_start_ts = pd.Timestamp(FACTORY_CAP_START)
factory_end_ts   = pd.Timestamp(FACTORY_CAP_END)
for fact, lim in FACTORY_LIMITS.items():
    sub = result_df[result_df["Usine"]==fact].copy()
    if sub.empty: continue
    sub_pic = sub[(sub["Date"] >= factory_start_ts) & (sub["Date"] <= factory_end_ts)]
    mx_pic  = sub_pic.groupby("Date")["Tonnes/Jour"].sum().max() if not sub_pic.empty else 0
    mx_all  = sub.groupby("Date")["Tonnes/Jour"].sum().max()
    cap_phys = FACTORY_CAPS[fact]
    lim_dur  = lim * 1.05   # contrainte dure réelle (marge 5%)
    # ✅ Comparer au seuil réel (LIMITE + marge faisabilité 5%)
    if mx_pic <= lim:
        ok_pic = "✅ DANS LIMITE"
    elif mx_pic <= lim_dur:
        depass = mx_pic - lim
        ok_pic = f"⚠️ +{depass:.0f}t au-dessus de LIMITE (dans marge 5%)"
    else:
        ok_pic = f"❌ HORS MARGE (max={lim_dur:.0f}t)"
    print(f"    {fact:<12}: max PIC={mx_pic:.0f}t/j | max hors-pic={mx_all:.0f}t/j | LIMITE={lim}t / cap={cap_phys}t  {ok_pic}")
print()
print("  Next: python migrate.py")
print(f"{'='*60}")