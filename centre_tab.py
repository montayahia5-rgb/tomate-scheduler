# ============================================================
# centre_tab.py — Dashboard & Upload pour Centres de collecte
# Version 2.0 — Upload fichiers + visualisation + Supabase
#
# Centres supportés : BACCARA · KERKOUANE · 428
# Rôle Streamlit   : "centre"
# ============================================================

import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import io
import datetime

# ── Mapping centre → commercial ──────────────────────────────
CENTRE_TO_COMMERCIAL = {
    "BACCARA":    "FEDI",
    "KERKOUANE":  "MAKKI BEN SALAH",
    "428":        "JILANI OBAY",
}
COMMERCIAL_TO_CENTRE = {v: k for k, v in CENTRE_TO_COMMERCIAL.items()}

USINES    = ["SICAM", "COMOCAP", "TUCAL", "ABIDA", "ELFALLEH"]
REGIONS   = ["CAP BON 1", "CAP BON 2", "NORD", "GAFSA / KASSRINE",
             "KAIROUAN", "SIDI BOUZID", "BOUFICHA"]
ACCESS    = ["PL/PPL", "PL/SEMI", "RM", "TRC/PPL", "TRC/PPL/PL", "PPL", "PL", "SEMI"]
VEH_TYPES = ["PL", "PPL", "SEMI", "TRACTEUR"]

DARK  = "#161b22"
DARK2 = "#0d1117"
BORDER = "#21262d"

# ─────────────────────────────────────────────────────────────
# SQL À EXÉCUTER DANS SUPABASE (1 seule fois)
# ─────────────────────────────────────────────────────────────
SQL_INIT = """
-- Table agriculteurs par centre
CREATE TABLE IF NOT EXISTS centre_agriculteurs (
    id              BIGSERIAL PRIMARY KEY,
    centre          TEXT NOT NULL,
    commercial      TEXT,
    nom             TEXT NOT NULL,
    hectares        NUMERIC,
    tonnage         NUMERIC NOT NULL,
    variete         TEXT,
    region          TEXT,
    usine           TEXT,
    accessibilite   TEXT,
    date_debut      DATE,
    date_fin        DATE,
    note            TEXT,
    uploaded_at     TIMESTAMPTZ DEFAULT NOW()
);
CREATE INDEX IF NOT EXISTS idx_ca_centre ON centre_agriculteurs(centre);

-- Table transport par centre
CREATE TABLE IF NOT EXISTS centre_transport (
    id              BIGSERIAL PRIMARY KEY,
    centre          TEXT NOT NULL,
    commercial      TEXT,
    usine           TEXT,
    type_vehicule   TEXT,
    nb_bennes       INTEGER,
    cap_tonne       NUMERIC,
    disponible_du   DATE,
    disponible_au   DATE,
    note            TEXT,
    uploaded_at     TIMESTAMPTZ DEFAULT NOW()
);
CREATE INDEX IF NOT EXISTS idx_ct_centre ON centre_transport(centre);
"""

# ─────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────
def _metric(label, value, color="#f0f6fc", delta=None, delta_label=""):
    delta_html = ""
    if delta is not None:
        dc = "#3dd68c" if delta >= 0 else "#ef5350"
        sign = "+" if delta >= 0 else ""
        delta_html = f"<div style='font-size:.7rem;color:{dc};margin-top:3px'>{sign}{delta:,.0f} {delta_label}</div>"
    return f"""<div style='background:{DARK};border:1px solid {BORDER};border-radius:10px;
padding:14px 18px;border-top:2px solid {color}'>
<div style='font-size:.7rem;color:#8b949e;text-transform:uppercase;letter-spacing:.06em'>{label}</div>
<div style='font-size:1.5rem;font-weight:700;color:#f0f6fc;margin-top:4px'>{value}</div>
{delta_html}</div>"""

def _norm_col(col):
    """Normalise un nom de colonne."""
    import unicodedata
    c = str(col).strip().lower()
    c = ''.join(ch for ch in unicodedata.normalize('NFD', c)
                if unicodedata.category(ch) != 'Mn')
    return c.replace(" ", "_").replace("/", "_").replace("(", "").replace(")", "").replace(".", "")

def _find(df, candidates):
    """Trouve la première colonne qui matche une liste de candidats normalisés."""
    norm_map = {_norm_col(c): c for c in df.columns}
    for cand in candidates:
        if _norm_col(cand) in norm_map:
            return norm_map[_norm_col(cand)]
    return None


# ─────────────────────────────────────────────────────────────
# PARSEURS UPLOAD
# ─────────────────────────────────────────────────────────────
def _parse_agriculteurs(file_obj, centre, commercial):
    """
    Parse un fichier Excel d'agriculteurs uploadé par un centre.
    Colonnes reconnues (flexibles) :
      - NOM / AGRICULTEUR
      - HECTARES / HA / NBR_HECTAR
      - TONNAGE / TONNAGE_TOTAL / TONNES
      - VARIETE / VARIÉTÉ
      - REGION / RÉGION
      - USINE
      - ACCESSIBILITE / ACCESSIBILITÉ / ACCES
      - DATE_DEBUT / DATE DEBUT RECOLTE
      - DATE_FIN / DATE FIN RECOLTE
      - NOTE / COMMENTAIRE
    """
    errors = []
    try:
        # Essayer header=0 puis header=3 (format référence interne)
        for hdr in [0, 1, 2, 3]:
            try:
                df = pd.read_excel(file_obj, sheet_name=0, header=hdr)
                df.columns = [str(c).strip() for c in df.columns]
                df = df.dropna(how='all')
                if len(df) > 0 and len(df.columns) >= 2:
                    break
            except Exception:
                continue

        # Identifier colonnes
        nom_col   = _find(df, ["nom","agriculteur","name","client"])
        ha_col    = _find(df, ["hectares","ha","nbr_hectar","nbrhect","hectare"])
        ton_col   = _find(df, ["tonnage","tonnage_total","tonnes","tonnage total","tonnage_total_t"])
        var_col   = _find(df, ["variete","variété","variety","var"])
        reg_col   = _find(df, ["region","région","zone_region","reg"])
        usi_col   = _find(df, ["usine","usine_destination","factory"])
        acc_col   = _find(df, ["accessibilite","accessibilité","acces","accessibilite_vehicule"])
        dd_col    = _find(df, ["date_debut","date debut","date_debut_recolte","debut_recolte","date debut recolte"])
        df_col    = _find(df, ["date_fin","date fin","date_fin_recolte","fin_recolte","date fin recolte"])
        note_col  = _find(df, ["note","commentaire","observation","notes"])

        if not nom_col:
            return None, ["❌ Colonne NOM introuvable. Vérifiez les en-têtes (NOM ou AGRICULTEUR)."]
        if not ton_col:
            return None, ["❌ Colonne TONNAGE introuvable. Vérifiez les en-têtes (TONNAGE ou TONNES)."]

        # Nettoyer
        df = df.dropna(subset=[nom_col])
        df[nom_col] = df[nom_col].astype(str).str.strip().str.upper()
        df = df[df[nom_col].str.len() > 2]
        df = df[~df[nom_col].str.startswith("TOTAL")]
        df = df[~df[nom_col].str.startswith("SOUS")]

        records = []
        for _, row in df.iterrows():
            nom     = str(row[nom_col]).strip().upper()
            tonnage = pd.to_numeric(row.get(ton_col, 0), errors="coerce")
            if pd.isna(tonnage) or tonnage <= 0:
                continue
            ha   = pd.to_numeric(row.get(ha_col, 0), errors="coerce") if ha_col else None
            var  = str(row.get(var_col, "")).strip() if var_col else ""
            reg  = str(row.get(reg_col, "CAP BON 1")).strip() if reg_col else "CAP BON 1"
            usi  = str(row.get(usi_col, "SICAM")).strip().upper() if usi_col else "SICAM"
            acc  = str(row.get(acc_col, "PL/PPL")).strip() if acc_col else "PL/PPL"
            note = str(row.get(note_col, "")).strip() if note_col else ""

            # Dates
            dd = None; df_ = None
            if dd_col:
                try:
                    dd = pd.to_datetime(row[dd_col], errors="coerce")
                    dd = dd.strftime("%Y-%m-%d") if not pd.isna(dd) else None
                except Exception:
                    pass
            if df_col:
                try:
                    df_ = pd.to_datetime(row[df_col], errors="coerce")
                    df_ = df_.strftime("%Y-%m-%d") if not pd.isna(df_) else None
                except Exception:
                    pass

            records.append({
                "centre":        centre,
                "commercial":    commercial,
                "nom":           nom,
                "hectares":      float(ha) if ha and not pd.isna(ha) else None,
                "tonnage":       float(tonnage),
                "variete":       var or None,
                "region":        reg,
                "usine":         usi,
                "accessibilite": acc,
                "date_debut":    dd,
                "date_fin":      df_,
                "note":          note or None,
            })

        if not records:
            return None, ["❌ Aucun agriculteur valide trouvé (tonnage > 0 requis)."]

        return pd.DataFrame(records), []

    except Exception as e:
        return None, [f"❌ Erreur lecture fichier : {e}"]


def _parse_transport(file_obj, centre, commercial):
    """
    Parse un fichier Excel de transport.
    Colonnes : USINE · TYPE_VEHICULE · NB_BENNES · CAP_TONNE · DISPO_DU · DISPO_AU · NOTE
    """
    try:
        df = pd.read_excel(file_obj, sheet_name=0, header=0)
        df.columns = [str(c).strip() for c in df.columns]
        df = df.dropna(how='all')

        usi_col = _find(df, ["usine","factory"])
        typ_col = _find(df, ["type_vehicule","type vehicule","type","vehicule"])
        nb_col  = _find(df, ["nb_bennes","nb bennes","nombre_bennes","nombre","nb"])
        cap_col = _find(df, ["cap_tonne","cap tonne","capacite","capacite_tonne","tonnage_benne"])
        du_col  = _find(df, ["disponible_du","dispo_du","date_debut","debut"])
        au_col  = _find(df, ["disponible_au","dispo_au","date_fin","fin"])
        note_col= _find(df, ["note","commentaire"])

        records = []
        for _, row in df.iterrows():
            nb = pd.to_numeric(row.get(nb_col, 0), errors="coerce") if nb_col else 0
            if pd.isna(nb) or nb <= 0:
                continue
            cap = pd.to_numeric(row.get(cap_col, 0), errors="coerce") if cap_col else 0
            usi = str(row.get(usi_col, "")).strip().upper() if usi_col else ""
            typ = str(row.get(typ_col, "PL")).strip().upper() if typ_col else "PL"

            dd = None
            if du_col:
                try:
                    d = pd.to_datetime(row[du_col], errors="coerce")
                    dd = d.strftime("%Y-%m-%d") if not pd.isna(d) else None
                except Exception:
                    pass
            da = None
            if au_col:
                try:
                    d = pd.to_datetime(row[au_col], errors="coerce")
                    da = d.strftime("%Y-%m-%d") if not pd.isna(d) else None
                except Exception:
                    pass

            records.append({
                "centre":          centre,
                "commercial":      commercial,
                "usine":           usi or None,
                "type_vehicule":   typ,
                "nb_bennes":       int(nb),
                "cap_tonne":       float(cap) if cap and not pd.isna(cap) else None,
                "disponible_du":   dd,
                "disponible_au":   da,
                "note": str(row.get(note_col,"")).strip() if note_col else None,
            })

        if not records:
            return None, ["❌ Aucune ligne de transport valide (nb_bennes > 0 requis)."]
        return pd.DataFrame(records), []
    except Exception as e:
        return None, [f"❌ Erreur lecture fichier transport : {e}"]


# ─────────────────────────────────────────────────────────────
# LOADERS SUPABASE
# ─────────────────────────────────────────────────────────────
@st.cache_data(ttl=15)
def _load_agri(centre, _v=0):
    try:
        sb = st.session_state.get("_sb_ref")
        if sb is None:
            return pd.DataFrame()
        data = sb.table("centre_agriculteurs").select("*").eq("centre", centre).execute().data
        return pd.DataFrame(data) if data else pd.DataFrame()
    except Exception:
        return pd.DataFrame()


@st.cache_data(ttl=15)
def _load_transport(centre, _v=0):
    try:
        sb = st.session_state.get("_sb_ref")
        if sb is None:
            return pd.DataFrame()
        data = sb.table("centre_transport").select("*").eq("centre", centre).execute().data
        return pd.DataFrame(data) if data else pd.DataFrame()
    except Exception:
        return pd.DataFrame()


def _get_planning(sb, centre):
    """Charge le planning du centre depuis la table planning."""
    try:
        comm = CENTRE_TO_COMMERCIAL.get(centre, "")
        data = sb.table("planning").select(
            "date,agriculteur,usine,tonnes_jour,type_vehicule,vehicules,note"
        ).eq("commercial", comm).execute().data
        if not data:
            return pd.DataFrame()
        df = pd.DataFrame(data)
        df["date"] = pd.to_datetime(df["date"], errors="coerce")
        df["tonnes_jour"] = pd.to_numeric(df["tonnes_jour"], errors="coerce").fillna(0)
        return df
    except Exception:
        return pd.DataFrame()


# ─────────────────────────────────────────────────────────────
# MODÈLE EXCEL À TÉLÉCHARGER
# ─────────────────────────────────────────────────────────────
def _generate_template(centre):
    """Génère un modèle Excel pré-rempli pour le centre."""
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ── Feuille Agriculteurs ──────────────────────────────────
    ws = wb.active
    ws.title = "Agriculteurs"
    ws.sheet_view.showGridLines = False

    hf = lambda h: PatternFill("solid", start_color=h, end_color=h)
    bf = lambda bold=True, white=True, size=10: Font(bold=bold, name="Calibri", size=size,
                                                       color="FFFFFF" if white else "000000")
    BD = Border(left=Side(style="thin", color="CCCCCC"),
                right=Side(style="thin", color="CCCCCC"),
                top=Side(style="thin", color="CCCCCC"),
                bottom=Side(style="thin", color="CCCCCC"))
    CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LFT = Alignment(horizontal="left", vertical="center")

    cols = [
        ("NOM", 28, "1F3864"),
        ("HECTARES", 12, "375623"),
        ("TONNAGE", 12, "375623"),
        ("VARIETE", 14, "4A235A"),
        ("REGION", 18, "4A235A"),
        ("USINE", 12, "7B3F00"),
        ("ACCESSIBILITE", 16, "0B4F6C"),
        ("DATE_DEBUT", 16, "C0392B"),
        ("DATE_FIN", 16, "C0392B"),
        ("NOTE", 24, "444444"),
    ]

    # Titre
    ws.merge_cells("A1:J1")
    ws["A1"] = f"Fichier Agriculteurs — Centre {centre} — Campagne 2026"
    ws["A1"].font = bf(True, True, 12)
    ws["A1"].fill = hf("0B4F6C")
    ws["A1"].alignment = CTR
    ws.row_dimensions[1].height = 30

    for ci, (col, w, c) in enumerate(cols, 1):
        cell = ws.cell(2, ci, value=col)
        cell.font = bf(True, True, 10)
        cell.fill = hf(c)
        cell.alignment = CTR
        cell.border = BD
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[2].height = 28

    # Exemples
    examples = [
        ("AHMED BEN ALI",      4.5,  180, "Savera",    "CAP BON 1", "SICAM",   "PL/PPL", "2026-07-01", "2026-07-20", ""),
        ("MOHAMED TRABELSI",   3.0,  120, "Dorra",     "CAP BON 1", "TUCAL",   "PL/PPL", "2026-07-05", "2026-07-25", "Bon accès"),
        ("STE EXEMPLE S.A",   25.0, 1000, "Savera",    "CAP BON 2", "SICAM",   "PL/PPL", "2026-06-25", "2026-08-10", "Multi-usine"),
    ]
    for ri, ex in enumerate(examples, 3):
        for ci, val in enumerate(ex, 1):
            c = ws.cell(ri, ci, value=val)
            c.border = BD
            c.fill = hf("F0F5FF" if ri % 2 == 1 else "FFFFFF")
            c.font = Font(name="Calibri", size=9)
            c.alignment = LFT if ci == 1 else CTR
        ws.row_dimensions[ri].height = 20

    ws.freeze_panes = "A3"

    # ── Feuille Transport ─────────────────────────────────────
    ws2 = wb.create_sheet("Transport")
    ws2.sheet_view.showGridLines = False
    ws2.merge_cells("A1:G1")
    ws2["A1"] = f"Fichier Transport — Centre {centre} — Campagne 2026"
    ws2["A1"].font = bf(True, True, 12)
    ws2["A1"].fill = hf("375623")
    ws2["A1"].alignment = CTR
    ws2.row_dimensions[1].height = 30

    t_cols = [
        ("USINE", 14, "375623"),
        ("TYPE_VEHICULE", 14, "1F3864"),
        ("NB_BENNES", 12, "7B3F00"),
        ("CAP_TONNE", 14, "7B3F00"),
        ("DISPONIBLE_DU", 16, "C0392B"),
        ("DISPONIBLE_AU", 16, "C0392B"),
        ("NOTE", 28, "444444"),
    ]
    for ci, (col, w, c) in enumerate(t_cols, 1):
        cell = ws2.cell(2, ci, value=col)
        cell.font = bf(True, True, 10)
        cell.fill = hf(c)
        cell.alignment = CTR
        cell.border = BD
        ws2.column_dimensions[get_column_letter(ci)].width = w
    ws2.row_dimensions[2].height = 28

    t_examples = [
        ("SICAM",  "PL",  12, 20, "2026-06-20", "2026-08-25", "Transport propre"),
        ("TUCAL",  "PPL",  6, 10, "2026-07-01", "2026-08-15", ""),
        ("COMOCAP","SEMI", 3, 30, "2026-07-01", "2026-07-31", "Location"),
    ]
    for ri, ex in enumerate(t_examples, 3):
        for ci, val in enumerate(ex, 1):
            c = ws2.cell(ri, ci, value=val)
            c.border = BD
            c.fill = hf("E8F5E9" if ri % 2 == 1 else "FFFFFF")
            c.font = Font(name="Calibri", size=9)
            c.alignment = CTR
        ws2.row_dimensions[ri].height = 20
    ws2.freeze_panes = "A3"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ─────────────────────────────────────────────────────────────
# FONCTION PRINCIPALE
# ─────────────────────────────────────────────────────────────
def render_centre_dashboard(sb, role_filter, centre_name):
    """Dashboard complet pour un centre de collecte."""

    # Stocker sb dans session pour les fonctions cachées
    st.session_state["_sb_ref"] = sb

    # Détecter le centre
    centre = role_filter.upper() if role_filter else centre_name.upper()
    # Normaliser
    for key in CENTRE_TO_COMMERCIAL:
        if key in centre:
            centre = key
            break
    commercial = CENTRE_TO_COMMERCIAL.get(centre, centre)

    # ── Version counter pour refresh ──────────────────────────
    if "centre_refresh" not in st.session_state:
        st.session_state["centre_refresh"] = 0

    # ── CSS ───────────────────────────────────────────────────
    st.markdown("""
    <style>
    [data-testid="stAppViewContainer"]{background:#0d1117}
    [data-testid="stSidebar"]{background:#161b22;border-right:1px solid #21262d}
    h1,h2,h3,h4{color:#f0f6fc}
    </style>""", unsafe_allow_html=True)

    # ── Header ────────────────────────────────────────────────
    st.markdown(f"""
    <div style='background:linear-gradient(90deg,#0B4F6C,#1a2332);
    border-radius:12px;padding:18px 24px;margin-bottom:20px;
    border-left:5px solid #00E5A0'>
    <div style='font-size:1.5rem;font-weight:900;color:#f0f6fc'>
    🏭 Centre {centre}
    </div>
    <div style='font-size:.85rem;color:#8b949e;margin-top:4px'>
    Commercial associé : <b style="color:#00E5A0">{commercial}</b>
    &nbsp;·&nbsp; Campagne Tomate 2026
    </div>
    </div>""", unsafe_allow_html=True)

    # ── Onglets ───────────────────────────────────────────────
    tab_dash, tab_upload, tab_agri, tab_transport, tab_planning, tab_sql = st.tabs([
        "📊 Tableau de bord",
        "📤 Upload fichiers",
        "🌾 Mes Agriculteurs",
        "🚛 Mon Transport",
        "📅 Planning OR-Tools",
        "🛠️ Config SQL",
    ])

    # ════════════════════════════════════════════════════════
    # TAB 1 — TABLEAU DE BORD
    # ════════════════════════════════════════════════════════
    with tab_dash:
        df_agri = _load_agri(centre, _v=st.session_state["centre_refresh"])
        df_transp = _load_transport(centre, _v=st.session_state["centre_refresh"])
        df_plan = _get_planning(sb, centre)

        if df_agri.empty and df_transp.empty:
            st.info("📭 Aucune donnée encore uploadée. Allez dans **📤 Upload fichiers** pour commencer.")
            st.markdown(f"""
            <div style='background:#1a2332;border-radius:10px;padding:20px;margin-top:16px'>
            <h3 style='color:#f0f6fc'>🚀 Démarrage rapide</h3>
            <ol style='color:#ccc;line-height:2'>
              <li>Allez dans <b>📤 Upload fichiers</b></li>
              <li>Téléchargez le <b>modèle Excel</b></li>
              <li>Remplissez vos agriculteurs et votre transport</li>
              <li>Uploadez le fichier rempli</li>
              <li>Revenez ici pour voir votre tableau de bord ✅</li>
            </ol>
            </div>""", unsafe_allow_html=True)
            return

        # KPIs agriculteurs
        if not df_agri.empty:
            df_agri["tonnage"] = pd.to_numeric(df_agri["tonnage"], errors="coerce").fillna(0)
            df_agri["hectares"] = pd.to_numeric(df_agri.get("hectares", 0), errors="coerce").fillna(0)
            tot_t = df_agri["tonnage"].sum()
            tot_h = df_agri["hectares"].sum()
            n_ag  = len(df_agri)
            rend  = round(tot_t / tot_h, 1) if tot_h > 0 else 0

            cols_k = st.columns(4)
            cols_k[0].markdown(_metric("Agriculteurs", n_ag, "#00E5A0"), unsafe_allow_html=True)
            cols_k[1].markdown(_metric("Tonnage total", f"{tot_t:,.0f} T", "#F5A623"), unsafe_allow_html=True)
            cols_k[2].markdown(_metric("Hectares", f"{tot_h:,.1f} ha", "#8B5CF6"), unsafe_allow_html=True)
            cols_k[3].markdown(_metric("Rendement moy.", f"{rend} t/ha", "#3B82F6"), unsafe_allow_html=True)

            # Répartition par usine
            c1, c2 = st.columns(2)
            with c1:
                if "usine" in df_agri.columns:
                    gu = df_agri.groupby("usine")["tonnage"].sum().reset_index()
                    fig = px.pie(gu, names="usine", values="tonnage",
                                 title="Tonnage par usine", hole=0.45,
                                 template="plotly_dark")
                    fig.update_layout(paper_bgcolor=DARK, height=300)
                    st.plotly_chart(fig, use_container_width=True)
            with c2:
                if "region" in df_agri.columns:
                    gr = df_agri.groupby("region")["tonnage"].sum().sort_values(ascending=True).reset_index()
                    fig2 = px.bar(gr, x="tonnage", y="region", orientation="h",
                                  title="Tonnage par région", template="plotly_dark",
                                  color="tonnage", color_continuous_scale="Viridis")
                    fig2.update_layout(paper_bgcolor=DARK, height=300, showlegend=False)
                    st.plotly_chart(fig2, use_container_width=True)

        # KPIs transport
        if not df_transp.empty:
            st.markdown("---")
            st.markdown("#### 🚛 Résumé transport déclaré")
            df_transp["nb_bennes"] = pd.to_numeric(df_transp["nb_bennes"], errors="coerce").fillna(0)
            df_transp["cap_tonne"] = pd.to_numeric(df_transp["cap_tonne"], errors="coerce").fillna(0)
            df_transp["cap_totale"] = df_transp["nb_bennes"] * df_transp["cap_tonne"]

            t1, t2, t3 = st.columns(3)
            t1.markdown(_metric("Total bennes", int(df_transp["nb_bennes"].sum()), "#F5A623"), unsafe_allow_html=True)
            t2.markdown(_metric("Capacité totale", f"{df_transp['cap_totale'].sum():,.0f} t/j", "#00E5A0"), unsafe_allow_html=True)
            t3.markdown(_metric("Types véhicules", df_transp["type_vehicule"].nunique(), "#8B5CF6"), unsafe_allow_html=True)

        # Planning snapshot
        if not df_plan.empty:
            st.markdown("---")
            st.markdown("#### 📅 Extrait du planning OR-Tools")
            peak_s = datetime.date(2026, 7, 1)
            peak_e = datetime.date(2026, 7, 15)
            daily = df_plan.groupby(df_plan["date"].dt.date)["tonnes_jour"].sum().reset_index()
            daily.columns = ["Date", "Tonnes/Jour"]
            fig3 = px.area(daily, x="Date", y="Tonnes/Jour",
                           title=f"Livraisons journalières — {centre}",
                           template="plotly_dark", color_discrete_sequence=["#00E5A0"])
            fig3.add_vrect(x0=str(peak_s), x1=str(peak_e), fillcolor="gold",
                           opacity=0.07, line_width=0,
                           annotation_text="⚡ PIC", annotation_position="top left",
                           annotation_font_color="#F5A623")
            fig3.update_layout(paper_bgcolor=DARK, height=280)
            st.plotly_chart(fig3, use_container_width=True)

    # ════════════════════════════════════════════════════════
    # TAB 2 — UPLOAD FICHIERS
    # ════════════════════════════════════════════════════════
    with tab_upload:
        st.markdown("### 📤 Upload de vos fichiers")

        # Modèle à télécharger
        col_tpl, col_info = st.columns([2, 3])
        with col_tpl:
            tpl_bytes = _generate_template(centre)
            st.download_button(
                "⬇️ Télécharger le modèle Excel",
                data=tpl_bytes,
                file_name=f"modele_centre_{centre.lower()}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True,
            )
        with col_info:
            st.markdown(f"""
            <div style='background:#1a2332;border-radius:8px;padding:12px;font-size:.85rem;color:#ccc'>
            📋 Le modèle contient 2 feuilles :<br>
            &nbsp;• <b>Agriculteurs</b> : NOM · HECTARES · TONNAGE · VARIETE · REGION · USINE · ACCESSIBILITE · DATE_DEBUT · DATE_FIN<br>
            &nbsp;• <b>Transport</b> : USINE · TYPE_VEHICULE · NB_BENNES · CAP_TONNE · DISPONIBLE_DU · DISPONIBLE_AU
            </div>""", unsafe_allow_html=True)

        st.markdown("---")

        # ── Upload agriculteurs ────────────────────────────────────
        st.markdown("#### 🌾 Ficher Agriculteurs")
        up_agri = st.file_uploader(
            "Choisir le fichier Excel agriculteurs",
            type=["xlsx","xls"],
            key="upload_agri_centre",
            help="Colonnes requises : NOM et TONNAGE. Les autres sont optionnelles."
        )

        if up_agri:
            df_parsed, errs = _parse_agriculteurs(up_agri, centre, commercial)
            if errs:
                for e in errs:
                    st.error(e)
            else:
                st.success(f"✅ {len(df_parsed)} agriculteurs détectés")
                st.dataframe(df_parsed, use_container_width=True, height=220, hide_index=True)

                col_save, col_cancel = st.columns(2)
                with col_save:
                    if st.button("💾 Sauvegarder dans Supabase",
                                 type="primary", use_container_width=True,
                                 key="save_agri"):
                        try:
                            # Supprimer anciens, insérer nouveaux
                            sb.table("centre_agriculteurs").delete().eq("centre", centre).execute()
                            records = df_parsed.where(pd.notnull(df_parsed), None).to_dict("records")
                            sb.table("centre_agriculteurs").insert(records).execute()
                            st.success(f"✅ {len(records)} agriculteurs sauvegardés !")
                            st.session_state["centre_refresh"] += 1
                            st.cache_data.clear()
                            st.rerun()
                        except Exception as e:
                            st.error(f"❌ Erreur Supabase : {e}")
                            st.info("💡 Exécutez le SQL dans l'onglet 🛠️ Config SQL si la table n'existe pas.")
                with col_cancel:
                    st.button("✖ Annuler", use_container_width=True)

        st.markdown("---")

        # ── Upload transport ───────────────────────────────────────
        st.markdown("#### 🚛 Fichier Transport")
        up_transp = st.file_uploader(
            "Choisir le fichier Excel transport",
            type=["xlsx","xls"],
            key="upload_transp_centre",
            help="Colonnes requises : TYPE_VEHICULE et NB_BENNES."
        )

        if up_transp:
            df_transp_p, errs_t = _parse_transport(up_transp, centre, commercial)
            if errs_t:
                for e in errs_t:
                    st.error(e)
            else:
                st.success(f"✅ {len(df_transp_p)} lignes de transport détectées")
                st.dataframe(df_transp_p, use_container_width=True, height=180, hide_index=True)

                if st.button("💾 Sauvegarder transport",
                             type="primary", use_container_width=True,
                             key="save_transport"):
                    try:
                        sb.table("centre_transport").delete().eq("centre", centre).execute()
                        records_t = df_transp_p.where(pd.notnull(df_transp_p), None).to_dict("records")
                        sb.table("centre_transport").insert(records_t).execute()
                        st.success(f"✅ {len(records_t)} lignes transport sauvegardées !")
                        st.session_state["centre_refresh"] += 1
                        st.cache_data.clear()
                        st.rerun()
                    except Exception as e:
                        st.error(f"❌ Erreur Supabase : {e}")
                        st.info("💡 Vérifiez l'onglet 🛠️ Config SQL.")

        # ── Effacer tout ──────────────────────────────────────────
        st.markdown("---")
        with st.expander("🗑️ Effacer toutes mes données (reset)"):
            st.warning("⚠️ Cette action supprime TOUS les agriculteurs et transport enregistrés pour votre centre.")
            confirm = st.checkbox("Je confirme la suppression")
            if confirm and st.button("🗑️ Effacer", type="primary"):
                try:
                    sb.table("centre_agriculteurs").delete().eq("centre", centre).execute()
                    sb.table("centre_transport").delete().eq("centre", centre).execute()
                    st.success("✅ Données effacées.")
                    st.session_state["centre_refresh"] += 1
                    st.cache_data.clear()
                    st.rerun()
                except Exception as e:
                    st.error(f"Erreur : {e}")

    # ════════════════════════════════════════════════════════
    # TAB 3 — MES AGRICULTEURS
    # ════════════════════════════════════════════════════════
    with tab_agri:
        df_agri = _load_agri(centre, _v=st.session_state["centre_refresh"])

        if df_agri.empty:
            st.info("Aucun agriculteur uploadé. Allez dans **📤 Upload fichiers**.")
        else:
            df_agri["tonnage"] = pd.to_numeric(df_agri["tonnage"], errors="coerce").fillna(0)
            df_agri["hectares"] = pd.to_numeric(df_agri.get("hectares", 0), errors="coerce").fillna(0)

            # Filtres rapides
            c_f1, c_f2 = st.columns(2)
            with c_f1:
                f_usine = st.multiselect("Usine",
                    options=sorted(df_agri["usine"].dropna().unique()),
                    default=sorted(df_agri["usine"].dropna().unique()))
            with c_f2:
                f_reg = st.multiselect("Région",
                    options=sorted(df_agri["region"].dropna().unique()),
                    default=sorted(df_agri["region"].dropna().unique()))

            df_f = df_agri[df_agri["usine"].isin(f_usine) & df_agri["region"].isin(f_reg)]

            # Stats filtrées
            k1, k2, k3 = st.columns(3)
            k1.metric("Agriculteurs", len(df_f))
            k2.metric("Tonnage", f"{df_f['tonnage'].sum():,.0f} T")
            k3.metric("Hectares", f"{df_f['hectares'].sum():,.1f} ha")

            # Tableau
            show_cols = [c for c in ["nom","usine","region","tonnage","hectares",
                                      "variete","accessibilite","date_debut","date_fin","note"]
                         if c in df_f.columns]
            st.dataframe(df_f[show_cols].sort_values("tonnage", ascending=False),
                         use_container_width=True, height=400, hide_index=True,
                         column_config={
                             "tonnage": st.column_config.NumberColumn("Tonnage (T)", format="%.0f T"),
                             "hectares": st.column_config.NumberColumn("Ha", format="%.1f"),
                         })

            # Export
            buf = io.BytesIO()
            df_f[show_cols].to_excel(buf, index=False)
            buf.seek(0)
            st.download_button("📊 Exporter Excel",
                data=buf.read(),
                file_name=f"agriculteurs_{centre}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True)

    # ════════════════════════════════════════════════════════
    # TAB 4 — MON TRANSPORT
    # ════════════════════════════════════════════════════════
    with tab_transport:
        df_t = _load_transport(centre, _v=st.session_state["centre_refresh"])

        if df_t.empty:
            st.info("Aucun transport uploadé. Allez dans **📤 Upload fichiers**.")
        else:
            df_t["nb_bennes"] = pd.to_numeric(df_t["nb_bennes"], errors="coerce").fillna(0)
            df_t["cap_tonne"] = pd.to_numeric(df_t["cap_tonne"], errors="coerce").fillna(0)
            df_t["cap_totale"] = df_t["nb_bennes"] * df_t["cap_tonne"]

            # KPIs
            k1, k2, k3, k4 = st.columns(4)
            k1.metric("Bennes totales", int(df_t["nb_bennes"].sum()))
            k2.metric("Capacité totale", f"{df_t['cap_totale'].sum():,.0f} t/j")
            k3.metric("Usines couvertes", df_t["usine"].nunique() if "usine" in df_t.columns else 0)
            k4.metric("Types véhicules", df_t["type_vehicule"].nunique())

            # Graphique
            if "type_vehicule" in df_t.columns:
                by_type = df_t.groupby("type_vehicule").agg(
                    nb=("nb_bennes","sum"), cap=("cap_totale","sum")).reset_index()
                fig_t = px.bar(by_type, x="type_vehicule", y="nb",
                               title="Bennes par type de véhicule", template="plotly_dark",
                               color="type_vehicule", text="nb")
                fig_t.update_layout(paper_bgcolor=DARK, height=280, showlegend=False)
                st.plotly_chart(fig_t, use_container_width=True)

            # Tableau
            show_t = [c for c in ["usine","type_vehicule","nb_bennes","cap_tonne","cap_totale",
                                   "disponible_du","disponible_au","note"] if c in df_t.columns]
            st.dataframe(df_t[show_t], use_container_width=True, height=300, hide_index=True)

    # ════════════════════════════════════════════════════════
    # TAB 5 — PLANNING OR-TOOLS
    # ════════════════════════════════════════════════════════
    with tab_planning:
        df_plan = _get_planning(sb, centre)

        if df_plan.empty:
            st.info(f"Aucun planning OR-Tools trouvé pour le commercial **{commercial}**.")
            st.markdown("Le planning sera visible ici une fois que le directeur aura lancé l'optimiseur.")
        else:
            df_plan["tonnes_jour"] = pd.to_numeric(df_plan["tonnes_jour"], errors="coerce").fillna(0)

            # KPIs
            k1, k2, k3 = st.columns(3)
            k1.metric("Total tonnes planifié", f"{df_plan['tonnes_jour'].sum():,.0f} T")
            k2.metric("Jours planifiés", df_plan["date"].dt.date.nunique())
            k3.metric("Agriculteurs", df_plan["agriculteur"].nunique() if "agriculteur" in df_plan.columns else 0)

            # Courbe
            daily = df_plan.groupby(df_plan["date"].dt.date)["tonnes_jour"].sum().reset_index()
            daily.columns = ["Date", "Tonnes/Jour"]
            fig_p = px.bar(daily, x="Date", y="Tonnes/Jour",
                           title=f"Planning journalier — {centre} (commercial: {commercial})",
                           template="plotly_dark", color_discrete_sequence=["#00E5A0"])
            fig_p.add_vrect(x0="2026-07-01", x1="2026-07-15", fillcolor="gold",
                            opacity=0.07, line_width=0,
                            annotation_text="⚡ PIC", annotation_position="top left",
                            annotation_font_color="#F5A623")
            fig_p.update_layout(paper_bgcolor=DARK, height=350)
            st.plotly_chart(fig_p, use_container_width=True)

            # Tableau
            show_p = [c for c in ["date","agriculteur","usine","tonnes_jour","type_vehicule","vehicules","note"]
                      if c in df_plan.columns]
            df_plan_disp = df_plan[show_p].sort_values("date")
            df_plan_disp["date"] = df_plan_disp["date"].dt.strftime("%Y-%m-%d")
            st.dataframe(df_plan_disp, use_container_width=True, height=300, hide_index=True)

    # ════════════════════════════════════════════════════════
    # TAB 6 — CONFIG SQL
    # ════════════════════════════════════════════════════════
    with tab_sql:
        st.markdown("### 🛠️ Configuration Supabase")
        st.markdown("Copiez ce SQL dans **Supabase → SQL Editor → Run** (1 seule fois) :")
        st.code(SQL_INIT, language="sql")
        st.info("Cette opération est idempotente (CREATE IF NOT EXISTS) — sans risque si exécutée plusieurs fois.")
        st.markdown("---")
        st.markdown("**Test de connexion :**")
        if st.button("🔌 Tester la connexion Supabase"):
            try:
                res = sb.table("centre_agriculteurs").select("id").limit(1).execute()
                st.success("✅ Table `centre_agriculteurs` accessible !")
            except Exception as e:
                st.error(f"❌ Erreur : {e}")
                st.warning("Exécutez le SQL ci-dessus dans Supabase d'abord.")