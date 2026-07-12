# -*- coding: utf-8 -*-
"""
agroeco_dashboard.py v2 — Dashboard Agroéconomique Tomate 2026
===============================================================
CORRECTIONS v2 :
  ✅ Tous les fichiers entrée = "centre" + "client" obligatoires
  ✅ Royal : date_debut_repiquage → date_debut_livraison
  ✅ Caisses vides : condition = date_debut_RECOLTE (fichier rectifié Supabase)
     1ère affectation = date_debut_recolte < 10 juillet  → avec caisses
     2ème affectation = date_debut_recolte ≥ 10 juillet  → sans caisses
"""

import streamlit as st

# ══════════════════════════════════════════════════════════════════
# DONNÉES RÉELLES 2026 — FICHIERS ORGANISÉS + DETAIL_VENTE SOTUSFA
# 220/220 Ha réels | 101 intrants réels | 119 estimés | STE 428 = 0
# STE BACCARA=SOCIETE BACCARA ET FILS (30,260 DT)
# STE KERKOUANE=STE COMPTOIR MS DU CAP BON (56,341 DT)
# STE 428 = 0 (Sotusfa non disponible)
# ══════════════════════════════════════════════════════════════════
_INTRANTS_2026 = {
    # Valeurs réelles — 5 DETAIL_VENTE + distribution ACHREF
    'ABDELFATEH BEN SLIMEN': 3295.119, 'ABDELHAKIM MEJRI': 2417.350,
    'ABDELKADER KALBOUSI': 2912.624, 'ABDELKARIM TRABELSSI': 17203.843,
    'ABDELMALAK NAJJAR': 6362.049, 'ABEDLAZIZ LAYARI': 19611.604,
    'ABEDRAZEK BEY': 14235.349, 'ABEDSATTAR HATBI': 6078.333,
    'ABEDSATTAR MATHLOUTHI': 12485.405, 'ABELSAMII MANSOURI': 9054.022,
    'ACHREF BEN SASSI': 9937.408, 'ACHREF HATBI': 2302.163,
    'ADEL JAZI': 6389.126, 'AHMED BALLAGUI': 15377.383,
    'AHMED BEN ALAYA': 30550.680, 'AHMED BEN CHIKH': 5723.274,
    'AHMED HMIDEN': 6094.325, 'AHMED IDRISSI': 26651.463,
    'AHMED MANSOURI': 9054.022, 'AHMED SASSI': 53528.566,
    'AKAREM HEMMEDI': 25452.185, 'ALAEDINE KILENI': 6634.099,
    'ALI KOTLI': 14806.682, 'AMAR GARMALAH': 65276.027,
    'ANIS DHAWADI': 4818.502, 'ANIS RAYES': 10555.796,
    'ARBI JABALI': 5403.997, 'AYMEN BEN OTHMEN': 4567.670,
    'AYMEN CHABEN': 7547.209, 'AYMEN HATTAB': 25190.721,
    'AYMEN SAIDI': 13552.400, 'AZOUZ BEN MASSOUD': 7230.849,
    'BADIA SAAFI': 24172.289, 'BILEL GHA 1': 14322.017,
    'BILEL GHA 2': 17187.021, 'BILEL GHA 3': 17187.021,
    'BILEL GHA 4': 15277.019, 'BILEL KEHIL': 8593.010,
    'BOUBAKER FILALI': 20692.349, 'CHOKRI MANSOURI': 9054.022,
    'DIVERS CLIENTS': 655.000, 'ELIFA MANSOURI': 9054.022,
    'EZZEDINE GUESMI': 224984.289, 'FAOUZI ANTRI': 335.949,
    'FETHI LEHBIBI': 22591.717, 'FETHI SDIRI': 26405.535,
    'HABIB BELWEAR': 79780.617, 'HABIB MAKHLOUF': 825.700,
    'HAFEDH MOSBEH': 153812.239, 'HAMED BEN YOUNIS': 13643.726,
    'HAMMADI BENZRIBIA': 3547.951, 'HAMMADI TRABISI': 3816.583,
    'HANI BEN KILANI': 4000.976, 'HASSAN BEN HAJD FRAJ': 7787.928,
    'HASSEN BEN ALIA': 19245.535, 'HEDI SLAMA': 80555.684,
    'HICHEM SAAFI': 4551.700, 'HSAN GARMALAH': 32638.013,
    'HSSIN HATBI': 2580.577, 'HSSINE BRINI': 7745.632,
    'IBRAHIM BEN BOUBAKER': 11639.821, 'IBRAHIM KILENI': 6463.247,
    'IMED TRABILSI': 10831.750, 'ISKANDER BEN SALAH': 11172.867,
    'JABER BEN DHIA': 10761.623, 'JAMEL GARMALAH': 16319.007,
    'JAMIL ALAYA': 15333.467, 'KAIS DHAOUI': 49777.537,
    'KAIS ELBAKOUCHE': 8483.440, 'KAMEL CHOUCHEN': 4231.974,
    'KAMEL TRABELSSI': 3529.787, 'KARIM GARMALAH 1': 65277.027,
    'KARIM GARMALAH 2': 65277.027, 'KHALED BELHAJ': 14690.013,
    'KHAMES JABALI': 8645.995, 'LAMINE MANSOURI': 12073.029,
    'LASSED NEILI': 10582.837, 'LAZHER HAJ MOULDI': 3442.652,
    'LOTFY HAJIJ': 4117.899, 'MAHER BELHAJ FRAJ': 9983.297,
    'MAHER BELHAJ SALAH': 33975.685, 'MAKRAM HAFFAR': 26458.648,
    'MAKREM MBARKI': 9054.022, 'MED ALI GARMALAH': 65277.027,
    'MED MARWENE MAJDOUB': 32099.873, 'MOHAMED ALI GHZELA': 7692.721,
    'MOHAMED ALI MBAREK': 8074.149, 'MOHAMED ALI SELMI': 9999.998,
    'MOHAMED AOUINI': 5554.825, 'MOHAMED BEDIA NEJI': 5627.279,
    'MOHAMED BEL MADHI': 6306.897, 'MOHAMED BEN HSSAN': 3313.895,
    'MOHAMED BEN MOUAOUIA': 88516.635, 'MOHAMED BEN SAID': 14228.353,
    'MOHAMED GARMALAH': 16319.007, 'MOHAMED GHARBI': 10466.098,
    'MOHAMED ILYES BEN OTHMEN': 4005.451, 'MOHAMED MANNOUBI': 1042.650,
    'MOHAMED RHIM': 2317.500, 'MOHAMED THAMER BEN ALAYA': 7939.499,
    'MOHAMED ZIADI': 2612.971, 'MONCEF ELMAJDOUB': 10635.486,
    'MOUAOUIA MOKTAR': 7677.326, 'MOUEZ BEN ISSA': 1408.000,
    'MOUEZ ESSAAFI': 18569.597, 'MOUNIR BEY': 35161.198,
    'MOURAD HEMMEDI': 35274.241, 'MOURAD MANSOURI': 45273.109,
    'NABIL BEN HSSAN': 6838.799, 'NADER BEN AICHA': 31896.916,
    'NAJIB BACCOUCH': 5230.390, 'NAJMEDDINE BEN SALAH': 11928.803,
    'NEGI ZAAFOURI': 119150.897, 'NIZAR BOUOUD': 2211.749,
    'NIZAR MANAA': 25567.943, 'NOOMEN ECHAGRAOUI': 143132.481,
    'NOUREDIN MANSOURI': 18109.044, 'OTHMEN DHIBI': 95170.590,
    'RAMDHAN MHEDHBI': 185047.213, 'RAMZI HAMDOUN': 7481.149,
    'RAMZI MATHLOUTHI': 7943.208, 'RIADH BEN SAID': 12705.045,
    'RIADH BEN ZBIR': 13647.227, 'ROMDHAN SAAFI': 7022.986,
    'SABER KHARBESH': 15052.009, 'SALEH BEN HAMOUDA': 12450.000,
    'SALEM EL MEJRI': 78995.336, 'SALEM LEGRERI': 1309.975,
    'SAMEH BACCOUCH': 12730.330, 'SAMI DAKHLAOUI': 13193.628,
    'SAMI FERGENI': 38404.338, 'SAMI KAAB': 28641.164,
    'SAMI LASMAR': 5846.750, 'SAMIR ATTIYA': 101107.132,
    'SASSI MANSOUR': 2674.646, 'SEBTI JABALI': 69169.956,
    'SLAH BANI': 7602.972, 'SLAH BEN ABDALLAH': 1590.164,
    'SLAH HATBI': 5534.132, 'SLIM MARZOUGUI': 37586.258,
    'SOFIENNE GHZELA': 2783.129, 'SOUHAIL BOUZANA': 27241.181,
    'STE BACCARA': 30260.485, 'STE KERKOUANE S A': 56340.518,
    'STE SEMAG': 938.196, 'TAHER MANSOURI': 9054.022,
    'TAHER MATHLOUTHI': 13023.147, 'TALEB JABLAH': 3241.998,
    'TAREK BEN ABDALAH': 3338.703, 'TAREK BEN NJI': 2504.900,
    'TAREK EL BAHRI': 5690.425, 'ZOUHAIR BAICH': 22268.364,
    'ZOUHAIR BEN ECHIK': 20549.127,

    # ── Ajouts ACHREF (distribution Ha) + FEDI manquants ──
    'ABDELKADER OMRANI': 14704.061,
    'ABDELKARIM SAAD': 9802.707,
    'ABDELRAOUF BOUALEGUE': 10782.978,
    'ABELSAMII MANSOURI': 8325.951,
    'AHMED BRAYKIA': 14704.061,
    'AHMED MANSOURI': 8325.951,
    'ALI LTIFI': 6135.021,
    'AMAR GARMALAH': 24540.085,
    'ANAS ZAYENI': 13497.047,
    'ARBI JABALI': 6404.663,
    'BASSEM ZIDI': 12253.384,
    'BECHA REDHWENI': 4901.354,
    'BILEL GHA 1': 13193.834,
    'BILEL GHA 2': 13193.834,
    'BILEL GHA 3': 32984.585,
    'BILEL GHA 4': 13193.834,
    'BILEL KEHIL': 4901.354,
    'BORNI BOUALEGUE': 7352.030,
    'CHIHEB OMRANI': 14704.061,
    'CHOKRI MANSOURI': 8325.951,
    'ELIFA MANSOURI': 8325.951,
    'FAYSEL GHOBTAN': 11763.248,
    'FEDI AMAYMIA': 6135.021,
    'HAFEDH MOSBEH': 61266.919,
    'HAMZA AMAYMIA': 12270.043,
    'HAYTHEM AMAYMIA': 12270.043,
    'HSAN GARMALAH': 12270.043,
    'IBRHIM GWEDRIA': 36810.128,
    'ILYES MANSOUR': 7352.030,
    'ISAMAIL ZIDI': 14704.061,
    'JAMEL GARMALAH': 6135.021,
    'KARIM AMAR': 4901.354,
    'KARIM GARMALAH 1': 24540.085,
    'KARIM GARMALAH 2': 24540.085,
    'KHAMES JABALI': 9606.994,
    'LAMINE MANSOURI': 12488.927,
    'LESWED TLILI': 30675.107,
    'LOAY GHOBTAN': 9802.707,
    'MAHER BOUALEGUE': 9802.707,
    'MAKREM MBARKI': 8325.951,
    'MED ALI GARMALAH': 24540.085,
    'MOHAMED GARMALAH': 6135.021,
    'MOHAMED LEHKIMI': 6447.252,
    'MOHAMED SLIMEN': 12270.043,
    'MOHSEN CHEWECH': 10782.978,
    'MOHSEN OMRANI': 9802.707,
    'MOURAD BELGACEM': 9802.707,
    'MOURAD MANSOURI': 41629.756,
    'NADER OMRANI': 14704.061,
    'NASREDIN ZIDI': 24506.768,
    'NOUREDIN MANSOURI': 16651.902,
    'RADHWEN AMAYMIA': 6135.021,
    'RADHWEN BOUALEGUE': 9802.707,
    'RASLEN BEN SALAH': 6397.428,
    'REBAH SMOUD': 6135.021,
    'RIDHA AMAYMIA': 12270.043,
    'SEBTI JABALI': 70451.289,
    'SLAH SAAD': 10782.978,
    'TAHER MANSOURI': 8325.951,
    'TALEB JABLAH': 6135.021,
    'WISSEM AMAYMIA': 24540.085,
    'YASIN MNASRI': 85773.687,
    'YASIN TLILI': 6135.021,
}

# ══════════════════════════════════════════════════════════════════
# AVANCES ET REPORTS RÉELS 2026 — source RECAP El Bourak
# 163 entrées avances | 61 entrées reports
# ══════════════════════════════════════════════════════════════════
_AVANCES_2026 = {
    'ABDELFATEH BEN SLIMEN': 900.00, 'ABDELHAMID ELMESSAI': 6702.50,
    'ABDELKADER KALBOUSI': 784.00, 'ABDELKARIM TRABELSSI': 3572.00,
    'ABDELMALAK NAJJAR': 6000.00, 'ABEDSATTAR HATBI': 668.00,
    'ABEDSATTAR MATHLOUTHI': 5392.50, 'ABEDLAZIZ LAYARI': 31608.00,
    'ABELSAMII MANSOURI': 4415.14, 'ACHREF BEN SASSI': 15484.50,
    'ACHREF HATBI': 5056.00, 'ADEL JAZI': 3228.00,
    'AHMED BALLAGUI': 10000.00, 'AHMED BEN ALAYA': 43050.50,
    'AHMED HMIDEN': 7105.00, 'AHMED IDRISSI': 4000.00,
    'AHMED MANSOURI': 4415.14, 'AHMED SASSI': 12000.00,
    'AKAREM HEMMEDI': 33541.00, 'ALAEDINE KILENI': 6300.00,
    'ALI KOTLI': 6257.00, 'AMAR GARMALAH': 45199.58,
    'ANIS RAYES': 6760.00, 'ARBI JABALI': 6184.57,
    'AYMEN BEN OTHMEN': 3480.00, 'AYMEN HATTAB': 8080.00,
    'AYMEN SAIDI': 18492.00, 'AZOUZ BEN MASSOUD': 3285.00,
    'BADIA SAAFI': 10000.00, 'BILEL GHA 1': 5920.95,
    'BILEL GHA 2': 7105.39, 'BILEL GHA 3': 7105.39,
    'BILEL GHA 4': 6315.77, 'BILEL KEHIL': 3552.49,
    'BOUBAKER FILALI': 15000.00, 'CHOKRI MANSOURI': 4415.14,
    'ELIFA MANSOURI': 4415.14, 'EZZEDINE GUESMI': 55000.00,
    'FAOUZI ANTRI': 4270.00, 'FETHI LEHBIBI': 16204.00,
    'FETHI SDIRI': 11500.00, 'HABIB BELWEAR': 30000.00,
    'HAFEDH MOSBEH': 131970.87, 'HAMED BEN YOUNIS': 4520.00,
    'HAMMADI TRABISI': 2000.00, 'HANI BEN KILANI': 1892.50,
    'HASSAN BEN HAJD FRAJ': 2000.00, 'HEDI SLAMA': 51000.00,
    'HICHEM SAAFI': 3000.00, 'HSAN GARMALAH': 22599.79,
    'HSSINE BRINI': 8925.00, 'IBRAHIM BEN BOUBAKER': 2000.00,
    'IBRAHIM KILENI': 1892.50, 'IMED AMDOUNI': 50000.00,
    'IMED TRABILSI': 2000.00, 'ISKANDER BEN SALAH': 4098.00,
    'ISSAM KOUKI': 0.00, 'JABER BEN DHIA': 4200.00,
    'JAMEL GARMALAH': 11299.90, 'JAMIL ALAYA': 5522.50,
    'KAIS DHAOUI': 33176.00, 'KAIS ELBAKOUCHE': 471.00,
    'KAMEL TRABELSSI': 2250.00, 'KARIM GARMALAH 1': 45200.28,
    'KARIM GARMALAH 2': 45200.28, 'KHALED BELHAJ': 2850.00,
    'KHAMES JABALI': 9894.86, 'LAMINE MANSOURI': 5887.35,
    'LASSED NEILI': 4544.00, 'LAZHER HAJ MOULDI': 1000.00,
    'LOTFI TRABELSI': 36000.00, 'MAHER BELHAJ FRAJ': 7904.00,
    'MAHER BELHAJ SALAH': 7000.00, 'MAKRAM HAFFAR': 750.50,
    'MAHMOUD MESSADI': 2200.00, 'MED ALI GARMALAH': 45200.28,
    'MED MARWENE MAJDOUB': 6000.00, 'MOEZ BEN ABDALLAH': 6000.00,
    'MOHAMED ALI MBAREK': 4756.50, 'MOHAMED ALI SELMI': 10000.00,
    'MOHAMED BEDIA NEJI': 10385.00, 'MOHAMED BEN HEDI MEHEMDI': 3392.50,
    'MOHAMED BEN HSSAN': 895.00, 'MOHAMED BEN MOUAOUIA': 180000.00,
    'MOHAMED BEN SAID': 5000.00, 'MOHAMED GARMALAH': 11299.90,
    'MOHAMED GHARBI': 7247.00, 'MOHAMED ILYES BEN OTHMEN': 1800.00,
    'MOHAMED AOUINI': 3000.00, 'MOHAMED LEHKIMI': 3716.00,
    'MOHAMED RHIM': 5000.00, 'MOHAMED THAMER BEN ALAYA': 8433.50,
    'MOHAMED ZIADI': 2800.00, 'MONCEF ELMAJDOUB': 7850.00,
    'MOUAOUIA MOKTAR': 4500.00, 'MOUEZ BEN ISSA': 3000.00,
    'MOUEZ ESSAAFI': 3190.00, 'MOHAMED ALI GHZELA': 6414.00,
    'MOUNIR BEY': 4500.00, 'MOURAD HEMMEDI': 35027.00,
    'MOURAD MANSOURI': 22077.18, 'NABIL BEN HSSAN': 3231.00,
    'NADER BEN AICHA': 17413.50, 'NAJMEDDINE BEN SALAH': 2755.00,
    'NEGI ZAAFOURI': 135854.00, 'NAJIB BACCOUCH': 2000.00,
    'NIZAR BOUOUD': 787.50, 'NIZAR MANAA': 9753.00,
    'NOOMEN ECHAGRAOUI': 33158.00, 'NOUREDIN MANSOURI': 8830.77,
    'OTHMEN DHIBI': 82293.00, 'RAMDHAN MHEDHBI': 123260.00,
    'RAMZI HAMDOUN': 1320.00, 'RAMZI MATHLOUTHI': 2956.00,
    'RASLEN BEN SALAH': 3475.00, 'RIADH BEN SAID': 9856.00,
    'RIADH BEN ZBIR': 19125.00, 'ROMDHAN SAAFI': 2285.00,
    'SABER KHARBESH': 1827.22, 'SALEH BEN HAMOUDA': 25730.50,
    'SALEM EL MEJRI': 65195.00, 'SALEM LEGRERI': 475.00,
    'SAMEH BACCOUCH': 12129.50, 'SAMI KAAB': 7500.00,
    'SAMI DAKHLAOUI': 7255.00, 'SAMI FERGENI': 26528.00,
    'SAMI LASMAR': 3192.00, 'SAMIR ATTIYA': 78097.00,
    'SASSI MANSOUR': 3270.00, 'SEBTI JABALI': 79161.14,
    'SLAH BANI': 4332.00, 'SLAH HATBI': 2711.00,
    'SLIM MARZOUGUI': 10000.00, 'SOFIENNE GHZELA': 1588.00,
    'SOUHAIL BOUZANA': 15200.00, 'STE 428 SERVICES AGRICOLES': 50000.00,
    'STE BACCARA': 210000.00, 'STE KERKOUANE S A': 150000.00,
    'STE SEMAG': 3891.00, 'TAHER MANSOURI': 4415.14,
    'TAHER MATHLOUTHI': 549.50, 'TALEB JABLAH': 3710.29,
    'TAREK EL BAHRI': 1000.00, 'ZOUHAIR BAICH': 10000.00,
    'ZOUHAIR BEN ECHIK': 3858.00, 'ABDELKADER YEDES': 2850.00,
}

_REPORTS_2026 = {
    'ABDELHAMID ELMESSAI': 4030.00, 'ABDELKARIM TRABELSSI': 376.00,
    'ABDELMALAK NAJJAR': 50.99, 'AHMED BEN ALAYA': 1109.00,
    'AHMED HMIDEN': 5154.77, 'AHMED SASSI': 2834.60,
    'BADIA SAAFI': 573.00, 'BOUBAKER FILALI': 3104.56,
    'FETHI LEHBIBI': 586.36, 'FETHI SDIRI': 79.30,
    'HASSEN BEN ALIA': 223.68, 'IMED TRABILSI': 1477.52,
    'ISKANDER BEN SALAH': 1000.00, 'LASSED NEILI': 17890.25,
    'LOTFI ZOGLAMI': 16318.00, 'MARWEN SELMI': 2289.96,
    'MOHAMED MANNOUBI': 9979.01, 'MOHAMED BEN HSSAN': 5031.62,
    'MOHAMED BEN SAID': 80.76, 'MOHAMED THAMER BEN ALAYA': 2033.44,
    'MOUAOUIA MOKTAR': 244.92, 'MOUEZ ESSAAFI': 553.00,
    'NACEUR SALLAMI': 74391.60, 'RAMDHAN MHEDHBI': 21891.87,
    'SALEM EL MEJRI': 104159.93, 'SAMI LASMAR': 173.20,
    'SLIM MARZOUGUI': 42667.29, 'SOUHAIL BOUZANA': 26344.58,
    'STE KERKOUANE S A': 1565.90, 'TAIEB MEDSIA': 46554.34,
    'ZOUHAIR BEN ECHIK': 529.90, 'ABDELKADER YEDES': 1200.00,
}


_PREVISION_2026 = {
    'ABDELFATEH BEN SLIMEN': {'ha': 1.0, 'ton': 75.0, 'region': 'CAP BON 1', 'zone': 'dar allouch', 'acces': 'TRC/PPL/PL', 'date_debut': '2026-07-20', 'date_fin': '2026-07-24', 'usine': 'COMOCAP', 'centre': 'nan'},
    'ABDELHAKIM MEJRI': {'ha': 1.5, 'ton': 112.5, 'region': 'CAP BON 1', 'zone': 'dar allouch', 'acces': 'TRC/PPL/PL', 'date_debut': '2026-07-15', 'date_fin': '2026-07-20', 'usine': 'COMOCAP', 'centre': 'nan'},
    'ABDELKADER KALBOUSI': {'ha': 3.0, 'ton': 180.0, 'region': 'CAP BON 2', 'zone': 'TEFELOUN', 'acces': 'PL', 'date_debut': '2026-07-15', 'date_fin': '2026-07-23', 'usine': 'COMOCAP', 'centre': 'nan'},
    'ABDELKADER MANNA': {'ha': 1.0, 'ton': 75.0, 'region': 'CAP BON 1', 'zone': 'dar allouch', 'acces': 'TRC/PPL/PL', 'date_debut': '2026-07-28', 'date_fin': '2026-08-01', 'usine': 'COMOCAP', 'centre': 'nan'},
    'ABDELKADER OMRANI': {'ha': 3.0, 'ton': 270.0, 'region': 'GAFSA / KASSRINE', 'zone': 'OULED OMRAN', 'acces': 'SEMI', 'date_debut': '2026-07-08', 'date_fin': '2026-07-15', 'usine': 'SICAM', 'centre': 'HAFEDH MOSBEH'},
    'ABDELKADER YEDES': {'ha': 25.0, 'ton': 1500.0, 'region': 'CAP BON 2', 'zone': 'KORBA/SOMAA', 'acces': 'PL', 'date_debut': '2026-07-01', 'date_fin': '2026-07-23', 'usine': 'ELFALLEH', 'centre': 'ABDELKADER YEDES'},
    'ABDELKARIM SAAD': {'ha': 2.0, 'ton': 180.0, 'region': 'GAFSA / KASSRINE', 'zone': 'SIDI AICH', 'acces': 'SEMI', 'date_debut': '2026-08-01', 'date_fin': '2026-08-05', 'usine': 'SICAM', 'centre': 'HAFEDH MOSBEH'},
    'ABDELKARIM TRABELSSI': {'ha': 4.0, 'ton': 300.0, 'region': 'CAP BON 1', 'zone': 'dar allouch', 'acces': 'TRC/PPL/PL', 'date_debut': '2026-07-09', 'date_fin': '2026-07-23', 'usine': 'COMOCAP', 'centre': 'nan'},
    'ABDELMALAK NAJJAR': {'ha': 4.0, 'ton': 300.0, 'region': 'CAP BON 1', 'zone': 'dar allouch', 'acces': 'TRC/PPL/PL', 'date_debut': '2026-07-09', 'date_fin': '2026-07-23', 'usine': 'COMOCAP', 'centre': 'nan'},
    'ABDELRAOUF BOUALEGUE': {'ha': 2.2, 'ton': 198.0, 'region': 'GAFSA / KASSRINE', 'zone': 'SIDI AICH', 'acces': 'SEMI', 'date_debut': '2026-08-08', 'date_fin': '2026-08-11', 'usine': 'SICAM', 'centre': 'HAFEDH MOSBEH'},
    'ABDESLEM BEN SOUISSI': {'ha': 2.0, 'ton': 150.0, 'region': 'CAP BON 1', 'zone': 'dar allouch', 'acces': 'PL/PPL', 'date_debut': '2026-07-09', 'date_fin': '2026-07-18', 'usine': 'COMOCAP', 'centre': 'nan'},
    'ABEDLAZIZ LAYARI': {'ha': 2.5, 'ton': 175.0, 'region': 'CAP BON 2', 'zone': 'DIAR HOJJEJ', 'acces': 'PL', 'date_debut': '2026-07-13', 'date_fin': '2026-07-21', 'usine': 'TUCAL', 'centre': 'nan'},
    'ABEDLAZIZ LAYARI RAMZI': {'ha': 6.0, 'ton': 420.0, 'region': 'CAP BON 2', 'zone': 'DIAR HOJJEJ', 'acces': 'PL', 'date_debut': '2026-07-03', 'date_fin': '2026-07-20', 'usine': 'SICAM', 'centre': 'nan'},
    'ABEDRAZEK BEY': {'ha': 7.0, 'ton': 490.0, 'region': 'CAP BON 2', 'zone': 'TEFELOUN', 'acces': 'PL', 'date_debut': '2026-06-22', 'date_fin': '2026-07-13', 'usine': 'TUCAL', 'centre': 'nan'},
    'ABEDSATTAR HATBI': {'ha': 1.5, 'ton': 105.0, 'region': 'CAP BON 2', 'zone': 'HTOUBA', 'acces': 'PL', 'date_debut': '2026-07-08', 'date_fin': '2026-07-12', 'usine': 'COMOCAP', 'centre': 'nan'},
    'ABEDSATTAR MATHLOUTHI': {'ha': 2.0, 'ton': 140.0, 'region': 'CAP BON 2', 'zone': 'DIAR HOJJEJ', 'acces': 'PL', 'date_debut': '2026-07-08', 'date_fin': '2026-07-14', 'usine': 'SICAM', 'centre': 'nan'},
    'ABELSAMII MANSOURI': {'ha': 1.0, 'ton': 90.0, 'region': 'GAFSA / KASSRINE', 'zone': 'MAJEL BELABESS', 'acces': 'SEMI', 'date_debut': '2026-08-01', 'date_fin': '2026-08-04', 'usine': 'TUCAL', 'centre': 'MOURAD MANSOURI'},
    'ACHREF BEN SASSI': {'ha': 4.0, 'ton': 280.0, 'region': 'CAP BON 2', 'zone': 'OUED CHIBA', 'acces': 'PL', 'date_debut': '2026-06-25', 'date_fin': '2026-07-12', 'usine': 'SICAM', 'centre': 'nan'},
    'ACHREF HATBI': {'ha': 2.0, 'ton': 140.0, 'region': 'CAP BON 2', 'zone': 'FARTOUNA', 'acces': 'PL', 'date_debut': '2026-07-20', 'date_fin': '2026-07-25', 'usine': 'COMOCAP', 'centre': 'nan'},
    'ADEL JAZI': {'ha': 1.5, 'ton': 105.0, 'region': 'CAP BON 2', 'zone': 'OUED CHIBA', 'acces': 'PL', 'date_debut': '2026-07-05', 'date_fin': '2026-07-09', 'usine': 'TUCAL', 'centre': 'nan'},
    'AHMED ATTIA': {'ha': 1.5, 'ton': 112.5, 'region': 'CAP BON 1', 'zone': 'dar allouch', 'acces': 'TRC/PPL', 'date_debut': '2026-07-22', 'date_fin': '2026-08-02', 'usine': 'COMOCAP', 'centre': 'nan'},
    'AHMED BALLAGUI': {'ha': 6.0, 'ton': 390.0, 'region': 'NORD', 'zone': 'Bou Salem', 'acces': 'PL/SEMI', 'date_debut': '2026-07-25', 'date_fin': '2026-08-04', 'usine': 'SICAM', 'centre': 'nan'},
    'AHMED BEN ALAYA': {'ha': 6.0, 'ton': 420.0, 'region': 'CAP BON 2', 'zone': 'OUED KHATEF', 'acces': 'PL', 'date_debut': '2026-07-13', 'date_fin': '2026-07-25', 'usine': 'SICAM', 'centre': 'nan'},
    'AHMED BEN CHIKH': {'ha': 2.0, 'ton': 150.0, 'region': 'CAP BON 1', 'zone': 'dar allouch', 'acces': 'TRC/PPL/PL', 'date_debut': '2026-07-13', 'date_fin': '2026-07-22', 'usine': 'COMOCAP', 'centre': 'nan'},
    'AHMED BRAYKIA': {'ha': 3.0, 'ton': 240.0, 'region': 'GAFSA / KASSRINE', 'zone': 'OULED ZID', 'acces': 'SEMI', 'date_debut': '2026-07-25', 'date_fin': '2026-08-01', 'usine': 'SICAM', 'centre': 'HAFEDH MOSBEH'},
    'AHMED HMIDEN': {'ha': 2.5, 'ton': 175.0, 'region': 'CAP BON 2', 'zone': 'TEFELOUN', 'acces': 'PL', 'date_debut': '2026-07-03', 'date_fin': '2026-07-12', 'usine': 'SICAM', 'centre': 'nan'},
    'AHMED IDRISSI': {'ha': 2.0, 'ton': 150.0, 'region': 'CAP BON 1', 'zone': 'dar allouch', 'acces': 'PL/PPL', 'date_debut': '2026-07-05', 'date_fin': '2026-07-14', 'usine': 'COMOCAP', 'centre': 'nan'},
    'AHMED MANSOURI': {'ha': 1.0, 'ton': 90.0, 'region': 'GAFSA / KASSRINE', 'zone': 'MAJEL BELABESS', 'acces': 'SEMI', 'date_debut': '2026-08-01', 'date_fin': '2026-08-04', 'usine': 'SICAM', 'centre': 'MOURAD MANSOURI'},
    'AHMED SASSI': {'ha': 1.0, 'ton': 75.0, 'region': 'CAP BON 1', 'zone': 'dar allouch', 'acces': 'TRC/PPL/PL', 'date_debut': '2026-07-14', 'date_fin': '2026-07-18', 'usine': 'COMOCAP', 'centre': 'nan'},
    'AKAREM HEMMEDI': {'ha': 6.0, 'ton': 480.0, 'region': 'KAIROUAN', 'zone': 'BATTEN', 'acces': 'RM', 'date_debut': '2026-06-25', 'date_fin': '2026-06-29', 'usine': 'SICAM', 'centre': 'nan'},
    'ALAEDINE KILENI': {'ha': 1.5, 'ton': 105.0, 'region': 'CAP BON 2', 'zone': 'BIR MASOUDA', 'acces': 'PL', 'date_debut': '2026-07-15', 'date_fin': '2026-07-19', 'usine': 'SICAM', 'centre': 'nan'},
    'ALI KOTLI': {'ha': 5.0, 'ton': 350.0, 'region': 'CAP BON 2', 'zone': 'KORBA', 'acces': 'PL', 'date_debut': '2026-07-10', 'date_fin': '2026-07-25', 'usine': 'SICAM', 'centre': 'nan'},
    'ALI LTIFI': {'ha': 1.0, 'ton': 90.0, 'region': 'GAFSA / KASSRINE', 'zone': 'AMAYMIA', 'acces': 'SEMI', 'date_debut': '2026-07-20', 'date_fin': '2026-08-07', 'usine': 'TUCAL', 'centre': 'KARIM GARMALAH'},
    'AMAR GARMALAH': {'ha': 4.0, 'ton': 360.0, 'region': 'GAFSA / KASSRINE', 'zone': 'AMAYMIA', 'acces': 'SEMI', 'date_debut': '2026-07-15', 'date_fin': '2026-07-20', 'usine': 'TUCAL', 'centre': 'KARIM GARMALAH'},
    'AMOR KHECHIN': {'ha': 24.0, 'ton': 1416.0, 'region': 'KAIROUAN', 'zone': 'BATTEN', 'acces': 'RM', 'date_debut': '2026-06-20', 'date_fin': '2026-06-25', 'usine': 'SICAM', 'centre': 'AMOR KHECHIN'},
    'ANAS ZAYENI': {'ha': 2.2, 'ton': 198.0, 'region': 'GAFSA / KASSRINE', 'zone': 'AMAYMIA', 'acces': 'SEMI', 'date_debut': '2026-07-01', 'date_fin': '2026-07-07', 'usine': 'SICAM', 'centre': 'KARIM GARMALAH'},
    'ANIS DHAWADI': {'ha': 1.0, 'ton': 75.0, 'region': 'CAP BON 1', 'zone': 'dar allouch', 'acces': 'TRC/PPL/PL', 'date_debut': '2026-07-10', 'date_fin': '2026-07-14', 'usine': 'COMOCAP', 'centre': 'nan'},
    'ANIS RAYES': {'ha': 3.5, 'ton': 245.0, 'region': 'CAP BON 2', 'zone': 'BIR MASOUDA', 'acces': 'PL', 'date_debut': '2026-06-25', 'date_fin': '2026-07-02', 'usine': 'TUCAL', 'centre': 'nan'},
    'ARBI JABALI': {'ha': 2.0, 'ton': 150.0, 'region': 'GAFSA / KASSRINE', 'zone': 'FERIANA', 'acces': 'SEMI', 'date_debut': '2026-08-20', 'date_fin': '2026-08-24', 'usine': 'ABIDA', 'centre': 'SEBTI JABALI'},
    'AYMEN BEN OTHMEN': {'ha': 2.5, 'ton': 150.0, 'region': 'CAP BON 2', 'zone': 'GARAT SASSI', 'acces': 'PL', 'date_debut': '2026-06-29', 'date_fin': '2026-07-06', 'usine': 'SICAM', 'centre': 'nan'},
    'AYMEN CHABEN': {'ha': 2.5, 'ton': 175.0, 'region': 'CAP BON 2', 'zone': 'DIAR HOJJEJ', 'acces': 'PL', 'date_debut': '2026-07-10', 'date_fin': '2026-07-18', 'usine': 'COMOCAP', 'centre': 'nan'},
    'AYMEN HATTAB': {'ha': 12.0, 'ton': 840.0, 'region': 'CAP BON 2', 'zone': 'OUED CHIBA', 'acces': 'PL', 'date_debut': '2026-07-15', 'date_fin': '2026-08-01', 'usine': 'TUCAL', 'centre': 'nan'},
    'AYMEN SAIDI': {'ha': 4.5, 'ton': 360.0, 'region': 'CAP BON 2', 'zone': 'TEFELOUN', 'acces': 'PL', 'date_debut': '2026-07-01', 'date_fin': '2026-07-13', 'usine': 'SICAM', 'centre': 'nan'},
    'AZAIZ BEN ISSA': {'ha': 5.0, 'ton': 375.0, 'region': 'CAP BON 1', 'zone': 'dar allouch', 'acces': 'TRC/PPL/PL', 'date_debut': '2026-07-18', 'date_fin': '2026-08-05', 'usine': 'SICAM', 'centre': 'nan'},
    'AZOUZ BEN MASSOUD': {'ha': 2.5, 'ton': 187.5, 'region': 'CAP BON 1', 'zone': 'dar allouch', 'acces': 'TRC/PPL', 'date_debut': '2026-07-09', 'date_fin': '2026-07-20', 'usine': 'COMOCAP', 'centre': 'nan'},
    'BADIA SAAFI': {'ha': 3.0, 'ton': 225.0, 'region': 'CAP BON 1', 'zone': 'dar allouch', 'acces': 'TRC/PPL/PL', 'date_debut': '2026-07-09', 'date_fin': '2026-07-21', 'usine': 'COMOCAP', 'centre': 'nan'},
    'BASSEM ZIDI': {'ha': 2.5, 'ton': 150.0, 'region': 'GAFSA / KASSRINE', 'zone': 'OULED ZID', 'acces': 'SEMI', 'date_debut': '2026-07-10', 'date_fin': '2026-07-15', 'usine': 'SICAM', 'centre': 'HAFEDH MOSBEH'},
    'BECHA REDHWENI': {'ha': 1.0, 'ton': 90.0, 'region': 'GAFSA / KASSRINE', 'zone': 'SIDI AICH', 'acces': 'SEMI', 'date_debut': '2026-08-02', 'date_fin': '2026-08-05', 'usine': 'ABIDA', 'centre': 'HAFEDH MOSBEH'},
    'BILEL GHA 1': {'ha': 2.0, 'ton': 160.0, 'region': 'GAFSA / KASSRINE', 'zone': 'SIDI AICH', 'acces': 'SEMI', 'date_debut': '2026-07-01', 'date_fin': '2026-07-04', 'usine': 'SICAM', 'centre': 'BILEL GHA'},
    'BILEL GHA 2': {'ha': 2.0, 'ton': 160.0, 'region': 'GAFSA / KASSRINE', 'zone': 'SIDI AICH', 'acces': 'SEMI', 'date_debut': '2026-07-10', 'date_fin': '2026-07-14', 'usine': 'SICAM', 'centre': 'BILEL GHA'},
    'BILEL GHA 3': {'ha': 5.0, 'ton': 400.0, 'region': 'GAFSA / KASSRINE', 'zone': 'SIDI AICH', 'acces': 'SEMI', 'date_debut': '2026-07-04', 'date_fin': '2026-07-10', 'usine': 'SICAM', 'centre': 'BILEL GHA'},
    'BILEL GHA 4': {'ha': 2.0, 'ton': 160.0, 'region': 'GAFSA / KASSRINE', 'zone': 'SIDI AICH', 'acces': 'SEMI', 'date_debut': '2026-07-14', 'date_fin': '2026-07-18', 'usine': 'ABIDA', 'centre': 'BILEL GHA'},
    'BILEL KEHIL': {'ha': 1.0, 'ton': 90.0, 'region': 'GAFSA / KASSRINE', 'zone': 'MANZEL GAMOUDI', 'acces': 'SEMI', 'date_debut': '2026-07-10', 'date_fin': '2026-07-14', 'usine': 'SICAM', 'centre': 'HAFEDH MOSBEH'},
    'BORNI BOUALEGUE': {'ha': 1.5, 'ton': 135.0, 'region': 'GAFSA / KASSRINE', 'zone': 'SIDI AICH', 'acces': 'SEMI', 'date_debut': '2026-08-02', 'date_fin': '2026-08-06', 'usine': 'ABIDA', 'centre': 'HAFEDH MOSBEH'},
    'BOUBAKER FILALI': {'ha': 4.0, 'ton': 340.0, 'region': 'KAIROUAN', 'zone': 'MENZEL MHIRI', 'acces': 'PL/SEMI', 'date_debut': '2026-07-08', 'date_fin': '2026-07-18', 'usine': 'ABIDA', 'centre': 'nan'},
    'CHIHEB OMRANI': {'ha': 3.0, 'ton': 210.0, 'region': 'GAFSA / KASSRINE', 'zone': 'OULED OMRAN', 'acces': 'SEMI', 'date_debut': '2026-07-02', 'date_fin': '2026-07-08', 'usine': 'SICAM', 'centre': 'HAFEDH MOSBEH'},
    'CHOKRI MANSOURI': {'ha': 1.0, 'ton': 90.0, 'region': 'GAFSA / KASSRINE', 'zone': 'MAJEL BELABESS', 'acces': 'SEMI', 'date_debut': '2026-08-14', 'date_fin': '2026-08-16', 'usine': 'TUCAL', 'centre': 'MOURAD MANSOURI'},
    'DIVERS CLIENT': {'ha': 6.0, 'ton': 400.0, 'region': 'CAP BON 1', 'zone': 'dar allouch', 'acces': 'PL', 'date_debut': '2026-07-10', 'date_fin': '2026-07-21', 'usine': 'SICAM', 'centre': 'DIVERS CLIENT'},
    'DIVERS CLIENTS': {'ha': 25.0, 'ton': 1800.0, 'region': 'GAFSA / KASSRINE', 'zone': 'FERIANA', 'acces': 'SEMI', 'date_debut': '2026-07-10', 'date_fin': '2026-07-27', 'usine': 'ABIDA', 'centre': 'DIVERS CLIENTS'},
    'ELIFA MANSOURI': {'ha': 1.0, 'ton': 90.0, 'region': 'GAFSA / KASSRINE', 'zone': 'MAJEL BELABESS', 'acces': 'SEMI', 'date_debut': '2026-08-06', 'date_fin': '2026-08-09', 'usine': 'SICAM', 'centre': 'MOURAD MANSOURI'},
    'EZZEDINE GUESMI': {'ha': 40.0, 'ton': 2345.0, 'region': 'KAIROUAN', 'zone': 'ZAAFRANA-ELKHADHRA', 'acces': 'PL/SEMI', 'date_debut': '2026-06-20', 'date_fin': '2026-07-07', 'usine': 'SICAM', 'centre': 'EZZEDINE GUESMI'},
    'FAOUZI ANTRI': {'ha': 1.5, 'ton': 105.0, 'region': 'CAP BON 2', 'zone': 'DIAR HOJJEJ', 'acces': 'PL', 'date_debut': '2026-07-15', 'date_fin': '2026-07-19', 'usine': 'COMOCAP', 'centre': 'nan'},
    'FAYSEL GHOBTAN': {'ha': 2.4, 'ton': 216.0, 'region': 'GAFSA / KASSRINE', 'zone': 'SIDI AICH', 'acces': 'SEMI', 'date_debut': '2026-07-22', 'date_fin': '2026-07-26', 'usine': 'SICAM', 'centre': 'HAFEDH MOSBEH'},
    'FEDI AMAYMIA': {'ha': 1.0, 'ton': 90.0, 'region': 'GAFSA / KASSRINE', 'zone': 'AMAYMIA', 'acces': 'SEMI', 'date_debut': '2026-07-05', 'date_fin': '2026-07-08', 'usine': 'SICAM', 'centre': 'KARIM GARMALAH'},
    'FETHI LEHBIBI': {'ha': 7.0, 'ton': 420.0, 'region': 'NORD', 'zone': 'bor amri', 'acces': 'PL/SEMI', 'date_debut': '2026-07-16', 'date_fin': '2026-07-27', 'usine': 'COMOCAP', 'centre': 'nan'},
    'FETHI SDIRI': {'ha': 7.0, 'ton': 490.0, 'region': 'NORD', 'zone': 'Gar Dimaou', 'acces': 'PL/SEMI', 'date_debut': '2026-07-25', 'date_fin': '2026-08-02', 'usine': 'SICAM', 'centre': 'nan'},
    'HABIB BELWEAR': {'ha': 18.0, 'ton': 1150.0, 'region': 'CAP BON 2', 'zone': 'lebna', 'acces': 'PL/PPL', 'date_debut': '2026-06-29', 'date_fin': '2026-07-17', 'usine': 'SICAM', 'centre': 'HABIB BELWEAR'},
    'HABIB MAKHLOUF': {'ha': 1.0, 'ton': 60.0, 'region': 'CAP BON 2', 'zone': 'GARAT SASSI', 'acces': 'PL', 'date_debut': '2026-07-25', 'date_fin': '2026-07-27', 'usine': 'COMOCAP', 'centre': 'nan'},
    'HAFEDH MOSBEH': {'ha': 12.5, 'ton': 1125.0, 'region': 'GAFSA / KASSRINE', 'zone': 'SIDI AICH', 'acces': 'RM', 'date_debut': '2026-07-25', 'date_fin': '2026-08-02', 'usine': 'SICAM', 'centre': 'HAFEDH MOSBEH'},
    'HAMED BEN YOUNIS': {'ha': 1.0, 'ton': 75.0, 'region': 'CAP BON 1', 'zone': 'dar allouch', 'acces': 'TRC/PPL/PL', 'date_debut': '2026-07-10', 'date_fin': '2026-07-14', 'usine': 'COMOCAP', 'centre': 'nan'},
    'HAMED HAMAMMI': {'ha': 1.0, 'ton': 75.0, 'region': 'CAP BON 1', 'zone': 'dar allouch', 'acces': 'TRC/PPL/PL', 'date_debut': '2026-07-10', 'date_fin': '2026-07-14', 'usine': 'COMOCAP', 'centre': 'nan'},
    'HAMMADI BENZRIBIA': {'ha': 1.0, 'ton': 75.0, 'region': 'CAP BON 1', 'zone': 'dar allouch', 'acces': 'PL/PPL', 'date_debut': '2026-07-23', 'date_fin': '2026-07-30', 'usine': 'SICAM', 'centre': 'nan'},
    'HAMMADI TRABISI': {'ha': 2.0, 'ton': 150.0, 'region': 'CAP BON 1', 'zone': 'dar allouch', 'acces': 'TRC/PPL/PL', 'date_debut': '2026-07-03', 'date_fin': '2026-07-12', 'usine': 'COMOCAP', 'centre': 'nan'},
    'HAMZA AMAYMIA': {'ha': 2.0, 'ton': 180.0, 'region': 'GAFSA / KASSRINE', 'zone': 'AMAYMIA', 'acces': 'SEMI', 'date_debut': '2026-07-31', 'date_fin': '2026-08-03', 'usine': 'SICAM', 'centre': 'KARIM GARMALAH'},
    'HANI BEN KILANI': {'ha': 1.5, 'ton': 112.5, 'region': 'CAP BON 1', 'zone': 'dar allouch', 'acces': 'PL/PPL', 'date_debut': '2026-07-15', 'date_fin': '2026-07-21', 'usine': 'COMOCAP', 'centre': 'nan'},
    'HASSAN BEN HAJD FRAJ': {'ha': 1.5, 'ton': 112.5, 'region': 'CAP BON 1', 'zone': 'dar allouch', 'acces': 'TRC/PPL', 'date_debut': '2026-07-23', 'date_fin': '2026-07-30', 'usine': 'COMOCAP', 'centre': 'nan'},
    'HASSEN BEN ALIA': {'ha': 15.0, 'ton': 960.0, 'region': 'CAP BON 2', 'zone': 'lebna', 'acces': 'PL/PPL', 'date_debut': '2026-07-02', 'date_fin': '2026-07-21', 'usine': 'TUCAL', 'centre': 'HASSEN BEN ALIA'},
    'HAYTHEM AMAYMIA': {'ha': 2.0, 'ton': 180.0, 'region': 'GAFSA / KASSRINE', 'zone': 'AMAYMIA', 'acces': 'SEMI', 'date_debut': '2026-08-03', 'date_fin': '2026-08-09', 'usine': 'ABIDA', 'centre': 'KARIM GARMALAH'},
    'HEDI SLAMA': {'ha': 14.0, 'ton': 1020.0, 'region': 'BOUFICHA', 'zone': 'SIDI KHELIFA', 'acces': 'PL', 'date_debut': '2026-06-29', 'date_fin': '2026-07-24', 'usine': 'COMOCAP', 'centre': 'nan'},
    'HICHEM SAAFI': {'ha': 2.0, 'ton': 150.0, 'region': 'CAP BON 1', 'zone': 'dar allouch', 'acces': 'TRC/PPL/PL', 'date_debut': '2026-07-01', 'date_fin': '2026-07-18', 'usine': 'SICAM', 'centre': 'nan'},
    'HICHEM TRABELSI': {'ha': 1.5, 'ton': 112.5, 'region': 'CAP BON 1', 'zone': 'dar allouch', 'acces': 'TRC/PPL/PL', 'date_debut': '2026-07-11', 'date_fin': '2026-07-16', 'usine': 'COMOCAP', 'centre': 'nan'},
    'HOUSSEM BRAYEK': {'ha': 1.0, 'ton': 75.0, 'region': 'CAP BON 1', 'zone': 'dar allouch', 'acces': 'TRC/PPL/PL', 'date_debut': '2026-07-10', 'date_fin': '2026-07-14', 'usine': 'COMOCAP', 'centre': 'nan'},
    'HSAN GARMALAH': {'ha': 2.0, 'ton': 180.0, 'region': 'GAFSA / KASSRINE', 'zone': 'AMAYMIA', 'acces': 'SEMI', 'date_debut': '2026-07-01', 'date_fin': '2026-07-05', 'usine': 'TUCAL', 'centre': 'KARIM GARMALAH'},
    'HSSIN HATBI': {'ha': 2.0, 'ton': 120.0, 'region': 'CAP BON 2', 'zone': 'FARTOUNA', 'acces': 'PL', 'date_debut': '2026-07-25', 'date_fin': '2026-07-30', 'usine': 'COMOCAP', 'centre': 'nan'},
    'HSSINE BRINI': {'ha': 1.0, 'ton': 75.0, 'region': 'CAP BON 1', 'zone': 'dar allouch', 'acces': 'TRC/PPL', 'date_debut': '2026-07-23', 'date_fin': '2026-07-30', 'usine': 'COMOCAP', 'centre': 'nan'},
    'IBRAHIM BEN BOUBAKER': {'ha': 1.6, 'ton': 120.0, 'region': 'CAP BON 1', 'zone': 'dar allouch', 'acces': 'TRC/PPL/PL', 'date_debut': '2026-07-23', 'date_fin': '2026-07-30', 'usine': 'COMOCAP', 'centre': 'nan'},
    'IBRAHIM KILENI': {'ha': 2.0, 'ton': 120.0, 'region': 'CAP BON 2', 'zone': 'DIAR HOJJEJ', 'acces': 'PL', 'date_debut': '2026-07-18', 'date_fin': '2026-07-23', 'usine': 'COMOCAP', 'centre': 'nan'},
    'IBRHIM GWEDRIA': {'ha': 6.0, 'ton': 540.0, 'region': 'GAFSA / KASSRINE', 'zone': 'AMAYMIA', 'acces': 'SEMI', 'date_debut': '2026-07-01', 'date_fin': '2026-07-15', 'usine': 'SICAM', 'centre': 'KARIM GARMALAH'},
    'ILYES MANSOUR': {'ha': 1.5, 'ton': 135.0, 'region': 'GAFSA / KASSRINE', 'zone': 'SIDI AICH', 'acces': 'SEMI', 'date_debut': '2026-07-10', 'date_fin': '2026-07-15', 'usine': 'SICAM', 'centre': 'HAFEDH MOSBEH'},
    'IMED AMDOUNI': {'ha': 10.0, 'ton': 600.0, 'region': 'NORD', 'zone': 'Jandouba', 'acces': 'PL/SEMI', 'date_debut': '2026-08-01', 'date_fin': '2026-08-11', 'usine': 'SICAM', 'centre': 'nan'},
    'IMED TRABILSI': {'ha': 2.0, 'ton': 150.0, 'region': 'CAP BON 1', 'zone': 'dar allouch', 'acces': 'TRC/PPL/PL', 'date_debut': '2026-07-01', 'date_fin': '2026-07-18', 'usine': 'COMOCAP', 'centre': 'nan'},
    'ISAMAIL ZIDI': {'ha': 3.0, 'ton': 180.0, 'region': 'GAFSA / KASSRINE', 'zone': 'OULED ZID', 'acces': 'SEMI', 'date_debut': '2026-07-10', 'date_fin': '2026-07-16', 'usine': 'SICAM', 'centre': 'HAFEDH MOSBEH'},
    'ISKANDER BEN SALAH': {'ha': 1.0, 'ton': 60.0, 'region': 'CAP BON 2', 'zone': 'TBAG', 'acces': 'PL', 'date_debut': '2026-07-13', 'date_fin': '2026-07-16', 'usine': 'SICAM', 'centre': 'nan'},
    'ISSAM KOUKI': {'ha': 15.0, 'ton': 900.0, 'region': 'NORD', 'zone': 'Sidi Ismail', 'acces': 'PL/SEMI', 'date_debut': '2026-07-19', 'date_fin': '2026-08-04', 'usine': 'SICAM', 'centre': 'nan'},
    'JABER BEN DHIA': {'ha': 3.5, 'ton': 210.0, 'region': 'CAP BON 2', 'zone': 'SOMAA', 'acces': 'PL', 'date_debut': '2026-07-10', 'date_fin': '2026-07-19', 'usine': 'COMOCAP', 'centre': 'nan'},
    'JALEL RHIM': {'ha': 2.0, 'ton': 160.0, 'region': 'CAP BON 2', 'zone': 'SOMAA', 'acces': 'PL', 'date_debut': '2026-06-24', 'date_fin': '2026-07-01', 'usine': 'SICAM', 'centre': 'nan'},
    'JAMEL GARMALAH': {'ha': 1.0, 'ton': 90.0, 'region': 'GAFSA / KASSRINE', 'zone': 'AMAYMIA', 'acces': 'SEMI', 'date_debut': '2026-08-03', 'date_fin': '2026-08-06', 'usine': 'TUCAL', 'centre': 'KARIM GARMALAH'},
    'JAMIL ALAYA': {'ha': 3.0, 'ton': 240.0, 'region': 'CAP BON 2', 'zone': 'DIAR HOJJEJ', 'acces': 'PL', 'date_debut': '2026-07-03', 'date_fin': '2026-07-09', 'usine': 'SICAM', 'centre': 'nan'},
    'KAIS DHAOUI': {'ha': 10.0, 'ton': 561.0, 'region': 'SIDI BOUZID', 'zone': 'SIDI BOUZID', 'acces': 'PL', 'date_debut': '2026-06-29', 'date_fin': '2026-07-10', 'usine': 'ABIDA', 'centre': 'nan'},
    'KAIS ELBAKOUCHE': {'ha': 2.0, 'ton': 150.0, 'region': 'CAP BON 1', 'zone': 'dar allouch', 'acces': 'TRC/PPL/PL', 'date_debut': '2026-07-28', 'date_fin': '2026-08-06', 'usine': 'COMOCAP', 'centre': 'nan'},
    'KAIS MHATLI': {'ha': 3.0, 'ton': 210.0, 'region': 'CAP BON 2', 'zone': 'SIDI HASSOUN', 'acces': 'PL', 'date_debut': '2026-07-13', 'date_fin': '2026-07-23', 'usine': 'SICAM', 'centre': 'nan'},
    'KAMEL CHOUCHEN': {'ha': 2.0, 'ton': 120.0, 'region': 'CAP BON 2', 'zone': 'OUED CHIBA', 'acces': 'PL', 'date_debut': '2026-07-15', 'date_fin': '2026-07-23', 'usine': 'SICAM', 'centre': 'nan'},
    'KAMEL TRABELSSI': {'ha': 1.0, 'ton': 75.0, 'region': 'CAP BON 1', 'zone': 'dar allouch', 'acces': 'TRC/PPL/PL', 'date_debut': '2026-07-18', 'date_fin': '2026-07-24', 'usine': 'COMOCAP', 'centre': 'nan'},
    'KARIM AMAR': {'ha': 1.0, 'ton': 90.0, 'region': 'GAFSA / KASSRINE', 'zone': 'OULED ZID', 'acces': 'SEMI', 'date_debut': '2026-07-20', 'date_fin': '2026-07-23', 'usine': 'ABIDA', 'centre': 'HAFEDH MOSBEH'},
    'KARIM GARMALAH 1': {'ha': 4.0, 'ton': 360.0, 'region': 'GAFSA / KASSRINE', 'zone': 'AMAYMIA', 'acces': 'SEMI', 'date_debut': '2026-07-12', 'date_fin': '2026-07-18', 'usine': 'SICAM', 'centre': 'KARIM GARMALAH'},
    'KARIM GARMALAH 2': {'ha': 4.0, 'ton': 360.0, 'region': 'GAFSA / KASSRINE', 'zone': 'AMAYMIA', 'acces': 'SEMI', 'date_debut': '2026-06-28', 'date_fin': '2026-07-06', 'usine': 'TUCAL', 'centre': 'KARIM GARMALAH'},
    'KHALED BELHAJ': {'ha': 24.0, 'ton': 1440.0, 'region': 'CAP BON 2', 'zone': 'GOURCHIN', 'acces': 'PL', 'date_debut': '2026-07-01', 'date_fin': '2026-07-18', 'usine': 'TUCAL', 'centre': 'KHALED BELHAJ'},
    'KHALED CHATER': {'ha': 10.0, 'ton': 800.0, 'region': 'CAP BON 2', 'zone': 'GROMBELIA', 'acces': 'PL', 'date_debut': '2026-07-20', 'date_fin': '2026-08-01', 'usine': 'TUCAL', 'centre': 'nan'},
    'KHAMES JABALI': {'ha': 3.0, 'ton': 225.0, 'region': 'GAFSA / KASSRINE', 'zone': 'FERIANA', 'acces': 'SEMI', 'date_debut': '2026-08-22', 'date_fin': '2026-08-29', 'usine': 'ABIDA', 'centre': 'SEBTI JABALI'},
    'LAMINE MANSOURI': {'ha': 1.5, 'ton': 135.0, 'region': 'GAFSA / KASSRINE', 'zone': 'MAJEL BELABESS', 'acces': 'SEMI', 'date_debut': '2026-08-16', 'date_fin': '2026-08-20', 'usine': 'ABIDA', 'centre': 'MOURAD MANSOURI'},
    'LASSED NEILI': {'ha': 2.0, 'ton': 80.0, 'region': 'CAP BON 2', 'zone': 'DIAR HOJJEJ/KHARREZ', 'acces': 'PL', 'date_debut': '2026-06-28', 'date_fin': '2026-07-01', 'usine': 'TUCAL', 'centre': 'nan'},
    'LAZHER HAJ MOULDI': {'ha': 1.0, 'ton': 75.0, 'region': 'CAP BON 1', 'zone': 'dar allouch', 'acces': 'TRC/PPL/PL', 'date_debut': '2026-07-10', 'date_fin': '2026-07-14', 'usine': 'COMOCAP', 'centre': 'nan'},
    'LESWED TLILI': {'ha': 5.0, 'ton': 450.0, 'region': 'GAFSA / KASSRINE', 'zone': 'AMAYMIA', 'acces': 'SEMI', 'date_debut': '2026-08-15', 'date_fin': '2026-08-25', 'usine': 'ABIDA', 'centre': 'KARIM GARMALAH'},
    'LOAY GHOBTAN': {'ha': 2.0, 'ton': 180.0, 'region': 'GAFSA / KASSRINE', 'zone': 'SIDI AICH', 'acces': 'SEMI', 'date_debut': '2026-07-20', 'date_fin': '2026-07-25', 'usine': 'SICAM', 'centre': 'HAFEDH MOSBEH'},
    'LOTFI TRABELSI': {'ha': 43.0, 'ton': 2600.0, 'region': 'NORD', 'zone': 'Jandouba', 'acces': 'PL/SEMI', 'date_debut': '2026-07-23', 'date_fin': '2026-08-10', 'usine': 'COMOCAP', 'centre': 'Lotfi Trabelsi'},
    'LOTFY HAJIJ': {'ha': 3.0, 'ton': 240.0, 'region': 'CAP BON 2', 'zone': 'ATHLETH/HTOUBA', 'acces': 'PL', 'date_debut': '2026-07-13', 'date_fin': '2026-07-24', 'usine': 'COMOCAP', 'centre': 'nan'},
    'MAHER BELHAJ FRAJ': {'ha': 3.5, 'ton': 245.0, 'region': 'CAP BON 2', 'zone': 'DIAR HOJJEJ', 'acces': 'PL', 'date_debut': '2026-07-12', 'date_fin': '2026-07-20', 'usine': 'COMOCAP', 'centre': 'nan'},
    'MAHER BELHAJ SALAH': {'ha': 5.0, 'ton': 375.0, 'region': 'CAP BON 1', 'zone': 'dar allouch', 'acces': 'TRC/PPL', 'date_debut': '2026-07-09', 'date_fin': '2026-07-27', 'usine': 'COMOCAP', 'centre': 'nan'},
    'MAHER BOUALEGUE': {'ha': 2.0, 'ton': 180.0, 'region': 'GAFSA / KASSRINE', 'zone': 'SIDI AICH', 'acces': 'SEMI', 'date_debut': '2026-08-01', 'date_fin': '2026-08-06', 'usine': 'SICAM', 'centre': 'HAFEDH MOSBEH'},
    'MAHMOUD MESSADI': {'ha': 4.0, 'ton': 280.0, 'region': 'CAP BON 2', 'zone': 'GARAT SASSI', 'acces': 'PL', 'date_debut': '2026-07-10', 'date_fin': '2026-07-18', 'usine': 'TUCAL', 'centre': 'nan'},
    'MAKRAM HAFFAR': {'ha': 8.0, 'ton': 640.0, 'region': 'CAP BON 2', 'zone': 'DIAR HOJJEJ', 'acces': 'PL', 'date_debut': '2026-07-05', 'date_fin': '2026-07-17', 'usine': 'TUCAL', 'centre': 'nan'},
    'MAKREM MBARKI': {'ha': 1.0, 'ton': 90.0, 'region': 'GAFSA / KASSRINE', 'zone': 'MAJEL BELABESS', 'acces': 'SEMI', 'date_debut': '2026-08-16', 'date_fin': '2026-08-19', 'usine': 'SICAM', 'centre': 'MOURAD MANSOURI'},
    'MED ALI GARMALAH': {'ha': 4.0, 'ton': 360.0, 'region': 'GAFSA / KASSRINE', 'zone': 'AMAYMIA', 'acces': 'SEMI', 'date_debut': '2026-07-01', 'date_fin': '2026-07-10', 'usine': 'SICAM', 'centre': 'KARIM GARMALAH'},
    'MED MARWENE MAJDOUB': {'ha': 2.5, 'ton': 187.5, 'region': 'CAP BON 1', 'zone': 'dar allouch', 'acces': 'PL/PPL', 'date_debut': '2026-07-28', 'date_fin': '2026-08-06', 'usine': 'TUCAL', 'centre': 'nan'},
    'MED TAHER': {'ha': 2.5, 'ton': 187.5, 'region': 'CAP BON 1', 'zone': 'dar allouch', 'acces': 'PL/PPL', 'date_debut': '2026-07-09', 'date_fin': '2026-07-21', 'usine': 'COMOCAP', 'centre': 'nan'},
    'MOEZ BEN ABDALLAH': {'ha': 5.0, 'ton': 325.0, 'region': 'NORD', 'zone': 'Sidi Othman', 'acces': 'PL/SEMI', 'date_debut': '2026-08-01', 'date_fin': '2026-08-10', 'usine': 'TUCAL', 'centre': 'nan'},
    'MOHAMED ALI GHZELA': {'ha': 2.0, 'ton': 160.0, 'region': 'CAP BON 2', 'zone': 'BENI AYECH', 'acces': 'PL', 'date_debut': '2026-07-15', 'date_fin': '2026-07-23', 'usine': 'TUCAL', 'centre': 'nan'},
    'MOHAMED ALI MBAREK': {'ha': 2.0, 'ton': 140.0, 'region': 'CAP BON 2', 'zone': 'ATHLETH', 'acces': 'PL', 'date_debut': '2026-07-10', 'date_fin': '2026-07-22', 'usine': 'COMOCAP', 'centre': 'nan'},
    'MOHAMED ALI SELMI': {'ha': 10.0, 'ton': 700.0, 'region': 'KAIROUAN', 'zone': 'KAIROUAN', 'acces': 'PL', 'date_debut': '2026-06-22', 'date_fin': '2026-07-10', 'usine': 'SICAM', 'centre': 'nan'},
    'MOHAMED AMAYMIA': {'ha': 2.0, 'ton': 180.0, 'region': 'GAFSA / KASSRINE', 'zone': 'AMAYMIA', 'acces': 'SEMI', 'date_debut': '2026-07-18', 'date_fin': '2026-07-22', 'usine': 'TUCAL', 'centre': 'KARIM GARMALAH'},
    'MOHAMED AOUINI': {'ha': 1.5, 'ton': 120.0, 'region': 'CAP BON 2', 'zone': 'TEFELOUN', 'acces': 'PL', 'date_debut': '2026-07-07', 'date_fin': '2026-07-12', 'usine': 'SICAM', 'centre': 'nan'},
    'MOHAMED BEDIA NEJI': {'ha': 4.0, 'ton': 320.0, 'region': 'CAP BON 2', 'zone': 'FRININ', 'acces': 'PL', 'date_debut': '2026-07-12', 'date_fin': '2026-07-22', 'usine': 'SICAM', 'centre': 'nan'},
    'MOHAMED BEL MADHI': {'ha': 5.0, 'ton': 400.0, 'region': 'CAP BON 2', 'zone': 'DIAR HOJJEJ', 'acces': 'PL', 'date_debut': '2026-07-07', 'date_fin': '2026-07-18', 'usine': 'SICAM', 'centre': 'nan'},
    'MOHAMED BEN HEDI MEHEMDI': {'ha': 2.5, 'ton': 187.5, 'region': 'CAP BON 1', 'zone': 'dar allouch', 'acces': 'PL/PPL', 'date_debut': '2026-07-28', 'date_fin': '2026-08-06', 'usine': 'SICAM', 'centre': 'nan'},
    'MOHAMED BEN HSSAN': {'ha': 1.0, 'ton': 80.0, 'region': 'CAP BON 2', 'zone': 'GARAT SASSI', 'acces': 'PL', 'date_debut': '2026-07-08', 'date_fin': '2026-07-11', 'usine': 'SICAM', 'centre': 'nan'},
    'MOHAMED BEN MOUAOUIA': {'ha': 90.0, 'ton': 4700.0, 'region': 'CAP BON 2', 'zone': 'menzel horr', 'acces': 'PL/PPL', 'date_debut': '2026-06-25', 'date_fin': '2026-07-24', 'usine': 'SICAM', 'centre': 'MOHAMED BEN MOUAOUIA'},
    'MOHAMED BEN SAID': {'ha': 2.5, 'ton': 200.0, 'region': 'CAP BON 2', 'zone': 'OUED CHIBA', 'acces': 'PL', 'date_debut': '2026-07-01', 'date_fin': '2026-07-10', 'usine': 'TUCAL', 'centre': 'nan'},
    'MOHAMED GARMALAH': {'ha': 1.0, 'ton': 90.0, 'region': 'GAFSA / KASSRINE', 'zone': 'AMAYMIA', 'acces': 'SEMI', 'date_debut': '2026-07-25', 'date_fin': '2026-07-29', 'usine': 'SICAM', 'centre': 'KARIM GARMALAH'},
    'MOHAMED GHARBI': {'ha': 3.0, 'ton': 240.0, 'region': 'CAP BON 2', 'zone': 'TEFELOUN', 'acces': 'PL', 'date_debut': '2026-07-05', 'date_fin': '2026-07-16', 'usine': 'SICAM', 'centre': 'nan'},
    'MOHAMED ILYES BEN OTHMEN': {'ha': 2.5, 'ton': 150.0, 'region': 'CAP BON 2', 'zone': 'GARAT SASSI', 'acces': 'PL', 'date_debut': '2026-07-08', 'date_fin': '2026-07-15', 'usine': 'SICAM', 'centre': 'nan'},
    'MOHAMED LEHKIMI': {'ha': 1.5, 'ton': 112.5, 'region': 'CAP BON 1', 'zone': 'dar allouch', 'acces': 'TRC/PPL/PL', 'date_debut': '2026-07-13', 'date_fin': '2026-07-19', 'usine': 'TUCAL', 'centre': 'nan'},
    'MOHAMED MANNOUBI': {'ha': 2.0, 'ton': 120.0, 'region': 'CAP BON 2', 'zone': 'BOUJRIDA', 'acces': 'PL', 'date_debut': '2026-07-30', 'date_fin': '2026-08-04', 'usine': 'SICAM', 'centre': 'nan'},
    'MOHAMED RHIM': {'ha': 2.0, 'ton': 80.0, 'region': 'CAP BON 2', 'zone': 'TEFELOUN', 'acces': 'PL', 'date_debut': '2026-07-10', 'date_fin': '2026-07-13', 'usine': 'SICAM', 'centre': 'nan'},
    'MOHAMED SLIMEN': {'ha': 2.0, 'ton': 180.0, 'region': 'GAFSA / KASSRINE', 'zone': 'AMAYMIA', 'acces': 'SEMI', 'date_debut': '2026-07-10', 'date_fin': '2026-07-16', 'usine': 'SICAM', 'centre': 'KARIM GARMALAH'},
    'MOHAMED THAMER BEN ALAYA': {'ha': 3.0, 'ton': 180.0, 'region': 'CAP BON 2', 'zone': 'GOURCHIN', 'acces': 'PL', 'date_debut': '2026-07-05', 'date_fin': '2026-07-19', 'usine': 'SICAM', 'centre': 'nan'},
    'MOHAMED ZIADI': {'ha': 1.0, 'ton': 60.0, 'region': 'CAP BON 2', 'zone': 'OUED CHIBA', 'acces': 'PL', 'date_debut': '2026-07-13', 'date_fin': '2026-07-15', 'usine': 'SICAM', 'centre': 'nan'},
    'MOHSEN CHEWECH': {'ha': 2.2, 'ton': 198.0, 'region': 'GAFSA / KASSRINE', 'zone': 'SIDI AICH', 'acces': 'SEMI', 'date_debut': '2026-07-10', 'date_fin': '2026-07-16', 'usine': 'SICAM', 'centre': 'HAFEDH MOSBEH'},
    'MOHSEN OMRANI': {'ha': 2.0, 'ton': 160.0, 'region': 'GAFSA / KASSRINE', 'zone': 'OULED OMRAN', 'acces': 'SEMI', 'date_debut': '2026-07-12', 'date_fin': '2026-07-17', 'usine': 'SICAM', 'centre': 'HAFEDH MOSBEH'},
    'MONCEF ELMAJDOUB': {'ha': 3.0, 'ton': 225.0, 'region': 'CAP BON 1', 'zone': 'dar allouch', 'acces': 'PL/PPL', 'date_debut': '2026-07-23', 'date_fin': '2026-08-03', 'usine': 'COMOCAP', 'centre': 'nan'},
    'MOUAOUIA MOKTAR': {'ha': 2.5, 'ton': 187.5, 'region': 'CAP BON 1', 'zone': 'dar allouch', 'acces': 'PL/PPL', 'date_debut': '2026-07-09', 'date_fin': '2026-07-17', 'usine': 'COMOCAP', 'centre': 'nan'},
    'MOUEZ BEN ISSA': {'ha': 4.0, 'ton': 300.0, 'region': 'CAP BON 1', 'zone': 'dar allouch', 'acces': 'PL/PPL', 'date_debut': '2026-07-09', 'date_fin': '2026-07-22', 'usine': 'COMOCAP', 'centre': 'nan'},
    'MOUEZ ESSAAFI': {'ha': 3.0, 'ton': 225.0, 'region': 'CAP BON 1', 'zone': 'dar allouch', 'acces': 'TRC/PPL/PL', 'date_debut': '2026-07-10', 'date_fin': '2026-07-20', 'usine': 'SICAM', 'centre': 'nan'},
    'MOUHAMED AOUINET': {'ha': 1.0, 'ton': 75.0, 'region': 'CAP BON 1', 'zone': 'dar allouch', 'acces': 'TRC/PPL/PL', 'date_debut': '2026-07-22', 'date_fin': '2026-07-26', 'usine': 'COMOCAP', 'centre': 'nan'},
    'MOUHAMED MESSII': {'ha': 3.5, 'ton': 262.5, 'region': 'CAP BON 1', 'zone': 'dar allouch', 'acces': 'PL/PPL', 'date_debut': '2026-07-13', 'date_fin': '2026-07-25', 'usine': 'COMOCAP', 'centre': 'nan'},
    'MOUHAMED TRABELSI': {'ha': 2.0, 'ton': 150.0, 'region': 'CAP BON 1', 'zone': 'dar allouch', 'acces': 'PL/PPL', 'date_debut': '2026-07-13', 'date_fin': '2026-07-19', 'usine': 'COMOCAP', 'centre': 'nan'},
    'MOUNIR BEY': {'ha': 13.5, 'ton': 870.0, 'region': 'CAP BON 2', 'zone': 'TEFELOUN/DIAR HOJJEJ', 'acces': 'PL', 'date_debut': '2026-07-05', 'date_fin': '2026-07-16', 'usine': 'COMOCAP', 'centre': 'nan'},
    'MOURAD BELGACEM': {'ha': 2.0, 'ton': 120.0, 'region': 'GAFSA / KASSRINE', 'zone': 'SIDI AICH', 'acces': 'SEMI', 'date_debut': '2026-07-14', 'date_fin': '2026-07-18', 'usine': 'SICAM', 'centre': 'HAFEDH MOSBEH'},
    'MOURAD HEMMEDI': {'ha': 8.5, 'ton': 720.0, 'region': 'KAIROUAN', 'zone': 'AWAMRIYA', 'acces': 'RM', 'date_debut': '2026-07-01', 'date_fin': '2026-07-07', 'usine': 'SICAM', 'centre': 'nan'},
    'MOURAD MANSOURI': {'ha': 5.0, 'ton': 450.0, 'region': 'GAFSA / KASSRINE', 'zone': 'MAJEL BELABESS', 'acces': 'SEMI', 'date_debut': '2026-08-16', 'date_fin': '2026-08-22', 'usine': 'SICAM', 'centre': 'MOURAD MANSOURI'},
    'NABIL BEN HSSAN': {'ha': 1.5, 'ton': 120.0, 'region': 'CAP BON 2', 'zone': 'DIAR HOJJEJ', 'acces': 'PL', 'date_debut': '2026-07-16', 'date_fin': '2026-07-21', 'usine': 'COMOCAP', 'centre': 'nan'},
    'NADER BEN AICHA': {'ha': 9.0, 'ton': 540.0, 'region': 'CAP BON 2', 'zone': 'TEFELOUN', 'acces': 'PL', 'date_debut': '2026-06-25', 'date_fin': '2026-07-12', 'usine': 'SICAM', 'centre': 'nan'},
    'NADER OMRANI': {'ha': 3.0, 'ton': 270.0, 'region': 'GAFSA / KASSRINE', 'zone': 'OULED OMRAN', 'acces': 'SEMI', 'date_debut': '2026-07-28', 'date_fin': '2026-08-04', 'usine': 'TUCAL', 'centre': 'HAFEDH MOSBEH'},
    'NAJIB BACCOUCH': {'ha': 1.0, 'ton': 75.0, 'region': 'CAP BON 1', 'zone': 'dar allouch', 'acces': 'TRC/PPL', 'date_debut': '2026-07-10', 'date_fin': '2026-07-14', 'usine': 'COMOCAP', 'centre': 'nan'},
    'NAJMEDDINE BEN SALAH': {'ha': 2.5, 'ton': 187.5, 'region': 'CAP BON 1', 'zone': 'dar allouch', 'acces': 'PL/PPL', 'date_debut': '2026-07-28', 'date_fin': '2026-08-09', 'usine': 'COMOCAP', 'centre': 'nan'},
    'NASREDIN ZIDI': {'ha': 5.0, 'ton': 450.0, 'region': 'GAFSA / KASSRINE', 'zone': 'MANZEL GAMOUDI', 'acces': 'RM', 'date_debut': '2026-07-15', 'date_fin': '2026-07-22', 'usine': 'SICAM', 'centre': 'HAFEDH MOSBEH'},
    'NEGI ZAAFOURI': {'ha': 28.5, 'ton': 2025.0, 'region': 'SIDI BOUZID', 'zone': 'ZAAFRIA', 'acces': 'PL/SEMI', 'date_debut': '2026-07-01', 'date_fin': '2026-07-26', 'usine': 'SICAM', 'centre': 'nan'},
    'NEJIB MECHRGUI': {'ha': 4.0, 'ton': 240.0, 'region': 'NORD', 'zone': 'Bellarigia', 'acces': 'PL/SEMI', 'date_debut': '2026-08-04', 'date_fin': '2026-08-10', 'usine': 'SICAM', 'centre': 'nan'},
    'NIZAR BOUOUD': {'ha': 1.0, 'ton': 60.0, 'region': 'CAP BON 2', 'zone': 'GOMBAR', 'acces': 'PL', 'date_debut': '2026-07-15', 'date_fin': '2026-07-20', 'usine': 'COMOCAP', 'centre': 'nan'},
    'NIZAR MANAA': {'ha': 9.5, 'ton': 712.5, 'region': 'CAP BON 1', 'zone': 'dar allouch', 'acces': 'PL/PPL', 'date_debut': '2026-07-23', 'date_fin': '2026-08-07', 'usine': 'SICAM', 'centre': 'nan'},
    'NOOMEN ECHAGRAOUI': {'ha': 13.0, 'ton': 975.0, 'region': 'CAP BON 1', 'zone': 'dar allouch', 'acces': 'TRC/PPL', 'date_debut': '2026-07-03', 'date_fin': '2026-07-23', 'usine': 'COMOCAP', 'centre': 'nan'},
    'NOUREDIN MANSOURI': {'ha': 2.0, 'ton': 180.0, 'region': 'GAFSA / KASSRINE', 'zone': 'MAJEL BELABESS', 'acces': 'SEMI', 'date_debut': '2026-08-04', 'date_fin': '2026-08-09', 'usine': 'SICAM', 'centre': 'MOURAD MANSOURI'},
    'OMAR HAMEMI': {'ha': 15.0, 'ton': 900.0, 'region': 'CAP BON 2', 'zone': 'KORBA', 'acces': 'PL', 'date_debut': '2026-06-25', 'date_fin': '2026-07-14', 'usine': 'TUCAL', 'centre': 'OMAR HAMEMI'},
    'OTHMEN DHIBI': {'ha': 25.0, 'ton': 1754.0, 'region': 'SIDI BOUZID', 'zone': 'OM ADHAM', 'acces': 'PL', 'date_debut': '2026-06-29', 'date_fin': '2026-07-25', 'usine': 'ABIDA', 'centre': 'nan'},
    'RADHWEN AMAYMIA': {'ha': 1.0, 'ton': 90.0, 'region': 'GAFSA / KASSRINE', 'zone': 'AMAYMIA', 'acces': 'SEMI', 'date_debut': '2026-07-25', 'date_fin': '2026-07-28', 'usine': 'TUCAL', 'centre': 'KARIM GARMALAH'},
    'RADHWEN BOUALEGUE': {'ha': 2.0, 'ton': 180.0, 'region': 'GAFSA / KASSRINE', 'zone': 'SIDI AICH', 'acces': 'SEMI', 'date_debut': '2026-08-04', 'date_fin': '2026-08-09', 'usine': 'SICAM', 'centre': 'HAFEDH MOSBEH'},
    'RAMDHAN MHEDHBI': {'ha': 25.0, 'ton': 1795.0, 'region': 'BOUFICHA', 'zone': 'SIDI SAIID', 'acces': 'PL', 'date_debut': '2026-06-28', 'date_fin': '2026-07-23', 'usine': 'TUCAL', 'centre': 'RAMDHAN MHEDHBI'},
    'RAMZI HAMDOUN': {'ha': 3.0, 'ton': 210.0, 'region': 'CAP BON 2', 'zone': 'TEFELOUN', 'acces': 'PL', 'date_debut': '2026-06-30', 'date_fin': '2026-07-07', 'usine': 'TUCAL', 'centre': 'nan'},
    'RAMZI MATHLOUTHI': {'ha': 3.0, 'ton': 210.0, 'region': 'CAP BON 2', 'zone': 'DIAR HOJJEJ', 'acces': 'PL', 'date_debut': '2026-07-05', 'date_fin': '2026-07-11', 'usine': 'TUCAL', 'centre': 'nan'},
    'RASLEN BEN SALAH': {'ha': 1.5, 'ton': 112.5, 'region': 'CAP BON 1', 'zone': 'dar allouch', 'acces': 'TRC/PPL', 'date_debut': '2026-07-28', 'date_fin': '2026-08-04', 'usine': 'COMOCAP', 'centre': 'nan'},
    'REBAH SMOUD': {'ha': 1.0, 'ton': 90.0, 'region': 'GAFSA / KASSRINE', 'zone': 'AMAYMIA', 'acces': 'SEMI', 'date_debut': '2026-08-15', 'date_fin': '2026-08-19', 'usine': 'ABIDA', 'centre': 'KARIM GARMALAH'},
    'RIADH BEN SAID': {'ha': 2.5, 'ton': 150.0, 'region': 'CAP BON 2', 'zone': 'OUED CHIBA', 'acces': 'PL', 'date_debut': '2026-07-15', 'date_fin': '2026-07-22', 'usine': 'SICAM', 'centre': 'nan'},
    'RIADH BEN ZBIR': {'ha': 4.0, 'ton': 280.0, 'region': 'CAP BON 2', 'zone': 'MENZEL HORR', 'acces': 'PL', 'date_debut': '2026-07-13', 'date_fin': '2026-07-23', 'usine': 'SICAM', 'centre': 'nan'},
    'RIADH KOUKI': {'ha': 16.5, 'ton': 1000.0, 'region': 'NORD', 'zone': 'Jandouba', 'acces': 'PL/SEMI', 'date_debut': '2026-07-22', 'date_fin': '2026-08-09', 'usine': 'SICAM', 'centre': 'Lotfi Trabelsi'},
    'RIDHA AMAYMIA': {'ha': 2.0, 'ton': 180.0, 'region': 'GAFSA / KASSRINE', 'zone': 'AMAYMIA', 'acces': 'SEMI', 'date_debut': '2026-08-10', 'date_fin': '2026-08-15', 'usine': 'SICAM', 'centre': 'KARIM GARMALAH'},
    'ROMDHAN SAAFI': {'ha': 1.0, 'ton': 75.0, 'region': 'CAP BON 1', 'zone': 'dar allouch', 'acces': 'TRC/PPL/PL', 'date_debut': '2026-07-07', 'date_fin': '2026-07-13', 'usine': 'COMOCAP', 'centre': 'nan'},
    'SABER KHARBESH': {'ha': 8.0, 'ton': 560.0, 'region': 'CAP BON 2', 'zone': 'KORBA', 'acces': 'PL', 'date_debut': '2026-07-01', 'date_fin': '2026-07-09', 'usine': 'TUCAL', 'centre': 'SABER KHARBESH'},
    'SALEH BEN HAMOUDA': {'ha': 20.0, 'ton': 1200.0, 'region': 'CAP BON 2', 'zone': 'MENZEL HORR', 'acces': 'PL', 'date_debut': '2026-07-01', 'date_fin': '2026-07-13', 'usine': 'ELFALLEH', 'centre': 'SALEH BEN HAMOUDA'},
    'SALEM EL MEJRI': {'ha': 23.0, 'ton': 1088.0, 'region': 'KAIROUAN', 'zone': 'CHEBIKA-ELHAWEREB', 'acces': 'PL', 'date_debut': '2026-06-28', 'date_fin': '2026-07-06', 'usine': 'COMOCAP', 'centre': 'SALEM EL MEJRI'},
    'SALEM LEGRERI': {'ha': 2.0, 'ton': 150.0, 'region': 'CAP BON 1', 'zone': 'dar allouch', 'acces': 'TRC/PPL', 'date_debut': '2026-07-23', 'date_fin': '2026-08-01', 'usine': 'COMOCAP', 'centre': 'nan'},
    'SAMEH BACCOUCH': {'ha': 2.5, 'ton': 200.0, 'region': 'CAP BON 2', 'zone': 'BELYES', 'acces': 'PL', 'date_debut': '2026-07-05', 'date_fin': '2026-07-11', 'usine': 'SICAM', 'centre': 'nan'},
    'SAMI DAKHLAOUI': {'ha': 3.0, 'ton': 225.0, 'region': 'CAP BON 1', 'zone': 'dar allouch', 'acces': 'PL/PPL', 'date_debut': '2026-07-28', 'date_fin': '2026-08-06', 'usine': 'SICAM', 'centre': 'nan'},
    'SAMI FERGENI': {'ha': 5.5, 'ton': 408.0, 'region': 'BOUFICHA', 'zone': 'SIDI SAIID', 'acces': 'PL', 'date_debut': '2026-07-03', 'date_fin': '2026-07-15', 'usine': 'SICAM', 'centre': 'nan'},
    'SAMI KAAB': {'ha': 3.0, 'ton': 225.0, 'region': 'CAP BON 1', 'zone': 'dar allouch', 'acces': 'PL/PPL', 'date_debut': '2026-07-23', 'date_fin': '2026-08-03', 'usine': 'COMOCAP', 'centre': 'nan'},
    'SAMI LASMAR': {'ha': 2.5, 'ton': 187.5, 'region': 'CAP BON 1', 'zone': 'dar allouch', 'acces': 'TRC/PPL/PL', 'date_debut': '2026-07-09', 'date_fin': '2026-07-19', 'usine': 'COMOCAP', 'centre': 'nan'},
    'SAMIR ATTIYA': {'ha': 27.0, 'ton': 1955.0, 'region': 'KAIROUAN', 'zone': 'ELHAWEREB-AIN BIDHA-HAFOUZ', 'acces': 'PL/SEMI', 'date_debut': '2026-06-28', 'date_fin': '2026-08-02', 'usine': 'ABIDA', 'centre': 'SAMIR ATTIYA'},
    'SASSI MANSOUR': {'ha': 1.3, 'ton': 78.0, 'region': 'CAP BON 2', 'zone': 'OUED CHIBA', 'acces': 'PL', 'date_debut': '2026-07-17', 'date_fin': '2026-07-20', 'usine': 'SICAM', 'centre': 'nan'},
    'SEBTI JABALI': {'ha': 22.0, 'ton': 1725.0, 'region': 'GAFSA / KASSRINE', 'zone': 'FERIANA', 'acces': 'SEMI', 'date_debut': '2026-08-15', 'date_fin': '2026-09-01', 'usine': 'SICAM', 'centre': 'SEBTI JABALI'},
    'SLAH BANI': {'ha': 2.0, 'ton': 120.0, 'region': 'CAP BON 2', 'zone': 'TEFELOUN', 'acces': 'PL', 'date_debut': '2026-06-27', 'date_fin': '2026-07-12', 'usine': 'SICAM', 'centre': 'nan'},
    'SLAH BEN ABDALLAH': {'ha': 0.7, 'ton': 42.0, 'region': 'CAP BON 2', 'zone': 'TEFELOUN', 'acces': 'PL', 'date_debut': '2026-07-16', 'date_fin': '2026-07-19', 'usine': 'COMOCAP', 'centre': 'nan'},
    'SLAH HATBI': {'ha': 1.5, 'ton': 120.0, 'region': 'CAP BON 2', 'zone': 'DIAR HOJJEJ', 'acces': 'PL', 'date_debut': '2026-07-12', 'date_fin': '2026-07-17', 'usine': 'SICAM', 'centre': 'nan'},
    'SLAH SAAD': {'ha': 2.2, 'ton': 198.0, 'region': 'GAFSA / KASSRINE', 'zone': 'SIDI AICH', 'acces': 'SEMI', 'date_debut': '2026-08-05', 'date_fin': '2026-08-10', 'usine': 'SICAM', 'centre': 'HAFEDH MOSBEH'},
    'SLIM MARZOUGUI': {'ha': 7.0, 'ton': 420.0, 'region': 'NORD', 'zone': 'Wed Mliz', 'acces': 'PL/SEMI', 'date_debut': '2026-07-25', 'date_fin': '2026-08-01', 'usine': 'SICAM', 'centre': 'nan'},
    'SOFIENNE GHZELA': {'ha': 2.0, 'ton': 140.0, 'region': 'CAP BON 2', 'zone': 'BENI AYECH', 'acces': 'PL', 'date_debut': '2026-07-17', 'date_fin': '2026-07-23', 'usine': 'COMOCAP', 'centre': 'nan'},
    'SOUHAIL BOUZANA': {'ha': 6.0, 'ton': 480.0, 'region': 'GAFSA / KASSRINE', 'zone': 'AMAYMIA', 'acces': 'SEMI', 'date_debut': '2026-07-06', 'date_fin': '2026-07-15', 'usine': 'SICAM', 'centre': 'SOUHAIEL BOUZANA'},
    'STE 428 SERVICES AGRICOLES': {'ha': 74.0, 'ton': 5000.0, 'region': 'NORD', 'zone': 'medjez beb', 'acces': 'PL/SEMI', 'date_debut': '2026-07-19', 'date_fin': '2026-07-31', 'usine': 'COMOCAP', 'centre': 'STE 428 SERVICES AGRICOLES'},
    'STE AGROBEST': {'ha': 17.0, 'ton': 1020.0, 'region': 'CAP BON 2', 'zone': 'KORBA', 'acces': 'PL', 'date_debut': '2026-07-01', 'date_fin': '2026-07-15', 'usine': 'TUCAL', 'centre': 'STE AGROBEST'},
    'STE BACCARA': {'ha': 120.0, 'ton': 7400.0, 'region': 'CAP BON 2', 'zone': 'menzel tamim', 'acces': 'PL/PPL', 'date_debut': '2026-06-28', 'date_fin': '2026-07-31', 'usine': 'SICAM', 'centre': 'STE BACCARA'},
    'STE KERKOUANE S A': {'ha': 50.0, 'ton': 3000.0, 'region': 'CAP BON 2', 'zone': 'menzel tamim', 'acces': 'PL/PPL', 'date_debut': '2026-07-09', 'date_fin': '2026-08-13', 'usine': 'COMOCAP', 'centre': 'STE KERKOUANE'},
    'STE SEMAG': {'ha': 8.0, 'ton': 612.0, 'region': 'KAIROUAN', 'zone': 'KHADHRA', 'acces': 'PL', 'date_debut': '2026-06-28', 'date_fin': '2026-07-10', 'usine': 'SICAM', 'centre': 'STE SMAG'},
    'TAHER MANSOURI': {'ha': 1.0, 'ton': 90.0, 'region': 'GAFSA / KASSRINE', 'zone': 'MAJEL BELABESS', 'acces': 'SEMI', 'date_debut': '2026-08-08', 'date_fin': '2026-08-11', 'usine': 'SICAM', 'centre': 'MOURAD MANSOURI'},
    'TAHER MATHLOUTHI': {'ha': 3.0, 'ton': 210.0, 'region': 'CAP BON 2', 'zone': 'TEFELOUN', 'acces': 'PL', 'date_debut': '2026-07-03', 'date_fin': '2026-07-13', 'usine': 'TUCAL', 'centre': 'nan'},
    'TALEB JABLAH': {'ha': 1.0, 'ton': 90.0, 'region': 'GAFSA / KASSRINE', 'zone': 'AMAYMIA', 'acces': 'SEMI', 'date_debut': '2026-07-21', 'date_fin': '2026-07-24', 'usine': 'TUCAL', 'centre': 'KARIM GARMALAH'},
    'TAREK BEN ABDALAH': {'ha': 3.0, 'ton': 225.0, 'region': 'CAP BON 1', 'zone': 'dar allouch', 'acces': 'TRC/PPL/PL', 'date_debut': '2026-07-10', 'date_fin': '2026-07-20', 'usine': 'COMOCAP', 'centre': 'nan'},
    'TAREK BEN NJI': {'ha': 1.0, 'ton': 60.0, 'region': 'CAP BON 2', 'zone': 'TEFELOUN', 'acces': 'PL', 'date_debut': '2026-07-10', 'date_fin': '2026-07-13', 'usine': 'COMOCAP', 'centre': 'nan'},
    'TAREK EL BAHRI': {'ha': 1.0, 'ton': 75.0, 'region': 'CAP BON 1', 'zone': 'dar allouch', 'acces': 'TRC/PPL', 'date_debut': '2026-07-13', 'date_fin': '2026-07-17', 'usine': 'COMOCAP', 'centre': 'nan'},
    'WISSEM AMAYMIA': {'ha': 4.0, 'ton': 360.0, 'region': 'GAFSA / KASSRINE', 'zone': 'AMAYMIA', 'acces': 'SEMI', 'date_debut': '2026-07-15', 'date_fin': '2026-07-20', 'usine': 'SICAM', 'centre': 'KARIM GARMALAH'},
    'YASIN MNASRI': {'ha': 17.5, 'ton': 1575.0, 'region': 'SIDI BOUZID', 'zone': 'BIR LAHFAY', 'acces': 'SEMI', 'date_debut': '2026-06-29', 'date_fin': '2026-07-19', 'usine': 'COMOCAP', 'centre': 'HAFEDH MOSBEH'},
    'YASIN TLILI': {'ha': 1.0, 'ton': 90.0, 'region': 'GAFSA / KASSRINE', 'zone': 'AMAYMIA', 'acces': 'SEMI', 'date_debut': '2026-07-25', 'date_fin': '2026-07-29', 'usine': 'TUCAL', 'centre': 'KARIM GARMALAH'},
    'ZOUHAIR BAICH': {'ha': 12.5, 'ton': 1000.0, 'region': 'CAP BON 2', 'zone': 'LEBNA/TAMEZRRAT', 'acces': 'PL', 'date_debut': '2026-07-01', 'date_fin': '2026-07-14', 'usine': 'SICAM', 'centre': 'nan'},
    'ZOUHAIR BEN ECHIK': {'ha': 2.5, 'ton': 187.5, 'region': 'CAP BON 1', 'zone': 'dar allouch', 'acces': 'TRC/PPL/PL', 'date_debut': '2026-07-09', 'date_fin': '2026-07-21', 'usine': 'COMOCAP', 'centre': 'nan'},
}
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
import io
from datetime import date

# ══════════════════════════════════════════════════════════════
# CONSTANTES
# ══════════════════════════════════════════════════════════════
DATE_CAISSE_LIMITE = date(2026, 7, 10)

# Paramètres caisses vides PAR USINE (modifiables dans l'UI)
# Format : {usine: {"nb_ha": nb_caisses/ha, "prix": DT/caisse, "type": description}}
CAISSES_USINE_DEFAULTS = {
    "SICAM":    {"nb_ha": 80,  "prix": 3.0,  "type": "Caisse plastique 25kg",  "cap_kg": 25},
    "TUCAL":    {"nb_ha": 80,  "prix": 3.0,  "type": "Caisse plastique 25kg",  "cap_kg": 25},
    "COMOCAP":  {"nb_ha": 60,  "prix": 2.5,  "type": "Bac tracteur (forfait)", "cap_kg": 30},
    "ABIDA":    {"nb_ha": 60,  "prix": 2.5,  "type": "Caisse plastique 25kg",  "cap_kg": 25},
    "ELFALLEH": {"nb_ha": 50,  "prix": 2.0,  "type": "Caisse métal 20kg",      "cap_kg": 20},
}
MO_TONNE_DEFAULT   = 50.0
DENSITE_STD = {
    "CAP BON 1":25000,"CAP BON 2":25000,"NORD":22000,
    "KAIROUAN":20000,"BOUFICHA":22000,
    "GAFSA / KASSRINE":18000,"SIDI BOUZID":18000,
}
FAM_NORM_MAP = {
    "engrais":"Engrais","engrais ":"Engrais",
    "fertilissant":"Fertilisant","fertilisant":"Fertilisant",
    "fongicide":"Fongicide","insecticide":"Insecticide",
    "irrigations ":"Irrigation","irrigations":"Irrigation",
    "irrigations turk":"Irrigation","irrigation":"Irrigation","irrigation ":"Irrigation",
    "herbicide":"Herbicide","divers":"Divers","divers ":"Divers",
    "materiel":"Matériel","traitement":"Traitement",
}

# ══════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════
def _norm(c):
    return (str(c).strip().lower()
            .replace("é","e").replace("è","e").replace("ê","e")
            .replace("â","a").replace("ô","o").replace("î","i")
            .replace("û","u").replace(" ","_").replace("/","_")
            .replace("(","").replace(")","").replace("°","")
            .replace("'",""))

def _find_header(raw, keywords, max_rows=8):
    """
    Trouve la ligne header en cherchant celle qui contient
    ≥2 mots-clés dans des CELLULES SÉPARÉES (évite les faux positifs sur
    les lignes titre qui contiennent tout en une seule cellule).
    """
    for i in range(min(max_rows, len(raw))):
        # Filtrer : seulement les cellules courtes (< 60 chars) = vraies colonnes
        cells = [_norm(str(v)) for v in raw.iloc[i].values
                 if pd.notna(v) and len(str(v).strip()) < 60]
        # Compter les mots-clés dans des CELLULES DIFFÉRENTES
        kw_cells = {kw: any(kw in c for c in cells) for kw in keywords}
        n_match = sum(kw_cells.values())
        if n_match >= 2:
            return i
    # Fallback : 1 seul match dans cellules courtes
    for i in range(min(max_rows, len(raw))):
        cells = [_norm(str(v)) for v in raw.iloc[i].values
                 if pd.notna(v) and len(str(v).strip()) < 60]
        if any(any(kw in c for c in cells) for kw in keywords):
            return i
    return 0

def _read_auto(file_obj, keywords):
    raw = pd.read_excel(file_obj, sheet_name=0, header=None)
    hr  = _find_header(raw, keywords)
    file_obj.seek(0)
    df  = pd.read_excel(file_obj, sheet_name=0, header=hr)
    df.columns = [_norm(c) for c in df.columns]
    return df

def _find_col(df, cands):
    for c in cands:
        if c in df.columns: return c
    return None

def _check_required(df, source_name, required=["centre","client"]):
    """Vérifie que les colonnes obligatoires sont présentes."""
    missing = [c for c in required if c not in df.columns]
    if missing:
        return False, f"⚠️ {source_name} — colonnes manquantes : {missing}"
    return True, ""

def _metric(label, value, color="#f0f6fc", delta=None, delta_label=""):
    dh = ""
    if delta is not None:
        dc = "#3dd68c" if delta >= 0 else "#ef5350"
        sign = "+" if delta >= 0 else ""
        dh = f"<div style='font-size:.75rem;color:{dc}'>{sign}{delta:,.1f} {delta_label}</div>"
    return f"""<div style='background:#161b22;border:1px solid #30363d;border-radius:10px;
padding:12px 16px;border-top:3px solid {color}'>
<div style='font-size:.7rem;color:#8b949e;text-transform:uppercase;letter-spacing:.05em'>{label}</div>
<div style='font-size:1.4rem;font-weight:700;color:#f0f6fc'>{value}</div>{dh}</div>"""

# ══════════════════════════════════════════════════════════════
# SUPABASE — Date début récolte par agriculteur
# ══════════════════════════════════════════════════════════════
@st.cache_data(ttl=300, show_spinner=False)
def load_date_debut_recolte(_sb):
    """
    Charge depuis Supabase (plan_rectifie_detail) la date MIN de livraison
    par agriculteur = date début récolte.
    C'est cette date qui détermine l'affectation des caisses vides :
      < 10 juillet  → 1ère affectation (avec caisses vides)
      ≥ 10 juillet  → 2ème affectation (sans caisses vides)
    """
    if sb is None:
        return pd.DataFrame()
    try:
        data = []
        offset = 0
        while True:
            batch = sb.table("plan_rectifie_detail").select(
                "agriculteur,date"
            ).range(offset, offset+999).execute().data
            if not batch: break
            data.extend(batch)
            if len(batch) < 1000: break
            offset += 1000
        if not data:
            return pd.DataFrame()
        df = pd.DataFrame(data)
        df["date"] = pd.to_datetime(df["date"], errors="coerce")
        df = df.dropna(subset=["date"])
        # Min date = date début récolte par agriculteur
        result = df.groupby("agriculteur")["date"].min().reset_index()
        result.columns = ["agriculteur","date_debut_recolte"]
        result["agriculteur"] = result["agriculteur"].astype(str).str.strip().str.upper()
        # Déduire l'affectation
        result["affectation_caisse"] = result["date_debut_recolte"].apply(
            lambda d: "1ère (avec caisses)" if pd.notna(d) and d.date() < DATE_CAISSE_LIMITE
                      else "2ème (sans caisses)")
        return result
    except Exception as e:
        return pd.DataFrame()


def load_prevision_juin(sb):
    """Tonnage prévu Juin depuis plan_rectifie_detail (somme par agriculteur)."""
    if sb is None:
        return pd.DataFrame()
    try:
        data = []
        offset = 0
        while True:
            batch = sb.table("plan_rectifie_detail").select(
                "agriculteur,tonnes"
            ).range(offset, offset+999).execute().data
            if not batch: break
            data.extend(batch)
            if len(batch) < 1000: break
            offset += 1000
        if not data:
            return pd.DataFrame()
        df = pd.DataFrame(data)
        df["tonnes"] = pd.to_numeric(df["tonnes"], errors="coerce").fillna(0)
        grp = df.groupby("agriculteur")["tonnes"].sum().reset_index()
        grp.columns = ["agriculteur","prevision_juin"]
        grp["agriculteur"] = grp["agriculteur"].astype(str).str.strip().str.upper()
        return grp
    except Exception:
        return pd.DataFrame()


# ══════════════════════════════════════════════════════════════
# PARSERS — "centre" et "client" obligatoires dans tous les fichiers
# ══════════════════════════════════════════════════════════════

def parse_bourak(file_obj):
    """
    BOURAK — Financement & Transport
    Colonnes OBLIGATOIRES : client (agriculteur) · centre
    Colonnes attendues    : commercial · ingenieur · region ·
                            hectares · avance · report
    """
    df = _read_auto(file_obj,
         ["client","responsable","ingenieur","centre","avance","report"])

    MAP = {
        "client":        ["client","agriculteur","nom"],
        "commercial":    ["responsable","commercial","resp"],
        "ingenieur":     ["ingenieur","ing","ingenieur_agronome"],
        "centre":        ["centre","centre_collecte"],
        "region":        ["region","zone"],
        "hectares":      ["hectares","ha","surface","nb_hectares","ha_reels"],
        "avance":        ["avance","avances","total_avance","montant_avance",
                          "avance_dt","avance_dinar"],
        "report":        ["report","reste","solde_precedent","non_paye","report_dt"],
        "plt_livres":    ["plt_livres","plateaux_livres","nb_plateaux_livres",
                          "plt_livre","nb_plt_livres"],
        "plt_retour":    ["plt_retour","plateaux_retour","nb_plateaux_retour",
                          "plt_ret","retour_plateaux"],
        "plt_perdus":    ["plt_perdus","plateaux_perdus","nb_plateaux_perdus",
                          "plt_perd"],
    }
    rename = {}
    for tgt, cands in MAP.items():
        for c in cands:
            if c in df.columns and tgt not in rename.values():
                rename[c] = tgt; break
    df = df.rename(columns=rename)

    # Vérification colonnes obligatoires
    if "client" not in df.columns:
        return None, "BOURAK — colonne client manquante"
    if "centre" not in df.columns:
        df["centre"] = ""

    for c in ["hectares","avance","report","plt_livres","plt_retour","plt_perdus"]:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)
    # Calculer plt_perdus si colonnes présentes
    if "plt_livres" in df.columns and "plt_retour" in df.columns:
        if "plt_perdus" not in df.columns or df["plt_perdus"].sum() == 0:
            df["plt_perdus"] = df["plt_livres"] - df["plt_retour"]
    df["client"] = df["client"].astype(str).str.strip()
    # Filtrer : vides, TOTAL, et séparateurs "── COMM ──" générés par les fichiers test
    df = df[~df["client"].str.upper().isin(["","NAN","TOTAL","SOUS-TOTAL"])]
    df = df[~df["client"].str.startswith("──")]
    df = df[~df["client"].str.startswith("--")]
    df = df[df["client"].str.len() > 2]
    # S'assurer que commercial est bien présent (depuis "responsable" ou "ingenieur")
    if "commercial" not in df.columns:
        for _try_col in ["responsable","ingenieur","Responsable","Commercial"]:
            if _try_col in df.columns:
                df["commercial"] = df[_try_col].astype(str).str.strip()
                break
        else:
            df["commercial"] = ""
    return df, ""


def parse_royal(file_obj):
    """
    ROYAL — Plants livrés (Pépinière)
    Colonnes OBLIGATOIRES : client · centre
    Colonnes attendues    : zone · variete · qte_livree · valeur ·
                            date_debut_livraison · date_fin_livraison ·
                            type_plateau · nb_plateaux

    NOTE : date_debut_livraison = date de la 1ère livraison de plants
           (≠ date début récolte — celle-ci vient du fichier rectifié)
    """
    df = _read_auto(file_obj,
         ["client","centre","variete","quantite","date","livraison"])

    MAP = {
        "client":               ["client","agriculteur","nom"],
        "centre":               ["centre","centre_collecte"],
        "zone":                 ["zone","destination","direction","localisation"],
        "variete":              ["variete","article","type_plant"],
        "qte_livree":           ["quantite_livree","qte_livree","qte","plants_livres","nb_plants"],
        "valeur_plants":        ["valeur","montant","total","prix_total","valeur_plants"],
        "date_debut_livraison": ["date_debut_livraison","date_premiere_livraison",
                                 "date_debut","debut_livraison","debut"],
        "date_fin_livraison":   ["date_fin_livraison","date_fin","fin_livraison","fin"],
        "type_plateau":         ["type_plateau","unite","plateau"],
        "nb_plateaux":          ["nb_plateaux","plateaux","qte_plateaux"],
    }
    rename = {}
    for tgt, cands in MAP.items():
        for c in cands:
            if c in df.columns and tgt not in rename.values():
                rename[c] = tgt; break
    df = df.rename(columns=rename)

    if "client" not in df.columns:
        return None, "ROYAL — colonne client manquante"
    if "centre" not in df.columns:
        df["centre"] = ""

    for c in ["qte_livree","valeur_plants","nb_plateaux"]:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)
    for c in ["date_debut_livraison","date_fin_livraison"]:
        if c in df.columns:
            df[c] = pd.to_datetime(df[c], errors="coerce")

    df["client"] = df["client"].astype(str).str.strip()
    df = df[~df["client"].str.upper().isin(["","NAN","TOTAL","CLIENT","AGRICULTEUR"])]
    df = df[~df["client"].str.startswith("──")]
    df = df[~df["client"].str.startswith("--")]
    df = df[df["client"].str.len() > 2]
    return df, ""


def parse_sotusfa(file_obj):
    """
    SOTUSFA — Engrais & Pesticides
    Colonnes OBLIGATOIRES : client · centre
    Colonnes attendues    : famille · article · qte · valeur
    """
    df = _read_auto(file_obj,
         ["client","agriculteur","centre","famille","article","valeur"])

    MAP = {
        "client":    ["client","agriculteur","nom"],
        "centre":    ["centre","centre_collecte"],
        "famille":   ["famille"],
        "article":   ["article","produit"],
        "qte":       ["qte","quantite"],
        "valeur":    ["total_ttc","total","valeur","montant"],
        "prix_u":    ["prix_un_ttc","prix_ttc","prix"],
    }
    rename = {}
    for tgt, cands in MAP.items():
        for c in cands:
            if c in df.columns and tgt not in rename.values():
                rename[c] = tgt; break
    df = df.rename(columns=rename)

    if "client" not in df.columns:
        return None, None, "SOTUSFA — colonne client/agriculteur manquante"
    if "centre" not in df.columns:
        df["centre"] = ""

    for c in ["qte","valeur","prix_u"]:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)
    df["client"] = df["client"].astype(str).str.strip()
    df = df[~df["client"].str.upper().isin(["","NAN","TOTAL"])]

    # Normaliser famille
    if "famille" in df.columns:
        df["famille_norm"] = df["famille"].astype(str).str.strip().str.lower()\
                             .map(FAM_NORM_MAP).fillna("Autre")

    # Pivot par (client + centre) + famille → une ligne par client
    if "famille_norm" in df.columns and "valeur" in df.columns:
        grp_cols = ["client"]
        if "centre" in df.columns:
            grp_cols.append("centre")
        pivot = df.groupby(grp_cols + ["famille_norm"])["valeur"].sum().unstack(
            fill_value=0).reset_index()
        # Normaliser noms colonnes (familles)
        new_cols = []
        for i, c in enumerate(pivot.columns):
            if i < len(grp_cols):
                new_cols.append(c)
            else:
                new_cols.append(_norm(str(c)))
        pivot.columns = new_cols
        pivot["total_intrants"] = pivot.select_dtypes("number").sum(axis=1)
    else:
        pivot = pd.DataFrame()

    return df, pivot, ""


def parse_quantite(file_obj):
    """
    TABLEAU QUANTITÉ — Plan livré / actif / extra
    Colonnes OBLIGATOIRES : client · centre
    Colonnes attendues    : qte_livree · qte_actif · qte_extra ·
                            tonnage_livre · prix_vente
    """
    df = _read_auto(file_obj,
         ["client","centre","livree","actif","extra","quantite","tonnage"])

    MAP = {
        "client":       ["client","agriculteur","nom"],
        "centre":       ["centre","centre_collecte"],
        "qte_livree":   ["quantite_livree","qte_livree","livree","plants_livres"],
        "qte_actif":    ["quantite_actif","qte_actif","actif","plants_actifs"],
        "qte_extra":    ["quantite_extra","qte_extra","extra","pertes"],
        "tonnage_livre":["tonnage_livre","tonnage","recolte","livraison_t",
                          "tonnage_plan","tonnage_plan_","tonnage_planif",
                          "tonnage_prevu","tonnage_planifie","tonnage_livre_t",
                          "volume","volume_t"],
        "prix_vente":   ["prix_vente","prix","prix_unitaire_vente",
                          "prix_vente_dt","prix_t","prix_tonne"],
        "commercial":   ["commercial","responsable","comm","ing"],
        # NOTE: hectares vient du Bourak — PAS de Quantite (éviter conflit de merge)
    }
    rename = {}
    for tgt, cands in MAP.items():
        for c in cands:
            if c in df.columns and tgt not in rename.values():
                rename[c] = tgt; break
    df = df.rename(columns=rename)

    if "client" not in df.columns:
        return None, "TABLEAU QUANTITÉ — colonne client manquante"
    if "centre" not in df.columns:
        df["centre"] = ""

    for c in ["qte_livree","qte_actif","qte_extra","tonnage_livre","prix_vente"]:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)
    if "qte_extra" not in df.columns and "qte_livree" in df.columns and "qte_actif" in df.columns:
        df["qte_extra"] = df["qte_livree"] - df["qte_actif"]
    df["client"] = df["client"].astype(str).str.strip()
    df = df[~df["client"].str.upper().isin(["","NAN","TOTAL"])]
    return df, ""


def parse_prevision(file_obj, col_name):
    """
    Prévision Déc ou Mai.
    Accepte le format réel (responsable région / AGRICULTEUR / TONNAGE)
    centre est OPTIONNEL dans ce fichier.
    """
    try:
        df = _read_auto(file_obj,
             ["agriculteur","client","tonnage","prevision","responsable"])

        MAP = {
            "client":     ["client","agriculteur","nom"],
            "commercial": ["responsable_region","responsable region",
                           "commercial","responsable","resp"],
            "centre":     ["centre","centre_collecte"],
            "region":     ["region"],
            col_name:     ["tonnage","prevision","tonnes","quantite",
                           "tonnage_total","total_tonnage"],
        }
        rename = {}
        for tgt, cands in MAP.items():
            for c in cands:
                if c in df.columns and tgt not in rename.values():
                    rename[c] = tgt; break
        df = df.rename(columns=rename)

        # client obligatoire
        if "client" not in df.columns:
            return None, f"PRÉVISION ({col_name}) — colonne client/AGRICULTEUR manquante"

        df[col_name] = pd.to_numeric(df.get(col_name, 0), errors="coerce").fillna(0)
        df["client"] = df["client"].astype(str).str.strip()

        # Filtrer lignes TOTAL, vides, sous-totaux
        df = df[~df["client"].str.upper().str.strip().isin(
            ["","NAN","TOTAL","TOTAL FEDI","TOTAL MEKKI","TOTAL KHALIL",
             "TOTAL MAKKI","TOTAL ACHREF","TOTAL JILANI","TOTAL MAKKI BEN SALAH",
             "TOTAL ACHREF AJLANI","TOTAL JILANI OBAY","SOUS-TOTAL"])]
        df = df[df[col_name] > 0]

        # centre optionnel — créer vide si absent
        if "centre" not in df.columns:
            df["centre"] = ""

        keep = ["client","centre",col_name]
        for extra in ["commercial","region"]:
            if extra in df.columns:
                keep.append(extra)
        df = df[keep]
        # ← CLEF : sommer par client pour éviter double comptage
        # (un agriculteur qui livre à 2 usines = 2 lignes dans le fichier → 1 après sum)
        num_cols = [col_name]
        df = df.groupby("client", as_index=False)[num_cols].sum()
        df["centre"] = ""  # centre non disponible après groupby
        return df, ""
    except Exception as e:
        return None, str(e)


# ══════════════════════════════════════════════════════════════
# FUSION ET CALCULS
# ══════════════════════════════════════════════════════════════

def merge_and_calculate(df_bourak, df_royal, df_sotusfa_raw,
                        df_sotusfa_pivot, df_quantite,
                        df_prev_dec, df_prev_mai, df_prev_juin,
                        df_dates_recolte, params):
    """
    Fusionne les 4 sources sur (client + centre) et calcule tous les indicateurs.

    Clé de jointure : client + centre (présents dans tous les fichiers)

    df_dates_recolte : DataFrame(agriculteur, date_debut_recolte, affectation_caisse)
                       issu de Supabase plan_rectifie_detail
                       ⇒ détermine 1ère/2ème affectation caisses vides
    """
    # ── Base = BOURAK ──────────────────────────────────────
    if df_bourak is not None and not df_bourak.empty:
        base = df_bourak.copy()
    elif df_quantite is not None and not df_quantite.empty:
        base = df_quantite[["client","centre"]].copy()
    else:
        return pd.DataFrame()

    def _upper(df, cols):
        for c in cols:
            if c in df.columns:
                df[c] = df[c].fillna("").astype(str).str.strip().str.upper()
                df[c] = df[c].replace({"NAN": "", "NONE": "", "NAT": ""})
        return df

    base = _upper(base, ["client","centre"])
    KEY = ["client","centre"]
    # Garder commercial depuis Bourak (vient du champ "responsable")
    if "commercial" not in base.columns:
        base["commercial"] = ""
    base["commercial"] = base["commercial"].fillna("").astype(str).str.strip()
    # Après chaque merge outer, préserver commercial depuis le côté gauche
    _comm_series = base.set_index("client")["commercial"].to_dict() if "client" in base.columns else {}
    # Plateaux depuis Bourak
    for _pc in ["plt_livres","plt_retour","plt_perdus"]:
        if _pc not in base.columns: base[_pc] = 0
        base[_pc] = pd.to_numeric(base[_pc], errors="coerce").fillna(0)

    # ── Merge ROYAL ────────────────────────────────────────
    if df_royal is not None and not df_royal.empty:
        r = _upper(df_royal.copy(), KEY)
        # Groupby Royal défensif — vérifier colonnes avant agg
        _ragg = {}
        if "qte_livree"          in r.columns: _ragg["qte_livree"]          = "sum"
        if "valeur_plants"       in r.columns: _ragg["valeur_plants"]       = "sum"
        if "date_debut_livraison"in r.columns: _ragg["date_debut_livraison"]= "min"
        if "date_fin_livraison"  in r.columns: _ragg["date_fin_livraison"]  = "max"
        r_grp_base = r.groupby(KEY).agg(_ragg).reset_index()
        # Colonnes mode (variete, zone) séparément
        for _mc in ["variete","zone"]:
            if _mc in r.columns:
                _mv = r.groupby(KEY)[_mc].agg(lambda x: x.mode()[0] if len(x)>0 else "").reset_index()
                r_grp_base = r_grp_base.merge(_mv, on=KEY, how="left")
        # Renommer pour compatibilité
        _rename_r = {"qte_livree":"qte_royal","date_debut_livraison":"date_debut_liv",
                     "date_fin_livraison":"date_fin_liv"}
        r_grp = r_grp_base.rename(columns={k:v for k,v in _rename_r.items() if k in r_grp_base.columns})
        # Consigne plateau
        prix_c = params.get("prix_consigne", {})
        if "type_plateau" in r.columns and "nb_plateaux" in r.columns:
            r["consigne_pl"] = r.apply(
                lambda row: row["nb_plateaux"] * prix_c.get(
                    str(row.get("type_plateau","")), 0), axis=1)
            r_cons = r.groupby(KEY)["consigne_pl"].sum().reset_index()
            r_grp  = r_grp.merge(r_cons, on=KEY, how="left")
            r_grp  = r_grp.rename(columns={"consigne_pl":"consigne_plateau"})
        else:
            r_grp["consigne_plateau"] = 0
        base = base.merge(r_grp, on=KEY, how="outer", suffixes=("","_r"))
        # Résoudre conflits _r
        for _col_r in [c for c in base.columns if c.endswith("_r")]:
            _col_orig = _col_r[:-2]
            if _col_orig in base.columns:
                base[_col_orig] = base[_col_orig].fillna(base[_col_r])
                base = base.drop(columns=[_col_r])
            else:
                base = base.rename(columns={_col_r: _col_orig})
        # Restaurer commercial perdu par le outer merge
        if "commercial" in base.columns:
            base["commercial"] = base["commercial"].fillna(
                base["client"].map(_comm_series)).fillna("")
        elif _comm_series:
            base["commercial"] = base["client"].map(_comm_series).fillna("")

    # ── Merge dates récolte (Supabase) ─────────────────────
    # ⚠️  La condition caisses vides = date_debut_RECOLTE (pas livraison)
    if df_dates_recolte is not None and not df_dates_recolte.empty:
        dr = df_dates_recolte.copy()
        dr["client"] = dr["agriculteur"].astype(str).str.strip().str.upper()
        base = base.merge(
            dr[["client","date_debut_recolte","affectation_caisse"]],
            on="client", how="left")
        base["affectation_caisse"] = base["affectation_caisse"].fillna(
            "2ème (sans caisses)")
    else:
        base["affectation_caisse"] = "2ème (sans caisses)"
        base["date_debut_recolte"] = pd.NaT

    # ── Merge SOTUSFA — données réelles 2026 en priorité ──────────
    def _sot_merge(df_raw, df_pivot, base_df):
        """
        Priorité 1 : Constante _INTRANTS_2026 (données réelles + estimées confirmées)
        Priorité 2 : Fichier Sotusfa uploadé (mapping direct + fuzzy)
        Priorité 3 : Distribution ACHREF groupes
        """
        """
        1. Map exact client → total_intrants depuis Sotusfa
        2. Fuzzy matching pour les noms légèrement différents
        3. Distribution proportionnelle pour ACHREF sous-membres
        """
        import re as _re, unicodedata as _uc

        def _cn(n):
            n = str(n).strip().upper()
            n = _re.sub(r"[(][^)]*[)]","",n)
            n = "".join(c for c in _uc.normalize("NFD",n) if _uc.category(c) != "Mn")
            n = _re.sub(r"[^A-Z0-9 ]"," ",n)
            return _re.sub(r"[ ]+"," ",n).strip()

        def _sco(a, b):
            wa, wb = set(a.split()), set(b.split())
            if not wa or not wb: return 0.0
            inter = len(wa & wb); union = len(wa | wb)
            sj = inter / union if union else 0
            sh, lo = (wa, wb) if len(wa) <= len(wb) else (wb, wa)
            sc = sum(1 for w in sh if any(lw.startswith(w[:4]) for lw in lo) and len(w) > 2) / max(len(sh), 1) * 0.85
            return max(sj, sc)

        # Groupes ACHREF — noms Sotusfa → membres référence (avec tonnages réels)
        ACHREF_GROUPES = {
            _cn("HAFEDH MESBEH"): [
                "HAFEDH MOSBEH","HAFEDH MOSBEH (ABDELSATER)"],
            _cn("ABDELKARIM GARMALLAH"): [
                "KARIM GARMALAH 1","KARIM GARMALAH 2","MED ALI GARMALAH",
                "AMAR GARMALAH (SEMI-SICAM)","AMAR GARMALAH (SEMI-TUCAL)",
                "HSAN GARMALAH","JAMEL GARMALAH","MOHAMED GARMALAH"],
            _cn("MOURAD MANSOURI"): [
                "MOURAD MANSOURI (SEMI-SICAM)","MOURAD MANSOURI (SEMI-TUCAL)",
                "NOUREDIN MANSOURI","ELIFA MANSOURI","TAHER MANSOURI",
                "ABELSAMII MANSOURI","CHOKRI MANSOURI","LAMINE MANSOURI","AHMED MANSOURI"],
            _cn("Sebti jaballi"): [
                "SEBTI JABALI (SEMI-SICAM)","SEBTI JABALI (SEMI-ABIDA)",
                "KHAMES JABALI","ARBI JABALI","TALEB JABLAH"],
            _cn("SOCIETE BILEL GHA SERVICE AGRICOLE"): [
                "BILEL GHA 1","BILEL GHA 2","BILEL GHA 3","BILEL GHA 4","BILEL KEHIL"],
            _cn("SOUHAIL BOUZENA"): ["SOUHAIL BOUZANA"],
        }
        # Distribution pré-calculée depuis tonnages réels référence ACHREF
        _ACHREF_DIST_PRECOMPUTED = {
            "HAFEDH MOSBEH":                153812,
            "HAFEDH MOSBEH (ABDELSATER)":   235845,
            "AMAR GARMALAH (SEMI-SICAM)":    32638,
            "KARIM GARMALAH 1":              65277,
            "MED ALI GARMALAH":              65277,
            "MOHAMED GARMALAH":              16319,
            "AMAR GARMALAH (SEMI-TUCAL)":    32638,
            "HSAN GARMALAH":                 32638,
            "JAMEL GARMALAH":                16319,
            "KARIM GARMALAH 2":              65277,
            "AHMED MANSOURI":                 9054,
            "ELIFA MANSOURI":                 9054,
            "MOURAD MANSOURI (SEMI-SICAM)":  15091,
            "NOUREDIN MANSOURI":             18109,
            "TAHER MANSOURI":                 9054,
            "ABELSAMII MANSOURI":             9054,
            "CHOKRI MANSOURI":                9054,
            "MOURAD MANSOURI (SEMI-TUCAL)":  30182,
            "LAMINE MANSOURI":               12073,
            "SEBTI JABALI (SEMI-SICAM)":     60524,
            "TALEB JABLAH":                   3242,
            "ARBI JABALI":                    5404,
            "KHAMES JABALI":                  8646,
            "SEBTI JABALI (SEMI-ABIDA)":      8646,
            "BILEL GHA 1":                   14322,
            "BILEL GHA 2":                   17187,
            "BILEL GHA 3":                   17187,
            "BILEL KEHIL":                    8593,
            "BILEL GHA 4":                   15277,
            "SOUHAIL BOUZANA":               27241,
        }
        _ACHREF_DIST_CN = {_cn(k):v for k,v in _ACHREF_DIST_PRECOMPUTED.items()}

        src = df_raw if (df_raw is not None and not df_raw.empty) else None
        if src is None and df_pivot is not None and not df_pivot.empty:
            src = df_pivot

        if src is None:
            # Pas de Sotusfa uploadé → utiliser uniquement _INTRANTS_2026
            base_df["total_intrants"] = base_df["client"].apply(
                lambda x: _INTRANTS_2026.get(_cn(str(x).strip()), 0.0))
            # Fuzzy fallback pour les non-trouvés
            _int2_keys = list(_INTRANTS_2026.keys())
            def _fuzz_int(x):
                ck = _cn(str(x).strip())
                if ck in _INTRANTS_2026: return _INTRANTS_2026[ck]
                best = max(_int2_keys, key=lambda k: _sco(ck,k), default=None)
                return _INTRANTS_2026[best] if best and _sco(ck,best)>=0.65 else 0.0
            base_df["total_intrants"] = base_df["client"].apply(_fuzz_int)
            return base_df

        _cli = next((c for c in src.columns if c in ["client","agriculteur","nom"]), None)
        _val = next((c for c in src.columns if c in ["valeur","total_ttc","total","montant","total_intrants"]), None)
        if not _cli or not _val:
            return base_df

        src = src.copy()
        src[_cli] = src[_cli].astype(str).str.strip().str.upper()
        src[_val] = pd.to_numeric(src[_val], errors="coerce").fillna(0)
        sot_agg = src.groupby(_cli)[_val].sum().to_dict()

        # Dict nettoyé
        sot_clean = {_cn(k): v for k, v in sot_agg.items()}
        sot_keys  = list(sot_clean.keys())

        # Tonnages depuis base pour distribution proportionnelle
        base_tons = {}
        if "tonnage_livre" in base_df.columns:
            for _,r in base_df.iterrows():
                k = _cn(str(r.get("client","")))
                base_tons[k] = float(pd.to_numeric(r.get("tonnage_livre",0), errors="coerce") or 0)

        result = {}
        for _, row in base_df.iterrows():
            client_raw = str(row.get("client","")).strip()
            ck = _cn(client_raw)

            # 0. PRIORITÉ ABSOLUE : données réelles/estimées confirmées 2026
            if ck in _INTRANTS_2026:
                result[client_raw] = _INTRANTS_2026[ck]
                continue

            # 1. Fuzzy sur _INTRANTS_2026 (seuil élevé = confiance)
            best_pre = max(_INTRANTS_2026.keys(), key=lambda k: _sco(ck,k), default=None)
            if best_pre and _sco(ck, best_pre) >= 0.70:
                result[client_raw] = _INTRANTS_2026[best_pre]
                continue

            # 2. Exact match Sotusfa uploadé
            if ck in sot_clean:
                result[client_raw] = sot_clean[ck]
                continue

            # 2. ACHREF : appartient à un groupe → distribution proportionnelle
            assigned = False
            for grp_k, membres in ACHREF_GROUPES.items():
                membres_cn = [_cn(m) for m in membres]
                if ck in membres_cn:
                    # 1. Distribution pré-calculée (basée sur tonnages référence ACHREF)
                    if ck in _ACHREF_DIST_CN:
                        result[client_raw] = _ACHREF_DIST_CN[ck]
                        assigned = True
                        break
                    # 2. Distribution dynamique si groupe dans Sotusfa
                    elif grp_k in sot_clean:
                        grp_total_intrants = sot_clean[grp_k]
                        membres_tons = {_cn(m): base_tons.get(_cn(m), 1.0) for m in membres}
                        tot = sum(membres_tons.values()) or 1.0
                        mon_ton = membres_tons.get(ck, 1.0)
                        result[client_raw] = round(grp_total_intrants * mon_ton / tot, 3)
                        assigned = True
                        break
            if assigned:
                continue

            # 3. Fuzzy match (seuil 0.55)
            best_k = max(sot_keys, key=lambda k: _sco(ck, k), default=None)
            if best_k and _sco(ck, best_k) >= 0.65:
                # Vérification supplémentaire : les noms sont vraiment proches
                result[client_raw] = sot_clean[best_k]
                continue

            # 4. Pas trouvé → 0
            result[client_raw] = 0.0

        # ── Assigner les intrants ──────────────────────────────────────
        base_df["total_intrants"] = base_df["client"].apply(
            lambda x: result.get(str(x).strip(), 0.0))

        return base_df

    base = _sot_merge(df_sotusfa_raw, df_sotusfa_pivot, base)

    # ── Merge QUANTITÉ ─────────────────────────────────────
    if df_quantite is not None and not df_quantite.empty:
        q = _upper(df_quantite.copy(), KEY)
        # Garder seulement les colonnes utiles de Quantite (pas hectares qui vient de Bourak)
        _q_keep = ["client","centre","qte_livree","qte_actif","qte_extra",
                   "tonnage_livre","prix_vente","commercial"]
        q = q[[c for c in _q_keep if c in q.columns]]
        base = base.merge(q, on=KEY, how="left", suffixes=("","_q"))
        # Résoudre conflits _q (garder valeur Bourak si présente)
        for _col_q in [c for c in base.columns if c.endswith("_q")]:
            _col_orig = _col_q[:-2]
            if _col_orig in base.columns:
                base[_col_orig] = base[_col_orig].fillna(base[_col_q])
                base = base.drop(columns=[_col_q])
            else:
                base = base.rename(columns={_col_q: _col_orig})

    # ── Merge PRÉVISIONS (concordance + fuzzy matching) ────
    for df_p, col in [(df_prev_dec,"prevision_dec"),
                      (df_prev_mai,"prevision_mai"),
                      (df_prev_juin,"prevision_juin")]:
        if df_p is not None and not df_p.empty and col in df_p.columns:
            p = df_p.copy()
            # Appliquer concordance sur les noms du fichier prévision
            if "client" in p.columns:
                p["client"] = p["client"].apply(
                    lambda x: _get_concordance_key(x) or x)
                p[col] = pd.to_numeric(p[col], errors="coerce").fillna(0)
                p = p.groupby("client")[col].sum().reset_index()
            # Merge avec fuzzy
            if "client" in p.columns and "client" in base.columns:
                base, n_m, n_t = _fuzzy_match_clients(base, p, col)
            else:
                base[col] = np.nan

    # ══ CALCULS ═══════════════════════════════════════════
    df = base.copy()

    # ── Plt Retour = Plt Livrés par défaut ───────────────────────
    # Si plt_retour = 0 (pas encore saisi), on considère
    # que tous les plateaux ont été retournés (pertes = 0)
    if "plt_livres" in df.columns:
        _plt_l_mc = pd.to_numeric(df["plt_livres"], errors="coerce").fillna(0)
        _plt_r_mc = pd.to_numeric(df.get("plt_retour", 0), errors="coerce").fillna(0)
        df["plt_retour"] = _plt_r_mc.where(_plt_r_mc > 0, _plt_l_mc)
        df["plt_perdus"]  = (_plt_l_mc - df["plt_retour"]).clip(lower=0)

    def g(col, d=0):
        """Getter sécurisé : retourne toujours une Series, jamais un scalaire."""
        if col in df.columns:
            return pd.to_numeric(df[col], errors="coerce").fillna(d)
        return pd.Series([d] * len(df), index=df.index, dtype=float)

    # Charges
    df["charge_plants"]   = g("valeur_plants")
    df["charge_intrants"] = g("total_intrants")
    df["avance_bourak"]   = g("avance")
    # Fallback avances depuis RECAP El Bourak
    try:
        import unicodedata as _uc_av, re as _re_av
        def _nrm_av(s):
            s = str(s).strip().upper()
            s = "".join(c for c in _uc_av.normalize("NFD",s) if _uc_av.category(c)!="Mn")
            return _re_av.sub(r"\s+"," ",_re_av.sub(r"[^A-Z0-9 ]"," ",s)).strip()
        _av_d = {_nrm_av(k): v for k,v in _AVANCES_2026.items()}
        _av_fb = df["client"].apply(lambda x: _av_d.get(_nrm_av(x), 0))
        df["avance_bourak"] = df["avance_bourak"].where(
            df["avance_bourak"].fillna(0) > 0, _av_fb)
    except Exception: pass
    # Charge totale = plants (Royal) + intrants (Sotusfa) + avance (Bourak)
    # RÈGLE : ne comptabiliser QUE les données réellement fournies
    df["charge_totale"] = (
        df["charge_plants"].fillna(0)       # Plants depuis Royal (0 si non fourni)
        + df["charge_intrants"].fillna(0)   # Intrants depuis Sotusfa (réel ou estimé confirmé)
        + df["avance_bourak"].fillna(0)     # Avance depuis RECAP (0 si non fourni)
    )
    # Assurer charge_intrants visible en DT dans l'export
    df["intrants_dt"] = df["charge_intrants"].fillna(0)

    # ── commercial : récupéré depuis Bourak (base) ───────────
    # La colonne "commercial" vient de parse_bourak (colonne "responsable")
    # Elle est déjà dans base depuis le début — on la normalise juste
    if "commercial" not in df.columns:
        # Pas dans df → chercher dans les sources sans merge supplémentaire
        for _src in [df_bourak, df_quantite]:
            if _src is not None and not _src.empty and "commercial" in _src.columns:
                _comm_map = (
                    _src[["client","commercial"]]
                    .copy()
                    .assign(client=lambda x: x["client"].astype(str).str.strip().str.upper())
                    .dropna(subset=["commercial"])
                    .query('commercial != ""')
                    .drop_duplicates("client")
                    .set_index("client")["commercial"]
                )
                df["commercial"] = df["client"].map(_comm_map).fillna("")
                if df["commercial"].ne("").any():
                    break
    if "commercial" not in df.columns:
        df["commercial"] = ""
    df["commercial"] = df["commercial"].fillna("").astype(str).str.strip()

    # Consigne caisse — PAR USINE (1ère affectation uniquement)
    caisses_par_usine = params.get("caisses_par_usine", {})
    # Fallback global si pas de config par usine
    _px_global  = params.get("prix_caisse", 3.0)
    _nb_global  = params.get("nb_caisses_ha", 80.0)

    def _calc_caisse(row):
        if not str(row.get("affectation_caisse","")).startswith("1ère"):
            return 0.0
        ha = float(row.get("hectares", 0) or 0)
        # Déterminer l'usine de l'agriculteur
        usine = str(row.get("usine", row.get("usine_livraison", ""))).upper().strip()
        # Chercher dans les usines connues
        cfg = None
        for u_key in caisses_par_usine:
            if u_key.upper() in usine or usine in u_key.upper():
                cfg = caisses_par_usine[u_key]
                break
        if cfg:
            return round(ha * cfg["nb_ha"] * cfg["prix"], 2)
        # Fallback global
        return round(ha * _nb_global * _px_global, 2)

    df["consigne_caisse"] = df.apply(_calc_caisse, axis=1)

    # Détail par usine pour affichage
    def _detail_caisse(row):
        if not str(row.get("affectation_caisse","")).startswith("1ère"):
            return "2ème — 0 DT"
        ha = float(row.get("hectares", 0) or 0)
        usine = str(row.get("usine", "")).upper().strip()
        for u_key, cfg in caisses_par_usine.items():
            if u_key.upper() in usine or usine in u_key.upper():
                nb = cfg["nb_ha"]; px = cfg["prix"]
                total = round(ha * nb * px, 0)
                return f"1ère — {int(ha*nb)} caisses × {px} DT = {total:,.0f} DT"
        nb = _nb_global; px = _px_global
        return f"1ère — {int(ha*nb)} caisses × {px} DT = {round(ha*nb*px):,.0f} DT"

    df["detail_caisse"] = df.apply(_detail_caisse, axis=1)

    df["consigne_plateau"] = g("consigne_plateau")
    # Consigne Plateau = 0 jusqu'à saisie des retours réels
    # (= décalage entre plateaux pris et retournés — donnée non encore disponible)
    df["consigne_plateau"] = df["consigne_plateau"].fillna(0)
    # Consigne caisse : 1ère affectation = ha × nb_caisses/ha × prix_caisse
    # Plants (calcul complet dans la section ci-dessous)
    # ── Plants et Ha (en premier car tout dépend de Ha) ──────────
    df["hectares"]    = g("hectares")      # Ha réels depuis Bourak
    # Enrichir Ha + variete + accessibilite + usine + zone depuis _PREVISION_2026
    # si absent ou 0 dans le fichier Bourak
    if "_ha_from_prev_done" not in dir():
        _base_ck = base["client"].astype(str).apply(
            lambda x: str(x).strip().upper()
        )
        import re as _re, unicodedata as _uc
        def _cn_local(n):
            n = str(n).strip().upper()
            n = _re.sub(r"[(][^)]*[)]","",n)
            n = "".join(c for c in _uc.normalize("NFD",n) if _uc.category(c)!="Mn")
            n = _re.sub(r"[^A-Z0-9 ]"," ",n)
            return _re.sub(r"[ ]+"," ",n).strip()
        _ha_prev = pd.to_numeric(df["hectares"].copy(), errors="coerce").fillna(0)
        for idx in df.index:
            if _ha_prev.loc[idx] <= 0:
                ck_i = _cn_local(str(df.loc[idx,"client"]))
                if ck_i in _PREVISION_2026:
                    _ha_prev.loc[idx] = _PREVISION_2026[ck_i].get("ha",0)
                else:
                    best_p = max(_PREVISION_2026.keys(),
                                 key=lambda k: sum(1 for w in set(ck_i.split()) & set(k.split())),
                                 default=None)
                    if best_p:
                        wck = set(ck_i.split()); wbp = set(best_p.split())
                        if len(wck&wbp)/max(len(wck|wbp),1) >= 0.60:
                            _ha_prev.loc[idx] = _PREVISION_2026[best_p].get("ha",0)
        df["hectares"] = _ha_prev
    # Enrichir variete, accessibilite, usine, zone depuis _PREVISION_2026
    for _col_prev, _key_prev in [("variete","variete_prev"),("acces","accessibilite"),
                                   ("usine","usine"),("zone","zone"),("region","region")]:
        _col_df = {"variete":"variete","acces":"acces","usine":"usine_prev",
                   "zone":"zone","region":"region"}.get(_col_prev, _col_prev)
        if _col_df not in df.columns or df[_col_df].astype(str).str.strip().isin(["","nan"]).all():
            df[_col_df] = df["client"].apply(
                lambda x: _PREVISION_2026.get(_cn_local(str(x)),{}).get(_col_prev,""))
    df["qte_livree"]  = g("qte_livree")    # Plants livrés
    df["qte_actif"]   = g("qte_actif")     # Plants actifs (pris racine)
    df["qte_extra"]   = g("qte_extra")     # Plants perdus
    df["qte_royal"]   = df["qte_livree"]   # alias

    _ha   = df["hectares"].fillna(0)
    _pl   = df["qte_livree"].fillna(0)
    _ha_s = _ha.where(_ha > 0, np.nan)    # NaN si ha=0 → résultats NaN→0
    _pl_s = _pl.where(_pl > 0, np.nan)

    # Taux prise et densité
    df["taux_prise"] = np.where(df["qte_livree"]>0,
                         (df["qte_actif"]/df["qte_livree"]*100).round(1), 0)
    df["densite_ha"] = (_pl / _ha_s).fillna(0).round(0)   # plants/ha

    # ── Prix vente ────────────────────────────────────────────────
    df["prix_vente"] = g("prix_vente")
    df["prix_vente"] = df["prix_vente"].where(df["prix_vente"]>0,
                        params.get("prix_vente_global", 270))

    # ── Tonnage livré et MO récolte ───────────────────────────────
    df["tonnage_livre"] = g("tonnage_livre")
    mo = params.get("mo_tonne", MO_TONNE_DEFAULT)
    df["mo_recolte"]    = (df["tonnage_livre"] * mo).round(0)

    # ── Consigne caisse + Caisses Vides ──────────────────────────
    _usine = params.get("usine_active", "SICAM")
    _pc = CAISSES_USINE_DEFAULTS.get(_usine, CAISSES_USINE_DEFAULTS.get("SICAM", {}))
    _nb_ha_c   = _pc.get("nb_ha", 80)     # caisses par hectare
    _prix_c    = _pc.get("prix", 3.0)     # prix consigne par caisse
    _is_1ere   = df["affectation_caisse"].astype(str).str.startswith("1ère")
    # Nb caisses vides nécessaires (pour la récolte)
    df["nb_caisses_vides"] = np.where(_is_1ere,
        (_ha * _nb_ha_c).round(0), 0)
    # Consigne caisse = nb_caisses × prix_consigne
    df["consigne_caisse"] = (df["nb_caisses_vides"] * _prix_c).round(0)

    # ── Charges totales ───────────────────────────────────────────
    charges_totales = (df["charge_totale"].fillna(0)
                     + df["consigne_plateau"].fillna(0)
                     + df["consigne_caisse"].fillna(0)
                     + df["mo_recolte"].fillna(0))
    df["charges_totales"] = charges_totales

    # ── Recouvrement ──────────────────────────────────────────────
    df["tonnage_recouvrement"] = np.where(df["prix_vente"]>0,
                                  (charges_totales / df["prix_vente"]).round(2), 0)

    # ── Charges à recouvrir (= tout ce que l'agri doit récupérer) ─
    df["charge_a_recouvrir"] = (charges_totales + df["report"].fillna(0)).round(0)

    # ── Indicateurs /ha ───────────────────────────────────────────
    df["recouvrement_ha"]   = (df["tonnage_recouvrement"] / _ha_s).fillna(0).round(2)
    df["rendement_ha_reel"] = (df["tonnage_livre"]        / _ha_s).fillna(0).round(1)
    df["cout_ha"]           = (df["charge_totale"].fillna(0) / _ha_s).fillna(0).round(0)
    df["cout_plant"]        = (df["charge_totale"].fillna(0) / _pl_s).fillna(0).round(4)

    # ── Prévision Mai ─────────────────────────────────────────────
    df["prevision_mai"]     = g("prevision_mai")
    # Fallback Prév. Mai: si valeur uploadée = 0, utiliser _PREVISION_2026["ton"]
    _pm_fb = {str(k).strip().upper(): float((v or {}).get("ton", 0) or 0)
              for k, v in _PREVISION_2026.items()}
    _mask_zero = df["prevision_mai"].fillna(0) == 0
    if _mask_zero.any():
        _fb = df.loc[_mask_zero, "client"].apply(
            lambda x: _pm_fb.get(str(x).strip().upper(), 0))
        df.loc[_mask_zero, "prevision_mai"] = _fb.values
    df["prevision_dec"]     = g("prevision_dec")
    df["prevision_juin"]    = g("prevision_juin")

    # ── Solde et valeur ───────────────────────────────────────────
    df["valeur_livree"] = (df["tonnage_livre"] * df["prix_vente"]).round(0)
    df["ecart_tonnage"] = (df["tonnage_livre"] - df["tonnage_recouvrement"]).round(2)
    df["solde_final"]   = (df["valeur_livree"] - charges_totales
                          - df["report"].fillna(0)).round(0)
    df["report"]        = g("report")

    # ── Ingénieur auto si absent ──────────────────────────────────
    if "ingenieur" not in df.columns or df["ingenieur"].fillna("").astype(str).eq("").all():
        df["ingenieur"] = ("ING. " + df["commercial"].astype(str).str[:8]).str.upper()

    # Alertes
    def _alerte(row):
        ecart      = row.get("ecart_tonnage", 0) or 0
        taux       = row.get("taux_prise", 100) or 100
        report_v   = row.get("report", 0) or 0
        charge     = row.get("charge_totale", 1) or 1
        prev_mai   = row.get("prevision_mai", 0) or 0
        recouvr    = row.get("tonnage_recouvrement", 0) or 0

        if ecart < -5:
            return "🔴 DÉFICIT RECOUVREMENT"
        if taux < 85:
            return "🔴 PRISE FAIBLE"
        if report_v > charge * 0.5:
            return "🔴 RISQUE FINANCIER"
        if recouvr > 0 and prev_mai > 0:
            ratio = prev_mai / recouvr
            if ratio < 0.90: return "🔴 PRÉVISION INSUFFISANTE"
            if ratio < 1.00: return "🟡 ATTENTION"
        if ecart >= 0:
            return "🟢 OK"
        return "🟡 ATTENTION"

    df["alerte"] = df.apply(_alerte, axis=1)

    # Renommer client → agriculteur pour affichage
    df = df.rename(columns={"client":"agriculteur"})
    return df


# ══════════════════════════════════════════════════════════════
# EXPORT EXCEL
# ══════════════════════════════════════════════════════════════

def export_excel(df, df_sotusfa_raw=None):
    """
    Export Excel exact — structure tirée du fichier de référence :
    35 colonnes · 7 groupes · 4 feuilles
    """
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule
    import io as _io, numpy as _np

    # ── helpers ─────────────────────────────────────────────────
    def hf(h): return PatternFill("solid", start_color=h, end_color=h)
    def bf(bold=True, white=False, size=10, color="000000"):
        return Font(bold=bold, name="Calibri", size=size,
                    color="FFFFFF" if white else color)
    T  = Side(style="thin",   color="CCCCCC")
    TM = Side(style="medium", color="444444")
    BD  = Border(left=T,  right=T,  top=T,  bottom=T)
    BDM = Border(left=TM, right=TM, top=TM, bottom=TM)
    CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LFT = Alignment(horizontal="left",   vertical="center")

    # ── Mapping : nom interne → nom affiché ─────────────────────
    # Ordre : colonnes multiples possibles (prend la première trouvée)
    COL_MAP = {
        # Nom affiché          : [noms internes possibles]
        "Agriculteur"         : ["Agriculteur","agriculteur","client"],
        "Commercial"          : ["Commercial","commercial"],
        "Ingénieur"           : ["Ingénieur","ingenieur"],
        "Centre"              : ["Centre","centre"],
        "Région"              : ["Région","region"],
        "Variété"             : ["Variété","variete","variete_prev"],
        "Accessibilité"       : ["Accessibilité","acces","accessibilite"],
        "Usine"               : ["Usine","usine","usine_prev"],
        "Zone"                : ["Zone","zone"],
        "Ha"                  : ["Ha","hectares"],
        "Plants Livrés"       : ["Plants Livrés","qte_royal"],
        "Plants Actifs"       : ["Plants Actifs","qte_actif"],
        "Extra (pertes)"      : ["Extra (pertes)","qte_extra"],
        "Taux prise %"        : ["Taux prise %","taux_prise"],
        "Densité/ha"          : ["Densité/ha","densite_ha"],
        "Plt Livrés"          : ["Plt Livrés","plt_livres","nb_plateaux"],
        "Plt Retour"          : ["Plt Retour","plt_retour"],
        "Plt Perdus"          : ["Plt Perdus","plt_perdus"],
        "Caisses Vides"       : ["Caisses Vides","nb_caisses_vides","nb_caisses","caisses_vides"],
        "Affectation"         : ["Affectation","affectation_caisse"],
        "Déb. Récolte"        : ["Déb. Récolte","date_debut_recolte"],
        "Plants (DT)"         : ["Plants (DT)","charge_plants","valeur_plants"],
        "Intrants (DT)"       : ["Intrants (DT)","charge_intrants","total_intrants",
                               "intrants_dt","intrants"],
        "Avance Bourak (DT)"  : ["Avance Bourak (DT)","avance_bourak"],
        "Charge Totale (DT)"  : ["Charge Totale (DT)","charge_totale"],
        "Consigne Plateau"    : ["Consigne Plateau","consigne_plateau"],
        "Report (DT)"         : ["Report (DT)","report"],
        "Consigne Caisse"     : ["Consigne Caisse","consigne_caisse"],
        "MO Récolte (DT)"     : ["MO Récolte (DT)","mo_recolte"],
        "Charges à recouvrir" : ["Charges à recouvrir","charge_a_recouvrir"],
        "Prév. Mai (T)"       : ["Prév. Mai (T)","prevision_mai"],
        "Livré (T)"           : ["Livré (T)","tonnage_livre"],
        "Prix Vente"          : ["Prix Vente","prix_vente"],
        "RECOUVREMENT (T)"    : ["RECOUVREMENT (T)","tonnage_recouvrement"],
        "Recouv./ha"          : ["Recouv./ha","recouvrement_ha"],
        "Écart (T)"           : ["Écart (T)","ecart_tonnage"],
        "T/ha réalisé"        : ["T/ha réalisé","rendement_ha_reel"],
        "Coût/ha"             : ["Coût/ha","cout_ha"],
        "Coût/plant"          : ["Coût/plant","cout_plant"],
        "Valeur Livrée"       : ["Valeur Livrée","valeur_livree"],
        "Solde Final"         : ["Solde Final","solde_final"],
        "Alerte"              : ["Alerte","alerte"],
    }

    # ── Structure EXACTE (groupes → colonnes dans l'ordre) ──────
    GROUPES = {
        "IDENTIFICATION": [
            "Agriculteur","Commercial","Ingénieur","Centre","Région"],
        "PLANT": [
            "Variété","Ha","Plants Livrés","Plants Actifs",
            "Extra (pertes)","Taux prise %","Densité/ha"],
        "PLATEAUX": [
            "Plt Livrés","Plt Retour"],
        "AFFECTATION CAISSES VIDES": [
            "Caisses Vides","Affectation","Déb. Récolte"],
        "CHARGES (DT)": [
            "Plants (DT)","Intrants (DT)","Avance Bourak (DT)",
            "Charge Totale (DT)","Consigne Plateau","Report (DT)",
            "Consigne Caisse","MO Récolte (DT)","Charges à recouvrir"],
        "PRÉVISIONS (T)": [
            "Prév. Mai (T)","Livré (T)"],
        "RECOUVREMENT": [
            "Prix Vente","RECOUVREMENT (T)","Recouv./ha","Écart (T)"],
        "RÉSULTAT": [
            "T/ha réalisé","Coût/ha","Coût/plant",
            "Valeur Livrée","Solde Final","Alerte"],
    }
    GRP_COLORS = {
        "IDENTIFICATION":            "1F3864",
        "PLANT":                     "1A5C2A",
        "PLATEAUX":                  "0B4F6C",
        "AFFECTATION CAISSES VIDES": "7B3F00",
        "CHARGES (DT)":              "8B0000",
        "PRÉVISIONS (T)":            "4A235A",
        "RECOUVREMENT":              "0B3954",
        "RÉSULTAT":                  "1B4332",
    }
    # Couleurs spéciales par colonne
    COL_SUBCOLORS = {
        "Taux prise %": "2D6A4F",
        "Densité/ha":   "1A5C2A",
        "Report (DT)":  "6B1212",
    }

    # ── Résoudre les valeurs depuis df ──────────────────────────
    # Pour chaque nom affiché, trouver la colonne dans df
    def resolve(df_, display_name):
        for internal in COL_MAP.get(display_name, [display_name]):
            if internal in df_.columns:
                return df_[internal]
        return pd.Series([""] * len(df_), index=df_.index)

    # Construire la liste ordonnée finale de colonnes à afficher
    all_display = []
    for grp_cols in GROUPES.values():
        for col_display in grp_cols:
            all_display.append(col_display)

    wb = Workbook()

    # ════════════════════════════════════════════════════════════
    # FEUILLE 1 — 📊 Dashboard
    # ════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "📊 Dashboard"
    ws.sheet_view.showGridLines = False

    ncols = len(all_display)

    # Ligne 1 : titre principal
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws["A1"] = "📊 DASHBOARD AGROÉCONOMIQUE TOMATE 2026 — v2"
    ws["A1"].font = bf(True, True, 13)
    ws["A1"].fill = hf("0a1628")
    ws["A1"].alignment = CTR
    ws.row_dimensions[1].height = 32

    # Ligne 2 : groupes (cellules fusionnées)
    # Ligne 3 : noms colonnes
    col_cursor = 1
    for grp_name, grp_cols in GROUPES.items():
        gc = GRP_COLORS[grp_name]
        nc = len(grp_cols)
        # Fusion groupe
        if nc > 1:
            ws.merge_cells(start_row=2, start_column=col_cursor,
                           end_row=2, end_column=col_cursor + nc - 1)
        g = ws.cell(2, col_cursor, value=grp_name)
        g.font = bf(True, True, 10); g.fill = hf(gc)
        g.alignment = CTR; g.border = BDM

        for col_display in grp_cols:
            # Sous-couleur si définie, sinon couleur du groupe
            sub_c = COL_SUBCOLORS.get(col_display, gc)
            h = ws.cell(3, col_cursor, value=col_display)
            h.font = bf(True, True, 9)
            h.fill = hf(sub_c)
            h.alignment = CTR
            h.border = BD
            # Largeur colonne adaptée
            _w = max(12, len(col_display) + 3)
            if col_display in ("Agriculteur",): _w = 30
            elif col_display in ("Commercial","Ingénieur","Centre"): _w = 18
            elif col_display in ("Plants (DT)","Intrants (DT)","Avance Bourak (DT)",
                                  "Charge Totale (DT)","Charges à recouvrir",
                                  "Consigne Plateau","Consigne Caisse","MO Récolte (DT)",
                                  "Report (DT)"): _w = 17
            elif col_display in ("RECOUVREMENT (T)","Valeur Livrée","Solde Final"): _w = 16
            ws.column_dimensions[get_column_letter(col_cursor)].width = _w
            col_cursor += 1

    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 30

    # Lignes données
    ALT_BG = {"🔴": "FFCDD2", "🟡": "FFF9C4", "🟢": "E8F5E9"}
    for ri, (_, row) in enumerate(df.iterrows()):
        r = ri + 4
        alerte_raw = str(resolve(df, "Alerte").iloc[ri] if ri < len(df) else "")
        emoji = alerte_raw[:2] if alerte_raw else ""
        row_bg = ALT_BG.get(emoji, "F0F5FF" if ri % 2 == 0 else "FFFFFF")

        for ci, col_display in enumerate(all_display, 1):
            series = resolve(df, col_display)
            val = series.iloc[ri] if ri < len(series) else ""
            if isinstance(val, float) and _np.isnan(val): val = ""

            c = ws.cell(r, ci, value=val)
            c.border = BD
            c.alignment = LFT if col_display == "Agriculteur" else CTR
            c.font = bf(col_display == "Agriculteur", size=9)

            # ── Style par colonne ──────────────────────────────
            if col_display == "Alerte":
                c.fill = hf(ALT_BG.get(emoji, row_bg))
                c.font = bf(True, size=9, color={
                    "🔴":"C0392B","🟡":"D4AC0D","🟢":"1E8449"}.get(emoji,"000000"))
            elif col_display == "Solde Final":
                try:
                    fv = float(val) if val != "" else 0
                    c.fill = hf("E8F5E9") if fv >= 0 else hf("FFEBEE")
                    c.font = bf(True, size=9, color="1E8449" if fv >= 0 else "C0392B")
                    c.number_format = '+#,##0 "DT";-#,##0 "DT";0'
                except: c.fill = hf(row_bg)
            elif col_display == "Écart (T)":
                try:
                    fv = float(val) if val != "" else 0
                    c.fill = hf("E8F5E9") if fv >= 0 else hf("FFEBEE")
                    c.number_format = '+#,##0.0;-#,##0.0;0'
                except: c.fill = hf(row_bg)
            elif col_display == "Taux prise %":
                c.fill = hf(row_bg)
                try:
                    tp = float(val)
                    if tp >= 90: c.fill = hf("E8F5E9")
                    elif tp >= 85: c.fill = hf("F0F4C3")
                    else: c.fill = hf("FFEBEE")
                except: pass
                c.number_format = "0.0"
            else:
                c.fill = hf(row_bg)
                # Formats numériques
                if isinstance(val, (int, float)) and val != "" and not _np.isnan(float(val) if isinstance(val,float) else 0):
                    if col_display in ("Ha","T/ha réalisé","Recouv./ha","Coût/plant"):
                        c.number_format = "0.00"
                    elif col_display in ("Plants Livrés","Plants Actifs","Extra (pertes)",
                                          "Densité/ha"):
                        c.number_format = "#,##0"
                    elif col_display in ("Prix Vente",):
                        c.number_format = "#,##0.0"
                    else:
                        c.number_format = "#,##0"

        ws.row_dimensions[r].height = 17

    # Ligne TOTAL
    tr = len(df) + 4
    SUM_COLS = {"Ha","Plants Livrés","Plants Actifs","Extra (pertes)",
                "Plants (DT)","Intrants (DT)","Avance Bourak (DT)",
                "Charge Totale (DT)","Consigne Plateau","Report (DT)",
                "Consigne Caisse","MO Récolte (DT)","Charges à recouvrir",
                "Prév. Mai (T)","Livré (T)","RECOUVREMENT (T)",
                "Écart (T)","Valeur Livrée","Solde Final"}
    for ci, col_display in enumerate(all_display, 1):
        c = ws.cell(tr, ci)
        c.fill = hf("1F3864"); c.font = bf(True, True, 9)
        c.border = BD; c.alignment = CTR
        if col_display == "Agriculteur":
            c.value = "TOTAL"; c.alignment = LFT
        elif col_display in SUM_COLS:
            cl = get_column_letter(ci)
            c.value = f"=SUM({cl}4:{cl}{tr-1})"
            c.number_format = "#,##0"
    ws.row_dimensions[tr].height = 22

    # Mise en forme conditionnelle taux prise
    tp_idx = all_display.index("Taux prise %") + 1 if "Taux prise %" in all_display else None
    if tp_idx:
        tl = get_column_letter(tp_idx)
        ws.conditional_formatting.add(
            f"{tl}4:{tl}{tr-1}",
            ColorScaleRule(start_type="num", start_color="FFCDD2",
                           mid_type="num",   mid_value=90, mid_color="FFF9C4",
                           end_type="num",   end_color="C8E6C9", end_value=97))
    ws.freeze_panes = "A4"

    # ════════════════════════════════════════════════════════════
    # FEUILLE 2 — 👤 Par Ingénieur
    # ════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("👤 Par Ingénieur")
    ws2.sheet_view.showGridLines = False

    _ic = next((c for c in ["Ingénieur","ingenieur"] if c in df.columns), None)
    _cc = next((c for c in ["Centre","centre"]       if c in df.columns), None)
    _ac = next((c for c in ["Agriculteur","agriculteur","client"] if c in df.columns), None)

    ws2.merge_cells("A1:L1")
    ws2["A1"] = "👤 SYNTHÈSE PAR INGÉNIEUR / CENTRE"
    ws2["A1"].font = bf(True, True, 12); ws2["A1"].fill = hf("0B4F6C")
    ws2["A1"].alignment = CTR; ws2.row_dimensions[1].height = 30

    if _ic and _ic in df.columns:
        _gk = [k for k in [_ic, _cc] if k and k in df.columns]
        _agg = {}
        if _ac: _agg["Agriculteurs"] = (_ac, "count")
        for _nc, _fc in [
            ("Ha","hectares"),("Plants Livrés","qte_royal"),
            ("Plants Actifs","qte_actif"),("Taux prise %","taux_prise"),
            ("Charge Totale (DT)","charge_totale"),("Livré (T)","tonnage_livre"),
            ("RECOUVREMENT (T)","tonnage_recouvrement"),("Écart (T)","ecart_tonnage"),
        ]:
            _fc_found = next((c for c in [_fc, _nc] if c in df.columns), None)
            if _fc_found:
                _agg[_nc] = (_fc_found, "mean" if _nc == "Taux prise %" else "sum")
        if "alerte" in df.columns:
            _agg["Alertes 🔴"] = ("alerte", lambda x: x.str.contains("🔴", na=False).sum())
        elif "Alerte" in df.columns:
            _agg["Alertes 🔴"] = ("Alerte", lambda x: x.str.contains("🔴", na=False).sum())
        try:
            _gi = df.groupby(_gk, as_index=False).agg(**_agg).round(1)
        except Exception:
            _gi = pd.DataFrame()

        if not _gi.empty:
            for ci, col in enumerate(_gi.columns, 1):
                h = ws2.cell(2, ci, value=col)
                h.font = bf(True, True, 10); h.fill = hf("0B4F6C")
                h.alignment = CTR; h.border = BD
                ws2.column_dimensions[get_column_letter(ci)].width = max(15, len(str(col)) + 4)
            ws2.row_dimensions[2].height = 28
            for ri, (_, row) in enumerate(_gi.iterrows()):
                r = ri + 3
                bg = "F0F5FF" if ri % 2 == 0 else "FFFFFF"
                for ci, val in enumerate(row.values, 1):
                    if isinstance(val, float) and _np.isnan(val): val = ""
                    c = ws2.cell(r, ci, value=val)
                    c.border = BD; c.fill = hf(bg); c.alignment = CTR
                    c.font = bf(False, size=9)
                    if ci <= len(_gk): c.alignment = LFT; c.font = bf(True, size=9)
                    if isinstance(val, (int, float)) and val != "":
                        c.number_format = "0.0" if list(_gi.columns)[ci-1] == "Taux prise %" else "#,##0"
    ws2.freeze_panes = "A3"

    # ════════════════════════════════════════════════════════════
    # FEUILLE 3 — 📦 Caisses Vides
    # ════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("📦 Caisses Vides")
    ws3.sheet_view.showGridLines = False

    _cv_src = [
        ("Agriculteur",["Agriculteur","agriculteur","client"]),
        ("Centre",     ["Centre","centre"]),
        ("Région",     ["Région","region"]),
        ("Affectation",["Affectation","affectation_caisse"]),
        ("Détail",     ["detail_caisse"]),
        ("Déb. Récolte",["Déb. Récolte","date_debut_recolte"]),
        ("Ha",         ["Ha","hectares"]),
        ("Consigne Caisse",  ["Consigne Caisse","consigne_caisse"]),
        ("Consigne Plateau", ["Consigne Plateau","consigne_plateau"]),
    ]
    _cv = [(disp, next((c for c in srcs if c in df.columns), None))
           for disp, srcs in _cv_src]
    _cv = [(d, s) for d, s in _cv if s]

    ws3.merge_cells(f"A1:{get_column_letter(len(_cv))}1")
    ws3["A1"] = "📦 CAISSES VIDES — Affectations & Consignes"
    ws3["A1"].font = bf(True, True, 12); ws3["A1"].fill = hf("7B3F00")
    ws3["A1"].alignment = CTR; ws3.row_dimensions[1].height = 30

    for ci, (disp, _) in enumerate(_cv, 1):
        h = ws3.cell(2, ci, value=disp)
        h.font = bf(True, True, 10); h.fill = hf("7B3F00")
        h.alignment = CTR; h.border = BD
        ws3.column_dimensions[get_column_letter(ci)].width = max(16, len(disp) + 4)
    ws3.row_dimensions[2].height = 28

    for ri, (_, row) in enumerate(df.iterrows()):
        r = ri + 3
        aff = str(row.get("affectation_caisse", row.get("Affectation", "")))
        bg = "FBE9E7" if "1ère" in aff else ("F0F5FF" if ri % 2 == 0 else "FFFFFF")
        for ci, (_, src_col) in enumerate(_cv, 1):
            val = row.get(src_col, "")
            if isinstance(val, float) and _np.isnan(val): val = ""
            c = ws3.cell(r, ci, value=val)
            c.border = BD; c.fill = hf(bg); c.alignment = CTR; c.font = bf(False, size=9)
            if ci == 1: c.alignment = LFT; c.font = bf(True, size=9)
            if isinstance(val, (int, float)) and val != "": c.number_format = "#,##0"
    ws3.freeze_panes = "A3"

    # ════════════════════════════════════════════════════════════
    # FEUILLE 4 — 📈 Prévisions
    # ════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("📈 Prévisions")
    ws4.sheet_view.showGridLines = False

    _pv_src = [
        ("Agriculteur",      ["Agriculteur","agriculteur","client"]),
        ("Centre",           ["Centre","centre"]),
        ("Prév. Déc (T)",    ["Prév. Déc (T)","prevision_dec"]),
        ("Prév. Mai (T)",    ["Prév. Mai (T)","prevision_mai"]),
        ("Prév. Juin (T)",   ["Prév. Juin (T)","prevision_juin"]),
        ("Livré (T)",        ["Livré (T)","tonnage_livre"]),
        ("RECOUVREMENT (T)", ["RECOUVREMENT (T)","tonnage_recouvrement"]),
        ("Recouv./ha",       ["Recouv./ha","recouvrement_ha"]),
        ("Écart (T)",        ["Écart (T)","ecart_tonnage"]),
    ]
    _pv = [(d, next((c for c in srcs if c in df.columns), None)) for d, srcs in _pv_src]
    _pv = [(d, s) for d, s in _pv if s]

    ws4.merge_cells(f"A1:{get_column_letter(len(_pv))}1")
    ws4["A1"] = "📈 PRÉVISIONS vs RÉALISÉ"
    ws4["A1"].font = bf(True, True, 12); ws4["A1"].fill = hf("4A235A")
    ws4["A1"].alignment = CTR; ws4.row_dimensions[1].height = 30

    for ci, (disp, _) in enumerate(_pv, 1):
        h = ws4.cell(2, ci, value=disp)
        h.font = bf(True, True, 10); h.fill = hf("4A235A")
        h.alignment = CTR; h.border = BD
        ws4.column_dimensions[get_column_letter(ci)].width = max(16, len(disp) + 4)
    ws4.row_dimensions[2].height = 28

    for ri, (_, row) in enumerate(df.iterrows()):
        r = ri + 3
        bg = "F0F5FF" if ri % 2 == 0 else "FFFFFF"
        for ci, (disp, src_col) in enumerate(_pv, 1):
            val = row.get(src_col, "")
            if isinstance(val, float) and _np.isnan(val): val = ""
            c = ws4.cell(r, ci, value=val)
            c.border = BD; c.alignment = CTR; c.font = bf(False, size=9)
            if disp == "Écart (T)" and isinstance(val, (int, float)) and val != "":
                try:
                    fv = float(val)
                    c.fill = hf("E8F5E9") if fv >= 0 else hf("FFEBEE")
                    c.number_format = "+#,##0.0;-#,##0.0;0"
                except: c.fill = hf(bg)
            else:
                c.fill = hf(bg)
                if isinstance(val, (int, float)) and val != "":
                    c.number_format = "0.00" if disp in ("Recouv./ha",) else "#,##0.0"
            if ci <= 2: c.alignment = LFT; c.font = bf(True, size=9)
    ws4.freeze_panes = "A3"

    buf = _io.BytesIO()
    wb.save(buf); buf.seek(0)
    return buf.read()



def _get_concordance_key(nom_ref):
    """Trouve le nom canonique correspondant via la table de concordance.
    La table est définie localement pour éviter tout NameError.
    """
    import unicodedata as _uc, re as _re

    # Table de concordance LOCALE (robuste — pas de variable globale requise)
    _CONC = {
        # KHALIL
        "NEJI ZAAFOURI":           "NEGI ZAAFOURI",
        "HEDI SLEMA":              "HEDI SLAMA",
        "SAMIR ATTIAA":            "SAMIR ATTIYA",
        "BOUBAKER FILELI":         "BOUBAKER FILALI",
        "KAIS EDHAOUI":            "KAIS DHAOUI",
        "EZZEDDIN ELGUESMI":       "EZZEDINE GUESMI",
        "MOURAD BEN SAID HAMMADI": "MOURAD HEMMEDI",
        "SAMI BEN AMOR FERJENI":   "SAMI FERGENI",
        "SALEM ELMEJRI":           "SALEM EL MEJRI",
        # MAKKI
        "ALI EL KOTLI":            "ALI KOTLI",
        "SASSI BEN MANSOUR":       "SASSI MANSOUR",
        "ABDELAZIZ LAYARI":        "ABEDLAZIZ LAYARI",
        "ABDERRAZEK BEY":          "ABEDRAZEK BEY",
        "MAKREM HAFFAR":           "MAKRAM HAFFAR",
        "SALAH BEN HAMOUDA":       "SALEH BEN HAMOUDA",
        "LASSAAD NEILI":           "LASSED NEILI",
        "ALAEDDINE BEN KILANI":    "ALAEDINE KILENI",
        "ADEL ALJAZI":             "ADEL JAZI",
        "MOHAMED BADIA NEJI":      "MOHAMED BEDIA NEJI",
        "SLAH BEN SLIMEN":         "SLAH BEN ABDALLAH",
        "ROMDHAN ELMEHEDEBI":      "RAMDHAN MHEDHBI",
        "AYMEN CHAABEN":           "AYMEN CHABEN",
        "SLAH BANNI":              "SLAH BANI",
        "SAMAH BACCOUCH":          "SAMEH BACCOUCH",
        "MOUHAMED GHARBI":         "MOHAMED GHARBI",
        "ZOUHAIR BEAICH":          "ZOUHAIR BAICH",
        # FEDI
        "ABDELFATEH BEN SLIMENE":  "ABDELFATEH BEN SLIMEN",
        "HAMED BEN YOUNES":        "HAMED BEN YOUNIS",
        "SAMI BEN HEDI KAAB":      "SAMI KAAB",
        "TAREK BEN ABDALLAH":      "TAREK BEN ABDALAH",
        "TAREK ELBAHRI":           "TAREK EL BAHRI",
        "SOCIETE BACCARA ET FILS": "STE BACCARA",
        "NEJIB BAKOUCHE":          "NAJIB BACCOUCH",
        "HASSEN BEN ALAYA":        "HASSEN BEN ALIA",
        "ANIS DHAOUADI":           "ANIS DHAWADI",
        "MAHER BELHAJ SALAH":      "MAHER BELHAJ FRAJ",
        "HANI BELKILANI":          "HANI BEN KILANI",
        "AHMED ELIDRISSI":         "AHMED IDRISSI",
        "HAMMADI BEN ZRIBIA":      "HAMMADI BENZRIBIA",
        "OSAMA KAAB":              "SAMI KAAB",
        "SOFYEN GHZALA":           "SOFIENNE GHZELA",
        "MOUHAMED ALI GHZALA":     "MOHAMED ALI GHZELA",
        "MOUHAMED ALI BELMADHI":   "MOHAMED BEL MADHI",
        "MED MANOUBI":             "MOHAMED MANNOUBI",
        # ACHREF (centres → sous-membres)
        "ABDELKARIM GARMALLAH":    "KARIM GARMALAH 1",
        "SOCIETE BILEL GHA SERVICE AGRICOLE": "BILEL GHA 1",
        "SEBTI JBALLAH":           "SEBTI JABALI",
        "SOUHAIEL BOUZ":           "SOUHAIL BOUZANA",
        "HAFEDH MESBEH":           "HAFEDH MOSBEH",
        "HAFEDH MOSBE":            "HAFEDH MOSBEH",
        "KARIM GARMAL":            "KARIM GARMALAH 1",
        # JILANI
        "SLIM MARZOUGUI":          "Slim Marzougui",
        "SLIM ELMARZOUGUI":        "Slim Marzougui",
        "AHMED BALAGUI":           "Ahmed Ballagui",
        "RIADH KOUKI":             "Riadh Kouki",
        "IMED AMDOU":              "Imed Amdouni",
        "NEJIB MECHRG":            "Nejib Mechrgui",
    }

    def _norm(s):
        s = str(s).strip().upper()
        s = ''.join(c for c in _uc.normalize('NFD', s)
                    if _uc.category(c) != 'Mn')
        s = _re.sub(r'[(][^)]*[)]', ' ', s)
        s = _re.sub(r'[^A-Z0-9 ]', ' ', s)
        return _re.sub(r'\s+', ' ', s).strip()

    nom_up    = str(nom_ref).strip().upper()
    nom_clean = _norm(nom_up)

    # 1. Recherche exacte
    for k, v in _CONC.items():
        if k.upper() == nom_up:
            return v
    # 2. Recherche normalisée
    for k, v in _CONC.items():
        if _norm(k) == nom_clean:
            return v
    return None

def _fuzzy_match_clients(df_base, df_prev, col_prev):
    """Merge fuzzy SANS double comptage. Un prev_key → un seul base_key."""
    import unicodedata as _uni, re as _re2

    def _clean(s):
        s = str(s).upper().strip()
        s = _re2.sub(r"\bSOCIETE\b", "STE", s)
        s = "".join(c for c in _uni.normalize("NFD", s) if _uni.category(c) != "Mn")
        s = _re2.sub(r"\([^)]*\)", " ", s)
        s = _re2.sub(r"[^A-Z0-9 ]", " ", s)
        return _re2.sub(r"\s+", " ", s).strip()

    def _score(a, b):
        wa, wb = set(a.split()), set(b.split())
        if not wa or not wb: return 0.0
        inter = len(wa & wb); union = len(wa | wb)
        sj = inter / union if union else 0
        sh, lo = (wa, wb) if len(wa) <= len(wb) else (wb, wa)
        sc = sum(1 for w in sh if any(lw.startswith(w[:4]) for lw in lo)
                 and len(w) > 2) / max(len(sh), 1) * 0.85
        return max(sj, sc)

    THRESHOLD = 0.38
    df_base = df_base.copy()
    df_prev = df_prev.copy()
    df_base["_km"] = df_base["client"].apply(_clean)
    df_prev["_km"] = df_prev["client"].apply(_clean)

    # Agréger prévisions par _km (sécurité contre doublons)
    prev_agg = df_prev.groupby("_km", as_index=False)[col_prev].sum()
    prev_dict = dict(zip(prev_agg["_km"], prev_agg[col_prev]))

    # Merge exact
    result = df_base.merge(prev_agg, on="_km", how="left")

    # Fuzzy pour non-matchés (bijectif : chaque prev_key → 1 base_key max)
    unmatched_mask = result[col_prev].isna()
    if unmatched_mask.any():
        used_prev = set(result.loc[~unmatched_mask, "_km"].values)
        avail = {k: v for k, v in prev_dict.items() if k not in used_prev}
        assigned = {}
        for bk in result.loc[unmatched_mask, "_km"].unique():
            best_sc = 0; best_v = None
            for pk, pv in avail.items():
                sc = _score(bk, pk)
                if sc > best_sc and sc >= THRESHOLD:
                    best_sc = sc; best_v = pv
            if best_v is not None:
                assigned[bk] = best_v
        for bk, val in assigned.items():
            result.loc[result["_km"] == bk, col_prev] = val

    result = result.drop(columns=["_km"])
    return result, result[col_prev].notna().sum(), len(result)


def _export_excel_table(df, sheet_title="Data",
                        header_text="Export", color_hex="1F3864"):
    """Excel formaté attractif pour n'importe quel DataFrame."""
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io as _io, numpy as _np
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]
    ws.sheet_view.showGridLines = False
    T = Side(style="thin", color="CCCCCC")
    BD = Border(left=T, right=T, top=T, bottom=T)
    CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LFT = Alignment(horizontal="left", vertical="center")
    def hf(h): return PatternFill("solid", start_color=h, end_color=h)
    def bf(bold=True, white=False, size=10, color=None):
        """bf local : supporte white=True (blanc) ET color hex explicite."""
        if white:
            final_color = "FFFFFF"
        elif color:
            final_color = str(color).lstrip("#")
        else:
            final_color = "000000"
        return Font(bold=bold, name="Calibri", size=size, color=final_color)
    nc = max(len(df.columns), 1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=nc)
    ws["A1"] = header_text
    ws["A1"].font = bf(True, True, 12)
    ws["A1"].fill = hf(color_hex)
    ws["A1"].alignment = CTR
    ws.row_dimensions[1].height = 30
    for ci, col in enumerate(df.columns, 1):
        c = ws.cell(2, ci, value=str(col))
        c.font = bf(True, True, 10)
        c.fill = hf(color_hex)
        c.alignment = CTR
        c.border = BD
        ws.column_dimensions[get_column_letter(ci)].width = max(14, len(str(col)) + 4)
    ws.row_dimensions[2].height = 28
    ALERTE_COLORS = {
        "🔴":("FFCDD2","C0392B"), "🟡":("FFF9C4","D4AC0D"),
        "🟢":("E8F5E9","1E8449"),
    }
    # Détecter colonnes numériques
    _num_cols = {col: i+1 for i, col in enumerate(df.columns)
                 if str(df[col].dtype).startswith(("int","float"))}

    for ri, (_, row) in enumerate(df.iterrows()):
        r = ri + 3
        # Couleur ligne selon alerte si présente
        alerte_val = str(row.get("alerte","")) if "alerte" in df.columns else ""
        if "🔴" in alerte_val: row_bg = "FFEBEE"
        elif "🟡" in alerte_val: row_bg = "FFF9E6"
        elif "🟢" in alerte_val: row_bg = "E8F5E9"
        else: row_bg = "F0F5FF" if ri % 2 == 0 else "FFFFFF"

        for ci, val in enumerate(row, 1):
            col_name = df.columns[ci-1]
            if isinstance(val, float) and _np.isnan(val):
                val = ""
            c = ws.cell(r, ci, value=val)
            c.border = BD
            c.alignment = LFT if ci == 1 else CTR
            c.font = bf(ci == 1, size=9)

            # Couleur spéciale selon colonne
            if col_name == "alerte" and val:
                for emoji,(bg2,fg2) in ALERTE_COLORS.items():
                    if emoji in str(val):
                        c.fill = hf(bg2)
                        c.font = bf(True, size=9, color=fg2)
                        break
                else:
                    c.fill = hf(row_bg)
            elif col_name in ("ecart_tonnage","solde_final") and val != "" and val is not None:
                try:
                    fv = float(val)
                    c.fill = hf("E8F5E9") if fv >= 0 else hf("FFEBEE")
                    c.font = bf(True, size=9, color="1E8449" if fv >= 0 else "C0392B")
                    c.number_format = "+#,##0;-#,##0;0"
                except (TypeError, ValueError):
                    c.fill = hf(row_bg)
            elif col_name == "taux_prise" and isinstance(val,(int,float)) and val==val:
                v = float(val)
                c.fill = hf("E8F5E9" if v>=90 else ("FFF9E6" if v>=85 else "FFEBEE"))
                c.number_format = "0.0"
            elif col_name == "affectation_caisse":
                c.fill = hf("FBE9E7") if "1ère" in str(val) else hf("E8F5E9")
            else:
                c.fill = hf(row_bg)

            if val != "" and val is not None and col_name not in ("ecart_tonnage","solde_final","taux_prise"):
                try:
                    fv2 = float(val)
                    if fv2 == fv2:  # pas NaN
                        c.number_format = "#,##0" if abs(fv2) >= 100 else "0.0"
                except (TypeError, ValueError):
                    pass
    num_ci = [i + 1 for i, col in enumerate(df.columns)
              if str(df[col].dtype).startswith(("int", "float"))]
    if num_ci:
        tr = len(df) + 3
        for ci in range(1, nc + 1):
            c = ws.cell(tr, ci)
            c.fill = hf(color_hex)
            c.font = bf(True, True)
            c.border = BD
            c.alignment = CTR
        ws.cell(tr, 1).value = "TOTAL"
        ws.cell(tr, 1).alignment = LFT
        for ci in num_ci:
            col_l = get_column_letter(ci)
            ws.cell(tr, ci).value = f"=SUM({col_l}3:{col_l}{tr-1})"
            ws.cell(tr, ci).number_format = "#,##0"
        ws.row_dimensions[tr].height = 22
    ws.freeze_panes = "A3"
    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ══════════════════════════════════════════════════════════════
# SESSION PERSISTANCE — Supabase (table agroeco_session)
# ══════════════════════════════════════════════════════════════

def _df_to_b64(df):
    """Sérialise un DataFrame en base64 gzip — robuste."""
    if df is None: return None
    try:
        if hasattr(df, "empty") and df.empty: return None
        df2 = df.copy()
        for col in df2.columns:
            dtype = str(df2[col].dtype)
            if "datetime" in dtype or "Timestamp" in dtype:
                df2[col] = df2[col].astype(str)
            elif "object" in dtype:
                df2[col] = df2[col].apply(
                    lambda x: str(x) if not isinstance(
                        x, (str, int, float, bool, type(None))) else x)
        df2 = df2.where(df2.notna(), other=None)
        raw = df2.to_json(orient="records", force_ascii=False, default_handler=str)
        import gzip as _gz, base64 as _b64mod
        compressed = _gz.compress(raw.encode("utf-8"), compresslevel=9)
        b64 = _b64mod.b64encode(compressed).decode("ascii")
        # Si trop gros, enlever les colonnes lourdes
        if len(b64) > 4_000_000:
            heavy = ["alerte","detail_caisse","meilleur_plan_variete",
                     "_s_rend","_s_int","_s_prise","_s_roi"]
            df3 = df2.drop(columns=[c for c in heavy if c in df2.columns], errors="ignore")
            raw2 = df3.to_json(orient="records", force_ascii=False, default_handler=str)
            b64 = _b64mod.b64encode(_gz.compress(raw2.encode(), compresslevel=9)).decode("ascii")
        return b64
    except Exception as _e:
        return None


def _b64_to_df(b64_str):
    """Désérialise un DataFrame depuis base64 gzip."""
    if not b64_str: return None
    try:
        import gzip as _gz, base64 as _b64mod
        compressed = _b64mod.b64decode(b64_str.encode("ascii"))
        raw = _gz.decompress(compressed).decode("utf-8")
        import json as _js
        records = _js.loads(raw)
        if not records: return None
        return pd.DataFrame(records)
    except Exception:
        return None


def save_session_to_supabase(sb, user_name, session_data):
    """Sauvegarde la session dans Supabase (clé partagée SHARED_2026)."""
    if sb is None:
        return False, "Supabase non disponible"
    try:
        SHARED_KEY = "SHARED_2026"
        payload = {
            "user_name":    SHARED_KEY,
            "saved_by":     str(user_name),
            "merged":       _df_to_b64(session_data.get("merged")),
            "bourak":       _df_to_b64(session_data.get("bourak")),
            "royal":        _df_to_b64(session_data.get("royal")),
            "sotusfa_raw":  _df_to_b64(session_data.get("sotusfa_raw")),
            "sotusfa_pivot":_df_to_b64(session_data.get("sotusfa_pivot")),
            "quantite":     _df_to_b64(session_data.get("quantite")),
            "prev_mai":     _df_to_b64(session_data.get("prev_mai")),
            "params":       __import__("json").dumps(
                session_data.get("params", {}), default=str),
            "saved_at":     pd.Timestamp.now().isoformat(),
        }
        sb.table("agroeco_session").delete().eq("user_name", SHARED_KEY).execute()
        sb.table("agroeco_session").insert(payload).execute()
        return True, ""
    except Exception as e:
        return False, f"{type(e).__name__}: {e}"


@st.cache_data(ttl=60, show_spinner=False)
def load_session_from_supabase(_sb, user_name="SHARED_2026"):
    """Charge la session partagée depuis Supabase."""
    if sb is None: return None
    try:
        SHARED_KEY = "SHARED_2026"
        rows = (sb.table("agroeco_session")
                  .select("*")
                  .eq("user_name", SHARED_KEY)
                  .order("saved_at", desc=True)
                  .limit(1)
                  .execute().data)
        if not rows: return None
        row = rows[0]
        return {
            "merged":       _b64_to_df(row.get("merged")),
            "bourak":       _b64_to_df(row.get("bourak")),
            "royal":        _b64_to_df(row.get("royal")),
            "sotusfa_raw":  _b64_to_df(row.get("sotusfa_raw")),
            "sotusfa_pivot":_b64_to_df(row.get("sotusfa_pivot")),
            "quantite":     _b64_to_df(row.get("quantite")),
            "prev_mai":     _b64_to_df(row.get("prev_mai")),
            "params":       __import__("json").loads(row.get("params") or "{}"),
            "saved_at":     row.get("saved_at", ""),
        }
    except Exception:
        return None


def _auto_save(sb, user_name):
    """Sauvegarde automatique silencieuse de tous les fichiers en session."""
    try:
        if sb is None: return
        save_session_to_supabase(sb, user_name or "directeur", {
            "merged":       st.session_state.get("abo_merged"),
            "bourak":       st.session_state.get("abo_bourak"),
            "royal":        st.session_state.get("abo_royal"),
            "sotusfa_raw":  st.session_state.get("abo_sotusfa_raw"),
            "sotusfa_pivot":st.session_state.get("abo_sotusfa_pivot"),
            "quantite":     st.session_state.get("abo_quantite"),
            "prev_mai":     st.session_state.get("abo_prev_mai"),
            "params":       st.session_state.get("abo_params", {}),
        })
    except Exception:
        pass  # Silencieux — ne pas bloquer l'UI
def render_agroeco_tab(sb=None, CURRENT_ROLE="directeur", CURRENT_NAME=""):

    st.markdown("""
<div style='background:#0a1a0a;border:1px solid #1E8449;border-radius:12px;
padding:16px 20px;margin-bottom:18px'>
  <div style='font-size:1.05rem;font-weight:700;color:#f0f6fc;margin-bottom:6px'>
    📊 Dashboard Agroéconomique — Tomate 2026
  </div>
  <div style='font-size:.82rem;color:#8b949e;line-height:1.8'>
    Clé de jointure : <b style='color:#FFD700'>client + centre</b> (obligatoires dans tous les fichiers) &nbsp;·&nbsp;
    Caisses vides : <b style='color:#FF9800'>date début RÉCOLTE</b> (fichier rectifié Supabase) &nbsp;·&nbsp;
    <b style='color:#ef5350'>Tonnage recouvrement</b> = (Charges + Consignes + MO 50DT/T) ÷ Prix vente
  </div>
</div>""", unsafe_allow_html=True)

    # Session state
    KEYS = ["abo_bourak","abo_royal","abo_sotusfa_raw","abo_sotusfa_pivot",
            "abo_quantite","abo_prev_dec","abo_prev_mai","abo_prev_juin",
            "abo_dates_recolte","abo_merged","abo_params","abo_errors",
            "abo_session_loaded"]
    for k in KEYS:
        if k not in st.session_state:
            st.session_state[k] = None

    # ── AUTO-RESTAURATION depuis Supabase ─────────────────────
    _is_admin_role = CURRENT_ROLE.lower() in ("directeur","admin")
    # Non-admins : toujours essayer de charger depuis Supabase (pas de cache)
    # Sauf si un recalcul forcé est en cours (_skip_supabase_restore=True)
    _skip_restore = st.session_state.get("_skip_supabase_restore", False)
    _should_restore = (
        st.session_state.get("abo_merged") is None and
        sb is not None and
        not _skip_restore and
        (not st.session_state.get("abo_session_loaded") or not _is_admin_role)
    )
    if _should_restore:
        st.session_state["abo_session_loaded"] = True
        with st.spinner("🔄 Chargement des données..."):
            _saved = load_session_from_supabase(sb)
        _has_any = bool(_saved and (
            _saved.get("merged") is not None or
            _saved.get("bourak") is not None or
            _saved.get("quantite") is not None))
        if _has_any:
            for _k, _sk in [
                ("abo_merged","merged"),("abo_bourak","bourak"),
                ("abo_royal","royal"),("abo_sotusfa_raw","sotusfa_raw"),
                ("abo_sotusfa_pivot","sotusfa_pivot"),("abo_quantite","quantite"),
                ("abo_prev_mai","prev_mai")]:
                if _saved.get(_sk) is not None:
                    st.session_state[_k] = _saved[_sk]
            if _saved.get("params"):
                st.session_state["abo_params"] = _saved["params"]
            _ts = str(_saved.get("saved_at",""))[:16].replace("T"," ")
            st.toast(f"✅ Session restaurée ({_ts})", icon="🔄")

    t0,t1,t2,t3,t4,t5,t6,t7,t8 = st.tabs([
        "⚙️ Paramètres & Import",
        "📋 Par Agriculteur",
        "👤 Par Ingénieur / Centre",
        "🗺️ Par Région",
        "🍅 Par Variété",
        "💊 Par Famille Intrant",
        "📈 Prévisions vs Réalisé",
        "🏆 Analyse Efficacité Pro",
        "🚛 Plan Récolte & Transport",
    ])

    # ══ TAB 0 — PARAMÈTRES ET IMPORT ══════════════════════
    with t0:
        # ── Contrôle d'accès : upload réservé au directeur ───
        _is_admin = (CURRENT_ROLE in ("directeur", "admin"))

        if not _is_admin:
            # Utilisateur non-admin : lecture seule, données depuis session partagée
            _role_label = {
                "commercial": "Commercial",
                "centre":     "Centre",
                "usine":      "Usine",
            }.get(CURRENT_ROLE.lower(), CURRENT_ROLE)

            _merged_ok = st.session_state.get("abo_merged") is not None
            if _merged_ok:
                _df_tmp = st.session_state["abo_merged"]
                _n_tot  = len(_df_tmp) if _df_tmp is not None else 0
                _ck_tmp = (_name_u.replace("CENTRE","").strip() or _name_u) if _role_l=="centre" else _name_u
                _n_filt = len(_df_tmp[_df_tmp["centre"].astype(str).str.upper().str.contains(_ck_tmp,na=False,regex=False)]) if (_role_l=="centre" and "centre" in _df_tmp.columns) else _n_tot
                st.success(f"✅ **{_role_label} : {CURRENT_NAME}** — {_n_filt} agriculteurs affichés (/{_n_tot} total)")
                col_r1, col_r2 = st.columns([3,1])
                col_r1.info(f"🔒 Upload réservé à l'administrateur · Filtre : **{_ck_tmp if _role_l in ('centre','commercial') else 'aucun (tout visible)'}**")
                if col_r2.button("🔄 Rafraîchir"):
                    if sb:
                        _sd = load_session_from_supabase(sb, "SHARED_2026")
                        if _sd and _sd.get("merged") is not None:
                            st.session_state["abo_merged"]        = _sd["merged"]
                            st.session_state["abo_bourak"]        = _sd.get("bourak")
                            st.session_state["abo_royal"]         = _sd.get("royal")
                            st.session_state["abo_sotusfa_raw"]   = _sd.get("sotusfa_raw")
                            st.session_state["abo_sotusfa_pivot"] = _sd.get("sotusfa_pivot")
                            st.session_state["abo_quantite"]      = _sd.get("quantite")
                            st.session_state["abo_prev_mai"]      = _sd.get("prev_mai")
                            st.cache_data.clear()
                            st.rerun()
                        else:
                            st.warning("Aucune session partagée trouvée.")
            else:
                st.warning("⏳ En attente des données — demandez à l'administrateur d'importer et sauvegarder les fichiers.")
                if st.button("🔄 Essayer de charger depuis Supabase"):
                    if sb:
                        _sd = load_session_from_supabase(sb, "SHARED_2026")
                        if _sd and _sd.get("merged") is not None:
                            st.session_state["abo_merged"]  = _sd["merged"]
                            for _k,_sk in [("abo_bourak","bourak"),("abo_royal","royal"),
                                           ("abo_sotusfa_raw","sotusfa_raw"),
                                           ("abo_sotusfa_pivot","sotusfa_pivot"),
                                           ("abo_quantite","quantite"),
                                           ("abo_prev_mai","prev_mai")]:
                                if _sd.get(_sk) is not None:
                                    st.session_state[_k] = _sd[_sk]
                            st.rerun()
                        else:
                            st.error("Aucune donnée disponible dans Supabase.")
            # ← PAS de return : on continue pour afficher les tabs filtrés

        if _is_admin:
            # ── Bouton recalcul forcé si données obsolètes ──────────
            # Bouton NUCLEAIRE : efface tout et repart de zéro
            _c1, _c2 = st.columns([1,2])
            with _c1:
                if st.button("🗑️ **RESET TOTAL**",
                             type="secondary", key="btn_reset_total",
                             help="Efface toutes les données Supabase + cache local + force recalcul"):
                    # 1. Supprimer la session sauvegardée dans Supabase
                    if sb:
                        try:
                            sb.table("shared_sessions").delete().eq(
                                "user_name","SHARED_2026").execute()
                        except Exception:
                            pass
                    # 2. Effacer tout le cache local
                    for _k in ["abo_merged","abo_session_loaded","abo_data_version",
                               "abo_bourak","abo_royal","abo_sotusfa_raw",
                               "abo_sotusfa_pivot","abo_quantite","abo_prev_mai",
                               "abo_params"]:
                        st.session_state.pop(_k, None)
                    # 3. Poser le flag anti-restore pour éviter rechargement Supabase
                    st.session_state["_skip_supabase_restore"] = True
                    st.cache_data.clear()
                    st.info("✅ Cache effacé. Déposez vos fichiers et cliquez **Fusionner**.")
                    st.rerun()

            # ── Paramètres (admin seulement) ─────────────────────
            st.markdown("### ⚙️ Paramètres de calcul")
            with st.form("params_form"):
                pc1,pc2,pc3 = st.columns(3)
                with pc1:
                    st.markdown("**💰 Prix vente global (DT/tonne)**")
                    st.caption("Utilisé si absent du tableau quantité")
                    prix_global = st.number_input("Prix vente DT/T",0.0,1000.0,240.0,10.0,key="px_g")
                with pc2:
                    st.markdown("**🔲 Consigne plateaux (DT/plateau)**")
                    p228pvc  = st.number_input("Pltx 228 PVC", 0.0,50.0,2.5,0.1,key="p1")
                    p228poly = st.number_input("Pltx 228 POLY",0.0,50.0,2.0,0.1,key="p2")
                    p160pvc  = st.number_input("Pltx 160 PVC", 0.0,50.0,2.0,0.1,key="p3")
                    p160poly = st.number_input("Pltx 160 POLY",0.0,50.0,1.8,0.1,key="p4")
                with pc3:
                    st.markdown("**📦 Caisses vides — MO récolte**")
                    mo_tonne = st.number_input("MO récolte (DT/T)",0.0,200.0,50.0,5.0,key="mo")
                    st.caption("Condition caisses : date début RÉCOLTE < 10 juil. → 1ère affectation")
                st.form_submit_button("✅ Appliquer les paramètres", use_container_width=True)

            # ── Caisses vides PAR USINE ───────────────────────────
            st.markdown("---")
            st.markdown("#### 📦 Caisses vides — Paramètres par usine")
            st.caption("1ère affectation (début récolte < 10 juillet) = caisses facturées | 2ème = 0 DT")

            caisses_par_usine = {}
            _saved_caisses = (st.session_state.get("abo_params") or {}).get("caisses_par_usine", {})
            usine_cols = st.columns(5)
            usine_names = ["SICAM","TUCAL","COMOCAP","ABIDA","ELFALLEH"]
            usine_colors = {"SICAM":"#F5A623","TUCAL":"#8B5CF6","COMOCAP":"#3B82F6",
                            "ABIDA":"#FF6B9D","ELFALLEH":"#00E5A0"}

            for ci2, usine in enumerate(usine_names):
                dft = CAISSES_USINE_DEFAULTS.get(usine, {"nb_ha":80,"prix":3.0,"type":"Caisse 25kg","cap_kg":25})
                saved_u = _saved_caisses.get(usine, dft)
                uc = usine_colors.get(usine,"#888")
                with usine_cols[ci2]:
                    st.markdown(f"<div style='background:#1a2332;border-radius:8px;padding:8px;"
                                f"border-top:3px solid {uc};margin-bottom:4px'>"
                                f"<b style='color:{uc};font-size:12px'>{usine}</b><br>"
                                f"<span style='font-size:10px;color:#aaa'>{dft['type']}</span>"
                                f"</div>", unsafe_allow_html=True)
                    nb_ha = st.number_input(f"Nb caisses/ha",
                        min_value=0.0, max_value=300.0,
                        value=float(saved_u.get("nb_ha", dft["nb_ha"])),
                        step=5.0, key=f"nb_c_{usine}")
                    prix_c = st.number_input(f"Prix/caisse (DT)",
                        min_value=0.0, max_value=20.0,
                        value=float(saved_u.get("prix", dft["prix"])),
                        step=0.25, key=f"px_c_{usine}")
                    cout_ha = round(nb_ha * prix_c, 2)
                    st.caption(f"→ **{cout_ha:.1f} DT/ha** (1ère affectation)")
                    caisses_par_usine[usine] = {"nb_ha": nb_ha, "prix": prix_c,
                                                "type": dft["type"], "cap_kg": dft["cap_kg"]}

            params = {
                "prix_vente_global": prix_global,
                "prix_consigne": {
                    "Pltx 228 PVC":p228pvc,"Pltx 228 POLY":p228poly,
                    "Pltx 160 PVC":p160pvc,"Pltx 160 POLY":p160poly,
                },
                "caisses_par_usine": caisses_par_usine,
                # Rétrocompat : valeurs globales = moyenne pondérée SICAM (usine principale)
                "prix_caisse":   caisses_par_usine.get("SICAM",{}).get("prix", 3.0),
                "nb_caisses_ha": caisses_par_usine.get("SICAM",{}).get("nb_ha", 80.0),
                "mo_tonne":      mo_tonne,
            }
            st.session_state["abo_params"] = params

            st.divider()
            st.markdown("### 📥 Import fichiers")
            st.markdown("""<div style='background:#161b22;border:1px solid #FFD700;
    border-radius:8px;padding:10px 14px;margin-bottom:14px;font-size:.85rem'>
    ⭐ <b style='color:#FFD700'>Colonnes OBLIGATOIRES dans tous les fichiers :</b>
    &nbsp;<code>centre</code> &nbsp;+&nbsp; <code>client</code>
    </div>""", unsafe_allow_html=True)

            fi1,fi2 = st.columns(2)
            fi3,fi4 = st.columns(2)

            def _upload_block(col, icon, name, color, desc, key, parse_fn, extra_args=()):
                with col:
                    st.markdown(f"""<div style='background:#111;border:1px solid #{color};
    border-radius:8px;padding:10px 14px;margin-bottom:8px'>
    <b style='color:#{color}'>{icon} {name}</b><br>
    <span style='font-size:.78rem;color:#aaa'>{desc}</span></div>""",
                        unsafe_allow_html=True)
                    f = st.file_uploader(name,type=["xlsx","xls"],
                                         key=key,label_visibility="collapsed")
                    if f:
                        try:
                            result = parse_fn(f, *extra_args)
                            if isinstance(result, tuple):
                                if len(result) == 2:
                                    df_res, msg = result
                                else:
                                    df_res, pivot, msg = result
                            else:
                                df_res, msg = result, ""
                            if msg:
                                st.error(msg)
                                return None
                            return f, df_res, pivot if len(result)==3 else None
                        except Exception as e:
                            st.error(f"Erreur {name}: {e}")
                    return None

            # BOURAK
            with fi1:
                st.markdown(f"""<div style='background:#111;border:1px solid #FF9800;
    border-radius:8px;padding:10px 14px;margin-bottom:8px'>
    <b style='color:#FF9800'>🚛 BOURAK</b> — Financement<br>
    <span style='font-size:.78rem;color:#aaa'>Obligatoire : <b>client · centre</b><br>
    Attendu : responsable · ingenieur · region · hectares · avance · report</span></div>""",
                    unsafe_allow_html=True)
                f_b = st.file_uploader("Bourak",type=["xlsx","xls"],
                                        key="up_b",label_visibility="collapsed")
                if f_b:
                    res = parse_bourak(f_b)
                    if isinstance(res, tuple): df_b, msg = res
                    else: df_b, msg = res, ""
                    if msg: st.error(msg)
                    else:
                        st.session_state["abo_bourak"] = df_b
                        tot_av = df_b["avance"].sum() if "avance" in df_b.columns else 0
                        st.success(f"✅ {len(df_b)} lignes · {tot_av:,.0f} DT avances")

            # ROYAL
            with fi2:
                st.markdown(f"""<div style='background:#111;border:1px solid #9C27B0;
    border-radius:8px;padding:10px 14px;margin-bottom:8px'>
    <b style='color:#9C27B0'>🌱 ROYAL</b> — Plants<br>
    <span style='font-size:.78rem;color:#aaa'>Obligatoire : <b>client · centre</b><br>
    Attendu : zone · variete · qte_livree · valeur · <b>date_debut_livraison</b> · date_fin</span></div>""",
                    unsafe_allow_html=True)
                f_r = st.file_uploader("Royal",type=["xlsx","xls"],
                                        key="up_r",label_visibility="collapsed")
                if f_r:
                    df_r, msg = parse_royal(f_r)
                    if msg: st.error(msg)
                    else:
                        st.session_state["abo_royal"] = df_r
                        st.success(f"✅ {len(df_r)} lignes")

            # SOTUSFA
            with fi3:
                st.markdown(f"""<div style='background:#111;border:1px solid #4CAF50;
    border-radius:8px;padding:10px 14px;margin-bottom:8px'>
    <b style='color:#4CAF50'>🌿 SOTUSFA</b> — Intrants<br>
    <span style='font-size:.78rem;color:#aaa'>Obligatoire : <b>client · centre</b><br>
    Attendu : famille · article · qte · valeur (DAP / fumure / fumier / pest.)</span></div>""",
                    unsafe_allow_html=True)
                f_s = st.file_uploader("Sotusfa",type=["xlsx","xls"],
                                        key="up_s",label_visibility="collapsed")
                if f_s:
                    df_s_raw, df_s_piv, msg = parse_sotusfa(f_s)
                    if msg: st.error(msg)
                    else:
                        st.session_state["abo_sotusfa_raw"]   = df_s_raw
                        st.session_state["abo_sotusfa_pivot"] = df_s_piv
                        tot = df_s_raw["valeur"].sum() if "valeur" in df_s_raw.columns else 0
                        st.success(f"✅ {len(df_s_raw)} lignes · {tot:,.0f} DT")

            # QUANTITÉ
            with fi4:
                st.markdown(f"""<div style='background:#111;border:1px solid #2196F3;
    border-radius:8px;padding:10px 14px;margin-bottom:8px'>
    <b style='color:#2196F3'>📊 QUANTITÉ</b> — Actif/Extra<br>
    <span style='font-size:.78rem;color:#aaa'>Obligatoire : <b>client · centre</b><br>
    Attendu : qte_livree · qte_actif · qte_extra · tonnage_livre · prix_vente</span></div>""",
                    unsafe_allow_html=True)
                f_q = st.file_uploader("Quantité",type=["xlsx","xls"],
                                        key="up_q",label_visibility="collapsed")
                if f_q:
                    df_q, msg = parse_quantite(f_q)
                    if msg: st.error(msg)
                    else:
                        st.session_state["abo_quantite"] = df_q
                        st.success(f"✅ {len(df_q)} agriculteurs")

            # ── Prévisions ──────────────────────────────────────
            st.divider()
            st.markdown("### 📅 Prévisions tonnage")
            pv1,pv2,pv3 = st.columns(3)
            with pv1:
                f_dec = st.file_uploader("📋 Prévision Décembre",
                                          type=["xlsx","xls"],key="up_dec")
                if f_dec:
                    df_d, msg = parse_prevision(f_dec, "prevision_dec")
                    if msg:
                        st.warning(f"⚠️ Déc (non bloquant): {msg}")
                        # Essayer quand même avec ce qu'on a
                        if df_d is not None and not df_d.empty:
                            st.session_state["abo_prev_dec"] = df_d
                            st.success(f"✅ Déc chargé malgré avertissement: {df_d['prevision_dec'].sum():,.0f} T")
                    elif df_d is not None and not df_d.empty:
                        st.session_state["abo_prev_dec"] = df_d
                        st.success(f"✅ Déc: {df_d['prevision_dec'].sum():,.0f} T")
            with pv2:
                f_mai = st.file_uploader("📋 Prévision Mai",
                                          type=["xlsx","xls"],key="up_mai")
                if f_mai:
                    df_m, msg = parse_prevision(f_mai, "prevision_mai")
                    if msg:
                        st.warning(f"⚠️ Mai (non bloquant): {msg}")
                        if df_m is not None and not df_m.empty:
                            st.session_state["abo_prev_mai"] = df_m
                            st.success(f"✅ Mai chargé malgré avertissement: {df_m['prevision_mai'].sum():,.0f} T")
                    elif df_m is not None and not df_m.empty:
                        st.session_state["abo_prev_mai"] = df_m
                        st.success(f"✅ Mai: {df_m['prevision_mai'].sum():,.0f} T")
            with pv3:
                st.markdown("**☁️ Juin — Supabase (fichier rectifié)**")
                c1,c2 = st.columns(2)
                with c1:
                    if st.button("🔄 Charger Juin", use_container_width=True):
                        df_j = load_prevision_juin(sb)
                        if not df_j.empty:
                            st.session_state["abo_prev_juin"] = df_j
                            st.success(f"✅ {df_j['prevision_juin'].sum():,.0f} T")
                        else:
                            st.warning("Aucune donnée Juin")
                with c2:
                    if st.button("🔄 Dates récolte", use_container_width=True):
                        df_dr = load_date_debut_recolte(sb)
                        if not df_dr.empty:
                            st.session_state["abo_dates_recolte"] = df_dr
                            n1 = (df_dr["affectation_caisse"].str.startswith("1ère")).sum()
                            st.success(f"✅ {n1} agriculteurs 1ère affectation")
                        else:
                            st.warning("Aucune date récolte")

            # Statut caisses vides
            dr = st.session_state.get("abo_dates_recolte")
            if dr is not None and not dr.empty:
                n1 = (dr["affectation_caisse"].str.startswith("1ère")).sum()
                n2 = len(dr) - n1
                st.info(f"📦 Caisses vides — **1ère affectation** (< 10 juil.) : {n1} agriculteurs · "
                        f"**2ème** (≥ 10 juil.) : {n2} agriculteurs")
            else:
                st.warning("⚠️ Dates de récolte non chargées → tous les agriculteurs "
                           "seront mis en 2ème affectation (sans caisses vides). "
                           "Cliquez '🔄 Dates récolte' ci-dessus.")

            # ── Fusionner ────────────────────────────────────────
            st.divider()
            if st.button("🔗 Fusionner et calculer",
                         type="primary",use_container_width=True):
                with st.spinner("Calcul en cours…"):
                    df_merged = merge_and_calculate(
                        st.session_state.get("abo_bourak"),
                        st.session_state.get("abo_royal"),
                        st.session_state.get("abo_sotusfa_raw"),
                        st.session_state.get("abo_sotusfa_pivot"),
                        st.session_state.get("abo_quantite"),
                        st.session_state.get("abo_prev_dec"),
                        st.session_state.get("abo_prev_mai"),
                        st.session_state.get("abo_prev_juin"),
                        st.session_state.get("abo_dates_recolte"),
                        st.session_state["abo_params"],
                    )
                if df_merged.empty:
                    st.error("❌ Aucune donnée fusionnée — vérifiez les fichiers.")
                else:
                    st.session_state["abo_merged"] = df_merged
                    n_r = (df_merged["alerte"].str.contains("🔴")).sum()
                    n_y = (df_merged["alerte"].str.contains("🟡")).sum()
                    n_g = (df_merged["alerte"].str.contains("🟢")).sum()

                    # ── AUTO-SAVE dans Supabase (session partagée) ──
                    _save_ok = False
                    if sb is not None:
                        try:
                            _save_ok, _save_err = save_session_to_supabase(
                                sb, CURRENT_NAME or "directeur", {
                                    "merged":       df_merged,
                                    "bourak":       st.session_state.get("abo_bourak"),
                                    "royal":        st.session_state.get("abo_royal"),
                                    "sotusfa_raw":  st.session_state.get("abo_sotusfa_raw"),
                                    "sotusfa_pivot":st.session_state.get("abo_sotusfa_pivot"),
                                    "quantite":     st.session_state.get("abo_quantite"),
                                    "prev_mai":     st.session_state.get("abo_prev_mai"),
                                    "params":       st.session_state.get("abo_params", {}),
                                })
                        except Exception as _se:
                            _save_ok = False; _save_err = str(_se)
                    _save_icon = "💾 sauvegardé auto" if _save_ok else "⚠️ non sauvegardé"

                    st.success(
                        f"✅ {len(df_merged)} agriculteurs · "
                        f"🔴 {n_r} critiques · 🟡 {n_y} attention · 🟢 {n_g} OK · {_save_icon}")

                    if not _save_ok and sb is not None:
                        st.warning(f"⚠️ Sauvegarde échouée : **{_save_err}**")

                    xl = export_excel(df_merged, st.session_state.get("abo_sotusfa_raw"))
                    st.download_button(
                        "📥 Télécharger Excel complet (4 feuilles)",
                        data=xl,
                        file_name="dashboard_agroeco_2026.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True)


    # ── Données fusionnées ─────────────────────────────────
    # ── df filtré selon le rôle de l'utilisateur ────────────────
    _df_all = st.session_state.get("abo_merged")
    _role_l = CURRENT_ROLE.lower() if CURRENT_ROLE else "directeur"
    _name_u = str(CURRENT_NAME).strip().upper()

    if _df_all is not None and not (hasattr(_df_all,"empty") and _df_all.empty):
        df = _df_all.copy()
        if _role_l == "commercial":
            if "commercial" in df.columns:
                df = df[df["commercial"].astype(str).str.upper()
                        .str.contains(_name_u, na=False, regex=False)]
        elif _role_l == "centre":
            # kerkouane → "KERKOUANE", baccara → "BACCARA", centre428 → "428"
            _ck = (_name_u.replace("CENTRE","").strip() or _name_u)
            if "centre" in df.columns:
                df = df[df["centre"].astype(str).str.upper()
                        .str.contains(_ck, na=False, regex=False)]
        # directeur / admin / usine → voient tout
    else:
        df = _df_all  # None ou vide


    # ══ POST-PROCESSING : Appliquer _INTRANTS_2026 et _PREVISION_2026 ══
    # Cette étape garantit les données réelles même avec un cache Supabase ancien
    import re as _re_pp, unicodedata as _uc_pp
    def _cn_pp(n):
        n = str(n).strip().upper()
        n = _re_pp.sub(r"[(][^)]*[)]","",n)
        n = "".join(c for c in _uc_pp.normalize("NFD",n) if _uc_pp.category(c)!="Mn")
        n = _re_pp.sub(r"[^A-Z0-9 ]"," ",n)
        return _re_pp.sub(r"[ ]+"," ",n).strip()

    def _sco_pp(a,b):
        wa,wb=set(a.split()),set(b.split())
        if not wa or not wb: return 0.0
        inter=len(wa&wb); union=len(wa|wb); sj=inter/union if union else 0
        sh,lo=(wa,wb) if len(wa)<=len(wb) else (wb,wa)
        sc=sum(1 for w in sh if any(lw.startswith(w[:4]) for lw in lo) and len(w)>2)/max(len(sh),1)*0.85
        return max(sj,sc)

    if df is not None and not (hasattr(df,"empty") and df.empty):
        _int_keys = list(_INTRANTS_2026.keys())
        _prv_keys = list(_PREVISION_2026.keys())
        _agri_col = next((c for c in ["agriculteur","client"] if c in df.columns), None)

        if _agri_col:
            def _get_intrant(nom):
                ck = _cn_pp(str(nom))
                if ck in _INTRANTS_2026: return _INTRANTS_2026[ck]
                best = max(_int_keys, key=lambda k: _sco_pp(ck,k), default=None)
                return _INTRANTS_2026[best] if best and _sco_pp(ck,best)>=0.65 else None

            def _get_prev(nom, key):
                ck = _cn_pp(str(nom))
                if ck in _PREVISION_2026: return _PREVISION_2026[ck].get(key,"")
                best = max(_prv_keys, key=lambda k: _sco_pp(ck,k), default=None)
                return _PREVISION_2026[best].get(key,"") if best and _sco_pp(ck,best)>=0.65 else ""

            # ── Intrants : calcul depuis _INTRANTS_2026 ──────────────
            _int_series = df[_agri_col].apply(_get_intrant)
            _mask_int   = _int_series.notna()
            # Mettre à jour TOUTES les colonnes intrants (interne + affichage)
            if "charge_intrants" not in df.columns:
                df["charge_intrants"] = 0.0
            df.loc[_mask_int, "charge_intrants"] = _int_series[_mask_int].astype(float)
            # Mettre à jour la colonne affichée "Intrants (DT)" directement
            _int_display_col = next((c for c in df.columns
                                    if c.strip().lower() in ["intrants (dt)","intrants(dt)"]), None)
            if _int_display_col:
                df.loc[_mask_int, _int_display_col] = _int_series[_mask_int].astype(float)

            # ── Recalculer Charge Totale avec intrants réels ──────────
            _cp_col  = next((c for c in df.columns if c.strip().lower() in ["plants (dt)","charge_plants","plants(dt)"]), None)
            _plants  = pd.to_numeric(df[_cp_col], errors="coerce").fillna(0) if _cp_col else pd.Series([0]*len(df), index=df.index, dtype=float)
            # Lire les intrants depuis charge_intrants (DÉJÀ mis à jour par _INTRANTS_2026)
            _intrant = pd.to_numeric(df["charge_intrants"], errors="coerce").fillna(0)
            _av_col  = next((c for c in df.columns if c.strip().lower() in ["avance bourak (dt)","avance_bourak","avance bourak"]), None)
            _avance  = pd.to_numeric(df[_av_col], errors="coerce").fillna(0) if _av_col else pd.Series([0]*len(df), index=df.index, dtype=float)
            df["charge_totale"]  = (_plants + _intrant + _avance).round(0)
            # Mettre à jour la colonne affichée
            _ct_col = next((c for c in df.columns if c.strip().lower() in ["charge totale (dt)","charge_totale"]), None)
            if _ct_col: df[_ct_col] = df["charge_totale"]

            # ── Prix vente (défaut 270) ───────────────────────────────
            _pv_col = next((c for c in df.columns if c.strip().lower() in ["prix vente","prix_vente"]), None)
            _pv = pd.to_numeric(df[_pv_col], errors="coerce").fillna(270) if _pv_col else pd.Series([270.0]*len(df), index=df.index, dtype=float)
            _pv = _pv.where(_pv>0, 270)

            # ── Recalculer Charges totales (avec consignes et MO) ────
            # Consigne Plateau = 0 (données non encore saisies — décalage plateaux pris/retournés)
            _cons_plt = pd.Series([0]*len(df), index=df.index, dtype=float)
            if "consigne_plateau" in df.columns:
                _cons_plt_raw = pd.to_numeric(df["consigne_plateau"], errors="coerce").fillna(0)
                # Ne garder que si la valeur vient d'une vraie saisie (pas calculée)
                # Pour l'instant = 0 jusqu'à réception des données
                _cons_plt = pd.Series([0]*len(df), index=df.index, dtype=float)
            _cc_col  = next((c for c in df.columns if c.strip().lower() in ["consigne caisse","consigne_caisse"]), None)
            _mo_col  = next((c for c in df.columns if c.strip().lower() in ["mo récolte (dt)","mo recolte (dt)","mo_recolte"]), None)
            _rp_col  = next((c for c in df.columns if c.strip().lower() in ["report (dt)","report_dt","report"]), None)
            _cons_c  = pd.to_numeric(df[_cc_col], errors="coerce").fillna(0) if _cc_col else pd.Series([0]*len(df), index=df.index, dtype=float)
            _mo      = pd.to_numeric(df[_mo_col], errors="coerce").fillna(0) if _mo_col else pd.Series([0]*len(df), index=df.index, dtype=float)
            _rep     = pd.to_numeric(df[_rp_col], errors="coerce").fillna(0) if _rp_col else pd.Series([0]*len(df), index=df.index, dtype=float)
            # Charges totales RÉELLES = charge_totale + consigne_caisse + MO
            # NOTE: consigne_plateau = 0 (données non fournies)
            _charges_tot = df["charge_totale"] + _cons_c + _mo
            df["charges_totales"]     = _charges_tot.round(0)
            df["charge_a_recouvrir"]  = (_charges_tot + _rep).round(0)
            df["tonnage_recouvrement"]= (_charges_tot / _pv.where(_pv>0,1)).round(2)

            # ── Recalculer Solde, Valeur, Écart ──────────────────────
            _ton_col = next((c for c in df.columns
                            if c.strip().lower() in ["livré (t)","livre_t","tonnage_livre",
                                                      "livré t","livret","livré(t)"]), None)
            _ton_livr = pd.to_numeric(df[_ton_col], errors="coerce").fillna(0) if _ton_col else pd.Series([0]*len(df), index=df.index, dtype=float)
            df["valeur_livree"] = (_ton_livr * _pv).round(0)
            df["ecart_tonnage"] = (_ton_livr - df["tonnage_recouvrement"]).round(2)
            df["solde_final"]   = (df["valeur_livree"] - _charges_tot - _rep).round(0)
            # Synchroniser toutes les colonnes affichées avec les valeurs calculées
            _col_sync = {
                "Charge Totale (DT)":     "charge_totale",
                "Intrants (DT)":          "charge_intrants",
                "Charges à recouvrir":    "charge_a_recouvrir",
                "RECOUVREMENT (T)":       "tonnage_recouvrement",
                "Recouv./ha":             "recouvrement_ha",
                "Valeur Livrée":          "valeur_livree",
                "Solde Final":            "solde_final",
                "Écart (T)":             "ecart_tonnage",
                "Coût/ha":                "cout_ha",
                "Coût/plant":             "cout_plant",
                "Densité/ha":             "densite_ha",
                "T/ha réalisé":           "rendement_ha_reel",
                "Plants Livrés":          "_plants_display",
                "Déb. Récolte":           "date_debut_recolte",
            }
            # Mettre à jour Plants Livrés affiché depuis plt_livres réels
            if "_pl_from_plt" in dir():
                df["_plants_display"] = _pl_from_plt.round(0)
            else:
                df["_plants_display"] = pd.Series([0]*len(df), index=df.index, dtype=float)
            # Prév. Mai = 0 (données mensuelles non disponibles)
            _pm_col = next((c for c in df.columns
                           if c.strip().lower() in ["prév. mai (t)","prev mai","prevision_mai"]), None)
            if _pm_col:
                df[_pm_col] = 0.0

            for _disp, _calc in _col_sync.items():
                if _disp in df.columns and _calc in df.columns:
                    df[_disp] = df[_calc]

            # ── Variété : effacer TOUS les faux noms (zones géographiques)
            # Vraies variétés tomate : Heinz, Savera, Tiger, Cobra, H2274, etc.
            # Fausses : dar allouch, amaymia, cap bon, gafsa, kairouan (=zones)
            _ZONES_KNOWN = {
                "dar allouch","amaymia","sidi aich","tefeloun","diar hojjej",
                "majel belabess","oued chiba","feriana","garat sassi","ouled omran",
                "batten","menzel tamim","el bourak","cap bon 1","cap bon 2",
                "gafsa / kassrine","gafsa","kairouan","cap bon","nabeul","kassrine",
                "sidi bouzid","bouficha","hafedh mosbeh","karim garmalah","mourad mansouri",
                "sebti jabali","souhail bouzana","bilel gha"
            }
            _VRAI_VAR_KEYWORDS = ["heinz","savera","tiger","cobra","perfect","dorra",
                                  "ercole","h1015","h2274","f1","momotaro","lyterno"]
            for _col in df.columns:
                if _col.strip().lower() in ["variété","variete"]:
                    _v = df[_col].astype(str).str.strip()
                    _v = _v.replace({"nan":"","NaN":"","None":"","NaT":""})
                    _is_zone = _v.str.lower().isin(_ZONES_KNOWN)
                    _has_var = _v.str.lower().str.contains(
                        "|".join(_VRAI_VAR_KEYWORDS), regex=True, na=False)
                    _has_num = _v.str.contains(r'[0-9]', regex=True, na=False)
                    # Garder seulement si c'est une vraie variété
                    _is_real = _has_var | (_has_num & ~_is_zone)
                    df.loc[~_is_real, _col] = ""

            # ── Recalculer les ratios /ha et /plant ───────────────────
            # Recherche insensible à la casse (Ha / ha / hectares / Hectares)
            _ha_col = next((c for c in df.columns
                           if c.strip().lower() in ["ha","hectares","superficie","nbre_ha"]), None)
            _pl_col = next((c for c in df.columns
                           if c.strip().lower() in ["plants livrés","plants_livres",
                                                     "qte_livree","qte livree","plants livres"]), None)

            _ha_pp = pd.to_numeric(df[_ha_col], errors="coerce").fillna(0) if _ha_col else pd.Series([0]*len(df), index=df.index, dtype=float)
            _ha_nz = _ha_pp.where(_ha_pp > 0, float("nan"))
            _pl_pp = pd.to_numeric(df[_pl_col], errors="coerce").fillna(0) if _pl_col else pd.Series([0]*len(df), index=df.index, dtype=float)
            _pl_nz = _pl_pp.where(_pl_pp > 0, float("nan"))

            df["cout_ha"]            = (df["charge_totale"] / _ha_nz).fillna(0).round(0)
            df["cout_plant"]         = (df["charge_totale"] / _pl_nz).fillna(0).round(4)
            # Densité/ha — priorité : nb_plateaux Royal > Plt Livrés Bourak > estimation
            _nb_plt_col = next((c for c in df.columns
                               if c.strip().lower() in ["nb_plateaux","nb plateaux","nb plts"]), None)
            _plt_col    = next((c for c in df.columns
                               if c.strip().lower() in ["plt livrés","plt_livres","plt livres"]), None)
            _nb_plt_v   = pd.to_numeric(df[_nb_plt_col], errors="coerce").fillna(0) if _nb_plt_col else pd.Series([0]*len(df), index=df.index, dtype=float)
            _plt_bourak = pd.to_numeric(df[_plt_col], errors="coerce").fillna(0) if _plt_col else pd.Series([0]*len(df), index=df.index, dtype=float)
            # Choisir la meilleure source : Royal nb_plateaux > Bourak plt_livres
            _plt_final  = _nb_plt_v.where(_nb_plt_v > 0, _plt_bourak)
            _pl_from_plt = _plt_final * 228
            _pl_nz_plt   = _pl_from_plt.where(_pl_from_plt > 0, float("nan"))
            _dens_raw    = (_pl_from_plt / _ha_nz).fillna(0)
            # Valider la densité (15000-40000 plants/ha réaliste pour tomate)
            _dens_ok     = _dens_raw.where((_dens_raw >= 15000) & (_dens_raw <= 40000), 25000)
            df["densite_ha"] = _dens_ok.round(0)
            _pl_nz = _pl_nz_plt if "_pl_nz_plt" in dir() else pd.Series([float("nan")]*len(df), index=df.index)
            df["rendement_ha_reel"]  = (_ton_livr / _ha_nz).fillna(0).round(1)
            df["recouvrement_ha"]    = (df["tonnage_recouvrement"] / _ha_nz).fillna(0).round(2)

            # Aussi mettre à jour les colonnes export (noms avec majuscule)
            for _src, _dst in [("cout_ha","Coût/ha"), ("cout_plant","Coût/plant"),
                                ("densite_ha","Densité/ha"), ("rendement_ha_reel","T/ha réalisé")]:
                if _src in df.columns and _dst in df.columns:
                    df[_dst] = df[_src]

            # ── Enrichir Accessibilité / Usine / Zone / Dates depuis _PREVISION_2026
            for _pk, _col_candidates in [
                ("acces",      ["Accessibilité","accessibilite"]),
                ("usine",      ["Usine","usine"]),
                ("zone",       ["Zone","zone"]),
                ("region",     ["Région","region"]),
                ("date_debut", ["Déb. Récolte","date_debut_recolte","deb_recolte","date_debut"]),
                ("date_fin",   ["Fin Récolte","date_fin_recolte","date_fin"]),
            ]:
                # Trouver la colonne correspondante dans df (insensible à la casse)
                _found_col = next((c for c in df.columns
                                  if c.strip().lower() in [x.lower() for x in _col_candidates]), None)
                if _found_col:
                    _is_empty = df[_found_col].astype(str).str.strip().isin(["","nan","NaN","NaT"]).all()
                    if _is_empty:
                        df[_found_col] = df[_agri_col].apply(lambda x: _get_prev(x, _pk))

            # Variété : conserver TOUTES les valeurs du fichier Royal
            # (Savera, Heinz 7709, Dorra, Ercole = toutes vraies variétés)
            # Ne rien effacer — si la colonne est remplie par l'utilisateur, garder
            _var_col = next((c for c in df.columns if c.strip().lower() in ["variété","variete"]), None)
            if _var_col:
                # Nettoyer uniquement les valeurs techniques (nan, NaN, None)
                df[_var_col] = df[_var_col].astype(str).str.strip().replace(
                    {"nan":"","NaN":"","None":"","NaT":""})

    def _no_data():
        if _df_all is None or (hasattr(_df_all,"empty") and _df_all.empty):
            if _is_admin_role:
                st.info("📥 Importez les fichiers dans l'onglet **⚙️ Paramètres & Import** "
                        "puis cliquez **Fusionner**.")
            else:
                st.warning(
                    f"⏳ **Aucune donnée disponible** pour votre profil ({CURRENT_ROLE} : {CURRENT_NAME})\n\n"
                    "Le directeur doit :\n"
                    "1. Se connecter avec son compte directeur\n"
                    "2. Déposer les fichiers dans ⚙️ Paramètres & Import\n"
                    "3. Cliquer **Fusionner** → la session est sauvegardée automatiquement\n"
                    "4. Revenir ici et cliquer 🔄 Rafraîchir"
                )
        else:
            n_total = len(_df_all) if _df_all is not None else 0
            _ck_debug = (_name_u.replace("CENTRE","").strip() or _name_u)
            st.warning(
                f"📭 **Aucun agriculteur** trouvé pour : {CURRENT_ROLE} = **{CURRENT_NAME}**\n\n"
                f"Recherche dans colonne \'centre\' : **'{_ck_debug}'**\n"
                f"({n_total} agriculteurs au total dans la base)"
            )

    # ══ TAB 1 — PAR AGRICULTEUR ════════════════════════════
    with t1:
        if df is None or df.empty:
            _no_data()
        else:
            kc = st.columns(6)
            kc[0].markdown(_metric("Agriculteurs",len(df)), unsafe_allow_html=True)
            kc[1].markdown(_metric("Charge à Recouvrir",f"{df['charge_totale'].sum():,.0f} DT",color="#FF9800"),unsafe_allow_html=True)
            kc[2].markdown(_metric("Recouvrement",f"{df['tonnage_recouvrement'].sum():,.1f} T",color="#ef5350"),unsafe_allow_html=True)

            # Tonnage réalisé : depuis Quantité ou prévision
            _ton_reel = df["tonnage_livre"].fillna(0).sum() if "tonnage_livre" in df.columns else 0
            _has_quantite = st.session_state.get("abo_quantite") is not None
            _ton_label = f"{_ton_reel:,.1f} T" if _has_quantite else "Non importé"
            _ton_color = "#4CAF50" if _has_quantite and _ton_reel > 0 else "#888"
            kc[3].markdown(_metric("Livré réel", _ton_label, color=_ton_color),
                           unsafe_allow_html=True)

            # Prévision Mai : total brut du fichier (pas juste les matchés)
            _prev_mai_brut = 0
            _df_prev_mai = st.session_state.get("abo_prev_mai")
            if _df_prev_mai is not None and "prevision_mai" in _df_prev_mai.columns:
                _prev_mai_brut = _df_prev_mai["prevision_mai"].sum()
            _prev_mai_merged = df["prevision_mai"].fillna(0).sum() if "prevision_mai" in df.columns else 0

            if _prev_mai_brut > 0:
                _pct_match = round(_prev_mai_merged / _prev_mai_brut * 100) if _prev_mai_brut > 0 else 0
                kc[4].markdown(_metric("Prévision Mai",
                    f"{_prev_mai_brut:,.0f} T",
                    color="#42A5F5",
                    delta=_prev_mai_merged - _prev_mai_brut,
                    delta_label=f"T matchés ({_pct_match}%)"),
                    unsafe_allow_html=True)
            else:
                n_crit=(df["alerte"].str.contains("🔴")).sum()
                kc[4].markdown(_metric("⚠️ Critiques",n_crit,color="#ef5350"),unsafe_allow_html=True)

            kc[5].markdown(_metric("Solde global",f"{df['solde_final'].sum():+,.0f} DT",
                color="#4CAF50" if df["solde_final"].sum()>=0 else "#ef5350"),unsafe_allow_html=True)

            # Avertissement si données fictives
            if not _has_quantite:
                st.warning("⚠️ **Tonnage réalisé = 0** — Le fichier **Tableau Quantité** n'est pas encore importé. "
                           "Les calculs de recouvrement et solde sont basés sur les prévisions uniquement.")
            if _prev_mai_brut > 0 and _prev_mai_merged < _prev_mai_brut * 0.5:
                st.info(f"ℹ️ **Prévision Mai** : {_prev_mai_brut:,.0f} T dans le fichier, "
                        f"mais seulement **{_prev_mai_merged:,.0f} T matchés** ({round(_prev_mai_merged/_prev_mai_brut*100)}%) "
                        f"car certains noms d'agriculteurs diffèrent entre les fichiers. "
                        f"Le total affiché dans les tableaux = valeurs matchées uniquement.")

            fc1,fc2,fc3,fc4 = st.columns(4)
            alerte_f = fc1.selectbox("Alerte",["Toutes","🔴","🟡","🟢"],key="t1a")
            comm_f   = fc2.selectbox("Commercial",
                ["Tous"]+sorted(df["commercial"].dropna().unique().tolist())
                if "commercial" in df.columns else ["Tous"],key="t1c")
            ing_f    = fc3.selectbox("Ingénieur",
                ["Tous"]+sorted(df["ingenieur"].dropna().unique().tolist())
                if "ingenieur" in df.columns else ["Tous"],key="t1i")
            ctr_f    = fc4.selectbox("Centre",
                ["Tous"]+sorted(df["centre"].dropna().unique().tolist())
                if "centre" in df.columns else ["Tous"],key="t1ct")

            df_f = df.copy()
            if alerte_f != "Toutes": df_f = df_f[df_f["alerte"].str.contains(alerte_f)]
            if comm_f   != "Tous" and "commercial" in df_f.columns: df_f = df_f[df_f["commercial"]==comm_f]
            if ing_f    != "Tous" and "ingenieur"  in df_f.columns: df_f = df_f[df_f["ingenieur"]== ing_f]
            if ctr_f    != "Tous" and "centre"     in df_f.columns: df_f = df_f[df_f["centre"]   == ctr_f]

            # Graphique écart tonnage
            if "ecart_tonnage" in df_f.columns and "agriculteur" in df_f.columns:
                df_c = df_f.dropna(subset=["ecart_tonnage"]).sort_values("ecart_tonnage")
                if not df_c.empty:
                    fig = go.Figure()
                    fig.add_trace(go.Bar(
                        y=df_c["agriculteur"], x=df_c["ecart_tonnage"],
                        orientation="h",
                        marker_color=["#ef5350" if v<0 else "#4CAF50"
                                      for v in df_c["ecart_tonnage"]],
                        text=df_c["ecart_tonnage"].apply(lambda v:f"{v:+.1f} T"),
                        textposition="outside"))
                    fig.add_vline(x=0,line_color="#888",line_width=1.5,
                                  line_dash="dash")
                    fig.update_layout(
                        title="Écart = Tonnage livré − Tonnage recouvrement",
                        template="plotly_dark",paper_bgcolor="#161b22",
                        plot_bgcolor="#0d1117",
                        height=max(350,len(df_c)*26+80),
                        margin=dict(l=220,r=80,t=40,b=30))
                    st.plotly_chart(fig, use_container_width=True)

            # ── Section Caisses Vides par usine ─────────────────
            if "affectation_caisse" in df_f.columns:
                st.markdown("#### 📦 Détail caisses vides par usine")
                # Récap par usine des consignes caisses
                caisse_cfg = st.session_state.get("abo_params",{}).get("caisses_par_usine",{})
                if caisse_cfg:
                    cv_cols = st.columns(len(caisse_cfg))
                    usine_colors_disp = {"SICAM":"#F5A623","TUCAL":"#8B5CF6",
                                         "COMOCAP":"#3B82F6","ABIDA":"#FF6B9D","ELFALLEH":"#00E5A0"}
                    for ci_u, (usine_u, cfg_u) in enumerate(caisse_cfg.items()):
                        # Filtrer agriculteurs de cette usine en 1ère affectation
                        _mask_1 = (df_f.get("affectation_caisse","").str.startswith("1ère")
                                   if hasattr(df_f.get("affectation_caisse",""),"str") else pd.Series([False]*len(df_f)))
                        _usine_mask = df_f.get("usine", pd.Series([""] * len(df_f))).astype(str).str.upper().str.contains(usine_u.upper(), na=False)
                        _agri_1ere = df_f[_mask_1 & _usine_mask] if "usine" in df_f.columns else df_f[_mask_1]
                        n_1ere = len(_agri_1ere)
                        total_cv = _agri_1ere["consigne_caisse"].sum() if "consigne_caisse" in _agri_1ere.columns else 0
                        cout_ha = round(cfg_u["nb_ha"] * cfg_u["prix"], 1)
                        uc2 = usine_colors_disp.get(usine_u,"#888")
                        with cv_cols[ci_u]:
                            st.markdown(f"""<div style='background:#1a2332;border-radius:10px;
padding:10px;text-align:center;border-top:3px solid {uc2}'>
<div style='font-size:13px;font-weight:bold;color:{uc2}'>{usine_u}</div>
<div style='font-size:11px;color:#aaa'>{cfg_u['type']}</div>
<div style='font-size:11px;color:#ccc;margin:4px 0'>
{cfg_u['nb_ha']:.0f} caisses/ha × {cfg_u['prix']:.2f} DT = <b style='color:#fff'>{cout_ha} DT/ha</b>
</div>
<div style='font-size:11px;color:#FFD700'>{n_1ere} agri. 1ère affect.</div>
<div style='font-size:13px;font-weight:bold;color:{"#FF7043" if total_cv>0 else "#4CAF50"}'>
{total_cv:,.0f} DT total</div>
</div>""", unsafe_allow_html=True)

                # Total global caisses
                if "consigne_caisse" in df_f.columns:
                    tot_cv = df_f["consigne_caisse"].sum()
                    n_1e = df_f["affectation_caisse"].str.startswith("1ère").sum() if "affectation_caisse" in df_f.columns else 0
                    n_2e = len(df_f) - n_1e
                    cc1,cc2,cc3 = st.columns(3)
                    cc1.markdown(_metric("Total consigne caisses",
                        f"{tot_cv:,.0f} DT", color="#FF7043"), unsafe_allow_html=True)
                    cc2.markdown(_metric("1ère affectation (< 10 juil.)",
                        f"{n_1e} agriculteurs", color="#FF9800"), unsafe_allow_html=True)
                    cc3.markdown(_metric("2ème affectation (≥ 10 juil.)",
                        f"{n_2e} agriculteurs — 0 DT", color="#4CAF50"), unsafe_allow_html=True)
                st.markdown("---")

            VIEW = [c for c in ["agriculteur","commercial","ingenieur","centre","variete",
                                  "hectares","affectation_caisse","detail_caisse","taux_prise",
                                  "report","charge_a_recouvrir","consigne_caisse","consigne_plateau",
                                  "mo_recolte","tonnage_recouvrement","tonnage_livre",
                                  "ecart_tonnage","valeur_livree","solde_final",
                                  "alerte"] if c in df_f.columns]
            st.dataframe(df_f[VIEW].round(1),
                use_container_width=True,hide_index=True,height=400,
                column_config={
                    "taux_prise":st.column_config.ProgressColumn(
                        "Taux prise %",min_value=0,max_value=100,format="%.1f%%"),
                    "ecart_tonnage":st.column_config.NumberColumn("Écart (T)",format="%+.1f"),
                    "solde_final":st.column_config.NumberColumn("Solde (DT)",format="%+,.0f"),
                })
            # Exports multiples tab1
            _dl1, _dl2 = st.columns(2)
            with _dl1:
                xl2 = export_excel(df_f, st.session_state.get("abo_sotusfa_raw"))
                st.download_button("📥 Excel complet (4 feuilles)",data=xl2,
                    file_name="agroeco_vue.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True, type="primary")
            with _dl2:
                _view_df = df_f[[c for c in VIEW if c in df_f.columns]].round(1)
                st.download_button("📥 Excel — Vue actuelle",
                    data=_export_excel_table(
                        _view_df, "Par Agriculteur",
                        "Tableau Agroéconomique par Agriculteur — Campagne 2026",
                        "1F3864"),
                    file_name="agroeco_par_agriculteur.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True)

    # ══ TAB 2 — PAR INGÉNIEUR / CENTRE ════════════════════
    with t2:
        if df is None or df.empty: _no_data()
        else:
            grp_col = st.radio("Regrouper par", ["Ingénieur","Centre","Ingénieur × Centre"],
                                horizontal=True, key="t2_grp")
            if grp_col == "Ingénieur":
                gc = ["ingenieur"] if "ingenieur" in df.columns else ["centre"]
            elif grp_col == "Centre":
                gc = ["centre"] if "centre" in df.columns else ["ingenieur"]
            else:
                gc = [c for c in ["ingenieur","centre"] if c in df.columns]

            if gc:
                try:
                    _ac2 = next((c for c in ["agriculteur","client"] if c in df.columns),None)
                    if "alerte" in df.columns:
                        df["_rouge2"] = df["alerte"].astype(str).str.contains("🔴",na=False).astype(int)
                    else:
                        df["_rouge2"] = 0
                    _g2d = {"_rouge2":"sum"}
                    if _ac2:                              _g2d[_ac2]                  = "count"
                    if "hectares"            in df.columns: _g2d["hectares"]            = "sum"
                    if "qte_actif"           in df.columns: _g2d["qte_actif"]           = "sum"
                    if "taux_prise"          in df.columns: _g2d["taux_prise"]          = "mean"
                    if "charge_totale"       in df.columns: _g2d["charge_totale"]       = "sum"
                    if "tonnage_recouvrement"in df.columns: _g2d["tonnage_recouvrement"]= "sum"
                    if "tonnage_livre"       in df.columns: _g2d["tonnage_livre"]       = "sum"
                    if "ecart_tonnage"       in df.columns: _g2d["ecart_tonnage"]       = "sum"
                    if "solde_final"         in df.columns: _g2d["solde_final"]         = "sum"
                    g2_raw = df.groupby(gc).agg(_g2d).reset_index().round(1)
                    _rn2 = {_ac2:"Agriculteurs","hectares":"Hectares","qte_actif":"Plants_actifs",
                            "taux_prise":"Taux_prise_moy","charge_totale":"Charge_totale",
                            "tonnage_recouvrement":"Recouvrement_T","tonnage_livre":"Livre_T",
                            "ecart_tonnage":"Ecart_T","solde_final":"Solde_DT","_rouge2":"Alertes_rouges"}
                    g2 = g2_raw.rename(columns={k:v for k,v in _rn2.items() if k in g2_raw.columns})
                except Exception as _eg2:
                    st.warning(f"Erreur graphique : {_eg2}")
                    g2 = pd.DataFrame()

                fig2 = go.Figure()
                x_col = gc[-1] if gc else "centre"
                fig2.add_trace(go.Bar(name="Recouvrement (T)",
                    x=g2[x_col],y=g2["Recouvrement_T"],marker_color="#ef5350"))
                fig2.add_trace(go.Bar(name="Livré réel (T)",
                    x=g2[x_col],y=g2["Livre_T"],marker_color="#4CAF50"))
                fig2.update_layout(barmode="group",template="plotly_dark",
                    paper_bgcolor="#161b22",plot_bgcolor="#0d1117",
                    height=340,title=f"Recouvrement vs Livré par {grp_col}")
                st.plotly_chart(fig2,use_container_width=True)
                st.dataframe(g2,use_container_width=True,hide_index=True)
                st.download_button(
                    "📥 Excel — Par Ingénieur/Centre",
                    data=_export_excel_table(g2,
                        "Par Ingenieur Centre",
                        f"Synthèse par {grp_col} — Campagne 2026",
                        "0B4F6C"),
                    file_name="agroeco_par_ingenieur.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True)

    # ══ TAB 3 — PAR RÉGION ════════════════════════════════
    with t3:
        if df is None or df.empty: _no_data()
        elif "region" not in df.columns:
            st.warning("Colonne 'region' absente — vérifiez le fichier Bourak.")
        else:
            # Agrégation par région — approche défensive sans named agg
            try:
                _df_rg = df.copy()
                # S'assurer que toutes les colonnes numériques sont bien numériques
                for _c in ["hectares","cout_ha","rendement_ha_reel","taux_prise",
                           "recouvrement_ha","tonnage_livre"]:
                    if _c in _df_rg.columns:
                        _df_rg[_c] = pd.to_numeric(_df_rg[_c], errors="coerce")
                # Colonne agriculteur
                _ac = next((c for c in ["agriculteur","client"] if c in _df_rg.columns), None)
                # Alerte rouge
                if "alerte" in _df_rg.columns:
                    _df_rg["_rouge"] = _df_rg["alerte"].astype(str).str.contains("🔴", na=False).astype(int)
                else:
                    _df_rg["_rouge"] = 0
                # Construction agg simple
                _rg_dict = {"_rouge": "sum"}
                if _ac: _rg_dict[_ac] = "count"
                for _c in ["hectares","cout_ha","rendement_ha_reel",
                           "taux_prise","recouvrement_ha","tonnage_livre"]:
                    if _c in _df_rg.columns:
                        _rg_dict[_c] = "mean" if _c not in ["hectares","tonnage_livre"] else "sum"
                rg_raw = _df_rg.groupby("region").agg(_rg_dict).reset_index().round(1)
                # Renommer
                _rename_map = {
                    _ac: "Agriculteurs", "hectares": "Hectares",
                    "cout_ha": "Cout_ha_moy", "rendement_ha_reel": "Rendement_moy",
                    "taux_prise": "Taux_prise_moy", "recouvrement_ha": "Recouvrement_ha",
                    "tonnage_livre": "Tonnage_total", "_rouge": "Alertes_rouges"
                }
                rg = rg_raw.rename(columns={k:v for k,v in _rename_map.items() if k in rg_raw.columns})
            except Exception as _e3:
                st.warning(f"Erreur agrégation région : {_e3}")
                rg = pd.DataFrame({"region":[]})

            fig3 = px.bar(rg,x="region",y="Rendement_moy",
                color="Rendement_moy",
                color_continuous_scale=["#ef5350","#FF9800","#4CAF50"],
                template="plotly_dark",text_auto=".1f",
                title="Rendement moyen (T/ha) par région")
            fig3.update_layout(paper_bgcolor="#161b22",plot_bgcolor="#0d1117",height=340)
            st.plotly_chart(fig3,use_container_width=True)
            st.dataframe(rg,use_container_width=True,hide_index=True)
            st.download_button(
                "📥 Excel — Par Région",
                data=_export_excel_table(rg,
                    "Par Region",
                    "Performance par Région — Campagne 2026",
                    "1A5C2A"),
                file_name="agroeco_par_region.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True)

    # ══ TAB 4 — PAR VARIÉTÉ ═══════════════════════════════
    with t4:
        if df is None or df.empty: _no_data()
        elif "variete" not in df.columns:
            st.warning("Colonne 'variete' absente — vérifiez le fichier Royal.")
        else:
            try:
                _df_v = df.dropna(subset=["variete"]).copy()
                _ac_v = next((c for c in ["agriculteur","client"] if c in _df_v.columns),None)
                _vagg = {}
                if _ac_v:                           _vagg[_ac_v]             = "count"
                if "hectares"       in _df_v.columns: _vagg["hectares"]       = "sum"
                if "densite_ha"     in _df_v.columns: _vagg["densite_ha"]     = "mean"
                if "rendement_ha_reel" in _df_v.columns: _vagg["rendement_ha_reel"] = "mean"
                if "taux_prise"     in _df_v.columns: _vagg["taux_prise"]     = "mean"
                if "cout_ha"        in _df_v.columns: _vagg["cout_ha"]        = "mean"
                if "tonnage_livre"  in _df_v.columns: _vagg["tonnage_livre"]  = "sum"
                vg_raw = _df_v.groupby("variete").agg(_vagg).reset_index()
                _rnv = {_ac_v:"Agriculteurs","hectares":"Hectares","densite_ha":"Densite_moy",
                        "rendement_ha_reel":"Rendement_moy","taux_prise":"Taux_prise",
                        "cout_ha":"Cout_ha_moy","tonnage_livre":"Tonnage_total"}
                vg = vg_raw.rename(columns={k:v for k,v in _rnv.items() if k in vg_raw.columns})
                if "Rendement_moy" in vg.columns:
                    vg = vg.sort_values("Rendement_moy",ascending=False).round(1)
            except Exception as _ev:
                st.warning(f"Erreur agrégation variété : {_ev}")
                vg = pd.DataFrame({"variete":[]})

            fig4 = px.bar(vg,x="variete",y="Rendement_moy",
                color="Rendement_moy",
                color_continuous_scale=["#ef5350","#FF9800","#4CAF50"],
                template="plotly_dark",text_auto=".1f",
                title="Rendement moyen (T/ha) par variété")
            fig4.update_layout(paper_bgcolor="#161b22",plot_bgcolor="#0d1117",height=340)
            st.plotly_chart(fig4,use_container_width=True)
            st.dataframe(vg,use_container_width=True,hide_index=True)
            st.download_button(
                "📥 Excel — Par Variété",
                data=_export_excel_table(vg,
                    "Par Variete",
                    "Performance par Variété — Campagne 2026",
                    "375623"),
                file_name="agroeco_par_variete.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True)

    # ══ TAB 5 — PAR FAMILLE INTRANT ═══════════════════════
    with t5:
        ds = st.session_state.get("abo_sotusfa_raw")
        if ds is None or ds.empty:
            _no_data()
            st.caption("Importez le fichier Sotusfa dans l'onglet ⚙️.")
        else:
            # Recalculer famille_norm si absent
            if "famille_norm" not in ds.columns and "famille" in ds.columns:
                ds = ds.copy()
                ds["famille_norm"] = ds["famille"].astype(str).str.strip().str.lower()\
                                     .map(FAM_NORM_MAP).fillna("Autre")
            if "famille_norm" not in ds.columns:
                st.warning("Colonne famille absente.")
            else:
                _cl = "client" if "client" in ds.columns else (
                      "agriculteur" if "agriculteur" in ds.columns else None)
                _agg = {"Valeur_DT":("valeur","sum")}
                if _cl: _agg["Nb_agri"] = (_cl,"nunique")
                fg = ds.groupby("famille_norm").agg(**_agg).reset_index()
                fg["Valeur_DT"] = fg["Valeur_DT"].round(0)
                fg["Part_pct"]  = (fg["Valeur_DT"]/fg["Valeur_DT"].sum()*100).round(1)
            fg = fg.sort_values("Valeur_DT",ascending=False)

            c5a,c5b = st.columns(2)
            with c5a:
                fig5 = px.pie(fg,names="famille_norm",values="Valeur_DT",hole=0.4,
                    template="plotly_dark",title="Dépenses intrants par famille (DT)",
                    color_discrete_sequence=px.colors.qualitative.Set2)
                fig5.update_layout(paper_bgcolor="#161b22",height=370)
                st.plotly_chart(fig5, use_container_width=True)
            with c5b:
                fig5b = px.bar(fg, x="famille_norm", y="Valeur_DT",
                    color="Valeur_DT",
                    color_continuous_scale=["#1A5C2A","#4CAF50","#A5D6A7"],
                    template="plotly_dark", text_auto=",.0f",
                    title="Dépenses par famille (DT)")
                fig5b.update_traces(textposition="outside", textfont_size=11)
                fig5b.update_layout(paper_bgcolor="#161b22",
                    plot_bgcolor="#0d1117", height=370,
                    xaxis_tickangle=-30)
                st.plotly_chart(fig5b, use_container_width=True)

            # Tableau + téléchargement
            fg_disp = fg.rename(columns={"famille_norm":"Famille","Part_pct":"Part %"})
            st.dataframe(fg_disp, use_container_width=True, hide_index=True,
                column_config={
                    "Valeur_DT": st.column_config.NumberColumn("Valeur (DT)", format="%,.0f"),
                    "Part %":    st.column_config.ProgressColumn(
                        "Part %", min_value=0, max_value=100, format="%.1f%%"),
                })
            st.download_button(
                "📥 Télécharger Famille Intrant (Excel)",
                data=_export_excel_table(fg_disp,
                    "Famille Intrant",
                    "Dépenses Intrants par Famille — Campagne 2026",
                    "1A5C2A"),
                file_name="famille_intrant_2026.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True)

            # ══ ANALYSE AGRONOMIQUE ══════════════════════════════════
            st.markdown("---")
            st.markdown("#### 🌱 Analyse Agronomique — Intrants vs Production")
            st.caption("DAP · Engrais · Fertilissants · Fongicides · Insecticides par agriculteur")

            # ── Données disponibles ──────────────────────────────
            _ds_agro = st.session_state.get("abo_sotusfa_raw")
            _df_main = df  # merged data

            if _ds_agro is None or _ds_agro.empty:
                st.warning("⚠️ Importez le fichier **Sotusfa** dans ⚙️ pour voir cette analyse.")
            else:
                import plotly.graph_objects as _go2
                import numpy as _np_agro

                _ds_agro = _ds_agro.copy()

                # Colonnes
                _cl_agro  = next((c for c in ["client","agriculteur"] if c in _ds_agro.columns), None)
                _fam_agro = next((c for c in ["famille","famille_norm"] if c in _ds_agro.columns), None)
                _val_agro = next((c for c in ["valeur","total_ttc","Total TTC"] if c in _ds_agro.columns), None)
                _qte_agro = next((c for c in ["qte","Qte"] if c in _ds_agro.columns), None)
                _art_agro = next((c for c in ["article","Article"] if c in _ds_agro.columns), None)

                if not _cl_agro or not _fam_agro:
                    st.info("Structure Sotusfa incompatible — colonnes client/famille manquantes.")
                else:
                    # Mapping familles réelles Sotusfa → catégories agro
                    _FAM_MAP = {
                        "engrais":      "🧪 DAP & Engrais",
                        "Engrais":      "🧪 DAP & Engrais",
                        "fertilissant": "🌿 Fertilissants",
                        "Fertilissant": "🌿 Fertilissants",
                        "fongicide":    "🛡️ Fongicides",
                        "Fongicide":    "🛡️ Fongicides",
                        "insecticide":  "🦟 Insecticides",
                        "Insecticide":  "🦟 Insecticides",
                        "IRRIGATIONS":  "💧 Irrigation",
                        "IRRIGATIONS TURK": "💧 Irrigation",
                        "HERBICIDE":    "🌾 Herbicides",
                        "Divers":       "📦 Divers",
                    }
                    _ds_agro["_cat"] = _ds_agro[_fam_agro].astype(str).str.strip().map(_FAM_MAP)
                    _ds_agro = _ds_agro[_ds_agro["_cat"].notna()]

                    _val_col = _val_agro if _val_agro else (_qte_agro if _qte_agro else None)
                    if _val_col:
                        _ds_agro[_val_col] = pd.to_numeric(_ds_agro[_val_col], errors="coerce").fillna(0)

                    # Pivot : 1 ligne par client, colonnes = catégories
                    if _val_col:
                        _pivot = _ds_agro.groupby([_cl_agro, "_cat"])[_val_col].sum()                                         .unstack("_cat", fill_value=0).reset_index()
                        _pivot.columns.name = None
                        _pivot = _pivot.rename(columns={_cl_agro: "client"})
                    else:
                        _pivot = _ds_agro.groupby([_cl_agro, "_cat"])["_cat"].count()                                         .unstack("_cat", fill_value=0).reset_index()                                         .rename(columns={_cl_agro: "client"})

                    # Tonnage : depuis df merged ou depuis prévision Mai
                    _ton_col = None
                    if _df_main is not None and not _df_main.empty:
                        _agri_col = next((c for c in ["agriculteur","client"] if c in _df_main.columns), None)
                        for tc in ["tonnage_livre","prevision_juin","prevision_mai","prevision_dec"]:
                            if tc in _df_main.columns and _df_main[tc].fillna(0).sum() > 0:
                                _ton_col = tc; break
                        if _agri_col and _ton_col:
                            # Construire la liste de colonnes en vérifiant leur existence
                            _keep_cols = [_agri_col, _ton_col]
                            for _extra in ["commercial", "region"]:
                                if _extra in _df_main.columns:
                                    _keep_cols.append(_extra)
                            _ton_df = _df_main[_keep_cols].copy()
                            _ton_df = _ton_df.rename(columns={_agri_col: "client"})
                            _ton_df[_ton_col] = pd.to_numeric(_ton_df[_ton_col], errors="coerce").fillna(0)
                            _ton_df = _ton_df[_ton_df[_ton_col] > 0]
                            _ma = _pivot.merge(_ton_df, on="client", how="inner")
                        else:
                            _ma = _pivot.copy()
                            _ton_col = None
                    else:
                        _ma = _pivot.copy()
                        _ton_col = None

                    _cats = [c for c in _ma.columns if c not in
                             ["client","commercial","region",_ton_col or "x"]]

                    # ── KPIs intrants ────────────────────────────────
                    st.markdown("**📊 Total intrants Sotusfa par catégorie**")
                    _kpi_cols = st.columns(min(len(_cats), 5))
                    _COLORS_AGR = {
                        "🧪 DAP & Engrais": "#42A5F5",
                        "🌿 Fertilissants": "#66BB6A",
                        "🛡️ Fongicides":    "#AB47BC",
                        "🦟 Insecticides":  "#FF7043",
                        "💧 Irrigation":    "#26C6DA",
                        "🌾 Herbicides":    "#FFA726",
                        "📦 Divers":        "#78909C",
                    }
                    for ci4, cat in enumerate(_cats):
                        if ci4 < 5:
                            tot_cat = _ma[cat].sum() if cat in _ma.columns else 0
                            _kpi_cols[ci4].metric(cat, f"{tot_cat:,.0f} DT")

                    # ── GRAPHIQUE 1 : Scatter corrélation ────────────
                    if _ton_col and _cats:
                        st.markdown(f"**📈 Corrélation Intrants → {_ton_col.replace('_',' ').title()}**")
                        _scatter_cats = [c for c in _cats if c in ["🧪 DAP & Engrais",
                                         "🌿 Fertilissants","🛡️ Fongicides","🦟 Insecticides"]]
                        if not _scatter_cats:
                            _scatter_cats = _cats[:4]

                        _sc_cols = st.columns(min(len(_scatter_cats), 3))
                        for ci5, cat in enumerate(_scatter_cats[:3]):
                            _df_sc = _ma[_ma[cat] > 0][["client", cat, _ton_col] +
                                        (["region"] if "region" in _ma.columns else [])].copy()
                            if len(_df_sc) < 3:
                                continue
                            _x = _df_sc[cat].values
                            _y = _df_sc[_ton_col].values
                            try:
                                _a, _b = _np_agro.polyfit(_x, _y, 1)
                                _r = float(_np_agro.corrcoef(_x, _y)[0,1])
                            except Exception:
                                _a = _b = _r = 0.0

                            _fig_sc = px.scatter(
                                _df_sc, x=cat, y=_ton_col,
                                hover_name="client",
                                color="region" if "region" in _df_sc.columns else None,
                                labels={cat: f"{cat} (DT)",
                                        _ton_col: "Tonnage (T)"},
                                title=f"{cat}<br><sup>r = {_r:.2f} | {len(_df_sc)} agriculteurs</sup>",
                                template="plotly_dark",
                                color_discrete_sequence=px.colors.qualitative.Set2,
                            )
                            # Ligne de tendance
                            if abs(_r) > 0.05 and _a != 0:
                                _xl = _np_agro.linspace(_x.min(), _x.max(), 50)
                                _fig_sc.add_scatter(
                                    x=_xl, y=_a*_xl+_b,
                                    mode="lines", name="Tendance",
                                    line=dict(color=_COLORS_AGR.get(cat,"#fff"),
                                              width=2.5, dash="dot"),
                                    showlegend=False,
                                )
                            _fig_sc.update_layout(
                                paper_bgcolor="#161b22",
                                plot_bgcolor="#0d1117",
                                height=340,
                                font=dict(color="#f0f6fc", size=10),
                                showlegend=("region" in _df_sc.columns),
                                legend=dict(font=dict(size=8)),
                                margin=dict(t=60,b=30,l=50,r=20),
                            )
                            with _sc_cols[ci5 % 3]:
                                st.plotly_chart(_fig_sc, use_container_width=True)

                    # ── GRAPHIQUE 2 : Top 20 barres horizontales ─────
                    st.markdown("**🏆 Top 20 Agriculteurs — Production vs Intrants**")
                    _sort_col = _ton_col if _ton_col and _ton_col in _ma.columns else                                 (_cats[0] if _cats else None)
                    if _sort_col:
                        _top20 = _ma.nlargest(20, _sort_col).sort_values(_sort_col, ascending=True)
                        _fig_top = _go2.Figure()

                        # Barre principale = tonnage
                        if _ton_col and _ton_col in _top20.columns:
                            _fig_top.add_trace(_go2.Bar(
                                y=_top20["client"], x=_top20[_ton_col],
                                name="Tonnage (T)", orientation="h",
                                marker_color="#FF9800", marker_opacity=0.9,
                                text=_top20[_ton_col].apply(lambda v: f"{v:,.0f}T"),
                                textposition="outside", textfont=dict(size=9),
                            ))

                        # Barres intrants (en % max pour superposition)
                        for cat in (_scatter_cats if _ton_col else _cats[:4]):
                            if cat not in _top20.columns: continue
                            _max_v = _top20[cat].max()
                            _max_t = _top20[_ton_col].max() if _ton_col in _top20.columns else 1
                            _norm = _top20[cat] / _max_v * _max_t * 0.4 if _max_v > 0 else 0
                            _fig_top.add_trace(_go2.Bar(
                                y=_top20["client"], x=_norm,
                                name=f"{cat} (normalisé)",
                                orientation="h",
                                marker_color=_COLORS_AGR.get(cat,"#888"),
                                marker_opacity=0.55,
                                visible="legendonly",
                                customdata=_top20[[cat]],
                                hovertemplate="%{y}<br>" + cat + ": %{customdata[0]:,.0f} DT<extra></extra>",
                            ))

                        _fig_top.update_layout(
                            barmode="overlay",
                            template="plotly_dark",
                            paper_bgcolor="#161b22",
                            plot_bgcolor="#0d1117",
                            height=max(400, len(_top20)*22),
                            title="Top 20 — Activer les intrants dans la légende pour comparer",
                            yaxis=dict(tickfont=dict(size=9)),
                            legend=dict(orientation="h", yanchor="bottom",
                                        y=1.01, font=dict(size=9)),
                            font=dict(color="#f0f6fc"),
                            margin=dict(l=200, r=80, t=80, b=30),
                        )
                        st.plotly_chart(_fig_top, use_container_width=True)

                    # ── TABLEAU ──────────────────────────────────────
                    st.markdown("**📋 Tableau complet**")
                    _tbl_cols = ["client"] + _cats +                                 ([_ton_col] if _ton_col else []) +                                 [c for c in ["commercial","region"] if c in _ma.columns]
                    _tbl_show = _ma[[c for c in _tbl_cols if c in _ma.columns]]                                .sort_values(_sort_col if _sort_col else _tbl_cols[1],
                                             ascending=False).round(0)
                    st.dataframe(_tbl_show, use_container_width=True,
                                 hide_index=True, height=350)
                    st.download_button(
                        "📥 Excel — Analyse Intrants vs Production",
                        data=_export_excel_table(
                            _tbl_show, "Analyse Intrants",
                            "Intrants vs Production — Campagne 2026", "1A5C2A"),
                        file_name="analyse_intrants_production_2026.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True)

    # ══ TAB 6 — PRÉVISIONS VS RÉALISÉ ═════════════════════    # ══ TAB 6 — PRÉVISIONS VS RÉALISÉ ═════════════════════
    with t6:
        if df is None or df.empty:
            _no_data()
        else:
            prev_exist = [c for c in ["prevision_dec","prevision_mai",
                                       "prevision_juin","tonnage_livre"]
                          if c in df.columns]
            # Afficher même avec une seule source
            if not any(c in df.columns for c in
                       ["prevision_dec","prevision_mai","prevision_juin"]):
                st.info("Importez un fichier de prévision (Déc ou Mai) "
                        "dans l'onglet ⚙️.")
            else:
                # Totaux par période
                labels_map = {
                    "prevision_dec":  "Prévision Déc",
                    "prevision_mai":  "Prévision Mai",
                    "prevision_juin": "Prévision Juin",
                    "tonnage_livre":  "Réalisé",
                }
                tots = {}
                for col, lbl in labels_map.items():
                    if col in df.columns:
                        val = df[col].fillna(0).sum()
                        if val > 0:
                            tots[lbl] = val

                if tots:
                    bar_colors = {
                        "Prévision Déc": "#78909C",
                        "Prévision Mai": "#42A5F5",
                        "Prévision Juin":"#26A69A",
                        "Réalisé":       "#FF9800",
                    }
                    LINE_STYLES = {
                        "Prévision Déc": dict(color="#78909C", width=2, dash="dot"),
                        "Prévision Mai": dict(color="#42A5F5", width=2, dash="dash"),
                        "Prévision Juin":dict(color="#26A69A", width=2, dash="dashdot"),
                        "Réalisé":       dict(color="#FF9800", width=3),
                    }
                    MARKER_SYMS = {
                        "Prévision Déc": "circle",
                        "Prévision Mai": "square",
                        "Prévision Juin":"triangle-up",
                        "Réalisé":       "diamond",
                    }

                    # ── GRAPHIQUE 1 : Courbes superposées par commercial ──
                    st.markdown("##### 📈 Courbes superposées par commercial — Déc vs Mai vs Réalisé")
                    st.caption("Chaque courbe = une prévision | Écarts verticaux = décalages entre versions")

                    # Construire données par commercial
                    _comms_all = sorted(df["commercial"].dropna().unique()) if "commercial" in df.columns else []
                    fig6_lines = go.Figure()

                    _col_map = {
                        "Prévision Déc":  "prevision_dec",
                        "Prévision Mai":  "prevision_mai",
                        "Prévision Juin": "prevision_juin",
                        "Réalisé":        "tonnage_livre",
                    }
                    _has_any = False
                    for lbl, col in _col_map.items():
                        if col in df.columns and df[col].fillna(0).sum() > 0:
                            _has_any = True
                            if "commercial" in df.columns:
                                _y = [df[df["commercial"]==c][col].fillna(0).sum() for c in _comms_all]
                                _x = list(_comms_all)
                            else:
                                _y = [df[col].fillna(0).sum()]
                                _x = ["TOTAL"]

                            # Barre + courbe superposée
                            fig6_lines.add_trace(go.Bar(
                                name=lbl, x=_x, y=_y,
                                marker_color=bar_colors.get(lbl,"#888"),
                                marker_opacity=0.65,
                                text=[f"{v:,.0f}T" for v in _y],
                                textposition="outside",
                                textfont=dict(size=9, color=bar_colors.get(lbl,"#fff")),
                            ))
                            fig6_lines.add_trace(go.Scatter(
                                name=f"Courbe {lbl}", x=_x, y=_y,
                                mode="lines+markers",
                                line=LINE_STYLES.get(lbl, dict(color="#fff", width=2)),
                                marker=dict(size=10, symbol=MARKER_SYMS.get(lbl,"circle"),
                                            color=bar_colors.get(lbl,"#888"),
                                            line=dict(width=2, color="#fff")),
                                showlegend=True,
                            ))

                    # Ligne recouvrement
                    if "tonnage_recouvrement" in df.columns:
                        recouv = df["tonnage_recouvrement"].fillna(0).sum()
                        if recouv > 0:
                            fig6_lines.add_hline(y=recouv,
                                line_dash="dot", line_color="#ef5350", line_width=2.5,
                                annotation_text=f"⚠️ Seuil recouvrement : {recouv:,.0f} T",
                                annotation_font_color="#ef5350", annotation_font_size=11,
                                annotation_position="top right")

                    fig6_lines.update_layout(
                        barmode="group", template="plotly_dark",
                        paper_bgcolor="#161b22", plot_bgcolor="#0d1117",
                        height=520,
                        title="<b>Prévisions vs Réalisé par Commercial</b>"
                              "<br><sup>Barres = volumes | Courbes = tendances | Écart vertical = décalage entre versions</sup>",
                        yaxis_title="Tonnes (T)",
                        yaxis=dict(gridcolor="#21262d"),
                        legend=dict(orientation="h", yanchor="bottom", y=1.02,
                                    traceorder="normal"),
                        font=dict(color="#f0f6fc"),
                        bargap=0.15, bargroupgap=0.04,
                    )
                    if _has_any:
                        st.plotly_chart(fig6_lines, use_container_width=True)

                    # ── GRAPHIQUE 2 : Radar statistique ──────────────
                    if len(_comms_all) >= 3:
                        st.markdown("##### 🕷️ Radar — Comparaison globale par commercial")
                        fig_radar = go.Figure()
                        _theta = list(_comms_all) + [_comms_all[0]]
                        for lbl, col in _col_map.items():
                            if col in df.columns and df[col].fillna(0).sum() > 0 and "commercial" in df.columns:
                                _r = [df[df["commercial"]==c][col].fillna(0).sum() for c in _comms_all]
                                _r += [_r[0]]
                                fig_radar.add_trace(go.Scatterpolar(
                                    r=_r, theta=_theta, fill="toself", name=lbl,
                                    line=dict(color=bar_colors.get(lbl,"#888"), width=2),
                                    fillcolor=bar_colors.get(lbl,"#888"),
                                    opacity=0.30,
                                ))
                        fig_radar.update_layout(
                            template="plotly_dark", paper_bgcolor="#161b22",
                            polar=dict(bgcolor="#0d1117",
                                       radialaxis=dict(gridcolor="#21262d"),
                                       angularaxis=dict(gridcolor="#21262d")),
                            height=420,
                            title="Radar — Répartition par commercial (toutes prévisions)",
                            legend=dict(orientation="h", yanchor="bottom", y=-0.2),
                            font=dict(color="#f0f6fc"),
                        )
                        st.plotly_chart(fig_radar, use_container_width=True)

                # KPIs comparaison
                if tots:
                    cols_k = st.columns(len(tots))
                    for i, (lbl, val) in enumerate(tots.items()):
                        ref = tots.get("Prévision Déc", val)
                        delta = val - ref if lbl != "Prévision Déc" and ref > 0 else None
                        cols_k[i].markdown(
                            _metric(lbl, f"{val:,.0f} T",
                                color=bar_colors.get(lbl, "#888"),
                                delta=delta),
                            unsafe_allow_html=True)

                # Tableau par agriculteur
                pv_cols = [c for c in [
                    "agriculteur","centre","commercial",
                    "prevision_dec","prevision_mai",
                    "prevision_juin","tonnage_livre",
                    "tonnage_recouvrement","ecart_tonnage","alerte"
                ] if c in df.columns]

                df_pv = df[pv_cols].copy()
                df_pv.columns = [c.replace("_"," ").replace("prevision","Prév.").title()
                                  for c in df_pv.columns]

                st.markdown("#### 📋 Tableau prévisions vs réalisé par agriculteur")
                st.dataframe(df_pv.round(1),
                    use_container_width=True,
                    hide_index=True,
                    height=400,
                    column_config={
                        c: st.column_config.NumberColumn(c, format="%,.1f")
                        for c in df_pv.select_dtypes("number").columns
                    })

                st.download_button(
                    "📥 Télécharger Prévisions vs Réalisé (Excel)",
                    data=_export_excel_table(
                        df_pv,
                        "Previsions",
                        "Prévisions vs Réalisé — Campagne 2026",
                        "4A235A"),
                    file_name="previsions_vs_realise_2026.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True)

    # ══════════════════════════════════════════════════════════
    # TAB 7 — ANALYSE EFFICACITÉ PROFESSIONNELLE
    # ══════════════════════════════════════════════════════════
    with t7:
        if df is None or df.empty:
            _no_data()
            st.caption("Fusionnez les données dans ⚙️ pour voir cette analyse.")
        else:
            import numpy as _npro
            import plotly.graph_objects as _gop

            st.markdown("""<div style='background:linear-gradient(90deg,#0a1628,#1a2332);
border-radius:12px;padding:16px 24px;margin-bottom:20px;border-left:4px solid #FFD700'>
<div style='font-size:1.1rem;font-weight:800;color:#f0f6fc'>
🏆 Analyse Efficacité Professionnelle — Campagne 2026</div>
<div style='font-size:.82rem;color:#8b949e;margin-top:6px'>
Score d'efficacité · Benchmark commerciaux · Matrice ROI · Recommandations automatiques
</div></div>""", unsafe_allow_html=True)

            _df7 = df.copy()
            for _col in ["hectares","tonnage_livre","taux_prise","charge_totale",
                         "valeur_livree","total_intrants","solde_final"]:
                if _col not in _df7.columns:
                    _df7[_col] = 0
                _df7[_col] = pd.to_numeric(_df7[_col], errors="coerce").fillna(0)
            for _sc2 in ["region","variete","commercial"]:
                if _sc2 not in _df7.columns:
                    _df7[_sc2] = ""

            # Métriques dérivées
            # Utiliser rendement_ha_reel (calculé dans merge_and_calculate)
            if "rendement_ha_reel" in _df7.columns and _df7["rendement_ha_reel"].fillna(0).gt(0).any():
                _df7["rendement_ha"] = _df7["rendement_ha_reel"].fillna(0)
            elif "rendement_ha_reel" in _df7.columns and _df7["rendement_ha_reel"].fillna(0).gt(0).any():
                _df7["rendement_ha"] = _df7["rendement_ha_reel"].fillna(0)
            else:
                _df7["rendement_ha"] = _df7.apply(
                    lambda r: round(r["tonnage_livre"]/r["hectares"],1) if r.get("hectares",0)>0 else 0, axis=1)
            _df7["cout_intrant_tonne"] = _df7.apply(
                lambda r: round(r["total_intrants"] / max(
                    r.get("tonnage_livre") or 0,
                    r.get("prevision_mai") or 0,
                    r.get("prevision_juin") or 0, 1), 1
                ) if r.get("total_intrants", 0) > 0 else 0, axis=1)
            _df7["cout_intrant_ha"] = _df7.apply(
                lambda r: round(r["total_intrants"]/r["hectares"],1) if r["hectares"]>0 else 0, axis=1)
            _df7["roi_pct"] = _df7.apply(
                lambda r: round((r["valeur_livree"]-r["charge_totale"])/r["charge_totale"]*100,1)
                          if r["charge_totale"]>0 else 0, axis=1)

            # ══ Score Efficacité ABSOLU (0-100) ══════════════════
            # Barèmes réels tomate industrielle Tunisie — NON relatif
            # Source : références AVFA / GIFruits Tunisie
            # ─────────────────────────────────────────────────────

            def _score_rendement(t_ha):
                """
                Barème rendement t/ha — calibré sur données réelles 2026.
                Vos agriculteurs sont dans la plage 60-110 t/ha.
                Données aberrantes (>120) → exclus (erreur de calcul ha).
                """
                try: v = float(t_ha)
                except: return 30.0
                if v > 110:   return None   # Aberrant (>110 t/ha) → exclu du score
                if v >= 90:   return 100.0  # Exceptionnel
                elif v >= 75: return 88.0   # Excellent
                elif v >= 65: return 74.0   # Très bon
                elif v >= 55: return 58.0   # Bon
                elif v >= 42: return 40.0   # Moyen
                elif v >= 28: return 22.0   # Faible
                else:         return 5.0    # Très faible

            def _score_taux_prise(tp):
                """Barème taux de prise % — vos données : 83-94%."""
                try: v = float(tp)
                except: return 40.0
                if v >= 93:   return 100.0  # Exceptionnel
                elif v >= 91: return 85.0   # Excellent
                elif v >= 88: return 70.0   # Très bon
                elif v >= 85: return 52.0   # Bon
                elif v >= 80: return 35.0   # Moyen
                else:         return 15.0   # Faible

            def _score_intrant(cout_tonne):
                """
                Barème coût intrant/tonne (DT/T).
                Si = 0 → données Sotusfa absentes → score neutre 50.
                """
                try: v = float(cout_tonne)
                except: return 50.0
                if v <= 0:    return 50.0   # Absent → neutre (ni bon ni mauvais)
                elif v <= 40: return 100.0  # Très économique
                elif v <= 60: return 80.0
                elif v <= 80: return 60.0
                elif v <= 100:return 40.0
                elif v <= 130:return 20.0
                else:         return 5.0    # Très coûteux

            def _score_roi(roi):
                """
                Score ROI — plafonné à 300% pour éviter l'inflation.
                Note : ROI élevé (~500%) est NORMAL car les charges ne
                comprennent que les avances société (plants + intrants + Bourak),
                PAS les frais réels de l'agriculteur (irrigation, labour...).
                → On plafonne à 300% pour neutraliser cet effet.
                """
                try: v = min(float(roi), 300.0)   # Plafond 300%
                except: return 0.0
                if v >= 200:  return 100.0
                elif v >= 100:return 75.0
                elif v >= 50: return 55.0
                elif v >= 0:  return 35.0
                else:         return 0.0   # Perte

            # ─── Application des barèmes ───────────────────────
            # Poids : Rendement 55% | Taux prise 35% | ROI 10%
            # (intrants à 0 = données absentes → poids redistribué)
            _df7["_s_rend"]  = _df7["rendement_ha"].apply(_score_rendement)
            _df7["_s_prise"] = _df7["taux_prise"].apply(_score_taux_prise)
            _df7["_s_int"]   = _df7["cout_intrant_tonne"].apply(_score_intrant)
            _df7["_s_roi"]   = _df7["roi_pct"].apply(_score_roi)

            # Marquer les données aberrantes (rendement > 120 t/ha)
            _aberrant = _df7["_s_rend"].isna()
            _df7.loc[_aberrant, "score_efficacite"] = float("nan")
            _df7.loc[_aberrant, "categorie"] = "⚠️ Données à vérifier"

            _valid = ~_aberrant
            _no_intrant = _df7["cout_intrant_tonne"].fillna(0).eq(0) & _valid
            _with_intrant = ~_df7["cout_intrant_tonne"].fillna(0).eq(0) & _valid

            # Sans intrants (0) : 55% rendement + 35% taux prise + 10% ROI
            if _no_intrant.any():
                _df7.loc[_no_intrant, "score_efficacite"] = (
                    _df7.loc[_no_intrant, "_s_rend"].fillna(0)  * 0.55 +
                    _df7.loc[_no_intrant, "_s_prise"] * 0.35 +
                    _df7.loc[_no_intrant, "_s_roi"]   * 0.10
                ).round(1)
            # Avec intrants : 45% rendement + 25% prise + 20% intrants + 10% ROI
            if _with_intrant.any():
                _df7.loc[_with_intrant, "score_efficacite"] = (
                    _df7.loc[_with_intrant, "_s_rend"].fillna(0)  * 0.45 +
                    _df7.loc[_with_intrant, "_s_prise"] * 0.25 +
                    _df7.loc[_with_intrant, "_s_int"]   * 0.20 +
                    _df7.loc[_with_intrant, "_s_roi"]   * 0.10
                ).round(1)

            # ─── Catégorie ABSOLUE ─────────────────────────────
            def _cat7(s):
                """
                Catégorie basée sur barèmes — calibrée sur vos données 2026.
                Score réel observé: 24-94/100 après révision.
                """
                if pd.isna(s):    return "⚠️ Données à vérifier"
                if s >= 85:       return "🏆 Excellent"
                elif s >= 70:     return "✅ Très bon"
                elif s >= 55:     return "✅ Bon"
                elif s >= 40:     return "⚠️ Moyen"
                else:             return "🔴 À améliorer"

            _df7["categorie"] = _df7["score_efficacite"].apply(_cat7)

            # ─── Note explicative ──────────────────────────────
            st.info("""
**📐 Méthode Score Efficacité — Campagne Tomate 2026**

**C'est quoi chaque indicateur ?**
- **Rendement t/ha** = Tonnes livrées ÷ Hectares réels. Vos données : 30-110 t/ha.
- **Taux prise %** = Plants actifs ÷ Plants livrés × 100 (% des plants qui ont produit). Vos données : 83-94%.
- **ROI %** = (Valeur livrée − Charges société) ÷ Charges société × 100. ⚠️ Note : ce ROI ne compte QUE les avances de la société (plants, intrants, avance Bourak), PAS les frais réels de l'agriculteur (irrigation, labour, location). C'est pour ça qu'il est élevé (200-600%). On le plafonne à 300% dans le calcul.
- **Coût maîtrisé** (radar commerciaux) = Score inversé du coût intrant/tonne. Moins l'agriculteur dépense en intrants pour produire 1 tonne, meilleur est le score.

**Barèmes calibrés sur vos données réelles :**

| Critère | Poids* | Barème (vos données 60-110 t/ha) |
|---|---|---|
| **Rendement t/ha** | 55% | <28→5 · 28-42→22 · 42-55→40 · 55-65→58 · 65-75→74 · 75-90→88 · 90+→100 |
| **Taux prise %** | 35% | <80→15 · 80-85→35 · 85-88→52 · 88-91→70 · 91-93→85 · 93+→100 |
| **ROI (plafonné 300%)** | 10% | <0→0 · 0-50→35 · 50-100→55 · 100-200→75 · 200+→100 |

*Sans données Sotusfa (coût=0). Avec : Rendement 45% · Prise 25% · Intrants 20% · ROI 10%.

**Catégories** : 🔴 <40 · ⚠️ 40-55 · ✅ 55-70 · ✅ 70-85 · 🏆 85+
⚠️ Rendement >110 t/ha = données suspectes (prob. erreur ha ACHREF) → exclu du score.
""")

            # ── KPIs ─────────────────────────────────────────
            st.markdown("### 📊 Indicateurs Clés")
            _kp = st.columns(5)
            _kp[0].markdown(_metric("Score moyen",f"{_df7['score_efficacite'].mean():.1f}/100",color="#FFD700"),unsafe_allow_html=True)
            _rend_pos = _df7["rendement_ha"][_df7["rendement_ha"]>0]
            _kp[1].markdown(_metric("Rendement moyen",f"{_rend_pos.mean():.1f} t/ha" if len(_rend_pos)>0 else "N/A",color="#4CAF50"),unsafe_allow_html=True)
            _cout_pos = _df7["cout_intrant_tonne"][_df7["cout_intrant_tonne"]>0]
            _kp[2].markdown(_metric("Coût intrant/tonne",f"{_cout_pos.mean():.0f} DT/T" if len(_cout_pos)>0 else "N/A",color="#FF9800"),unsafe_allow_html=True)
            _kp[3].markdown(_metric("ROI moyen",f"{_df7['roi_pct'].mean():+.1f}%",
                color="#4CAF50" if _df7["roi_pct"].mean()>=0 else "#ef5350"),unsafe_allow_html=True)
            _kp[4].markdown(_metric("Excellents (≥75)",f"{(_df7['score_efficacite']>=75).sum()} agri",color="#FFD700"),unsafe_allow_html=True)

            # ── Matrice Efficacité ────────────────────────────
            st.markdown("---")
            st.markdown("### 🔷 Matrice Efficacité — Coût Intrant vs Rendement")
            st.caption("4 quadrants : Efficace · Surinvesti · Potentiel · Inefficace")
            _dfm = _df7[(_df7["cout_intrant_ha"]>0)&(_df7["rendement_ha"]>0)].copy()
            if not _dfm.empty:
                _med_c = _dfm["cout_intrant_ha"].median()
                _med_r = _dfm["rendement_ha"].median()
                _ac = next((c for c in ["agriculteur","client"] if c in _dfm.columns),None)
                _CAT_COL = {"🏆 Excellent":"#FFD700","✅ Bon":"#4CAF50","⚠️ Moyen":"#FF9800","🔴 À améliorer":"#ef5350"}
                _figm = _gop.Figure()
                for _cat7v, _cc in _CAT_COL.items():
                    _sub = _dfm[_dfm["categorie"]==_cat7v]
                    if _sub.empty: continue
                    _figm.add_trace(_gop.Scatter(
                        x=_sub["cout_intrant_ha"], y=_sub["rendement_ha"],
                        mode="markers", name=_cat7v,
                        marker=dict(color=_cc,size=10,opacity=0.85,line=dict(width=1,color="#fff")),
                        text=_sub[_ac] if _ac else None,
                        hovertemplate="<b>%{text}</b><br>Coût: %{x:,.0f} DT/ha<br>Rend: %{y:.1f} t/ha<extra></extra>",
                    ))
                _figm.add_hline(y=_med_r,line_dash="dash",line_color="#555",line_width=1.5,
                    annotation_text=f"Médiane {_med_r:.1f}t/ha",annotation_font_color="#999")
                _figm.add_vline(x=_med_c,line_dash="dash",line_color="#555",line_width=1.5,
                    annotation_text=f"Médiane {_med_c:.0f}DT/ha",annotation_font_color="#999")
                _xmax = _dfm["cout_intrant_ha"].quantile(0.9)
                _ymax = _dfm["rendement_ha"].quantile(0.9)
                for _ql,_xa,_ya,_qc in [
                    ("⭐ EFFICACE",_med_c*0.3,_ymax*0.9,"#4CAF50"),
                    ("💸 SURINVESTI",_xmax*0.75,_ymax*0.9,"#FF9800"),
                    ("🔍 POTENTIEL",_med_c*0.3,_med_r*0.3,"#42A5F5"),
                    ("❌ INEFFICACE",_xmax*0.75,_med_r*0.3,"#ef5350")]:
                    _figm.add_annotation(x=_xa,y=_ya,text=_ql,showarrow=False,
                        font=dict(size=10,color=_qc),bgcolor="rgba(0,0,0,0.5)",
                        bordercolor=_qc,borderwidth=1,borderpad=4)
                _figm.update_layout(template="plotly_dark",paper_bgcolor="#161b22",
                    plot_bgcolor="#0d1117",height=460,
                    xaxis_title="Coût intrants / ha (DT)",yaxis_title="Rendement (t/ha)",
                    legend=dict(orientation="h",y=1.02,font=dict(size=10)),font=dict(color="#f0f6fc"))
                st.plotly_chart(_figm,use_container_width=True)

            # ── Benchmark Commerciaux ─────────────────────────
            st.markdown("---")
            st.markdown("### 👔 Benchmark Commerciaux — Radar 5 axes")
            if _df7["commercial"].ne("").any():
                _cb = _df7.groupby("commercial").agg(
                    Rend_moy=("rendement_ha",lambda x: round(x[x>0].mean(),1) if (x>0).any() else 0),
                    Taux_prise=("taux_prise",lambda x: round(x[x>0].mean(),1) if (x>0).any() else 0),
                    Cout_T=("cout_intrant_tonne",lambda x: round(x[x>0].mean(),0) if (x>0).any() else 0),
                    ROI_pos=("roi_pct",lambda x: round((x>0).mean()*100,0)),
                    Score=("score_efficacite","mean"),
                    Nb=("score_efficacite","count"),
                    Tonnage=("tonnage_livre","sum"),
                ).reset_index().round(1)

                def _n100(s,inv=False):
                    mn,mx=s.min(),s.max()
                    if mx==mn: return pd.Series([50.0]*len(s),index=s.index)
                    n=(s-mn)/(mx-mn)*100
                    return 100-n if inv else n

                _rb = _cb.copy()
                _rb["nRend"]=_n100(_rb["Rend_moy"])
                _rb["nPrise"]=_n100(_rb["Taux_prise"])
                _rb["nCout"]=_n100(_rb["Cout_T"],inv=True)
                _rb["nROI"]=_n100(_rb["ROI_pos"])
                _rb["nScore"]=_n100(_rb["Score"])

                _CCOL={"KHALIL":"#F5A623","KHALIL MAIRECH":"#F5A623","MAKKI BEN SALAH":"#00E5A0",
                       "FEDI":"#3B82F6","JILANI OBAY":"#FF6B9D","ACHREF AJLANI":"#8B5CF6"}
                _theta7=["Rendement/ha","Taux prise","Coût maîtrisé","ROI agri","Score global"]

                _figr = _gop.Figure()
                for _,_rw in _rb.iterrows():
                    _cm = str(_rw["commercial"])
                    _rv = [_rw["nRend"],_rw["nPrise"],_rw["nCout"],_rw["nROI"],_rw["nScore"],_rw["nRend"]]
                    _figr.add_trace(_gop.Scatterpolar(
                        r=_rv,theta=_theta7+[_theta7[0]],fill="toself",name=_cm,
                        line=dict(color=_CCOL.get(_cm,"#888"),width=2),
                        fillcolor=_CCOL.get(_cm,"#888"),opacity=0.2))
                _figr.update_layout(template="plotly_dark",paper_bgcolor="#161b22",
                    polar=dict(bgcolor="#0d1117",
                               radialaxis=dict(gridcolor="#21262d",range=[0,100]),
                               angularaxis=dict(gridcolor="#21262d")),
                    height=420,title="Radar Efficacité — 5 Commerciaux",
                    legend=dict(orientation="h",y=-0.15,font=dict(size=10)),font=dict(color="#f0f6fc"))

                _cb_sorted = _cb.sort_values("Score",ascending=True)
                _figb7 = _gop.Figure()
                _figb7.add_trace(_gop.Bar(
                    y=_cb_sorted["commercial"],x=_cb_sorted["Score"],orientation="h",
                    marker_color=[_CCOL.get(c,"#888") for c in _cb_sorted["commercial"]],
                    text=_cb_sorted["Score"].apply(lambda v: f"{v:.1f}/100"),
                    textposition="outside",textfont=dict(size=11)))
                _figb7.update_layout(template="plotly_dark",paper_bgcolor="#161b22",
                    plot_bgcolor="#0d1117",height=260,title="Classement Score",
                    xaxis=dict(range=[0,100]),font=dict(color="#f0f6fc"),
                    margin=dict(l=160,r=80,t=50,b=20))

                _cr1,_cr2 = st.columns([3,2])
                with _cr1: st.plotly_chart(_figr,use_container_width=True)
                with _cr2:
                    st.plotly_chart(_figb7,use_container_width=True)
                    st.dataframe(_cb[["commercial","Score","Rend_moy","Taux_prise",
                                      "ROI_pos","Nb","Tonnage"]]                        .rename(columns={"commercial":"Commercial","Score":"Score/100",
                            "Rend_moy":"Rend(t/ha)","Taux_prise":"Taux prise",
                            "ROI_pos":"% ROI positif","Nb":"Nb agri","Tonnage":"Tonnage(T)"})                        .sort_values("Score/100",ascending=False),
                        hide_index=True,use_container_width=True)

            # ── Analyse par Variété ───────────────────────────
            if _df7["variete"].ne("").any() and (_df7["rendement_ha"]>0).any():
                st.markdown("---")
                st.markdown("### 🍅 Efficacité par Variété")
                try:
                    _df7v = _df7[_df7["rendement_ha"]>0].copy()
                    _evd  = {}
                    if "rendement_ha"     in _df7v.columns: _evd["rendement_ha"]     = ["mean","count"]
                    if "score_efficacite" in _df7v.columns: _evd["score_efficacite"] = "mean"
                    if "cout_intrant_tonne" in _df7v.columns: _evd["cout_intrant_tonne"] = "mean"
                    _vg_raw = _df7v.groupby("variete").agg(_evd)
                    _vg_raw.columns = ['_'.join(c).strip('_') if isinstance(c,tuple) else c
                                       for c in _vg_raw.columns]
                    _vg_raw = _vg_raw.reset_index()
                    # Renommer colonnes plates
                    _rn_ev = {"rendement_ha_mean":"Rend_moy","rendement_ha_count":"Nb",
                              "score_efficacite_mean":"Score_moy",
                              "cout_intrant_tonne_mean":"Cout_moy",
                              "rendement_ha":"Rend_moy","score_efficacite":"Score_moy"}
                    _vg = _vg_raw.rename(columns={k:v for k,v in _rn_ev.items() if k in _vg_raw.columns})
                    if "Rend_moy" in _vg.columns:
                        _vg = _vg.sort_values("Rend_moy",ascending=False).round(1)
                except Exception as _evv:
                    _vg = pd.DataFrame({"variete":[]})
                _vg["Recommandation"] = ["⭐ MEILLEURE" if i==0
                    else ("✅ Bonne" if r["Rend_moy"]>=_vg["Rend_moy"].median() else "💡 À optimiser")
                    for i,(_,r) in enumerate(_vg.iterrows())]
                _vg_c1,_vg_c2 = st.columns(2)
                with _vg_c1:
                    _figv=_gop.Figure()
                    _figv.add_trace(_gop.Bar(x=_vg["variete"],y=_vg["Rend_moy"],
                        marker_color=["#FFD700" if i==0 else "#42A5F5" for i in range(len(_vg))],
                        text=_vg["Rend_moy"].apply(lambda v: f"{v:.1f} t/ha"),textposition="outside"))
                    _figv.update_layout(template="plotly_dark",paper_bgcolor="#161b22",
                        plot_bgcolor="#0d1117",height=320,title="Rendement moyen par variété",
                        yaxis_title="t/ha",font=dict(color="#f0f6fc"))
                    st.plotly_chart(_figv,use_container_width=True)
                with _vg_c2:
                    st.dataframe(_vg.rename(columns={"variete":"Variété","Rend_moy":"Rend(t/ha)",
                        "Score_moy":"Score/100","Cout_moy":"Coût/T(DT)","Nb":"Nb agri"}),
                        hide_index=True,use_container_width=True,height=320)

            # ── Tableau complet avec score ────────────────────
            st.markdown("---")
            st.markdown("### 📋 Tableau Score Efficacité Complet")
            _ac7 = next((c for c in ["agriculteur","client"] if c in _df7.columns),None)
            _v7 = [c for c in [_ac7,"commercial","region","variete","score_efficacite",
                "categorie","rendement_ha","taux_prise","cout_intrant_tonne","roi_pct","solde_final"]
                if c and c in _df7.columns]
            _df7d = _df7[_v7].sort_values("score_efficacite",ascending=False).round(1)
            st.dataframe(_df7d,hide_index=True,use_container_width=True,height=380,
                column_config={
                    "score_efficacite":st.column_config.ProgressColumn("Score/100",min_value=0,max_value=100,format="%.1f"),
                    "roi_pct":st.column_config.NumberColumn("ROI%",format="%+.1f%%"),
                    "rendement_ha":st.column_config.NumberColumn("Rend(t/ha)",format="%.1f"),
                    "cout_intrant_tonne":st.column_config.NumberColumn("Coût/T(DT)",format="%.0f"),
                    "solde_final":st.column_config.NumberColumn("Solde(DT)",format="%+,.0f"),
                })

            # ── Top & Bottom ──────────────────────────────────
            _tb1,_tb2 = st.columns(2)
            with _tb1:
                st.markdown("**⭐ Top 10**")
                for idx,(_,r) in enumerate(_df7d.head(10).iterrows()):
                    _med = ["🥇","🥈","🥉","4️⃣","5️⃣","6️⃣","7️⃣","8️⃣","9️⃣","🔟"][idx]
                    _nm7 = r[_ac7] if _ac7 and _ac7 in r else str(r.name)
                    st.caption(f"{_med} **{_nm7}** — Score {r['score_efficacite']:.0f}/100 | Rend {r['rendement_ha']:.1f}t/ha | ROI {r['roi_pct']:+.0f}%")
            with _tb2:
                st.markdown("**🔴 À améliorer (10 derniers)**")
                for _,r in _df7d.tail(10).sort_values("score_efficacite").iterrows():
                    _nm7 = r[_ac7] if _ac7 and _ac7 in r else str(r.name)
                    _cause = ("faible rendement" if r["rendement_ha"]<15
                              else ("coût élevé" if r["cout_intrant_tonne"]>80 else "taux prise bas"))
                    st.caption(f"⚠️ **{_nm7}** — {r['score_efficacite']:.0f}/100 | Cause : {_cause}")

            # ── Recommandations ───────────────────────────────
            st.markdown("---")
            st.markdown("### 💡 Recommandations Automatiques")
            _rc1,_rc2 = st.columns(2)
            with _rc1:
                if _df7["variete"].ne("").any() and (_df7["rendement_ha"]>0).any():
                    _bv = _df7[_df7["rendement_ha"]>0].groupby("variete")["rendement_ha"].mean().idxmax()
                    st.info(f"🌱 **Variété recommandée : {_bv}** — meilleur rendement moyen. Priorité pour la prochaine campagne.")
                if _df7["commercial"].ne("").any():
                    _bc = _df7.groupby("commercial")["score_efficacite"].mean().idxmax()
                    st.success(f"👔 **Meilleur commercial : {_bc}** — partager ses méthodes de suivi avec les autres équipes.")
            with _rc2:
                _n_sous = (_df7["score_efficacite"]<35).sum()
                _n_sur  = (_df7["cout_intrant_ha"]>_df7["cout_intrant_ha"].quantile(0.75)).sum()
                if _n_sous > 0:
                    st.warning(f"⚠️ **{_n_sous} agriculteurs** ont un score < 35/100 — nécessitent un accompagnement terrain urgent.")
                if _n_sur > 0:
                    st.error(f"💸 **{_n_sur} agriculteurs** surinvestissent en intrants (quartile supérieur) sans rendement proportionnel — rationaliser les doses DAP/fongicides.")

            st.download_button("📥 Excel — Analyse Efficacité Complète",
                data=_export_excel_table(_df7d.rename(columns={
                    "score_efficacite":"Score/100","categorie":"Catégorie",
                    "rendement_ha":"Rend(t/ha)","cout_intrant_tonne":"Coût/T(DT)",
                    "roi_pct":"ROI%","solde_final":"Solde(DT)"}),
                    "Analyse Efficacite","Score Efficacité & ROI — 2026","FFD700"),
                file_name="analyse_efficacite_2026.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True)



    # ══ TAB 8 — PLAN RÉCOLTE & TRANSPORT ═══════════════════
    with t8:
        st.markdown("## 🚛 Plan Récolte & Transport")
        st.info("**Ingénieur = Commercial** (même personne)  ·  **Transport = Accessibilité** (PL / PPL / SEMI)")

        # Récupérer les données prévision
        _df_prev = st.session_state.get("abo_prev_mai")
        _df_merge = df if df is not None and not (hasattr(df,"empty") and df.empty) else None

        if _df_merge is None and _df_prev is None:
            st.info("📥 Importez le fichier **Plan Récolte** (Prévisions) dans ⚙️ Paramètres & Import")
            st.markdown("**Format attendu :** `PLAN_Recolte_Centre_2026.xlsx` avec les colonnes :")
            st.code("Client · Centre · Ha · Rendement_Ha · Usine · Date_Debut · Date_Fin · Ingénieur · Transport")
            st.stop()

        # Utiliser le df_merged si disponible, sinon prévisions
        _src = _df_merge.copy() if _df_merge is not None else pd.DataFrame()

        # ── Section 1 : Par Commercial (= Ingénieur) ─────────
        st.markdown("### 👤 Par Commercial / Ingénieur (même personne)")
        if _src is not None and not _src.empty:
            _comm_col = next((c for c in ["commercial","responsable","ingenieur"] if c in _src.columns), None)
            if _comm_col:
                _comm_list = sorted(_src[_comm_col].dropna().astype(str).unique().tolist())
                sel_comm = st.selectbox("Filtrer par commercial / ingénieur",
                                        ["Tous"] + _comm_list, key="plan_comm")
                _ing_df = _src.copy()
                if sel_comm != "Tous":
                    _ing_df = _ing_df[_ing_df[_comm_col].astype(str) == sel_comm]

                _show_cols = [c for c in [_comm_col,"client","centre","hectares",
                              "rendement_ha_reel","tonnage_livre",
                              "date_debut_recolte","date_fin_recolte",
                              "usine","acces"] if c in _ing_df.columns]
                if _show_cols:
                    _disp = _ing_df[_show_cols].copy()
                    _disp.columns = [{"commercial":"Commercial / Ingénieur",
                        "responsable":"Commercial / Ingénieur","ingenieur":"Ingénieur",
                        "client":"Client","centre":"Centre","hectares":"Ha",
                        "rendement_ha_reel":"T/ha","tonnage_livre":"Tonnage(T)",
                        "date_debut_recolte":"Déb. Récolte","date_fin_recolte":"Fin Récolte",
                        "usine":"Usine","acces":"Transport / Accès"}.get(c,c) for c in _show_cols]
                    st.dataframe(_disp, use_container_width=True, hide_index=True)

                # Résumé par commercial — approche défensive
                try:
                    _ragg = {}
                    _ac = next((c for c in ["client","agriculteur"] if c in _src.columns), None)
                    if _ac: _ragg[_ac] = "count"
                    if "hectares"     in _src.columns: _ragg["hectares"]     = "sum"
                    if "tonnage_livre" in _src.columns: _ragg["tonnage_livre"] = "sum"
                    if _ragg:
                        _rsum_raw = _src.groupby(_comm_col).agg(_ragg).reset_index().round(1)
                        _rn = {_comm_col:"Commercial / Ingénieur",
                               _ac:"Agriculteurs","hectares":"Ha","tonnage_livre":"Tonnage(T)"}
                        _rsum_raw.rename(columns={k:v for k,v in _rn.items() if k in _rsum_raw.columns}, inplace=True)
                        st.dataframe(_rsum_raw, use_container_width=True, hide_index=True)
                except Exception as _e8a:
                    st.caption(f"Résumé indisponible : {_e8a}")
            else:
                st.info("Colonne commercial/ingénieur absente dans les données.")

        st.divider()

        # ── Section 2 : Planning par Usine/Date ─────────────
        st.markdown("### 🏭 Calendrier récolte par usine")
        if _df_merge is not None and not _df_merge.empty:
            _usine_col = next((c for c in _src.columns if "usine" in c.lower()), None)
            _date_col  = next((c for c in _src.columns if "date_debut" in c.lower() or "deb_recolt" in c.lower()), None)
            if _usine_col and _date_col:
                _cal = _src.dropna(subset=[_date_col]).copy()
                _cal["_date"] = pd.to_datetime(_cal[_date_col], errors="coerce")
                _cal = _cal.dropna(subset=["_date"]).sort_values("_date")
                for usine, grp in _cal.groupby(_usine_col):
                    n = len(grp)
                    ha  = pd.to_numeric(grp["hectares"],     errors="coerce").sum() if "hectares"      in grp.columns else 0
                    ton = pd.to_numeric(grp["tonnage_livre"],errors="coerce").sum() if "tonnage_livre" in grp.columns else 0
                    with st.expander(f"🏭 {usine} — {n} agriculteurs · {ha:.0f} ha · {ton:.0f} T"):
                        show_cols = [c for c in ["client","hectares","tonnage_livre",_date_col,"acces"] if c in grp.columns]
                        if show_cols:
                            st.dataframe(grp[show_cols], use_container_width=True, hide_index=True)
                        else:
                            st.dataframe(grp, use_container_width=True, hide_index=True)
            else:
                st.info("Colonnes Usine ou Date début récolte absentes.")

        st.divider()

        # ── Section 3 : Transport = Accessibilité ──────────
        st.markdown("### 🚛 Transport = Accessibilité (PL / PPL / SEMI)")
        st.caption("Le type de transport détermine le type de véhicule : PL = camion plateau libre · PPL = avec pente · SEMI = semi-remorque")
        _acc_col = next((c for c in ["acces","accessibilite","accessibilité"] if c in _src.columns), None)
        if _acc_col and not _src.empty:
            _acc_list = sorted(_src[_acc_col].dropna().astype(str).unique().tolist())
            sel_acc = st.selectbox("Filtrer par transport / accessibilité",
                                   ["Tous"] + _acc_list, key="plan_acc")
            _trans_df = _src.copy()
            if sel_acc != "Tous":
                _trans_df = _trans_df[_trans_df[_acc_col].astype(str) == sel_acc]

            # Tableau détail
            _tcols = [c for c in [_acc_col,"client","hectares","tonnage_livre",
                                   "date_debut_recolte","date_fin_recolte","usine","zone"]
                      if c in _trans_df.columns]
            if _tcols:
                _td = _trans_df[_tcols].copy()
                _td.columns = [{"acces":"Transport / Accès","accessibilite":"Transport / Accès",
                    "client":"Client","hectares":"Ha","tonnage_livre":"Tonnage(T)",
                    "date_debut_recolte":"Déb. Récolte","date_fin_recolte":"Fin Récolte",
                    "usine":"Usine","zone":"Zone"}.get(c,c) for c in _tcols]
                st.dataframe(_td, use_container_width=True, hide_index=True)

            # Résumé par type transport
            try:
                _tagg = {}
                _tc = next((c for c in ["client","agriculteur"] if c in _src.columns), None)
                if _tc: _tagg[_tc] = "count"
                if "hectares"      in _src.columns: _tagg["hectares"]      = "sum"
                if "tonnage_livre" in _src.columns: _tagg["tonnage_livre"] = "sum"
                if _tagg:
                    _rsum_t_raw = _src.groupby(_acc_col).agg(_tagg).reset_index().round(1)
                    _rnt = {_acc_col:"Transport / Accès", _tc:"Agriculteurs",
                            "hectares":"Ha","tonnage_livre":"Tonnage(T)"}
                    _rsum_t_raw.rename(columns={k:v for k,v in _rnt.items() if k in _rsum_t_raw.columns}, inplace=True)
                    st.dataframe(_rsum_t_raw, use_container_width=True, hide_index=True)
            except Exception as _e8b:
                st.caption(f"Résumé transport indisponible : {_e8b}")
        else:
            st.info("ℹ️ Colonne Accessibilité / Transport absente.")