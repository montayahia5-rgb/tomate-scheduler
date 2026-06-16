"""
Patch final dashboard_phase10.py — 2 corrections precises:
1. Supprime le bloc mort RECTIF_COMM_DAILY (ancien import v4 + 5 fonctions _rectif_*)
2. Corrige le bouton export Tab1 pour utiliser _planning_export (avec Disponibles/Manquants)
   au lieu de p (planning brut).

Usage: python fix_dashboard_final.py
"""
import shutil, sys, ast

FILE = "dashboard_phase10.py"

def main():
    try:
        with open(FILE, "r", encoding="utf-8") as f:
            src = f.read()
    except FileNotFoundError:
        print(f"ERREUR: {FILE} introuvable dans ce dossier.")
        sys.exit(1)

    shutil.copy2(FILE, FILE.replace(".py", "_BACKUP_FINAL.py"))
    print(f"Backup cree: {FILE.replace('.py','_BACKUP_FINAL.py')}")

    original_len = len(src)
    applied = []

    # ── PATCH 1 : Supprimer le bloc mort RECTIF_COMM_DAILY ────────────
    OLD_DEAD_BLOCK = '''# Import données rectifiées (SOURCE PRINCIPALE — reference_interne 13/06/2026)
try:
    from comparaison_tab import (
        RECTIF_COMM_DAILY, RECTIF_USINE_DAILY,
        RECTIF_COMM_DICT, RECTIF_USINE_DICT,
        RECTIF_STATS_COMM, RECTIF_STATS_USINE,
    )
    RECTIF_AVAILABLE = True
except ImportError:
    RECTIF_AVAILABLE = False
    # Fallback vide si comparaison_tab absent
    RECTIF_COMM_DAILY = {}; RECTIF_USINE_DAILY = {}
    RECTIF_COMM_DICT  = {}; RECTIF_USINE_DICT  = {}
    RECTIF_STATS_COMM = {}; RECTIF_STATS_USINE = {}

# ── Helpers données rectifiées ─────────────────────────────
def _rectif_total_daily():
    """DataFrame total journalier depuis données rectifiées (tous commerciaux)"""
    import pandas as _pd
    rows = []
    for comm, vals in RECTIF_COMM_DAILY.items():
        for i, v in enumerate(vals):
            if v > 0:
                d = _pd.Timestamp("2026-06-20") + _pd.Timedelta(days=i)
                rows.append({"Date": d, "Tonnes/Jour": float(v), "Commercial": comm})
    if not rows:
        return _pd.DataFrame(columns=["Date","Tonnes/Jour","Commercial"])
    df = _pd.DataFrame(rows)
    return df.groupby("Date")["Tonnes/Jour"].sum().reset_index()

def _rectif_comm_daily_df():
    """DataFrame par commercial depuis données rectifiées"""
    import pandas as _pd
    rows = []
    for comm, vals in RECTIF_COMM_DAILY.items():
        for i, v in enumerate(vals):
            if v > 0:
                d = _pd.Timestamp("2026-06-20") + _pd.Timedelta(days=i)
                rows.append({"Date": d, "Commercial": comm, "Tonnes/Jour": float(v)})
    if not rows:
        return _pd.DataFrame(columns=["Date","Commercial","Tonnes/Jour"])
    return _pd.DataFrame(rows)

def _rectif_one_comm_daily(comm):
    """DataFrame journalier pour un commercial depuis données rectifiées"""
    import pandas as _pd
    vals = RECTIF_COMM_DAILY.get(comm, [])
    rows = [{"Date": _pd.Timestamp("2026-06-20")+_pd.Timedelta(days=i), "Tonnes/Jour": float(v)}
            for i, v in enumerate(vals) if v > 0]
    return _pd.DataFrame(rows) if rows else _pd.DataFrame(columns=["Date","Tonnes/Jour"])

def _rectif_usine_daily_df():
    """DataFrame par usine depuis données rectifiées"""
    import pandas as _pd
    rows = []
    for usine, vals in RECTIF_USINE_DAILY.items():
        for i, v in enumerate(vals):
            if v > 0:
                d = _pd.Timestamp("2026-06-20") + _pd.Timedelta(days=i)
                rows.append({"Date": d, "Usine": usine, "Tonnes/Jour": float(v)})
    if not rows:
        return _pd.DataFrame(columns=["Date","Usine","Tonnes/Jour"])
    return _pd.DataFrame(rows)

def _build_rectif_wide_excel(planning_df_ort=None):
    """
    Génère Excel avec colonnes dates (structure wide) basé sur données rectifiées.
    Colonnes: Date | Commercial | Agriculteur | Usine | Tonnes/Jour | Type Véhicule |
              Véhicules Requis | Disponibles | Manquants (à louer) | Nb Voyages |
              Pic de Récolte | 20/06/2026 | 21/06/2026 | ... | 25/08/2026
    """
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io as _io, pandas as _pd

    SEASON_DATES = list(_pd.date_range("2026-06-20","2026-08-25",freq="D"))
    PIC_S_ = _pd.Timestamp("2026-07-01").date()
    PIC_E_ = _pd.Timestamp("2026-07-15").date()

    wb = Workbook(); wb.remove(wb.active)
    ws = wb.create_sheet("Planning Rectifié (Large)")

    HDR_FILL = PatternFill("solid", start_color="1F3864", end_color="1F3864")
    HDR_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=9)
    PIC_FILL = PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC")
    GRN_FILL = PatternFill("solid", start_color="E2EFDA", end_color="E2EFDA")
    ALT_FILL = PatternFill("solid", start_color="F8F9FA", end_color="F8F9FA")
    WHT_FILL = PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")
    RED_FILL = PatternFill("solid", start_color="FFC7CE", end_color="FFC7CE")
    THIN     = Side(style="thin", color="CCCCCC")
    BORD     = Border(left=THIN,right=THIN,top=THIN,bottom=THIN)
    CTR      = Alignment(horizontal="center",vertical="center")
    LFT      = Alignment(horizontal="left",  vertical="center")

    COMM_FILLS = {"FEDI":"DEEBF7","MAKKI BEN SALAH":"E2EFDA","KHALIL":"FFF2CC",
                  "ACHREF AJLANI":"EDEDED","JILANI OBAY":"FCE4D6"}
    FLEET      = {"SICAM":{"PL":48,"PPL":6,"SEMI":13},
                  "TUCAL":{"PL":17,"PPL":0,"SEMI":2},
                  "COMOCAP":{"PL":6,"PPL":14,"SEMI":3},
                  "ABIDA":{"PL":1,"PPL":0,"SEMI":2},
                  "ELFALLEH":{"PL":0,"PPL":2,"SEMI":0}}

    FIXED_HDRS = ["Date","Commercial","Agriculteur","Usine","Tonnes/Jour",
                  "Type Véhicule","Véhicules Requis","Disponibles","Manquants (à louer)",
                  "Nb Voyages","Pic de Récolte"]
    DATE_HDRS  = [d.strftime("%d/%m/%Y") for d in SEASON_DATES]
    ALL_HDRS   = FIXED_HDRS + DATE_HDRS
    N = len(ALL_HDRS)

    # Titre
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=min(N,35))
    ws.cell(1,1).value = "Planning Rectifié — SOURCE: reference_interne 13/06/2026"
    ws.cell(1,1).font  = Font(bold=True,color="FFFFFF",name="Calibri",size=12)
    ws.cell(1,1).fill  = HDR_FILL; ws.cell(1,1).alignment = CTR
    ws.row_dimensions[1].height = 28

    # En-têtes
    for ci, h in enumerate(ALL_HDRS, 1):
        c = ws.cell(2, ci); c.value = h; c.border = BORD; c.alignment = CTR
        if ci <= len(FIXED_HDRS):
            c.fill = HDR_FILL; c.font = HDR_FONT
        else:
            d_obj = SEASON_DATES[ci-len(FIXED_HDRS)-1].date()
            c.fill = PIC_FILL if PIC_S_<=d_obj<=PIC_E_ else PatternFill("solid",start_color="1F4E79",end_color="1F4E79")
            c.font = Font(bold=True,color="FFFFFF",name="Calibri",size=8)
    ws.row_dimensions[2].height = 36

    comm_order = ["FEDI","MAKKI BEN SALAH","KHALIL","ACHREF AJLANI","JILANI OBAY"]
    data_row   = 3

    for comm in comm_order:
        rectif = RECTIF_COMM_DICT.get(comm, {})
        cfill  = PatternFill("solid", start_color=COMM_FILLS.get(comm,"F0F0F0"), end_color=COMM_FILLS.get(comm,"F0F0F0"))

        # Utiliser planning OR-Tools pour structure agriculteurs si disponible
        if planning_df_ort is not None and not planning_df_ort.empty and "Commercial" in planning_df_ort.columns:
            sub = planning_df_ort[planning_df_ort["Commercial"]==comm].copy()
            if not sub.empty and "Agriculteur" in sub.columns:
                sub["Date"] = _pd.to_datetime(sub["Date"],errors="coerce")
                grps = sub.groupby(["Agriculteur","Usine"] if "Usine" in sub.columns else ["Agriculteur"])
                for key, grp in grps:
                    agri  = key[0] if isinstance(key,tuple) else key
                    usine = key[1] if isinstance(key,tuple) and len(key)>1 else ""
                    ort_d = {}
                    if "Date" in grp.columns and "Tonnes/Jour" in grp.columns:
                        for _, r in grp.iterrows():
                            if _pd.notna(r["Date"]):
                                ort_d[str(r["Date"].date())] = float(r.get("Tonnes/Jour",0) or 0)
                    r0 = grp.iloc[0]
                    tv  = str(r0.get("Type Véhicule","") or "").upper()
                    vr  = str(r0.get("Véhicules Requis","") or "")
                    nv  = str(r0.get("Nb Voyages","") or "")
                    vr_n= int(_pd.to_numeric(vr,errors="coerce") or 0)
                    dispo= FLEET.get(usine.upper(),{}).get(tv,0)
                    manq = max(0,vr_n-dispo)
                    total_ort = sum(ort_d.values())
                    is_pic = any(PIC_S_<=_pd.Timestamp(k).date()<=PIC_E_ for k,v in ort_d.items() if v>0)

                    alt = data_row%2==0; bf = ALT_FILL if alt else WHT_FILL
                    fv = ["",comm,agri,usine,round(total_ort,0) if total_ort>0 else "",
                          tv,vr,dispo if vr_n>0 else "",manq if vr_n>0 else "",nv,
                          "⚡ PIC" if is_pic else ""]
                    for ci, val in enumerate(fv,1):
                        cell=ws.cell(data_row,ci); cell.value=val; cell.border=BORD; cell.alignment=CTR
                        if ci==2: cell.fill=cfill; cell.font=Font(bold=True,name="Calibri",size=9)
                        elif ci==9 and isinstance(val,int) and val>0:
                            cell.fill=RED_FILL; cell.font=Font(bold=True,name="Calibri",size=9,color="9C0006")
                        else: cell.fill=bf; cell.font=Font(name="Calibri",size=9)
                    for di,d in enumerate(SEASON_DATES):
                        dk=str(d.date()); val=ort_d.get(dk,"")
                        if val==0: val=""
                        ci2=len(FIXED_HDRS)+di+1
                        cell=ws.cell(data_row,ci2)
                        cell.value=int(val) if val!="" else ""; cell.border=BORD; cell.alignment=CTR
                        is_pic_d=PIC_S_<=d.date()<=PIC_E_
                        cell.fill=PIC_FILL if (is_pic_d and val!="") else (GRN_FILL if val!="" else bf)
                        cell.font=Font(name="Calibri",size=8)
                    ws.row_dimensions[data_row].height=15; data_row+=1
                continue  # prochain commercial

        # Fallback: ligne agrégée rectifiée
        total_rect=sum(float(v) for v in rectif.values())
        alt=data_row%2==0; bf=ALT_FILL if alt else WHT_FILL
        fv=["",comm,"(total commercial)","Toutes usines",round(total_rect,0),"","","","","",
            "⚡ PIC" if any(PIC_S_<=_pd.Timestamp(k).date()<=PIC_E_ for k,v in rectif.items() if float(v)>0) else ""]
        for ci,val in enumerate(fv,1):
            cell=ws.cell(data_row,ci); cell.value=val; cell.border=BORD; cell.alignment=CTR
            cell.fill=cfill if ci==2 else bf; cell.font=Font(name="Calibri",size=9,bold=(ci==2))
        for di,d in enumerate(SEASON_DATES):
            dk=str(d.date()); val=float(rectif.get(dk,0) or 0)
            ci2=len(FIXED_HDRS)+di+1
            cell=ws.cell(data_row,ci2)
            cell.value=int(val) if val>0 else ""; cell.border=BORD; cell.alignment=CTR
            is_pic_d=PIC_S_<=d.date()<=PIC_E_
            cell.fill=PIC_FILL if (is_pic_d and val>0) else (GRN_FILL if val>0 else bf)
            cell.font=Font(name="Calibri",size=8)
        ws.row_dimensions[data_row].height=15; data_row+=1

    # Largeurs colonnes
    for ci,w in enumerate([12,16,28,12,11,12,11,11,13,10,11],1):
        ws.column_dimensions[get_column_letter(ci)].width=w
    for di in range(len(SEASON_DATES)):
        ws.column_dimensions[get_column_letter(len(FIXED_HDRS)+di+1)].width=8
    ws.freeze_panes="L3"

    # Onglet récap journalier
    ws2=wb.create_sheet("Recap Journalier Rectifié")
    ws2.merge_cells(start_row=1,start_column=1,end_row=1,end_column=8)
    ws2.cell(1,1).value="Récapitulatif journalier — Données Rectifiées (reference_interne)"
    ws2.cell(1,1).fill=HDR_FILL; ws2.cell(1,1).font=Font(bold=True,color="FFFFFF",name="Calibri",size=11)
    ws2.cell(1,1).alignment=CTR; ws2.row_dimensions[1].height=24

    comm_order2=["FEDI","MAKKI BEN SALAH","KHALIL","ACHREF AJLANI","JILANI OBAY"]
    hdrs2=["Date","Jour"]+[c.split()[0] for c in comm_order2]+["TOTAL"]
    for ci,h in enumerate(hdrs2,1):
        cell=ws2.cell(2,ci); cell.value=h; cell.fill=HDR_FILL; cell.font=HDR_FONT
        cell.alignment=CTR; cell.border=BORD
    ws2.row_dimensions[2].height=20
    DAYS_FR2=["Lun","Mar","Mer","Jeu","Ven","Sam","Dim"]
    for ri,d in enumerate(SEASON_DATES):
        dk=d.date(); is_pic=PIC_S_<=dk<=PIC_E_
        vals=[float(RECTIF_COMM_DICT.get(c,{}).get(str(dk),0) or 0) for c in comm_order2]
        total_j=sum(vals)
        if total_j==0: continue
        base=PIC_FILL if is_pic else (ALT_FILL if ri%2==0 else WHT_FILL)
        row_vals=[d.strftime("%d/%m/%Y"),DAYS_FR2[d.weekday()]]+[int(v) if v>0 else "" for v in vals]+[int(total_j)]
        for ci,val in enumerate(row_vals,1):
            cell=ws2.cell(ri+3,ci); cell.value=val; cell.border=BORD; cell.alignment=CTR
            cell.fill=base; cell.font=Font(name="Calibri",size=9,bold=(ci==8))
        ws2.row_dimensions[ri+3].height=15
    for ci,w in enumerate([12,6,12,12,12,12,12,14],1):
        ws2.column_dimensions[get_column_letter(ci)].width=w
    ws2.freeze_panes="A3"

    buf=_io.BytesIO(); wb.save(buf); buf.seek(0); return buf.read()

# ── Fin helpers ─────────────────────────────────────────────

'''

    NEW_BLOCK = ''

    if OLD_DEAD_BLOCK in src:
        src = src.replace(OLD_DEAD_BLOCK, NEW_BLOCK, 1)
        applied.append("PATCH 1 — bloc mort RECTIF_COMM_DAILY supprime (import + 5 fonctions)")
    else:
        print("AVERTISSEMENT: PATCH 1 — bloc non trouve exactement (peut-etre deja nettoye)")

    # ── PATCH 2 : Fix bouton export Tab1 ────────────────────────────
    OLD_BTN = '''                st.download_button(
                    "📆 Planning RECTIFIÉ (colonnes dates)",
                    data=generate_planning_wide_excel(p),'''
    NEW_BTN = '''                st.download_button(
                    "📆 Planning RECTIFIÉ (colonnes dates)",
                    data=generate_planning_wide_excel(_planning_export),'''

    if OLD_BTN in src:
        src = src.replace(OLD_BTN, NEW_BTN, 1)
        applied.append("PATCH 2 — bouton export Tab1 utilise _planning_export (avec Disponibles/Manquants)")
    else:
        print("AVERTISSEMENT: PATCH 2 — ligne bouton non trouvee exactement")

    with open(FILE, "w", encoding="utf-8") as f:
        f.write(src)

    print(f"\\nTaille: {original_len} -> {len(src)} chars ({len(src)-original_len:+d})")
    print(f"{len(applied)}/2 patches appliques:")
    for a in applied:
        print(f"  OK {a}")

    # Validation syntaxe
    try:
        ast.parse(src)
        print("\\nSyntaxe Python validee — fichier pret.")
    except SyntaxError as e:
        print(f"\\nERREUR DE SYNTAXE ligne {e.lineno}: {e.msg}")
        print("Restaure le backup et previens-moi avant de relancer streamlit.")
        sys.exit(1)

if __name__ == "__main__":
    main()